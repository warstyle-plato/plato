"""Тесты выгрузки полной модели (/report/model).

Проверяется и структура ZIP, и то, что книги открываются как настоящий XLSX:
листы, формулы и посчитанные значения. openpyxl — только для проверки,
в рантайме не нужен. Запуск: python3 -m pytest tests -q
"""

from __future__ import annotations

import io
import json
import sys
import xml.etree.ElementTree as ET
import zipfile
from functools import lru_cache
from pathlib import Path

import pytest
from fastapi import HTTPException

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import main as _wrapper  # noqa: E402

main = _wrapper.core

openpyxl = pytest.importorskip("openpyxl", reason="openpyxl нужен только для проверки выгрузки")
from openpyxl import load_workbook  # noqa: E402

PHASING = {
    "enabled": True,
    "user_enabled": True,
    "phase_count": 3,
    "target_size_sqm": 70000,
    "phase_gap_months": 12,
    "cost_inflation_pct": 8,
    "sales_price_inflation_pct": 8,
}


@lru_cache(maxsize=8)
def _build_cached(phasing_key: str, project_name: str):
    return main.build_model_archive(
        main.DEFAULT_INPUTS, main.TEP_DEFAULT, [], json.loads(phasing_key),
        project_name=project_name,
    )


def build(phasing=None, project_name="Мытищи"):
    # Сборка тянет шаблон ПЛАТО на 113 708 формул: без кеша набор тестов
    # пересобирал бы его десятки раз.
    return _build_cached(json.dumps(phasing or {}, sort_keys=True), project_name)


def archive(phasing=None, project_name="Мытищи") -> zipfile.ZipFile:
    content, _ = build(phasing, project_name)
    return zipfile.ZipFile(io.BytesIO(content))


def workbook(zip_file: zipfile.ZipFile, name: str, *, values: bool = False):
    return load_workbook(io.BytesIO(zip_file.read(name)), data_only=values)


# --- состав архива ----------------------------------------------------------

def test_single_archive_layout():
    content, filename = build()
    assert filename.endswith(".zip")
    assert "модель" in filename
    names = zipfile.ZipFile(io.BytesIO(content)).namelist()
    # Сначала живая модель на шаблоне, следом детализация расчёта.
    assert names == ["00_Модель_Мытищи.xlsx", "90_Детализация_Мытищи.xlsx", "README.txt"]


def test_phased_archive_has_consolidator_and_phase_files():
    content, filename = build(PHASING)
    assert "очереди" in filename
    names = zipfile.ZipFile(io.BytesIO(content)).namelist()
    assert names[0] == "00_Модель_консолидация_Мытищи.xlsx"
    assert [name for name in names if name.startswith("0") and name != names[0]] == [
        "01_Модель_О1.xlsx", "02_Модель_О2.xlsx", "03_Модель_О3.xlsx",
    ]
    assert [name for name in names if name.startswith("9")] == [
        "90_Детализация_консолидация.xlsx",
        "91_Детализация_О1.xlsx", "92_Детализация_О2.xlsx", "93_Детализация_О3.xlsx",
    ]
    assert "README.txt" in names


def test_readme_describes_consolidator_only_for_phases():
    single = archive().read("README.txt").decode("utf-8")
    phased = archive(PHASING).read("README.txt").decode("utf-8")
    assert "Консолидатор" not in single
    assert "Консолидатор" in phased
    assert "SUMIF" in phased
    assert "не суммируются" in phased


def test_project_name_is_sanitised():
    _, filename = build(project_name='Проект/ЖК "Юг": 1')
    assert "/" not in filename and ":" not in filename and '"' not in filename


# --- одиночная модель -------------------------------------------------------

def test_single_workbook_sheets():
    wb = workbook(archive(), "90_Детализация_Мытищи.xlsx")
    assert wb.sheetnames == [
        "Сводка", "Вводные", "ТЭП", "Выручка", "Расходы", "ВРИ",
        "Помесячно", "Расходы помесячно", "Продажи помесячно",
        "Финансирование поквартально", "Расходы поквартально", "Продажи поквартально",
        "Денежный поток", "Календарь",
    ]


def test_monthly_sheet_matches_engine():
    result = main.calculate(main.CalcRequest(inputs=main.DEFAULT_INPUTS, tep=main.TEP_DEFAULT, rates=[]))
    finance_rows = result["finance"]["rows"]
    sheet = workbook(archive(), "90_Детализация_Мытищи.xlsx")["Помесячно"]
    # шапка + строка «Итого»
    assert sheet.max_row == len(finance_rows) + 4
    assert sheet.max_column == len(main._MODEL_FINANCE_COLUMNS) + 1
    assert sheet.cell(row=4, column=1).value == finance_rows[0]["month"]
    assert sheet.cell(row=4, column=2).value == pytest.approx(finance_rows[0]["sales"] / 1e6, rel=1e-6)


def test_monthly_totals_are_formulas_with_cached_values():
    result = main.calculate(main.CalcRequest(inputs=main.DEFAULT_INPUTS, tep=main.TEP_DEFAULT, rates=[]))
    expected = sum(row["sales"] for row in result["finance"]["rows"]) / 1e6
    zip_file = archive()
    formulas = workbook(zip_file, "90_Детализация_Мытищи.xlsx")["Помесячно"]
    values = workbook(zip_file, "90_Детализация_Мытищи.xlsx", values=True)["Помесячно"]
    last = formulas.max_row
    assert str(formulas.cell(row=last, column=2).value).startswith("=SUM(")
    assert values.cell(row=last, column=2).value == pytest.approx(expected, rel=1e-9)


def test_balances_and_rates_are_not_summed():
    sheet = workbook(archive(), "90_Детализация_Мытищи.xlsx")["Помесячно"]
    header = [cell.value for cell in sheet[3]]
    last = sheet.max_row
    for label in ("Остаток ПФ", "Ключевая ставка", "Эскроу", "Накопленная база налога"):
        column = header.index(label) + 1
        assert sheet.cell(row=last, column=column).value in (None, "")


def test_summary_sheet_carries_key_metrics():
    result = main.calculate(main.CalcRequest(inputs=main.DEFAULT_INPUTS, tep=main.TEP_DEFAULT, rates=[]))
    sheet = workbook(archive(), "90_Детализация_Мытищи.xlsx")["Сводка"]
    found = {}
    for row in sheet.iter_rows(min_col=1, max_col=2, values_only=True):
        if row[0] and row[1] is not None:
            found[str(row[0])] = row[1]
    assert found["Выручка"] == pytest.approx(result["summary"]["revenue"] / 1e6, rel=1e-6)
    assert found["LLCR"] == pytest.approx(result["summary"]["llcr"], rel=1e-6)
    assert found["Маржинальность"] == pytest.approx(result["summary"]["margin"], rel=1e-6)


def test_inputs_sheet_exports_every_field_with_key():
    sheet = workbook(archive(), "90_Детализация_Мытищи.xlsx")["Вводные"]
    keys = {row[4] for row in sheet.iter_rows(min_row=5, max_col=5, values_only=True) if row[4]}
    model_keys = {field[0] for _, fields in main.FIELD_GROUPS for field in fields}
    assert model_keys.issubset(keys)


def test_tep_total_row_is_a_formula():
    zip_file = archive()
    formulas = workbook(zip_file, "90_Детализация_Мытищи.xlsx")["ТЭП"]
    values = workbook(zip_file, "90_Детализация_Мытищи.xlsx", values=True)["ТЭП"]
    result = main.calculate(main.CalcRequest(inputs=main.DEFAULT_INPUTS, tep=main.TEP_DEFAULT, rates=[]))
    last = formulas.max_row
    assert formulas.cell(row=last, column=1).value == "Итого"
    assert str(formulas.cell(row=last, column=2).value).startswith("=SUM(")
    assert values.cell(row=last, column=2).value == pytest.approx(result["tep"]["total"]["gns"], rel=1e-6)


def test_cashflow_running_total():
    result = main.calculate(main.CalcRequest(inputs=main.DEFAULT_INPUTS, tep=main.TEP_DEFAULT, rates=[]))
    zip_file = archive()
    formulas = workbook(zip_file, "90_Детализация_Мытищи.xlsx")["Денежный поток"]
    values = workbook(zip_file, "90_Детализация_Мытищи.xlsx", values=True)["Денежный поток"]
    assert str(formulas.cell(row=4, column=5).value).startswith("=SUM($B$4:")
    total = sum(result["cashflow"]["project"]) / 1e6
    assert values.cell(row=formulas.max_row, column=5).value == pytest.approx(total, rel=1e-9)


# --- очереди и консолидатор -------------------------------------------------

def test_consolidator_contains_phase_sheets():
    wb = workbook(archive(PHASING), "90_Детализация_консолидация.xlsx")
    assert wb.sheetnames[:3] == ["Сводка", "Сравнение очередей", "Консолидация помесячно"]
    assert wb.sheetnames[3:6] == ["1. О1", "2. О2", "3. О3"]


def test_consolidation_uses_live_sumif_over_phase_sheets():
    sheet = workbook(archive(PHASING), "90_Детализация_консолидация.xlsx")["Консолидация помесячно"]
    formula = str(sheet.cell(row=5, column=2).value)
    assert formula.startswith("=SUMIF('1. О1'!$A:$A,$A5,")
    assert "'2. О2'" in formula and "'3. О3'" in formula


def test_consolidation_values_equal_sum_of_phases():
    bundle = main._run_authoritative_model(main.DEFAULT_INPUTS, main.TEP_DEFAULT, [], PHASING)
    expected: dict[str, float] = {}
    for phase in bundle["phases"]:
        for row in phase["result"]["finance"]["rows"]:
            expected[row["month"]] = expected.get(row["month"], 0.0) + (row.get("sales") or 0.0)
    values = workbook(archive(PHASING), "90_Детализация_консолидация.xlsx", values=True)["Консолидация помесячно"]
    checked = 0
    for row in values.iter_rows(min_row=5, max_col=2, values_only=True):
        month = row[0]
        if month in expected:
            assert row[1] == pytest.approx(expected[month] / 1e6, rel=1e-9)
            checked += 1
    assert checked == len(expected)


def test_consolidation_covers_every_phase_month():
    bundle = main._run_authoritative_model(main.DEFAULT_INPUTS, main.TEP_DEFAULT, [], PHASING)
    months = {
        row["month"]
        for phase in bundle["phases"]
        for row in phase["result"]["finance"]["rows"]
    }
    values = workbook(archive(PHASING), "90_Детализация_консолидация.xlsx", values=True)["Консолидация помесячно"]
    exported = {
        row[0] for row in values.iter_rows(min_row=5, max_col=1, values_only=True)
        if row[0] and row[0] != "Итого"
    }
    assert exported == months


def test_phase_comparison_totals():
    bundle = main._run_authoritative_model(main.DEFAULT_INPUTS, main.TEP_DEFAULT, [], PHASING)
    expected = sum(item["revenue"] for item in bundle["comparison"]) / 1e6
    zip_file = archive(PHASING)
    formulas = workbook(zip_file, "90_Детализация_консолидация.xlsx")["Сравнение очередей"]
    values = workbook(zip_file, "90_Детализация_консолидация.xlsx", values=True)["Сравнение очередей"]
    last = formulas.max_row
    assert formulas.cell(row=last, column=1).value == "Итого"
    assert values.cell(row=last, column=3).value == pytest.approx(expected, rel=1e-9)


def test_phase_files_hold_their_own_model():
    bundle = main._run_authoritative_model(main.DEFAULT_INPUTS, main.TEP_DEFAULT, [], PHASING)
    phase_rows = bundle["phases"][0]["result"]["finance"]["rows"]
    wb = workbook(archive(PHASING), "91_Детализация_О1.xlsx")
    assert "Помесячно" in wb.sheetnames
    assert wb["Помесячно"].max_row == len(phase_rows) + 4


def test_single_mode_has_no_consolidation_sheet():
    wb = workbook(archive(), "90_Детализация_Мытищи.xlsx")
    assert "Консолидация помесячно" not in wb.sheetnames
    assert "Сравнение очередей" not in wb.sheetnames


def test_one_phase_project_exports_as_single_model():
    content, filename = build({"enabled": True, "user_enabled": True, "phase_count": 1})
    assert "модель" in filename
    assert zipfile.ZipFile(io.BytesIO(content)).namelist() == [
        "00_Модель_Мытищи.xlsx", "90_Детализация_Мытищи.xlsx", "README.txt",
    ]


# --- эндпоинт ---------------------------------------------------------------

# --- помесячная детализация финмодели ---------------------------------------

def test_monthly_detail_reconciles_with_totals():
    result = main.calculate(main.CalcRequest(inputs=main.DEFAULT_INPUTS, tep=main.TEP_DEFAULT, rates=[]))
    monthly = result["monthly"]
    assert len(monthly["months"]) == len(result["cashflow"]["months"])
    assert sum(item["total"] for item in monthly["costs"]) == pytest.approx(result["capex"]["total"], rel=1e-9)
    assert sum(item["total"] for item in monthly["revenue"]) == pytest.approx(result["revenue"]["total"], rel=1e-9)
    assert sum(monthly["commercial_costs"]) == pytest.approx(result["commercial_costs"], rel=1e-9)


def test_monthly_costs_add_up_month_by_month():
    monthly = main.calculate(
        main.CalcRequest(inputs=main.DEFAULT_INPUTS, tep=main.TEP_DEFAULT, rates=[])
    )["monthly"]
    for index in range(len(monthly["months"])):
        assert sum(item["values"][index] for item in monthly["costs"]) == pytest.approx(
            monthly["capex_total"][index], abs=0.5
        )


def test_monthly_articles_are_labelled_and_ordered():
    monthly = main.calculate(
        main.CalcRequest(inputs=main.DEFAULT_INPUTS, tep=main.TEP_DEFAULT, rates=[])
    )["monthly"]
    labels = [item["label"] for item in monthly["costs"]]
    assert "Основное строительство, наземная часть" in labels
    assert "Земельные правоотношения / смена ВРИ" in labels
    assert labels.index("ИРД и согласования") < labels.index("Резерв")


def test_cost_multiplier_scales_the_detail():
    inputs = {**main.DEFAULT_INPUTS, "scenario_cost_multiplier": 1.1}
    result = main.calculate(main.CalcRequest(inputs=inputs, tep=main.TEP_DEFAULT, rates=[]))
    assert sum(item["total"] for item in result["monthly"]["costs"]) == pytest.approx(
        result["capex"]["total"], rel=1e-9
    )


def test_export_has_detail_sheets():
    wb = workbook(archive(), "90_Детализация_Мытищи.xlsx")
    assert "Расходы помесячно" in wb.sheetnames
    assert "Продажи помесячно" in wb.sheetnames


def test_detail_sheet_matches_the_engine():
    result = main.calculate(main.CalcRequest(inputs=main.DEFAULT_INPUTS, tep=main.TEP_DEFAULT, rates=[]))
    monthly = result["monthly"]
    zip_file = archive()
    formulas = workbook(zip_file, "90_Детализация_Мытищи.xlsx")["Расходы помесячно"]
    values = workbook(zip_file, "90_Детализация_Мытищи.xlsx", values=True)["Расходы помесячно"]
    assert formulas.max_column == len(monthly["months"]) + 2
    header = [cell.value for cell in formulas[4]]
    assert header[2] == monthly["months"][0]
    totals = {
        row[0]: row[1]
        for row in values.iter_rows(min_row=5, max_col=2, values_only=True)
        if row[0] and row[1] is not None
    }
    for item in monthly["costs"]:
        assert totals[item["label"]] == pytest.approx(item["total"] / 1e6, abs=0.01)
    # итог по строке — формула, а не константа
    assert str(formulas.cell(row=5, column=2).value).startswith("=SUM(C5:")


def test_sales_detail_sheet_holds_revenue_and_volumes():
    result = main.calculate(main.CalcRequest(inputs=main.DEFAULT_INPUTS, tep=main.TEP_DEFAULT, rates=[]))
    values = workbook(archive(), "90_Детализация_Мытищи.xlsx", values=True)["Продажи помесячно"]
    # «Квартиры» встречаются дважды: в блоке выручки и в блоке объёмов
    pairs = [
        (row[0], row[1])
        for row in values.iter_rows(min_row=1, max_col=2, values_only=True)
        if row[0] and row[1] is not None
    ]
    labels = [row[0] for row in values.iter_rows(min_row=1, max_col=1, values_only=True) if row[0]]
    assert "Выручка" in labels and "Реализованные объёмы" in labels
    apartments_revenue = next(item for item in result["monthly"]["revenue"] if item["key"] == "apartments")
    apartments_volume = next(item for item in result["monthly"]["quantity"] if item["key"] == "apartments")
    flats = [value for label, value in pairs if label == "Квартиры"]
    assert flats[0] == pytest.approx(apartments_revenue["total"] / 1e6, abs=0.01)
    assert flats[1] == pytest.approx(apartments_volume["total"], abs=0.01)


def test_phase_files_carry_their_own_detail():
    wb = workbook(archive(PHASING), "91_Детализация_О1.xlsx")
    assert "Расходы помесячно" in wb.sheetnames


# --- поквартальная сводка ---------------------------------------------------

def test_quarter_grouping():
    assert main._quarter_label("2027-01-01") == "2027-Q1"
    assert main._quarter_label("2027-12-01") == "2027-Q4"
    groups = main._quarter_groups(["2027-01-01", "2027-02-01", "2027-03-01", "2027-04-01"])
    assert [label for label, _ in groups] == ["2027-Q1", "2027-Q2"]
    assert [len(indexes) for _, indexes in groups] == [3, 1]


def test_quarterly_costs_equal_monthly_costs():
    result = main.calculate(main.CalcRequest(inputs=main.DEFAULT_INPUTS, tep=main.TEP_DEFAULT, rates=[]))
    monthly = result["monthly"]
    values = workbook(archive(), "90_Детализация_Мытищи.xlsx", values=True)["Расходы поквартально"]
    totals = {
        row[0]: row[1]
        for row in values.iter_rows(min_row=5, max_col=2, values_only=True)
        if row[0] and row[1] is not None
    }
    for item in monthly["costs"]:
        assert totals[item["label"]] == pytest.approx(item["total"] / 1e6, abs=0.01)


def test_quarterly_columns_are_quarters():
    result = main.calculate(main.CalcRequest(inputs=main.DEFAULT_INPUTS, tep=main.TEP_DEFAULT, rates=[]))
    quarters = {main._quarter_label(month) for month in result["monthly"]["months"]}
    sheet = workbook(archive(), "90_Детализация_Мытищи.xlsx")["Расходы поквартально"]
    header = [cell.value for cell in sheet[4]][2:]
    assert set(header) == quarters
    assert len(header) < len(result["monthly"]["months"])


def test_quarterly_finance_sums_flows_and_keeps_balances():
    result = main.calculate(main.CalcRequest(inputs=main.DEFAULT_INPUTS, tep=main.TEP_DEFAULT, rates=[]))
    rows = result["finance"]["rows"]
    groups = main._quarter_groups([row["month"] for row in rows])
    values = workbook(archive(), "90_Детализация_Мытищи.xlsx", values=True)["Финансирование поквартально"]
    header = [cell.value for cell in values[4]]
    sales_column = header.index("Продажи (поступления)") + 1
    balance_column = header.index("Остаток ПФ") + 1
    for offset, (label, indexes) in enumerate(groups):
        row_number = 5 + offset
        assert values.cell(row=row_number, column=1).value == label
        expected_sales = sum(rows[index]["sales"] for index in indexes) / 1e6
        assert values.cell(row=row_number, column=sales_column).value == pytest.approx(expected_sales, abs=0.01)
        expected_balance = rows[indexes[-1]]["pf_balance"] / 1e6
        assert values.cell(row=row_number, column=balance_column).value == pytest.approx(expected_balance, abs=0.01)


def test_quarterly_finance_totals_only_flows():
    sheet = workbook(archive(), "90_Детализация_Мытищи.xlsx")["Финансирование поквартально"]
    header = [cell.value for cell in sheet[4]]
    last = sheet.max_row
    assert sheet.cell(row=last, column=1).value == "Итого"
    for label in ("Остаток ПФ", "Ключевая ставка", "Эскроу"):
        assert sheet.cell(row=last, column=header.index(label) + 1).value in (None, "")


def test_phase_files_carry_quarterly_sheets():
    wb = workbook(archive(PHASING), "91_Детализация_О1.xlsx")
    assert "Расходы поквартально" in wb.sheetnames
    assert "Финансирование поквартально" in wb.sheetnames


def test_endpoint_returns_zip_attachment():
    response = main.report_model(main.ModelExportRequest(
        inputs=main.DEFAULT_INPUTS, tep=main.TEP_DEFAULT, project_name="Мишина",
    ))
    assert response.media_type == "application/zip"
    assert "attachment" in response.headers["content-disposition"]
    assert response.body[:2] == b"PK"


def test_partial_payload_is_filled_with_model_defaults():
    # неполные вводные (например, из Telegram) дополняются базовыми значениями
    response = main.report_model(main.ModelExportRequest(inputs={}, tep={}, project_name="Проверка"))
    zip_file = zipfile.ZipFile(io.BytesIO(response.body))
    sheet = workbook(zip_file, "90_Детализация_Проверка.xlsx", values=True)["Сводка"]
    found = {str(row[0]): row[1] for row in sheet.iter_rows(min_col=1, max_col=2, values_only=True) if row[0]}
    assert found["Выручка"] > 0


def test_build_failure_is_reported_in_russian(monkeypatch):
    def boom(*args, **kwargs):
        raise RuntimeError("движок недоступен")

    monkeypatch.setattr(main, "build_model_archive", boom)
    with pytest.raises(HTTPException) as exc:
        main.report_model(main.ModelExportRequest(inputs=main.DEFAULT_INPUTS, tep=main.TEP_DEFAULT))
    assert exc.value.status_code == 500
    assert "Не удалось собрать модель" in str(exc.value.detail)


def test_telegram_sends_model_archive_after_pdf(monkeypatch):
    monkeypatch.setenv("TELEGRAM_BOT_TOKEN", "test-token")
    documents: list[dict] = []
    monkeypatch.setattr(main, "_telegram_send_message", lambda *a, **kw: {"ok": True})
    monkeypatch.setattr(main, "_build_developaid_pdf", lambda payload: b"%PDF-1.4 test")
    monkeypatch.setattr(
        main,
        "_telegram_send_document_bytes",
        lambda chat_id, content, filename, caption="", content_type="application/octet-stream":
            documents.append({"filename": filename, "content": content, "type": content_type}),
    )
    result = main.calculate(main.CalcRequest(inputs=main.DEFAULT_INPUTS, tep=main.TEP_DEFAULT, rates=[]))
    session = main._telegram_session(4242, [])
    main.telegram_result(main.TelegramResultRequest(session=session, summary={
        "project_name": "Мишина",
        "report_payload": {
            "result": result,
            "inputs": main.DEFAULT_INPUTS,
            "tep": main.TEP_DEFAULT,
            "rates": [],
            "phasing": {},
            "scenario": "base",
            "project_name": "Мишина",
        },
    }))
    kinds = [item["filename"].rsplit(".", 1)[-1] for item in documents]
    assert kinds == ["pdf", "zip"]
    archive_item = documents[-1]
    assert archive_item["type"] == "application/zip"
    assert "Мишина" in archive_item["filename"]
    assert zipfile.ZipFile(io.BytesIO(archive_item["content"])).namelist()[-1] == "README.txt"


def test_export_is_available_through_wrapper():
    routes = {getattr(route, "path", "") for route in _wrapper.app.routes}
    assert "/report/model" in routes


if __name__ == "__main__":
    raise SystemExit(pytest.main([__file__, "-q"]))

# --- диаграммы --------------------------------------------------------------

def test_expense_structure_has_a_bar_chart():
    sheet = workbook(archive(), "90_Детализация_Мытищи.xlsx")["Расходы"]
    charts = sheet._charts
    assert len(charts) == 1
    chart = charts[0]
    assert type(chart).__name__ == "BarChart"
    assert len(chart.series) == 1
    # Диаграмма смотрит на строки структуры расходов, а не на итоги CAPEX.
    values = chart.series[0].val.numRef.f
    assert values.startswith("'Расходы'!$B$")


def test_monthly_sheet_has_a_debt_and_escrow_line_chart():
    book = workbook(archive(), "90_Детализация_Мытищи.xlsx")
    chart = book["Помесячно"]._charts[0]
    assert type(chart).__name__ == "LineChart"
    assert len(chart.series) == 3
    columns = {main._xlsx_column_name(index + 1): key
               for index, (key, _, _) in enumerate(main._MODEL_FINANCE_COLUMNS)}
    for series in chart.series:
        column = series.val.numRef.f.split("$")[1]
        assert columns[column] in {"pf_balance", "escrow", "bridge_balance"}
        assert series.cat.strRef.f.startswith("'Помесячно'!$A$")


def test_phase_comparison_has_llcr_and_revenue_charts():
    book = workbook(archive(PHASING), "90_Детализация_консолидация.xlsx")
    charts = book["Сравнение очередей"]._charts
    assert len(charts) == 2
    assert charts[0].series[0].val.numRef.f.startswith("'Сравнение очередей'!$I$")
    assert len(charts[1].series) == 2


def test_chart_parts_are_well_formed_and_declared():
    content, _ = build()
    inner = zipfile.ZipFile(io.BytesIO(zipfile.ZipFile(
        io.BytesIO(content)).read("90_Детализация_Мытищи.xlsx")))
    names = inner.namelist()
    types = inner.read("[Content_Types].xml").decode("utf-8")
    for name in names:
        if name.startswith(("xl/charts/", "xl/drawings/")):
            ET.fromstring(inner.read(name))
    assert "xl/charts/chart1.xml" in names
    assert "xl/drawings/drawing1.xml" in names
    assert "/xl/charts/chart1.xml" in types
    assert "/xl/drawings/drawing1.xml" in types
    # Лист должен ссылаться на рисунок через собственные отношения.
    assert any(name.startswith("xl/worksheets/_rels/") for name in names)


# --- живая модель против детализации ----------------------------------------

def test_archive_leads_with_the_live_template_model():
    """Первым в архиве идёт шаблон ПЛАТО: он и есть модель на формулах."""
    book = workbook(archive(), "00_Модель_Мытищи.xlsx")
    formulas = sum(
        1
        for sheet in book.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    )
    # Две формулы шаблона заменены датами окна рассрочки на листе «ЗУ».
    assert formulas == 113_706
    assert len(book.sheetnames) == 27
    assert book.calculation.fullCalcOnLoad is True


def test_detail_workbook_is_not_a_model_and_says_so():
    readme = archive().read("README.txt").decode("utf-8")
    assert "живая модель на шаблоне ПЛАТО" in readme
    assert "Это НЕ модель" in readme
    assert "Правка любой вводной пересчитывает книгу целиком" in readme


def test_every_phase_gets_its_own_live_model():
    names = archive(PHASING).namelist()
    models = [name for name in names if "_Модель_" in name]
    assert len(models) == 4  # консолидация плюс три очереди
    for name in models:
        book = workbook(archive(PHASING), name)
        assert len(book.sheetnames) == 27


def test_missing_template_leaves_the_detail_and_explains(monkeypatch, tmp_path):
    monkeypatch.setattr(main, "_PLATO_TEMPLATE_PATH", tmp_path / "нет.xlsx")
    content, _ = main.build_model_archive(
        main.DEFAULT_INPUTS, main.TEP_DEFAULT, [], {}, project_name="Без шаблона"
    )
    zip_file = zipfile.ZipFile(io.BytesIO(content))
    names = zip_file.namelist()
    assert names == ["90_Детализация_Без шаблона.xlsx", "README.txt"]
    readme = zip_file.read("README.txt").decode("utf-8")
    assert "Живая модель на шаблоне ПЛАТО не собрана" in readme
