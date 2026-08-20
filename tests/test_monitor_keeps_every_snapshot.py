"""Монитор хранит снимки выгрузок, а не последнее состояние.

Почти всё, что мы разбирали на Гродненской, оказалось не ошибкой методики, а
тем, что файлы — снимки разных моментов. Реестр книги отставал от реестра РСС на
два месяца, и разрыв в 454,9 млн ₽ дважды был принят за расхождение подходов.

Прошлое к тому же переписывают: между РСС на 30.06 и на 20.08 из мая ушло
124,2 млн ₽, а в июнь пришло 38,1. Ни один отдельный файл этого не показывает —
только пара снимков.

Отсюда проверки: снимок не перезаписывается, продажи живут отдельно от РСС,
переписанное прошлое видно, а маршруты закрыты тем же входом, что проекты.
"""

from __future__ import annotations

import base64
import datetime
import io

import pytest

from openpyxl import Workbook

import developaid_monitor as monitor


@pytest.fixture(autouse=True)
def _isolated(tmp_path, monkeypatch):
    monkeypatch.setattr(monitor, "_SNAPSHOT_DIR", tmp_path / "monitor")


def _estimate_book(acts):
    """Книга РСС в миниатюре: только реестр актов, остального монитору хватит."""
    book = Workbook()
    sheet = book.active
    sheet.title = "Расчет стоимости строительства"
    sheet.cell(row=9, column=1, value="Код")
    sheet.cell(row=12, column=4, value="Всего инвестиционные расходы")
    sheet.cell(row=12, column=5, value=1000.0)
    for name in ("Реестр договоров", "Реестр платежей", "Реестр выполненных работ"):
        book.create_sheet(name)
    works = book["Реестр выполненных работ"]
    for offset, (day, amount) in enumerate(acts):
        line = 9 + offset
        works.cell(row=line, column=2, value="Акт")
        works.cell(row=line, column=3, value=f"{offset + 1} от {day}")
        works.cell(row=line, column=7, value="2.2.1.4")
        works.cell(row=line, column=9, value=amount)
        works.cell(row=line, column=11, value="СтройКо")
    blob = io.BytesIO()
    book.save(blob)
    return blob.getvalue()


def test_a_snapshot_is_never_overwritten():
    """Второй файл на ту же дату — это спор, а не обновление.

    Перезаписать снимок значит потерять единственную возможность увидеть, что
    прошлое переписали.
    """
    data = _estimate_book([("15.05.2026", 100e6)])
    monitor.store_estimate("Гродненская", data, "2026-06-30")

    with pytest.raises(FileExistsError):
        monitor.store_estimate("Гродненская", data, "2026-06-30")


def test_the_snapshot_date_is_the_export_date_not_today():
    """Файл приносят и через неделю после того, как его выгрузили."""
    monitor.store_estimate("Гродненская", _estimate_book([]), "2026-06-30")

    assert monitor.snapshots("Гродненская")["estimate"] == ["2026-06-30"]


def test_a_bad_date_is_refused():
    """Без даты снимка история бессмысленна — гадать не из чего."""
    with pytest.raises(ValueError):
        monitor.store_estimate("Гродненская", _estimate_book([]), "июнь")


def test_sales_live_apart_from_the_estimate():
    """РСС обновляют еженедельно, книгу — раз в месяц.

    Ждать книгу значит не знать про продажи ничего: на Гродненской они
    отставали на пять месяцев.
    """
    monitor.store_sales("Гродненская", [
        {"month": "2026-08", "units": 4, "area": 240.0, "revenue": 160e6},
    ], "2026-08-20")
    stored = monitor.snapshots("Гродненская")

    assert stored["sales"] == ["2026-08-20"]
    assert stored["estimate"] == []


def test_rewritten_history_is_visible_between_snapshots():
    """Акт, переехавший из мая в июнь, виден только сравнением двух снимков."""
    monitor.store_estimate("Гродненская", _estimate_book([
        ("15.05.2026", 272.7e6), ("15.06.2026", 76.6e6)]), "2026-06-30")
    monitor.store_estimate("Гродненская", _estimate_book([
        ("15.05.2026", 148.4e6), ("15.06.2026", 114.7e6)]), "2026-08-20")

    moved = monitor.moved_between_snapshots(
        "Гродненская", "2026-06-30", "2026-08-20")
    by_month = {row["month"]: row for row in moved["rows"]}

    assert by_month["2026-05"]["delta"] == pytest.approx(-124.3e6)
    assert by_month["2026-06"]["delta"] == pytest.approx(38.1e6)
    assert len(moved["rewritten"]) == 2


def test_a_new_month_is_not_a_rewrite():
    """Месяц, которого в первом снимке не было, — новые данные, а не правка."""
    monitor.store_estimate("Гродненская",
                           _estimate_book([("15.05.2026", 100e6)]), "2026-06-30")
    monitor.store_estimate("Гродненская",
                           _estimate_book([("15.05.2026", 100e6),
                                           ("15.07.2026", 97.6e6)]), "2026-08-20")

    moved = monitor.moved_between_snapshots(
        "Гродненская", "2026-06-30", "2026-08-20")

    assert moved["rewritten"] == []


def test_a_view_reads_the_snapshot_not_later_than_asked():
    """Тренд строится по снимкам: каждая точка видит только своё прошлое."""
    monitor.store_estimate("Гродненская",
                           _estimate_book([("15.05.2026", 100e6)]), "2026-06-30")
    monitor.store_estimate("Гродненская",
                           _estimate_book([("15.05.2026", 100e6),
                                           ("15.07.2026", 200e6)]), "2026-08-20")

    early = monitor.build("Гродненская", cut="2026-09-01", upto="2026-06-30")
    late = monitor.build("Гродненская", cut="2026-09-01")

    assert early["source"]["estimate"] == "2026-06-30"
    assert late["source"]["estimate"] == "2026-08-20"
    assert late["money"]["accepted"] > early["money"]["accepted"]


def test_a_project_name_cannot_escape_its_folder():
    """Имя проекта приходит от человека — путь из него строить нельзя."""
    monitor.store_estimate("../../etc", _estimate_book([]), "2026-06-30")
    written = list((monitor._SNAPSHOT_DIR).glob("*/estimate/*.xlsx"))

    assert written
    assert all(monitor._SNAPSHOT_DIR in path.parents for path in written)


def test_the_routes_are_hidden_and_gated():
    """Здесь сметы, договоры и контрагенты — открытым это быть не может.

    Маршруты не показываются в схеме и живут за тем же входом, что проекты.
    """
    import main_legacy as engine

    routes = [route for route in engine.app.routes
              if "/monitor" in getattr(route, "path", "")]

    assert routes
    assert not any(getattr(route, "include_in_schema", True) for route in routes)


def test_the_service_takes_a_file_from_memory():
    """Файл приходит телом запроса — писать его на диск ради чтения незачем."""
    import developaid_actuals as actuals

    blob = io.BytesIO(_estimate_book([("15.05.2026", 100e6)]))
    works = actuals.read_completed_works(blob)

    assert works["total"] == pytest.approx(100e6)
    assert works["rows"][0]["date"] == datetime.date(2026, 5, 15)

def _schedule_book():
    """Очищенный ГПР в миниатюре."""
    book = Workbook()
    sheet = book.active
    sheet.title = "ГПР"
    sheet.cell(row=4, column=1, value="ID")
    row = ["1", "1.16.1", "СМР", "Корпус 1", "Кладка", "Работа", 0.5,
           "2026-07-01", "2026-09-30", "В работе", 60, "", "", "", "", "",
           "2.2.1.4.", "Фундаментная плита", "Наследовано"]
    for column, value in enumerate(row, 1):
        sheet.cell(row=5, column=column, value=value)
    blob = io.BytesIO()
    book.save(blob)
    return blob.getvalue()


def _pm_book():
    """Выгрузка планировщика в миниатюре: базовый план и факт для той же работы."""
    book = Workbook()
    sheet = book.active
    sheet.title = "Таблица_задач1"
    for column, title in {1: "Ид", 3: "Название_задачи", 6: "Уровень_структуры",
                          9: "Начало", 10: "Окончание",
                          19: "Фактическое_начало", 20: "Фактическое_окончание",
                          21: "Базовое_начало", 22: "Базовое_окончание",
                          36: "СДР", 39: "Суммарная_задача",
                          49: "Статус"}.items():
        sheet.cell(row=1, column=column, value=title)
    values = {1: "1", 3: "Кладка", 6: 3, 9: "01 Июль 2026 9:00",
              10: "30 Сентябрь 2026 18:00", 19: "Ср 01.07.26", 20: "НД",
              21: "01 Июнь 2026 9:00", 22: "31 Июль 2026 18:00",
              36: "1.16.1", 39: "Нет", 49: ""}
    for column, value in values.items():
        sheet.cell(row=2, column=column, value=value)
    blob = io.BytesIO()
    book.save(blob)
    return blob.getvalue()


def test_a_schedule_snapshot_is_a_pair_and_is_never_overwritten():
    """ГПР несёт код РСС, выгрузка планировщика — базовый план: хранятся парой."""
    monitor.store_schedule("Гродненская", _schedule_book(), _pm_book(),
                           "2026-07-23")

    assert monitor.snapshots("Гродненская")["schedule"] == ["2026-07-23"]
    with pytest.raises(FileExistsError):
        monitor.store_schedule("Гродненская", _schedule_book(), None,
                               "2026-07-23")


def test_the_gantt_measures_slip_from_the_baseline():
    """Отставание — от базового плана; факт идущей работы кончается срезом."""
    monitor.store_schedule("Гродненская", _schedule_book(), _pm_book(),
                           "2026-07-23")

    gantt = monitor.gantt("Гродненская", cut="2026-08-15")
    bar = gantt["bars"][0]

    assert bar["slip_days"] == 61
    assert bar["fact"] == ["2026-07-01", "2026-08-15"]
    assert gantt["source"] == {"schedule": "2026-07-23", "with_baseline": True}


def test_the_gantt_without_the_pm_export_says_so():
    """Без выгрузки планировщика Гант рисует текущий план и честно молчит о сдвиге."""
    monitor.store_schedule("Гродненская", _schedule_book(), None, "2026-07-23")

    gantt = monitor.gantt("Гродненская", cut="2026-08-15")

    assert gantt["source"]["with_baseline"] is False
    assert gantt["bars"][0]["slip_days"] is None
    assert gantt["deadline"] == {"known": False}

def _sales_book(rows):
    """Лист «План продаж» в миниатюре: период, метка, лоты, м², цена."""
    import developaid_actuals as actuals

    book = Workbook()
    sheet = book.active
    sheet.title = "План продаж"
    for offset, (month, mark, units, area, price) in enumerate(rows):
        line = actuals._SALES_FIRST_ROW + offset
        sheet.cell(row=line, column=actuals._SALES_COLUMNS["period"] + 1,
                   value=month)
        sheet.cell(row=line, column=actuals._SALES_COLUMNS["mark"] + 1,
                   value=mark)
        sheet.cell(row=line, column=actuals._SALES_COLUMNS["units"] + 1,
                   value=units)
        sheet.cell(row=line, column=actuals._SALES_COLUMNS["area"] + 1,
                   value=area)
        sheet.cell(row=line, column=actuals._SALES_COLUMNS["price"] + 1,
                   value=price)
    blob = io.BytesIO()
    book.save(blob)
    return blob.getvalue()


def test_sales_can_arrive_as_the_book_fact_rows_only():
    """Из книги берутся только строки ФАКТ: план книги — не продажи."""
    data = _sales_book([
        (datetime.date(2026, 3, 1), "ФАКТ", 5, 300.0, 650_000.0),
        (datetime.date(2026, 9, 1), "ПЛАН", 8, 500.0, 900_000.0),
    ])
    monitor.store_sales_file("Гродненская", data, "2026-07-31")

    view_sales = monitor.snapshots("Гродненская")["sales"]
    assert view_sales == ["2026-07-31"]

    import json
    stored = json.loads((monitor._SNAPSHOT_DIR / "Гродненская" / "sales"
                         / "2026-07-31.json").read_text(encoding="utf-8"))
    assert len(stored["rows"]) == 1
    assert stored["rows"][0]["month"] == "2026-03-01"
    assert stored["rows"][0]["revenue"] == pytest.approx(300.0 * 650_000.0)


def test_a_book_without_fact_rows_is_refused():
    data = _sales_book([(datetime.date(2026, 9, 1), "ПЛАН", 8, 500.0, 900_000.0)])

    with pytest.raises(ValueError):
        monitor.store_sales_file("Гродненская", data, "2026-07-31")
