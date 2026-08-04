"""Parity-блок ПРОВЕРОК и общий идентификатор расчёта в PDF и книге.

Финал плана паритета: после выравнивания методик книга обязана сама
кричать о будущем разъезде с движком. Билдер пишет контрольные числа
движка в C76:C84 листа ПРОВЕРКИ, книга считает свои в B, и вердикт листа
даёт СБОЙ при расхождении сверх допуска. Пара «PDF + книга» одного
расчёта несёт один отпечаток вводных — сверка начинается с него, а не со
спора, одна ли это версия.

Запуск: python3 -m pytest tests -q
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
sys.path.insert(0, str(Path(__file__).resolve().parent))

import main as wrapper  # noqa: E402
from xlsx_eval import Evaluator  # noqa: E402

core = wrapper.core


def _scenario():
    inputs = dict(core.DEFAULT_INPUTS)
    tep = {key: dict(value) for key, value in core.TEP_DEFAULT.items()}
    return inputs, tep


def test_the_parity_block_confirms_the_engine():
    """На честной сборке все девять строк паритета — OK, вердикт не падает."""
    sys.setrecursionlimit(400000)
    inputs, tep = _scenario()
    content, _, meta = core.build_project_workbook(
        inputs, tep, [], {}, project_name="Паритет")
    assert meta["missing"] == []
    book = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
    evaluator = Evaluator(book)
    for row in range(76, 85):
        assert book["ПРОВЕРКИ"][f"C{row}"].value is not None, f"цель строки {row}"
        assert evaluator.cell("ПРОВЕРКИ", f"F{row}") == "OK", \
            str(book["ПРОВЕРКИ"][f"A{row}"].value)
    assert evaluator.cell("ПРОВЕРКИ", "B3") in (
        "ПРОЙДЕНО", "ПРОЙДЕНО С ПРЕДУПРЕЖДЕНИЯМИ")


def test_the_verdict_fails_on_a_forged_target():
    """Сдвинутая контрольная цель роняет строку в FAIL, а вердикт — в СБОЙ:
    расхождение поверхностей больше не проходит молча."""
    sys.setrecursionlimit(400000)
    inputs, tep = _scenario()
    hints = {"parity": {**core._v4_parity_targets(
        core._run_authoritative_model(inputs, tep, [], {})["consolidated"])}}
    hints["parity"]["net_profit_mln"] += 500.0  # «движок» смещён на полмиллиарда
    content, _, _ = core.build_project_workbook(
        inputs, tep, [], {}, project_name="Сбой", finance_hints=hints)
    evaluator = Evaluator(openpyxl.load_workbook(io.BytesIO(content), data_only=False))
    assert evaluator.cell("ПРОВЕРКИ", "F81") == "FAIL"
    assert evaluator.cell("ПРОВЕРКИ", "B3") == "СБОЙ"


def test_without_targets_the_rows_stay_silent():
    """Без движковых чисел (finance_hints={}) строки паритета пусты и в
    вердикт не входят — книга по собственным формулам остаётся живой."""
    sys.setrecursionlimit(400000)
    inputs, tep = _scenario()
    content, _, _ = core.build_project_workbook(
        inputs, tep, [], {}, project_name="Тихо", finance_hints={})
    book = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
    evaluator = Evaluator(book)
    assert book["ПРОВЕРКИ"]["C76"].value in (None, "")
    assert evaluator.cell("ПРОВЕРКИ", "F76") in (None, "")
    assert evaluator.cell("ПРОВЕРКИ", "B3") in (
        "ПРОЙДЕНО", "ПРОЙДЕНО С ПРЕДУПРЕЖДЕНИЯМИ")


def test_the_pdf_and_the_book_share_one_fingerprint():
    pypdf = pytest.importorskip("pypdf")
    sys.setrecursionlimit(400000)
    inputs, tep = _scenario()
    bundle = core._run_authoritative_model(inputs, tep, [], {})
    fingerprint = core._calculation_fingerprint(inputs, tep, {})
    content, _, _ = core.build_project_workbook(
        inputs, tep, [], {}, project_name="Отпечаток", finance_hints={})
    book = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
    assert fingerprint in str(book["Источники"]["D16"].value)
    payload = {"result": bundle["consolidated"], "inputs": inputs, "tep": tep,
               "rates": [], "phasing": {}, "scenario": "base"}
    text = pypdf.PdfReader(io.BytesIO(core._build_developaid_pdf(payload))).pages[0].extract_text()
    assert fingerprint in text
