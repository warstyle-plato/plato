"""Рассрочка ВРИ в книге считается по методике движка.

Владелец, 04.09.2026: «расхождение осталось» — на его проекте книга показывала
CAPEX 18 780,4 млн ₽ против 18 716,1 у движка. Разрыв 64,3 млн разложился на
две половины, и обе — «методику меняют в двух местах»:

* шаблон начинал рассрочку ПЕРИОДОМ ПОЗЖЕ даты обязательства, хотя правило
  движка с 08.2026 обратное. Число платежей сохранялось, поэтому расходилось не
  тело, а проценты: 317,9 млн ₽ против 269,1 у движка;
* строка 15 CAPEX несла весь денежный платёж по ВРИ, а база резерва — это
  `SUM(B15:B29)`: книга брала 5% и с процентов. Движок берёт резерв раньше, чем
  проценты становятся статьёй, — это ещё 15,9 млн.

Запуск: python3 -m pytest tests/test_the_vri_installment_matches_the_engine.py -q
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "tests"))

import main_legacy as core  # noqa: E402
from test_book_interest_horizon_follows_the_engine import BASE, tep_of_a_real_project  # noqa: E402

openpyxl = pytest.importorskip("openpyxl")

# Рассрочка на три года поквартально с процентами — тот же режим, что у владельца.
INSTALLMENT = {
    "land_rights_cost_mln": 1630.486,
    "vri_required": True,
    "vri_payment_mode": "installment",
    "vri_installment_years": 3,
    "vri_periodicity_months": 3,
    "vri_interest_enabled": "1",
    "vri_interest_spread_pp": 3.0,
    "vri_obligation_date_mode": "at_rns",
}


def _book(**overrides):
    content, _, meta = core.build_project_workbook(
        {**BASE, **INSTALLMENT, **overrides}, tep_of_a_real_project(), [], {}, project_name="П")
    return openpyxl.load_workbook(io.BytesIO(content)), meta["missing"]


def _evaluator(wb):
    from xlsx_eval import Evaluator

    sys.setrecursionlimit(400000)
    return Evaluator(wb)


def test_the_first_instalment_falls_on_the_obligation_date() -> None:
    """Первый взнос — в месяц обязательства, и процентов на нём ещё нет."""
    wb, missing = _book()
    assert not [m for m in missing if "ВРИ" in m], missing
    ev = _evaluator(wb)
    ws = wb["ВРИ"]
    from openpyxl.utils import get_column_letter as col

    obligation = ev.cell("ВРИ", "B7")
    months = [i for i in range(120)
              if float(ev.cell("ВРИ", f"{col(4 + i)}12") or 0) > 0]
    assert months, "график рассрочки пуст"
    first = ev.cell("ВРИ", f"{col(4 + months[0])}3")
    assert first == obligation, (first, obligation)
    # На первом платеже процентов нет: копить их ещё не на чем.
    assert float(ev.cell("ВРИ", f"{col(4 + months[0])}13") or 0) == 0
    # Платежей ровно двенадцать — три года поквартально.
    assert len(months) == 12, months
    assert ws["A12"].value == "Погашение основного долга"


def test_the_interest_matches_the_engine() -> None:
    """Проценты книги совпадают с движком: одиннадцать начислений, не двенадцать."""
    wb, _ = _book()
    ev = _evaluator(wb)
    engine = core.calculate(core.CalcRequest(
        inputs={**BASE, **INSTALLMENT}, tep=tep_of_a_real_project(), rates=[]))
    theirs = float(engine["vri"]["totals"]["interest"]) / 1e6
    ours = float(ev.cell("ВРИ", "B13"))
    assert abs(ours - theirs) <= max(1.0, theirs * 0.005), (ours, theirs)
    # Тело обязательства книга гасит целиком.
    assert abs(float(ev.cell("ВРИ", "B12")) - INSTALLMENT["land_rights_cost_mln"]) < 0.01


def test_the_reserve_does_not_grow_on_the_interest() -> None:
    """Проценты стоят своей строкой и в базу резерва не входят."""
    wb, _ = _book()
    ev = _evaluator(wb)
    principal = float(ev.cell("ВРИ", "B12"))
    interest = float(ev.cell("ВРИ", "B13")) + float(ev.cell("ВРИ", "B14"))
    assert wb["CAPEX"]["A38"].value == "Проценты и обеспечение по рассрочке ВРИ"
    assert abs(float(ev.cell("CAPEX", "B15")) - principal) < 0.01
    assert abs(float(ev.cell("CAPEX", "B38")) - interest) < 0.01
    # База резерва — строки 15..29 плюс 31: строки 38 в ней нет по построению.
    formula = wb["CAPEX"]["B30"].value
    assert "B38" not in formula, formula
    engine = core.calculate(core.CalcRequest(
        inputs={**BASE, **INSTALLMENT}, tep=tep_of_a_real_project(), rates=[]))
    reserve = float(engine["capex"]["reserve"]) / 1e6
    assert abs(float(ev.cell("CAPEX", "B30")) - reserve) <= max(1.0, reserve * 0.01)


def test_the_book_keeps_parity_on_an_installment_project() -> None:
    """Те строки ПРОВЕРОК, что расходились у владельца, сходятся."""
    wb, _ = _book()
    ev = _evaluator(wb)
    ws = wb["ПРОВЕРКИ"]
    checked = 0
    for row in range(60, 100):
        name = str(ws.cell(row, 1).value or "")
        if "аритет" not in name:
            continue
        fact, expected, tolerance = (ev.cell("ПРОВЕРКИ", f"B{row}"),
                                     ws.cell(row, 3).value, ws.cell(row, 5).value)
        if expected in (None, "") or tolerance in (None, ""):
            continue
        assert abs(float(fact) - float(expected)) <= float(tolerance), (name, fact, expected)
        checked += 1
    assert checked >= 6
