"""Единая книга проекта: DevelopAid v4, заполненная текущими вводными.

На сайте жили две выгрузки — ZIP с детализацией значениями и ZIP с шаблоном
ПЛАТО, — и рядом они читались как разные модели одного проекта. Теперь кнопка
одна и отдаёт книгу v4: весь расчёт — живые формулы, сюда пишутся только
значения листа «Вводные». Книга несёт диаграммы, которые openpyxl при
перезаписи теряет, поэтому значения правятся прямо в XML внутри zip.

Запуск: python3 -m pytest tests -q
"""

from __future__ import annotations

import io
import sys
import zipfile
from datetime import datetime
from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import main as wrapper  # noqa: E402

core = wrapper.core
client = TestClient(wrapper.app)


def build(inputs=None, tep=None, phasing=None, name="Тест"):
    x = {**core.DEFAULT_INPUTS, **(inputs or {})}
    t = {k: dict(v) for k, v in core.TEP_DEFAULT.items()}
    for key, row in (tep or {}).items():
        t.setdefault(key, {}).update(row)
    return core.build_project_workbook(x, t, [], phasing or {}, project_name=name)


@pytest.fixture(scope="module")
def default_book():
    content, filename, meta = build()
    sheet = openpyxl.load_workbook(io.BytesIO(content), data_only=False)["Вводные"]
    return content, filename, meta, sheet


# --- сборка ----------------------------------------------------------------

def test_every_input_finds_its_cell(default_book):
    """Поле без соответствия — потерянные данные, а не мелочь."""
    _, _, meta, _ = default_book
    assert meta["missing"] == []


def test_the_charts_survive_the_fill(default_book):
    """Ради диаграмм книга и правится в XML, а не через openpyxl."""
    content, _, _, _ = default_book
    names = zipfile.ZipFile(io.BytesIO(content)).namelist()
    assert any("charts/chart" in name for name in names), "диаграммы потеряны"


def test_the_filename_is_a_single_xlsx_not_an_archive(default_book):
    _, filename, _, _ = default_book
    assert filename.endswith(".xlsx")


# --- значения --------------------------------------------------------------

def test_percent_points_become_fractions(default_book):
    """Движок хранит 25, книга ждёт 0,25 — иначе налог посчитается в 2500%."""
    _, _, _, sheet = default_book
    assert sheet["B21"].value == pytest.approx(0.25)   # налог на прибыль
    assert sheet["B63"].value == pytest.approx(0.85)   # доля продаж до РВЭ
    assert sheet["B25"].value == pytest.approx(0.06)   # спред БРИДЖ
    assert sheet["B33"].value == pytest.approx(0.14)   # ключевая на старте


def test_the_tep_lands_in_the_first_queue(default_book):
    _, _, _, sheet = default_book
    assert sheet["B88"].value == "Да"
    assert sheet["B89"].value == "Нет"
    assert sheet["W88"].value == pytest.approx(core.TEP_DEFAULT["apartments"]["gns"], rel=1e-6)
    assert sheet["Z88"].value == pytest.approx(core.TEP_DEFAULT["apartments"]["saleable"], rel=1e-6)
    assert sheet["AB88"].value == pytest.approx(
        core.TEP_DEFAULT["underground_parking"]["units"], rel=1e-6)


def test_the_project_start_drives_the_horizon(default_book):
    _, _, _, sheet = default_book
    assert sheet["B8"].value == datetime(2027, 1, 1)
    assert sheet["D88"].value == datetime(2027, 1, 1)


def test_the_formulas_of_the_book_are_not_overwritten(default_book):
    """Пишутся только значения: срок продаж и целевая ставка остаются формулами."""
    _, _, _, sheet = default_book
    assert str(sheet["H88"].value).startswith("=")
    assert str(sheet["B34"].value).startswith("=")


def test_the_bridge_limit_is_computed_not_hardcoded(default_book):
    """Фиксированный лимит 3 500 млн резал выборку БРИДЖ на крупном проекте."""
    _, _, _, sheet = default_book
    formula = str(sheet["B24"].value)
    assert formula.startswith("=")
    assert "CAPEX" in formula


def test_the_density_factor_is_one_at_export(default_book):
    """Базовый потенциал равен применяемому: ТЭП уезжает ровно как в проекте."""
    _, _, _, sheet = default_book
    assert sheet["K8"].value == "Нет"
    assert sheet["K11"].value == pytest.approx(30000)


def test_the_vri_switch_follows_the_payment_not_the_flag():
    """Выключатель ВРИ — нулевая плата: заданная сумма платится всегда."""
    _, _, _, sheet = _book({"vri_required": False, "land_rights_cost_mln": 500})
    assert sheet["B74"].value == "Да"
    _, _, _, sheet = _book({"vri_required": True, "land_rights_cost_mln": 0})
    assert sheet["B74"].value == "Нет"


def _book(inputs):
    content, filename, meta = build(inputs)
    return content, filename, meta, openpyxl.load_workbook(
        io.BytesIO(content), data_only=False)["Вводные"]


def test_an_installment_is_written_in_the_words_of_the_book():
    _, _, _, sheet = _book({"vri_payment_mode": "installment", "vri_installment_years": 3})
    assert sheet["B75"].value == "3 года"


def test_the_office_block_carries_dates_and_terms():
    _, _, _, sheet = _book({"offices_enabled": True, "offices_gba_sqm": 21700,
                            "offices_months": 24, "offices_residual_months": 6})
    assert sheet["K20"].value == "Да"
    assert sheet["K23"].value == pytest.approx(21700)
    assert sheet["K33"].value == pytest.approx(30)
    assert sheet["K27"].value == datetime(2028, 7, 1)


# --- очереди ---------------------------------------------------------------

def test_the_phases_split_the_tep_by_their_weights():
    phasing = {
        "enabled": True, "phase_count": 2,
        "phases": [{"start_offset_months": 0, "construction_months": 24},
                   {"start_offset_months": 12, "construction_months": 30}],
        "products": {"apartments": [0.6, 0.4], "ground_commercial": [0.5, 0.5],
                     "underground_parking": [0.7, 0.3], "storage": [1, 0]},
        "shared_cash": {"purchase": [1, 0], "land_rights": [0.5, 0.5],
                        "social_compensation": [0, 1]},
        "cost_inflation_pct": 8, "sales_price_inflation_pct": 8,
    }
    content, _, meta = build(phasing=phasing)
    sheet = openpyxl.load_workbook(io.BytesIO(content), data_only=False)["Вводные"]
    total = core.TEP_DEFAULT["apartments"]["gns"]
    assert meta["phased"] is True
    assert sheet["B88"].value == "Да" and sheet["B89"].value == "Да"
    assert sheet["B90"].value == "Нет"
    assert sheet["W88"].value == pytest.approx(total * 0.6, rel=1e-6)
    assert sheet["W89"].value == pytest.approx(total * 0.4, rel=1e-6)
    assert sheet["D89"].value == datetime(2028, 1, 1)
    assert sheet["F89"].value == pytest.approx(30)
    assert sheet["Q89"].value == pytest.approx(0.5)
    # Индексация очереди — множителем цены к сдвигу старта, как в движке.
    assert sheet["S89"].value == pytest.approx(1.08)
    assert sheet["AE89"].value == pytest.approx(0.0)


def test_percent_weights_are_normalized_to_shares():
    """Страница хранит веса очередей в процентах ([40, 32, 28]): записанные
    как есть, они раздували ГНС очереди в сорок раз, а долю покупки — в сто."""
    phasing = {
        "enabled": True, "phase_count": 3,
        "phases": [{"start_offset_months": 12 * i, "construction_months": 24}
                   for i in range(3)],
        "products": {"apartments": [40, 32, 28], "ground_commercial": [40, 32, 28],
                     "underground_parking": [40, 32, 28], "storage": [40, 32, 28]},
        "shared_cash": {"purchase": [100, 0, 0], "land_rights": [100, 0, 0],
                        "social_compensation": [100, 0, 0]},
        "cost_inflation_pct": 8, "sales_price_inflation_pct": 8,
    }
    content, _, _ = build(phasing=phasing)
    sheet = openpyxl.load_workbook(io.BytesIO(content), data_only=False)["Вводные"]
    total = core.TEP_DEFAULT["apartments"]["gns"]

    assert sheet["W88"].value == pytest.approx(total * 0.40, rel=1e-6)
    assert sheet["W90"].value == pytest.approx(total * 0.28, rel=1e-6)
    assert sheet["P88"].value == pytest.approx(1.0)
    assert sheet["U88"].value == pytest.approx(0.40)


def test_social_construction_is_not_lost_by_the_book():
    """Книга знает один канал соцнагрузки — компенсацию. Строительство садов
    и школ (у Мытищ — миллиарды) иначе выпадало из расходов, завышая LLCR."""
    _, _, _, sheet = _book({
        "social_mode": "Строительство", "social_compensation_mln": 100,
        "kindergarten_places": 465, "kindergarten_cost_mln_per_place": 2.75,
        "school_places": 975, "school_cost_mln_per_place": 3,
        "clinic_capacity": 0,
    })
    # Компенсация в режиме «Строительство» не платится: движок считает
    # social_total только из строительства, и книга обязана совпадать.
    assert sheet["B17"].value == pytest.approx(465 * 2.75 + 975 * 3)


def test_social_construction_is_indexed_to_its_queue():
    """Движок строит соцобъекты в своих очередях с инфляцией затрат: сады во
    второй (×1,08), школа в третьей (×1,166). Базовая сумма занижала
    социалку книги на 12% против движка."""
    phasing = {
        "enabled": True, "phase_count": 3,
        "phases": [{"start_offset_months": 12 * i, "construction_months": 24}
                   for i in range(3)],
        "products": {k: [40, 32, 28] for k in ("apartments", "ground_commercial",
                                               "underground_parking", "storage")},
        "shared_allocation": {}, "cost_inflation_pct": 8,
        "sales_price_inflation_pct": 8,
    }
    content, _, _ = build(
        {"social_mode": "Строительство", "social_compensation_mln": 0,
         "kindergarten_places": 453, "kindergarten_cost_mln_per_place": 2.75,
         "school_places": 950, "school_cost_mln_per_place": 3,
         "clinic_capacity": 124, "clinic_cost_mln_per_unit": 3},
        phasing=phasing)
    sheet = openpyxl.load_workbook(io.BytesIO(content), data_only=False)["Вводные"]

    expected = (453 * 2.75 * 1.08 + 950 * 3 * 1.08 ** 2 + 124 * 3 * 1.08)
    assert sheet["B17"].value == pytest.approx(expected, rel=1e-6)


def test_a_pure_compensation_stays_as_entered():
    _, _, _, sheet = _book({"social_mode": "Денежная компенсация",
                            "social_compensation_mln": 575.379,
                            "kindergarten_places": 465})
    assert sheet["B17"].value == pytest.approx(575.379)


def test_stale_formula_caches_are_stripped(default_book):
    """Просмотрщики формулы не считают: кэшированные результаты шаблона
    показывали числа проекта, под который книга собиралась в прошлый раз."""
    import re as _re
    content, _, _, _ = default_book
    archive = zipfile.ZipFile(io.BytesIO(content))
    stale = 0
    for name in archive.namelist():
        if name.startswith("xl/worksheets/") and name.endswith(".xml"):
            text = archive.read(name).decode("utf-8")
            stale += len(_re.findall(r"</x:f><x:v>", text))
            stale += len(_re.findall(r"<x:f[^>]*/><x:v>", text))
    assert stale == 0, f"в книге осталось {stale} кэшированных результатов формул"


# --- книга считает так же, как движок ---------------------------------------

def test_the_book_passes_its_own_checks_and_matches_the_engine():
    """Книга вся из формул, и ошибка в любой уедет к аналитику молча. Поэтому
    заполненная книга пересчитывается своим вычислителем: собственные проверки
    книги не дают FAIL, а выручка и LLCR сходятся с движком. Так был пойман
    двойной счёт офисов на листе ТЭП и индексация цены первой очереди."""
    import sys
    sys.setrecursionlimit(300000)
    from xlsx_eval import Evaluator

    # Цена выше умолчания: паритет проверяется на проекте, который гасит долг.
    # На слабом проекте книга честно даёт FAIL «финальный долг не ноль» — это
    # её бизнес-чек, а не расхождение с движком.
    inputs = {**core.DEFAULT_INPUTS, "apartment_price_th": 500, "commercial_price_th": 500}
    tep = {k: dict(v) for k, v in core.TEP_DEFAULT.items()}
    phasing = {
        "enabled": True, "phase_count": 2,
        "phases": [{"start_offset_months": 0, "construction_months": 24},
                   {"start_offset_months": 12, "construction_months": 24}],
        "products": {k: [55, 45] for k in ("apartments", "ground_commercial",
                                           "underground_parking", "storage")},
        "shared_allocation": {"purchase": [100, 0], "land_rights": [60, 40],
                              "social_compensation": [100, 0]},
        "cost_inflation_pct": 8, "sales_price_inflation_pct": 8,
    }
    content, _, meta = core.build_project_workbook(
        inputs, tep, [], phasing, project_name="Паритет")
    assert meta["missing"] == []

    book = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
    evaluator = Evaluator(book)
    status = evaluator.cell("ПРОВЕРКИ", "B3")
    if status not in ("PASS", "PASS WITH WARNINGS"):
        broken = [
            f'{book["ПРОВЕРКИ"][f"A{row}"].value}: '
            f'факт={evaluator.cell("ПРОВЕРКИ", f"B{row}")} '
            f'ожид={evaluator.cell("ПРОВЕРКИ", f"C{row}")}'
            for row in range(6, 62)
            if book["ПРОВЕРКИ"][f"A{row}"].value
            and evaluator.cell("ПРОВЕРКИ", f"F{row}") == "FAIL"
        ]
        raise AssertionError(f"книга не проходит свои проверки: {broken}")

    engine = core.calculate_phased(core.PhasedCalcRequest(
        inputs=inputs, tep=tep, rates=[], phasing=phasing))
    summary = engine["consolidated"]["summary"]
    book_revenue = float(evaluator.cell("CF", "B6"))
    book_llcr = float(evaluator.cell("CF", "B30"))

    assert book_revenue == pytest.approx(summary["revenue"] / 1e6, rel=0.02), \
        "выручка книги разошлась с движком больше чем на 2%"
    assert book_llcr == pytest.approx(summary["llcr"], rel=0.05), \
        "LLCR книги разошёлся с движком больше чем на 5%"


def test_the_tep_sheet_does_not_double_count_the_objects():
    """«ИТОГО ЖИЛЫЕ ОЧЕРЕДИ» ссылался на CF!B6, который уже включает офисы,
    ТЦ и наземный паркинг, — и лист ТЭП считал объекты дважды."""
    template = openpyxl.load_workbook(core._V4_TEMPLATE_PATH, data_only=False)
    assert str(template["ТЭП"]["G28"].value).replace(" ", "") == "=SUM(G8,G14,G20,G26)"


def test_a_fourth_queue_gets_its_own_cf_sheet():
    """Книга несёт четыре очереди: CF_4, строка 91 на «Вводных», свой блок
    продаж — четырёхочередной проект больше не режется."""
    phasing = {
        "enabled": True, "phase_count": 4,
        "phases": [{"start_offset_months": 12 * i, "construction_months": 24}
                   for i in range(4)],
        "products": {k: [32, 26, 22, 20] for k in ("apartments", "ground_commercial",
                                                   "underground_parking", "storage")},
        "shared_allocation": {"purchase": [100, 0, 0, 0], "land_rights": [40, 30, 20, 10],
                              "social_compensation": [100, 0, 0, 0]},
        "cost_inflation_pct": 8, "sales_price_inflation_pct": 8,
    }
    content, _, meta = build(phasing=phasing)
    book = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
    sheet = book["Вводные"]
    total = core.TEP_DEFAULT["apartments"]["gns"]

    assert meta["missing"] == []
    assert "CF_4" in book.sheetnames
    assert sheet["B91"].value == "Да"
    assert sheet["W91"].value == pytest.approx(total * 0.20, rel=1e-6)
    assert sheet["S91"].value == pytest.approx(1.08 ** 3, rel=1e-6)
    assert book["КОНСОЛИДАТОР"]["A8"].value == "КОНСОЛИДИРОВАНО"


def test_objects_are_not_double_counted_by_the_idle_fourth_queue():
    """Аллокация объектов очереди 4 — клон третьей: признак очереди в ней
    хранится в ОБЪЕКТЫ!$B$8/$B$36/$B$64, и не заменённая тройка дублировала
    офисы третьей очереди в спящий CF_4 — плюс 15 млрд выручки из воздуха."""
    import sys
    sys.setrecursionlimit(300000)
    from xlsx_eval import Evaluator

    inputs = {**core.DEFAULT_INPUTS, "offices_enabled": True}
    tep = {k: dict(v) for k, v in core.TEP_DEFAULT.items()}
    phasing = {
        "enabled": True, "phase_count": 3,
        "phases": [{"start_offset_months": 12 * i, "construction_months": 24}
                   for i in range(3)],
        "products": {k: [40, 32, 28] for k in ("apartments", "ground_commercial",
                                               "underground_parking", "storage")},
        "shared_allocation": {},
        "discrete": {"offices": 3, "standalone_retail": 2, "above_parking": 2},
        "cost_inflation_pct": 8, "sales_price_inflation_pct": 8,
    }
    content, _, _ = core.build_project_workbook(inputs, tep, [], phasing, project_name="Т")
    book = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
    evaluator = Evaluator(book)

    offices = float(evaluator.cell("ОБЪЕКТЫ", "B24"))
    assert offices > 0, "офисы включены, но выручки нет"
    assert float(evaluator.cell("ОБЪЕКТЫ", "B116")) == pytest.approx(0.0, abs=0.01), \
        "аллокация спящей четвёртой очереди дублирует объекты третьей"
    assert float(evaluator.cell("CF_4", "B10")) == pytest.approx(0.0, abs=0.01), \
        "выключенная четвёртая очередь несёт выручку"
    # сквозной чек книги: выручка продуктов равна консолидированному CF
    assert evaluator.cell("ПРОВЕРКИ", "F48") == "OK"


def test_a_fifth_queue_is_folded_into_the_fourth():
    """Пятая очередь в книгу не помещается — сливается в четвёртую с
    предупреждением, а не теряется и не размазывается."""
    phasing = {
        "enabled": True, "phase_count": 5,
        "phases": [{"start_offset_months": 12 * i, "construction_months": 24}
                   for i in range(5)],
        "products": {k: [28, 22, 19, 16, 15] for k in ("apartments", "ground_commercial",
                                                       "underground_parking", "storage")},
        "shared_allocation": {},
        "cost_inflation_pct": 8, "sales_price_inflation_pct": 8,
    }
    content, _, meta = build(phasing=phasing)
    sheet = openpyxl.load_workbook(io.BytesIO(content), data_only=False)["Вводные"]
    total = core.TEP_DEFAULT["apartments"]["gns"]

    assert any("слиты в четвёртую" in str(item) for item in meta["missing"])
    assert sheet["W91"].value == pytest.approx(total * 0.31, rel=1e-6), \
        "объём пятой очереди не слит в четвёртую"


def test_the_objects_inherit_their_queue_from_the_phasing():
    """Шаблонная «очередь 1» сажала офисы в первую очередь, а движок ведёт их
    в третьей: продаваемая площадь очередей расходилась с отчётом."""
    phasing = {
        "enabled": True, "phase_count": 3,
        "phases": [{"start_offset_months": 12 * i, "construction_months": 24}
                   for i in range(3)],
        "products": {k: [40, 32, 28] for k in ("apartments", "ground_commercial",
                                               "underground_parking", "storage")},
        "shared_allocation": {},
        "discrete": {"offices": 3, "standalone_retail": 2, "above_parking": 2},
        "cost_inflation_pct": 8, "sales_price_inflation_pct": 8,
    }
    content, _, _ = build(phasing=phasing)
    sheet = openpyxl.load_workbook(io.BytesIO(content), data_only=False)["Вводные"]

    assert sheet["K21"].value == pytest.approx(3)   # офисы
    assert sheet["K41"].value == pytest.approx(2)   # ТЦ
    assert sheet["K61"].value == pytest.approx(2)   # наземный паркинг
    # Даты объекта сдвигаются на сдвиг старта его очереди, как в движке:
    # сырая мастер-дата строила офисы третьей очереди на два года раньше,
    # CAPEX падал в БРИДЖ до РнС, и пик завышался почти вдвое.
    assert sheet["K27"].value == datetime(2030, 7, 1)   # офисы: 2028-07 + 24 мес
    assert sheet["K47"].value == datetime(2029, 7, 1)   # ТЦ: 2028-07 + 12 мес

    single, _, _ = build()  # без очередей всё в первой
    sheet_single = openpyxl.load_workbook(io.BytesIO(single), data_only=False)["Вводные"]
    assert sheet_single["K21"].value == pytest.approx(1)


def test_the_report_carries_a_pdf_comparable_unit_revenue():
    """«Средняя цена реализации» книги — площадные продукты без паркинга, а
    PDF делит всю выручку очереди: цифры расходились на цену паркинга и
    читались как ошибка. Строка-мостик считает по определению PDF."""
    template = openpyxl.load_workbook(core._V4_TEMPLATE_PATH, data_only=False)["ОТЧЕТ"]

    assert "как в PDF" in str(template["A73"].value)
    for cell, consolidator in (("B73", "G4"), ("C73", "G5"), ("D73", "G6"),
                               ("E73", "G7"), ("F73", "G8")):
        formula = str(template[cell].value)
        assert f"'КОНСОЛИДАТОР'!{consolidator}" in formula, f"{cell} не делит выручку консолидатора"


def test_both_surfaces_carry_the_pure_apartment_unit_price():
    """Смешанные периметры удельных («вся выручка» против «площадных») не
    сверить глазами. Чисто квартирная цена на м² продаваемой — общий
    знаменатель: колонка «в т.ч. квартиры» в PDF и строка 95 книги."""
    import inspect
    module = inspect.getsource(core)
    assert "в т.ч. квартиры" in module, "в PDF нет квартирной колонки"

    phasing = {
        "enabled": True, "phase_count": 2,
        "phases": [{"start_offset_months": 0, "construction_months": 24},
                   {"start_offset_months": 12, "construction_months": 24}],
        "products": {k: [55, 45] for k in ("apartments", "ground_commercial",
                                           "underground_parking", "storage")},
        "shared_allocation": {}, "cost_inflation_pct": 8,
        "sales_price_inflation_pct": 8,
    }
    bundle = core.calculate_phased(core.PhasedCalcRequest(
        inputs=dict(core.DEFAULT_INPUTS),
        tep={k: dict(v) for k, v in core.TEP_DEFAULT.items()},
        rates=[], phasing=phasing))
    row = (bundle.get("comparison") or [{}])[0]
    assert row.get("apartment_price_th"), "сравнение очередей не несёт цену квартир"
    assert row.get("apartment_saleable_sqm"), "нет квартирной продаваемой для итога"

    # Строка стоит внутри таблицы удельных — сразу после «Средней цены
    # реализации», как колонка в PDF, — а не отдельным блоком под отчётом.
    template = openpyxl.load_workbook(core._V4_TEMPLATE_PATH, data_only=False)["ОТЧЕТ"]
    assert template["A60"].value == "Средняя цена реализации"
    assert template["A61"].value == "в т.ч. квартиры"
    for cell, sales in (("B61", "B16"), ("C61", "B39"), ("D61", "B62")):
        formula = str(template[cell].value)
        assert f"'Продажи'!{sales}" in formula and "$L$" in formula, \
            f"{cell} книги не делит квартирную выручку на квартирную продаваемую"
    assert template["B61"].number_format == template["B60"].number_format, \
        "квартирная строка без числового формата удельных"
    # Сдвиг не разорвал ссылки соседних блоков: колонка проекта после
    # добавления «Очереди 4» переехала из E в F, и G-блок читает её оттуда.
    assert str(template["G9"].value) == "=F62"
    assert str(template["G21"].value) == "=F70"


def test_the_queue_price_multiplier_carries_the_phase_indexation():
    """Индексация очередей — множителем к сдвигу старта, как в движке:
    книга вела годовую инфляцию от даты базы цен и индексировала даже
    первую очередь — на +12% против движка."""
    phasing = {
        "enabled": True, "phase_count": 2,
        "phases": [{"start_offset_months": 0, "construction_months": 24},
                   {"start_offset_months": 12, "construction_months": 24}],
        "products": {k: [55, 45] for k in ("apartments", "ground_commercial",
                                           "underground_parking", "storage")},
        "shared_allocation": {}, "cost_inflation_pct": 8,
        "sales_price_inflation_pct": 8,
    }
    content, _, _ = build(phasing=phasing)
    sheet = openpyxl.load_workbook(io.BytesIO(content), data_only=False)["Вводные"]

    assert sheet["S88"].value == pytest.approx(1.0)
    assert sheet["S89"].value == pytest.approx(1.08)
    assert sheet["AE88"].value == pytest.approx(0.0)
    assert sheet["AE89"].value == pytest.approx(0.0)


# --- поверхности -----------------------------------------------------------

def test_the_endpoint_returns_a_single_workbook():
    response = client.post("/report/workbook", json={
        "inputs": dict(core.DEFAULT_INPUTS),
        "tep": {k: dict(v) for k, v in core.TEP_DEFAULT.items()},
        "rates": [], "phasing": {}, "project_name": "Ручной", "scenario": "base",
    })
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    assert response.content[:2] == b"PK"


def test_the_page_offers_one_model_download():
    """Зачем и шаблон, и зип? Кнопка одна, и она отдаёт книгу v4."""
    assert "Скачать модель (Excel)" in core.PAGE
    assert "/report/workbook" in core.PAGE
    assert "Выгрузить в шаблон ПЛАТО" not in core.PAGE
    assert "exportPlatoTemplate" not in core.PAGE


def test_the_bot_attachment_is_the_same_workbook():
    """Бот и сайт отдают одну книгу, а не каждый свою."""
    import inspect
    source = inspect.getsource(core._telegram_send_attachments)
    assert "build_project_workbook" in source
    assert "build_model_archive" not in source


# --- четвёртая очередь как драйвер, а не только листы -----------------------

def test_the_fourth_queue_drives_the_whole_book_not_just_its_sheets():
    """О4 существовала листами, но не драйвером: INDEX по очередям упирался в
    строку 90, объект в четвёртой очереди ронял ТЭП и аллокацию, лимиты БРИДЖ
    и ПФ не видели CAPEX четвёртого блока, а проектные агрегаты ОТЧЁТа
    складывали три блока продаж из четырёх."""
    import sys
    sys.setrecursionlimit(400000)
    from xlsx_eval import Evaluator

    inputs = {**core.DEFAULT_INPUTS, "apartment_price_th": 500,
              "commercial_price_th": 500, "offices_enabled": True}
    tep = {k: dict(v) for k, v in core.TEP_DEFAULT.items()}
    phasing = {
        "enabled": True, "phase_count": 4,
        "phases": [{"start_offset_months": 12 * i, "construction_months": 24}
                   for i in range(4)],
        "products": {k: [32, 26, 22, 20] for k in ("apartments", "ground_commercial",
                                                   "underground_parking", "storage")},
        "shared_allocation": {"purchase": [100, 0, 0, 0], "land_rights": [40, 30, 20, 10],
                              "social_compensation": [100, 0, 0, 0]},
        "discrete": {"offices": 4, "standalone_retail": 2, "above_parking": 2},
        "cost_inflation_pct": 8, "sales_price_inflation_pct": 8,
    }
    content, _, meta = core.build_project_workbook(inputs, tep, [], phasing,
                                                   project_name="О4-драйвер")
    assert meta["missing"] == []
    book = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
    evaluator = Evaluator(book)

    assert book["Вводные"]["K21"].value == pytest.approx(4)
    assert evaluator.cell("ПРОВЕРКИ", "B3") in ("PASS", "PASS WITH WARNINGS"), [
        f'{book["ПРОВЕРКИ"][f"A{row}"].value}: '
        f'факт={evaluator.cell("ПРОВЕРКИ", f"B{row}")} '
        f'ожид={evaluator.cell("ПРОВЕРКИ", f"C{row}")}'
        for row in range(6, 76)
        if book["ПРОВЕРКИ"][f"A{row}"].value
        and evaluator.cell("ПРОВЕРКИ", f"F{row}") == "FAIL"
    ]
    assert float(evaluator.cell("ОБЪЕКТЫ", "B116")) > 0, \
        "аллокация офисов четвёртой очереди пуста"

    engine = core.calculate_phased(core.PhasedCalcRequest(
        inputs=inputs, tep=tep, rates=[], phasing=phasing))
    summary = engine["consolidated"]["summary"]
    assert float(evaluator.cell("CF", "B6")) == pytest.approx(
        summary["revenue"] / 1e6, rel=0.02)
    assert float(evaluator.cell("ОТЧЕТ", "F61")) == pytest.approx(
        summary["average_apartment_price_th"], rel=0.02), \
        "«в т.ч. квартиры» проекта не сходится с движком"


def test_the_project_aggregates_include_the_fourth_block():
    """Формулы-агрегаты, где четвёртый блок терялся: лимиты, ОТЧЁТ, чеки."""
    template = openpyxl.load_workbook(core._V4_TEMPLATE_PATH, data_only=False)
    flat = lambda cell: str(cell.value).replace(" ", "")

    # витринный ID четвёртой очереди и признак ТЦ в консолидации были «3»:
    # ТЦ третьей очереди попадал в площадь четвёртой
    assert template["Вводные"]["A91"].value == 4
    assert "$K$41=4" in flat(template["КОНСОЛИДАТОР"]["F7"])
    assert "$K$41=3" not in flat(template["КОНСОЛИДАТОР"]["F7"])
    assert "H91" in flat(template["ПРОВЕРКИ"]["B57"])

    # контур финансирования: суммы долей лимитов, даты объектов против старта
    # их очереди, пик ПФ в пределах лимита с рефинансируемым БРИДЖем
    assert "U88:U91" in flat(template["ПРОВЕРКИ"]["B69"])
    assert "V88:V91" in flat(template["ПРОВЕРКИ"]["B70"])
    assert "$K$27" in flat(template["ПРОВЕРКИ"]["B71"])
    assert "$D$88:$D$91" in flat(template["ПРОВЕРКИ"]["B71"])
    assert "'CF_4'!B83" in flat(template["ПРОВЕРКИ"]["B72"])
    assert "$B$24" in flat(template["ПРОВЕРКИ"]["B72"])
    assert "F6:F75" in flat(template["ПРОВЕРКИ"]["B3"])

    assert "'CAPEX'!$B$137" in flat(template["Вводные"]["B26"])
    assert "'Продажи'!B85" in flat(template["ОТЧЕТ"]["F61"])
    assert "$L$88:$L$91" in flat(template["ОТЧЕТ"]["F61"])
    assert "'Продажи'!B85" in flat(template["ОТЧЕТ"]["E46"])
    assert "'ОБЪЕКТЫ'!B116" in flat(template["ПРОВЕРКИ"]["B50"])
    assert "'ОБЪЕКТЫ'!B120" in flat(template["ПРОВЕРКИ"]["B51"])
    assert "$D$83:$DS$83" in flat(template["ОТЧЕТ"]["B87"]), \
        "темп квартир проекта не видит четвёртый блок продаж"

    # выпадающие списки: объект можно поставить в четвёртую очередь,
    # переключатель и ограничения тренда действуют и на строку 91
    validations = {}
    for dv in template["Вводные"].data_validations.dataValidation:
        validations[str(dv.sqref)] = dv.formula1
    assert validations.get("K21") == '"1,2,3,4"'
    assert validations.get("K41") == '"1,2,3,4"'
    assert validations.get("K61") == '"1,2,3,4"'
    assert "B88:B91" in validations
    assert "AD88:AF91" in validations

    # удельный блок ОТЧЁТа: колонка «Проект» делила итоги на данные О4
    # (артефакт сдвига колонок E→F), колонка О3 несла расходы и долг О4
    # (артефакт дубль-пасса CF_3→CF_4), GBA очереди захватывала соседку,
    # а «объёмы» продаж О3/О4 включали строку цены двухстрочным диапазоном
    assert flat(template["ОТЧЕТ"]["F60"]) == "=IFERROR(F59*1000/F57,0)"
    assert "/F58" in flat(template["ОТЧЕТ"]["F62"])
    assert "/F57" in flat(template["ОТЧЕТ"]["F63"])
    assert "($B$5-F59))*1000/F57" in flat(template["ОТЧЕТ"]["F64"])
    assert "/F57" in flat(template["ОТЧЕТ"]["F73"])
    for cell in ("D63", "D64", "D71", "D72"):
        assert "CF_4" not in flat(template["ОТЧЕТ"][cell]), \
            f"колонка О3 ({cell}) несёт данные четвёртой очереди"
    assert "'CF_4'!$D$38:$DS$38,'CF_4'!$D$39:$DS$39" in flat(template["ОТЧЕТ"]["F71"])
    assert "$I$88:$K$88" in flat(template["ОТЧЕТ"]["B58"])
    assert "$I$89:$K$89" in flat(template["ОТЧЕТ"]["C58"])
    assert "$I$90:$K$90" in flat(template["ОТЧЕТ"]["D58"])
    for cell in ("D65", "D68", "E65", "E68", "F65", "F68"):
        formula = flat(template["ОТЧЕТ"][cell])
        for two_rows in ("$D$63:$DS$64", "$D$66:$DS$67", "$D$69:$DS$70",
                         "$D$86:$DS$87", "$D$89:$DS$90", "$D$92:$DS$93"):
            assert two_rows not in formula, \
                f"{cell}: в объём продаж попадает строка цены ({two_rows})"

    # модель без единой включённой очереди — FAIL, а не нулевые даты
    assert "COUNTIF('Вводные'!B88:B91" in flat(template["ПРОВЕРКИ"]["B73"])

    # сводные сроки не смотрят на выключенные очереди
    assert flat(template["КОНСОЛИДАТОР"]["C7"]).startswith(
        "=IF('Вводные'!$B$91=\"Да\"")
    assert flat(template["КОНСОЛИДАТОР"]["E4"]).startswith(
        "=IF('Вводные'!$B$88=\"Да\"")

    # INDEX по очередям обязан дотягиваться до строки 91, а кэш шаблона —
    # отсутствовать: смесь старых значений и пустых клонов О4 выглядела как
    # посчитанная книга для всего, что читает значения без пересчёта
    import re as _re, zipfile as _zip
    with _zip.ZipFile(core._V4_TEMPLATE_PATH) as z:
        for name in z.namelist():
            if not name.startswith("xl/worksheets/"):
                continue
            text = z.read(name).decode()
            leftovers = _re.findall(
                r"INDEX\('Вводные'!\$[A-Z]{1,2}\$88:\$[A-Z]{1,2}\$90,", text)
            assert not leftovers, f"{name}: INDEX всё ещё упирается в строку 90"
            cached = len(_re.findall(r"</x:f><x:v>", text)) + \
                len(_re.findall(r"<x:f[^>]*/><x:v>", text))
            assert cached == 0, f"{name}: {cached} кэшированных значений в шаблоне"


def test_the_builder_bridge_limit_sees_the_fourth_queue_capex(default_book):
    _, _, _, sheet = default_book
    assert "'CAPEX'!$B$137" in str(sheet["B24"].value)


def test_the_social_construction_is_spread_like_the_engine():
    """Социалка строительством платилась в книге одним куском за месяц до РнС,
    и пик БРИДЖа выходил на 17% выше движкового: движок строит соцобъекты
    месяцами. Теперь билдер пишет старт и окно (B18/E18), книга размазывает
    сумму равномерно, и пики сходятся."""
    import sys
    sys.setrecursionlimit(400000)
    from xlsx_eval import Evaluator

    inputs = {**core.DEFAULT_INPUTS, "apartment_price_th": 500,
              "commercial_price_th": 500}
    tep = {k: dict(v) for k, v in core.TEP_DEFAULT.items()}
    content, _, _ = core.build_project_workbook(inputs, tep, [], {},
                                                project_name="Бридж")
    book = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
    evaluator = Evaluator(book)

    assert book["Вводные"]["E18"].value == pytest.approx(24)
    social = [evaluator.cell("CAPEX", f"{col}31") or 0 for col in
              (openpyxl.utils.get_column_letter(i) for i in range(4, 130))]
    active = [v for v in social if float(v or 0) > 0]
    assert len(active) == 24, "социалка не размазана по месяцам строительства"

    engine = core.calculate_phased(core.PhasedCalcRequest(
        inputs=inputs, tep=tep, rates=[], phasing={}))
    peak_engine = engine["consolidated"]["finance"]["peak_bridge"] / 1e6
    balances = [float(evaluator.cell("CF_1", f"{col}36") or 0) for col in
                (openpyxl.utils.get_column_letter(i) for i in range(4, 130))]
    assert max(balances) == pytest.approx(peak_engine, rel=0.05), \
        "пик БРИДЖа книги разошёлся с движком больше чем на 5%"


def test_the_rows_of_every_sheet_are_ordered_and_unique(default_book):
    """Excel отказывался открывать книгу: хирургия четвёртой очереди дописала
    строку 91 «Вводных» в конец sheetData, а в СРОКАХ оставила две строки 35.
    openpyxl такое молча терпит, поэтому structura проверяется по XML."""
    import re as _re
    content, _, _, _ = default_book
    with zipfile.ZipFile(io.BytesIO(content)) as z:
        for name in z.namelist():
            if not (name.startswith("xl/worksheets/") and name.endswith(".xml")):
                continue
            rows = [int(m.group(1)) for m in
                    _re.finditer(r'<x:row r="(\d+)"', z.read(name).decode())]
            assert rows == sorted(rows), f"{name}: строки не по порядку"
            assert len(rows) == len(set(rows)), f"{name}: дубли строк"


def test_the_management_follows_the_direct_costs_not_the_calendar():
    """Управление проектом (5% прямых затрат) размазывалось равномерно от
    старта до РВЭ: полный штаб с первого дня клал в БРИДЖ 478 млн при
    реальных тратах до РнС в одно проектирование. Теперь и движок, и книга
    ведут управление по профилю прямых затрат месяца."""
    inputs = {**core.DEFAULT_INPUTS, "apartment_price_th": 500,
              "commercial_price_th": 500}
    tep = {k: dict(v) for k, v in core.TEP_DEFAULT.items()}
    engine = core.calculate_phased(core.PhasedCalcRequest(
        inputs=inputs, tep=tep, rates=[], phasing={}))
    monthly = engine["consolidated"]["monthly"]
    months = monthly["months"]
    rows = engine["consolidated"]["finance"]["rows"]
    permit = next(r["month"] for r in rows if r["pf_draw"] > 0)
    management = next(c for c in monthly["costs"]
                      if "Управление" in str(c.get("label")))
    values = management.get("values") or []
    before = sum(v for m, v in zip(months, values) if m < permit)
    total = sum(values)
    assert total > 0
    assert before / total < 0.15, \
        "управление до РнС всё ещё платится равномерным календарём"

    template = openpyxl.load_workbook(core._V4_TEMPLATE_PATH, data_only=False)
    formula = str(template["CAPEX"]["D28"].value).replace(" ", "")
    assert "SUM(D16:D26)" in formula, "книга не ведёт управление по профилю"


def test_the_shared_costs_are_paid_by_the_cash_schedule():
    """Доли P/Q/R — кассовый график: ВРИ платится перед первым РнС, а не
    тремя кусками перед РнС каждой очереди. Экономическая аллокация движка
    (shared_allocation) на кассу книги больше не влияет."""
    phasing = {
        "enabled": True, "phase_count": 3,
        "phases": [{"start_offset_months": 12 * i, "construction_months": 24}
                   for i in range(3)],
        "products": {k: [40, 32, 28] for k in ("apartments", "ground_commercial",
                                               "underground_parking", "storage")},
        "shared_allocation": {"land_rights": [40, 30, 30]},
        "cost_inflation_pct": 8, "sales_price_inflation_pct": 8,
    }
    content, _, _ = build(phasing=phasing)
    sheet = openpyxl.load_workbook(io.BytesIO(content), data_only=False)["Вводные"]
    assert sheet["Q88"].value == pytest.approx(1.0)
    assert sheet["Q89"].value == pytest.approx(0.0)
    assert sheet["Q90"].value == pytest.approx(0.0)


def test_the_cadastre_77_09_regression_social_is_construction_only():
    """Регресс по 77:09:0004014:13. Билдер складывал денежную компенсацию
    (580,668) со стройкой соцобъектов (193,25 = ДОУ 19×2,75 + СОШ 38×3 +
    поликлиника 9×3): книга несла 773,918 млн социалки против 193,25 у
    движка — минус 648 млн EBITDA, минус 5% LLCR, плюс 11% к пику БРИДЖа.
    Движок в режиме «Строительство» компенсацию не платит вовсе."""
    import sys
    sys.setrecursionlimit(400000)
    from xlsx_eval import Evaluator

    inputs = {**core.DEFAULT_INPUTS,
              "purchase_price_mln": 790.0,
              "land_rights_cost_mln": 1267.539,
              "apartment_price_th": 700.0, "commercial_price_th": 700.0,
              "parking_price_th": 5000.0,
              "main_above_th_per_sqm": 190.0, "main_under_th_per_sqm": 190.0,
              "social_mode": "Строительство",
              "social_compensation_mln": 580.668,
              "kindergarten_places": 19, "kindergarten_cost_mln_per_place": 2.75,
              "school_places": 38, "school_cost_mln_per_place": 3.0,
              "clinic_capacity": 9, "clinic_cost_mln_per_unit": 3.0}
    tep = {k: dict(v) for k, v in core.TEP_DEFAULT.items()}
    tep["apartments"] = {"label": "Квартиры", "gns": 21415.0, "saleable": 13920.0}
    tep["ground_commercial"] = {"label": "Коммерция 1 эт.", "gns": 1367.0,
                                "saleable": 1230.0}
    tep["underground_parking"] = {"label": "Подземный паркинг", "gns": 3185.0,
                                  "units": 91.0}
    tep["storage"] = {"label": "Кладовые", "gns": 0.0, "units": 0.0}

    social = 19 * 2.75 + 38 * 3.0 + 9 * 3.0  # 193.25
    content, _, _ = core.build_project_workbook(inputs, tep, [], {},
                                                project_name="77:09")
    book = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
    assert book["Вводные"]["B17"].value == pytest.approx(social, abs=0.01), \
        "компенсация добавлена поверх строительства соцобъектов"

    engine = core.calculate_phased(core.PhasedCalcRequest(
        inputs=inputs, tep=tep, rates=[], phasing={}))
    summary = engine["consolidated"]["summary"]
    assert summary["social_payment"] / 1e6 == pytest.approx(social, abs=0.01)

    # постатейно: книга и движок на одних вводных
    evaluator = Evaluator(book)
    b = lambda cell, sheet="ОТЧЕТ": float(evaluator.cell(sheet, cell))
    assert b("B31", "CAPEX") == pytest.approx(social, abs=0.01), "социалка CAPEX"
    assert b("B8") == pytest.approx(summary["ebitda"] / 1e6, rel=0.02), "EBITDA"
    assert b("B11") == pytest.approx(summary["profit_tax"] / 1e6, rel=0.05), "налог"
    assert b("B12") == pytest.approx(summary["net_profit"] / 1e6, rel=0.03), "ЧП"
    assert b("B19") == pytest.approx(summary["llcr"], abs=0.01), "LLCR"


def test_the_default_object_queues_match_on_both_surfaces():
    """Без phasing.discrete движок сажал офисы в первую очередь, а билдер
    книги — в третью: на Мытищах 13,5 млрд выручки офисов молча жили в
    разных очередях двух поверхностей. Дефолты единые: офисы — 3, ТЦ и
    паркинг — 2, с кэпом по числу очередей."""
    inputs = {**core.DEFAULT_INPUTS, "offices_enabled": True}
    tep = {k: dict(v) for k, v in core.TEP_DEFAULT.items()}
    phasing = {
        "enabled": True, "phase_count": 3,
        "phases": [{"start_offset_months": 12 * i, "construction_months": 24}
                   for i in range(3)],
        "products": {k: [40, 32, 28] for k in ("apartments", "ground_commercial",
                                               "underground_parking", "storage")},
        "shared_allocation": {}, "cost_inflation_pct": 8,
        "sales_price_inflation_pct": 8,
    }
    engine = core.calculate_phased(core.PhasedCalcRequest(
        inputs=inputs, tep=tep, rates=[], phasing=phasing))
    offices_by_phase = [
        next((p["revenue"] for p in ph["result"]["report"]["products"]
              if "фис" in str(p.get("label"))), 0.0)
        for ph in engine["phases"]
    ]
    assert offices_by_phase[0] == 0 and offices_by_phase[1] == 0
    assert offices_by_phase[2] > 0, "движок без discrete не сажает офисы в третью"

    content, _, _ = core.build_project_workbook(inputs, tep, [], phasing,
                                                project_name="Дефолт очередей")
    sheet = openpyxl.load_workbook(io.BytesIO(content), data_only=False)["Вводные"]
    assert sheet["K21"].value == pytest.approx(3)

    # две очереди: дефолт капится, объект не пропадает
    ph2 = dict(phasing, phase_count=2, phases=phasing["phases"][:2],
               products={k: [55, 45] for k in phasing["products"]})
    engine2 = core.calculate_phased(core.PhasedCalcRequest(
        inputs=inputs, tep=tep, rates=[], phasing=ph2))
    offices2 = [
        next((p["revenue"] for p in ph["result"]["report"]["products"]
              if "фис" in str(p.get("label"))), 0.0)
        for ph in engine2["phases"]
    ]
    assert sum(offices2) > 0, "офисы пропали при дефолтной очереди больше числа фаз"
    assert offices2[-1] > 0


def test_the_social_construction_cash_follows_the_objects_queues():
    """Кассовая доля соцнагрузки «всё в первую очередь» верна для денежной
    компенсации, а стройка платится там, где стоит объект: на Мытищах с
    5,4 млрд соцобъектов первая очередь книги брала весь платёж, уходила
    в дефолт с 2,2 млрд непогашенного ПФ — при гасящем долг движке."""
    import sys
    sys.setrecursionlimit(400000)
    from xlsx_eval import Evaluator

    inputs = {**core.DEFAULT_INPUTS, "apartment_price_th": 500,
              "commercial_price_th": 500, "social_mode": "Строительство",
              "kindergarten_places": 700, "kindergarten_cost_mln_per_place": 2.75,
              "school_places": 825, "school_cost_mln_per_place": 3.0}
    tep = {k: dict(v) for k, v in core.TEP_DEFAULT.items()}
    phasing = {
        "enabled": True, "phase_count": 3,
        "phases": [{"start_offset_months": 12 * i, "construction_months": 24}
                   for i in range(3)],
        "products": {k: [40, 32, 28] for k in ("apartments", "ground_commercial",
                                               "underground_parking", "storage")},
        "shared_allocation": {}, "cost_inflation_pct": 8,
        "sales_price_inflation_pct": 8,
    }
    content, _, _ = core.build_project_workbook(inputs, tep, [], phasing,
                                                project_name="Соцдоли")
    sheet = openpyxl.load_workbook(io.BytesIO(content), data_only=False)["Вводные"]
    shares = [float(sheet[f"R{r}"].value or 0) for r in (88, 89, 90)]
    assert sum(shares) == pytest.approx(1.0, abs=1e-9)
    assert shares[0] < 1.0, "вся социалка строительством по-прежнему в первой очереди"

    engine = core.calculate_phased(core.PhasedCalcRequest(
        inputs=inputs, tep=tep, rates=[], phasing=phasing))
    by_phase = [0.0, 0.0, 0.0]
    for i, ph in enumerate(engine["phases"]):
        by_phase[i] = ph["result"]["summary"]["social_payment"] / 1e6
    total = sum(by_phase)
    for i in range(3):
        assert shares[i] == pytest.approx(by_phase[i] / total, abs=0.02), \
            f"доля соцнагрузки очереди {i + 1} разошлась с движком"
