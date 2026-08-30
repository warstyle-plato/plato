"""Книга учится переносу непогашенного долга между очередями.

Движок переносит долг с очереди, которой раскрытого эскроу не хватило, на ПФ
следующей (владелец, 27–29.08.2026). Книга об этом не знала: с включённым
признаком отчёт и Excel-выгрузка разошлись бы на видимую величину, и оба
выглядели бы достоверно.

Строки 64 и 65 листов CF в шаблоне пусты, поэтому ряд заводится, не сдвигая ни
одной ссылки: вставить строку в занятое место нельзя — поедет всё, что на неё
ссылается. Ровно по этой причине зачёт переданных метров когда-то лёг в ячейку
льготы, а не в свою.

Запуск: python3 -m pytest tests/test_the_workbook_carries_the_debt.py -q
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
sys.path.insert(0, str(Path(__file__).resolve().parent))

import main as _wrapper  # noqa: E402

core = _wrapper.core


def _phasing(carry: bool) -> dict:
    return {
        "enabled": True, "mode": "phased", "user_enabled": True,
        "phase_count": 2, "phase_gap_months": 12,
        "phases": [{"name": "О1", "start_offset_months": 0, "construction_months": 24},
                   {"name": "О2", "start_offset_months": 12, "construction_months": 24}],
        "products": {key: [35, 65] for key in
                     ("apartments", "ground_commercial", "underground_parking", "storage")},
        "shared_cash": {}, "shared_allocation": {}, "social_objects": [],
        "carry_debt_forward": carry,
    }


def _book(carry: bool):
    """Проект, где первая очередь не гасит свой ПФ раскрытым эскроу."""
    inputs = {**core.DEFAULT_INPUTS, "purchase_price_mln": 12000,
              "project_start": "2027-01-01", "ird_months": 12,
              "construction_months": 24, "apartment_price_th": 700}
    tep = {key: dict(row) for key, row in core.TEP_DEFAULT.items()}
    content, _, meta = core.build_project_workbook(
        inputs, tep, [], _phasing(carry), project_name="Перенос долга")
    assert meta["missing"] == [], meta["missing"]
    return openpyxl.load_workbook(io.BytesIO(content), data_only=False)


def _as_date(value):
    """Вычислитель отдаёт даты объектами, а не серийными числами."""
    from datetime import date, datetime
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _evaluator(book):
    sys.setrecursionlimit(400000)
    from xlsx_eval import Evaluator
    return Evaluator(book)


@pytest.fixture(scope="module")
def carried():
    book = _book(True)
    return book, _evaluator(book)


@pytest.fixture(scope="module")
def plain():
    book = _book(False)
    return book, _evaluator(book)


def test_the_workbook_moves_the_debt_to_the_next_queue(carried):
    book, ev = carried
    passed = float(ev.cell("CF_1", "B65"))
    accepted = float(ev.cell("CF_2", "B64"))
    assert passed > 1_000, "предохранитель: на этих вводных О1 обязана не погасить долг"
    assert accepted == pytest.approx(passed), (
        "принято следующей очередью обязано совпасть с переданным до копейки: "
        "иначе обязательство размножилось или пропало")
    assert float(ev.cell("CF_1", "B47")) == pytest.approx(0.0, abs=1e-6), (
        "у передавшей очереди долга больше нет — должник сменился")


def test_without_the_flag_the_workbook_behaves_exactly_as_before(plain):
    """Признак выключен — книга обязана считать ровно как раньше.

    Правка трогает формулы строк 40, 42, 43, 45, 46, 47 и 61 всех листов CF.
    Если с выключенным переносом хоть одна изменит значение, паритет поедет на
    проектах, где перенос никто не включал.
    """
    book, ev = plain
    assert float(ev.cell("CF_1", "B65")) == pytest.approx(0.0)
    assert float(ev.cell("CF_2", "B64")) == pytest.approx(0.0)
    assert float(ev.cell("CF_1", "B47")) > 1_000, (
        "без переноса непогашенный долг обязан остаться на первой очереди")


def test_the_closed_line_neither_lends_nor_collects(carried):
    """НКЛ закрыт после того, как долг с него ушёл.

    Довод владельца: «как он может там висеть, если по договору НКЛ должна
    быть закрыта — юридически он или дефолтный, или в воздухе».
    """
    book, ev = carried
    from openpyxl.utils import get_column_letter
    rve = _as_date(ev.cell("CF_1", "B8"))
    assert rve is not None, "РВЭ очереди не вычислился"
    after = []
    for index in range(4, 124):
        column = get_column_letter(index)
        month = _as_date(ev.cell("CF_1", f"{column}3"))
        if month is not None and month > rve:
            after.append(column)
    assert after, "предохранитель: у очереди обязаны быть месяцы после РВЭ"
    for column in after[:12]:
        assert float(ev.cell("CF_1", f"{column}45")) == pytest.approx(0.0), (
            f"выборка ПФ в {column} после РВЭ: линия закрыта, выбирать нечего")
        assert float(ev.cell("CF_1", f"{column}46")) == pytest.approx(0.0), (
            f"погашение ПФ в {column} после РВЭ: остаточные продажи остаются "
            "застройщику")


def test_the_debt_lands_in_the_month_the_escrow_opened(carried):
    """Дата приёма — РВЭ передавшей очереди, а не открытие ПФ принимающей."""
    book, ev = carried
    from openpyxl.utils import get_column_letter
    rve = _as_date(ev.cell("CF_1", "B8"))
    permit = _as_date(ev.cell("CF_2", "B7"))
    expected = max(rve, permit)
    landed = [get_column_letter(i) for i in range(4, 124)
              if abs(float(ev.cell("CF_2", f"{get_column_letter(i)}64") or 0)) > 1e-9]
    assert len(landed) == 1, f"долг обязан приехать одним месяцем, а не {len(landed)}"
    month = _as_date(ev.cell("CF_2", f"{landed[0]}3"))
    assert (month.year, month.month) == (expected.year, expected.month), (
        f"месяц приёма {month} не совпал с раскрытием эскроу передавшей "
        f"очереди {expected}")
    # И это НЕ дата открытия ПФ принимающей — иначе тест проходит и на прежней,
    # неверной методике.
    assert (expected.year, expected.month) != (permit.year, permit.month), (
        "предохранитель: на этих вводных две даты обязаны различаться")


def test_the_parity_row_watches_the_transfer():
    """Перенос итоги проекта почти не двигает — он меняет, КТО платит.

    Прежние строки паритета смотрят на итоги, поэтому книга могла бы переносить
    не то и не тогда при вердикте «ПРОЙДЕНО». Своя строка обязана быть — и
    обязана входить в вердикт листа.
    """
    book = _book(True)
    checks = book["ПРОВЕРКИ"]
    row = core._V4_CARRY_PARITY_ROW
    assert "перевед" in str(checks[f"A{row}"].value).lower() or "переданный" in str(
        checks[f"A{row}"].value)
    assert float(checks[f"C{row}"].value) > 1_000, "контрольное число движка пусто"
    assert f"F6:F{row}" in str(book["ПРОВЕРКИ"]["B3"].value), (
        "новая строка не входит в вердикт листа: она красная, а лист «ПРОЙДЕНО»")


def test_the_flag_carries_the_engine_decision_not_the_intent():
    """Признак в книге = «перенос ПРИМЕНЁН», а не «пользователь его включил».

    Гейт по общему LLCR решает движок, и книга этого LLCR не знает. Полагаясь
    на одно намерение, она перенесла бы долг там, где банк отказал, и показала
    бы очередь рассчитавшейся, пока отчёт зовёт её дефолтной.
    """
    row = int(core._V4_CARRY_FLAG_CELL[1:])
    on = _book(True)["Вводные"]
    off = _book(False)["Вводные"]
    assert on[f"B{row}"].value == "Да", "на этом проекте перенос обязан примениться"
    assert off[f"B{row}"].value == "Нет"
    assert off[f"D{row}"].value == "carry_debt_applied", (
        "ключ рядом со значением — по нему ячейку находят глазами")


def test_a_refused_transfer_does_not_reach_the_workbook_as_applied():
    """Гейт отказал — книга обязана НЕ переносить, как и движок."""
    inputs = {**core.DEFAULT_INPUTS, "purchase_price_mln": 6000,
              "apartment_price_th": 430, "project_start": "2027-01-01",
              "ird_months": 12, "construction_months": 24}
    phasing = {**_phasing(True), "phase_count": 3,
               "phases": [{"name": f"О{i + 1}", "start_offset_months": 12 * i,
                           "construction_months": 24} for i in range(3)],
               "products": {key: [35, 35, 30] for key in
                            ("apartments", "ground_commercial",
                             "underground_parking", "storage")}}
    tep = {key: dict(row) for key, row in core.TEP_DEFAULT.items()}
    bundle = core.calculate_phased(core.PhasedCalcRequest(
        inputs=inputs, tep=tep, rates=[], phasing=phasing))
    assert (bundle.get("debt_carry") or {}).get("applied") is False, (
        "предохранитель: на этих вводных гейт обязан отказать")
    content, _, meta = core.build_project_workbook(
        inputs, tep, [], phasing, project_name="Отказ гейта")
    assert meta["missing"] == [], meta["missing"]
    book = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
    row = int(core._V4_CARRY_FLAG_CELL[1:])
    assert book["Вводные"][f"B{row}"].value == "Нет"
