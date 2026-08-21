"""Финансирование книги живёт в горизонте движка и берёт ту же базу комиссии.

Обе поверхности считают один проект, и книга сама объявляет расхождение
строкой «Паритет: стоимость финансирования» на листе ПРОВЕРКИ. У проекта
владельца от 06.08.2026 она показывала FAIL: 1833,2 млн в книге против
1741,6 у движка — и ровно на ту же разницу расходились чистая прибыль и
маржинальность, при совпадающих до копейки выручке, CAPEX, EBITDA, пике ПФ
и непогашенном долге. Причин оказалось две.

Первая: строка 42 CF-листов начисляла проценты ПФ до «РВЭ + срок продаж»
(H88..H91). Срок продаж отсчитывается от РнС, а не от РВЭ, а горизонт
движка — РВЭ + MAX(остаточные + 3, 12). На проекте, где долг не погашен,
книга накручивала лишние месяцы процентов на остаток.

Вторая: строка 57 брала базу комиссии выдачи БРИДЖа как покупка + П + РД,
без денежной социальной компенсации, которую движок в расчётный лимит
БРИДЖа включает.

Ни то ни другое не ловилось: прежние проверки паритета шли на вводных, где
соцнагрузка строится (компенсация 0), а долг гасится в срок — обе ветки
книги на них молчат.

Запуск: python3 -m pytest tests -q
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

import openpyxl
import pytest
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
sys.path.insert(0, str(Path(__file__).resolve().parent))

import main as wrapper  # noqa: E402
from xlsx_eval import Evaluator  # noqa: E402

core = wrapper.core


@pytest.fixture(scope="module")
def defaulting_case():
    """Проект с денежной соцкомпенсацией и непогашенным долгом на конец."""
    sys.setrecursionlimit(400000)
    x = dict(core.DEFAULT_INPUTS)
    x.update({
        "project_start": "2027-01-01",
        "purchase_price_mln": 250,
        "land_rights_cost_mln": 971.423,
        "vri_payment_mode": "lump",
        "vri_installment_years": 0,
        "social_mode": "Денежная компенсация",
        "social_compensation_mln": 777.336,
        "social_comp_date": "2028-06-01",
        "kindergarten_places": 0, "school_places": 0, "clinic_capacity": 0,
        "social_dou_gba_sqm": 0, "social_school_gba_sqm": 0,
        "social_clinic_gba_sqm": 0,
        "main_above_th_per_sqm": 190, "main_under_th_per_sqm": 190,
        "apartment_price_th": 450, "commercial_price_th": 450,
        "parking_price_th": 3000,
        "pace_adjustment_pct": 1,
        "rate_start_date": "2027-01-01",
        "limit_fee_pct": 0.5, "reservation_fee_pct": 0.5,
        "underground_manual_gns_sqm": 6475, "underground_manual_spaces": 185,
    })
    tep = {key: dict(value) for key, value in core.TEP_DEFAULT.items()}
    tep["apartments"].update(gns=33018, total_area=33018, useful=21462,
                             saleable=21462, units=0)
    tep["ground_commercial"].update(gns=2108, total_area=2108, useful=1897,
                                    saleable=1897, units=0)
    tep["underground_parking"].update(gns=6475, total_area=6475, units=185)
    tep["storage"].update(gns=0, total_area=0, units=0)
    tep["kindergarten"].update(gns=0, total_area=0, transfer=0, units=0)

    bundle = core._run_authoritative_model(x, tep, [], {"enabled": False})
    content, _, _ = core.build_project_workbook(
        x, tep, [], {"enabled": False}, finance_hints={})
    book = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
    return bundle["consolidated"], Evaluator(book)


def test_the_case_is_the_one_that_used_to_hide_the_gap(defaulting_case):
    """Предохранитель: если вводные перестанут давать непогашенный долг и
    денежную компенсацию, тесты ниже станут зелёными ни о чём."""
    result, _ = defaulting_case
    assert float(result["finance"]["ending_pf"]) > 0
    assert float(result["finance"]["bridge_fee"]) > 0
    assert float(result["summary"]["social_payment"]) > 0


def test_the_book_and_the_engine_agree_on_the_cost_of_financing(defaulting_case):
    """Допуск — процент, и остаток объяснён.

    Движок с 21.08.2026 резервирует деньги под будущие платежи ВРИ: пока
    обязательство живо, касса не раздаётся собственнику и долг гасится позже.
    Книга этого не умеет — там waterfall собран формулами, и резерва в них нет.
    Отсюда устойчивая разница около 0,7% по стоимости финансирования: движок
    платит чуть больше процентов. Разница видна в блоке «ПАРИТЕТ С ДВИЖКОМ»
    самой книги, а не спрятана, и стоит в открытых задачах — учить книгу
    резерву надо отдельной работой, это не подгонка допуска.
    """
    result, evaluator = defaulting_case
    engine = float(result["summary"]["financing_cost"]) / 1e6
    book = float(evaluator.cell("CF_1", "B74") or 0)
    assert book == pytest.approx(engine, rel=0.01)
    # Книга обязана быть ДЕШЕВЛЕ: у неё нет резерва, значит долг гасится раньше.
    # Дороже — это уже другая ошибка, и допуском её прикрывать нельзя.
    assert book <= engine, "книга дороже движка — резервом это не объясняется"


def test_the_interest_stops_where_the_model_ends(defaulting_case):
    """За горизонтом движка книга не считает ничего: месяц после последнего
    модельного — уже ноль, иначе непогашенный остаток копит проценты в
    пустоте."""
    result, evaluator = defaulting_case
    modelled = len(result["finance"]["rows"])
    interest = [float(evaluator.cell("CF_1", f"{get_column_letter(col)}42") or 0)
                for col in range(4, 124)]
    assert interest[modelled - 1] > 0
    assert sum(interest[modelled:]) == 0


def test_the_bridge_fee_counts_the_cash_social_payment(defaulting_case):
    """Расчётный лимит БРИДЖа — покупка + П + РД + денежная компенсация;
    комиссия выдачи считается от него же."""
    result, evaluator = defaulting_case
    finance = result["finance"]
    engine = (float(finance["bridge_fee"])
              + float(finance["pf_reservation_fee"])) / 1e6
    book = float(evaluator.cell("CF_1", "B57") or 0)
    assert book == pytest.approx(engine, abs=0.5)


def test_the_template_carries_both_corrections():
    """Шаблон — исходник этих формул: правка в нём, а не при выгрузке."""
    template = openpyxl.load_workbook(core._V4_TEMPLATE_PATH, data_only=False)
    social_rows = {"CF_1": "$B$31", "CF_2": "$B$65",
                   "CF_3": "$B$99", "CF_4": "$B$133"}
    for sheet, social in social_rows.items():
        interest = str(template[sheet]["D42"].value)
        assert "MAX('Вводные'!$B$69+3,12)+1" in interest, sheet
        assert "$H$8" not in interest and "$H$9" not in interest, sheet
        fee = str(template[sheet]["D57"].value)
        assert f"'CAPEX'!{social}" in fee, sheet
        # Признак обязателен: B17 книги несёт социалку и стройкой тоже, а её
        # движок в лимит БРИДЖа не берёт.
        assert '\'Вводные\'!$B$37="Денежная компенсация"' in fee, sheet
    assert template["Вводные"]["D37"].value == "social_mode"
