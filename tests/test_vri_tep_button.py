"""Кнопка бота «Посчитать ВРИ и ТЭП»: регион → (для МО плотность) → участок →
карточка + файл.

МО считается полностью (РНГП, УПКС, Кд); Москва — по формулам калькулятора
ГлавАПУ, восстановленным из его выгрузок (СПП 94/6, НП 90%, население
33 м²/чел, соцпотребность на тысячу жителей); машино-места и плата за ВРИ
не реверсятся — их считает калькулятор в мини-приложении, и карточка честно
об этом говорит. Файл — формат калькулятора ГлавАПУ: его читает наш же
парсер, и числа совпадают.

Запуск: python3 -m pytest tests -q
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import main as wrapper  # noqa: E402

core = wrapper.core


def test_the_mo_branch_counts_vri_and_returns_a_readable_file():
    """Файл — полный 91-строчный формат калькулятора с секциями: первый
    вариант был огрызком из 13 строк, без секции ВРИ и с «ДОО 0 млн.руб»."""
    result = core.vri_tep_quick("mo", "", site_area_ha=22.423,
                                district="Городской округ Мытищи",
                                density_sqm_per_ha=8700)
    assert "Московская область" in result["card"]
    assert "плата за смену ВРИ — 4 643,9" in result["card"]
    import io
    import openpyxl
    book = openpyxl.load_workbook(io.BytesIO(result["file"]))
    assert book.sheetnames == ["ТЭП", "МПТ", "Машино-места", "Параметры территории"]
    sheet = book["ТЭП"]
    assert sheet.max_row == 91, "формат неполный — калькулятор ГлавАПУ ведёт 91 строку"
    labels = {str(sheet.cell(row=r, column=1).value): r for r in range(2, 92)}
    assert sheet.cell(row=labels["44"], column=4).value == "4 643,921"
    assert sheet.cell(row=labels["18"], column=4).value == "453"
    assert sheet.cell(row=labels["22"], column=4).value == "950"
    parsed = core.parse_glavapu_xlsx(result["file"], result["filename"])
    normalized = parsed["normalized"]
    assert normalized["site_area_ha"] == pytest.approx(22.423)
    assert normalized["apartment_area_sqm"] == pytest.approx(195080.0, rel=0.001)
    assert normalized["change_vri_mln"] == pytest.approx(4643.921, rel=0.001)
    assert normalized["actual_kindergarten_places"] == pytest.approx(453)


def test_the_msk_branch_reproduces_the_calculator_export(monkeypatch):
    """Московская ветка воспроизводит настоящую выгрузку калькулятора ГлавАПУ
    (эталон — файл по 77:09:0004014:13 от 16.08.2026, население 422): четыре
    листа, итоги в строках секций, спорт/торговля/озеленение по нормативам на
    тысячу жителей с округлением вверх, компенсация по городским ставкам,
    паркинг по методике августа 2026 (одно место на 90 м² НП жилых × К1),
    плата за ВРИ с квартальной индексацией базовой стоимости."""
    monkeypatch.setattr(core, "analyze_cadastral_territory", lambda req: {
        "territory": {"area_ha": 0.651, "district": "Савеловский",
                      "cadastral_quarter": "77:09:0004014"},
        "coefficients": {"rail": 0.75, "business_outside_ttc": 0.5,
                         "rent": 0.1281, "base_cost_zh_high": 229036.29},
    })
    result = core.vri_tep_quick("msk", "77:09:0004014:13")
    assert "Москва" in result["card"]
    assert "мини-приложении" in result["card"]
    assert "580,7 млн ₽" in result["card"]
    assert "паркинг — 178 м/м" in result["card"]
    # Формула Df.calcOwn с индексацией базовой на площади 0,651 га (СПП
    # 22 785 против 22 782 в эталоне из-за округления площади): 1 289,9 млн ₽.
    assert "плата за смену ВРИ — 1 289,9 млн ₽" in result["card"]
    import io
    import openpyxl
    book = openpyxl.load_workbook(io.BytesIO(result["file"]))
    assert book.sheetnames == ["ТЭП", "МПТ", "Машино-места", "Параметры территории"], \
        "выгрузка калькулятора ведёт четыре листа именно в этом порядке"
    sheet = book["ТЭП"]
    assert sheet.max_row == 91
    labels = {str(sheet.cell(row=r, column=1).value): r for r in range(2, 92)}
    cell = lambda code: sheet.cell(row=labels[code], column=4).value
    # Числа эталона: население 422, квартир 201, соцпотребность 19/38/9 (6+3).
    assert cell("4") == "422"
    assert cell("5") == "201"
    assert cell("3") == "31,5"
    assert cell("12") == "0,651 (100,0%)"
    assert [cell(c) for c in ("30", "31", "32", "33", "34")] == \
        ["19", "38", "9", "6", "3"]
    # Спорт, торговля и озеленение — нормативы на тысячу жителей, копейка в
    # копейку с эталоном (округление вверх до последнего знака).
    assert [cell(c) for c in ("35", "36", "36.1", "36.2")] == \
        ["0,0410", "0,338", "0,136", "0,203"]
    assert [cell(c) for c in ("37", "38", "39", "40", "41")] == \
        ["0,114", "0,043", "0,051", "0,064", "0,038"]
    assert [cell(c) for c in ("57", "58", "59", "60")] == \
        ["0,2110", "0,0211", "0,0043", "0,0296"]
    # Компенсация — как в эталонной выгрузке от 01.08.2026.
    assert [cell(c) for c in ("54", "55", "56")] == \
        ["188,414", "294,540", "97,714"]
    comp_row = next(r for r in range(2, 92)
                    if "компенсации за социальные" in str(sheet.cell(row=r, column=2).value or ""))
    assert sheet.cell(row=comp_row, column=4).value == "580,668"
    # Машино-места — методика августа 2026: эталонная выгрузка от 16.08.2026
    # по этому участку даёт 161+17+6 и 1+3 кратковременных.
    assert [cell(c) for c in ("42", "42.1", "42.2", "42.3", "43")] == \
        ["184", "161", "17", "6", "4"]
    mm_sheet = book["Машино-места"]
    assert [mm_sheet.cell(row=2, column=c).value for c in range(4, 9)] == \
        ["179", "0", "161", "17", "1"]
    assert [mm_sheet.cell(row=3, column=c).value for c in range(4, 9)] == \
        ["9", "6", "0", "0", "3"]
    # Плата за смену ВРИ — формула Df.calcOwn с квартальной индексацией
    # базовой: 1,8964 × СПП × 0,1281 × 229 036,29 × 1,0175 / 1,00001. Эталон
    # от 16.08.2026 по точной площади 0,65091 га даёт 1 289,735; здесь
    # площадь округлена до 0,651.
    expected_vri = round(1.8964 * 0.651 * 35000 * 0.1281 * 229036.29
                         * core._GLAVAPU_VRI_BASE_INDEXATION / 1.00001 / 1e6, 3)
    expected_text = f"{expected_vri:,.3f}".replace(",", " ").replace(".", ",")
    assert cell("44") == expected_text == "1 289,919"
    vri_row = next(r for r in range(2, 92)
                   if "стоимости смены ВРИ" in str(sheet.cell(row=r, column=2).value or ""))
    assert sheet.cell(row=vri_row, column=4).value == expected_text
    # МПТ: 1 367 м² нежилой СПП по 36 м² на рабочее место.
    assert book["МПТ"].cell(row=2, column=4).value == "38"
    parsed = core.parse_glavapu_xlsx(result["file"], result["filename"])
    normalized = parsed["normalized"]
    assert normalized["site_area_ha"] == pytest.approx(0.651)
    assert normalized["social_compensation_total_mln"] == pytest.approx(580.668, abs=0.01)
    assert normalized["suggested_social_mode"] == "Денежная компенсация"
    assert normalized["parking_permanent"] == pytest.approx(161)
    assert normalized["parking_guest"] == pytest.approx(17)
    assert normalized["change_vri_mln"] == pytest.approx(1289.919, abs=0.001)


def test_the_second_export_confirms_the_methodology(monkeypatch):
    """Вторая точка — Гродненская 77:07:0008006:3 (население 377, выгрузка от
    16.08.2026). Формула, выведенная по одному участку, — подгонка: здесь та же
    методика обязана сойтись на другом квартале, с другими К2, арендой и
    базовой стоимостью. Эталон: постоянные 144, гостевые 15, приобъектные 7,
    ДОО 17, школа 34, поликлиника 5+3, плата за ВРИ 851,315 млн ₽ (у нас
    851,602 — расхождение 0,03% от округления площади до 0,581 га)."""
    monkeypatch.setattr(core, "analyze_cadastral_territory", lambda req: {
        "territory": {"area_ha": 0.581, "district": "Можайский",
                      "cadastral_quarter": "77:07:0008006"},
        "coefficients": {"rail": 0.75, "business_outside_ttc": 0.7,
                         "rent": 0.1161, "base_cost_zh_high": 186939.25},
    })
    result = core.vri_tep_quick("msk", "77:07:0008006:3")
    import io
    import openpyxl
    sheet = openpyxl.load_workbook(io.BytesIO(result["file"]))["ТЭП"]
    labels = {str(sheet.cell(row=r, column=1).value): r for r in range(2, 92)}
    cell = lambda code: sheet.cell(row=labels[code], column=4).value
    assert cell("4") == "377"
    assert [cell(c) for c in ("30", "31", "32", "33", "34")] == \
        ["17", "34", "8", "5", "3"], "поликлиника: 13,3 на этом населении дала бы 6 взрослых"
    assert [cell(c) for c in ("42", "42.1", "42.2", "42.3")] == \
        ["166", "144", "15", "7"]
    normalized = core.parse_glavapu_xlsx(result["file"], result["filename"])["normalized"]
    assert normalized["parking_permanent"] == pytest.approx(144)
    assert normalized["change_vri_mln"] == pytest.approx(851.602, abs=0.001)
    assert abs(normalized["change_vri_mln"] / 851.315 - 1) < 0.001, \
        "плата разошлась с эталонной выгрузкой больше чем на 0,1%"


def test_the_third_point_pins_the_mixed_clinic_norm(monkeypatch):
    """Третья точка — 77:03:0006003:98 (население 970, дрейф от 16.08.2026).

    Смешанная поликлиника — свой норматив 19 пос./смену на тысячу, а не сумма
    взрослой и детской: город дал 19 при наших частях 13+7, и компенсация
    разошлась ровно в 20/19 (190,814 против 200,857 — плашка дрейфа). На
    населении 377 и 422 суммы совпадали со смешанной случайно."""
    monkeypatch.setattr(core, "analyze_cadastral_territory", lambda req: {
        "territory": {"area_ha": 1.4968, "district": "Головинский",
                      "cadastral_quarter": "77:03:0006003"},
        "coefficients": {"rail": 0.75, "business_outside_ttc": 0.5,
                         "rent": 0.1281, "base_cost_zh_high": 229036.29},
    })
    result = core.vri_tep_quick("msk", "77:03:0006003:98")
    import io
    import openpyxl
    sheet = openpyxl.load_workbook(io.BytesIO(result["file"]))["ТЭП"]
    labels = {str(sheet.cell(row=r, column=1).value): r for r in range(2, 92)}
    cell = lambda code: sheet.cell(row=labels[code], column=4).value
    assert cell("4") == "970"
    assert cell("32") == "19", "смешанная поликлиника — норматив 19/1000, не сумма 13+7"
    assert [cell(c) for c in ("33", "34")] == ["13", "7"]


def test_a_missing_coefficient_keeps_the_parking_honest(monkeypatch):
    """Без К1/К2 и базовой стоимости квартала ни машино-места, ни плата за
    ВРИ не выдумываются — нули и честная отсылка в мини-приложение."""
    monkeypatch.setattr(core, "analyze_cadastral_territory", lambda req: {
        "territory": {"area_ha": 0.651}, "coefficients": {},
    })
    result = core.vri_tep_quick("msk", "77:09:0004014:13")
    assert "Плату за смену ВРИ и машино-места считает калькулятор" in result["card"]
    parsed = core.parse_glavapu_xlsx(result["file"], result["filename"])
    assert (parsed["normalized"]["parking_permanent"] or 0) == 0
    assert (parsed["normalized"]["change_vri_mln"] or 0) == 0


def test_both_branches_attach_the_filled_developaid_template(monkeypatch):
    """Экспорт — ещё и в формате шаблона DevelopAid: тот же файл, что бот
    предлагает заполнить и загрузить, — его правят и возвращают в расчёт."""
    monkeypatch.setattr(core, "analyze_cadastral_territory", lambda req: {
        "territory": {"area_ha": 0.651, "district": "Савеловский"},
        "coefficients": {"rail": 0.75, "business_outside_ttc": 0.5},
    })
    msk = core.vri_tep_quick("msk", "77:09:0004014:13")
    parsed = core.parse_manual_tep_xlsx(msk["template_file"], msk["template_filename"])
    assert parsed["inputs"]["social_compensation_mln"] == pytest.approx(580.668, abs=0.01)
    assert parsed["inputs"]["social_mode"] == "Денежная компенсация"
    assert parsed["tep"]["apartments"]["saleable"] == pytest.approx(13921.6, rel=0.001)
    assert parsed["tep"]["underground_parking"]["units"] == pytest.approx(178)

    mo = core.vri_tep_quick("mo", "", site_area_ha=22.423,
                            district="Городской округ Мытищи",
                            density_sqm_per_ha=8700)
    parsed = core.parse_manual_tep_xlsx(mo["template_file"], mo["template_filename"])
    assert parsed["inputs"]["land_rights_cost_mln"] == pytest.approx(4643.921, rel=0.001)
    assert parsed["tep"]["apartments"]["saleable"] == pytest.approx(195080.1, rel=0.001)
    assert parsed["tep"]["kindergarten"]["units"] == pytest.approx(453)
    # Квартиры Подмосковья меряет РНГП области, а не число без источника: было
    # 3 321 по литералу 58,75 (он не брался ниоткуда), стало 3 318 по 28 м² на
    # человека при размере домовладения 2,1. Утверждение — про ДЕЛИТЕЛЬ, а не
    # про запомненное число: литерал разошёлся бы с движком молча.
    flat, basis = core.average_flat_sqm("mo")
    assert "РНГП Московской области" in basis, basis
    assert parsed["tep"]["apartments"]["units"] == pytest.approx(
        parsed["tep"]["apartments"]["saleable"] / flat, rel=0.001)


def test_the_msk_branch_finds_the_parcel_by_address(monkeypatch):
    """«Мишина 46 Москва» падал: анализ территории ждёт кадастры. Теперь
    адрес сперва проходит через тот же поиск ЕГРН, что и основной бот."""
    seen: dict[str, str] = {}
    monkeypatch.setattr(core, "land_lookup_via_core", lambda query: seen.update(
        {"query": query}) or {"results": [
            {"kind": "land", "cadastral_number": "77:09:0004014:13"},
            {"kind": "building", "cadastral_number": "77:09:0004014:1000"},
        ]})
    analyzed: dict[str, list] = {}
    monkeypatch.setattr(core, "analyze_cadastral_territory", lambda req: analyzed.update(
        {"numbers": list(req.cadastral_numbers)}) or {
        "territory": {"area_ha": 0.651, "district": "Савеловский"},
        "coefficients": {},
    })
    result = core.vri_tep_quick("msk", "Мишина 46 Москва")
    assert seen["query"] == "Мишина 46 Москва"
    assert analyzed["numbers"] == ["77:09:0004014:13"], \
        "в анализ должны уходить только земельные участки"
    assert "0,6510" in result["card"]


def test_the_msk_address_without_a_parcel_reads_like_advice(monkeypatch):
    monkeypatch.setattr(core, "land_lookup_via_core", lambda query: {"results": []})
    with pytest.raises(Exception) as err:
        core.vri_tep_quick("msk", "несуществующий адрес")
    assert "кадастровый номер" in str(getattr(err.value, "detail", err.value))


def test_the_bot_flow_asks_region_then_takes_the_parcel(monkeypatch, tmp_path):
    sent: list[dict] = []
    monkeypatch.setattr(wrapper, "_STATE_DIR", tmp_path)
    monkeypatch.setattr(wrapper, "_send_message",
                        lambda chat_id, text, **kw: sent.append(
                            {"text": text, "markup": kw.get("reply_markup")}))
    documents: list[str] = []
    monkeypatch.setattr(core, "_telegram_send_document_bytes",
                        lambda chat_id, data, name, **kw: documents.append(name))
    monkeypatch.setattr(core, "vri_tep_quick",
                        lambda region, query, **kw: {
                            "card": f"карточка {region} {query}",
                            "file": b"PK", "filename": "f.xlsx"})

    wrapper._start_vritep(42)
    assert "Выберите регион" in sent[-1]["text"]
    wrapper._vritep_ask_input(42, "mo")
    assert wrapper._vritep_region(42) == "mo_density", \
        "МО начинается с вопроса о плотности"
    assert "по умолчанию" in sent[-1]["text"]
    handled = wrapper._vritep_handle_text(42, "35")
    assert handled
    assert wrapper._vritep_region(42) == "mo"
    handled = wrapper._vritep_handle_text(42, "50:12:0100131:497")
    assert handled
    assert any("карточка mo 50:12:0100131:497" in item["text"] for item in sent)
    assert documents == ["f.xlsx"]
    assert wrapper._vritep_region(42) == "", "состояние не снято после расчёта"


def test_the_manual_area_and_district_are_parsed(monkeypatch, tmp_path):
    monkeypatch.setattr(wrapper, "_STATE_DIR", tmp_path)
    calls: list[dict] = []
    monkeypatch.setattr(wrapper, "_send_message", lambda *a, **kw: None)
    monkeypatch.setattr(core, "_telegram_send_document_bytes", lambda *a, **kw: None)
    monkeypatch.setattr(core, "vri_tep_quick",
                        lambda region, query, **kw: calls.append(
                            {"region": region, "query": query, **kw}) or {
                            "card": "к", "file": b"PK", "filename": "f.xlsx"})
    wrapper._vritep_ask_input(42, "mo")
    wrapper._vritep_handle_text(42, "10,5 га Городской округ Мытищи")
    assert calls[-1]["site_area_ha"] == pytest.approx(10.5)
    assert calls[-1]["district"] == "Городской округ Мытищи"
    assert calls[-1]["query"] == ""
    # Участок прислали, минуя вопрос о плотности — берётся умолчание 35
    # тыс. м² СПП/га, то есть ≈ 21 385 м² квартир на гектар.
    assert calls[-1]["density_sqm_per_ha"] == pytest.approx(35 * 1000 * 0.94 * 0.65)


def test_the_button_lives_in_both_menus():
    """Стартовое меню бота живёт в движке, меню помощи — в обёртке: кнопка
    была только во втором, и /start её не показывал."""
    legacy = open("main_legacy.py", encoding="utf-8").read()
    assert legacy.count('"vritep_start"') >= 1, "кнопки нет в стартовом меню"
    wrapper_src = open("main.py", encoding="utf-8").read()
    assert wrapper_src.count('"vritep_start"') >= 2, \
        "кнопка или колбэк пропали из обёртки"


def test_the_density_is_parsed_from_the_bot_text(monkeypatch, tmp_path):
    monkeypatch.setattr(wrapper, "_STATE_DIR", tmp_path)
    calls: list[dict] = []
    monkeypatch.setattr(wrapper, "_send_message", lambda *a, **kw: None)
    monkeypatch.setattr(core, "_telegram_send_document_bytes", lambda *a, **kw: None)
    monkeypatch.setattr(core, "vri_tep_quick",
                        lambda region, query, **kw: calls.append(
                            {"region": region, "query": query, **kw}) or {
                            "card": "к", "file": b"PK", "filename": "f.xlsx"})
    wrapper._vritep_ask_input(42, "mo")
    wrapper._vritep_handle_text(42, "22,4 га Городской округ Мытищи плотность 8700")
    assert calls[-1]["density_sqm_per_ha"] == pytest.approx(8700)
    assert calls[-1]["site_area_ha"] == pytest.approx(22.4)
    assert calls[-1]["district"] == "Городской округ Мытищи"


def test_the_density_step_understands_both_metrics(monkeypatch, tmp_path):
    """Ответ на вопрос о плотности: до 1000 — тыс. м² СПП/га по метрике
    ГлавАПУ, больше — уже м² квартир/га; «квартир 8700» — явная метрика РНГП."""
    monkeypatch.setattr(wrapper, "_STATE_DIR", tmp_path)
    calls: list[dict] = []
    monkeypatch.setattr(wrapper, "_send_message", lambda *a, **kw: None)
    monkeypatch.setattr(core, "_telegram_send_document_bytes", lambda *a, **kw: None)
    monkeypatch.setattr(core, "vri_tep_quick",
                        lambda region, query, **kw: calls.append(
                            {"region": region, "query": query, **kw}) or {
                            "card": "к", "file": b"PK", "filename": "f.xlsx"})
    assert wrapper._vritep_mo_density("35") == pytest.approx(35 * 1000 * 0.94 * 0.65)
    assert wrapper._vritep_mo_density("8 700") == pytest.approx(8700)
    assert wrapper._vritep_mo_density("квартир 8700") == pytest.approx(8700)
    assert wrapper._vritep_mo_density("50:12:0100131:497") is None

    wrapper._vritep_ask_input(42, "mo")
    wrapper._vritep_handle_text(42, "14,3")
    wrapper._vritep_handle_text(42, "10 га Городской округ Мытищи")
    assert calls[-1]["density_sqm_per_ha"] == pytest.approx(14.3 * 1000 * 0.94 * 0.65)


def test_the_native_menu_and_the_command_open_the_button(monkeypatch, tmp_path):
    """Нативное меню Telegram объявляется один раз — в движке. Списка было
    два (движок при вебхуке, обёртка на старте), побеждал последний, и
    /vritep из меню пропадал, хотя команда работала."""
    commands = {item["command"] for item in core.TELEGRAM_BOT_COMMANDS}
    assert "vritep" in commands
    # Команды движка и обёртки объявляются в одном месте — но не все попадают в
    # само меню: оно сведено к пяти решениям, остальные живут рядом, в
    # TELEGRAM_EXTRA_COMMANDS, и перечисляются помощью. Единственность источника
    # от этого не меняется, а вот «всё обязано быть в меню» больше не условие.
    declared = commands | {item["command"] for item in core.TELEGRAM_EXTRA_COMMANDS}
    assert {"address", "comment", "status"} <= declared, \
        "единый список обязан покрывать команды и движка, и обёртки"
    wrapper_src = open("main.py", encoding="utf-8").read()
    assert wrapper_src.count('{"command"') == 0, \
        "у обёртки не должно быть собственного списка команд"
    assert "core.TELEGRAM_BOT_COMMANDS" in wrapper_src
    legacy_src = open("main_legacy.py", encoding="utf-8").read()
    assert legacy_src.count('"commands": TELEGRAM_BOT_COMMANDS') == 1
    monkeypatch.setattr(wrapper, "_STATE_DIR", tmp_path)
    sent: list[str] = []
    monkeypatch.setattr(wrapper, "_send_message",
                        lambda chat_id, text, **kw: sent.append(text))
    wrapper._handle_message({"chat": {"id": 42}, "from": {"id": 42},
                             "text": "/vritep"})
    assert sent and "Выберите регион" in sent[-1]


def test_the_chat_accepts_both_of_its_own_file_formats(monkeypatch):
    """Подпись к выгрузке обещает «загрузить как обычный ТЭП», а чат принимал
    только шаблон DevelopAid — файл ГлавАПУ падал с ошибкой структуры.
    Теперь бот принимает оба собственных формата."""
    result = core.vri_tep_quick("mo", "", site_area_ha=22.423,
                                district="Городской округ Мытищи",
                                density_sqm_per_ha=8700)
    sent: list[str] = []
    monkeypatch.setattr(core, "_telegram_send_message",
                        lambda chat_id, text, **kw: sent.append(text))
    monkeypatch.setattr(core, "_telegram_web_app_url",
                        lambda *a, **kw: "https://example.org/")

    monkeypatch.setattr(core, "_telegram_download_document",
                        lambda doc: (result["file"], result["filename"]))
    core._telegram_handle_manual_document(42, {"file_name": result["filename"]})
    assert "Файл калькулятора ГлавАПУ распознан" in sent[-1]
    assert "22,4230 га" in sent[-1]

    monkeypatch.setattr(core, "_telegram_download_document",
                        lambda doc: (result["template_file"], result["template_filename"]))
    core._telegram_handle_manual_document(42, {"file_name": result["template_filename"]})
    assert "Ручной ТЭП распознан" in sent[-1]
