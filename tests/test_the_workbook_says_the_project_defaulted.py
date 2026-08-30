"""Книга называет дефолт там же, где его называет экран.

Владелец, 30.08.2026: «в экселе мы где-то увидим что проект дефолт». Строка у
книги была — «Непогашенный долг на конец проекта (дефолт, если > 0)», — но с
тем же изъяном, который на экране чинился отдельно: она смотрит на КОНЕЦ
горизонта. Банк ждёт погашения в дату раскрытия эскроу, и если остаточные
продажи закрывают долг годом позже, эта строка показывает ноль — книга молчит
ровно там, где дефолт и произошёл.

Нехватку книга считает по очередям (строка 29 листов CF), но одной суммой её
показывать нельзя: та же величина возникает при законном переносе долга
следующей очереди, а это не дефолт — банк его принял. Поэтому из суммы
вычитается переданное.

Запуск: python3 -m pytest tests/test_the_workbook_says_the_project_defaulted.py -q
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT.parent))
sys.path.insert(0, str(ROOT))

import main as _wrapper  # noqa: E402
from xlsx_eval import Evaluator  # noqa: E402

core = _wrapper.core


def _scenario(carry: bool) -> tuple[dict, dict, dict]:
    inputs = dict(core.DEFAULT_INPUTS)
    inputs.update(purchase_price_mln=12000, project_start="2027-01-01",
                  ird_months=12, construction_months=24, apartment_price_th=900)
    tep = {key: dict(row) for key, row in core.TEP_DEFAULT.items()}
    phasing = {
        "enabled": True, "mode": "phased", "user_enabled": True,
        "phase_count": 2, "phase_gap_months": 12,
        "phases": [{"name": "О1", "start_offset_months": 0, "construction_months": 24},
                   {"name": "О2", "start_offset_months": 12, "construction_months": 24}],
        "products": {key: [35, 65] for key in
                     ("apartments", "ground_commercial", "underground_parking")},
        "shared_cash": {}, "shared_allocation": {}, "social_objects": [],
        "carry_debt_forward": carry,
    }
    return inputs, tep, phasing


def _built(carry: bool) -> tuple[dict, Evaluator, list]:
    sys.setrecursionlimit(400000)
    inputs, tep, phasing = _scenario(carry)
    bundle = core.calculate_phased(core.PhasedCalcRequest(
        inputs=inputs, tep=tep, rates=[], phasing=phasing))
    content, _, meta = core.build_project_workbook(
        inputs, tep, [], phasing, project_name="Дефолт")
    return bundle, Evaluator(openpyxl.load_workbook(io.BytesIO(content))), meta["missing"]


@pytest.fixture(scope="module")
def defaulted() -> tuple[dict, Evaluator, list]:
    bundle, evaluator, missing = _built(carry=False)
    financing = (bundle["consolidated"].get("report") or {}).get("financing") or {}
    assert financing.get("default_date"), (
        "предохранитель: на этих вводных дефолт обязан быть")
    assert float(financing.get("ending_pf") or 0.0) < 500_000, (
        "предохранитель: долг обязан быть закрыт к концу горизонта — иначе "
        "сработала бы прежняя строка 22, и новая ни при чём")
    return bundle, evaluator, missing


def test_the_row_lands_without_a_single_shifted_reference(defaulted) -> None:
    """Пустая строка шаблона занимается, соседняя панель не трогается."""
    _, _, missing = defaulted
    assert missing == [], missing


def test_the_old_row_is_silent_here(defaulted) -> None:
    """Ради чего всё: прежняя строка книги в этом случае показывает ноль."""
    _, evaluator, _ = defaulted
    assert float(evaluator.cell("ОТЧЕТ", "B22") or 0.0) < 0.5


def test_the_new_row_names_the_unpaid_amount(defaulted) -> None:
    bundle, evaluator, _ = defaulted
    book = float(evaluator.cell("ОТЧЕТ", "B21"))
    phases = [phase["result"]["finance"] for phase in bundle["phases"]]
    engine = (sum(f.get("rve_unpaid", 0.0) for f in phases)
              - sum(f.get("debt_carried_out", 0.0) for f in phases)) / 1e6
    assert engine > 500, "предохранитель: сравнивать не с чем"
    assert book == pytest.approx(engine, rel=0.005), (
        f"книга {book:,.1f} против движка {engine:,.1f}")


def test_a_transferred_debt_is_not_a_default() -> None:
    """Та же нехватка, но долг принят следующей очередью — строка молчит.

    Без вычитания переданного книга объявляла бы дефолтом законный перенос,
    и строка перестала бы что-либо значить: она стояла бы почти всегда.
    """
    bundle, evaluator, missing = _built(carry=True)
    assert missing == []
    carried = bundle.get("debt_carry") or {}
    assert carried.get("applied") is True, (
        "предохранитель: перенос обязан сработать, иначе ветка не та")
    phases = [phase["result"]["finance"] for phase in bundle["phases"]]
    assert sum(f.get("rve_unpaid", 0.0) for f in phases) > 500_000_000, (
        "предохранитель: нехватка обязана быть — иначе ноль в книге ничего "
        "не доказывает")
    assert float(evaluator.cell("ОТЧЕТ", "B21") or 0.0) < 0.5


def test_the_label_says_what_the_number_means() -> None:
    """Число без подписи в книге — та же голая цифра, что и на экране."""
    inputs, tep, phasing = _scenario(carry=False)
    content, _, _ = core.build_project_workbook(
        inputs, tep, [], phasing, project_name="Подпись")
    sheet = openpyxl.load_workbook(io.BytesIO(content))["ОТЧЕТ"]
    label = str(sheet["A21"].value or "")
    assert "дефолт" in label, label
    assert "не передано следующей очереди" in label, label
    # Соседняя панель темпов продаж в тех же строках цела.
    assert sheet["F21"].value == "Темп продаж в штуках"
