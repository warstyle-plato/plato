"""Календарь соцстройки книга считает сама, а не хранит 480 чисел.

Владелец: «эксель должен работать почти как движок… если что-то меняешь
где-то, все должно меняться так же как в движке» (03.09.2026), и следом —
жалоба клиента: «у вас книга просто придаток к веб-сервису, она не
самостоятельный продукт».

Ревизия нашла в четырёх строках CAPEX 480 чисел: помесячный график стройки
садика, школы и поликлиники. Сдвинь в книге начало объекта или срок стройки —
не менялось ничего, а движок строит каждый объект в СВОЕЙ очереди её
календарём. Выразить это формулой было нечем: ячеек объектов на «Вводных» не
существовало вовсе, только общая строка A17 «Социальный платёж».

Теперь блок есть — строка на пару «тип × очередь» с местами, ценой места,
началом, сроком и множителем инфляции, — а строка CAPEX читает его формулой.

Запуск: python3 -m pytest tests/test_the_social_calendar_lives_in_the_book.py -q
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "tests"))

import main_legacy as core  # noqa: E402

openpyxl = pytest.importorskip("openpyxl")

from openpyxl.utils import get_column_letter  # noqa: E402

BASE = {**core.DEFAULT_INPUTS, "purchase_price_mln": 12000,
        "project_start": "2027-01-01", "ird_months": 12,
        "construction_months": 24, "apartment_price_th": 700,
        "kindergarten_places": 250, "school_places": 800, "clinic_capacity": 100}

PHASING = {
    "enabled": True, "mode": "phased", "user_enabled": True,
    "phase_count": 3, "phase_gap_months": 12,
    "phases": [{"name": f"О{index + 1}", "start_offset_months": 12 * index,
                "construction_months": 24} for index in range(3)],
    "products": {key: [30, 35, 35] for key in
                 ("apartments", "ground_commercial", "underground_parking")},
    "shared_cash": {}, "shared_allocation": {},
    "social_objects": [{"type": "kindergarten", "capacity": 250, "phase": 1},
                       {"type": "school", "capacity": 800, "phase": 2},
                       {"type": "clinic", "capacity": 100, "phase": 3}],
    "carry_debt_forward": False,
}

CAPEX_SOCIAL_ROWS = (31, 65, 99, 133)


def book(phasing=None, **overrides):
    content, _, _ = core.build_project_workbook(
        {**BASE, **overrides}, core.TEP_DEFAULT, [], phasing or PHASING, project_name="Соц")
    return openpyxl.load_workbook(io.BytesIO(content), data_only=False)


def evaluated(workbook):
    from xlsx_eval import Evaluator

    sys.setrecursionlimit(400000)
    return Evaluator(workbook)


def monthly(evaluator, row: int) -> list[float]:
    return [evaluator.cell("CAPEX", f"{get_column_letter(4 + index)}{row}")
            for index in range(120)]


@pytest.fixture(scope="module")
def phased():
    return book()


def test_the_monthly_row_is_a_formula_not_a_number(phased):
    """480 чисел — это график, который ничему не подчиняется."""
    for row in CAPEX_SOCIAL_ROWS:
        for column_index in (0, 17, 60, 119):
            coord = f"{get_column_letter(4 + column_index)}{row}"
            value = phased["CAPEX"][coord].value
            assert isinstance(value, str) and value.startswith("="), coord


def test_the_block_has_a_row_for_every_pair(phased):
    """Строка есть и у пустой пары: включить объект в книге иначе было бы негде,
    а формула CAPEX ссылалась бы в пустоту."""
    sheet = phased["Вводные"]
    labels = [str(sheet[f"A{row}"].value or "") for row in range(1, sheet.max_row + 1)]
    for label in ("ДОО", "СОШ", "Поликлиника"):
        for phase in range(1, 5):
            assert f"{label} — очередь {phase}" in labels, (label, phase)


def test_the_book_builds_each_object_in_its_own_queue(phased):
    """Ровно то, что делает движок: садик в первой, школа во второй, поликлиника
    в третьей — каждый своим календарём, а не общей долей от одной даты.

    Стройка объекта начинается не раньше РнС своей очереди — БРИДЖ стройку не
    финансирует (владелец, 04.09.2026); здесь ИРД 12 месяцев, шаг очередей 12.
    """
    evaluator = evaluated(phased)
    windows = {}
    for index, row in enumerate(CAPEX_SOCIAL_ROWS):
        months = [month for month, value in enumerate(monthly(evaluator, row))
                  if abs(value) > 1e-9]
        windows[index] = (months[0], months[-1], len(months)) if months else None
    assert windows[0] == (12, 35, 24), windows[0]     # ДОО, 24 месяца с РнС первой
    assert windows[1] == (24, 53, 30), windows[1]     # СОШ, 30 месяцев с РнС второй
    assert windows[2] == (36, 59, 24), windows[2]     # поликлиника, с РнС третьей
    assert windows[3] is None


def test_the_total_still_equals_the_declared_social_load(phased):
    """Сумма месяцев обязана сойтись с B17 — иначе книга строит не то, что объявила."""
    evaluator = evaluated(phased)
    total = sum(evaluator.cell("CAPEX", f"B{row}") for row in CAPEX_SOCIAL_ROWS)
    assert total == pytest.approx(evaluator.cell("Вводные", "B17"), abs=0.01)


def test_moving_the_start_inside_excel_moves_the_schedule(phased):
    """Ради чего правка: сдвиг начала объекта прямо в книге двигает график.

    Проверяется пересчётом книги, а не подстановкой числа движка: число на
    этом месте осталось бы прежним при любой правке."""
    evaluator = evaluated(phased)
    before = [month for month, value in enumerate(monthly(evaluator, 31))
              if abs(value) > 1e-9]

    moved = book()
    sheet = moved["Вводные"]
    row = next(number for number in range(1, sheet.max_row + 1)
               if str(sheet[f"A{number}"].value or "") == "ДОО — очередь 1")
    sheet[f"D{row}"] = sheet[f"D{row}"].value + 6      # начало на полгода позже
    after = [month for month, value in enumerate(monthly(evaluated(moved), 31))
             if abs(value) > 1e-9]
    assert after and after[0] > before[0], (before[:2], after[:2])


def test_stretching_the_term_inside_excel_thins_the_months(phased):
    """Срок вдвое длиннее — расход месяца вдвое меньше, итог тот же."""
    evaluator = evaluated(phased)
    before = monthly(evaluator, 31)

    longer = book()
    sheet = longer["Вводные"]
    row = next(number for number in range(1, sheet.max_row + 1)
               if str(sheet[f"A{number}"].value or "") == "ДОО — очередь 1")
    sheet[f"E{row}"] = sheet[f"E{row}"].value * 2
    after = monthly(evaluated(longer), 31)

    assert sum(after) == pytest.approx(sum(before), abs=0.01)
    assert max(after) == pytest.approx(max(before) / 2, rel=1e-6)


def test_the_places_drive_the_cost(phased):
    """Мест вдвое больше — стоимость объекта вдвое выше: цена места ячейкой рядом."""
    doubled = book()
    sheet = doubled["Вводные"]
    row = next(number for number in range(1, sheet.max_row + 1)
               if str(sheet[f"A{number}"].value or "") == "ДОО — очередь 1")
    base = evaluated(phased).cell("CAPEX", "B31")
    sheet[f"B{row}"] = sheet[f"B{row}"].value * 2
    assert evaluated(doubled).cell("CAPEX", "B31") == pytest.approx(base * 2, rel=1e-6)


def test_the_entered_payment_date_is_the_cell_and_the_rule_is_the_reader():
    """Дата компенсации — вводная, а обрезка по РнС живёт в формуле читателя.

    Прежде в клетку писался уже обрезанный результат: правка РнС в книге его
    не двигала, а из числа не было видно, что за ним стоит правило."""
    entered = book(phasing={}, social_mode=core.SOCIAL_MODE_BOTH,
                   social_comp_date="2028-01-01", social_compensation_mln=500)
    assert entered["Вводные"]["B18"].value.strftime("%Y-%m-%d") == "2028-01-01"

    sheet = entered["Вводные"]
    cash = next(number for number in range(1, sheet.max_row + 1)
                if str(sheet[f"A{number}"].value or "").startswith("Денежная компенсация"))
    formula = sheet[f"D{cash}"].value
    assert isinstance(formula, str) and "MIN" in formula and "$B$18" in formula


def test_the_deadline_wins_over_a_late_entered_date():
    """Методика движка: компенсация не выходит за бридж-период."""
    late = book(phasing={}, social_mode=core.SOCIAL_MODE_BOTH,
                social_comp_date="2035-01-01", social_compensation_mln=500)
    evaluator = evaluated(late)
    sheet = late["Вводные"]
    cash = next(number for number in range(1, sheet.max_row + 1)
                if str(sheet[f"A{number}"].value or "").startswith("Денежная компенсация"))
    # Вычислитель отдаёт то серийный номер, то дату — приводим к одной мере.
    def serial(value):
        if hasattr(value, "isoformat"):
            return core._v4_excel_serial(str(value)[:10])
        return float(value)

    paid = serial(evaluator.cell("Вводные", f"D{cash}"))
    permit = serial(evaluator.cell("CF_1", "B7"))
    assert paid < core._v4_excel_serial("2035-01-01"), paid
    assert paid < permit, (paid, permit)
