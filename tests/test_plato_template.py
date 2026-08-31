"""Тесты выгрузки в шаблон ПЛАТО.

Заполняются только листы-вводные, формулы шаблона не трогаются.
Запуск: python3 -m pytest tests -q
"""

from __future__ import annotations

import io
import sys
import zipfile
from pathlib import Path

import pytest
from fastapi import HTTPException

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import main as _wrapper  # noqa: E402

main = _wrapper.core

pytest.importorskip("openpyxl")
from openpyxl import load_workbook  # noqa: E402

TEMPLATE = Path(__file__).resolve().parent.parent / "templates" / "PLATO_template.xlsx"

pytestmark = pytest.mark.skipif(not TEMPLATE.is_file(), reason="шаблон ПЛАТО не поставляется")


def count_formulas(source) -> int:
    workbook = load_workbook(source, data_only=False)
    return sum(
        1
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    )


@pytest.fixture(scope="module")
def filled():
    content, report = main.fill_plato_template(main.DEFAULT_INPUTS, main.TEP_DEFAULT)
    return content, report, load_workbook(io.BytesIO(content), data_only=False)


def test_every_mapped_field_is_found(filled):
    _, report, _ = filled
    assert report["missing"] == []
    assert report["filled_count"] >= 80


def test_formulas_are_not_touched(filled):
    content, _, _ = filled
    # Баланс по листу «ЗУ»: две формулы окна платежей заменены датами,
    # три недостающие формулы первых месяцев рассрочки достроены.
    # Плюс четыре поля соцнагрузки: шаблон считал их сам с листа «Расчет ВРИ
    # (ТЭП)» и расходился с движком — по три сценарных колонки на поле.
    social = len(main._PLATO_OVERRIDE_TEMPLATE_FORMULA) * 3
    # Минус одна: в первой помесячной колонке погашения ПФ формула заменена
    # нулём — накопительный диапазон там вырождается в ссылку на себя.
    assert count_formulas(io.BytesIO(content)) == count_formulas(str(TEMPLATE)) - 2 + 3 - social - 1


def test_land_sheet_gets_the_vri_installment_window(filled):
    _, _, workbook = filled
    sheet = workbook["ЗУ"]
    labels = {
        str(sheet.cell(row=row, column=2).value or "").strip(): row
        for row in range(56, 70)
    }
    first = sheet.cell(row=labels["Первый"], column=3).value
    last = sheet.cell(row=labels["Последний"], column=3).value
    assert first is not None and last is not None
    assert last >= first
    assert sheet.cell(row=labels["Доля оплаты"], column=3).value == pytest.approx(1.0)
    assert str(sheet.cell(row=labels["В месяц"], column=3).value).startswith("=C")


def test_foreign_parcel_data_is_cleared(filled):
    """В шаблоне остались участок и суммы чужого проекта — их быть не должно."""
    _, _, workbook = filled
    sheet = workbook["ЗУ"]
    for row in range(22, 50):
        value = sheet.cell(row=row, column=3).value
        if isinstance(value, str):
            assert "Лётная" not in value
            assert "77:07:0013002:8740" not in value
        if isinstance(value, (int, float)):
            assert value != pytest.approx(1_512_108_174.46)


def test_all_sheets_survive(filled):
    _, _, workbook = filled
    source = load_workbook(str(TEMPLATE), data_only=False)
    assert workbook.sheetnames == source.sheetnames


def test_inputs_land_in_all_three_scenarios(filled):
    _, _, workbook = filled
    sheet = workbook["Вводные"]
    rows = {
        str(sheet.cell(row=row, column=2).value or "").strip(): row
        for row in range(1, sheet.max_row + 1)
    }
    row = rows["Стартовая цена квартир"]
    for column in (4, 5, 6):
        assert sheet.cell(row=row, column=column).value == pytest.approx(
            main.DEFAULT_INPUTS["apartment_price_th"]
        )


def test_percentages_are_written_as_shares(filled):
    _, _, workbook = filled
    sheet = workbook["Вводные"]
    for label, key in (("Налог на прибыль", "profit_tax_pct"), ("Маркетинг", "marketing_pct")):
        row = next(
            row for row in range(1, sheet.max_row + 1)
            if str(sheet.cell(row=row, column=2).value or "").strip() == label
        )
        assert sheet.cell(row=row, column=5).value == pytest.approx(main.DEFAULT_INPUTS[key] / 100)


def test_dates_are_written_as_dates(filled):
    _, _, workbook = filled
    sheet = workbook["Вводные"]
    row = next(
        row for row in range(1, sheet.max_row + 1)
        if str(sheet.cell(row=row, column=2).value or "").strip() == "Начало проекта"
    )
    value = sheet.cell(row=row, column=5).value
    assert value.strftime("%Y-%m-%d") == main.DEFAULT_INPUTS["project_start"]


def test_tep_sheet_receives_model_areas(filled):
    _, _, workbook = filled
    sheet = workbook["Расчет ВРИ (ТЭП)"]
    rows = {
        str(sheet.cell(row=row, column=2).value or "").strip(): row
        for row in range(1, sheet.max_row + 1)
    }
    assert sheet.cell(row=rows["Площадь квартир"], column=4).value == pytest.approx(
        main.TEP_DEFAULT["apartments"]["saleable"], abs=0.01
    )
    assert sheet.cell(row=rows["СПП жилая"], column=4).value == pytest.approx(
        main.TEP_DEFAULT["apartments"]["gns"], abs=0.01
    )
    # ТЭП!I33 шаблона складывает постоянные и гостевые. Значит в «постоянные»
    # идёт число мест за вычетом гостевых, иначе гостевые считаются дважды, —
    # а сумма двух строк обязана дать построенный паркинг целиком.
    total = main.TEP_DEFAULT["underground_parking"]["units"]
    guest = sheet.cell(row=rows["Гостевые парковки"], column=4).value
    permanent = sheet.cell(row=rows["Постоянные парковки"], column=4).value
    assert guest == pytest.approx(round(total / 11.0), abs=0.01)
    assert permanent + guest == pytest.approx(total, abs=0.01)


def test_price_growth_target_reproduces_the_model_monthly_growth(filled):
    """Шаблон выводит месячный рост из целевого совокупного — пишем цель обратным счётом.

    Пока эта строка не заполнялась, в шаблоне оставались 30% сценария, а модель
    считала по своим 1,5% в месяц: на 24 месяцах продаж выручка расходилась
    примерно на четверть.
    """
    _, _, workbook = filled
    sheet = workbook["Вводные"]
    rows = {main._plato_normalize(sheet.cell(row=r, column=2).value): r
            for r in range(1, sheet.max_row + 1)}
    row = rows["целевой совокупный рост цены от старта продаж до рвэ"]
    monthly = main.DEFAULT_INPUTS["monthly_growth_pre_pct"] / 100
    months = main.DEFAULT_INPUTS["construction_months"] - main.DEFAULT_INPUTS["sales_lag_months"]
    for column in (4, 5, 6):
        target = sheet.cell(row=row, column=column).value
        assert target == pytest.approx((1 + monthly) ** months - 1, rel=1e-6)
        # Обратный ход формулы шаблона должен вернуть ровно наш месячный рост.
        assert (1 + target) ** (1 / months) - 1 == pytest.approx(monthly, rel=1e-9)


def test_standalone_objects_get_their_growth_and_dates(filled):
    _, _, workbook = filled
    sheet = workbook["Вводные"]
    found = {}
    for row in range(1, sheet.max_row + 1):
        block = main._plato_normalize(sheet.cell(row=row, column=1).value)
        label = main._plato_normalize(sheet.cell(row=row, column=2).value)
        if block.startswith("мфоц") and label == "ежемесячный рост цены до рвэ":
            found["pre"] = sheet.cell(row=row, column=4).value
        if block.startswith("мфоц") and label == "ежемесячный рост цены после рвэ":
            found["post"] = sheet.cell(row=row, column=4).value
    assert found["pre"] == pytest.approx(main.DEFAULT_INPUTS["offices_growth_pre_pct"] / 100)
    assert found["post"] == pytest.approx(main.DEFAULT_INPUTS["offices_growth_post_pct"] / 100)


def test_project_name_replaces_the_template_leftover():
    """В шаблоне в шапке ОТЧЕТа зашит чужой проект — он не должен уезжать заказчику."""
    stale = load_workbook(TEMPLATE, data_only=False)["ОТЧЕТ"]["C1"].value
    content, _ = main.fill_plato_template(
        main.DEFAULT_INPUTS, main.TEP_DEFAULT, project_name="Мытищи"
    )
    assert load_workbook(io.BytesIO(content))["ОТЧЕТ"]["C1"].value == "Мытищи" != stale


def test_vri_cost_comes_from_the_model(filled):
    _, _, workbook = filled
    sheet = workbook["Расчет ВРИ (ТЭП)"]
    row = next(
        row for row in range(1, sheet.max_row + 1)
        if str(sheet.cell(row=row, column=2).value or "").strip() == "Многоквартирная жилые здания"
    )
    assert sheet.cell(row=row, column=4).value == pytest.approx(
        main.DEFAULT_INPUTS["land_rights_cost_mln"], abs=0.01
    )


def test_workbook_recalculates_on_open(filled):
    _, _, workbook = filled
    assert workbook.calculation.fullCalcOnLoad is True


def test_changed_input_reaches_the_template():
    content, _ = main.fill_plato_template(
        {**main.DEFAULT_INPUTS, "apartment_price_th": 999}, main.TEP_DEFAULT
    )
    sheet = load_workbook(io.BytesIO(content), data_only=False)["Вводные"]
    row = next(
        row for row in range(1, sheet.max_row + 1)
        if str(sheet.cell(row=row, column=2).value or "").strip() == "Стартовая цена квартир"
    )
    assert sheet.cell(row=row, column=5).value == 999


# --- архив ------------------------------------------------------------------

def test_missing_template_is_reported(tmp_path):
    with pytest.raises(HTTPException) as exc:
        main.fill_plato_template(
            main.DEFAULT_INPUTS, main.TEP_DEFAULT, template_path=tmp_path / "нет.xlsx"
        )
    assert exc.value.status_code == 503
    assert "шаблон" in str(exc.value.detail).lower()


if __name__ == "__main__":
    raise SystemExit(pytest.main([__file__, "-q"]))


def test_first_two_columns_of_the_vri_installment_are_repaired(filled):
    """В шаблоне D65 задан статикой и роняет всю плату в первый месяц модели,
    а D66/E66 пустые — проценты первых двух месяцев не начисляются."""
    _, _, workbook = filled
    sheet = workbook["ЗУ"]
    assert sheet["D65"].value == "=IF(D19=$C$63,$C$60,0)"
    assert sheet["E65"].value == "=IF(E19=$C$63,$C$60-SUM($D64:D64),0)"
    assert sheet["D66"].value == "=IF(D19>=$C$62,$C$60*D61/12,0)"
    assert sheet["E66"].value.startswith("=IF(E19>=$C$62,($C$60-SUM($D64:D65))*E61/12")
    # Статическая ссылка на лист ТЭП убрана.
    assert "Расчет ВРИ (ТЭП)'!D73" not in str(sheet["D65"].value)
