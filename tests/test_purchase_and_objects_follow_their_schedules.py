"""График платежей за покупку и графики продаж объектов — в движке и в книге.

«Добавить график платежей в стоимость покупки, с суммами и сроками: что-то
платится сразу, часть по графику, или в цену входит выкуп чужих объектов по
КРТ со своим графиком. Офисник и ТЦ — чтобы можно было настроить график продаж
(все 100% на старте или долю сразу, долю потом) и лестницу цены по этапам»
(владелец, 02.09.2026). Лестница — этапами строительной готовности, по
образцу полей «Рост цены — этап 1–4» шаблона ПЛАТО («не сносить этапы, а
сделать на их примере»): этап k — готовность k×25%.

Методику меняют в двух местах — в движке и в книге v4: строки покупки CAPEX и
строки объёма и цены листа ОБЪЕКТЫ читают тот же график из блока внизу
«Вводных». Паритет проверяется вычислителем формул, а не глазами.

Запуск: python3 -m pytest tests/test_purchase_and_objects_follow_their_schedules.py -q
"""

from __future__ import annotations

import io
import sys
from datetime import date
from pathlib import Path

import openpyxl
import pytest
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "tests"))

import main_legacy as core  # noqa: E402
from xlsx_eval import Evaluator  # noqa: E402

SCRATCH = ROOT / "tests" / "_scratch_schedules"


def _inputs(**extra):
    x = dict(core.DEFAULT_INPUTS)
    x.update({"purchase_price_mln": 1000})
    x.update(extra)
    return x


def _tep():
    return {key: dict(value) for key, value in core.TEP_DEFAULT.items()}


# --- разбор ------------------------------------------------------------------

def test_the_schedule_is_read_in_both_forms_and_a_mix_is_refused() -> None:
    items, percent, warnings = core.parse_month_schedule("30%@0; 40%@6, 30%@12")
    assert items == [(30.0, 0), (40.0, 6), (30.0, 12)] and percent and not warnings
    items, percent, _ = core.parse_month_schedule("500@0; 300,5@12")
    assert items == [(500.0, 0), (300.5, 12)] and not percent
    # Десятичная запятая внутри доли не путается с разделителем пар.
    items, _, _ = core.parse_month_schedule("12,5%@6; 87,5%@12")
    assert items == [(12.5, 6), (87.5, 12)]
    mixed, _, warnings = core.parse_month_schedule("30%@0; 500@6")
    assert mixed == [] and "смешивает" in warnings[0]
    garbage, _, warnings = core.parse_month_schedule("через полгода половину")
    assert garbage == [] and "не распознан" in warnings[0]
    assert core.parse_month_schedule("") == ([], False, [])


def test_shares_are_normalised_and_amount_remainders_are_named() -> None:
    start, end = date(2027, 1, 1), date(2031, 1, 1)
    plan, warnings = core.purchase_payment_plan(1e9, "30%@0; 30%@6", start, end)
    assert [round(a / 1e6) for _, a in plan] == [500, 500]
    assert any("приведены к 100%" in w for w in warnings)
    plan, warnings = core.purchase_payment_plan(1e9, "300@6; 300@12", start, end)
    assert plan[0] == (start, pytest.approx(400e6)) and any("остаток" in w for w in warnings)
    # Платёж за горизонтом не исчезает — ложится в последний месяц и назван.
    plan, warnings = core.purchase_payment_plan(1e9, "50%@0; 50%@120", start, end)
    assert plan[-1][0] == end and any("за горизонтом" in w for w in warnings)
    # Пустой график — всё в дату сделки, как было.
    assert core.purchase_payment_plan(1e9, "", start, end) == ([(start, 1e9)], [])


# --- движок ------------------------------------------------------------------

def test_the_purchase_is_paid_by_the_schedule_and_the_total_is_kept() -> None:
    op = core.build_operating_model(_inputs(purchase_schedule="30%@0; 40%@6; 30%@12"), _tep(), [])
    paid = {when.isoformat(): round(value / 1e6, 3) for when, value in op["capex_by_article"]["purchase"].items()}
    assert paid == {"2027-01-01": 300.0, "2027-07-01": 400.0, "2028-01-01": 300.0}
    assert sum(op["capex_by_article"]["purchase"].values()) == pytest.approx(1e9)
    rows = op["purchase_schedule"]["rows"]
    assert [row["offset_months"] for row in rows] == [0, 6, 12]
    assert sum(row["share"] for row in rows) == pytest.approx(1.0)
    plain = core.build_operating_model(_inputs(), _tep(), [])
    assert plain["capex_by_article"]["purchase"] == {date(2027, 1, 1): pytest.approx(1e9)}
    assert plain["purchase_schedule"]["custom"] is False


def test_a_deferred_purchase_needs_less_bridge_at_the_start() -> None:
    upfront = core._run_authoritative_model(_inputs(), _tep(), [], {})["consolidated"]
    deferred = core._run_authoritative_model(
        _inputs(purchase_schedule="20%@0; 80%@12"), _tep(), [], {})["consolidated"]
    # Лимит — формула банка от всей цены: не меняется. Пик выборки — меняется.
    assert deferred["report"]["financing"]["calculated_bridge"] == pytest.approx(
        upfront["report"]["financing"]["calculated_bridge"])
    assert deferred["finance"]["peak_bridge"] <= upfront["finance"]["peak_bridge"]
    assert deferred["report"]["purchase"]["custom"] is True


def test_the_object_sells_by_its_profile_and_prices_by_its_ladder() -> None:
    op = core.build_operating_model(_inputs(
        offices_enabled=True, offices_sales_profile="60%@0; 40%@12",
        offices_growth_stage1_pct=10, offices_growth_stage2_pct=10), _tep(), [])
    qty = {when.isoformat(): round(v, 1) for when, v in sorted(op["quantity_product_schedules"]["offices"].items())}
    rev = {when.isoformat(): round(v / 1e6, 1) for when, v in sorted(op["revenue_product_schedules"]["offices"].items())}
    # 6 000 м² × 60% по стартовой 500; 40% через год — стройка 24 месяца с
    # июля 2028, готовность 25% в январе и 50% в июле 2029: 500 × 1,1 × 1,1.
    assert qty == {"2028-07-01": 3600.0, "2029-07-01": 2400.0}
    assert rev == {"2028-07-01": 1800.0, "2029-07-01": 1452.0}
    notes = op["object_schedule_notes"]["offices"]
    assert notes["profile_applied"] and notes["steps_applied"] and not notes["warnings"]
    assert notes["price_ladder"] == "этап 1 · 25% · 01.2029: +10%; этап 2 · 50% · 07.2029: +10%"
    assert [row["date"] for row in notes["price_stages"]] == [
        "2029-01-01", "2029-07-01", "2030-01-01", "2030-07-01"]


def test_everything_at_the_start_is_one_month() -> None:
    op = core.build_operating_model(_inputs(retail_enabled=True, retail_sales_profile="100%@0"), _tep(), [])
    assert len(op["quantity_product_schedules"]["standalone_retail"]) == 1


def test_an_empty_profile_changes_nothing() -> None:
    before = core.build_operating_model(_inputs(offices_enabled=True), _tep(), [])
    after = core.build_operating_model(_inputs(offices_enabled=True, offices_sales_profile="",
                                               offices_growth_stage1_pct=0, offices_growth_stage3_pct=""), _tep(), [])
    assert before["revenue_product_schedules"]["offices"] == after["revenue_product_schedules"]["offices"]


def test_sums_in_a_profile_are_refused_and_named() -> None:
    op = core.build_operating_model(_inputs(offices_enabled=True, offices_sales_profile="3000@0; 3000@6"), _tep(), [])
    notes = op["object_schedule_notes"]["offices"]
    assert notes["profile_applied"] is False and any("долями" in w for w in notes["warnings"])
    product = [p for p in core._run_authoritative_model(
        _inputs(offices_enabled=True, offices_sales_profile="3000@0; 3000@6"), _tep(), [], {}
    )["consolidated"]["report"]["products"] if p["key"] == "offices"][0]
    assert product["schedule_warnings"], "оговорка разбора до отчёта не доехала"


# --- книга -------------------------------------------------------------------

def _book(x):
    data, _name, meta = core.build_project_workbook(x, _tep(), [], {}, project_name="график")
    assert meta["missing"] == [], meta["missing"]
    return Evaluator(openpyxl.load_workbook(io.BytesIO(data), data_only=False))


def _months(ev, sheet, row, first=4, count=60):
    out = {}
    for column in range(first, first + count):
        letter = get_column_letter(column)
        value = float(ev.cell(sheet, f"{letter}{row}") or 0)
        if abs(value) > 1e-6:
            when = ev.cell("Ставки", f"{letter}3")
            out[str(when)[:10]] = round(value, 3)
    return out


def test_the_book_pays_the_purchase_by_the_same_schedule() -> None:
    x = _inputs(purchase_schedule="30%@0; 40%@6; 30%@12")
    ev = _book(x)
    assert _months(ev, "CAPEX", 14) == {"2027-01-01": 300.0, "2027-07-01": 400.0, "2028-01-01": 300.0}
    # Блок графика лежит в «Вводных» и подписан — его правят в книге.
    values = [str(ev.workbook["Вводные"].cell(row=r, column=1).value or "")
              for r in range(1, ev.workbook["Вводные"].max_row + 1)]
    assert "ГРАФИК ПЛАТЕЖЕЙ ЗА ПОКУПКУ" in values


def test_the_book_sells_the_object_by_the_same_profile_and_ladder() -> None:
    x = _inputs(offices_enabled=True, offices_sales_profile="60%@0; 40%@12",
                offices_growth_stage1_pct=10, offices_growth_stage2_pct=10)
    ev = _book(x)
    assert _months(ev, "ОБЪЕКТЫ", 22) == {"2028-07-01": 3600.0, "2029-07-01": 2400.0}
    assert _months(ev, "ОБЪЕКТЫ", 24) == {"2028-07-01": 1800.0, "2029-07-01": 1452.0}
    assert float(ev.cell("ОБЪЕКТЫ", "B24")) == pytest.approx(3252.0)


def test_amounts_in_the_purchase_schedule_reach_the_book_with_the_remainder() -> None:
    ev = _book(_inputs(purchase_schedule="300@6; 300@12"))
    assert _months(ev, "CAPEX", 14) == {"2027-01-01": 400.0, "2027-07-01": 300.0, "2028-01-01": 300.0}


def test_an_unrecognised_template_row_goes_to_missing(monkeypatch) -> None:
    """Формула шаблона не та — книга не считает молча иначе, чем движок."""
    xml = '<x:c r="D14" s="1" t="n"><x:f>SOMETHING($B$14)</x:f></x:c>'
    missing: list[str] = []
    out = core._v4_apply_purchase_schedule(xml, [("\'Вводные\'!$B$130", "\'Вводные\'!$C$130")], missing)
    assert out == xml and any("очередь 1" in item for item in missing)


# --- поверхности -------------------------------------------------------------

def test_the_page_has_the_fields_and_prints_the_schedules() -> None:
    fields = {f[0]: f for _group, items in core.FIELD_GROUPS for f in items}
    for key in ("purchase_schedule", "offices_sales_profile",
                "retail_sales_profile", "above_parking_sales_profile"):
        # Тип поля — «schedule»: график вводится ячейками, а хранится той же
        # строкой, которую читают и движок, и книга. Лестница цены сюда не
        # входит: у неё именованные поля этапов по строительной готовности —
        # «как в квартирах блок Этап и процент» (владелец, 02.09.2026).
        assert key in fields and fields[key][3] == "schedule", key
        assert core.DEFAULT_INPUTS[key] == ""
    for prefix in ("", "offices_", "retail_", "above_parking_"):
        for stage in (1, 2, 3, 4):
            key = f"{prefix}growth_stage{stage}_pct"
            assert key in fields and fields[key][3] == "number", key
            assert fields[key][1] == f"Рост цены — этап {stage}"
            assert core.DEFAULT_INPUTS[key] == 0
    page = core.PAGE
    assert "График платежей за покупку" in page
    assert "p.sales_profile" in page and "p.price_ladder" in page and "p.schedule_warnings" in page


def test_the_pdf_prints_the_schedule_and_the_object_notes() -> None:
    pypdf = pytest.importorskip("pypdf")
    x = _inputs(purchase_schedule="30%@0; 40%@6; 30%@12", offices_enabled=True,
                offices_sales_profile="60%@0; 40%@12", offices_growth_stage1_pct=10)
    result = core._run_authoritative_model(x, _tep(), [], {})["consolidated"]
    pdf = core._build_developaid_pdf({"result": result, "inputs": x, "tep": _tep(), "rates": [],
                                      "phasing": {}, "scenario": "base", "project_name": "Т"})
    text = " ".join(" ".join((page.extract_text() or "").split())
                    for page in pypdf.PdfReader(io.BytesIO(pdf)).pages)
    assert "График платежей за покупку" in text and "01.07.2027" in text
    assert "профиль продаж 60%@0; 40%@12" in text
    assert "лестница цены этап 1 · 25% · 01.2029: +10%" in text


# --- квартиры ----------------------------------------------------------------

STAGES = dict(growth_stage1_pct=5, growth_stage2_pct=5, growth_stage3_pct=10, growth_stage4_pct=0)


def test_the_apartment_ladder_prices_all_core_products() -> None:
    """«Делать и квартирам — на них и строилась эта идея» (владелец, 02.09.2026).
    Одна лестница на четыре основных продукта, как и ежемесячный рост.
    Стройка 24 месяца с июля 2028: этапы в январе и июле 2029 и 2030."""
    op = core.build_operating_model(_inputs(**STAGES), _tep(), [])
    for key in ("apartments", "ground_commercial", "underground_parking", "storage"):
        rev, qty = op["revenue_product_schedules"][key], op["quantity_product_schedules"][key]
        if not rev:
            continue
        price = {when.isoformat(): rev[when] / qty[when] for when in rev}
        start_price = price["2028-07-01"]
        assert price["2028-12-01"] == pytest.approx(start_price)
        assert price["2029-01-01"] == pytest.approx(start_price * 1.05)
        assert price["2029-07-01"] == pytest.approx(start_price * 1.05 * 1.05)
        assert price["2030-01-01"] == pytest.approx(start_price * 1.05 * 1.05 * 1.10)
        # Этап 4 нулевой; после РВЭ (июль 2030) продолжает ежемесячный рост.
        assert price["2030-08-01"] == pytest.approx(start_price * 1.05 * 1.05 * 1.10 * 1.0025)
    assert op["object_schedule_notes"]["core"]["steps_applied"] is True
    product = [p for p in core._run_authoritative_model(
        _inputs(**STAGES), _tep(), [], {})["consolidated"]["report"]["products"]
        if p["key"] == "apartments"][0]
    assert product["price_ladder"].startswith("этап 1 · 25% · 01.2029: +5%") and product["schedule_applied"]


def test_the_stage_months_follow_the_construction_term() -> None:
    """Этап — четверть срока строительства, не зашитый месяц: 30 месяцев дают
    8/15/23/30, и ни один этап не наступает раньше начала стройки."""
    factor, rows, warnings = core.stage_ladder_factor([2, 0, 0, 3], date(2028, 7, 1), 30)
    assert [row["date"] for row in rows] == [
        date(2029, 3, 1), date(2029, 10, 1), date(2030, 6, 1), date(2031, 1, 1)]
    assert factor(date(2028, 7, 1)) == 1.0
    assert factor(date(2029, 3, 1)) == pytest.approx(1.02)
    assert factor(date(2031, 1, 1)) == pytest.approx(1.02 * 1.03)
    assert not warnings
    assert core.stage_ladder_factor([0, "", None, 0], date(2028, 7, 1), 24)[0] is None
    _factor, _rows, warnings = core.stage_ladder_factor(["много", 5, 0, 0], date(2028, 7, 1), 24)
    assert warnings and "этап 1" in warnings[0]


def test_the_book_prices_apartments_by_the_same_ladder() -> None:
    x = _inputs(**STAGES)
    op = core.build_operating_model(x, _tep(), [])
    rev, qty = op["revenue_product_schedules"]["apartments"], op["quantity_product_schedules"]["apartments"]
    engine = {when.isoformat(): round(rev[when] / qty[when] / 1000, 3) for when in rev}
    ev = _book(x)
    book = {}
    for column in range(4, 64):
        letter = get_column_letter(column)
        if float(ev.cell("Продажи", f"{letter}14") or 0) > 1e-6:
            book[str(ev.cell("Ставки", f"{letter}3"))[:10]] = round(float(ev.cell("Продажи", f"{letter}15")), 3)
    assert book == engine
    # Блок этапов подписан готовностью, месяц в нём не хранится — его считает
    # формула из срока строительства очереди.
    values = [str(ev.workbook["Вводные"].cell(row=r, column=1).value or "")
              for r in range(1, ev.workbook["Вводные"].max_row + 1)]
    assert "Этап 1 · готовность 25%" in values


def test_the_stage_fields_are_the_ladder_and_the_engine_reads_them() -> None:
    """«Не сносить этапы, а сделать на их примере» (владелец, 02.09.2026):
    поля шаблона ПЛАТО остались и стали рабочими, текстовой лестницы нет."""
    fields = {f[0]: f for _group, items in core.FIELD_GROUPS for f in items}
    assert "price_steps" not in fields and "offices_price_steps" not in fields
    assert "price_steps" not in core.DEFAULT_INPUTS
    for key in ("growth_stage1_pct", "growth_stage2_pct", "growth_stage3_pct", "growth_stage4_pct"):
        assert key in fields and key in core.DEFAULT_INPUTS
        assert key not in core._M2_TEMPLATE_ONLY_INPUTS, f"{key} помечен как нечитаемый движком"
    # Заданный этап замещает ежемесячный рост до РВЭ: без него цену ведёт
    # 1,5% в месяц, с ним — одна ступень.
    flat = _inputs(monthly_growth_pre_pct=0)
    before = core.build_operating_model(flat, _tep(), [])["revenue_by_product"]["apartments"]
    after = core.build_operating_model({**flat, "growth_stage2_pct": 10}, _tep(), [])["revenue_by_product"]["apartments"]
    assert after > before
    # Шаблон ПЛАТО принимает этапы своими строками «Вводных».
    assert {label for label, _key, _kind in core._PLATO_INPUT_MAP} >= {
        "Рост цены — этап 1", "Рост цены — этап 2", "Рост цены — этап 3", "Рост цены — этап 4"}
