"""Кабинет принимает файл ЦФ и отдаёт свод продаж одним вызовом.

Один файл — один разбор. ЦФ несёт и «Контрактацию», и «1С_Факт», и сверка между
ними идёт этим же вызовом: просить загрузить один файл дважды значит однажды
получить два разных файла и показать их как один проект — правило кабинета,
записанное ещё для плана продаж.

Отсутствие проводок 1С — не ошибка: у выгрузки без этого листа есть
контрактация, и это законный свод. Причина идёт строкой рядом, а не пятисоткой
поверх удавшегося разбора.

Запуск: python3 -m pytest tests/test_the_cabinet_takes_the_project_file.py -q
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

import pytest
from fastapi.testclient import TestClient

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

KEY = "cabinet-probe-key"


@pytest.fixture()
def client(monkeypatch):
    monkeypatch.setenv("MARKET_CABINET_KEY", KEY)
    import main_registry
    api = TestClient(main_registry.app)
    api.post("/cabinet/login", content=f"key={KEY}")
    return api


def _workbook() -> bytes:
    """Лист «Контрактация» той же формы, что в выгрузке ЦФ.

    Реестра владельца в репозитории нет и не будет: в нём ФИО покупателей.
    """
    from openpyxl import Workbook
    book = Workbook()
    page = book.active
    page.title = "Контрактация"
    page["A3"], page["E3"] = "Название проекта:", "Проверочный ЖК"
    header = ["Квартал", "Год", "Месяц", "Объект недвижимости", "Корпус",
              "Тип объекта недвижимости", "Вид отделки", "Проектная S", "Договор",
              "Тип договора", "Состояние договора", "Сумма договор", "Дата договора",
              "Покупатель", "Шт", "Цена ДДУ за кв. м", "Вариант оплаты",
              "Оплачено премии ОП ", "Брокеры", "% брокера", "Наименование брокера",
              "% оплаты", "остаток к оплате", "оплачено эскроу всего"]
    for column, title in enumerate(header, start=1):
        page.cell(row=6, column=column, value=title)
    rows = [
        ["3Q 2025", 2025, 45900, "Квартира", "корп.1 кв.1", "жилая", "без отделки",
         40.0, "ДДУ №1-1/П от 12.08.2025", "ДДУ", "Действующий", 24_000_000,
         "12.08.2025", "Иванов Иван", 1, 600_000, "100% оплата", 0,
         1_200_000, 0.05, "Брокер ООО", 1.0, 0, 24_000_000],
        ["3Q 2025", 2025, 45900, "Квартира", "корп.1 кв.2", "жилая", "без отделки",
         60.0, "ДДУ №1-2/П от 20.08.2025", "ДДУ", "Действующий", 36_000_000,
         "20.08.2025", "Ромашка ООО", 1, 600_000, "Рассрочка 20% ПВ", 0,
         0, None, "", 0.25, 27_000_000, 9_000_000],
    ]
    for index, values in enumerate(rows, start=7):
        for column, value in enumerate(values, start=1):
            page.cell(row=index, column=column, value=value)
    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def test_the_summary_comes_back_from_one_upload(client) -> None:
    """Файл телом запроса — как у плана продаж: multipart тянет лишнюю зависимость."""
    answer = client.post("/cabinet/contracting", content=_workbook())
    assert answer.status_code == 200, answer.text
    got = answer.json()
    assert got["project"] == "Проверочный ЖК"
    assert got["total"]["contracts"] == 2
    assert got["total"]["amount"] == 60_000_000


def test_a_missing_ledger_is_a_note_not_a_failure(client) -> None:
    """У выгрузки без листа проводок есть контрактация, и это законный свод."""
    got = client.post("/cabinet/contracting", content=_workbook()).json()
    assert got["terminated"] == []
    assert any("1С" in str(note) for note in got["missing"]), got["missing"]


def test_an_empty_body_is_refused(client) -> None:
    assert client.post("/cabinet/contracting", content=b"").status_code == 422


def test_a_file_without_the_sheet_says_which_sheets_there_are(client) -> None:
    """Отказ обязан называть причину: лист чужой и может быть переименован."""
    from openpyxl import Workbook
    book = Workbook()
    book.active.title = "Не тот лист"
    buffer = io.BytesIO()
    book.save(buffer)
    answer = client.post("/cabinet/contracting", content=buffer.getvalue())
    assert answer.status_code == 422
    assert "Контрактация" in answer.json()["detail"]


def test_the_cabinet_is_closed_without_the_key(monkeypatch) -> None:
    """Свод показывает продажи проекта — без ключа он не открывается."""
    monkeypatch.setenv("MARKET_CABINET_KEY", KEY)
    import main_registry
    api = TestClient(main_registry.app)
    assert api.post("/cabinet/contracting", content=_workbook()).status_code == 401


def test_a_non_ascii_key_is_refused_not_a_crash(monkeypatch) -> None:
    """Ключ с кириллицей ронял вход пятисоткой.

    `compare_digest` не работает со строками, где есть не-ASCII, а
    `/cabinet/login` звал его напрямую — там, где сам кабинет объясняет
    проблему словами, логин отвечал ошибкой сервера. Отказ и поломка выглядели
    одинаково.
    """
    monkeypatch.setenv("MARKET_CABINET_KEY", "ключ-кириллицей")
    import market_search.cabinet as cabinet
    assert cabinet.key_accepted("ключ-кириллицей") is True
    assert cabinet.key_accepted("другой") is False
