"""Кэш-свип: банк забирает долю поступлений очереди, передавшей свой долг.

Владелец, 31.08.2026: «пока строится 2 очередь, копятся деньги от первой. В
первой не погасили 10 ярдов, они ушли на вторую. Вторая не смогла погасить 5.
Но есть продажи первой после перевода долга — возможно их хватит, и тогда эти
деньги банк заберёт в погашение».

До этой правки не учитывались ни накопленная касса проекта, ни остаточные
продажи передавшей очереди: при переносе её линия закрывается, и продажи
уходили застройщику целиком, пока долг ехал дальше. Касса проекта в движке
была, но тратилась только на расходы до РнС следующей очереди и только при
стратегии «единая касса» — на погашение долга не шла никогда.

Механика — договорная (владелец, 31.08.2026): «обычно бывает аналог кэш-свипа.
80% проданного забирает сразу банк, ему их надо напрямую перечислять в
погашение долга, а 20% оставляет застройщик» — и эти 20% остаются резервом
проекта, а не карманом. Доля берётся от ВЫРУЧКИ, а не от свободной кассы:
банк смотрит на поступления на счёт.

Разовое сметание в дату раскрытия, сделанное первым заходом, снято: два
механизма на одно явление в этом проекте всегда расходились, а владелец сказал
прямо — «два механизма это сложновато для восприятия».

Запуск: python3 -m pytest tests/test_accumulated_sales_repay_the_carried_debt.py -q
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT.parent))

import main as _wrapper  # noqa: E402

core = _wrapper.core


def _bundle(sweep: bool) -> dict:
    inputs = dict(core.DEFAULT_INPUTS)
    # Доля продаж до РВЭ занижена нарочно: эскроу к раскрытию не наполняется,
    # а остаток продаётся после — ровно та картина, где накопленное решает.
    inputs.update(purchase_price_mln=12000, project_start="2027-01-01", ird_months=12,
                  construction_months=24, apartment_price_th=900,
                  share_before_rve_pct=45, residual_sales_months=18)
    tep = {key: dict(row) for key, row in core.TEP_DEFAULT.items()}
    phasing = {
        "enabled": True, "mode": "phased", "user_enabled": True,
        "phase_count": 2, "phase_gap_months": 12,
        "phases": [{"name": "О1", "start_offset_months": 0, "construction_months": 24},
                   {"name": "О2", "start_offset_months": 12, "construction_months": 24}],
        "products": {key: [35, 65] for key in
                     ("apartments", "ground_commercial", "underground_parking")},
        "shared_cash": {}, "shared_allocation": {}, "social_objects": [],
        "carry_debt_forward": True, "sweep_project_cash": sweep,
    }
    return core.calculate_phased(core.PhasedCalcRequest(
        inputs=inputs, tep=tep, rates=[], phasing=phasing))


@pytest.fixture(scope="module")
def without() -> dict:
    bundle = _bundle(False)
    second = bundle["phases"][1]["result"]["finance"]
    assert float(second.get("rve_unpaid") or 0.0) > 1e9, (
        "предохранитель: без сметания вторая очередь обязана не дотянуть, "
        "иначе сравнивать не с чем")
    assert second.get("default_date"), "предохранитель: дефолт обязан быть"
    return bundle


@pytest.fixture(scope="module")
def with_sweep() -> dict:
    return _bundle(True)


def test_it_is_on_unless_switched_off() -> None:
    """Умолчание — включено (владелец, 31.08.2026: «делать базово включенным»).

    И отдельно: отсутствующий ключ — «не задано», а не «снято». Проект,
    сохранённый до появления признака, приезжает без него, и прочитать это
    как выключенный значило бы молча считать его по прежней методике.
    """
    assert core._sweep_project_cash_enabled({}) is True
    assert core._sweep_project_cash_enabled({"sweep_project_cash": None}) is True
    assert core._sweep_project_cash_enabled({"sweep_project_cash": False}) is False
    # Экран читает признак тем же правилом, иначе галочка и расчёт разойдутся.
    assert "phasing.sweep_project_cash!==false" in core.PAGE
    assert "sweep_project_cash:true" in core.PAGE


def test_switching_it_off_brings_the_default_back(without) -> None:
    """Предохранитель: снятая галочка обязана возвращать прежнюю картину."""
    finance = without["phases"][1]["result"]["finance"]
    assert float(finance.get("project_cash_sweep") or 0.0) == 0.0
    assert finance.get("default_date")


def test_a_project_without_a_transfer_is_untouched() -> None:
    """Свип без переноса не бывает: долг остаётся на своей очереди.

    Умолчание «включено» поэтому ничего не меняет там, где перенос выключен —
    а он выключен по умолчанию. Без этой проверки включённый признак выглядел
    бы правкой всех многоочередных проектов разом.
    """
    inputs = dict(core.DEFAULT_INPUTS)
    tep = {key: dict(row) for key, row in core.TEP_DEFAULT.items()}
    bundle = core.calculate_phased(core.PhasedCalcRequest(
        inputs=inputs, tep=tep, rates=[],
        phasing={"enabled": True, "mode": "phased", "user_enabled": True,
                 "phase_count": 2, "phase_gap_months": 12,
                 "phases": [{"name": "О1"}, {"name": "О2"}],
                 "shared_cash": {}, "shared_allocation": {}, "social_objects": []}))
    for phase in bundle["phases"]:
        assert float(phase["result"]["finance"].get("bank_sweep_out") or 0.0) == 0.0


def test_the_sweep_closes_the_gap(without, with_sweep) -> None:
    gap = float(without["phases"][1]["result"]["finance"]["rve_unpaid"])
    applied = float(with_sweep["phases"][1]["result"]["finance"]["project_cash_sweep"])
    assert applied > gap, (applied, gap)
    assert not with_sweep["phases"][1]["result"]["finance"].get("default_date"), (
        "дефолт остался, хотя свип закрыл нехватку")


def test_the_share_is_taken_from_revenue(with_sweep) -> None:
    """80% поступлений, а не свободной кассы: банк смотрит на счёт."""
    first = with_sweep["phases"][0]["result"]
    rve = str(first["dates"]["rve"])
    sales = sum(float(row.get("sales") or 0.0) for row in first["finance"]["rows"]
                if str(row["month"]) > rve)
    out = float(first["finance"]["bank_sweep_out"])
    assert out == pytest.approx(sales * 0.8, rel=1e-6), (out, sales)
    assert float(first["finance"]["cash_sweep_pct"]) == pytest.approx(80.0)


def test_what_the_bank_did_not_need_comes_back(with_sweep) -> None:
    """Банк забирает не больше, чем ему должны, — остальное проекту.

    Иначе деньги растворяются: из кассы передавшей очереди вычтены, а долг
    ими не гасится.
    """
    out = float(with_sweep["phases"][0]["result"]["finance"]["bank_sweep_out"])
    finance = with_sweep["phases"][1]["result"]["finance"]
    applied = float(finance["project_cash_sweep"])
    returned = float(finance["bank_sweep_returned"])
    assert returned > 0, "предохранитель: на этих вводных возврат обязан быть"
    assert out == pytest.approx(applied + returned, abs=1_000_000.0)


def test_the_money_comes_from_the_earlier_queue_after_its_own_rve(with_sweep) -> None:
    """Это продажи О1 ПОСЛЕ передачи долга — иначе брать было бы нечего."""
    first = with_sweep["phases"][0]["result"]
    swept_month = with_sweep["phases"][1]["result"]["finance"]["project_cash_sweep_month"]
    assert swept_month > str(first["dates"]["rve"]), (swept_month, first["dates"]["rve"])
    # У передавшей очереди линия закрыта, и её продажи после РВЭ — свободная
    # касса проекта, а не погашение по своей линии.
    assert float(first["finance"]["ending_pf"]) == pytest.approx(0.0, abs=1.0)
    assert float(first["finance"]["debt_carried_out"]) > 1e9


def test_the_money_leaves_the_queue_that_earned_it(with_sweep) -> None:
    """Деньги ушли банку — поток на капитал передавшей очереди их теряет.

    Иначе они посчитаны дважды: остались бы у застройщика и одновременно
    погасили долг следующей очереди.
    """
    first = with_sweep["phases"][0]["result"]
    rve = str(first["dates"]["rve"])
    swept_months = [row for row in first["finance"]["rows"]
                    if float(row.get("bank_sweep_out") or 0.0) > 0]
    assert swept_months, "свип не пошёл вовсе"
    assert all(str(row["month"]) > rve for row in swept_months), (
        "свип начался раньше передачи долга")
    cashflow = first["cashflow"]
    for row in swept_months[:3]:
        index = cashflow["months"].index(str(row["month"]))
        equity = float(cashflow["equity"][index])
        without_sweep = equity + float(row["bank_sweep_out"])
        assert without_sweep > equity, "касса не потеряла отданное банку"


def test_the_debt_falls_month_by_month_not_in_one_lump(with_sweep) -> None:
    """Помесячно — в этом вся разница: долг гасится раньше, процентов меньше."""
    rows = [row for row in with_sweep["phases"][1]["result"]["finance"]["rows"]
            if float(row.get("project_cash_sweep") or 0.0) > 0]
    assert len(rows) >= 6, f"свип пришёл {len(rows)} раз — это не помесячно"


def test_it_makes_the_project_cheaper_not_richer(without, with_sweep) -> None:
    """Долг гасится раньше — процентов меньше. Выручка при этом та же.

    Проверка от обратного: если бы сметённое считалось доходом, выросла бы
    выручка, а она обязана совпасть до копейки.
    """
    a, b = without["consolidated"]["summary"], with_sweep["consolidated"]["summary"]
    assert b["revenue"] == pytest.approx(a["revenue"], rel=1e-9)
    assert b["financing_cost"] < a["financing_cost"]
    assert b["net_profit"] > a["net_profit"]


def test_the_plate_names_what_closed_the_gap() -> None:
    """Долг, исчезнувший без объяснения, читается как ошибка расчёта."""
    page = core.PAGE
    # Фраза собирается конкатенацией и целиком в исходнике не лежит; готовый
    # текст плашки гоняет через node test_the_rve_plate_speaks_by_queue.py.
    assert "банк забирал" in page
    assert "поступлений" in page
    # И признак на вкладке «Очерёдность», иначе включить его негде.
    assert 'id="phaseSweepCash"' in page
    assert "sweep_project_cash" in page


# --- книга ---------------------------------------------------------------------

def _book(sweep: bool, carry: bool = True):
    import io
    import openpyxl
    sys.path.insert(0, str(ROOT))
    from xlsx_eval import Evaluator
    sys.setrecursionlimit(400000)
    inputs = dict(core.DEFAULT_INPUTS)
    inputs.update(purchase_price_mln=12000, project_start="2027-01-01", ird_months=12,
                  construction_months=24, apartment_price_th=900,
                  share_before_rve_pct=45, residual_sales_months=18)
    tep = {key: dict(row) for key, row in core.TEP_DEFAULT.items()}
    phasing = {
        "enabled": True, "mode": "phased", "user_enabled": True,
        "phase_count": 2, "phase_gap_months": 12,
        "phases": [{"name": "О1", "start_offset_months": 0, "construction_months": 24},
                   {"name": "О2", "start_offset_months": 12, "construction_months": 24}],
        "products": {key: [35, 65] for key in
                     ("apartments", "ground_commercial", "underground_parking")},
        "shared_cash": {}, "shared_allocation": {}, "social_objects": [],
        "carry_debt_forward": carry, "sweep_project_cash": sweep, "cash_sweep_pct": 80,
    }
    content, _, meta = core.build_project_workbook(
        inputs, tep, [], phasing, project_name="Свип")
    return Evaluator(openpyxl.load_workbook(io.BytesIO(content))), meta


def test_the_workbook_sweeps_the_same_money(with_sweep) -> None:
    """Книга считает свип своей формулой — и обязана сойтись с движком.

    Иначе выгрузка и отчёт скажут разное про один проект, а это в модели, с
    которой идут в банк, худший из исходов.
    """
    book, meta = _book(sweep=True)
    assert meta["missing"] == [], meta["missing"]
    engine = float(with_sweep["phases"][1]["result"]["finance"]["project_cash_sweep"]) / 1e6
    assert float(book.cell("CF_2", "B26")) == pytest.approx(engine, rel=0.005)
    assert book.cell("ПРОВЕРКИ", "B3") == "ПРОЙДЕНО"


def test_the_workbook_share_comes_from_the_engine() -> None:
    """Доля лежит в книге числом: читатель видит, на чём посчитано."""
    book, _ = _book(sweep=True)
    assert float(book.cell("Вводные", "F92")) == pytest.approx(0.8)
    off, _ = _book(sweep=False)
    assert float(off.cell("Вводные", "F92") or 0.0) == 0.0
    assert float(off.cell("CF_2", "B26") or 0.0) == 0.0


def test_the_workbook_verdict_holds_in_every_combination() -> None:
    """Предохранитель: правка не должна ронять книгу там, где свипа нет."""
    for sweep, carry in ((True, True), (False, True), (False, False)):
        book, meta = _book(sweep=sweep, carry=carry)
        assert meta["missing"] == [], (sweep, carry, meta["missing"])
        assert book.cell("ПРОВЕРКИ", "B3") == "ПРОЙДЕНО", (sweep, carry)


def test_the_limit_check_knows_the_carried_debt_is_not_a_draw() -> None:
    """Принятый долг лимита не выбирает — проверка книги теперь это знает.

    Прежде она сравнивала пик ПФ с лимитом как есть и на любом проекте с
    переносом кричала «превышение», которого по методике нет. Кричащая зря
    проверка хуже отсутствующей: её перестают читать.
    """
    book, _ = _book(sweep=False, carry=True)
    assert book.cell("ПРОВЕРКИ", "F72") == "OK"
    source = Path(core.__file__).read_text(encoding="utf-8")
    assert "_v4_relax_limit_check_for_carried_debt" in source
