"""Proposal parser remains available, but proposals are not Project Monitor inputs.

Recovery schedules can still be parsed for other analytical workflows. The
weekly Project Monitor has a fixed PM/GPR baseline and is updated from RSS
6.1.2 only; it must never replace that baseline with a proposal.
"""

from __future__ import annotations

import datetime
import io

import pytest
from openpyxl import Workbook

import developaid_actuals as actuals


def _proposal_book(sheet="наше предложение"):
    book = Workbook()
    page = book.active
    page.title = sheet
    for column, name in enumerate(["ФАКТ ИЮЛЬ", "август", "сентябрь"], 6):
        page.cell(row=1, column=column, value=name)
    page.cell(row=3, column=3, value="Материал")
    page.cell(row=3, column=6, value=999e6)  # корпусная строка: не итог
    page.cell(row=10, column=5, value="Материал")
    page.cell(row=10, column=6, value=341.8e6)
    page.cell(row=10, column=7, value=146.9e6)
    page.cell(row=11, column=5, value="СМР")
    page.cell(row=11, column=7, value=9.0e6)
    page.cell(row=11, column=8, value=32.6e6)
    blob = io.BytesIO(); book.save(blob); blob.seek(0); return blob


def test_acceptance_is_the_smr_row_not_the_money():
    proposal = actuals.read_proposal(
        _proposal_book(), "наше предложение", "2026-07", "2.2.2.6.")
    assert proposal["code"] == "2.2.2.6"
    assert proposal["acceptance"] == {
        datetime.date(2026, 8, 1): pytest.approx(9.0e6),
        datetime.date(2026, 9, 1): pytest.approx(32.6e6),
    }
    assert proposal["payments_total"] == pytest.approx(530.3e6)


def test_the_corps_rows_are_not_summed():
    proposal = actuals.read_proposal(
        _proposal_book(), "наше предложение", "2026-07", "2.2.2.6")
    assert proposal["payments"][datetime.date(2026, 7, 1)] == pytest.approx(341.8e6)


def test_a_sheet_without_totals_is_refused():
    book = Workbook(); page = book.active; page.title = "наше предложение"
    page.cell(row=1, column=6, value="июль")
    blob = io.BytesIO(); book.save(blob); blob.seek(0)
    with pytest.raises(ValueError):
        actuals.read_proposal(blob, "наше предложение", "2026-07", "2.2.2.6")


def test_the_first_month_comes_from_outside():
    with pytest.raises(ValueError):
        actuals.read_proposal(_proposal_book(), "наше предложение", "", "2.2.2.6")


def test_apply_proposals_still_works_as_a_generic_helper():
    """Kept for non-Monitor uses; Monitor itself no longer calls this helper."""
    programme = {
        "by_code": {"2.2.2.6": {datetime.date(2026, 8, 1): 200e6}},
        "leaves": {"2.2.2.6"}, "months": [datetime.date(2026, 8, 1)],
        "first": datetime.date(2026, 8, 1), "last": datetime.date(2026, 8, 1),
    }
    replaced = actuals.apply_proposals(programme, [{
        "code": "2.2.2.6", "taken_at": "2026-08-11",
        "acceptance": {"2026-08": 9.0e6, "2026-09": 32.6e6},
    }])
    assert replaced["by_code"]["2.2.2.6"] == {
        datetime.date(2026, 8, 1): pytest.approx(9.0e6),
        datetime.date(2026, 9, 1): pytest.approx(32.6e6),
    }
