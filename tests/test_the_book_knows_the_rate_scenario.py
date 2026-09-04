"""Книга считает ту кривую ключевой ставки, что и движок, — включая сценарий.

Владелец, 04.09.2026: «ты эксель доделал? все ячейки?» Считать пришлось
дважды: первый счёт занизил охват, потому что ключи очередей стоят подстрокой
в шапке колонки, а блоки графиков рисуются только при заполненном поле.

Настоящая находка была одна, и она про числа, а не про подписи. Ячейка `B6` —
общий драйвер сценария книги — закреплена движком словом «Base», и это ВЕРНО:
в колонку Base он пишет ПРИМЕНЁННЫЕ множители цены и затрат. Но по тому же
`B6` выбиралась и целевая ключевая ставка, а её сценарий — отдельная вводная
`rate_scenario`. Проект на консервативном сценарии книга вела по базовой
целевой: 12,03 → 9,00% против 12,82 → 11,00% у движка, до 2 п.п. по всей
длине. Тесты паритета в эту ветку не заходят — они идут на умолчаниях, а там
сценарий как раз базовый.

Рядом две вводные стояли ЧИСЛАМИ внутри формулы листа «Ставки» (`EXP(-2*…)` и
смещение `+5`): числа движок подставлял верные, но править их в книге было
негде — «если величина в книге не выражается формулой, значит не хватает не
формулы, а вводной».

Запуск: python3 -m pytest tests/test_the_book_knows_the_rate_scenario.py -q
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "tests"))

import main_legacy as core  # noqa: E402

import v4_inputs  # noqa: E402

openpyxl = pytest.importorskip("openpyxl")
from openpyxl.utils import get_column_letter  # noqa: E402

from xlsx_eval import Evaluator  # noqa: E402

SCENARIOS = ("base", "high", "low")


def book(scenario: str, **over):
    inputs = {**core.DEFAULT_INPUTS, "rate_scenario": scenario, **over}
    content, _, meta = core.build_project_workbook(
        inputs, core.TEP_DEFAULT, [], {}, project_name="Ставка")
    assert meta["missing"] == [], meta["missing"]
    return content, inputs


def cells(content) -> dict[str, tuple[str, object]]:
    """Ключ движка → (координата, значение). Ключ живёт рядом со значением."""
    sheet = v4_inputs.inputs(openpyxl.load_workbook(io.BytesIO(content), data_only=False))
    out: dict[str, tuple[str, object]] = {}
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip():
                out.setdefault(cell.value.strip(),
                               (cell.coordinate, sheet[f"B{cell.row}"].value))
    return out


@pytest.mark.parametrize("scenario", SCENARIOS)
def test_the_curve_matches_the_engine_under_every_scenario(scenario) -> None:
    """Тот самый случай, ради которого проверка написана: на умолчаниях
    сценарий базовый, и расхождение видно только на двух других."""
    content, inputs = book(scenario)
    engine = [row["key_rate"] for row in
              core._run_authoritative_model(inputs, core.TEP_DEFAULT, [], {})
              ["consolidated"]["finance"]["rows"]]
    evaluator = Evaluator(openpyxl.load_workbook(io.BytesIO(content), data_only=False))
    read = [float(evaluator.cell("Ставки", f"{get_column_letter(4 + i)}5"))
            for i in range(len(engine))]
    for month, (mine, theirs) in enumerate(zip(engine, read)):
        assert theirs == pytest.approx(mine, abs=1e-9), (scenario, month)


def test_the_scenarios_are_not_the_same_curve() -> None:
    """Предохранитель: сойдись все три в одну кривую — проверка выше не
    значила бы ничего, а именно так книга и вела себя до правки."""
    ends = {}
    for scenario in SCENARIOS:
        rows = core._run_authoritative_model(
            {**core.DEFAULT_INPUTS, "rate_scenario": scenario},
            core.TEP_DEFAULT, [], {})["consolidated"]["finance"]["rows"]
        ends[scenario] = rows[-1]["key_rate"]
    assert len(set(round(v, 6) for v in ends.values())) == 3, ends
    assert ends["high"] > ends["base"] > ends["low"], ends


def test_the_book_names_the_scenario_by_its_own_column() -> None:
    """У книги имя про исход проекта, у движка — про уровень ставки: высокая
    ставка это Downside, а не Upside. Перепутанные, они дали бы кривую
    соседнего сценария — и она выглядела бы посчитанной."""
    for scenario, column in (("base", "Base"), ("high", "Downside"), ("low", "Upside")):
        content, _ = book(scenario)
        coord, value = cells(content)["rate_scenario"]
        assert value == column, (scenario, value)


def test_b6_still_carries_the_applied_multipliers() -> None:
    """`B6` трогать нельзя: он ведёт множители цены и затрат, и движок пишет
    ПРИМЕНЁННЫЕ в колонку Base. Уведи его за сценарием ставки — и книга взяла
    бы чужие множители и чужую задержку старта."""
    for scenario in SCENARIOS:
        content, _ = book(scenario)
        sheet = v4_inputs.inputs(openpyxl.load_workbook(io.BytesIO(content), data_only=False))
        assert sheet["B6"].value == "Base", scenario


def test_the_shape_and_the_start_date_are_cells_not_numbers_in_a_formula() -> None:
    """Правка в книге обязана что-то менять: число внутри формулы неотличимо
    от вводной, пока его не попробуешь исправить."""
    import re
    import zipfile

    content, _ = book("base")
    named = cells(content)
    for key in ("rate_curve_shape", "rate_start_date", "rate_scenario",
                "rate_target_base_pct", "rate_target_low_pct", "rate_target_high_pct"):
        assert key in named, key
    archive = zipfile.ZipFile(io.BytesIO(content))
    formula = ""
    for name in archive.namelist():
        if not (name.endswith(".xml") and "sheet" in name):
            continue
        text = archive.read(name).decode("utf-8", "replace")
        found = re.search(r"<x:f>(MAX\(0,'Параметры модели'!\$B\$34[^<]*)</x:f>", text)
        if found:
            formula = found.group(1)
            break
    assert formula, "кривая ключевой ставки не найдена"
    # Форма и смещение — ссылки на «Вводные», а не литералы.
    assert formula.count("'Параметры модели'!$B$") >= 5, formula
    assert "EXP(-2*" not in formula and "EXP(-2)" not in formula, formula


def test_every_input_has_a_home_in_the_book() -> None:
    """Счёт «чего в книге нет» считает то, о чём знает: блоки графиков и
    лестницы цен рисуются только при заполненном поле, поэтому меряем на
    вводных, где они заполнены."""
    filled = {
        "purchase_schedule": "60%@0; 40%@12",
        "offices_sales_profile": "50%@0; 50%@12",
        "retail_sales_profile": "50%@0; 50%@12",
        "above_parking_sales_profile": "50%@0; 50%@12",
        "offices_enabled": True, "retail_enabled": True,
        **{f"{prefix}growth_stage{step}_pct": 5
           for prefix in ("", "offices_", "retail_", "above_parking_")
           for step in (1, 2, 3, 4)},
    }
    content, _ = book("base", **filled)
    workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
    texts = {cell.value.strip() for name in workbook.sheetnames
             for row in workbook[name].iter_rows() for cell in row
             if isinstance(cell.value, str) and cell.value.strip()}
    homeless = [key for key in sorted(core.DEFAULT_INPUTS)
                if key not in texts and not any(key in text for text in texts)]
    assert homeless == [], homeless
