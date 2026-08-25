import datetime

import pytest

import developaid_monitor_scenarios as scenarios


def pm():
    d = datetime.date
    return {"known": True, "rnv_id": "3", "tasks": {
        "1": {"id": "1", "name": "Фасады", "start": d(2026, 1, 1),
              "finish": d(2026, 1, 31), "duration_days": 30,
              "predecessors": [], "free_float_days": 0, "total_float_days": 0},
        "2": {"id": "2", "name": "Отделка", "start": d(2026, 2, 1),
              "finish": d(2026, 3, 1), "duration_days": 28,
              "predecessors": [{"id": "1", "type": "FS", "lag_days": 0}],
              "free_float_days": 0, "total_float_days": 0},
        "3": {"id": "3", "name": "Получение РНВ", "start": d(2026, 3, 1),
              "finish": d(2026, 3, 1), "duration_days": 0,
              "predecessors": [{"id": "2", "type": "FS", "lag_days": 0}],
              "free_float_days": 0, "total_float_days": 0},
    }}


def test_delay_is_propagated_to_rnv():
    d = datetime.date
    view = {"schedule": {"rows": [
        {"id": "1", "forecast_finish": "2026-01-31", "code": "2.2.1"},
        {"id": "2", "forecast_finish": "2026-03-01", "code": "2.2.2"},
        {"id": "3", "forecast_finish": "2026-03-01", "code": ""},
    ]}}
    current, changed = scenarios._scenario_seeds(
        view, pm(), d(2026, 1, 15), "delay_wbs", ["1"], 20, 20
    )
    tasks = scenarios.graph._propagate(pm(), changed)
    assert tasks["1"]["forecast_finish"] == d(2026, 2, 20)
    assert scenarios._network_rnv(pm(), tasks) == d(2026, 3, 20)


def test_current_pace_uses_rss_evidence_but_not_as_physical_fact():
    d = datetime.date
    row = {"rss_accepted_ratio": .5, "rss_act_cost_rate_3m": .1}
    # Формула одна на оба контура и возвращает пару «дата, способ»:
    # способ подписывает строку, дата двигает сеть.
    assert scenarios._pace_finish(row, d(2026, 1, 1))[0] == d(2026, 6, 2)


def test_current_pace_does_not_turn_mixed_rss_21_into_finish_date():
    d = datetime.date
    row = {
        "code": "2.1", "plan_start": "2024-01-01",
        "rss_accepted_ratio": .8, "rss_act_cost_rate_3m": .02,
    }
    assert scenarios._pace_finish(row, d(2026, 8, 22))[0] is None


def test_forecast_coverage_says_why_rss_21_is_excluded():
    d = datetime.date
    view = {"schedule": {"rows": [
        {"code": "2.1", "plan_start": "2024-01-01", "plan_finish": "2027-09-30",
         "rss_accepted_ratio": .8, "rss_act_cost_rate_3m": .02},
        {"code": "2.2.1", "plan_start": "2026-01-01", "plan_finish": "2027-03-01",
         "rss_accepted_ratio": .5, "rss_act_cost_rate_3m": .1},
    ]}}
    result = scenarios._forecast_coverage(view, d(2026, 8, 22))
    assert result["active_articles"] == 1
    assert result["observed_articles"] == 1
    assert result["excluded_articles"] == ["2.1"]
    assert "смешивает" in result["exclusion_note"]


def test_forecast_drivers_explain_marginal_rnv_impact():
    d = datetime.date
    view = {"cut": "2026-01-15", "schedule": {"rows": [{
        "id": "1", "code": "2.2.1", "plan_start": "2026-01-01",
        "rss_accepted_ratio": .5, "rss_act_cost_rate_3m": .1,
    }]}}
    seeds = {"1": d(2026, 2, 20)}
    current = scenarios.graph._propagate(pm(), seeds)
    drivers = scenarios._forecast_drivers(
        view, pm(), seeds, scenarios._network_rnv(pm(), current)
    )
    assert drivers[0]["id"] == "1"
    assert drivers[0]["local_delay_days"] == 20
    assert drivers[0]["rnv_impact_days"] == 19
    assert drivers[0]["rss_codes"] == ["2.2.1"]


def test_parallel_late_chains_share_joint_rnv_impact():
    """Две параллельные опаздывающие ветки: поодиночке РНВ держит соседняя
    (одиночное влияние — ноль), а связка объясняет весь сдвиг."""
    d = datetime.date
    two_pm = {"known": True, "rnv_id": "9", "tasks": {
        "a": {"id": "a", "name": "Монолит К2", "start": d(2026, 1, 1),
              "finish": d(2026, 3, 1), "duration_days": 59,
              "predecessors": [], "free_float_days": 0, "total_float_days": 0},
        "b": {"id": "b", "name": "Монолит К3", "start": d(2026, 1, 1),
              "finish": d(2026, 3, 1), "duration_days": 59,
              "predecessors": [], "free_float_days": 0, "total_float_days": 0},
        "9": {"id": "9", "name": "Получение РНВ", "start": d(2026, 3, 1),
              "finish": d(2026, 3, 1), "duration_days": 0,
              "predecessors": [{"id": "a", "type": "FS", "lag_days": 0},
                               {"id": "b", "type": "FS", "lag_days": 0}],
              "free_float_days": 0, "total_float_days": 0},
    }}
    view = {"cut": "2026-03-15", "schedule": {"rows": [
        {"id": "a", "code": "2.2.2.1"}, {"id": "b", "code": "2.2.2.2"},
    ]}}
    seeds = {"a": d(2026, 5, 1), "b": d(2026, 5, 1)}
    current_rnv = scenarios._network_rnv(
        two_pm, scenarios.graph._propagate(two_pm, seeds))
    drivers = scenarios._forecast_drivers(view, two_pm, seeds, current_rnv)
    by_id = {item["id"]: item for item in drivers}

    assert by_id["a"]["rnv_impact_days"] == 0
    assert by_id["b"]["rnv_impact_days"] == 0
    # Вся задержка 61 день распределена внутри связки без остатка.
    assert sum(item["rnv_joint_days"] for item in drivers) == 61
    assert all(item["rnv_joint_days"] >= 0 for item in drivers)


def test_delayed_article_need_is_moved_without_changing_amount():
    articles = {"2.2.1.4": {"rss_limit": 10, "monthly_need": {
        "2026-08-01": 30, "2026-09-01": 40,
    }}}
    shifted = scenarios._shift_articles(articles, {"2.2.1": 40})
    assert shifted["2.2.1.4"]["monthly_need"] == {
        "2026-09-01": 30.0, "2026-10-01": 40.0,
    }
    assert sum(shifted["2.2.1.4"]["monthly_need"].values()) == pytest.approx(70)


def test_past_need_is_not_resurrected_by_scenario_rephasing():
    articles = {"2.2.1.4": {"rss_limit": 10, "monthly_need": {
        "2026-06-01": 100, "2026-08-01": 30,
    }}}
    shifted = scenarios._shift_articles(
        articles, {"2.2.1": 90}, datetime.date(2026, 8, 21)
    )
    assert shifted["2.2.1.4"]["monthly_need"] == {
        "2026-06-01": 100.0,
        "2026-10-01": 30.0,
    }


def test_acceleration_recovers_current_delay_but_not_before_baseline():
    d = datetime.date
    # Задержка объявляется так, как её признаёт общий контур: утверждённым
    # rebaseline. Голый `forecast_finish` — это уже учтённый в PM план, и
    # пересевать им сеть значило бы считать одну задержку дважды. Прежде
    # сценарный движок так и делал, а верхняя карточка — нет; отсюда и разные
    # даты РНВ на Гродненской.
    view = {"schedule": {"rows": [{
        "id": "1", "forecast_finish": "2026-04-15", "code": "2.2.1",
        "forecast_source": "approved_rebaseline",
    }]}}
    _, changed = scenarios._scenario_seeds(
        view, pm(), d(2026, 1, 15), "accelerate_wbs", ["1"], 0, 100
    )
    assert changed["1"] == d(2026, 3, 1)


def test_monitor_v2_shows_scenario_comparison_in_primary_ui():
    from developaid_monitor_page import MONITOR_PAGE
    assert "Cost control" in MONITOR_PAGE
    assert "Управленческий Гант" in MONITOR_PAGE
    assert "Платон · сценарный анализ" in MONITOR_PAGE
    assert "data-scenario=\"current_pace\"" in MONITOR_PAGE
    assert "Текущий утверждённый план" in MONITOR_PAGE


def test_scenario_route_is_registered_behind_monitor_access_gate():
    import main
    route = next(r for r in main.app.routes if getattr(r, "path", "") == "/monitor/scenario")
    assert "POST" in route.methods


def test_doubts_list_unclosed_monolith_and_recompute_without_it(tmp_path, monkeypatch):
    """Платон сомневается вслух: перечень с причиной и срок «если верны».

    Котлован с прошедшим на месяцы сроком при статье, актированной на 80%, —
    это незаполненная отметка, а не стройка; сеть без его seed даёт реальный
    прогноз.
    """
    import datetime
    import io
    import developaid_monitor as monitor
    import developaid_monitor_scenarios as scen
    import developaid_monitor_schedule_graph as graph
    from openpyxl import Workbook

    monkeypatch.setattr(monitor, "_SNAPSHOT_DIR", tmp_path)

    book = Workbook(); sheet = book.active; sheet.title = "ГПР"
    header = ["ID", "WBS", "Раздел", "Объект", "Наименование работ", "Тип строки",
              "% выполнения", "Начало", "Окончание", "Статус", "Длительность р.д.",
              "Предшественники", "Связанный тендер", "Окончание тендера",
              "Резерв до начала работ", "Увязка", "Код РСС", "Статья РСС", "Основание привязки"]
    for c, value in enumerate(header, 1):
        sheet.cell(row=4, column=c, value=value)
    d = datetime.date
    rows = [
        (1, "Корпус 1", "Разработка котлована", 0.0,
         d(2025, 8, 1), d(2025, 10, 28), "Просрочено", "2.2.1.1"),
        (2, "Корпус 1", "Кровля", 0.1,
         d(2026, 6, 1), d(2027, 3, 1), "В работе", "2.2.2.3"),
    ]
    for i, (rid, obj, name, pr, start, finish, status, code) in enumerate(rows):
        values = [rid, str(rid), "СМР", obj, name, "Работа", pr, start, finish,
                  status, 10, "", "", "", "", "", code, "", ""]
        for c, value in enumerate(values, 1):
            sheet.cell(row=5 + i, column=c, value=value)
    blob = io.BytesIO(); book.save(blob)
    monitor.store_schedule("Кутузов", blob.getvalue(), None, "2026-07-23")

    def fake_estimate(_):
        return {"rows": [
            {"code": "2.2.1.1", "parent": "", "estimate": 100e6},
            {"code": "2.2.2.3", "parent": "", "estimate": 100e6},
        ], "by_code": {"2.2.1.1": {"estimate": 100e6},
                       "2.2.2.3": {"estimate": 100e6}},
            "total": {"estimate": 200e6}}

    def fake_works(_):
        return {"rows": [
            {"code": "2.2.1.1", "amount": 80e6, "construction": True,
             "date": datetime.date(2026, 8, 1)},
        ], "total": 80e6}

    monkeypatch.setattr(monitor.actuals, "read_estimate", fake_estimate)
    monkeypatch.setattr(monitor.actuals, "read_completed_works", fake_works)
    monkeypatch.setattr(monitor.actuals, "read_payments", lambda _: {"rows": [], "total": 0.0})
    monkeypatch.setattr(monitor.actuals, "read_contracts", lambda _: {"rows": [], "total": 0.0})
    (tmp_path / "Кутузов" / "estimate").mkdir(parents=True, exist_ok=True)
    fake = Workbook(); fs = fake.active; fs.title = "Расчет стоимости строительства"
    fs.cell(row=9, column=1, value="Код")
    fs.cell(row=12, column=4, value="Всего инвестиционные расходы")
    fs.cell(row=12, column=5, value=200e6)
    fake.save(tmp_path / "Кутузов" / "estimate" / "2026-08-20.xlsx")

    monkeypatch.setattr(graph, "_load_pm", lambda _: {
        "known": True, "source": "pm.xlsx", "rnv_id": "9",
        "tasks": {
            "1": {"id": "1", "name": "Разработка котлована",
                  "start": d(2025, 8, 1), "finish": d(2025, 10, 28),
                  "duration_days": 60, "predecessors": [],
                  "free_float_days": 0, "total_float_days": 0},
            "2": {"id": "2", "name": "Кровля",
                  "start": d(2026, 6, 1), "finish": d(2027, 3, 1),
                  "duration_days": 200,
                  "predecessors": [{"id": "1", "type": "FS", "lag_days": 0}],
                  "free_float_days": 0, "total_float_days": 0},
            "9": {"id": "9", "name": "РНВ",
                  "start": d(2027, 9, 30), "finish": d(2027, 9, 30),
                  "duration_days": 0,
                  "predecessors": [{"id": "2", "type": "FS", "lag_days": 0}],
                  "free_float_days": 0, "total_float_days": 0},
        }})

    result = scen.doubts("Кутузов", "2026-08-24")

    ids = [row["id"] for row in result["rows"]]
    assert "1" in ids
    assert "2" not in ids
    doubted = next(row for row in result["rows"] if row["id"] == "1")
    assert "80%" in doubted["reason"]
    assert result["doubt_rnv"] <= result["current_rnv"]
    assert (result["recovered_days"] or 0) >= 0
