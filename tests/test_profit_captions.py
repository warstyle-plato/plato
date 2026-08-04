"""Словарь показателей: кассовая против аллоцированной прибыли, пики БРИДЖа.

«О1: −416 млн в книге против +427 в PDF» и «пик 7,49 против 6,25 млрд»
были сравнением разных показателей: кассовая прибыль CF-листа против
аллоцированной, тело долга против остатка с капитализацией. Числа верны —
подписи молчали. Теперь книга называет кассовые строки кассовыми, PDF
печатает словарь аллоцированной прибыли, а пик БРИДЖа книги назван телом.

Запуск: python3 -m pytest tests -q
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import main as wrapper  # noqa: E402

core = wrapper.core


def test_the_book_names_cash_rows():
    template = openpyxl.load_workbook(core._V4_TEMPLATE_PATH, data_only=False)
    for sheet in ("CF_1", "CF_2", "CF_3", "CF_4"):
        ws = template[sheet]
        assert "кассовая" in str(ws["A74"].value)
        assert "кассовая" in str(ws["A77"].value)
        assert "тело долга" in str(ws["A82"].value), \
            "после payable-механики тело БРИДЖа больше не несёт капитализацию"
    assert "кассовая" in str(template["КОНСОЛИДАТОР"]["L3"].value)


def test_the_pdf_prints_the_allocation_dictionary():
    pypdf = pytest.importorskip("pypdf")
    inputs = dict(core.DEFAULT_INPUTS)
    tep = {key: dict(value) for key, value in core.TEP_DEFAULT.items()}
    phasing = {"enabled": True, "phase_count": 2, "phase_gap_months": 12}
    bundle = core._run_authoritative_model(inputs, tep, [], phasing)
    payload = {"result": bundle["consolidated"], "inputs": inputs, "tep": tep,
               "rates": [], "phasing": bundle.get("phasing") or phasing,
               "scenario": "base", "project_name": "Словарь"}
    text = "\n".join(page.extract_text() or "" for page in pypdf.PdfReader(
        io.BytesIO(core._build_developaid_pdf(payload))).pages)
    assert "кассовая, как в CF-листах" in text.replace("\n", " ")
    assert "Аллоцированная прибыль" in text
    assert "как в книге" not in text, \
        "книга показывает тело долга — подпись «как в книге» устарела"
