"""Заполненный шаблон ТЭП на сайте: раньше он молча превращался в прочерки.

Владелец заполнил шаблон, который сам бот и выдаёт по `/template`, и загрузил
его на страницу. Страница ответила «Файл распознан» и показала таблицу из
прочерков: площадь территории «—», площадь квартир «—», смена ВРИ «—».

Причина в том, что страница знала один формат. Единственное поле загрузки
слало файл в `/import/glavapu`, разбор ГлавАПУ не отказывался от чужой книги,
не находил ни одного своего показателя и возвращал набор из одних None. Пустой
разбор считался успехом — та же ошибка, что была в боте, только здесь она не
приводила к сбою, а показывала пустоту с видом ответа.

При этом всё нужное уже существовало: `/import/manual-tep` разбирает шаблон
целиком, а `applyTelegramManualTep` умеет разложить его по вкладкам — этим
путём приходят проекты из Telegram. Не хватало одной попытки.

Проверено в живом Chromium: файл владельца даёт квартиры 29 308,89 м² и
355 штук, коммерцию 2 451,09 м², паркинг 266 мест, участок 2,084 га.

Запуск: python3 -m pytest tests -q
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import main_legacy as core  # noqa: E402


def upload_source() -> str:
    match = re.search(r"async function uploadGlavapu\(\)\{.*?\n\}", core.PAGE, re.S)
    assert match, "функция загрузки не найдена"
    return match.group(0)


# --- свой шаблон пробуется первым ---------------------------------------------

def test_the_page_tries_our_own_template_first():
    """Шаблон строгий — он несёт версию, и его разбор либо узнаёт файл, либо
    отказывается. Поэтому он и идёт первым: ошибиться нельзя."""
    body = upload_source()
    assert "/import/manual-tep" in body
    assert body.index("/import/manual-tep") < body.index("/import/glavapu")


def test_a_recognised_template_is_applied_not_just_shown():
    """Сводку с кнопкой «применить» строит разбор ГлавАПУ по своей структуре.
    У шаблона структура другая, и показывать его тем же блоком нечем — значит
    применяем сразу, тем же путём, что и проект из Telegram."""
    body = upload_source()
    assert "applyTelegramManualTep" in body


def test_the_person_learns_which_format_was_recognised():
    """«Файл распознан» без имени формата не даёт понять, что произошло."""
    body = upload_source()
    assert "Шаблон ТЭП DevelopAid применён" in body


# --- пустой разбор не успех ----------------------------------------------------

def test_an_empty_glavapu_parse_is_not_a_success():
    """Ни одного числа значит «формат не наш», а не «в файле нули»."""
    body = upload_source()
    assert "Формат файла не распознан" in body
    assert "typeof v==='number'" in body.replace(" ", "").replace("\n", "") \
        or "typeof v==='number'" in body


def test_the_refusal_says_what_to_do():
    body = upload_source()
    assert "Скачайте шаблон" in body


def test_the_field_names_both_formats():
    """Прежде подпись обещала только файл калькулятора ГлавАПУ."""
    body = upload_source()
    assert "шаблон ТЭП DevelopAid" in body and "ГлавАПУ" in body


# --- разбор шаблона на сервере -------------------------------------------------

def a_filled_template() -> bytes:
    """Тот же формат, что бот отдаёт по /template."""
    import io

    openpyxl = pytest.importorskip("openpyxl")
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "ТЭП DevelopAid"
    sheet["A4"], sheet["B4"] = "Версия шаблона", core.MANUAL_TEP_TEMPLATE_VERSION
    sheet["A5"], sheet["B5"] = "Название проекта", "Проверка"
    sheet["A6"], sheet["B6"] = "Регион / город", "Москва"
    sheet["A7"], sheet["B7"] = "Площадь территории", 2.084
    headers = ["Код", "Продукт", "ГНС / СПП, м²", "Общая площадь, м²",
               "Полезная площадь, м²", "Продаваемая площадь, м²",
               "Передаваемая площадь, м²", "Количество"]
    for index, title in enumerate(headers):
        sheet.cell(row=12, column=1 + index, value=title)
    sheet.append([])
    # Шаблон требует все строки продуктов: отсутствующая читается как порча
    # структуры, а не как «этого продукта нет».
    products = [("apartments", 43201.96, 29308.89, 355),
                ("ground_commercial", 2634.85, 2451.09, 0),
                ("standalone_retail", 0, 0, 0),
                ("offices", 0, 0, 0),
                ("above_parking", 0, 0, 0),
                ("underground_parking", 0, 0, 266),
                ("storage", 0, 0, 0),
                ("kindergarten", 0, 0, 0),
                ("school", 0, 0, 0),
                ("clinic", 0, 0, 0)]
    for row, (code, gns, saleable, units) in enumerate(products, start=13):
        sheet.cell(row=row, column=1, value=code)
        sheet.cell(row=row, column=3, value=gns)
        sheet.cell(row=row, column=4, value=gns)
        sheet.cell(row=row, column=5, value=saleable)
        sheet.cell(row=row, column=6, value=saleable)
        sheet.cell(row=row, column=8, value=units)
    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def test_the_template_parser_reads_the_whole_file():
    """То, чего страница не спрашивала, хотя оно работало."""
    parsed = core.parse_manual_tep_xlsx(a_filled_template(), "шаблон.xlsx")
    assert parsed["site_area_ha"] == pytest.approx(2.084)
    assert parsed["tep"]["apartments"]["saleable"] == pytest.approx(29308.89)
    assert parsed["tep"]["apartments"]["units"] == 355
    assert parsed["tep"]["ground_commercial"]["saleable"] == pytest.approx(2451.09)
    assert parsed["tep"]["underground_parking"]["units"] == 266
    # The issued DevelopAid_TEP_2 workbook predates this product row. It must
    # continue to load and receives a safe zero until a new template is issued.
    assert parsed["tep"]["other_mandatory"] == {
        "gns": 0.0, "total_area": 0.0, "useful": 0.0,
        "saleable": 0.0, "transfer": 0.0, "units": 0.0,
    }


def test_transferred_area_is_not_sold_from_a_file():
    """Передаваемая не продаётся — правило одно на все шаблоны.

    На странице его держит обработчик поля: правка передаваемой убирает
    столько же из продаваемой. У файла обработчика нет, колонки независимы,
    и заполненные обе целиком продали бы переданные муниципалитету метры.
    Проверка срабатывает только при нарушении инварианта, поэтому вычесть
    дважды она не может.
    """
    import io as _io
    import openpyxl as _openpyxl

    book = _openpyxl.load_workbook(_io.BytesIO(a_filled_template()))
    sheet = book.active
    # Квартиры: общая 43 201,96, продаваемая 29 308,89. Вписываем 20 000 м²
    # передаваемых, не трогая продаваемую: вместе они больше общей, то есть
    # переданные метры остались в продаже.
    for row in range(13, 30):
        if str(sheet.cell(row=row, column=1).value or "") == "apartments":
            sheet.cell(row=row, column=7, value=20000)
            break
    buffer = _io.BytesIO()
    book.save(buffer)

    parsed = core.parse_manual_tep_xlsx(buffer.getvalue(), "шаблон.xlsx")
    apartments = parsed["tep"]["apartments"]
    assert apartments["transfer"] == pytest.approx(20000)
    assert apartments["saleable"] == pytest.approx(43201.96 - 20000, abs=0.01)
    assert apartments["saleable"] + apartments["transfer"] <= apartments["total_area"] + 1
    # Молчаливое исправление неотличимо от отсутствующего.
    assert any("передаваемая не продаётся" in note for note in parsed["notes"])


def test_a_clean_template_keeps_its_saleable_area():
    """Инвариант не трогает файл, который его и не нарушал."""
    parsed = core.parse_manual_tep_xlsx(a_filled_template(), "шаблон.xlsx")
    assert parsed["tep"]["apartments"]["saleable"] == pytest.approx(29308.89)
    assert parsed["notes"] == []


# --- шаблон скачивается там же, где загружается ---------------------------------

def test_the_page_offers_the_template_for_download():
    """Отказ обещал «кнопку ниже», а её на странице не было ни одной: шаблон
    выдавал только бот командой `/template`, и человек с сайта узнать об этом
    ниоткуда не мог."""
    assert 'href="/templates/tep"' in core.PAGE


def test_the_download_stands_next_to_the_upload():
    """Скачать и загрузить — один шаг, и разнесённые по странице они не
    складываются в него."""
    upload = core.PAGE.index('id="glavapuFile"')
    link = core.PAGE.index('href="/templates/tep"')
    assert 0 < link - upload < 400, "ссылка на шаблон далеко от поля загрузки"


def test_the_refusal_does_not_promise_a_button_that_is_not_there():
    body = upload_source()
    assert "кнопкой ниже" not in body


# --- пустой свой шаблон не выдаётся за чужой ------------------------------------

def test_an_empty_template_keeps_its_own_reason():
    """Владелец загрузил наш же пустой шаблон и прочитал, что файл не наш:
    разбор шаблона отказал, страница молча пошла в ГлавАПУ и объявила формат
    неопознанным. Причина была известна на первом шаге."""
    body = upload_source()
    assert "X-DevelopAid-Template" in body
    assert "Заполните жёлтые ячейки" in body


def test_the_server_marks_whose_file_it_was():
    """Признак нужен странице, чтобы решить, пробовать ли второй формат."""
    import io

    from fastapi import HTTPException

    openpyxl = pytest.importorskip("openpyxl")
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "ТЭП DevelopAid"
    sheet["A4"], sheet["B4"] = "Версия шаблона", core.MANUAL_TEP_TEMPLATE_VERSION
    buffer = io.BytesIO()
    book.save(buffer)

    with pytest.raises(ValueError) as ours:
        core.parse_manual_tep_xlsx(buffer.getvalue(), "наш.xlsx")
    assert not isinstance(ours.value, core.ManualTepFormatError), \
        "опознанная версия — уже наш файл, а не чужой формат"


def test_a_file_without_our_sheet_is_a_foreign_format():
    """Только это и даёт право пробовать разбор ГлавАПУ."""
    import io

    openpyxl = pytest.importorskip("openpyxl")
    book = openpyxl.Workbook()
    book.active.title = "Лист1"
    buffer = io.BytesIO()
    book.save(buffer)
    with pytest.raises(core.ManualTepFormatError):
        core.parse_manual_tep_xlsx(buffer.getvalue(), "чужая.xlsx")


def test_the_untouched_template_says_what_is_missing():
    """Тот самый файл, что владелец скачал у бота и загрузил не заполнив."""
    import base64

    blank = base64.b64decode(
        core.MANUAL_TEP_TEMPLATE_B64_PATH.read_text(encoding="ascii").strip(), validate=True)
    with pytest.raises(ValueError) as exc:
        core.parse_manual_tep_xlsx(blank, core.MANUAL_TEP_TEMPLATE_FILENAME)
    assert not isinstance(exc.value, core.ManualTepFormatError)
    assert "ГНС" in str(exc.value)


def test_a_foreign_workbook_is_refused_by_the_template_parser():
    """Отказ разбора шаблона — то, что переводит файл на путь ГлавАПУ."""
    import io

    openpyxl = pytest.importorskip("openpyxl")
    book = openpyxl.Workbook()
    book.active.title = "ТЭП DevelopAid"
    book.active["A1"] = "Чужая книга без версии"
    buffer = io.BytesIO()
    book.save(buffer)
    with pytest.raises(ValueError):
        core.parse_manual_tep_xlsx(buffer.getvalue(), "чужая.xlsx")
