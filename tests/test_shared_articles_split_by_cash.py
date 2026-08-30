"""ИРД, проектирование, подготовка и наружные сети делятся кассовыми долями.

Движок берёт эти четыре статьи по проекту целиком и раскладывает между
очередями кассовыми долями (`shared_cash`; умолчание — веса очередей, со
страницы приходит front-loaded пресет). Книга считала каждую от площадей
СВОЕЙ очереди — то есть по долям ТЭП.

Пока доли близки, разница пряталась в допуске паритета. На проекте с длинным
ИРД и крупной покупкой она вылезла в 686 млн ₽ пика БРИДЖа и 406 млн стоимости
финансирования — 4% при совпадающих выручке и CAPEX. Нашлось помесячным рядом:
отношение выборок 306,1/194,8 = 1,571, а это ровно 55/35 — кассовая доля против
доли ТЭП.

Запуск: python3 -m pytest tests/test_shared_articles_split_by_cash.py -q
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
sys.path.insert(0, str(Path(__file__).resolve().parent))

import main as _wrapper  # noqa: E402

core = _wrapper.core

# Проект, на котором разрыв виден: длинный БРИДЖ под крупную покупку, доли ТЭП
# (35/65) заметно отличаются от кассовых (55/45).
INPUTS = {**core.DEFAULT_INPUTS, "purchase_price_mln": 12000,
          "project_start": "2027-01-01", "ird_months": 12,
          "construction_months": 24, "apartment_price_th": 700}
PHASING = {
    "enabled": True, "mode": "phased", "user_enabled": True,
    "phase_count": 2, "phase_gap_months": 12,
    "phases": [{"name": "О1", "start_offset_months": 0, "construction_months": 24},
               {"name": "О2", "start_offset_months": 12, "construction_months": 24}],
    "products": {key: [35, 65] for key in
                 ("apartments", "ground_commercial", "underground_parking", "storage")},
    "shared_cash": {}, "shared_allocation": {}, "social_objects": [],
}


@pytest.fixture(scope="module")
def book():
    tep = {key: dict(row) for key, row in core.TEP_DEFAULT.items()}
    content, _, meta = core.build_project_workbook(
        INPUTS, tep, [], PHASING, project_name="Кассовые доли")
    assert meta["missing"] == [], meta["missing"]
    return openpyxl.load_workbook(io.BytesIO(content), data_only=False)


@pytest.fixture(scope="module")
def evaluated(book):
    sys.setrecursionlimit(400000)
    from xlsx_eval import Evaluator
    return Evaluator(book)


def test_the_shares_differ_from_the_tep_shares():
    """Предохранитель. Совпади кассовые доли с долями ТЭП — проверки ниже
    проходили бы и на сломанной книге."""
    cash = core.phase_cash_default_weights(2)["ird"]
    assert cash == [55.0, 45.0]
    assert PHASING["products"]["apartments"] == [35, 65], (
        "доли ТЭП обязаны отличаться от кассовых, иначе тест ничего не ловит")


def test_the_workbook_carries_the_cash_shares(book):
    inputs = book["Вводные"]
    for column in ("AJ", "AK", "AL", "AM"):
        assert inputs[f"{column}88"].value == pytest.approx(0.55)
        assert inputs[f"{column}89"].value == pytest.approx(0.45)
        assert inputs[f"{column}90"].value == pytest.approx(0.0)


def test_the_articles_take_the_project_area_not_the_queue_area(book):
    """Статья считается от площади ПРОЕКТА, помноженной на кассовую долю.

    Прежде в формуле стояли площади самой очереди — та же величина, что делит
    ТЭП, и оттуда бралась доля.
    """
    capex = book["CAPEX"]
    for _key, row, column in core._V4_SHARED_CASH_ARTICLES:
        formula = str(capex[f"B{row}"].value)
        assert core._V4_PROJECT_AREA in formula, f"B{row}: нет площади проекта"
        assert f"'Вводные'!${column}$88" in formula, f"B{row}: нет кассовой доли"
        assert core._V4_QUEUE_AREA.format(row=88) not in formula, (
            f"B{row}: осталась площадь очереди — статья по-прежнему делится по ТЭП")


def test_the_book_and_the_engine_now_agree(evaluated, book):
    """Ради чего всё: паритет листа ПРОВЕРКИ на этом проекте.

    До правки красными были стоимость финансирования, налог, чистая прибыль и
    пик БРИДЖа. Строка «Финальный долг» остаётся красной по существу — проект
    и правда не гасит долг, — и в вердикт она входит, поэтому смотрим строки
    паритета поимённо.
    """
    failed = []
    for row in range(76, 86):
        label = book["ПРОВЕРКИ"][f"A{row}"].value
        if not label:
            continue
        if evaluated.cell("ПРОВЕРКИ", f"F{row}") == "FAIL":
            failed.append((label, evaluated.cell("ПРОВЕРКИ", f"B{row}"),
                           evaluated.cell("ПРОВЕРКИ", f"C{row}")))
    assert not failed, failed


def test_the_unpaid_debt_agrees_too(evaluated):
    """Непогашенный долг книги сходится с движком.

    Он и был главным следствием: книга видела 3 086 млн там, где движок 4 229,
    и на этой разнице строился весь дальнейший разъезд.
    """
    bundle = core.calculate_phased(core.PhasedCalcRequest(
        inputs=INPUTS, tep={k: dict(v) for k, v in core.TEP_DEFAULT.items()},
        rates=[], phasing=PHASING))
    engine = float(bundle["phases"][0]["result"]["finance"]["ending_pf"]) / 1e6
    # Порог держит смысл сценария — очередь не гасит долг, — а не конкретное
    # число: ставка подземки с 30.08.2026 равна 0,8 наземной, и долг тут же
    # упал с 4 229 до 3 862. Предохранитель на это и сработал.
    assert engine > 3_000, "предохранитель: очередь обязана не погасить долг"
    assert float(evaluated.cell("CF_1", "B47")) == pytest.approx(engine, rel=0.005)
