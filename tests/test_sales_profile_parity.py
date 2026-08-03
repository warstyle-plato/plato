"""Профиль продаж книги повторяет движок: ramp до РВЭ, объекты без тренда.

Книга продавала с экспоненциальным трендом 1%/мес на всём окне (дефолт
AD88:AD91 шаблона, билдер его не трогал) и применяла квартирную сезонность
к отдельно стоящим объектам. Движок ведёт квартиры линейным ramp'ом
(pace_adjustment_pct, только до РВЭ), а объекты — ровным темпом без
сезонности. На Мытищах эти профили давали +36 млн по квартирам и +87 млн
по офисам ниоткуда: объём уезжал в поздние дорогие месяцы. Теперь вес
темпа в книге — движковый ramp, тренд пишет билдер из pace_adjustment_pct,
объекты продаются равномерно; выручка сходится в ноль по каждому продукту.

Запуск: python3 -m pytest tests -q
"""

from __future__ import annotations

import io
import sys
import zipfile
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
sys.path.insert(0, str(Path(__file__).resolve().parent))

import main as wrapper  # noqa: E402
from xlsx_eval import Evaluator  # noqa: E402

core = wrapper.core


def test_the_template_uses_linear_ramp_and_keeps_objects_flat():
    """Формулы шаблона: вес темпа очередей — линейный MIN(1, n/срок),
    в ОБЪЕКТЫ нет ни тренда AD, ни квартирной сезонности."""
    archive = zipfile.ZipFile(core._V4_TEMPLATE_PATH)
    sales = archive.read("xl/worksheets/sheet5.xml").decode("utf-8")
    assert "*'Вводные'!$H$7*MIN(1," in sales, "ramp движка в Продажах"
    assert ")^MAX(0,(12*(YEAR(D$3)-YEAR($B$7))" not in sales.replace(
        "MIN(1,(12*(YEAR(D$3)-YEAR($B$7))", ""), "экспонента должна уйти"
    objects = archive.read("xl/worksheets/sheet18.xml").decode("utf-8")
    assert "$AD$88" not in objects, "объекты продаются без тренда темпа"
    assert "'Вводные'!$B$67" not in objects, "сезонность к объектам не применяется"


def test_the_builder_writes_the_engine_pace():
    inputs = dict(core.DEFAULT_INPUTS)
    inputs["pace_adjustment_pct"] = 25
    content, _, _ = core.build_project_workbook(
        inputs, {key: dict(value) for key, value in core.TEP_DEFAULT.items()},
        [], {}, finance_hints={})
    sheet = openpyxl.load_workbook(io.BytesIO(content), data_only=False)["Вводные"]
    for row in (88, 89, 90, 91):
        assert sheet[f"AD{row}"].value == pytest.approx(0.25)


def test_the_book_revenue_matches_the_engine_per_product():
    """Выручка каждого продукта в книге равна движковой — включая офисы
    третьей очереди с индексацией старта и паркинг целыми местами."""
    sys.setrecursionlimit(400000)
    inputs = dict(core.DEFAULT_INPUTS)
    inputs["offices_enabled"] = True
    tep = {key: dict(value) for key, value in core.TEP_DEFAULT.items()}
    phasing = {"enabled": True, "phase_count": 3, "phase_gap_months": 12,
               "discrete": {"offices": 3}}
    bundle = core._run_authoritative_model(inputs, tep, [], phasing)
    engine = {p["key"]: p["revenue"] / 1e6
              for p in bundle["consolidated"]["report"]["products"]}

    content, _, _ = core.build_project_workbook(
        inputs, tep, [], bundle.get("phasing") or phasing, finance_hints={})
    book = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
    evaluator = Evaluator(book)
    for cell, key in (("E46", "apartments"), ("E47", "ground_commercial"),
                      ("E48", "underground_parking"), ("E50", "offices")):
        value = float(evaluator.cell("ОТЧЕТ", cell) or 0)
        assert value == pytest.approx(engine.get(key, 0.0), abs=0.05), key
