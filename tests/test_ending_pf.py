"""Непогашенный долг ПФ виден на каждой поверхности, а не только внутри движка.

Движок всегда считал остаток долга ПФ на конец проекта (`ending_pf`), но
наружу его не выводил: сводка, PDF и карточка бота показывали безупречный
отчёт при непогашенных миллиардах — дефолт по проектному финансированию
выглядел как обычный слабый LLCR. Книга при этом продолжала начислять
проценты и после конца проекта, наращивая долг до конца горизонта.

Запуск: python3 -m pytest tests -q
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import main as wrapper  # noqa: E402

core = wrapper.core

NA = "N/A (долг не погашен)"


def weak_single():
    """Проект, который долг не гасит: цена квартир 245 против дефолтных 350
    оставляет около 5,4 млрд ₽ непогашенного ПФ."""
    inputs = dict(core.DEFAULT_INPUTS)
    inputs["apartment_price_th"] = 245
    tep = {key: dict(value) for key, value in core.TEP_DEFAULT.items()}
    return inputs, tep


# --- движок -----------------------------------------------------------------

def test_the_engine_exposes_the_unpaid_debt():
    inputs, tep = weak_single()
    result = core._run_authoritative_model(inputs, tep, [], {})["consolidated"]

    ending = result["finance"]["ending_pf"]
    assert ending > 0, "дефолтные вводные должны оставлять долг непогашенным"
    assert result["summary"]["ending_pf"] == pytest.approx(ending)
    assert result["report"]["financing"]["ending_pf"] == pytest.approx(ending)


def test_the_phases_sum_their_unpaid_debt():
    inputs, tep = weak_single()
    phasing = {"enabled": True, "phase_count": 2, "phase_gap_months": 12, "phases": []}
    bundle = core._run_authoritative_model(inputs, tep, [], phasing)
    result = bundle["consolidated"]

    total = sum(r["result"]["finance"]["ending_pf"] for r in bundle["phases"])
    assert result["summary"]["ending_pf"] == pytest.approx(total)
    assert result["report"]["financing"]["ending_pf"] == pytest.approx(total)


# --- PDF --------------------------------------------------------------------

def test_the_pdf_names_the_unpaid_debt():
    pypdf = pytest.importorskip("pypdf")
    inputs, tep = weak_single()
    bundle = core._run_authoritative_model(inputs, tep, [], {})
    data = core._build_developaid_pdf({
        "result": bundle["consolidated"], "inputs": inputs, "tep": tep,
        "rates": [], "phasing": {}, "scenario": "base", "project_name": "Дефолт",
    })
    reader = pypdf.PdfReader(io.BytesIO(data))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    assert "Непогашенный долг ПФ" in text.replace("\n", " ")


# --- карточка бота ----------------------------------------------------------

def card_text(monkeypatch, summary) -> str:
    sent: list[str] = []
    monkeypatch.setattr(core, "_telegram_verify_session", lambda s: {"chat_id": 42, "cad": []})
    monkeypatch.setattr(core, "_telegram_user_allowed", lambda c: True)
    monkeypatch.setattr(core, "_telegram_send_message",
                        lambda chat_id, text, **kw: sent.append(text))
    monkeypatch.setattr(core, "_telegram_web_app_url", lambda *a, **k: "https://example.org/")
    core.telegram_result(core.TelegramResultRequest(session="s", summary={
        "purchase_price_mln": 6500, "net_profit_mln": 900, "llcr": 1.3,
        "revenue_mln": 12000, "total_expenses_mln": 11000,
        **summary,
    }))
    assert sent, "карточка не отправлена"
    return sent[0]


def test_the_card_warns_about_a_default(monkeypatch):
    text = card_text(monkeypatch, {"ending_pf_mln": 5885.0})
    assert "Долг ПФ не погашается" in text
    assert core._telegram_money_mln(5885.0) in text
    assert "риск дефолта" in text


def test_a_repaid_project_gets_no_warning(monkeypatch):
    text = card_text(monkeypatch, {"ending_pf_mln": 0.0})
    assert "Долг ПФ не погашается" not in text


def test_the_page_sends_the_ending_debt_to_the_card():
    assert "ending_pf_mln:Number(f.ending_pf||0)/1e6" in core.PAGE


# --- книга ------------------------------------------------------------------

def test_the_book_stops_accruals_and_declares_the_default():
    """После конца проекта очереди проценты ПФ не начисляются, остаток долга
    назван непогашенным, IRR/NPV не притворяются доходностью при дефолте."""
    sys.setrecursionlimit(400000)
    from xlsx_eval import Evaluator

    inputs, tep = weak_single()
    content, _, meta = core.build_project_workbook(
        inputs, tep, [], {}, project_name="Дефолт")
    assert meta["missing"] == []

    book = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
    evaluator = Evaluator(book)

    debt = float(evaluator.cell("CF_1", "B86"))
    assert debt > 0, "дефолтные вводные должны оставлять долг в книге"
    assert float(evaluator.cell("CF_1", "DS47")) == pytest.approx(debt, rel=1e-6)
    assert float(evaluator.cell("CF_1", "DS42")) == pytest.approx(0.0, abs=1e-9), \
        "проценты ПФ продолжают начисляться после конца проекта"

    assert evaluator.cell("CF_1", "B80") == NA
    assert evaluator.cell("CF_1", "B81") == NA
    assert evaluator.cell("КОНСОЛИДАТОР", "N8") == NA
    assert evaluator.cell("КОНСОЛИДАТОР", "O8") == NA
    assert float(evaluator.cell("ОТЧЕТ", "B22")) == pytest.approx(debt, rel=1e-6)
    assert evaluator.cell("ПРОВЕРКИ", "F23") == "FAIL", \
        "чек финального долга обязан замечать дефолт"


def test_the_template_labels_the_debt_row():
    template = openpyxl.load_workbook(core._V4_TEMPLATE_PATH, data_only=False)
    for sheet in ("CF_1", "CF_2", "CF_3", "CF_4"):
        assert template[sheet]["A86"].value == "Непогашенный долг на конец проекта"
    assert "Непогашенный долг" in str(template["ОТЧЕТ"]["A22"].value)


def test_the_two_bridge_peaks_are_separate_indicators():
    """Книга ведёт остаток БРИДЖа с капитализацией процентов, движок «пиком»
    называет тело долга: одинаковое слово читалось как расхождение моделей
    (2 952 против 2 740 на 77:09). Теперь это два разных показателя."""
    inputs = dict(core.DEFAULT_INPUTS)
    tep = {key: dict(value) for key, value in core.TEP_DEFAULT.items()}
    result = core._run_authoritative_model(inputs, tep, [], {})["consolidated"]
    financing = result["report"]["financing"]
    fin = result["finance"]
    assert financing["bridge_peak_capitalized"] == pytest.approx(
        fin["peak_bridge"] + fin["transferred_bridge_interest"])

    phasing = {"enabled": True, "phase_count": 2, "phase_gap_months": 12, "phases": []}
    phased = core._run_authoritative_model(inputs, tep, [], phasing)["consolidated"]
    assert phased["report"]["financing"]["bridge_peak_capitalized"] >= \
        phased["report"]["financing"]["actual_bridge"]

    template = openpyxl.load_workbook(core._V4_TEMPLATE_PATH, data_only=False)
    for sheet in ("CF_1", "CF_2", "CF_3", "CF_4"):
        assert "капитализацией процентов" in str(template[sheet]["A82"].value)


def test_the_pdf_shows_both_bridge_peaks():
    pypdf = pytest.importorskip("pypdf")
    inputs, tep = weak_single()
    bundle = core._run_authoritative_model(inputs, tep, [], {})
    data = core._build_developaid_pdf({
        "result": bundle["consolidated"], "inputs": inputs, "tep": tep,
        "rates": [], "phasing": {}, "scenario": "base", "project_name": "Пики",
    })
    reader = pypdf.PdfReader(io.BytesIO(data))
    text = " ".join(page.extract_text() or "" for page in reader.pages).replace("\n", " ")
    assert "тело долга" in text
    assert "капитализацией процентов" in text
