"""Финансирование книги повторяет механику движка: payable, а не тело.

Книга капитализировала проценты и комиссии ПФ в тело долга и вела офисный
эскроу до РВЭ самого офиса; движок копит начисления на отдельном счёте
«к уплате», гасит его кассой в РВЭ очереди, а объектные поступления после
РВЭ очереди гасят ПФ напрямую. На Мытищах это давало −580 млн стоимости
финансирования и +3,2 млрд пикового ПФ. Теперь CF-листы ведут payable
(строки 37 и 48), уплату (53) и кассовые комиссии выдачи (57); стоимость
финансирования очереди — кассовая (B74), налоговая база вычитает уплату,
как движок.

Запуск: python3 -m pytest tests -q
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
sys.path.insert(0, str(Path(__file__).resolve().parent))

import main as wrapper  # noqa: E402
from xlsx_eval import Evaluator  # noqa: E402

core = wrapper.core


@pytest.fixture(scope="module")
def phased_case():
    sys.setrecursionlimit(400000)
    inputs = dict(core.DEFAULT_INPUTS)
    inputs["offices_enabled"] = True
    tep = {key: dict(value) for key, value in core.TEP_DEFAULT.items()}
    phasing = {"enabled": True, "phase_count": 2, "phase_gap_months": 12,
               "discrete": {"offices": 2}}
    bundle = core._run_authoritative_model(inputs, tep, [], phasing)
    content, _, _ = core.build_project_workbook(
        inputs, tep, [], bundle.get("phasing") or phasing, finance_hints={})
    book = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
    return bundle, Evaluator(book)


def test_the_financing_cost_matches_the_engine(phased_case):
    """Кассовая стоимость финансирования очереди — в пределах процента от
    движковой: базы процентов, payable и комиссии согласованы."""
    bundle, evaluator = phased_case
    for index, sheet in enumerate(("CF_1", "CF_2")):
        engine = float(bundle["phases"][index]["result"]["finance"]["financing_cost"]) / 1e6
        book = float(evaluator.cell(sheet, "B74") or 0)
        assert book == pytest.approx(engine, rel=0.02), sheet


def test_the_issuance_fees_are_cash_and_exact(phased_case):
    """Комиссии выдачи: БРИДЖ от расчётного лимита фазы (покупка+П+РД),
    резервирование ПФ от лимита очереди — в ноль с движком."""
    bundle, evaluator = phased_case
    for index, sheet in enumerate(("CF_1", "CF_2")):
        finance = bundle["phases"][index]["result"]["finance"]
        engine = (float(finance["pf_reservation_fee"]) + float(finance["bridge_fee"])) / 1e6
        book = float(evaluator.cell(sheet, "B57") or 0)
        assert book == pytest.approx(engine, abs=0.5), sheet


def test_the_pf_body_carries_no_capitalization(phased_case):
    """Тело ПФ книги равно движковому: проценты копятся в payable (48),
    а не в теле — раньше пик ПФ выходил на миллиарды больше."""
    from openpyxl.utils import get_column_letter
    bundle, evaluator = phased_case
    finance = bundle["phases"][0]["result"]["finance"]
    engine_peak = float(finance["peak_pf"]) / 1e6
    balances = [float(evaluator.cell("CF_1", f"{get_column_letter(c)}47") or 0)
                for c in range(4, 124)]
    assert max(balances) == pytest.approx(engine_peak, rel=0.02)


def test_the_template_rows_exist():
    template = openpyxl.load_workbook(core._V4_TEMPLATE_PATH, data_only=False)
    for sheet in ("CF_1", "CF_2", "CF_3", "CF_4"):
        ws = template[sheet]
        assert "к уплате" in str(ws["A37"].value)
        assert "к уплате" in str(ws["A48"].value)
        assert "уплата" in str(ws["A53"].value)
        assert "Комиссии выдачи" in str(ws["A57"].value)
        assert str(ws["B74"].value).startswith("=SUM(B53,B57)")
        # Налоговая база вычитает кассовую уплату, как движок.
        assert "D53" in str(ws["D22"].value) and "D57" in str(ws["D22"].value)
