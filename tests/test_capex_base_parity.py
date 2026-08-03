"""Процентные статьи CAPEX книги считаются от движковых баз.

Книга начисляла генподрядчика и техзаказчика на весь стройблок (ИРД,
проектирование, сети…), а движок — на СМР, социалку и отдельные объекты
(works_base). Управление проектом в книге включало «сдачу и ввод», у движка
её в базе нет. Резерв книги не видел объектов. На Мытищах это давало ±205
и ±286 млн по генподрядчику разных очередей и −331 млн резерва третьей.
Теперь базы книги повторяют движок: works_base = СМР + соцнагрузка +
объекты очереди; резерв — все статьи с ВРИ и объектами, без покупки.

Запуск: python3 -m pytest tests -q
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
sys.path.insert(0, str(Path(__file__).resolve().parent))

import main as wrapper  # noqa: E402
from xlsx_eval import Evaluator  # noqa: E402

core = wrapper.core

_ARTICLES = (
    (27, "Технический заказчик / стройконтроль"),
    (28, "Управление проектом"),
    (29, "Вознаграждение генподрядчика"),
    (30, "Резерв"),
)


def test_the_percent_articles_match_the_engine_per_queue():
    """Генподрядчик, техзаказчик, управление и резерв каждой очереди — в ноль
    с движковыми статьями; офисы третьей очереди входят в базы."""
    sys.setrecursionlimit(400000)
    inputs = dict(core.DEFAULT_INPUTS)
    inputs["offices_enabled"] = True
    inputs["technical_supervision_pct"] = 5
    tep = {key: dict(value) for key, value in core.TEP_DEFAULT.items()}
    phasing = {"enabled": True, "phase_count": 3, "phase_gap_months": 12,
               "discrete": {"offices": 3}}
    bundle = core._run_authoritative_model(inputs, tep, [], phasing)

    content, _, _ = core.build_project_workbook(
        inputs, tep, [], bundle.get("phasing") or phasing, finance_hints={})
    book = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
    evaluator = Evaluator(book)

    for index, base in enumerate((0, 34, 68)):
        report = bundle["phases"][index]["result"]["report"]
        engine = {row["label"]: row["value"] / 1e6
                  for row in report["construction_costs"]}
        for offset, label in _ARTICLES:
            value = float(evaluator.cell("CAPEX", f"B{offset + base}") or 0)
            assert value == pytest.approx(engine.get(label, 0.0), abs=0.05), \
                f"О{index + 1}: {label}"


def test_the_capex_total_matches_the_engine():
    """CAPEX очереди в книге равен движковому: базы процентов, ВРИ в резерве
    и объектный CAPEX согласованы, расхождение — только сериализация."""
    sys.setrecursionlimit(400000)
    inputs = dict(core.DEFAULT_INPUTS)
    tep = {key: dict(value) for key, value in core.TEP_DEFAULT.items()}
    phasing = {"enabled": True, "phase_count": 2, "phase_gap_months": 12}
    bundle = core._run_authoritative_model(inputs, tep, [], phasing)
    content, _, _ = core.build_project_workbook(
        inputs, tep, [], bundle.get("phasing") or phasing, finance_hints={})
    book = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
    evaluator = Evaluator(book)
    for index, base in enumerate((0, 34)):
        engine = float(bundle["phases"][index]["result"]["summary"]["capex"]) / 1e6
        value = float(evaluator.cell("CAPEX", f"B{32 + base}") or 0)
        assert value == pytest.approx(engine, abs=0.5), f"очередь {index + 1}"
