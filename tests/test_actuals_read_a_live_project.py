"""Факт действующего проекта читается так, как его отдают, а не как хочется.

Проверки написаны по тому, на чём разбор ломался на настоящих файлах
Гродненской: описка в кодах РСС, число текстом из 1С, дата акта внутри его
номера, год «204» и главы, сравниваемые с листьями.

Настоящие книги проекта сюда не кладутся — в них контрагенты и договоры.
Каждая проверка строит свою маленькую книгу той же формы.
"""

from __future__ import annotations

import datetime

import pytest

from openpyxl import Workbook

import developaid_actuals as actuals


def _estimate_book(path, rows):
    """Книга РСС той же формы: заголовок на 9-й строке, итог отдельной строкой."""
    book = Workbook()
    sheet = book.active
    sheet.title = "Расчет стоимости строительства"
    sheet.cell(row=9, column=1, value="Код")
    sheet.cell(row=9, column=4, value="Статья затрат")
    for offset, row in enumerate(rows):
        line = 12 + offset
        sheet.cell(row=line, column=1, value=row[0])
        sheet.cell(row=line, column=4, value=row[1])
        for column, value in zip((5, 6, 7, 11), row[2:]):
            sheet.cell(row=line, column=column, value=value)
    book.save(path)
    return path


def _register_book(path, rows):
    """Книга финансовой модели: лист «факт», заголовок на 12-й строке."""
    book = Workbook()
    sheet = book.active
    sheet.title = "факт"
    sheet.cell(row=_HEADER, column=11, value="Код банк")
    for offset, row in enumerate(rows):
        line = _HEADER + 1 + offset
        sheet.cell(row=line, column=11, value=row["code"])
        sheet.cell(row=line, column=13, value=row.get("bdds", ""))
        sheet.cell(row=line, column=19, value=row.get("paid_date"))
        sheet.cell(row=line, column=20, value=row.get("paid"))
        sheet.cell(row=line, column=22, value=row.get("act"))
    book.save(path)
    return path


_HEADER = 12


def _works_book(path, rows):
    """Книга РСС, лист актов: данные с девятой строки."""
    book = Workbook()
    sheet = book.active
    sheet.title = "Реестр выполненных работ"
    for offset, row in enumerate(rows):
        line = 9 + offset
        sheet.cell(row=line, column=2, value=row.get("document", "Акт"))
        sheet.cell(row=line, column=3, value=row.get("number_and_date"))
        sheet.cell(row=line, column=7, value=row.get("code"))
        sheet.cell(row=line, column=9, value=row.get("amount"))
        sheet.cell(row=line, column=11, value=row.get("contractor", "Подрядчик"))
    book.save(path)
    return path


# Форма РСС Гродненской в миниатюре: у главы 2 две подглавы, и подстроки
# внутренних инженерных пронумерованы 2.2.3.x под заголовком 2.3.
# Строка итога кода не имеет — он объявлен статьёй, как в настоящем РСС.
_ESTIMATE_ROWS = [
    (None,       "Всего инвестиционные расходы",   300.0, 300.0, 200.0, 150.0),
    ("1",        "Глава 1. Земля",                 100.0, 100.0, 100.0, 100.0),
    ("1.1.",     "Приобретение прав",              100.0, 100.0, 100.0, 100.0),
    ("2",        "Стоимость СМР",                  200.0, 200.0, 100.0,  50.0),
    ("2.2.",     "Основные объекты",               140.0, 140.0,  80.0,  50.0),
    ("2.2.1.",   "Подземная часть",                 90.0,  90.0,  50.0,  30.0),
    ("2.2.2.",   "Надземная часть",                 50.0,  50.0,  30.0,  20.0),
    ("2.3.",     "Внутренние инженерные системы",    60.0,  60.0,  20.0,   0.0),
    ("2.2.3.1.", "Вертикальный транспорт",           25.0,  25.0,  10.0,   0.0),
    ("2.2.3.2.", "Сантехнические системы",           35.0,  35.0,  10.0,   0.0),
]


def _estimate_fixture(tmp_path):
    return _estimate_book(tmp_path / "ssr.xlsx", _ESTIMATE_ROWS)


def test_the_tree_follows_the_rows_not_the_code(tmp_path):
    """`2.2.3.1` под заголовком `2.3` — ребёнок `2.3`, а не главы `2.2`.

    По префиксу кода он уехал бы в чужую главу и посчитался бы дважды, а сама
    `2.3` осталась бы листом. В РСС Гродненской это не гипотеза, а факт.
    """
    estimate = actuals.read_estimate(_estimate_fixture(tmp_path))
    by_code = estimate["by_code"]

    assert by_code["2.2.3.1"]["parent"] == "2.3"
    assert by_code["2.2.3.2"]["parent"] == "2.3"
    assert by_code["2.3"]["parent"] == "2"
    assert by_code["2.3"]["is_leaf"] is False
    assert by_code["2.2.1"]["is_leaf"] is True

    children = sum(row["estimate"] for row in estimate["rows"]
                   if row["parent"] == "2.3")
    assert children == pytest.approx(by_code["2.3"]["estimate"])


def test_the_total_is_declared_not_summed(tmp_path):
    """Итог берётся объявленной строкой: лист несёт и главы, и их подстроки."""
    estimate = actuals.read_estimate(_estimate_fixture(tmp_path))

    assert estimate["total"]["estimate"] == pytest.approx(300.0)
    assert estimate["total"]["paid"] == pytest.approx(200.0)
    # Суммирование всего листа дало бы двойной счёт — и это не итог.
    assert sum(row["estimate"] for row in estimate["rows"]) > 300.0


def test_a_chapter_is_compared_with_a_chapter(tmp_path):
    """Реестр платит по листьям, РСС хранит агрегат — сравнивать надо уровень.

    Без подъёма по дереву глава «Стоимость СМР» показывала бы расхождение на
    всю свою сумму: реестр в неё саму не платит никогда.
    """
    estimate = actuals.read_estimate(_estimate_fixture(tmp_path))
    register = actuals.read_register(_register_book(tmp_path / "fm.xlsx", [
        {"code": "2.2.1.", "paid": 50.0, "bdds": "2.2.2.3"},
        {"code": "2.2.2.", "paid": 30.0, "bdds": "2.2.2.4"},
        {"code": "2.2.3.1.", "paid": 10.0, "bdds": "2.2.2.7"},
        {"code": "2.2.3.2.", "paid": 10.0, "bdds": "2.2.2.7"},
    ]))
    report = actuals.reconcile(estimate, register)
    rolled = {row["code"]: row for row in report["by_code"]}

    assert rolled["2"]["paid_register"] == pytest.approx(100.0)
    assert rolled["2"]["paid_delta"] == pytest.approx(0.0)
    assert rolled["2.2"]["paid_register"] == pytest.approx(80.0)
    assert rolled["2.3"]["paid_register"] == pytest.approx(20.0)
    assert not any("нет в РСС" in warning for warning in report["warnings"])


def test_a_number_written_as_text_is_still_a_number(tmp_path):
    """«43 212,18» с неразрывным пробелом — обычная выгрузка 1С, а не мусор."""
    register = actuals.read_register(_register_book(tmp_path / "fm.xlsx", [
        {"code": "1.1.", "paid": "43\xa0212,18"},
        {"code": "1.1.", "paid": "1 000,50"},
    ]))

    assert register["paid"] == pytest.approx(44212.68)


def test_the_act_date_hides_inside_the_act_number(tmp_path):
    """Дата берётся после слова «от»: номер акта сам полон цифр с точками.

    Наивный поиск первого похожего на дату куска находил в «1960210 от 45485»
    номер, а не дату, и уводил акт в 204-й год.
    """
    works = actuals.read_completed_works(_works_book(tmp_path / "ssr.xlsx", [
        {"code": "2.2.1.", "amount": 10.0, "number_and_date": "516 от 19.10.2024"},
        {"code": "2.2.1.", "amount": 20.0,
         "number_and_date": "32-03-24 от 09.09.2024"},
        {"code": "2.2.1.", "amount": 30.0, "number_and_date": "бн от 02.10.2023"},
    ]))
    dates = sorted(row["date"] for row in works["rows"])

    assert dates == [datetime.date(2023, 10, 2), datetime.date(2024, 9, 9),
                     datetime.date(2024, 10, 19)]
    assert works["dated"] == pytest.approx(60.0)


def test_a_year_that_cannot_be_is_not_a_date(tmp_path):
    """«от 01.12.204» — описка источника. Такой месяц в свод не пускается."""
    works = actuals.read_completed_works(_works_book(tmp_path / "ssr.xlsx", [
        {"code": "2.2.1.", "amount": 10.0, "number_and_date": "7 от 01.12.204"},
    ]))

    assert works["rows"][0]["date"] is None
    assert works["undated"] == pytest.approx(10.0)


def test_paying_the_city_is_not_construction_progress(tmp_path):
    """Плата в ДГИ, комиссия банка и ФОТ — деньги, а не акт КС.

    На Гродненской они составляли 881,2 млн ₽ «выполненного». Сложенные с
    актами, они показали бы строительную готовность, которой нет.
    """
    works = actuals.read_completed_works(_works_book(tmp_path / "ssr.xlsx", [
        {"code": "2.2.1.", "amount": 100.0, "contractor": "ООО Подрядчик",
         "number_and_date": "1 от 05.05.2025"},
        {"code": "1.6.", "amount": 470.0, "contractor": "УФК по г. Москве (ДГИ)",
         "number_and_date": "бн от 05.05.2025"},
        {"code": "3.5.", "amount": 90.0, "contractor": "ПАО Сбербанк",
         "number_and_date": "бн от 05.05.2025"},
    ]))

    assert works["total"] == pytest.approx(660.0)
    assert works["dated"] == pytest.approx(660.0)
    assert works["construction_dated"] == pytest.approx(100.0)


def test_money_and_volume_keep_separate_clocks(tmp_path):
    """Аванс уходит раньше акта — слитая строка не покажет ни того ни другого."""
    register = actuals.read_register(_register_book(tmp_path / "fm.xlsx", [
        {"code": "2.2.1.", "paid": 60.0,
         "paid_date": datetime.datetime(2025, 4, 15)},
        {"code": "2.2.1.", "paid": 40.0,
         "paid_date": datetime.datetime(2025, 6, 10)},
    ]))
    works = actuals.read_completed_works(_works_book(tmp_path / "ssr.xlsx", [
        {"code": "2.2.1.", "amount": 70.0, "number_and_date": "1 от 20.06.2025"},
    ]))
    series = {row["month"]: row for row in actuals.monthly(register, works)["series"]}

    assert series["2025-04"]["paid"] == pytest.approx(60.0)
    assert series["2025-04"]["accepted"] == pytest.approx(0.0)
    assert series["2025-06"]["accepted"] == pytest.approx(70.0)
    assert series["2025-06"]["paid_cumulative"] == pytest.approx(100.0)


def test_a_gap_between_the_sources_is_reported_not_smoothed(tmp_path):
    """Расхождение источников — результат сводки, а не повод выбрать один."""
    estimate = actuals.read_estimate(_estimate_fixture(tmp_path))
    register = actuals.read_register(_register_book(tmp_path / "fm.xlsx", [
        {"code": "1.1.", "paid": 100.0},
        {"code": "2.2.1.", "paid": 20.0},
    ]))
    report = actuals.reconcile(estimate, register)

    assert report["total"]["paid_estimate"] == pytest.approx(200.0)
    assert report["total"]["paid_register"] == pytest.approx(120.0)
    assert any("разрыв" in warning for warning in report["warnings"])


def test_a_payment_without_a_known_code_is_not_lost(tmp_path):
    """`#N/A` в коде — сорванный поиск в источнике, и деньги за ним настоящие."""
    register = actuals.read_register(_register_book(tmp_path / "fm.xlsx", [
        {"code": "#N/A", "paid": 5.0},
        {"code": "1.1.", "paid": 10.0},
    ]))

    assert register["unmapped_paid"] == pytest.approx(5.0)
    assert register["paid"] == pytest.approx(15.0)
