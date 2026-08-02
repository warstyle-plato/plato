"""Крошечный вычислитель формул Excel для проверки выгружаемых книг.

Пересчитать книгу в песочнице нечем: openpyxl формулы не считает, LibreOffice
здесь сломан, а `formulas` спотыкается на функциях шаблона. Поэтому правки
формул проверялись чужим пересчётом — то есть на слово. Для книги, которую мы
собираем сами, этого мало: она вся из формул, и ошибка в любой уедет к
аналитику молча.

Здесь ровно тот набор, которым книга и написана: ссылки, диапазоны, арифметика,
сравнения и девять функций. Ничего больше он не умеет и уметь не должен —
формула, для которой не хватает возможностей, должна ломать тест, а не
пролезать.

Запуск: python3 -m pytest tests -q
"""

from __future__ import annotations

import calendar as _calendar
import datetime as _dt
import math as _math
import re
from typing import Any, Callable

from openpyxl.utils import column_index_from_string, get_column_letter

_TOKEN = re.compile(
    r"""
    (?P<string>"(?:[^"]|"")*")
  | (?P<number>\d+(?:\.\d+)?(?:[eE][+-]?\d+)?)
  | (?P<sheet>(?:'[^']+'|[^\s,()<>=+\-*/&:!]+)!)
  | (?P<cell>\$?[A-Za-z]{1,3}\$?\d+)
  | (?P<op><>|<=|>=|[-+*/^<>=&,():])
  | (?P<name>[A-Za-z_][A-Za-z_0-9.]*)
  | (?P<space>\s+)
    """,
    re.VERBOSE,
)


class FormulaError(Exception):
    """Формула, которую вычислитель не понимает, — повод упасть, а не гадать."""


def _as_number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, bool):
        return 1.0 if value else 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, _dt.datetime):
        value = value.date()
    if isinstance(value, _dt.date):
        # Серийная дата Excel: 1899-12-30 — нулевой день книги.
        return float((value - _dt.date(1899, 12, 30)).days)
    raise FormulaError(f"не число: {value!r}")


def _as_date(value: Any) -> _dt.date:
    if isinstance(value, _dt.datetime):
        return value.date()
    if isinstance(value, _dt.date):
        return value
    return _dt.date(1899, 12, 30) + _dt.timedelta(days=int(_as_number(value)))


def _edate(value: Any, offset: Any) -> _dt.date:
    start = _as_date(value)
    total = start.year * 12 + start.month - 1 + int(_as_number(offset))
    year, month = divmod(total, 12)
    day = min(start.day, _calendar.monthrange(year, month + 1)[1])
    return _dt.date(year, month + 1, day)


def _ceiling(value: float, step: float) -> float:
    if step == 0:
        return 0.0
    return _math.ceil(value / step) * step


def _flatten(values: Any) -> list[Any]:
    if isinstance(values, list):
        out: list[Any] = []
        for item in values:
            out.extend(_flatten(item))
        return out
    return [values]


def _numbers(values: Any) -> list[float]:
    return [_as_number(v) for v in _flatten(values) if v is not None and v != ""]


class RangeValue(list):
    """Диапазон, помнящий свою форму: INDEX(range, r, c) без неё не собрать."""

    def __init__(self, values: list[Any], rows: int, cols: int) -> None:
        super().__init__(values)
        self.rows = rows
        self.cols = cols


def _excel_round(value: float, digits: float, up: bool = False) -> float:
    n = int(digits)
    scale = 10.0 ** n
    scaled = value * scale
    if up:
        result = _math.ceil(abs(scaled)) * (1 if scaled >= 0 else -1)
    else:
        # Excel округляет половину от нуля, а не к чётному, как round().
        result = _math.floor(abs(scaled) + 0.5) * (1 if scaled >= 0 else -1)
    return result / scale


def _index(args: list[Any]) -> Any:
    values = args[0]
    row = int(_as_number(args[1])) if len(args) > 1 else 1
    col = int(_as_number(args[2])) if len(args) > 2 else 0
    if isinstance(values, RangeValue) and col:
        return values[(row - 1) * values.cols + (col - 1)]
    flat = _flatten(values)
    position = max(row, col, 1)
    if not 1 <= position <= len(flat):
        raise FormulaError(f"INDEX за пределами диапазона: {position} из {len(flat)}")
    return flat[position - 1]


def _match(args: list[Any]) -> int:
    target = args[0]
    flat = _flatten(args[1])
    mode = int(_as_number(args[2])) if len(args) > 2 else 1
    if mode != 0:
        raise FormulaError("MATCH поддержан только с точным совпадением (0)")
    for position, item in enumerate(flat, start=1):
        if _compare("=", item, target):
            return position
    raise FormulaError(f"MATCH: {target!r} не найдено")


def _countif(args: list[Any]) -> int:
    criteria = args[1]
    return sum(1 for item in _flatten(args[0]) if _compare("=", item, criteria))


def _npv(args: list[Any]) -> float:
    rate = _as_number(args[0])
    flows = _numbers(args[1:])
    return sum(flow / (1 + rate) ** period for period, flow in enumerate(flows, start=1))


def _xirr(args: list[Any]) -> float:
    flows = [_as_number(v) for v in _flatten(args[0])]
    days = [_as_number(v) for v in _flatten(args[1])]
    if len(flows) != len(days) or not flows:
        raise FormulaError("XIRR: потоки и даты разной длины")
    base = days[0]

    def value(rate: float) -> float:
        return sum(f / (1 + rate) ** ((d - base) / 365.0) for f, d in zip(flows, days))

    low, high = -0.9999, 100.0
    f_low, f_high = value(low), value(high)
    if f_low * f_high > 0:
        raise _DivZero()  # IRR не существует — Excel вернул бы #NUM!, ловится IFERROR
    for _ in range(200):
        mid = (low + high) / 2
        f_mid = value(mid)
        if abs(f_mid) < 1e-9:
            return mid
        if f_low * f_mid <= 0:
            high, f_high = mid, f_mid
        else:
            low, f_low = mid, f_mid
    return (low + high) / 2


FUNCTIONS: dict[str, Callable[[list[Any]], Any]] = {
    "SUM": lambda args: sum(_numbers(args)),
    "MAX": lambda args: max(_numbers(args) or [0.0]),
    "MIN": lambda args: min(_numbers(args) or [0.0]),
    "ABS": lambda args: abs(_as_number(args[0])),
    "AND": lambda args: all(bool(_as_number(a)) for a in _flatten(args)),
    "OR": lambda args: any(bool(_as_number(a)) for a in _flatten(args)),
    "IF": lambda args: args[1] if bool(_as_number(args[0]))
    else (args[2] if len(args) > 2 else False),
    "SUMPRODUCT": lambda args: sum(
        a * b for a, b in zip(*[[_as_number(v) for v in _flatten(arg)] for arg in args])
    ),
    "YEAR": lambda args: _as_date(args[0]).year,
    "DAY": lambda args: _as_date(args[0]).day,
    "EXP": lambda args: _math.exp(_as_number(args[0])),
    "CEILING": lambda args: _ceiling(_as_number(args[0]), _as_number(args[1])),
    "CEILING.MATH": lambda args: _ceiling(_as_number(args[0]),
                                          _as_number(args[1]) if len(args) > 1 else 1.0),
    "EDATE": lambda args: _edate(args[0], args[1]),
    "MONTH": lambda args: _as_date(args[0]).month,
    "MOD": lambda args: _as_number(args[0]) - _as_number(args[1]) * _math.floor(
        _as_number(args[0]) / _as_number(args[1])),
    "ROUND": lambda args: _excel_round(_as_number(args[0]), _as_number(args[1])),
    "ROUNDUP": lambda args: _excel_round(_as_number(args[0]), _as_number(args[1]), up=True),
    "COUNT": lambda args: len(_numbers(args)),
    "COUNTIF": _countif,
    "INDEX": _index,
    "MATCH": _match,
    "TEXT": lambda args: _text(args[0]),
    "NPV": _npv,
    "XIRR": _xirr,
}


class Evaluator:
    """Считает лист openpyxl: значения как есть, формулы — разбором.

    Ленивый: значение ячейки вычисляется при обращении и запоминается. Цикл
    ссылок ловится и превращается в ошибку — в книге его быть не должно.
    """

    def __init__(self, workbook) -> None:
        self.workbook = workbook
        self._cache: dict[tuple[str, str], Any] = {}
        self._stack: set[tuple[str, str]] = set()
        self._tokens: list[tuple[str, str]] = []
        self._pos = 0
        self._sheet = ""

    # --- доступ к ячейкам --------------------------------------------------

    def cell(self, sheet: str, address: str) -> Any:
        address = address.replace("$", "").upper()
        key = (sheet, address)
        if key in self._cache:
            return self._cache[key]
        if key in self._stack:
            raise FormulaError(f"круговая ссылка на {sheet}!{address}")
        raw = self.workbook[sheet][address].value
        self._stack.add(key)
        try:
            value = self.evaluate(raw, sheet) if isinstance(raw, str) and raw.startswith("=") else raw
        finally:
            self._stack.discard(key)
        self._cache[key] = value
        return value

    def _range(self, sheet: str, start: str, end: str) -> list[Any]:
        start, end = start.replace("$", "").upper(), end.replace("$", "").upper()
        c1, r1 = _split(start)
        c2, r2 = _split(end)
        out = []
        for row in range(min(r1, r2), max(r1, r2) + 1):
            for column in range(min(c1, c2), max(c1, c2) + 1):
                out.append(self.cell(sheet, f"{get_column_letter(column)}{row}"))
        return RangeValue(out, abs(r2 - r1) + 1, abs(c2 - c1) + 1)

    # --- разбор ------------------------------------------------------------

    def evaluate(self, formula: str, sheet: str) -> Any:
        # Разбор рекурсивный: ссылка внутри формулы уводит в другую формулу,
        # поэтому состояние разборщика сохраняется и возвращается.
        saved = (self._tokens, self._pos, self._sheet)
        self._tokens, self._pos, self._sheet = _tokenize(formula.lstrip("=")), 0, sheet
        try:
            value = self._expression()
            if self._pos != len(self._tokens):
                raise FormulaError(f"хвост после разбора: {formula}")
            return value
        finally:
            self._tokens, self._pos, self._sheet = saved

    def _peek(self) -> tuple[str, str] | None:
        return self._tokens[self._pos] if self._pos < len(self._tokens) else None

    def _take(self) -> tuple[str, str]:
        token = self._tokens[self._pos]
        self._pos += 1
        return token

    def _expression(self) -> Any:
        left = self._concat()
        while (token := self._peek()) and token[0] == "op" and token[1] in ("=", "<>", "<", "<=", ">", ">="):
            self._take()
            right = self._concat()
            left = _compare(token[1], left, right)
        return left

    def _concat(self) -> Any:
        left = self._sum()
        while (token := self._peek()) and token == ("op", "&"):
            self._take()
            left = f"{_text(left)}{_text(self._sum())}"
        return left

    def _sum(self) -> Any:
        left = self._product()
        while (token := self._peek()) and token[0] == "op" and token[1] in "+-":
            self._take()
            right = self._product()
            left = _as_number(left) + _as_number(right) if token[1] == "+" \
                else _as_number(left) - _as_number(right)
        return left

    def _product(self) -> Any:
        left = self._power()
        while (token := self._peek()) and token[0] == "op" and token[1] in "*/":
            self._take()
            right = _as_number(self._power())
            if token[1] == "/":
                if right == 0:
                    raise _DivZero()
                left = _as_number(left) / right
            else:
                left = _as_number(left) * right
        return left

    def _power(self) -> Any:
        left = self._unary()
        while (token := self._peek()) and token == ("op", "^"):
            self._take()
            left = _as_number(left) ** _as_number(self._unary())
        return left

    def _unary(self) -> Any:
        token = self._peek()
        if token and token[0] == "op" and token[1] in "+-":
            self._take()
            value = _as_number(self._unary())
            return -value if token[1] == "-" else value
        return self._atom()

    def _atom(self) -> Any:
        token = self._take()
        kind, text = token
        if kind == "number":
            return float(text)
        if kind == "string":
            return text[1:-1].replace('""', '"')
        if kind == "op" and text == "(":
            value = self._expression()
            self._expect(")")
            return value
        if kind == "sheet":
            sheet = text[:-1].strip("'")
            return self._reference(sheet, self._take())
        if kind == "cell":
            return self._reference(self._sheet, token)
        if kind == "name":
            return self._call(text.upper().removeprefix("_XLFN."))
        raise FormulaError(f"неожиданный токен {text!r}")

    def _reference(self, sheet: str, token: tuple[str, str]) -> Any:
        if token[0] != "cell":
            raise FormulaError(f"после имени листа ожидалась ячейка, а не {token[1]!r}")
        start = token[1]
        if (nxt := self._peek()) and nxt == ("op", ":"):
            self._take()
            end = self._take()
            if end[0] == "sheet":
                # «Лист!A1:Лист!B2» — легальная запись Excel; лист обязан
                # совпадать с началом диапазона.
                if end[1][:-1].strip("'") != sheet:
                    raise FormulaError("диапазон через два разных листа")
                end = self._take()
            return self._range(sheet, start, end[1])
        return self.cell(sheet, start)

    def _call(self, name: str) -> Any:
        self._expect("(")
        args: list[Any] = []
        if self._peek() != ("op", ")"):
            while True:
                if name in ("IF", "IFERROR"):
                    args.append(self._lazy())
                else:
                    args.append(self._expression())
                if self._peek() == ("op", ","):
                    self._take()
                    continue
                break
        self._expect(")")
        if name == "IFERROR":
            try:
                value = args[0]()
            except (_DivZero, ZeroDivisionError, FormulaError):
                return args[1]()
            return value
        if name == "IF":
            condition = bool(_as_number(args[0]()))
            if condition:
                return args[1]()
            return args[2]() if len(args) > 2 else False
        function = FUNCTIONS.get(name)
        if function is None:
            raise FormulaError(f"функция {name} вычислителю неизвестна")
        return function(args)

    def _lazy(self) -> Callable[[], Any]:
        """Аргумент IF/IFERROR считается только если до него дошло."""
        start = self._pos
        depth = 0
        while self._pos < len(self._tokens):
            kind, text = self._tokens[self._pos]
            if kind == "op" and text == "(":
                depth += 1
            elif kind == "op" and text == ")":
                if depth == 0:
                    break
                depth -= 1
            elif kind == "op" and text == "," and depth == 0:
                break
            self._pos += 1
        piece = self._tokens[start:self._pos]
        sheet = self._sheet

        def run() -> Any:
            saved_tokens, saved_pos, saved_sheet = self._tokens, self._pos, self._sheet
            self._tokens, self._pos, self._sheet = piece, 0, sheet
            try:
                value = self._expression()
                if self._pos != len(piece):
                    raise FormulaError("хвост в аргументе")
                return value
            finally:
                self._tokens, self._pos, self._sheet = saved_tokens, saved_pos, saved_sheet

        return run

    def _expect(self, symbol: str) -> None:
        token = self._take()
        if token != ("op", symbol):
            raise FormulaError(f"ожидалось {symbol!r}, встретилось {token[1]!r}")


class _DivZero(Exception):
    """#DIV/0! — ловится IFERROR, как в Excel."""


def _split(address: str) -> tuple[int, int]:
    match = re.fullmatch(r"([A-Z]{1,3})(\d+)", address)
    if not match:
        raise FormulaError(f"не адрес: {address}")
    return column_index_from_string(match.group(1)), int(match.group(2))


def _text(value: Any) -> str:
    return "" if value is None else str(value)


def _compare(operator: str, left: Any, right: Any) -> bool:
    if isinstance(left, str) or isinstance(right, str):
        a, b = _text(left), _text(right)
    else:
        a, b = _as_number(left), _as_number(right)
    return {
        "=": a == b, "<>": a != b, "<": a < b,
        "<=": a <= b, ">": a > b, ">=": a >= b,
    }[operator]


def _tokenize(formula: str) -> list[tuple[str, str]]:
    tokens: list[tuple[str, str]] = []
    position = 0
    while position < len(formula):
        match = _TOKEN.match(formula, position)
        if not match:
            raise FormulaError(f"не разобрано с позиции {position}: {formula[position:]!r}")
        position = match.end()
        kind = match.lastgroup
        if kind == "space":
            continue
        tokens.append((kind, match.group()))
    return tokens
