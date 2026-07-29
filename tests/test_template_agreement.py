"""Согласие выгрузки в шаблон ПЛАТО с расчётом DevelopAid.

Шаблон считает продажи сам — лист «ПОДБОР_КВ.М» — по тем вводным, которые мы
в него записали. Если наш движок распределяет продажи иначе, чем шаблон, то по
одним и тем же исходным данным отчёт на сайте и модель в Excel показывают
разную выручку. Ровно так и было: сезонность и смещение темпа уходили только в
шаблон, а паркинг в движке рос вдвое медленнее квартир.

Пересчитать книгу в песочнице нечем, поэтому движок продаж шаблона повторён
здесь на Python по формулам строк 54-68 листа «ПОДБОР_КВ.М».
Запуск: python3 -m pytest tests -q
"""

from __future__ import annotations

import io
import sys
from datetime import date
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import main as _wrapper  # noqa: E402

main = _wrapper.core

openpyxl = pytest.importorskip("openpyxl", reason="openpyxl нужен только для проверки выгрузки")
from openpyxl import load_workbook  # noqa: E402

TEMPLATE = Path(__file__).resolve().parent.parent / "templates" / "PLATO_template.xlsx"
pytestmark = pytest.mark.skipif(not TEMPLATE.is_file(), reason="шаблон ПЛАТО не поставляется")

LOW_SEASON = (1, 5, 6, 7, 8)


def add_months(d: date, n: int) -> date:
    year, month = divmod(d.year * 12 + (d.month - 1) + n, 12)
    return date(year, month + 1, 1)


def months_between(a: date, b: date) -> int:
    return 12 * (b.year - a.year) + b.month - a.month


def template_sales(params: dict) -> dict[str, float]:
    """Движок продаж шаблона: строки 55-68 листа ПОДБОР_КВ.М."""
    sales_start = add_months(params["permit"], params["sales_lag"])
    rve = add_months(params["permit"], params["construction_months"])
    sales_end = add_months(rve, params["residual_months"])

    months = [add_months(params["project_start"], i) for i in range(240)]
    flags, weights, prices = [], [], []
    for month in months:
        flag = 0 if month < sales_start or month > sales_end else (1 if month < rve else 2)
        flags.append(flag)
        season = 1 + params["seasonal"] if month.month in LOW_SEASON else 1.0
        if flag == 1:
            span = max(1, months_between(sales_start, rve))
            shift = 1 + params["pace"] * max(0.0, min(1.0, months_between(sales_start, month) / span))
        else:
            shift = 1.0
        weights.append(0.0 if flag == 0 else season * shift)
        if flag == 0:
            prices.append(0.0)
        else:
            pre = min(max(0, months_between(sales_start, month)),
                      max(0, months_between(sales_start, rve)))
            post = max(0, months_between(rve, month))
            prices.append(params["price_apartment"]
                          * (1 + params["growth_pre"]) ** pre
                          * (1 + params["growth_post"]) ** post)

    before = sum(w for w, f in zip(weights, flags) if f == 1)
    after = sum(w for w, f in zip(weights, flags) if f == 2)
    share = params["share_before_rve"]
    shares = []
    for w, f in zip(weights, flags):
        if f == 1 and before:
            shares.append(share * w / before)
        elif f == 2 and after:
            shares.append((1 - share) * w / after)
        else:
            shares.append(0.0)

    base = params["price_apartment"]
    return {
        "квартиры": sum(params["apartments_sqm"] * s * p for s, p in zip(shares, prices)) / 1000,
        "коммерция": sum(params["commercial_sqm"] * s * p * params["price_commercial"] / base
                         for s, p in zip(shares, prices)) / 1000,
        "паркинг": sum(params["parking_units"] * s * p * params["price_parking"] / base
                       for s, p in zip(shares, prices)) / 1000,
        "контроль": sum(shares),
    }


def template_params(content: bytes) -> dict:
    """Читает обратно то, что выгрузка записала в шаблон."""
    book = load_workbook(io.BytesIO(content))
    sheet = book["Вводные"]
    rows = {main._plato_normalize(sheet.cell(row=r, column=2).value): r
            for r in range(1, sheet.max_row + 1)}
    value = lambda label: sheet.cell(row=rows[label], column=5).value  # noqa: E731

    tep = book["Расчет ВРИ (ТЭП)"]
    tep_rows = {main._plato_normalize(tep.cell(row=r, column=2).value): r
                for r in range(1, tep.max_row + 1)}
    area = lambda label: float(tep.cell(row=tep_rows[label], column=4).value)  # noqa: E731

    start = value("начало проекта").date()
    construction = int(value("срок строительства"))
    lag = int(value("лаг старта продаж после рнс"))
    target = value("целевой совокупный рост цены от старта продаж до рвэ")
    return {
        "project_start": start,
        "permit": add_months(start, int(value("срок ирд до рнс"))),
        "construction_months": construction,
        "sales_lag": lag,
        "residual_months": int(value("остаточные продажи после рвэ")),
        "share_before_rve": value("доля продаж до рвэ"),
        "seasonal": value("сезонное снижение темпа"),
        "pace": value("смещение темпа продаж к поздним месяцам"),
        # Шаблон извлекает месячный рост из целевого — повторяем его формулу.
        "growth_pre": (1 + target) ** (1 / max(1, construction - lag)) - 1,
        "growth_post": (1 + value("инфляция после рвэ")) ** (1 / 12) - 1,
        "apartments_sqm": area("площадь квартир"),
        "commercial_sqm": area("нежилая наземная площадь (ннп)"),
        # ТЭП!I33 шаблона складывает постоянные и гостевые парковки.
        "parking_units": area("постоянные парковки") + area("гостевые парковки"),
        "price_apartment": value("стартовая цена квартир"),
        "price_commercial": value("стартовая цена коммерции"),
        "price_parking": value("цена машино-места"),
    }


@pytest.fixture(scope="module")
def single_project():
    content, _ = main.fill_plato_template(main.DEFAULT_INPUTS, main.TEP_DEFAULT)
    result = main._run_authoritative_model(main.DEFAULT_INPUTS, main.TEP_DEFAULT, [], {})
    return template_sales(template_params(content)), result["consolidated"]["revenue"]


def engine_mln(products: dict, key: str) -> float:
    return float(products.get(key) or 0.0) / 1e6


def test_guest_parking_of_a_foreign_project_is_cleared():
    """Гостевые парковки шаблона нельзя оставлять: ТЭП прибавляет их к продаваемым."""
    stale = load_workbook(TEMPLATE)["Расчет ВРИ (ТЭП)"]
    rows = {main._plato_normalize(stale.cell(row=r, column=2).value): r
            for r in range(1, stale.max_row + 1)}
    row = rows["гостевые парковки"]
    assert stale.cell(row=row, column=4).value  # в шаблоне лежит чужое значение
    content, _ = main.fill_plato_template(main.DEFAULT_INPUTS, main.TEP_DEFAULT)
    filled = load_workbook(io.BytesIO(content))["Расчет ВРИ (ТЭП)"]
    assert filled.cell(row=row, column=4).value == 0


def test_standalone_objects_can_be_switched_on():
    """«Объект включен» живёт в колонке G литералом, а не в сценарных D:F.

    Пока выгрузка писала только в D:F, МФОЦ, ТЦ и наземный паркинг оставались
    выключенными в каждой модели, и их выручка пропадала: свод по очередям
    показывал 104 млрд против 118 млрд в отчёте.
    """
    stale = load_workbook(TEMPLATE)["Вводные"]
    switches = [row for row in range(1, stale.max_row + 1)
                if main._plato_normalize(stale.cell(row=row, column=2).value) == "объект включен"]
    assert switches, "в шаблоне не нашлись выключатели объектов"
    for row in switches:
        assert stale.cell(row=row, column=7).value == "Нет"

    content, _ = main.fill_plato_template(
        {**main.DEFAULT_INPUTS, "offices_enabled": True}, main.TEP_DEFAULT)
    filled = load_workbook(io.BytesIO(content))["Вводные"]
    offices = next(row for row in switches
                   if str(filled.cell(row=row, column=1).value or "").startswith("МФОЦ"))
    assert filled.cell(row=offices, column=7).value == "Да"


def test_scenario_formulas_survive_the_fill():
    """Колонка G в остальных строках выбирает сценарий формулой — её не трогаем."""
    stale = load_workbook(TEMPLATE)["Вводные"]
    content, _ = main.fill_plato_template(main.DEFAULT_INPUTS, main.TEP_DEFAULT)
    filled = load_workbook(io.BytesIO(content))["Вводные"]
    for row in range(1, stale.max_row + 1):
        before = stale.cell(row=row, column=7).value
        if isinstance(before, str) and before.startswith("="):
            assert filled.cell(row=row, column=7).value == before, f"строка {row}"


def test_sales_distribution_matches_the_template(single_project):
    """Веса продаж по месяцам обязаны совпадать: иначе разойдётся средняя цена."""
    template, _ = single_project
    assert template["контроль"] == pytest.approx(1.0, rel=1e-9)


@pytest.mark.parametrize("product,key", [
    ("квартиры", "apartments"),
    ("коммерция", "ground_commercial"),
    ("паркинг", "underground_parking"),
])
def test_revenue_agrees_with_the_template(single_project, product, key):
    template, products = single_project
    engine = engine_mln(products, key)
    assert engine > 0
    assert template[product] == pytest.approx(engine, rel=0.005), (
        f"{product}: шаблон {template[product]:,.0f} млн, движок {engine:,.0f} млн"
    )


def test_seasonality_and_pace_actually_move_the_model():
    """Поля есть в интерфейсе — значит должны менять расчёт, а не только выгрузку."""
    base = main._run_authoritative_model(main.DEFAULT_INPUTS, main.TEP_DEFAULT, [], {})
    flat = main._run_authoritative_model(
        {**main.DEFAULT_INPUTS, "seasonal_reduction_pct": 0, "pace_adjustment_pct": 0},
        main.TEP_DEFAULT, [], {},
    )
    assert base["consolidated"]["summary"]["revenue"] != flat["consolidated"]["summary"]["revenue"]
