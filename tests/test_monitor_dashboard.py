import datetime
from pathlib import Path

import pytest

import developaid_monitor as monitor
import developaid_monitor_dashboard as dashboard


def test_dated_construction_acts_exclude_nonconstruction_and_undated_rows(monkeypatch):
    monkeypatch.setattr(dashboard.actuals, "read_completed_works", lambda _: {"rows": [
        {"code": "2.2.1", "amount": 100.0, "construction": True, "date": datetime.date(2026, 8, 10)},
        {"code": "2.2.1", "amount": 500.0, "construction": False, "date": datetime.date(2026, 8, 10)},
        {"code": "2.2.1", "amount": 300.0, "construction": True, "date": None},
        {"code": "3.1", "amount": 200.0, "construction": True, "date": datetime.date(2026, 8, 10)},
    ]})
    assert dashboard._physical_smr(Path("rss.xlsx"), {}, datetime.date(2026, 8, 20)) == 100.0


def test_article_surplus_does_not_cross_fund_another_article():
    articles = {
        "2.2.2.1": {
            "rss_limit": 100.0, "paid_at_baseline": 0.0,
            "monthly_need": {"2026-08-01": 0.0},
        },
        "2.2.3.3": {
            "rss_limit": 0.0, "paid_at_baseline": 0.0,
            "monthly_need": {"2026-08-01": 50.0},
        },
    }
    result = dashboard._article_waterfall(
        articles, reserve=60.0, cut=datetime.date(2026, 8, 21)
    )
    assert result["opening_bank_remaining"] == pytest.approx(100.0)
    assert result["monthly_reserve_draw"]["2026-08-01"] == pytest.approx(50.0)
    assert result["reserve_balance"] == pytest.approx(10.0)
    assert result["additional_financing"] == pytest.approx(0.0)


def test_grodnenskaya_reserve_control_benchmark_exhausts_in_november():
    articles = {
        "2.2.3.3": {
            "rss_limit": 0.0,
            "paid_at_baseline": 0.0,
            "monthly_need": {
                "2026-08-01": 2_000_000.0,
                "2026-09-01": 66_330_343.0,
                "2026-10-01": 100_680_310.59827147,
                "2026-11-01": 208_530_177.3909351,
                "2026-12-01": 290_632_872.24154127,
            },
        }
    }
    result = dashboard._article_waterfall(
        articles,
        reserve=306_141_482.2601274,
        cut=datetime.date(2026, 8, 21),
    )

    assert result["reserve_start"] == datetime.date(2026, 8, 1)
    assert result["monthly_reserve_balance"]["2026-08-01"] == pytest.approx(304_141_482.2601274)
    assert result["monthly_reserve_balance"]["2026-09-01"] == pytest.approx(237_811_139.2601274)
    assert result["monthly_reserve_balance"]["2026-10-01"] == pytest.approx(137_130_828.66185594)
    assert result["reserve_exhaustion"].year == 2026
    assert result["reserve_exhaustion"].month == 11
    assert result["monthly_unfunded"]["2026-11-01"] == pytest.approx(71_399_348.72907916)
    assert result["additional_financing"] == pytest.approx(362_032_220.9706204)


def test_funding_risk_uses_article_waterfall(monkeypatch):
    monkeypatch.setattr(dashboard, "_finance_baseline", lambda _: {
        "known": True,
        "source": "finance.xlsx",
        "approved": 1000.0,
        "completion_need_at_baseline": 900.0,
        "paid_at_baseline": 100.0,
        "reserve": 100.0,
        "reserve_parts": {"2.8": 40.0, "2.9": 60.0},
        "monthly_need": {},
        "tail_after_apr": 600.0,
        "articles": {
            "2.2.3.3": {
                "rss_limit": 50.0,
                "paid_at_baseline": 0.0,
                "monthly_need": {
                    "2026-08-01": 70.0,
                    "2026-09-01": 100.0,
                },
            }
        },
    })
    monkeypatch.setattr(dashboard.actuals, "read_estimate", lambda _: {
        "rows": [], "by_code": {"2": {"estimate": 700.0}, "3": {"estimate": 100.0}}
    })
    monkeypatch.setattr(dashboard, "_payment_total_ch23", lambda *_: 200.0)
    monkeypatch.setattr(dashboard, "_payment_by_code", lambda _: {})
    view = {"schedule": {"approved_end": "2027-09-30", "forecast_end": None}}

    result = dashboard._funding_risk(
        "demo", Path("rss.xlsx"), datetime.date(2026, 8, 20), view
    )

    assert result["reserve_start"] == "2026-08-01"
    assert result["reserve_exhaustion"].startswith("2026-09-")
    assert result["additional_financing"] == pytest.approx(20.0)
    assert result["reserve"] == pytest.approx(100.0)
    assert "постатейный waterfall" in result["method"]


def test_page_has_funding_and_correct_cost_evidence_language():
    from developaid_monitor_page import MONITOR_PAGE
    assert "Исчерпание резерва" in MONITOR_PAGE
    assert "Непокрытая потребность" in MONITOR_PAGE
    assert "Cost control" in MONITOR_PAGE
    assert "КС / EAC proxy" in MONITOR_PAGE
    assert "Утверждённый РНВ" in MONITOR_PAGE
    assert "Прогноз РНВ" in MONITOR_PAGE


def test_sales_merge_book_fact_with_manual_rows(tmp_path, monkeypatch):
    """Строка руками перекрывает тот же месяц книги, а не дублирует его.

    Книга обновляется раз в месяц и отстаёт: «в августе продано 4 лота»
    приходит словами раньше выгрузки. На Кутузове книга знала 41 лот по март,
    август пришёл строкой.
    """
    import io
    import json
    from openpyxl import Workbook

    monkeypatch.setattr(monitor, "_SNAPSHOT_DIR", tmp_path)

    import developaid_actuals as actuals

    book = Workbook()
    sheet = book.active
    sheet.title = "План продаж"
    for offset, (month, mark, units, area, price) in enumerate([
        (datetime.date(2026, 3, 1), "ФАКТ", 3, 300.0, 650_000.0),
        (datetime.date(2026, 8, 1), "ФАКТ", 1, 50.0, 600_000.0),
        (datetime.date(2026, 9, 1), "ПЛАН", 9, 500.0, 900_000.0),
    ]):
        line = actuals._SALES_FIRST_ROW + offset
        sheet.cell(row=line, column=actuals._SALES_COLUMNS["period"] + 1, value=month)
        sheet.cell(row=line, column=actuals._SALES_COLUMNS["mark"] + 1, value=mark)
        sheet.cell(row=line, column=actuals._SALES_COLUMNS["units"] + 1, value=units)
        sheet.cell(row=line, column=actuals._SALES_COLUMNS["area"] + 1, value=area)
        sheet.cell(row=line, column=actuals._SALES_COLUMNS["price"] + 1, value=price)
    blob = io.BytesIO()
    book.save(blob)

    monitor.store_sales_file("Гродненская", blob.getvalue(), "2026-07-31")
    monitor.store_sales("Гродненская", [
        {"month": "2026-08", "units": 4, "area": 240.0, "revenue": 160e6},
    ], "2026-08-20")

    snapshot = dashboard._sales_snapshot("Гродненская", datetime.date(2026, 8, 23))

    assert snapshot["known"]
    assert snapshot["last_fact"] == "2026-08"
    # август из строк (4 лота), а не из книги (1 лот); план книги не факт
    assert snapshot["total_units"] == pytest.approx(3 + 4)
    assert snapshot["total_revenue"] == pytest.approx(300.0 * 650_000.0 + 160e6)


def test_manual_sales_rows_are_not_ignored(tmp_path, monkeypatch):
    monkeypatch.setattr(monitor, "_SNAPSHOT_DIR", tmp_path)
    stored = monitor.store_sales("Гродненская", [
        {"month": "2026-08", "units": 4, "area": 240.0, "revenue": 160e6},
    ], "2026-08-20")

    assert "ignored_by_monitor" not in stored
    assert stored["months"] == 1

    with pytest.raises(ValueError):
        monitor.store_sales("Гродненская", [{"units": 4}], "2026-08-21")


def test_missing_finance_book_is_a_reason_not_a_zero(tmp_path, monkeypatch):
    """Ноль, показанный вместо «книги нет», неотличим от посчитанного нуля."""
    monkeypatch.setattr(monitor, "_SNAPSHOT_DIR", tmp_path)
    baseline = dashboard._finance_baseline("Гродненская")

    assert baseline["known"] is False
    assert "не загружен" in baseline["reason"]
