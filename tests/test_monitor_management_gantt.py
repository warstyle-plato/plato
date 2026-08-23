import datetime

import pytest

import developaid_monitor_manager as manager


def test_rss_codes_use_natural_numeric_order():
    codes = ["2.2.3.10", "2.4.1", "2.2.3.2", "2.2.3.9", "2.1"]
    assert sorted(codes, key=manager._natural) == [
        "2.1", "2.2.3.2", "2.2.3.9", "2.2.3.10", "2.4.1"
    ]


def test_rss_21_is_split_by_wbs_meaning_not_used_as_one_schedule_phase():
    prep = {"wbs": "1.16.1.2.4", "name": "Организация строительной площадки"}
    pos = {"wbs": "1.16.1.3.1", "name": "Работы, предусмотренные ПОС"}

    assert manager._schedule_bucket(prep, "2.1") == (
        "Подготовка", "Подготовка территории"
    )
    assert manager._schedule_bucket(pos, "2.1") == (
        "Организация стройплощадки / ПОС", "ПОС и временная инфраструктура"
    )


def test_completed_works_are_cost_evidence_not_physical_wbs_percent():
    cut = datetime.date(2026, 8, 1)
    estimate = {
        "rows": [{"code": "2.2.1.4", "parent": ""}],
        "by_code": {"2.2.1.4": {"estimate": 100.0}},
    }
    works = {"rows": [
        {"code": "2.2.1.4", "construction": True,
         "date": datetime.date(2026, 7, 15), "amount": 40.0},
        {"code": "2.2.1.4", "construction": False,
         "date": datetime.date(2026, 7, 20), "amount": 50.0},
        {"code": "2.2.1.4", "construction": True,
         "date": datetime.date(2026, 8, 15), "amount": 30.0},
    ]}

    result = manager._metrics(estimate, works, "2.2.1.4", cut)

    assert result["accepted"] == pytest.approx(40.0)
    assert result["accepted_ratio"] == pytest.approx(0.4)
    assert "progress" not in result


def test_money_pace_can_never_push_a_wbs_finish_to_2031(monkeypatch):
    monkeypatch.setattr(manager, "_baseline_status", lambda _: {
        "100": {"closed": False, "status": "В работе"}
    })
    schedule = {"rows": [{
        "id": "100",
        "wbs": "1.16.2.5",
        "plan_finish": "2027-03-31",
        "forecast_finish": "2031-06-01",
        "actual_progress": 0.829,
        "rate_3m": 0.002,
        "delta_days": 1523,
    }]}

    manager._sanitize_base_schedule("demo", schedule)
    row = schedule["rows"][0]

    assert row["forecast_finish"] == "2027-03-31"
    assert row["delta_days"] == 0
    assert row["actual_progress"] is None
    assert row["rss_accepted_ratio"] == pytest.approx(0.829)
    assert row["progress_kind"] == "accepted_cost_ratio"


def test_closed_wbs_stays_closed_even_if_cost_ratio_is_below_one(monkeypatch):
    monkeypatch.setattr(manager, "_baseline_status", lambda _: {
        "914": {"closed": True, "status": "Завершено"}
    })
    schedule = {"rows": [{
        "id": "914",
        "wbs": "1.16.1.2.4",
        "plan_finish": "2025-06-09",
        "forecast_finish": "2031-01-01",
        "actual_progress": 0.829,
        "rate_3m": 0.0,
        "delta_days": 2000,
    }]}

    manager._sanitize_base_schedule("demo", schedule)
    row = schedule["rows"][0]

    assert row["baseline_closed"] is True
    assert row["forecast_finish"] == "2025-06-09"
    assert row["status"] == "ЗАВЕРШЕНО ПО УТВЕРЖДЕННОМУ ГПР"


def test_own_gpr_percent_is_the_floor_of_task_readiness(monkeypatch):
    """КС статьи — готовность всей статьи, а не задачи.

    «Разработка котлована», физически пройденная, наследовала статейные 67,8%
    и давала +472 дня, из которых +83 доезжали до РНВ по FS-цепочке. Свой
    процент задачи из утверждённого ГПР — нижняя граница: задача не может
    быть менее готова, чем принято в baseline.
    """
    monkeypatch.setattr(manager, "_baseline_status", lambda _: {
        "1393": {"closed": False, "status": "Просрочено", "progress": 0.9},
    })
    schedule = {"rows": [{
        "id": "1393",
        "wbs": "1.16.2.1",
        "plan_finish": "2025-10-28",
        "forecast_finish": "2027-02-12",
        "actual_progress": 0.678,
        "rate_3m": 0.02,
        "delta_days": 472,
    }]}

    manager._sanitize_base_schedule("demo", schedule)
    row = schedule["rows"][0]

    assert row["rss_accepted_ratio"] == pytest.approx(0.9)
    assert row["progress_kind"] == "accepted_cost_ratio_floor_gpr"


def test_a_task_at_full_gpr_percent_is_closed_without_the_word(monkeypatch):
    """100% в ГПР значит «сделано», даже если статус забыли перевести."""
    import developaid_monitor as monitor

    monkeypatch.setattr(monitor, "_read_baseline_gpr", lambda _: {"works": [
        {"id": "7", "status": "Просрочено", "progress": 1.0},
        {"id": "8", "status": "В работе", "progress": 0.55},
        {"id": "9", "status": "Завершено", "progress": None},
    ]})

    out = manager._baseline_status("demo")

    assert out["7"]["closed"] is True
    assert out["8"]["closed"] is False
    assert out["8"]["progress"] == pytest.approx(0.55)
    assert out["9"]["closed"] is True


def test_gpr_percent_written_as_hundred_scale_is_normalized(monkeypatch):
    import developaid_monitor as monitor

    monkeypatch.setattr(monitor, "_read_baseline_gpr", lambda _: {"works": [
        {"id": "5", "status": "В работе", "progress": 86.0},
    ]})

    out = manager._baseline_status("demo")

    assert out["5"]["progress"] == pytest.approx(0.86)
    assert out["5"]["closed"] is False


def _gpr_book(rows):
    """Миниатюрный ГПР: лист, шапка в 4-й строке, работы с 5-й."""
    import io
    from openpyxl import Workbook

    book = Workbook()
    sheet = book.active
    sheet.title = "ГПР"
    header = ["ID", "WBS", "Раздел", "Объект", "Наименование работ", "Тип строки",
              "% выполнения", "Начало", "Окончание", "Статус", "Длительность р.д.",
              "Предшественники", "Связанный тендер", "Окончание тендера",
              "Резерв до начала работ", "Увязка", "Код РСС", "Статья РСС",
              "Основание привязки"]
    for c, value in enumerate(header, 1):
        sheet.cell(row=4, column=c, value=value)
    for i, (rid, name, progress, start, finish, status, code) in enumerate(rows):
        line = 5 + i
        values = [rid, str(rid), "СМР", "Корпус 1", name, "Работа", progress,
                  start, finish, status, 10, "", "", "", "", "", code, "", ""]
        for c, value in enumerate(values, 1):
            sheet.cell(row=line, column=c, value=value)
    blob = io.BytesIO()
    book.save(blob)
    return blob.getvalue()


def test_weekly_schedule_fact_overlays_percent_over_the_baseline(tmp_path, monkeypatch):
    """План зафиксирован baseline; выполнение приезжает еженедельным ГПР-фактом.

    На Кутузове котлован стоял в baseline нулём и десять месяцев тянул прогноз
    на +463 дня; свежий ГПР со 100% снимает его без правки baseline.
    """
    import datetime
    import developaid_monitor as monitor

    monkeypatch.setattr(monitor, "_SNAPSHOT_DIR", tmp_path)
    d1 = datetime.date(2025, 8, 23)
    d2 = datetime.date(2025, 10, 28)
    monitor.store_schedule("Кутузов", _gpr_book([
        (1393, "Разработка котлована", 0.0, d1, d2, "Просрочено", "2.2.1.1"),
        (1500, "Кровля", 0.0, d1, d2, "Просрочено", "2.2.2.3"),
    ]), None, "2026-07-23")

    before = manager._baseline_status("Кутузов")
    assert before["1393"]["closed"] is False

    stored = monitor.store_schedule_fact("Кутузов", _gpr_book([
        (1393, "Разработка котлована", 1.0, d1, d2, "Завершено", "2.2.1.1"),
        (1500, "Кровля", 0.4, d1, d2, "В работе", "2.2.2.3"),
    ]), "2026-08-20")
    assert stored["completed"] == 1

    after = manager._baseline_status("Кутузов")
    assert after["1393"]["closed"] is True
    assert after["1500"]["closed"] is False
    assert after["1500"]["progress"] == pytest.approx(0.4)


def test_a_schedule_fact_snapshot_is_never_overwritten(tmp_path, monkeypatch):
    import datetime
    import developaid_monitor as monitor

    monkeypatch.setattr(monitor, "_SNAPSHOT_DIR", tmp_path)
    data = _gpr_book([(1, "Работа", 0.5,
                       datetime.date(2026, 1, 1), datetime.date(2026, 6, 1),
                       "В работе", "2.2.1.1")])
    monitor.store_schedule_fact("Кутузов", data, "2026-08-20")

    with pytest.raises(FileExistsError):
        monitor.store_schedule_fact("Кутузов", data, "2026-08-20")


def test_the_baseline_plan_dates_stay_even_with_a_fact_snapshot(tmp_path, monkeypatch):
    """ГПР-факт двигает выполнение, но не план: план — это baseline."""
    import datetime
    import developaid_monitor as monitor

    monkeypatch.setattr(monitor, "_SNAPSHOT_DIR", tmp_path)
    d1, d2 = datetime.date(2025, 8, 23), datetime.date(2025, 10, 28)
    monitor.store_schedule("Кутузов", _gpr_book([
        (1393, "Разработка котлована", 0.0, d1, d2, "Просрочено", "2.2.1.1"),
    ]), None, "2026-07-23")
    monitor.store_schedule_fact("Кутузов", _gpr_book([
        (1393, "Разработка котлована", 1.0,
         datetime.date(2026, 1, 1), datetime.date(2026, 7, 1),
         "Завершено", "2.2.1.1"),
    ]), "2026-08-20")

    baseline = monitor._read_baseline_gpr("Кутузов")
    work = baseline["works"][0]
    assert work["start"] == d1
    assert work["finish"] == d2
