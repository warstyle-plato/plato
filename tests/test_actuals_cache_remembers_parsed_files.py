"""Разобранный файл помнится, а не читается заново на каждый взгляд.

Один срез монитора читал один и тот же РСС двадцать раз — смета ×4, платежи
×3, акты ×5 — и отвечал 29–39 секунд при лимите nginx в 60. Кэш по
(путь, mtime, размер) сводит это к одному разбору; наружу уходит копия,
чтобы правка результата не портила кэш.
"""

from __future__ import annotations

import io

import pytest

from openpyxl import Workbook

import developaid_actuals as actuals


def _estimate_book(path, amount=1000.0):
    book = Workbook()
    sheet = book.active
    sheet.title = "Расчет стоимости строительства"
    sheet.cell(row=9, column=1, value="Код")
    sheet.cell(row=10, column=1, value="2.2")
    sheet.cell(row=10, column=5, value=amount)
    sheet.cell(row=12, column=4, value="Всего инвестиционные расходы")
    sheet.cell(row=12, column=5, value=amount)
    book.save(path)


def test_the_same_file_is_parsed_once(tmp_path, monkeypatch):
    path = tmp_path / "rss.xlsx"
    _estimate_book(path)

    import openpyxl

    calls = {"n": 0}
    original = openpyxl.load_workbook

    def counting(*args, **kwargs):
        calls["n"] += 1
        return original(*args, **kwargs)

    # _sheet импортирует load_workbook из openpyxl при каждом вызове
    monkeypatch.setattr(openpyxl, "load_workbook", counting)
    first = actuals.read_estimate(path)
    second = actuals.read_estimate(path)

    assert calls["n"] == 1
    assert first == second


def test_a_mutated_result_does_not_poison_the_cache(tmp_path):
    path = tmp_path / "rss.xlsx"
    _estimate_book(path)

    first = actuals.read_estimate(path)
    first["rows"].clear()
    first["total"]["estimate"] = -1.0

    second = actuals.read_estimate(path)
    assert second["total"]["estimate"] == pytest.approx(1000.0)
    assert second["rows"]


def test_a_rewritten_file_is_read_fresh(tmp_path):
    path = tmp_path / "rss.xlsx"
    _estimate_book(path, amount=1000.0)
    assert actuals.read_estimate(path)["total"]["estimate"] == pytest.approx(1000.0)

    import os
    _estimate_book(path, amount=2000.0)
    stat = os.stat(path)
    os.utime(path, ns=(stat.st_atime_ns, stat.st_mtime_ns + 1_000_000))

    assert actuals.read_estimate(path)["total"]["estimate"] == pytest.approx(2000.0)


def test_a_stream_is_not_cached():
    blob = io.BytesIO()
    book = Workbook()
    sheet = book.active
    sheet.title = "Расчет стоимости строительства"
    sheet.cell(row=9, column=1, value="Код")
    sheet.cell(row=12, column=4, value="Всего инвестиционные расходы")
    sheet.cell(row=12, column=5, value=500.0)
    book.save(blob)
    blob.seek(0)

    assert actuals.read_estimate(blob)["total"]["estimate"] == pytest.approx(500.0)
