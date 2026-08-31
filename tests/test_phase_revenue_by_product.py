"""Выручка очереди разложена по продуктам — и на экране, и в книге.

В сравнении очередей выручка стояла одной строкой. Она не отвечает на вопрос,
чем очередь живёт: у одной весь объём в квартирах, у другой треть в паркинге и
ОСЗ, а маржа и риск у них разные (владелец, 23.08.2026). Числа для разбивки
лежали в отчёте каждой очереди и просто не выводились.

Разбивка не считается заново: берётся `report.products` очереди. Второй счёт
той же выручки однажды разошёлся бы с первым, и обе строки выглядели бы верными.

Запуск: python3 -m pytest tests/test_phase_revenue_by_product.py -q
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT.parent))
sys.path.insert(0, str(ROOT))

import main as wrapper  # noqa: E402

core = wrapper.core


@pytest.fixture(scope="module")
def bundle():
    inputs = dict(core.DEFAULT_INPUTS)
    # Офисы и ОСЗ включены нарочно: с одними квартирами разбивка ничего не
    # показывает, а вся её польза — в очереди со смешанным составом.
    inputs.update({"offices_enabled": True, "retail_enabled": True})
    tep = {key: dict(value) for key, value in core.TEP_DEFAULT.items()}
    return core._run_authoritative_model(
        inputs, tep, [], {"enabled": True, "phase_count": 2, "phase_gap_months": 12})


# --- разбивка едет из расчёта -------------------------------------------------

def test_every_queue_carries_its_revenue_by_product(bundle):
    for item in bundle["comparison"]:
        assert item["revenue_by_product"], f"{item['name']}: разбивки нет"
        assert item["saleable_by_product"], f"{item['name']}: объёма продаж нет"


def test_the_parts_add_up_to_the_queue_revenue(bundle):
    """Сумма продуктов — это и есть выручка очереди, а не близкое к ней число."""
    for item in bundle["comparison"]:
        assert sum(item["revenue_by_product"].values()) == pytest.approx(
            item["revenue"], rel=1e-9), item["name"]


def test_the_queues_add_up_to_the_project(bundle):
    """По каждому продукту сумма очередей сходится со сводом."""
    consolidated = {str(p["key"]): float(p.get("revenue") or 0.0)
                    for p in bundle["consolidated"]["report"]["products"]}
    for key, total in consolidated.items():
        by_queue = sum(float((item["revenue_by_product"] or {}).get(key) or 0.0)
                       for item in bundle["comparison"])
        assert by_queue == pytest.approx(total, rel=1e-9), key


def test_the_split_is_not_counted_a_second_time():
    """Числа берутся из отчёта очереди, а не считаются здесь заново."""
    import inspect

    # Цикл очередей живёт в однопроходной `_calculate_phased_once`:
    # `calculate_phased` стала обёрткой с переносом долга между очередями.
    source = inspect.getsource(core._calculate_phased_once)
    block = source[source.index('"revenue_by_product"'):]
    block = block[:block.index('"cash_shared_cost"')]
    assert 'report' in block and 'products' in block
    assert "price" not in block and "*" not in block, "разбивка снова считается своей формулой"


def test_a_discrete_object_lands_in_one_queue_only(bundle):
    """Офисы и ОСЗ размещаются дискретно — выручка обязана стоять в своей
    очереди целиком, а не размазаться по всем."""
    for key in ("offices", "standalone_retail"):
        values = [float((item["revenue_by_product"] or {}).get(key) or 0.0)
                  for item in bundle["comparison"]]
        assert sum(1 for v in values if v > 0) == 1, f"{key}: {values}"


# --- экран --------------------------------------------------------------------

def test_the_screen_lists_the_products_under_revenue():
    page = core.PAGE
    found = re.search(r"\nfunction renderPhaseComparison\(.*?\n\}", page, re.S)
    assert found, "функция сравнения очередей на странице не найдена"
    body = found.group(0)
    assert "revenue_by_product" in body
    # Строки продуктов стоят сразу под выручкой, а не в конце таблицы.
    revenue = body.index("['Выручка',")
    assert body.index("...prodRows") > revenue
    assert body.index("['Цена реализации на м² продаваемой'") > body.index("...prodRows")


def test_the_screen_hides_products_without_revenue():
    """Семь нулевых строк — это шум, а не полнота."""
    body = re.search(r"\nfunction renderPhaseComparison\(.*?\n\}", core.PAGE, re.S).group(0)
    assert "filter(k=>c.some(x=>Number((x.revenue_by_product||{})[k]||0)>0))" in body


# --- книга --------------------------------------------------------------------
#
# Разбивка жила в выгрузке детализации, а её сняли вместе с архивом очередей
# (владелец, 30.08.2026: рабочая книга одна). Требование от этого не исчезло —
# проверять его надо там, где книга теперь одна: в КОНСОЛИДАТОРЕ книги v4.

BOOK_INPUTS = {"offices_enabled": True, "retail_enabled": True}
BOOK_PHASING = {"enabled": True, "mode": "phased", "user_enabled": True,
                "phase_count": 2, "phase_gap_months": 12,
                "phases": [{"name": "О1", "start_offset_months": 0},
                           {"name": "О2", "start_offset_months": 12}],
                "shared_cash": {}, "shared_allocation": {}, "social_objects": []}


def _book():
    import io as _io
    import openpyxl
    sys.setrecursionlimit(400000)
    inputs = {**core.DEFAULT_INPUTS, **BOOK_INPUTS}
    tep = {key: dict(value) for key, value in core.TEP_DEFAULT.items()}
    content, _, meta = core.build_project_workbook(
        inputs, tep, [], dict(BOOK_PHASING), project_name="Разбивка")
    assert meta["missing"] == [], meta["missing"]
    sheet = openpyxl.load_workbook(_io.BytesIO(content))["КОНСОЛИДАТОР"]
    columns = {}
    for column in range(17, 30):
        title = sheet.cell(row=3, column=column).value
        if isinstance(title, str) and title.startswith("Выручка · "):
            columns[title[len("Выручка · "):-len(", млн ₽")]] = column
    return sheet, columns, content


def test_the_workbook_has_the_same_columns(bundle):
    _, columns, _ = _book()
    assert columns, "в книге разбивки нет — книга и экран скажут разное"
    labels = {p["label"] for p in bundle["consolidated"]["report"]["products"]}
    for title in columns:
        assert title in labels, title


def test_the_workbook_numbers_match_the_engine(bundle):
    """Книга считает разбивку своими формулами — и обязана сойтись с движком."""
    import io as _io
    import openpyxl
    from xlsx_eval import Evaluator
    _, columns, content = _book()
    book = Evaluator(openpyxl.load_workbook(_io.BytesIO(content)))
    by_label = {p["label"]: p["key"] for p in bundle["consolidated"]["report"]["products"]}
    checked = 0
    for title, column in columns.items():
        letter = openpyxl.utils.get_column_letter(column)
        for index, row in enumerate(bundle["comparison"]):
            engine = float((row.get("revenue_by_product") or {}).get(by_label[title]) or 0.0) / 1e6
            assert book.cell("КОНСОЛИДАТОР", f"{letter}{4 + index}") == pytest.approx(
                engine, abs=1.0, rel=0.005), (title, row["name"])
            checked += 1
    assert checked >= 6, "проверено слишком мало — выборка не та"


def test_the_workbook_total_row_sums_each_product():
    """Итог книги — сумма очередей её же формулой, а не второй счёт."""
    sheet, columns, _ = _book()
    assert columns
    for column in columns.values():
        letter = __import__("openpyxl").utils.get_column_letter(column)
        assert sheet.cell(row=8, column=column).value == f"=SUM({letter}4:{letter}7)"


def test_a_project_without_extra_products_gets_no_empty_columns():
    """Колонка заводится под продукт, у которого выручка есть.

    Семь нулевых колонок — шум, а не полнота: то же правило, что на экране.
    """
    import io as _io
    import openpyxl
    sys.setrecursionlimit(400000)
    inputs = dict(core.DEFAULT_INPUTS)   # офисы и ОСЗ выключены
    tep = {key: dict(value) for key, value in core.TEP_DEFAULT.items()}
    content, _, meta = core.build_project_workbook(
        inputs, tep, [], dict(BOOK_PHASING), project_name="Только жильё")
    assert meta["missing"] == [], meta["missing"]
    sheet = openpyxl.load_workbook(_io.BytesIO(content))["КОНСОЛИДАТОР"]
    titles = [sheet.cell(row=3, column=c).value for c in range(17, 30)]
    titles = [t for t in titles if isinstance(t, str)]
    assert titles, "разбивки нет вовсе"
    for absent in ("Офисы", "Коммерция ОСЗ", "Наземный паркинг", "Кладовые"):
        assert not any(absent in t for t in titles), (absent, titles)
