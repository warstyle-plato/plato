"""Надбавка к ключевой ступенчатая — по фактическому покрытию эскроу.

В договорах НКЛ ставка на покрытую эскроу часть долга идёт лестницей: BVX003 —
3,47% при покрытии 100–110%, 1,75% при 110–120%, 0,03% при 120–130%, дальше
0,01%. У нас была одна ставка на всё покрытие, и модель завышала проценты
сильным проектам. Таблица индивидуальна для каждого НКЛ, поэтому «типовую»
зашивать нельзя: ступени вводятся полем, пустое поле оставляет прежнее
поведение.

Отдельно проверяется достижимость: на вводных по умолчанию покрытие упирается
в 0,76×, и тест, написанный на них, прошёл бы, ни разу не зайдя в лестницу.

Запуск: python3 -m pytest tests/test_escrow_rate_steps.py -q
"""

from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import main as wrapper  # noqa: E402

core = wrapper.core

LADDER = "100:3,47; 110:1,75; 120:0,03; 130:0,01"

# Проект, где эскроу перекрывает долг: дешёвый вход, дорогие метры, продажи до
# РВЭ. На вводных по умолчанию покрытие до единицы не доходит вовсе.
STRONG = dict(purchase_price_mln=500.0, apartment_price_th=700,
              commercial_price_th=700, share_before_rve_pct=95,
              land_rights_cost_mln=0.0, vri_required=False,
              social_compensation_mln=0.0, kindergarten_places=0)


def _run(**extra):
    inputs = {**core.DEFAULT_INPUTS, **STRONG, **extra}
    tep = {key: dict(value) for key, value in core.TEP_DEFAULT.items()}
    return core.calculate(core.CalcRequest(inputs=inputs, tep=tep, rates=[]))


def test_the_ladder_is_read_the_way_a_contract_is_written():
    """Строку переписывают из договора руками — разбор это терпит."""
    assert core.pf_special_steps(LADDER) == [
        (1.0, 0.0347), (1.1, 0.0175), (1.2, 0.0003), (1.3, 0.0001)]
    # Проценты и разы — одна и та же лестница.
    assert core.pf_special_steps("1,0:3,47 1,1:1,75") == [(1.0, 0.0347), (1.1, 0.0175)]
    assert core.pf_special_steps("110% - 1.75%\n100% - 3.47%") == [
        (1.0, 0.0347), (1.1, 0.0175)], "порядок ввода лестницу не задаёт"
    assert core.pf_special_steps("") == []
    assert core.pf_special_steps("ставка по договору") == []


def test_below_the_first_step_the_usual_rate_applies():
    """Лестница в договоре начинается со стопроцентного покрытия. Ниже —
    обычная специальная ставка, а не выдуманная за банк ступень."""
    steps = core.pf_special_steps(LADDER)
    assert core.pf_special_rate_at(0.0, steps, 0.045) == 0.045
    assert core.pf_special_rate_at(0.99, steps, 0.045) == 0.045
    assert core.pf_special_rate_at(1.0, steps, 0.045) == 0.0347
    assert core.pf_special_rate_at(1.19, steps, 0.045) == 0.0175
    assert core.pf_special_rate_at(5.0, steps, 0.045) == 0.0001
    # Ступеней нет — ставка одна, как было.
    assert core.pf_special_rate_at(2.0, [], 0.045) == 0.045


def test_the_test_project_actually_reaches_the_ladder():
    """Зелёный тест не значит, что ветка верна — возможно, до неё не доходят."""
    finance = _run()["finance"]
    coverage = [row["coverage"] for row in finance["rows"] if row["pf_balance"] > 0]
    assert max(coverage) > 1.3, max(coverage)


def test_an_empty_field_keeps_the_single_rate():
    """Пустое поле — прежнее поведение: одна ставка на весь срок.

    С 0.19.35 умолчание поля — лестница НКЛ Сбера (решение владельца), поэтому
    «как было» задаётся теперь явно пустой строкой, а не отсутствием правки.
    """
    plain = _run(pf_special_steps="")["finance"]
    assert plain["pf_special_steps"] == []
    stepped = _run()["finance"]
    assert stepped["pf_interest"] < plain["pf_interest"], (
        "умолчание — лестница, и на сильном проекте она дешевле одной ставки")


def test_the_ladder_makes_a_strong_project_cheaper():
    """Ради этого всё и затевалось: одна ставка завышала проценты сильным."""
    plain = _run(pf_special_steps="")["finance"]
    stepped = _run(pf_special_steps=LADDER)["finance"]
    assert stepped["pf_interest"] < plain["pf_interest"] * 0.5, (
        plain["pf_interest"], stepped["pf_interest"])
    assert stepped["avg_pf_rate"] < plain["avg_pf_rate"]


def test_the_default_is_the_sber_ladder_and_it_is_named():
    """Умолчание — числа из договора, и договор назван.

    Прежде поле было пустым: таблица у каждого НКЛ своя, и типовую не зашивали.
    Владелец решил иначе (20.08.2026): «ставим базово по умолчанию то, что у
    Сбера, а человек может вручную вбить или оставить». Раз в каждый новый
    проект уезжают числа конкретного договора, договор обязан быть назван — иначе
    они читаются как норма.
    """
    assert core.DEFAULT_INPUTS["pf_special_steps"] == core.PF_SPECIAL_STEPS_DEFAULT
    assert core.pf_special_steps(core.PF_SPECIAL_STEPS_DEFAULT) == [
        (1.0, 0.0347), (1.1, 0.0175), (1.2, 0.0003), (1.3, 0.0001)]
    assert "400F00BVX003" in core.PF_SPECIAL_STEPS_SOURCE


def test_the_default_changes_nothing_where_the_coverage_never_gets_there():
    """Умолчание не двигает проекты, не дотягивающие до первой ступени.

    Ниже первой ступени действует обычная специальная ставка, а на вводных по
    умолчанию покрытие до 1× не доходит. Значит, смена умолчания не переписала
    экономику всем разом — и это проверяется, а не предполагается.
    """
    tep = {key: dict(value) for key, value in core.TEP_DEFAULT.items()}
    def run(steps):
        inputs = {**core.DEFAULT_INPUTS, "pf_special_steps": steps}
        return core.calculate(core.CalcRequest(inputs=inputs, tep=tep, rates=[]))["finance"]
    plain, stepped = run(""), run(core.PF_SPECIAL_STEPS_DEFAULT)
    coverage = [row["coverage"] for row in stepped["rows"] if row["pf_balance"] > 0]
    assert coverage and max(coverage) < 1.0, max(coverage or [0])
    assert plain["pf_interest"] == stepped["pf_interest"]
    assert plain["avg_pf_rate"] == stepped["avg_pf_rate"]


def test_the_report_says_which_step_worked_and_how_long():
    """Средняя ставка без лестницы объяснима, с лестницей — нет."""
    finance = _run(pf_special_steps=LADDER)["finance"]
    steps = finance["pf_special_steps"]
    assert [step["coverage_from_pct"] for step in steps] == [100.0, 110.0, 120.0, 130.0]
    assert [step["rate_pct"] for step in steps] == [3.47, 1.75, 0.03, 0.01]
    months = sum(step["months"] for step in steps)
    with_pf = len([row for row in finance["rows"] if row["pf_balance"] > 0])
    assert 0 < months <= with_pf, (months, with_pf)


def test_the_monthly_row_carries_the_rate_of_its_own_step():
    """Помесячная таблица должна отвечать «почему столько» без вводных."""
    finance = _run(pf_special_steps=LADDER)["finance"]
    steps = core.pf_special_steps(LADDER)
    checked = 0
    for row in finance["rows"]:
        if row["pf_balance"] <= 0 or row["coverage"] <= 0:
            continue
        expected = core.pf_special_rate_at(
            row["coverage"], steps, core.DEFAULT_INPUTS["pf_special_pct"] / 100)
        assert abs(row["pf_special_rate"] - expected) < 1e-12, row["month"]
        checked += 1
    assert checked, "месяцев с ПФ не нашлось — проверять было нечего"


def test_the_steps_reach_the_report_block():
    financing = _run(pf_special_steps=LADDER)["report"]["financing"]
    assert financing["pf_special_steps"], financing.keys()
    assert financing["pf_special_steps"][0]["rate_pct"] == 3.47


def test_the_pdf_prints_the_ladder():
    financing = _run(pf_special_steps=LADDER)["report"]["financing"]
    rows = core._pdf_pf_step_rows(financing)
    assert rows and rows[0][0].startswith("Ступень: покрытие от 100")
    assert "3,47%" in rows[0][1] and "мес." in rows[0][1]
    # Ступеней нет — строк нет: пустая таблица читалась бы как их отсутствие
    # в договоре, а не как невведённые вводные.
    assert core._pdf_pf_step_rows({"pf_special_steps": []}) == []


def test_the_field_is_declared_where_the_page_takes_it():
    fields = [item[0] for group in core.FIELD_GROUPS for item in group[1]]
    assert "pf_special_steps" in fields
    assert core.DEFAULT_INPUTS["pf_special_steps"] == core.PF_SPECIAL_STEPS_DEFAULT
    finance_group = next(group for group in core.FIELD_GROUPS if group[0] == "Финансирование")
    field = next(item for item in finance_group[1] if item[0] == "pf_special_steps")
    # Лестница рисуется таблицей диапазонов, как в НКЛ (владелец, 20.08.2026);
    # хранится по-прежнему строкой того же формата — второго хранилища нет.
    assert field[3] == "pf_steps", "лестница — таблица диапазонов, не текстовое поле"
    page = core.PAGE
    assert "renderPfStepsEditor" in page and "pfStepAdd" in page
    assert "и выше" in page, "верхний диапазон назван, а не подразумевается"
    assert "Пусто" in field[2], "пустое поле — прежнее поведение, и это сказано"
    assert "по умолчанию" in field[2], "откуда взялись числа в поле — сказано там же"


def test_the_formula_for_excel_matches_the_engine():
    """Методику правят в двух местах — движок и книга. Формулу собирает тот же
    код, что считает ступени, иначе они разойдутся молча."""
    steps = core.pf_special_steps(LADDER)
    formula = core.pf_special_steps_formula("C53", steps, "0.045")
    assert formula == (
        "IF(C53>=1.3,0.0001,IF(C53>=1.2,0.0003,"
        "IF(C53>=1.1,0.0175,IF(C53>=1,0.0347,0.045))))")
    assert core.pf_special_steps_formula("C53", [], "0.045") == "0.045"


def test_the_generated_workbook_carries_the_ladder():
    import io
    import re
    import zipfile

    import openpyxl

    inputs = {**core.DEFAULT_INPUTS, **STRONG, "pf_special_steps": LADDER}
    tep = {key: dict(value) for key, value in core.TEP_DEFAULT.items()}
    data, _meta = core.build_plato_model_v2(inputs, tep, None, "Ступени")
    book = openpyxl.load_workbook(io.BytesIO(data))
    ws = book["Вводные"]
    # Лестница — вводная и в книге: числовые строки, которые можно править
    # (владелец, 20.08.2026: «сейчас-то надбавка — это вводная»). Текстовое поле
    # разбирается при сборке, его правка в Excel ничего не пересчитывает —
    # пересчитывают эти строки.
    ladder_rows = {}
    for row in range(1, ws.max_row + 1):
        label = str(ws.cell(row=row, column=1).value or "")
        if label.startswith("Ступень"):
            ladder_rows[label] = (row, ws.cell(row=row, column=2).value)
    assert ladder_rows["Ступень 1 — покрытие от"][1] == 1
    assert ladder_rows["Ступень 4 — ставка"][1] == 0.0001
    # Помесячная формула ссылается на эти ячейки, а не несёт числа в себе.
    edge_row = ladder_rows["Ступень 4 — покрытие от"][0]
    found = 0
    for name in book.sheetnames:
        for row in book[name].iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and f"Вводные!$B${edge_row}" in value:
                    found += 1
    assert found > 10, found
    del re, zipfile  # прежний тест искал зашитые числа в XML — их больше нет


def test_the_plato_template_gets_the_ladder_too():
    """Строка 57 листа «КРЕДИТЫ» — та самая, на которой книга уже расходилась
    с движком по ставке ПФ (360,3 млн ₽ против 746,5)."""
    openpyxl = __import__("openpyxl")
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "КРЕДИТЫ"
    for column in ("C", "D"):
        sheet[f"{column}55"] = f"=IF({column}$3<$D61,IF({column}53>1,{column}57,{column}56),{column}$13)"
        sheet[f"{column}56"] = f"={column}54*(1-{column}53)+{column}57*{column}53"
        sheet[f"{column}57"] = 0.045
    filled: list = []
    missing: list = []
    core._plato_apply_pf_rate_methodology(
        book, filled, missing, core.pf_special_steps(LADDER))
    assert not missing, missing
    # Без листа «Вводные» в этой мини-книге ссылаться некуда — числа в формуле.
    assert sheet["C57"].value.startswith("=IF(C53>=1.3,0.0001")
    assert "0.045" in sheet["C57"].value, "ниже первой ступени — обычная ставка"
    assert any("ступени" in str(item.get("label", "")) for item in filled), filled


def test_the_plato_ladder_is_editable_cells_not_baked_numbers():
    """В настоящем шаблоне лестница — вводная: ячейки внизу «Вводных».

    Человек работает книгой и правит лестницу как данные, а не как формулу по
    всем месячным колонкам двух очередей. Вставить строки в середину листа
    нельзя — openpyxl при сдвиге не переписывает формулы, а под блоком
    финансирования вся карта записи фиксированными адресами, — поэтому блок
    стоит в конце листа, где свободно и куда никто не ссылается.
    """
    openpyxl = __import__("openpyxl")
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "КРЕДИТЫ"
    inputs_sheet = book.create_sheet("Вводные")
    inputs_sheet["A1"] = "ВВОДНЫЕ"
    for column in ("C", "D"):
        sheet[f"{column}55"] = f"=IF({column}$3<$D61,IF({column}53>1,{column}57,{column}56),{column}$13)"
        sheet[f"{column}56"] = f"={column}54*(1-{column}53)+{column}57*{column}53"
        sheet[f"{column}57"] = 0.045
    filled: list = []
    missing: list = []
    core._plato_apply_pf_rate_methodology(
        book, filled, missing, core.pf_special_steps(LADDER))
    assert not missing, missing
    labels = [str(inputs_sheet.cell(row=row, column=1).value or "")
              for row in range(1, inputs_sheet.max_row + 1)]
    assert any(label.startswith("СТУПЕНИ СТАВКИ ПФ") for label in labels)
    edge_rows = [row for row in range(1, inputs_sheet.max_row + 1)
                 if str(inputs_sheet.cell(row=row, column=1).value or "").endswith("покрытие от")]
    assert len(edge_rows) == 4
    assert inputs_sheet.cell(row=edge_rows[0], column=2).value == 1.0
    # Формула ссылается на ячейки; пустой порог читается нулём и гасится
    # защитой «порог > 0» — иначе стёртая ступень срабатывала бы всегда.
    formula = sheet["C57"].value
    assert f"Вводные!$B${edge_rows[0]}" in formula
    assert f"AND(Вводные!$B${edge_rows[0]}>0" in formula
    assert formula.rstrip(")").endswith("0.045"), "ниже первой ступени — обычная ставка"


def test_a_template_that_does_not_take_the_ladder_is_named_out_loud():
    """Ступени заданы, а в книгу не легли — расхождение методик. Молчать
    нельзя: книга посчитает по одной ставке, отчёт по лестнице, и оба будут
    выглядеть достоверно."""
    openpyxl = __import__("openpyxl")
    book = openpyxl.Workbook()
    book.active.title = "КРЕДИТЫ"
    filled: list = []
    missing: list = []
    core._plato_apply_pf_rate_methodology(
        book, filled, missing, core.pf_special_steps(LADDER))
    assert missing, "переписанный шаблон обязан попасть в missing"


def test_the_page_prints_the_ladder_with_its_own_code():
    """Строки ступеней рисует настоящий код страницы, а не его пересказ."""
    import json
    import re
    import shutil
    import subprocess

    import pytest

    node = shutil.which("node")
    if not node:
        pytest.skip("node недоступен")
    match = re.search(r"function pfStepRows\(f\)\{.*?\n\}", core.PAGE, re.S)
    assert match, "функция ступеней на странице не найдена"
    financing = {"pf_special_steps": [
        {"coverage_from_pct": 100.0, "rate_pct": 3.47, "months": 2},
        {"coverage_from_pct": 130.0, "rate_pct": 0.01, "months": 21}]}
    script = (
        "function row(label,value){return `<tr><td>${label}</td><td>${value}</td></tr>`}\n"
        + match.group(0) + "\n"
        f"console.log(JSON.stringify([pfStepRows({json.dumps(financing)}),"
        "pfStepRows({}),pfStepRows({pf_special_steps:[]})]));\n"
    )
    done = subprocess.run([node, "-e", script], capture_output=True, text=True, timeout=30)
    assert done.returncode == 0, done.stderr
    stepped, missing, empty = json.loads(done.stdout)
    assert "покрытие от 100%" in stepped and "3,47%" in stepped and "2 мес." in stepped
    assert "покрытие от 130%" in stepped and "21 мес." in stepped
    # Расчёт без ступеней и расчёт постарше — обе строки пустые, а не «0».
    assert missing == "" and empty == ""
