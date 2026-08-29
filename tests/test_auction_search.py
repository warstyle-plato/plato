from datetime import date, datetime, timedelta
import io
from urllib.parse import parse_qs, urlparse
from zoneinfo import ZoneInfo

from auction_search.adapters.lot_online import LotOnlineAdapter
from auction_search.api import _discovery_adapters, _handoff_land_cadastres, _xlsx
from auction_search.bridge import auction_page_with_handoff, install_page_bridge
from auction_search.classifier import classify_lot
from auction_search.developaid_mapper import build_developaid_seed
from auction_search.documents import DocumentAuthorizationRequired, _looks_like_login_page, _request_headers
from auction_search.krt import extract_krt_obligations, extract_krt_program
from auction_search.krt_pipeline import enrich_krt_from_official_documents
from auction_search.models import (
    AuctionDocument,
    AuctionLot,
    AuctionSource,
    KrtProgramItem,
    LotKind,
    Provenance,
    SourceKind,
)
from auction_search.preset_mapper import build_project_preset
from auction_search.service import AuctionSearchService


def source():
    return AuctionSource(
        platform=SourceKind.LOT_ONLINE,
        lot_url="https://catalog.lot-online.ru/example",
        external_lot_id="RAD-TEST",
        fetched_at="2026-08-21T18:00:00Z",
    )


def test_classifier_krt_overrides_generic_land_wording():
    assert classify_lot("Земельные участки. Право на заключение договора о комплексном развитии территории") == LotKind.KRT


def test_generic_right_to_lease_is_not_krt():
    assert classify_lot("Право на заключение договора аренды земельного участка") == LotKind.LAND_LEASE


def test_development_project_company_share_is_property_complex():
    title = "Продажа 100% доли юрлица — собственника проекта комплекса жилой застройки"
    assert classify_lot(title) == LotKind.PROPERTY_COMPLEX
    verbose_title = "Продажа доли в размере 100 (сто) % уставного капитала компании, владеющей зданием"
    assert classify_lot(verbose_title) == LotKind.PROPERTY_COMPLEX


def test_non_development_company_share_is_not_promoted_to_property_complex():
    title = "Продажа доли в размере 100% уставного капитала ООО ПРАЧЕЧНАЯ"
    assert classify_lot(title) == LotKind.OTHER


def test_small_ijs_is_filtered_out():
    lot = AuctionLot(
        source=source(),
        lot_kind=LotKind.LAND_SALE,
        title="Участок",
        land_area_sqm=900,
        permitted_use="ИЖС",
    )
    assert AuctionSearchService.is_development_relevant(lot) is False


def test_legacy_50_cadastral_prefix_is_not_excluded_for_moscow():
    lot = AuctionLot(
        source=source(),
        lot_kind=LotKind.LAND_SALE,
        title="Коммунарка",
        land_area_sqm=173_494,
        cadastral_numbers=["50:21:0120316:1221"],
        permitted_use="Многофункциональные общественные центры",
        address="Москва, Коммунарка",
    )
    assert AuctionSearchService.is_development_relevant(lot) is True


def test_rad_discovery_uses_official_public_catalogue_filters():
    url = LotOnlineAdapter._discovery_url()
    parsed = urlparse(url)
    query = parse_qs(parsed.query)
    assert parsed.hostname == "catalog.lot-online.ru"
    assert query["category_id"] == ["2"]
    assert query["dispatch"] == ["categories.view"]
    assert query["filter_fields[is_archive]"] == ["false"]
    assert query["q"] == ["москва"]
    assert query["items_per_page"] == ["96"]


def test_rad_history_is_opt_in_and_checks_project_company_shares():
    url = LotOnlineAdapter._discovery_url(category_id="85", include_archive=True)
    query = parse_qs(urlparse(url).query)
    assert query["category_id"] == ["85"]
    assert query["filter_fields[is_archive]"] == ["true"]
    assert LotOnlineAdapter._discovery_url() != url


def test_rad_current_project_shares_are_disabled_by_default(monkeypatch):
    adapter = LotOnlineAdapter()
    captured = {}

    def discover_urls(**kwargs):
        captured.update(kwargs)
        return []

    monkeypatch.setattr(adapter, "_discover_candidate_urls", discover_urls)
    assert adapter.discover_moscow() == []
    assert captured["category_ids"] == ("2",)
    assert captured["include_archive"] is False


def test_rad_current_project_shares_can_be_enabled_explicitly(monkeypatch):
    adapter = LotOnlineAdapter(include_project_shares=True)
    captured = {}

    def discover_urls(**kwargs):
        captured.update(kwargs)
        return []

    monkeypatch.setattr(adapter, "_discover_candidate_urls", discover_urls)
    assert adapter.discover_moscow() == []
    assert captured["category_ids"] == ("2", "85")
    assert captured["include_archive"] is False


def test_rad_current_discovery_rejects_test_and_deadline_less_catalogue_leaks(monkeypatch):
    adapter = LotOnlineAdapter()
    monkeypatch.setattr(
        adapter,
        "_discover_candidate_urls",
        lambda **_kwargs: ["test", "archive", "current"],
    )
    future = (datetime.now(ZoneInfo("Europe/Moscow")) + timedelta(days=2)).isoformat()
    lots = {
        "test": AuctionLot(
            source=source(),
            lot_kind=LotKind.LAND_SALE,
            title="[Тест] Тестовый лот |",
            address="Москва",
            application_deadline=future,
        ),
        "archive": AuctionLot(
            source=source(),
            lot_kind=LotKind.LAND_SALE,
            title="Прв аук 12:35 23.03.2021 |",
            address="Москва",
            application_deadline=None,
        ),
        "current": AuctionLot(
            source=source(),
            lot_kind=LotKind.LAND_SALE,
            title="Продажа земельного участка в Коммунарке",
            address="Москва",
            application_deadline=future,
        ),
    }
    monkeypatch.setattr(adapter, "fetch_lot", lots.__getitem__)

    discovered = adapter.discover_moscow()

    assert [lot.title for lot in discovered] == ["Продажа земельного участка в Коммунарке"]


def test_explicit_platform_test_lot_is_screened_as_noise():
    lot = AuctionLot(
        source=source(),
        lot_kind=LotKind.KRT,
        title="[Тест] Тестовый лот |",
        address="Москва",
    )

    screening = AuctionSearchService.screen_lot(lot)

    assert screening["development_relevant"] is False
    assert screening["rating"] == "Шум"
    assert "тестовая карточка ЭТП" in screening["exclusion_reasons"]
    assert "platform_test_lot" in screening["relevance_flags"]


def test_an_apartment_is_not_presented_as_a_development_complex():
    lot = AuctionLot(
        source=source(),
        lot_kind=LotKind.PROPERTY_COMPLEX,
        title="Продажа квартиры площадью 97,4 кв. м",
        address="Москва, ул. Примерная, д. 1, кв. 31",
    )

    screening = AuctionSearchService.screen_lot(lot)

    assert screening["development_relevant"] is False
    assert "residential_unit" in screening["relevance_flags"]
    assert "квартира" in " ".join(screening["exclusion_reasons"])


def test_a_non_residential_building_is_not_mistaken_for_a_flat():
    lot = AuctionLot(
        source=source(),
        lot_kind=LotKind.PROPERTY_COMPLEX,
        title="Нежилое помещение и земельный участок",
    )

    screening = AuctionSearchService.screen_lot(lot)

    assert screening["development_relevant"] is True
    assert "residential_unit" not in screening["relevance_flags"]


def test_api_runtime_flag_controls_project_share_discovery(monkeypatch):
    monkeypatch.delenv("AUCTION_LOTONLINE_PROJECT_SHARES_DISCOVERY", raising=False)
    default_adapter = _discovery_adapters("lot_online")[0]
    assert default_adapter.include_project_shares is False

    monkeypatch.setenv("AUCTION_LOTONLINE_PROJECT_SHARES_DISCOVERY", "true")
    enabled_adapter = _discovery_adapters("lot_online")[0]
    assert enabled_adapter.include_project_shares is True

    monkeypatch.setenv("AUCTION_LOTONLINE_PROJECT_SHARES_DISCOVERY", "false")
    disabled_adapter = next(
        adapter for adapter in _discovery_adapters("all")
        if isinstance(adapter, LotOnlineAdapter)
    )
    assert disabled_adapter.include_project_shares is False


def test_rad_publication_date_is_parsed_in_moscow_time():
    published = LotOnlineAdapter._published_at("Опубликовано: На lot-online.ru: 18.05.2026 17:01")
    assert published is not None
    assert published.date() == date(2026, 5, 18)
    assert published.utcoffset().total_seconds() == 3 * 60 * 60


def test_rad_title_is_scoped_after_publication_metadata():
    text = (
        "Каталог Земельный участок со зданием Опубликовано: "
        "На lot-online.ru: 07.07.2026 16:38 "
        "Продажа доли в размере 100 (сто) % уставного капитала ООО РИВЬЕРА ПАРК "
        "Начальная цена 750 000 000 ₽"
    )
    assert LotOnlineAdapter._extract_title(text) == (
        "Продажа доли в размере 100 (сто) % уставного капитала ООО РИВЬЕРА ПАРК"
    )


def test_rad_title_drops_catalogue_navigation_chrome():
    text = (
        "Опубликовано: На lot-online.ru: 20.08.2026 10:00 "
        "объекты Список сравнения Сохранённые шаблоны поиска Расширенный поиск "
        "Земельный участок пл. 800,0 кв.м. с жилым домом пл. 43,3 кв.м., Москва, Рассудово "
        "Начальная цена 12 000 000 ₽"
    )
    title = LotOnlineAdapter._extract_title(text)
    assert title.startswith("Земельный участок пл. 800,0 кв.м.")
    assert "Список сравнения" not in title
    assert "Расширенный поиск" not in title


def test_rad_land_area_prefers_parcel_over_house_and_normalizes_decimal_comma():
    text = "Земельный участок пл. 800,0 кв.м. с жилым домом площадью 43,3 кв.м."
    area, raw = LotOnlineAdapter._extract_land_area(text)
    assert area == 800
    assert raw == "800,0 кв.м"


def test_small_rad_site_with_residential_house_is_noise_with_auditable_reasons():
    lot = AuctionLot(
        source=source(),
        lot_kind=LotKind.LAND_SALE,
        title="Земельный участок пл. 800,0 кв.м. с жилым домом пл. 43,3 кв.м.",
        address="Москва, Новофёдоровское, Рассудово",
        land_area_sqm=800,
    )
    screening = AuctionSearchService.screen_lot(lot)
    assert screening["development_relevant"] is False
    assert screening["rating"] == "Шум"
    assert "малый участок с жилым домом" in screening["exclusion_reasons"]
    assert screening["platon_explanation"]["grounding"] == "selection_reasons_and_official_lot_fields_only"
    assert lot.selection_reasons[:3] == ["Москва", "продажа земли", "площадь 800 м²"]


def test_history_keeps_relisted_procedures_with_same_cadastre():
    first = AuctionLot(
        source=AuctionSource(SourceKind.LOT_ONLINE, "https://catalog.lot-online.ru/1", "old", "now"),
        lot_kind=LotKind.PROPERTY_COMPLEX,
        title="old",
        cadastral_numbers=["77:05:0003002:54"],
    )
    relist = AuctionLot(
        source=AuctionSource(SourceKind.LOT_ONLINE, "https://catalog.lot-online.ru/2", "new", "now"),
        lot_kind=LotKind.PROPERTY_COMPLEX,
        title="new",
        cadastral_numbers=["77:05:0003002:54"],
    )
    assert AuctionSearchService._deduplicate_history([first, relist]) == [first, relist]


def test_rad_history_filters_cards_by_publication_window(monkeypatch):
    adapter = LotOnlineAdapter()
    captured = {}

    def discover_urls(**kwargs):
        captured.update(kwargs)
        return ["https://catalog.lot-online.ru/index.php?dispatch=products.view&product_id=old"]

    def fetch(url):
        return AuctionLot(
            source=AuctionSource(SourceKind.LOT_ONLINE, url, "old", "now"),
            lot_kind=LotKind.PROPERTY_COMPLEX,
            title="Проект жилой застройки, г. Москва",
            address="г. Москва",
            raw={"region": "Москва", "page_text": "На lot-online.ru: 18.05.2026 17:01"},
        )

    monkeypatch.setattr(adapter, "_discover_candidate_urls", discover_urls)
    monkeypatch.setattr(adapter, "fetch_lot", fetch)
    lots = adapter.discover_moscow_history(date(2026, 2, 21), date(2026, 8, 21))
    assert len(lots) == 1
    assert captured["category_ids"] == ("2", "85")
    assert captured["include_archive"] is True
    assert lots[0].raw["published_at"].startswith("2026-05-18T17:01")


def test_rad_catalogue_extracts_only_official_product_cards():
    links = [
        ("index.php?dispatch=products.view&product_id=1759924", "Коммунарка"),
        ("https://catalog.lot-online.ru/index.php?dispatch=products.view&product_id=1759924", "дубль"),
        ("index.php?dispatch=categories.view&category_id=2", "категория"),
        ("https://example.com/index.php?dispatch=products.view&product_id=1", "чужой сайт"),
    ]
    urls = LotOnlineAdapter._catalog_lot_urls(LotOnlineAdapter._discovery_url(), links)
    assert len(urls) == 1
    assert "product_id=1759924" in urls[0]
    assert urlparse(urls[0]).hostname == "catalog.lot-online.ru"


def test_rad_moscow_confirmation_does_not_confuse_moscow_region():
    moscow = AuctionLot(
        source=source(), lot_kind=LotKind.LAND_SALE, title="Коммунарка", address="г. Москва, Коммунарка"
    )
    region = AuctionLot(
        source=source(), lot_kind=LotKind.LAND_SALE, title="участок", address="Московская область, Одинцово"
    )
    assert LotOnlineAdapter._confirmed_moscow(moscow) is True
    assert LotOnlineAdapter._confirmed_moscow(region) is False


def test_rad_public_offer_schedule_is_structured():
    # Same official table shape as the Kommunarka RAD lot; no network in unit test.
    text = (
        "Время начала периода, начала приема заявок Время окончания приема заявок "
        "Время окончания периода Величина изменения Предложение Сумма задатка Время внесения задатка "
        "22.07.2026 00:00 27.08.2026 14:00 31.08.2026 00:00 "
        "0.00 1 100 250 000.00 165 037 500.00 27.08.2026 14:00 "
        "31.08.2026 00:00 01.09.2026 14:00 05.09.2026 00:00 "
        "79 438 050.00 1 020 811 950.00 153 121 792.50 01.09.2026 14:00"
    )
    periods = LotOnlineAdapter._price_schedule(
        text,
        lot_url="https://catalog.lot-online.ru/test",
        fetched_at="2026-08-21T18:00:00Z",
    )
    assert len(periods) == 2
    assert periods[0].price_rub == 1_100_250_000
    assert periods[0].deposit_rub == 165_037_500
    assert periods[0].application_deadline.startswith("2026-08-27T14:00")
    assert periods[1].price_rub == 1_020_811_950


def test_public_first_document_headers_do_not_require_credentials(monkeypatch):
    monkeypatch.delenv("AUCTION_LOTONLINE_COOKIE", raising=False)
    headers, authenticated = _request_headers("https://catalog.lot-online.ru/file.pdf")
    assert authenticated is False
    assert "Cookie" not in headers


def test_service_account_session_is_read_only_from_runtime_secret(monkeypatch):
    monkeypatch.setenv("AUCTION_ROSELTORG_COOKIE", "session=fake-test-value")
    headers, authenticated = _request_headers("https://www.roseltorg.ru/file.pdf")
    assert authenticated is True
    assert headers["Cookie"] == "session=fake-test-value"


def test_html_login_form_is_detected_as_authorization_boundary():
    html = b'<html><form><input type="password"><button>Login</button></form></html>'
    assert _looks_like_login_page("https://www.roseltorg.ru/login", "text/html", html) is True


def test_krt_auth_required_is_not_mistaken_for_no_obligations(monkeypatch):
    lot = AuctionLot(
        source=source(),
        lot_kind=LotKind.KRT,
        title="КРТ Тест",
        documents=[AuctionDocument(title="Проект договора КРТ.pdf", url="https://catalog.lot-online.ru/doc.pdf")],
    )

    def require_auth(_document):
        raise DocumentAuthorizationRequired("login required")

    monkeypatch.setattr("auction_search.krt_pipeline.extract_document_paragraphs", require_auth)
    enriched = enrich_krt_from_official_documents(lot)
    assert enriched.raw["krt_auth_required"] is True
    assert enriched.raw["krt_extraction_complete"] is False
    assert enriched.documents[0].auth_required is True
    assert enriched.documents[0].access_status == "auth_required"


def test_krt_program_is_separate_from_investor_obligation():
    text = ["Предельная площадь жилой застройки составляет 180 000 кв. м."]
    program = extract_krt_program(
        text,
        source_url="https://official-etp.example/lot/1",
        source_document="Решение КРТ.pdf",
        fetched_at="2026-08-21T18:00:00Z",
    )
    obligations = extract_krt_obligations(
        text,
        source_url="https://official-etp.example/lot/1",
        source_document="Решение КРТ.pdf",
        fetched_at="2026-08-21T18:00:00Z",
    )
    assert program
    assert program[0].category == "housing"
    assert program[0].area_sqm == 180_000
    assert obligations == []


def test_krt_obligation_keeps_source_provenance():
    text = ["Инвестор обязан построить общеобразовательную школу на 775 мест и передать объект городу Москве безвозмездно."]
    items = extract_krt_obligations(
        text,
        source_url="https://official-etp.example/lot/1",
        source_document="Проект договора КРТ.pdf",
        fetched_at="2026-08-21T18:00:00Z",
    )
    assert items
    assert items[0].quantity == 775
    assert items[0].provenance.source_document == "Проект договора КРТ.pdf"
    assert items[0].estimated_cost_rub is None
    assert items[0].transfer_free_of_charge is True
    assert items[0].recipient == "город Москва"


def test_krt_mapper_does_not_replace_terms_with_glavapu():
    lot = AuctionLot(source=source(), lot_kind=LotKind.KRT, title="КРТ Тест")
    seed = build_developaid_seed(lot)
    assert seed["krt"]["tep_source_policy"] == "official_krt_documents"
    assert seed["krt"]["glavapu_role"] == "validation_only"


def test_ordinary_land_does_not_autoinsert_vri_payment():
    lot = AuctionLot(source=source(), lot_kind=LotKind.LAND_SALE, title="Обычная продажа")
    seed = build_developaid_seed(lot)
    assert seed["ordinary_land"]["vri_change_payment_rub"] is None


def test_ordinary_auction_preset_prefills_purchase_price_and_cadastre():
    lot = AuctionLot(
        source=source(),
        lot_kind=LotKind.LAND_SALE,
        title="Коммунарка",
        address="Москва, Коммунарка",
        cadastral_numbers=["50:21:0120316:1221"],
        land_area_sqm=173_494,
        current_price_rub=1_100_250_000,
        permitted_use="Многофункциональные общественные центры",
    )
    preset = build_project_preset(lot)
    assert preset["schema_version"] == "developaid.project_preset.v4"
    assert preset["auction_import"]["filled_inputs"]["purchase_price_mln"] == 1100.25
    assert preset["project"]["cadastral_numbers"] == ["50:21:0120316:1221"]
    assert preset["planning"]["objects"] == []
    assert preset["open_items"]


def test_handoff_excludes_building_cadastre_without_changing_lot_data():
    lot = AuctionLot(
        source=source(), lot_kind=LotKind.PROPERTY_COMPLEX, title="ОСЗ с участком",
        cadastral_numbers=["77:02:0019005:1000", "77:02:0019005:4"],
        current_price_rub=100_000_000,
    )
    preset = build_project_preset(lot)
    context = {
        "buildings": [{"cadastral_number": "77:02:0019005:1000"}],
        "land_parcels": [{"cadastral_number": "77:02:0019005:4"}],
    }
    selected = _handoff_land_cadastres(preset, context)
    assert selected == ["77:02:0019005:4"]
    assert preset["project"]["cadastral_numbers"] == ["77:02:0019005:4"]
    assert preset["land"]["cadastral_numbers"] == ["77:02:0019005:4"]
    assert lot.cadastral_numbers == ["77:02:0019005:1000", "77:02:0019005:4"]


def test_krt_preset_maps_only_unambiguous_program_products():
    prov = Provenance(
        source_url="https://catalog.lot-online.ru/doc.pdf",
        source_document="Решение КРТ.pdf",
        raw_value="test",
    )
    lot = AuctionLot(
        source=source(),
        lot_kind=LotKind.KRT,
        title="КРТ Тест",
        current_price_rub=900_000_000,
        krt_program=[
            KrtProgramItem(category="housing", title="Жилая застройка 180 000 м²", area_sqm=180_000, provenance=prov),
            KrtProgramItem(category="office", title="Офисы 40 000 м²", area_sqm=40_000, provenance=prov),
            KrtProgramItem(category="public_business", title="МФК 25 000 м²", area_sqm=25_000, provenance=prov),
        ],
    )
    preset = build_project_preset(lot)
    objects = preset["planning"]["objects"]
    assert len(objects) == 2
    assert objects[0]["residential_part_m2"] == 180_000
    assert objects[1]["building_type"] == "office_from_krt"
    assert any("public_business" in item for item in preset["open_items"])
    assert preset["import_rules"]["do_not_replace_krt_terms_with_glavapu"] is True


def test_auction_list_page_gets_handoff_script_once():
    page = "<html><body>auction</body></html>"
    bridged = auction_page_with_handoff(page)
    assert "developaid-auction-list-handoff-v1" in bridged
    assert auction_page_with_handoff(bridged) == bridged


def test_auction_ui_does_not_render_missing_area_as_zero():
    from auction_search.ui import auctions_page

    page = auctions_page()
    assert "n!==null&&n!==undefined&&n!==''" in page


def test_excel_export_separates_land_and_building_as_numbers():
    from openpyxl import load_workbook

    raw = _xlsx([{
        "section": "Торги", "name": "ОСЗ с участком", "cadastre": "77:1, 77:2",
        "type": "ЗИК", "land_area_sqm": 3650, "building_area_sqm": 785,
        "price": 158_636_700, "score": 24, "url": "https://example.test/lot/1",
    }])
    sheet = load_workbook(io.BytesIO(raw)).active
    headers = [cell.value for cell in sheet[1]]
    assert "Площадь участка, м²" in headers
    assert "Площадь здания/ОКС, м²" in headers
    assert sheet.cell(2, headers.index("Площадь участка, м²") + 1).value == 3650
    assert sheet.cell(2, headers.index("Площадь здания/ОКС, м²") + 1).value == 785
    assert sheet.cell(2, headers.index("Цена, ₽") + 1).data_type == "n"
    assert sheet.freeze_panes == "A2"


def test_krt_excel_export_keeps_territory_and_program_areas_separate():
    from openpyxl import load_workbook

    raw = _xlsx([{
        "section": "КРТ", "name": "Тестовая территория", "okrug": "ЗАО",
        "district": "Кунцево", "krt_area_ha": 23.5, "total_gfa_sqm": 74470,
        "housing_gfa_sqm": 61000, "score": 87, "status": "Планируемый",
    }])
    sheet = load_workbook(io.BytesIO(raw)).active
    headers = [cell.value for cell in sheet[1]]
    assert sheet.cell(2, headers.index("Площадь КРТ, га") + 1).value == 23.5
    assert sheet.cell(2, headers.index("Общий объём, м²") + 1).value == 74470
    assert sheet.cell(2, headers.index("Жильё, м²") + 1).value == 61000


def test_auction_ui_names_city_discovery_and_shows_source_funnel():
    from auction_search.ui import auctions_page

    page = auctions_page()
    assert '<option value="investmoscow">Торги Москвы → ЭТП</option>' in page
    assert "без подтверждённой ЭТП" in page
    assert "нужны адаптеры" in page


def test_model_page_bridge_is_idempotent():
    class Core:
        PAGE = "<html><body>model</body></html>"

    core = Core()
    assert install_page_bridge(core) is True
    assert "developaid-auction-preset-bridge-v1" in core.PAGE
    first = core.PAGE
    assert install_page_bridge(core) is False
    assert core.PAGE == first
