"""Убыток одной очереди уменьшает базу остальных — и книга считает так же.

Очереди — один проект и один налогоплательщик (решение владельца, 24.08.2026),
поэтому убыток убыточной очереди зачитывается в базе прибыльных с половинным
ограничением ст. 283 НК.

Тест написан задним числом, и это его главное содержание. Правка уехала в
движок, полный набор остался ЗЕЛЁНЫМ, а книга всё это время считала налог
по-старому: на проекте с убыточной первой очередью она давала 4 177,6 млн
против 844,9 у движка — разрыв в 3 332 млн, которого не видел никто. Ни один
тест не доходил до убыточной очереди: во всех сценариях набора обе очереди
прибыльны, а там зачитывать нечего.

Ровно та ловушка, что записана в правилах проекта: зелёный набор не значит,
что ветка кода верна, — возможно, до неё просто не доходят. Поэтому здесь
проверяется не «налог считается», а «налог считается ТАМ, где включается
зачёт»: первая очередь обязана быть убыточной, иначе тест проверяет не то,
ради чего написан.

Запуск: python3 -m pytest tests/test_a_losing_queue_shares_its_loss.py -q
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(Path(__file__).resolve().parent))

import main_legacy as core  # noqa: E402
from xlsx_eval import Evaluator  # noqa: E402


def _scenario() -> tuple[dict, dict, dict]:
    """Дорогой вход загоняет первую очередь в минус — там и живёт зачёт."""
    inputs = dict(core.DEFAULT_INPUTS)
    inputs.update(apartment_price_th=650, commercial_price_th=650,
                  parking_price_th=5000, purchase_price_mln=26000)
    tep = {key: dict(value) for key, value in core.TEP_DEFAULT.items()}
    phasing = {
        "enabled": True, "phase_count": 2, "phase_gap_months": 12,
        "phases": [
            {"name": "О1", "start_offset_months": 0, "construction_months": 24},
            {"name": "О2", "start_offset_months": 12, "construction_months": 24},
        ],
        "social_objects": [],
        "discrete": {"offices": 2, "standalone_retail": 2, "above_parking": 2},
    }
    return inputs, tep, phasing


@pytest.fixture(scope="module")
def phased():
    inputs, tep, phasing = _scenario()
    return core.calculate_phased(core.PhasedCalcRequest(
        inputs=inputs, tep=tep, phasing=phasing))


def test_the_first_queue_really_is_at_a_loss(phased) -> None:
    """Предохранитель самого теста.

    Подкрути кто-нибудь вводные так, что обе очереди станут прибыльными, —
    и тест продолжит зеленеть, ничего больше не проверяя. Так и вышло с
    остальным набором.
    """
    first = phased["comparison"][0]
    assert first["net_profit"] < 0, "первая очередь должна быть убыточной"
    assert first.get("profit_tax", 0) == 0, "убыточная очередь налога не платит"


def test_the_loss_of_one_queue_lowers_the_tax_of_the_others(phased) -> None:
    """Убыток не выбрасывается: без зачёта налог был бы кратно больше."""
    consolidated = phased["consolidated"]["summary"]["profit_tax"]
    alone = sum(max(0.0, float(row.get("profit_tax") or 0.0))
                for row in phased["comparison"])
    assert consolidated == pytest.approx(alone, rel=1e-6), \
        "сумма очередей обязана равняться своду — иначе таблица не сходится"
    # Без зачёта прибыльная очередь заплатила бы со всей своей базы.
    second = phased["comparison"][1]
    assert consolidated < abs(second["net_profit"]) * 0.25


def test_the_workbook_agrees_with_the_engine_on_a_losing_queue() -> None:
    """Та самая сверка, которой не было.

    До правки книга платила 4 177,6 против 844,9 — 3 332 млн разницы при
    зелёном наборе.
    """
    sys.setrecursionlimit(400000)
    inputs, tep, phasing = _scenario()
    engine = core.calculate_phased(core.PhasedCalcRequest(
        inputs=inputs, tep=tep, phasing=phasing))
    content, _, meta = core.build_project_workbook(
        inputs, tep, [], phasing, project_name="Убыточная очередь")
    assert meta["missing"] == []
    evaluator = Evaluator(openpyxl.load_workbook(io.BytesIO(content)))
    book_tax = evaluator.cell("КОНСОЛИДАТОР", "K8")
    engine_tax = engine["consolidated"]["summary"]["profit_tax"] / 1e6
    assert book_tax == pytest.approx(engine_tax, rel=0.005), (
        f"книга {book_tax:,.1f} против движка {engine_tax:,.1f}")


def test_the_queue_rows_add_up_to_the_consolidated_tax() -> None:
    """Строка очереди и итог — одно и то же число, разложенное по долям."""
    sys.setrecursionlimit(400000)
    inputs, tep, phasing = _scenario()
    content, _, _ = core.build_project_workbook(
        inputs, tep, [], phasing, project_name="Доли")
    evaluator = Evaluator(openpyxl.load_workbook(io.BytesIO(content)))
    rows = sum(evaluator.cell("КОНСОЛИДАТОР", f"K{n}") for n in range(4, 8))
    assert rows == pytest.approx(evaluator.cell("КОНСОЛИДАТОР", "K8"), rel=1e-6)


def test_the_half_limit_is_declared_once() -> None:
    """Ограничение ст. 283 объявлено в движке одним числом.

    Вторая копия «половины» в книге разошлась бы с первой при первой же
    правке — так уже расходились ставка ПФ и профиль управления.
    """
    assert core._LOSS_CARRY_USE_LIMIT == 0.5
