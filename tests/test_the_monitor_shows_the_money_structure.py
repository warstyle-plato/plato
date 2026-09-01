"""РСС, утверждённый бюджет, дефицит общий и структурный — раздельно.

«Непонятно, почему ДДС-остаток 1,6 при дефиците потока 3,6… Нет структуры
интуитивно понятно — РСС такой, утверждённый бюджет такой-то. Остаток лимитов на
завершение по РСС такой-то, по утверждённому бюджету — такой-то. Дефицит такой-то
общий и такой-то структурный… потребность в доп финансировании нужно выделить по
сумме и сроку» (владелец, 29.08.2026).

Структурный дефицит — это когда общей суммы по РСС хватает, а лежит она на
статьях, куда потребность не относится: свободный лимит одной статьи не
закрывает дефицит другой. Движок это уже считал (`additional_financing`), но на
экране два разных числа из двух разных контуров стояли порознь без подписей и
читались как одно.

И гарантийные удержания: «в РСС банка сумма по договору берётся общая, но ГУ до
момента погашения ПФ не заплатятся — по сути это скрытый резерв».

Запуск: python3 -m pytest tests/test_the_monitor_shows_the_money_structure.py -q
"""

from __future__ import annotations

import datetime
import io
import sys
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import developaid_monitor_retention as retention  # noqa: E402

PAGE = (ROOT / "developaid_monitor_page.py").read_text(encoding="utf-8")
DASH = (ROOT / "developaid_monitor_dashboard.py").read_text(encoding="utf-8")


def test_the_structure_names_both_contours_and_both_deficits() -> None:
    body = PAGE[PAGE.index("function fundingStructure("):]
    body = body[: body.index("\n}\n")]
    for line in ("Общая сметная стоимость глав 2–3", "Остаток лимитов на завершение",
                 "Утверждённый бюджет глав 2–3",
                 "Остаток потребности по утверждённому бюджету",
                 "Надо достроить по утверждённой модели", "Есть: остаток лимитов + резерв",
                 "row('ДЕФИЦИТ'", "Структурный дефицит внутри лимитов"):
        assert line in body, f"в структуре нет строки «{line}»"
    # Срок — половина ответа: «нужно 2,2 млрд» и «нужно с марта» — разные новости.
    assert "additional_financing_from" in body
    assert "Первый месяц нехватки" in body


def test_the_need_carries_the_month_it_starts() -> None:
    assert '"additional_financing_from": unfunded_from' in DASH
    assert '"additional_financing_from": waterfall["additional_financing_from"]' in DASH


def test_a_bigger_structural_deficit_is_explained_not_just_shown() -> None:
    """Денег хватает, а лежат не там — это повод к перераспределению."""
    body = PAGE[PAGE.index("function fundingStructure("):]
    assert "Структурный дефицит больше общего" in body


def _register(rows: list[dict]) -> bytes:
    from openpyxl import Workbook

    book = Workbook()
    sheet = book.active
    sheet.title = "Незачтенный аванс "
    head = ["№ п/п", "Заказчик", "Контрагент", "Реквизиты договора",
            "Стоимость Договора ", "Размер ГУ, %", "Сумма ГУ",
            "Выплачено ГУ по состточнию на", "Остаток к выплате",
            "Дата выплаты ГУ", "Примечание", "Дата окончания работ"]
    for column, title in enumerate(head, start=1):
        sheet.cell(row=3, column=column, value=title)
    for index, row in enumerate(rows, start=4):
        for column, key in enumerate(
                ["n", "customer", "counterparty", "contract", "contract_amount",
                 "share", "amount", "paid", "left", "due", "note", "works_end"], start=1):
            sheet.cell(row=index, column=column, value=row.get(key))
    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def test_the_total_row_of_the_register_is_not_added_to_its_own_rows() -> None:
    """Итог, посчитанный вместе со строками, удваивает резерв."""
    data = _register([
        {"contract_amount": 100.0, "amount": 5.0, "paid": 0.0, "left": 5.0},
        {"counterparty": "Базис ООО", "contract": "382-ГРД", "contract_amount": 60.0,
         "share": 5, "amount": 3.0, "paid": 3.0, "left": 0.0, "due": "16.02.2026"},
        {"counterparty": "НУР ООО", "contract": "386-ГРД", "contract_amount": 40.0,
         "share": 5, "amount": 2.0, "paid": 0.0, "left": 2.0, "due": "30.09.2032"},
    ])
    got = retention.read_retention(data)
    assert len(got["rows"]) == 2, "итоговая строка попала в реестр"
    assert got["file_total"]["amount"] == 5.0

    summary = retention.summary(got, horizon="2027-12-31")
    assert summary["amount"] == 5.0 and summary["left"] == 2.0
    assert summary["contract_amount"] == 100.0
    # За горизонтом стройки — тот самый скрытый резерв.
    assert summary["deferred_after_horizon"] == 2.0


def test_a_second_payout_stage_does_not_double_the_contract() -> None:
    """Строка без контрагента, но со своим сроком, — продолжение договора."""
    data = _register([
        {"counterparty": "КЛОДО ООО", "contract": "412-ГРД", "contract_amount": 48.0,
         "share": 3, "amount": 0.7, "paid": 0.0, "left": 0.7, "due": "30.09.2027"},
        {"contract_amount": 48.0, "amount": 0.7, "paid": 0.0, "left": 0.7,
         "due": "30.09.2032"},
    ])
    got = retention.read_retention(data)
    assert len(got["rows"]) == 2
    assert got["rows"][1]["continued"] is True
    summary = retention.summary(got, horizon="2027-12-31")
    assert summary["contract_amount"] == 48.0, "стоимость договора сложена дважды"
    assert round(summary["left"], 2) == 1.4


def test_without_a_horizon_the_deferred_part_is_not_invented() -> None:
    """Без горизонта вопрос «что за ним» ответа не имеет — и это не ноль."""
    data = _register([
        {"counterparty": "НУР ООО", "contract": "386", "contract_amount": 40.0,
         "share": 5, "amount": 2.0, "paid": 0.0, "left": 2.0, "due": "30.09.2032"},
    ])
    summary = retention.summary(retention.read_retention(data))
    assert summary["deferred_after_horizon"] is None


def test_a_register_without_its_head_is_refused_by_name() -> None:
    from openpyxl import Workbook

    book = Workbook()
    book.active["A1"] = "просто таблица"
    buffer = io.BytesIO()
    book.save(buffer)
    with pytest.raises(retention.RetentionUnreadable) as failed:
        retention.read_retention(buffer.getvalue())
    assert "Сумма ГУ" in str(failed.value)


def test_our_sum_is_checked_against_the_file_total() -> None:
    """Строка, выпавшая из формулы реестра, иначе не видна никому."""
    # Числа настоящего порядка: допуск расхождения — рубль, и на копейках
    # проверять его бессмысленно.
    data = _register([
        {"contract_amount": 90_000_000.0, "amount": 4_000_000.0,
         "paid": 0.0, "left": 4_000_000.0},
        {"counterparty": "А", "contract": "1", "contract_amount": 50_000_000.0,
         "share": 5, "amount": 2_500_000.0, "paid": 0.0, "left": 2_500_000.0,
         "due": "30.09.2032"},
        {"counterparty": "Б", "contract": "2", "contract_amount": 40_000_000.0,
         "share": 5, "amount": 2_000_000.0, "paid": 0.0, "left": 2_000_000.0,
         "due": "30.09.2032"},
    ])
    summary = retention.summary(retention.read_retention(data), horizon="2027-12-31")
    bad = summary["file_total_mismatch"]
    assert "amount" in bad and round(bad["amount"]["delta"], 2) == 500_000.0
    assert "contract_amount" not in bad, "стоимость договоров сошлась — молчим"


def test_the_hidden_reserve_reaches_the_screen_and_the_store() -> None:
    import developaid_monitor as monitor

    assert hasattr(monitor, "store_retention") and hasattr(monitor, "latest_retention")
    assert '"retention": _retention(project, rnv)' in DASH
    body = PAGE[PAGE.index("function fundingStructure("):]
    assert "справочно, дефицит они не уменьшают" in body
    assert "Из них после ввода" in body
    assert "в стройке эти деньги не будут потрачены" in body
    # Реестра нет — строки нет вовсе: «не загружали» и «удержаний нет» разное.
    assert "const gu=f.retention;" in body
    # И маршрут, которым он приезжает.
    import main_legacy

    assert "/monitor/retention" in {getattr(route, "path", "") for route in main_legacy.app.routes}


def test_a_stored_register_is_read_back(tmp_path, monkeypatch) -> None:
    import developaid_monitor as monitor

    # Каталог снимков читается один раз при импорте: `DATA_DIR` в окружении
    # теста уже опоздал, и проверка писала в рабочие данные репозитория — они
    # и уехали в коммит. Подменяется сам каталог.
    monkeypatch.setattr(monitor, "_SNAPSHOT_DIR", tmp_path / "monitor")
    data = _register([
        {"counterparty": "НУР ООО", "contract": "386", "contract_amount": 40.0,
         "share": 5, "amount": 2.0, "paid": 0.0, "left": 2.0, "due": "30.09.2032"},
    ])
    stored = monitor.store_retention("Проверочный", data, "2026-08-29", "гу.xlsx")
    assert stored["contracts"] == 1
    back = monitor.latest_retention("Проверочный")
    assert back and back["taken_at"] == "2026-08-29"
    assert retention.summary(back, horizon=datetime.date(2027, 12, 31))["left"] == 2.0


def test_the_upload_stands_with_the_other_uploads() -> None:
    """Реестр ГУ — такой же источник проекта, как РСС и продажи.

    Поле уехало в верхнюю панель, к «Проекту» и «Срезу»: вставка искала
    ближайший `<div class="field">` перед кнопкой продаж, а в блоке загрузок
    таких нет вовсе — разметка там своя. «Почему загрузку ГУ вынесли за пределы
    всех загрузок?» (владелец, 30.08.2026) — верный вопрос: источник, стоящий
    в стороне от источников, читается как отдельная кнопка неизвестно чего.
    """
    page = PAGE
    weekly = page[page.index("<b>Еженедельно</b>"):]
    weekly = weekly[: weekly.index('<div class="msg"')]
    assert 'id="retention"' in weekly, "загрузка ГУ стоит не с остальными загрузками"
    assert 'id="retentionBtn"' in weekly
    assert "Реестр гарантийных удержаний" in weekly, "и блок называет этот источник"
    # А в верхней панели её нет: там управление срезом, а не файлы.
    controls = page[page.index('<div class="controls">'):]
    controls = controls[: controls.index("</div></div>")]
    assert "retention" not in controls


def test_the_tests_do_not_write_into_the_repository() -> None:
    """Проверка, оставляющая файлы в `data/`, уезжает с ними в коммит.

    Так и вышло: снимок реестра ГУ проверочного проекта дважды попал в
    репозиторий. Каталог снимков читается при импорте, и `DATA_DIR`,
    выставленный в тесте, до него уже не доходит.
    """
    import developaid_monitor as monitor

    left = sorted(path.name for path in (ROOT / "data" / "monitor").glob("*"))
    assert "Проверочный" not in left, "проверка снова пишет в рабочие данные"
    assert monitor._SNAPSHOT_DIR.name == "monitor"


def test_the_need_comes_from_the_model_and_the_bank_is_the_source() -> None:
    """«РСС — это то, что даёт банк. Утверждённая модель — сколько надо реально,
    чтобы построить» (владелец, 30.08.2026).

    Значит потребность берётся из модели, лимиты банка с резервом — источник, а
    главный дефицит есть разница. Прежде потребностью считалась банковская
    колонка «Средства на завершение», и дефицит выходил вчетверо меньше
    настоящего: 0,2 млрд вместо 2,2.
    """
    import developaid_monitor_dashboard as dash

    funding = {"known": True, "remaining_need": 1.66e9, "bank_remaining": 1.15e9,
               "reserve": 306.1e6, "approved_remaining": 3.66e9,
               "monthly_unfunded": {}, "additional_financing": 0.2e9}
    said = dash._summary({"dashboard": {}}, funding, {})[0]
    assert "По утверждённой модели достроить стоит 3,66 млрд ₽" in said
    assert "По РСС осталось 1,46 млрд ₽" in said
    assert "оставшийся лимит и есть то, что банк готов дать" in said
    assert "бюджет всей стройки, а не банковская доля" in said
    assert "Дефицит 2,20 млрд ₽" in said
    # Банковский остаток остаётся справочным и назван взглядом банка.
    assert "его взгляд по РСС, а не потребность стройки" in said
    # Источник назван своим именем, а не чужим.
    assert "ДДС" not in said

    # Без модели потребность не выдумывается.
    blind = dash._summary({"dashboard": {}}, {**funding, "approved_remaining": 0}, {})[0]
    assert "Утверждённая модель не прочитана" in blind


def test_the_screen_puts_the_model_against_the_bank() -> None:
    """На экране те же роли: модель — потребность, банк — источник, разница —
    дефицит; банковский остаток стоит справочной строкой."""
    body = PAGE[PAGE.index("function fundingStructure("):]
    body = body[: body.index("\n}\n")]
    assert "Надо достроить по утверждённой модели" in body
    assert "Есть: остаток лимитов + резерв" in body
    assert "row('ДЕФИЦИТ'" in body
    assert "mainGap=modelNeed==null?null:Math.max(0,modelNeed-fuel)" in body
    assert "«Средства на завершение» по РСС" in body
    assert "взгляд банка на остаток, не потребность стройки" in body
    # Структурный дефицит остаётся, но он про статьи внутри лимитов.
    assert "Структурный дефицит внутри лимитов" in body
    # Без книги модель не выдумывается.
    assert "сколько реально надо, сказать нечем" in body


def test_the_structural_deficit_says_how_it_is_counted_and_from_what() -> None:
    """«Структурный дефицит — непонятно откуда 1,44, как это посчитано? И он
    считается от РСС или утверждённой модели?» (владелец, 01.09.2026)."""
    body = PAGE[PAGE.index("function fundingStructure("):]
    body = body[: body.index("\n}\n")]
    assert "Помесячно по программе РСС" in body
    assert "утверждённая модель здесь не участвует" in body
    assert "Нехватку гасит резерв 2.8/2.9" in body
    # Строка «разница двух остатков» снята: два почти одинаковых числа под
    # разными именами читались как бред, а не как оговорка.
    assert "Разница двух остатков потребности" not in body


def test_the_bank_column_is_reconciled_not_just_shown() -> None:
    """«Почему там 1,66, если остаток лимитов и резервов 1,46??? Откуда ещё 200
    млн взялось?» — разница раскладывается по главам, резерву и статьям с
    перерасходом, а несведённое названо остатком."""
    import developaid_monitor_dashboard as dash

    waterfall = {
        "opening_bank_remaining": 1.40e9, "opening_article_deficit": 0.05e9,
        "articles": [
            {"code": "2.1", "chapter": "2", "has_programme": True, "opening_limit_raw": 0.9e9},
            {"code": "2.2", "chapter": "2", "has_programme": False, "opening_limit_raw": 0.2e9},
            {"code": "2.3", "chapter": "2", "has_programme": True, "opening_limit_raw": -0.05e9},
            {"code": "3.1", "chapter": "3", "has_programme": False, "opening_limit_raw": 0.3e9},
        ],
    }
    check = dash._bank_need_check(1.66e9, 0.06e9, waterfall)
    assert check["bank_column"] == pytest.approx(1.66e9)
    assert check["ours"] == pytest.approx(1.46e9)
    assert check["by_chapter"] == {"2": pytest.approx(1.1e9), "3": pytest.approx(0.3e9)}
    assert check["no_programme"] == pytest.approx(0.5e9)
    assert check["overpaid_clipped"] == pytest.approx(0.05e9)
    # 1,66 − (1,40 − 0,05 + 0,06): то, чего мы объяснить не можем, названо.
    assert check["residual"] == pytest.approx(0.25e9)

    body = PAGE[PAGE.index("function fundingStructure("):]
    body = body[: body.index("\n}\n")]
    assert "Сверка с колонкой банка" in body
    for key in ("check.bank_column", "check.ours", "check.by_chapter",
                "check.no_programme", "check.overpaid_clipped", "check.residual"):
        assert key in body, f"экран не показывает {key}"
    assert "Не сошлось с колонкой банка" in body
    assert "С колонкой банка сходится" in body


def test_the_limits_cover_both_chapters_like_the_bank_column(tmp_path) -> None:
    """«Это точно 2 глава? Или в разделе РСС 2, а в дефиците 2 и 3?»

    Было именно так: статьи водопада брались только из главы 2 и только с
    программой, а «Средства на завершение», модель и оплаченное — по итоговой
    строке глав 2–3. «Есть: остаток лимитов + резерв» выходило на 200 млн ₽
    меньше колонки банка, и разница была главой 3 и статьями без программы.
    """
    from openpyxl import Workbook

    import developaid_monitor_dashboard as dash

    book = Workbook()
    ws = book.active
    ws.title = "Расчет стоимости строительства"
    for c, v in {1: "Код", 4: "Общая сметная стоимость", 7: "Утвержденная фин.модель проекта"}.items():
        ws.cell(row=9, column=c, value=v)
    for c, v in {9: "Оплачено по состояни. На 17.07.2026",
                 11: "Средства на завершение согласно бюджету",
                 13: "производстввенная программа"}.items():
        ws.cell(row=8, column=c, value=v)
    ws.cell(row=9, column=13, value="Август")
    ws.cell(row=9, column=14, value="Сентябрь")
    rows = [
        ("2", "Глава 2", 700.0, 100.0, None, None),
        ("2.1", "С программой", 300.0, 50.0, 40.0, 60.0),
        ("2.2", "Без программы", 300.0, 50.0, None, None),
        ("2.8", "Резерв", 100.0, 0.0, None, None),
        ("3", "Глава 3", 200.0, 20.0, None, None),
        ("3.1", "Прочие затраты", 200.0, 20.0, None, None),
    ]
    for offset, (code, name, limit, paid, aug, sep) in enumerate(rows, start=10):
        ws.cell(row=offset, column=1, value=code)
        ws.cell(row=offset, column=2, value=name)
        ws.cell(row=offset, column=4, value=limit)
        ws.cell(row=offset, column=9, value=paid)
        ws.cell(row=offset, column=11, value=limit - paid)
        if aug is not None:
            ws.cell(row=offset, column=13, value=aug)
            ws.cell(row=offset, column=14, value=sep)
    total = 10 + len(rows)
    ws.cell(row=total, column=2, value="Всего инвестиционные расходы глава 2, 3")
    ws.cell(row=total, column=7, value=1200.0)
    ws.cell(row=total, column=9, value=120.0)
    ws.cell(row=total, column=11, value=780.0)
    path = tmp_path / "finance.xlsx"
    book.save(path)

    baseline = dash._read_finance_baseline(path)
    assert baseline["known"], baseline
    articles = baseline["articles"]
    assert set(articles) == {"2.1", "2.2", "3.1"}, "статьи не по главам 2–3"
    assert articles["2.2"]["has_programme"] is False
    assert articles["3.1"]["chapter"] == "3"
    assert articles["2.1"]["has_programme"] is True
    # Резерв — не статья, а главы не считаются вместе со своими статьями.
    assert baseline["reserve"] == pytest.approx(100.0)
    waterfall = dash._article_waterfall(articles, baseline["reserve"], datetime.date(2026, 8, 20))
    # Остаток лимитов теперь тот же контур, что колонка банка: 250 + 250 + 180.
    assert waterfall["opening_bank_remaining"] == pytest.approx(680.0)
    check = dash._bank_need_check(baseline["completion_need_at_baseline"], baseline["reserve"], waterfall)
    assert check["ours"] == pytest.approx(780.0)
    assert check["residual"] == pytest.approx(0.0, abs=1e-6)




def test_the_estimate_column_is_not_called_a_credit_limit() -> None:
    """Суммы ПФ в РСС нет — она из НКЛ (владелец, 30.08.2026).

    Строка «Лимит РСС» показывала колонку «Общая сметная стоимость»
    (`_ESTIMATE_COLUMNS["estimate"] = 4`, сумма по главам 2 и 3), то есть
    сметную стоимость всей стройки под именем кредитной линии. Та же
    ошибка, что подпись «Даёт банк» под остатком РСС: число посчитано
    верно и прочитано неверно, потому что названо чужим именем. При этом
    оставшийся лимит — это ровно то, что банк готов дать, и так и сказано.
    """
    body = PAGE[PAGE.index("function fundingStructure("):]
    body = body[: body.index("\n}\n")]
    assert "Общая сметная стоимость глав 2–3" in body
    assert "Лимит РСС" not in body, "сметная стоимость названа кредитным лимитом"
    assert "колонка РСС, а не лимит кредитной линии" in body
    assert "суммы ПФ в нём нет, она из НКЛ" in body
    assert "это и есть то, что банк готов дать" in body
