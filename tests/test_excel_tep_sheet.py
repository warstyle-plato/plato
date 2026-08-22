"""ТЭП полной Excel-книги читается и правится в отдельном безопасном листе."""

from __future__ import annotations

import io

import openpyxl

import main as wrapper
from excel_tep_sheet import SHEET_NAME, add_tep_project_sheet, find_technical_tep_table


core = wrapper.core


def _book_with_storage(units: float = 123):
    inputs = dict(core.DEFAULT_INPUTS)
    tep = {key: dict(value) for key, value in core.TEP_DEFAULT.items()}
    tep["storage"]["units"] = units
    raw, _meta = core.build_plato_model_v2(inputs, tep, None, "ТЭП")
    before = openpyxl.load_workbook(io.BytesIO(raw), data_only=False)
    data = add_tep_project_sheet(raw, tep, core)
    after = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    return tep, before, after


def test_a_separate_tep_sheet_does_not_move_existing_rows():
    """Новый интерфейс не вставляет строки в старые листы."""
    _tep, before, after = _book_with_storage()
    assert SHEET_NAME in after.sheetnames
    for name in before.sheetnames:
        assert name in after.sheetnames
        assert after[name].max_row == before[name].max_row, name
        assert after[name].max_column == before[name].max_column, name


def test_the_tep_sheet_is_the_product_table_the_user_expects():
    tep, _before, book = _book_with_storage(137)
    ws = book[SHEET_NAME]
    headers = [ws.cell(row=4, column=col).value for col in range(1, 11)]
    assert headers[:7] == [
        "Продукт",
        "ГНС, м²",
        "Общая площадь, м²",
        "Полезная площадь, м²",
        "Продаваемая площадь, м²",
        "Передаваемая площадь, м²",
        "Количество, шт.",
    ]
    labels = {ws.cell(row=row, column=1).value: row for row in range(5, ws.max_row + 1)}
    storage_row = labels[core.TEP_DEFAULT["storage"]["label"]]
    assert ws.cell(row=storage_row, column=7).value == 137
    assert "цена кладовой" in str(ws.cell(row=storage_row, column=10).value).lower()

    apartment_row = labels[core.TEP_DEFAULT["apartments"]["label"]]
    assert ws.cell(row=apartment_row, column=8).value == f"=IFERROR(C{apartment_row}/B{apartment_row},0)"
    assert ws.cell(row=apartment_row, column=9).value == f"=IFERROR(E{apartment_row}/C{apartment_row},0)"
    assert tep["apartments"]["gns"] > 0


def test_the_old_technical_table_points_to_the_new_inputs():
    """Книга продолжает считать через прежние адреса, но источником стал новый лист."""
    tep, before, after = _book_with_storage(211)
    labels = {key: str((tep.get(key) or {}).get("label") or core.TEP_DEFAULT[key]["label"])
              for key in tep}
    old_ws, old_rows = find_technical_tep_table(before, labels)
    new_ws, new_rows = find_technical_tep_table(after, labels)
    assert old_ws is not None, "в исходной книге не найдена рабочая таблица ТЭП"
    assert new_ws is not None and new_ws.title == old_ws.title
    assert new_rows == old_rows, "строки рабочей таблицы ТЭП не должны сдвигаться"

    storage_new_row = next(
        row for row in range(5, after[SHEET_NAME].max_row + 1)
        if after[SHEET_NAME].cell(row=row, column=1).value == core.TEP_DEFAULT["storage"]["label"]
    )
    assert new_ws.cell(row=new_rows["storage"], column=7).value == (
        f"='{SHEET_NAME}'!$G${storage_new_row}"
    )

    apartment_new_row = next(
        row for row in range(5, after[SHEET_NAME].max_row + 1)
        if after[SHEET_NAME].cell(row=row, column=1).value == core.TEP_DEFAULT["apartments"]["label"]
    )
    assert new_ws.cell(row=new_rows["apartments"], column=5).value == (
        f"='{SHEET_NAME}'!$E${apartment_new_row}"
    )
