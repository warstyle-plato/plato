"""Финансовая модель DevelopAid считает то же, что и движок, — своими формулами.

Шаблон ПЛАТО сводили с расчётом восемью правками, и каждая проверялась чужим
пересчётом: посчитать книгу здесь нечем. Так жить нельзя — правка формулы
уезжала к аналитику непроверенной.

Эта книга собирается с нуля: тринадцать листов — вводные, ТЭП, сроки, продажи,
себестоимость, ставки, эскроу, кредитование, налоги, CF, LLCR, ВРИ и отчёт.
Значениями приходят только драйверы сценария; весь финансовый контур —
формулы, и здесь они считаются прямо в тесте и сверяются с движком помесячно.
Расхождение допускается только на уровне порядка сложения чисел с плавающей
точкой.

Заодно закреплено, ради чего книга и собиралась: аналитик меняет вводную, и
книга пересчитывается сама, а не показывает выгруженное движком число.

Запуск: python3 -m pytest tests -q
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT.parent))
sys.path.insert(0, str(ROOT))

import main as wrapper  # noqa: E402

# Вычислитель формул идёт по ссылкам рекурсией: линия ПФ, закрытая после
# РВЭ, читает нехватку по предыдущим месяцам, и цепочка стала длиннее предела
# по умолчанию. Книге это не мешает — Excel считает итеративно; предел
# поднимается так же, как в проверках книги v4.
sys.setrecursionlimit(400000)
from xlsx_eval import Evaluator, FormulaError  # noqa: E402

core = wrapper.core
openpyxl = pytest.importorskip("openpyxl")

from openpyxl.utils import get_column_letter  # noqa: E402

# Книга считает в миллионах. Драйверы графиков движок отдаёт округлёнными до
# копеек, поэтому доли периода и
# индексы цены несут копеечную погрешность: 1e-3 млн ₽ — это тысяча рублей на
# сорока миллиардах. Всё, что больше, — уже методика.
KOPECK = 1e-3


# Сверка книги с движком должна идти на прибыльном проекте: у убыточного налог
# равен нулю, и проверки вроде «ставка налога двигает LLCR» перестают что-либо
# проверять, оставаясь зелёными. Прежде прибыль давали сами умолчания, но после
# сверки удельных ставок с банковским бюджетом комфорт по 350 тыс ₽/м² на них не
# сходится — поэтому сценарий задаётся здесь явно, а не достаётся по случайности.
BOOK_INPUTS = {**core.DEFAULT_INPUTS, "apartment_price_th": 650,
               "commercial_price_th": 650, "parking_price_th": 5000}


@pytest.fixture(scope="module")
def book():
    data, meta = core.build_plato_model_v2(
        dict(BOOK_INPUTS), dict(core.TEP_DEFAULT), [], project_name="Проверка",
    )
    workbook = openpyxl.load_workbook(io.BytesIO(data))
    engine = core.calculate(core.CalcRequest(
        inputs=dict(BOOK_INPUTS), tep=dict(core.TEP_DEFAULT), rates=[],
    ))
    return workbook, Evaluator(workbook), engine, meta


def close(got: float, expected: float) -> bool:
    return abs(got - expected) <= max(KOPECK, abs(expected) * 1e-9)


def column(index: int) -> str:
    return get_column_letter(2 + index)


# --- состав книги ----------------------------------------------------------

def test_the_workbook_has_the_sheets_a_model_is_made_of(book):
    """Три листа — это выгрузка, а не финмодель."""
    workbook, _, _, _ = book

    assert workbook.sheetnames == [
        "ОТЧЁТ", "Вводные", "ТЭП", "СРОКИ", "ПРОДАЖИ", "СЕБЕСТОИМОСТЬ", "СТАВКИ",
        "ЭСКРОУ", "КРЕДИТОВАНИЕ", "НАЛОГИ", "CF", "LLCR", "ВРИ",
    ]


# --- помесячный расчёт -----------------------------------------------------

# Лист и строка книги -> та же величина в строке движка и её масштаб.
MONTHLY = [
    ("ЭСКРОУ", "escrow", "balance", "escrow", 1e6),
    ("ЭСКРОУ", "escrow", "release", "escrow_release", 1e6),
    ("КРЕДИТОВАНИЕ", "credit", "bridge_draw", "bridge_draw", 1e6),
    ("КРЕДИТОВАНИЕ", "credit", "bridge_balance", "bridge_balance", 1e6),
    ("КРЕДИТОВАНИЕ", "credit", "bridge_interest", "bridge_interest", 1e6),
    ("КРЕДИТОВАНИЕ", "credit", "bridge_cap", "bridge_capitalization", 1e6),
    ("КРЕДИТОВАНИЕ", "credit", "bridge_refinance", "bridge_repayment", 1e6),
    ("КРЕДИТОВАНИЕ", "credit", "pf_draw", "pf_draw", 1e6),
    ("КРЕДИТОВАНИЕ", "credit", "pf_repayment", "pf_repayment", 1e6),
    ("КРЕДИТОВАНИЕ", "credit", "pf_balance", "pf_balance", 1e6),
    ("КРЕДИТОВАНИЕ", "credit", "coverage", "coverage", 1.0),
    ("КРЕДИТОВАНИЕ", "credit", "pf_rate", "pf_rate", 1.0),
    ("КРЕДИТОВАНИЕ", "credit", "pf_interest", "pf_interest", 1e6),
    ("КРЕДИТОВАНИЕ", "credit", "pf_cap", "pf_interest_capitalization", 1e6),
    ("КРЕДИТОВАНИЕ", "credit", "limit_fee", "limit_fee", 1e6),
    ("КРЕДИТОВАНИЕ", "credit", "interest_payment", "interest_payment", 1e6),
    ("НАЛОГИ", "tax", "tax", "profit_tax", 1e6),
]


@pytest.mark.parametrize("sheet,grid,row_key,engine_key,scale", MONTHLY)
def test_the_monthly_row_matches_the_engine(book, sheet, grid, row_key, engine_key, scale):
    _, evaluator, engine, meta = book
    line = meta["layout"][grid][row_key]

    for index, item in enumerate(engine["finance"]["rows"]):
        got = evaluator.cell(sheet, f"{column(index)}{line}")
        expected = float(item.get(engine_key) or 0.0) / scale
        assert close(got, expected), (
            f"{sheet}·{row_key} в {item['month']}: книга {got}, движок {expected}"
        )


@pytest.mark.parametrize("row_key,engine_key", [("project", "project"), ("equity", "equity")])
def test_the_cash_flow_matches_the_engine(book, row_key, engine_key):
    _, evaluator, engine, meta = book
    line = meta["layout"]["cf"][row_key]
    series = engine["cashflow"][engine_key]

    for index, month in enumerate(engine["cashflow"]["months"]):
        got = evaluator.cell("CF", f"{column(index)}{line}")
        assert close(got, float(series[index]) / 1e6), f"CF·{row_key} в {month}"


def test_the_coverage_is_taken_before_the_escrow_is_released(book):
    """На РВЭ покрытие считается по накопленному эскроу, а не по нулю после раскрытия.

    Это тот самый порядок внутри месяца: сначала выборка ПФ, потом покрытие,
    ставка и проценты, и только затем раскрытие и погашение. Возьми покрытие
    после раскрытия — на РВЭ оно обнулится, ставка подскочит к базовой, и
    проценты уедут вверх на весь остаток проекта.
    """
    _, evaluator, engine, meta = book
    rve = engine["dates"]["rve"]
    index = [i for i, r in enumerate(engine["finance"]["rows"]) if r["month"] == rve]
    assert index, "месяц РВЭ не найден в расчёте"
    letter = column(index[0])

    coverage = evaluator.cell("КРЕДИТОВАНИЕ", f"{letter}{meta['layout']['credit']['coverage']}")
    escrow_after = evaluator.cell("ЭСКРОУ", f"{letter}{meta['layout']['escrow']['balance']}")

    assert escrow_after == pytest.approx(0.0, abs=KOPECK), "эскроу на РВЭ не раскрыто"
    assert coverage > 0.5, f"покрытие на РВЭ схлопнулось до {coverage}"
    assert coverage == pytest.approx(engine["finance"]["rows"][index[0]]["coverage"])


def test_the_escrow_holds_the_money_until_the_permit_to_occupy(book):
    """До РВЭ выручка на эскроу и в поток на собственный капитал не попадает."""
    _, evaluator, engine, meta = book
    rve = engine["dates"]["rve"]
    cash_in = meta["layout"]["cf"]["cash_in"]

    for index, item in enumerate(engine["finance"]["rows"]):
        if item["month"] >= rve:
            continue
        assert evaluator.cell("CF", f"{column(index)}{cash_in}") == pytest.approx(0.0), (
            f"в {item['month']} деньги дошли до акционера раньше раскрытия эскроу"
        )


# --- отчёт -----------------------------------------------------------------

def test_the_report_agrees_with_its_own_control_column(book):
    """Колонка D — разница между формулами книги и расчётом DevelopAid."""
    workbook, evaluator, _, _ = book
    sheet = workbook["ОТЧЁТ"]

    checked = 0
    for row in range(4, 4 + len(core._M2_REPORT_KEYS)):
        expected = sheet.cell(row=row, column=3).value
        got = evaluator.cell("ОТЧЁТ", f"B{row}")
        assert close(got, float(expected)), (
            f"{sheet.cell(row=row, column=1).value}: книга {got}, движок {expected}"
        )
        assert close(evaluator.cell("ОТЧЁТ", f"D{row}"), 0.0)
        checked += 1

    assert checked == len(core._M2_REPORT_KEYS)


def test_the_llcr_sheet_shows_its_own_arithmetic(book):
    """LLCR — не одно число в отчёте, а числитель, знаменатель и их состав."""
    _, evaluator, engine, _ = book

    # Строки сдвинулись на одну: между налогом на прибыль и выборкой ПФ встал
    # НДС — он такой же денежный расход и обязан уходить из числителя.
    numerator = evaluator.cell("LLCR", "B10")
    denominator = evaluator.cell("LLCR", "B13")
    llcr = evaluator.cell("LLCR", "B14")

    assert close(numerator, engine["finance"]["llcr_numerator"] / 1e6)
    assert close(denominator, engine["finance"]["llcr_denominator"] / 1e6)
    assert llcr == pytest.approx(engine["summary"]["llcr"], rel=1e-7)
    assert evaluator.cell("ОТЧЁТ", f"B{4 + core._M2_REPORT_KEYS.index('llcr')}") == pytest.approx(llcr)


# --- живая книга -----------------------------------------------------------

def test_changing_a_spread_moves_the_report(book):
    """Ради этого книга и собиралась: вводная меняется — считает книга, не движок."""
    workbook, _, _, meta = book
    fresh = _reopen(workbook)
    row = 4 + core._M2_REPORT_KEYS.index("financing_cost")
    was = Evaluator(fresh).cell("ОТЧЁТ", f"B{row}")

    line = meta["layout"]["inputs"]["pf_spread_pp"]
    fresh["Вводные"][f"B{line}"] = fresh["Вводные"][f"B{line}"].value + 0.02

    after = Evaluator(fresh).cell("ОТЧЁТ", f"B{row}")
    # Порог относительный, а не «плюс 100»: спред входит только в базовую ставку,
    # и чем выше покрытие эскроу, тем меньше её вес в средневзвешенной. На хорошо
    # покрытом проекте два пункта спреда двигают проценты меньше чем на процент —
    # абсолютный порог мерил бы не проходимость вводной, а покрытие. Полпроцента
    # взяты вдвое ниже наблюдаемых 0,995%: порог вплотную к факту уже подвёл, а
    # численный шум здесь на семь порядков меньше.
    assert after > was * 1.005, f"плюс 2 п.п. к спреду ПФ не сдвинули проценты: {was} → {after}"


def test_the_rate_sheet_feeds_the_credit_sheet(book):
    """Спред живёт на «Вводных», ставка — на «СТАВКАХ», кредит их читает."""
    workbook, _, _, meta = book
    fresh = _reopen(workbook)
    rate_row = meta["layout"]["credit"]["pf_base_rate"]
    was = Evaluator(fresh).cell("КРЕДИТОВАНИЕ", f"C{rate_row}")

    line = meta["layout"]["inputs"]["pf_spread_pp"]
    fresh["Вводные"][f"B{line}"] = fresh["Вводные"][f"B{line}"].value + 0.01

    assert Evaluator(fresh).cell("КРЕДИТОВАНИЕ", f"C{rate_row}") == pytest.approx(was + 0.01)


def test_the_tax_rate_reaches_the_llcr(book):
    workbook, _, _, meta = book
    fresh = _reopen(workbook)
    row = 4 + core._M2_REPORT_KEYS.index("llcr")
    was = Evaluator(fresh).cell("ОТЧЁТ", f"B{row}")

    fresh["Вводные"][f"B{meta['layout']['inputs']['profit_tax_pct']}"] = 0.0

    assert Evaluator(fresh).cell("ОТЧЁТ", f"B{row}") > was


# Драйверы — это сценарий: объёмы, выручка по продуктам, статьи затрат, прогноз
# ставки, налоговая маржа. Всё прочее на помесячных листах обязано быть формулой.
DRIVER_ROWS = {
    "ПРОДАЖИ": (),
    "СЕБЕСТОИМОСТЬ": ("cost_land_rights", "cost_social", "debt", "index"),
    "СТАВКИ": (),
    "ЭСКРОУ": (),
    "КРЕДИТОВАНИЕ": (),
    # «vat» — такой же сценарный ряд, как маржа: НДС считает движок, книга
    # его получает значениями и тратит деньгами.
    "НАЛОГИ": ("margin", "adjust", "vat"),
    "CF": (),
}
GRID_OF = {"ПРОДАЖИ": "sales", "СЕБЕСТОИМОСТЬ": "costs", "СТАВКИ": "rates",
           "ЭСКРОУ": "escrow", "КРЕДИТОВАНИЕ": "credit", "НАЛОГИ": "tax", "CF": "cf"}


@pytest.mark.parametrize("sheet", list(DRIVER_ROWS))
def test_the_financial_block_is_formulas_and_not_numbers(book, sheet):
    """Захардкоженное число вместо формулы — это возврат к мёртвой выгрузке."""
    workbook, _, engine, meta = book
    page = workbook[sheet]
    months = len(engine["finance"]["rows"])
    drivers = DRIVER_ROWS[sheet]

    checked = 0
    for key, line in meta["layout"][GRID_OF[sheet]].items():
        if any(key == name or key.startswith(name) for name in drivers):
            continue
        for index in range(months):
            value = page.cell(row=line, column=2 + index).value
            assert isinstance(value, str) and value.startswith("="), (
                f"{sheet}·{key} в колонке {column(index)} — не формула: {value!r}"
            )
        checked += 1

    assert checked, f"на листе {sheet} не осталось ни одной расчётной строки"


def test_every_formula_is_understood(book):
    """Вычислитель узкий нарочно: непонятная формула должна ломать тест."""
    workbook, evaluator, _, _ = book

    total = 0
    for name in workbook.sheetnames:
        sheet = workbook[name]
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    total += 1
                    try:
                        evaluator.cell(name, cell.coordinate)
                    except FormulaError as exc:
                        pytest.fail(f"{name}!{cell.coordinate}: {exc}\n{cell.value}")

    assert total > 1000, f"формул подозрительно мало: {total}"


def test_the_price_drives_the_revenue(book):
    """Выручка — это объём из ТЭП на цену с «Вводных», а не выгруженный ряд.

    Прежде продажи лежали в книге числами: правка цены квадратного метра ничего
    не двигала, и модель нельзя было использовать по назначению — посмотреть,
    что будет при другой цене.
    """
    workbook, _, _, meta = book
    fresh = _reopen(workbook)
    row = 4 + core._M2_REPORT_KEYS.index("revenue")
    was = Evaluator(fresh).cell("ОТЧЁТ", f"B{row}")

    line = meta["layout"]["inputs"]["apartment_price_th"]
    fresh["Вводные"][f"B{line}"] = fresh["Вводные"][f"B{line}"].value * 1.1

    after = Evaluator(fresh).cell("ОТЧЁТ", f"B{row}")
    assert after > was * 1.05, f"плюс 10% к цене квартир не сдвинули выручку: {was} → {after}"


def test_the_area_drives_the_revenue(book):
    """Площадь живёт на «ТЭП», и продажи считаются от неё."""
    workbook, _, _, meta = book
    fresh = _reopen(workbook)
    row = 4 + core._M2_REPORT_KEYS.index("revenue")
    was = Evaluator(fresh).cell("ОТЧЁТ", f"B{row}")

    sheet = fresh["ТЭП"]
    line = [r for r in range(5, sheet.max_row + 1) if sheet.cell(r, 1).value == "Квартиры"][0]
    sheet.cell(row=line, column=5).value = sheet.cell(row=line, column=5).value * 0.5

    assert Evaluator(fresh).cell("ОТЧЁТ", f"B{row}") < was * 0.7


def test_the_unit_rate_drives_the_capex(book):
    """Себестоимость — ставка на базу: тысячи рублей за метр на ГНС."""
    workbook, _, _, meta = book
    fresh = _reopen(workbook)
    row = 4 + core._M2_REPORT_KEYS.index("capex")
    was = Evaluator(fresh).cell("ОТЧЁТ", f"B{row}")

    line = meta["layout"]["inputs"]["main_above_th_per_sqm"]
    fresh["Вводные"][f"B{line}"] = fresh["Вводные"][f"B{line}"].value + 10

    after = Evaluator(fresh).cell("ОТЧЁТ", f"B{row}")
    assert after > was + 1000, f"плюс 10 тыс. ₽/м² не сдвинули CAPEX: {was} → {after}"


def test_the_marketing_share_follows_the_revenue(book):
    """Маркетинг и продажи — процент от выручки, а не отдельный ряд чисел."""
    workbook, _, _, meta = book
    fresh = _reopen(workbook)
    row = 4 + core._M2_REPORT_KEYS.index("operating")
    was = Evaluator(fresh).cell("ОТЧЁТ", f"B{row}")

    line = meta["layout"]["inputs"]["marketing_pct"]
    fresh["Вводные"][f"B{line}"] = fresh["Вводные"][f"B{line}"].value * 2

    assert Evaluator(fresh).cell("ОТЧЁТ", f"B{row}") > was * 1.3


def test_the_term_of_the_permit_is_calculated_not_typed(book):
    """РнС и РВЭ — это старт проекта плюс сроки ИРД и строительства."""
    workbook, _, engine, meta = book
    fresh = _reopen(workbook)
    evaluator = Evaluator(fresh)

    permit = evaluator.cell("Вводные", f"B{meta['layout']['inputs']['permit']}")
    rve = evaluator.cell("Вводные", f"B{meta['layout']['inputs']['rve']}")

    assert permit.isoformat()[:10] == engine["dates"]["permit"]
    assert rve.isoformat()[:10] == engine["dates"]["rve"]


def test_a_longer_permit_stage_moves_the_construction_window(book):
    """Сдвинули срок ИРД — поехал и график освоения, и стоимость долга."""
    workbook, _, _, meta = book
    fresh = _reopen(workbook)
    row = 4 + core._M2_REPORT_KEYS.index("financing_cost")
    was = Evaluator(fresh).cell("ОТЧЁТ", f"B{row}")

    line = meta["layout"]["inputs"]["ird_months"]
    fresh["Вводные"][f"B{line}"] = fresh["Вводные"][f"B{line}"].value + 6

    assert Evaluator(fresh).cell("ОТЧЁТ", f"B{row}") != pytest.approx(was)


def test_every_input_of_the_model_is_on_the_sheet(book):
    """Куцые вводные — это книга, в которой нечего менять."""
    _, _, _, meta = book
    on_sheet = set(meta["layout"]["inputs"])

    expected = {field[0] for _, fields in core.FIELD_GROUPS for field in fields}
    assert not expected - on_sheet, f"на «Вводных» нет полей: {sorted(expected - on_sheet)}"
    assert len(expected) > 100


# Вводные, которые обязаны двигать отчёт, и показатель, по которому это видно.
DRIVES = [
    ("apartment_price_th", 1.1, "revenue"),
    ("commercial_price_th", 1.2, "revenue"),
    ("parking_price_th", 1.3, "revenue"),
    ("share_before_rve_pct", 0.9, "financing_cost"),
    ("monthly_growth_pre_pct", 1.5, "revenue"),
    ("monthly_growth_post_pct", 3.0, "revenue"),
    ("seasonal_reduction_pct", 2.0, "financing_cost"),
    ("pace_adjustment_pct", 2.0, "financing_cost"),
    ("residual_sales_months", 2.0, "revenue"),
    ("sales_lag_months", 3.0, "revenue"),
    ("main_above_th_per_sqm", 1.2, "capex"),
    ("main_under_th_per_sqm", 1.2, "capex"),
    ("utilities_th_per_sqm", 1.5, "capex"),
    ("gc_fee_pct", 1.5, "capex"),
    ("reserve_pct", 1.5, "capex"),
    ("project_management_pct", 1.5, "capex"),
    ("technical_supervision_pct", 1.5, "capex"),
    ("marketing_pct", 2.0, "operating"),
    ("selling_pct", 2.0, "operating"),
    ("ird_months", 1.5, "financing_cost"),
    ("construction_months", 1.5, "financing_cost"),
    ("bridge_spread_pp", 1.5, "financing_cost"),
    ("pf_spread_pp", 1.5, "financing_cost"),
    ("pf_special_pct", 2.0, "financing_cost"),
    ("limit_fee_pct", 2.0, "financing_cost"),
    ("reservation_fee_pct", 2.0, "financing_cost"),
    ("profit_tax_pct", 0.5, "profit_tax"),
    ("purchase_price_mln", 2.0, "capex"),
    ("rate_start_pct", 1.5, "financing_cost"),
    ("rate_target_base_pct", 1.5, "financing_cost"),
    ("rate_normalization_months", 2.0, "financing_cost"),
    ("rate_curve_shape", 3.0, "financing_cost"),
]


@pytest.mark.parametrize("key,factor,indicator", DRIVES)
def test_the_input_reaches_the_report(book, key, factor, indicator):
    """Вводная, которая ничего не двигает, — это подпись, а не параметр.

    Сезонность и смещение темпа особенно: они уже были в интерфейсе и на
    расчёт не влияли — уходили только в шаблон, и два расчёта по одним и тем же
    вводным давали разную выручку.
    """
    workbook, _, _, meta = book
    fresh = _reopen(workbook)
    row = 4 + core._M2_REPORT_KEYS.index(indicator)
    was = Evaluator(fresh).cell("ОТЧЁТ", f"B{row}")

    line = meta["layout"]["inputs"][key]
    value = fresh["Вводные"][f"B{line}"].value
    fresh["Вводные"][f"B{line}"] = (value or 1.0) * factor if value else factor

    after = Evaluator(fresh).cell("ОТЧЁТ", f"B{row}")
    assert after != pytest.approx(was, rel=1e-9), (
        f"«{key}» не двигает «{indicator}»: {was} → {after}"
    )


def test_the_share_before_the_permit_is_exactly_what_was_asked(book):
    """Веса нормируются до и после РВЭ отдельно — доля не плывёт за сезонностью."""
    _, evaluator, engine, meta = book
    rve = engine["dates"]["rve"]
    line = meta["layout"]["sales"]["quantity_apartments"]

    before = after = 0.0
    for index, item in enumerate(engine["finance"]["rows"]):
        volume = evaluator.cell("ПРОДАЖИ", f"{column(index)}{line}")
        if item["month"] < rve:
            before += volume
        else:
            after += volume

    share = core.n(core.DEFAULT_INPUTS, "share_before_rve_pct") / 100
    assert before / (before + after) == pytest.approx(share, rel=1e-6)


def test_the_seasonal_months_sell_less(book):
    """Январь, май–август — месяцы пониженного спроса, как в шаблоне ПЛАТО."""
    _, evaluator, engine, meta = book
    line = meta["layout"]["sales"]["season_apartments"]

    for index, item in enumerate(engine["finance"]["rows"]):
        month = int(item["month"][5:7])
        got = evaluator.cell("ПРОДАЖИ", f"{column(index)}{line}")
        expected = 1 + core.n(core.DEFAULT_INPUTS, "seasonal_reduction_pct") / 100 \
            if month in core._M2_SEASONAL_LOW_MONTHS else 1.0
        assert got == pytest.approx(expected), f"сезонность в {item['month']}"


def test_the_fields_the_engine_never_reads_are_marked(book):
    """Помеченные поля обязаны быть ровно теми, что до расчёта не доходят.

    Этап роста цены и инфляция после РВЭ живут на странице и уезжают в шаблон
    ПЛАТО, а движок их не читает. Без пометки аналитик правит их в книге, ничего
    не происходит, и виноватой выглядит книга.
    """
    workbook, _, _, meta = book
    sheet = workbook["Вводные"]

    inputs = dict(core.DEFAULT_INPUTS)
    inputs.update({"offices_enabled": True, "retail_enabled": True,
                   "above_parking_enabled": True, "purchase_price_mln": 700})

    def outcome(data):
        result = core.calculate(core.CalcRequest(
            inputs=dict(data), tep=dict(core.TEP_DEFAULT), rates=[]))
        summary = result["summary"]
        return tuple(round(summary[k], 4) for k in
                     ("revenue", "capex", "financing_cost", "profit_tax", "llcr"))

    reference = outcome(inputs)
    for key in core._M2_TEMPLATE_ONLY_INPUTS:
        trial = dict(inputs)
        trial[key] = core.n(inputs, key, 0.0) * 1.5 + 7
        assert outcome(trial) == reference, (
            f"«{key}» помечена как неучаствующая, но расчёт от неё меняется"
        )
        line = meta["layout"]["inputs"][key]
        note = str(sheet.cell(row=line, column=3).value or "")
        assert "не участвует" in note, f"«{key}» не помечена на «Вводных»"


# Сценарии, на которых книга обязана сходиться с движком: одних умолчаний мало —
# объекты КРТ, рассрочка ВРИ и денежная компенсация идут по своим ветвям.
SCENARIOS = {
    "умолчания": {},
    "объекты КРТ и цена входа": {
        "offices_enabled": True, "retail_enabled": True,
        "above_parking_enabled": True, "purchase_price_mln": 700,
    },
    "рассрочка ВРИ": {
        "vri_payment_mode": "installment", "vri_installment_years": 6,
        "purchase_price_mln": 300,
    },
    "денежная компенсация соцобъектов": {
        "social_mode": "Денежная компенсация", "social_compensation_mln": 580.7,
    },
    "консервативный сценарий ставки": {"rate_scenario": "high"},
    "оптимистичный сценарий ставки": {"rate_scenario": "low"},
    "проценты БРИДЖ выплачиваются": {
        "bridge_interest_mode": "Выплата при рефинансировании",
    },
}


@pytest.mark.parametrize("name", list(SCENARIOS))
def test_the_report_holds_on_other_scenarios(name):
    """Сходимость на умолчаниях ничего не значит, если ветка одна."""
    inputs = dict(core.DEFAULT_INPUTS)
    inputs.update(SCENARIOS[name])
    data, _ = core.build_plato_model_v2(inputs, dict(core.TEP_DEFAULT), [], project_name=name)
    workbook = openpyxl.load_workbook(io.BytesIO(data))
    evaluator = Evaluator(workbook)
    sheet = workbook["ОТЧЁТ"]

    for row in range(4, 4 + len(core._M2_REPORT_KEYS)):
        expected = float(sheet.cell(row=row, column=3).value)
        got = evaluator.cell("ОТЧЁТ", f"B{row}")
        assert close(got, expected), (
            f"{name}: {sheet.cell(row=row, column=1).value} — книга {got}, движок {expected}"
        )


def test_the_reserve_is_not_charged_on_the_purchase_price(name="цена входа"):
    """Цена входа в базу резерва не входит: движок берёт процент не от неё.

    Семьсот миллионов покупки давали тридцать пять миллионов резерва из ниоткуда,
    и вслед за ними разъезжались CAPEX, прибыль и LLCR.
    """
    inputs = dict(core.DEFAULT_INPUTS)
    inputs["purchase_price_mln"] = 700
    data, meta = core.build_plato_model_v2(inputs, dict(core.TEP_DEFAULT), [], project_name=name)
    workbook = openpyxl.load_workbook(io.BytesIO(data))
    evaluator = Evaluator(workbook)

    row = 4 + core._M2_REPORT_KEYS.index("capex")
    expected = float(workbook["ОТЧЁТ"].cell(row=row, column=3).value)
    assert close(evaluator.cell("ОТЧЁТ", f"B{row}"), expected)


def test_the_key_rate_curve_matches_the_engine(book):
    """Прогноз ставки собирается формулой, а не приезжает рядом чисел."""
    _, evaluator, engine, meta = book
    line = meta["layout"]["rates"]["key_rate"]

    for index, item in enumerate(engine["finance"]["rows"]):
        got = evaluator.cell("СТАВКИ", f"{column(index)}{line}")
        assert got == pytest.approx(float(item["key_rate"]), abs=1e-12), (
            f"ключевая ставка в {item['month']}"
        )


def test_the_rate_scenario_is_chosen_in_the_workbook(book):
    """Сценарий ставки — выбор из списка, и он двигает всю кривую."""
    workbook, _, _, meta = book
    fresh = _reopen(workbook)
    line = meta["layout"]["rates"]["key_rate"]
    was = Evaluator(fresh).cell("СТАВКИ", f"BD{line}")

    fresh["Вводные"][f"B{meta['layout']['inputs']['rate_scenario']}"] = "Оптимистичный"

    assert Evaluator(fresh).cell("СТАВКИ", f"BD{line}") < was


# Поля, у которых на странице выбор из списка: в книге у них обязан быть
# тот же список, иначе аналитик впишет своё и ветка расчёта уйдёт молча.
CHOICE_FIELDS = [
    ("vri_region", ["Москва", "Московская область"]),
    ("land_right", ["Собственность", "Аренда"]),
    ("vri_payment_mode", ["Единовременно", "Рассрочка"]),
    ("vri_periodicity_months", ["Ежемесячно", "Ежеквартально", "Раз в полгода", "Раз в год"]),
    ("vri_relief_mode", ["Нет", "Доля от суммы", "Фиксированная сумма"]),
    ("vri_financing_mode", ["Как весь проект", "Заданные доли"]),
    # Третья форма — и стройка, и компенсация: у Румянцева школа с садиком
    # строятся, а за стадион платят деньгами (решение владельца 14.08.2026).
    ("social_mode", ["Строительство", "Денежная компенсация",
                     "Строительство и компенсация"]),
    ("bridge_interest_mode", ["Капитализация в ПФ", "Выплата при рефинансировании"]),
    ("rate_scenario", ["Консервативный", "Базовый", "Оптимистичный"]),
    ("vri_required", ["Да", "Нет"]),
    ("offices_enabled", ["Да", "Нет"]),
]


@pytest.mark.parametrize("key,expected", CHOICE_FIELDS)
def test_the_choice_fields_offer_the_same_list_as_the_page(book, key, expected):
    workbook, _, _, meta = book
    sheet = workbook["Вводные"]
    address = f"B{meta['layout']['inputs'][key]}"

    lists = [rule for rule in sheet.data_validations.dataValidation
             if rule.type == "list" and address in str(rule.sqref)]
    assert lists, f"у «{key}» нет выпадающего списка"

    offered = lists[0].formula1.strip('"').split(",")
    assert offered == expected, f"«{key}»: в книге {offered}, на странице {expected}"
    assert sheet[address].value in expected, (
        f"«{key}» хранит {sheet[address].value!r}, чего нет в списке"
    )


def test_every_choice_on_the_page_is_a_choice_in_the_workbook(book):
    """Ни одно поле выбора не должно остаться свободным вводом."""
    workbook, _, _, meta = book
    sheet = workbook["Вводные"]
    validated = {
        address
        for rule in sheet.data_validations.dataValidation if rule.type == "list"
        for address in str(rule.sqref).split()
    }

    for _, fields in core.FIELD_GROUPS:
        for field in fields:
            key, kind = field[0], field[3]
            # Свободный ввод бывает и осознанным: строка «text» — не выбор из
            # списка, а данные, которых мы не знаем заранее (лестница ставок
            # переписывается из конкретного НКЛ). Сторожим здесь именно поля
            # выбора: у них список известен, и книга обязана его повторить.
            if kind in ("number", "date", "text", "pf_steps", "schedule"):
                continue
            # Неизвестный тип поля не молчит: он либо выбор, либо свободный ввод,
            # и решать это должен тот, кто его завёл. Прежде такое поле молча
            # считалось выбором, и проверка падала с диагнозом «остался свободным
            # вводом» — то есть не о том.
            assert kind in ("select", "checkbox", "finance_select"), (
                f"«{key}»: тип поля «{kind}» не назван ни выбором, ни свободным вводом")
            address = f"B{meta['layout']['inputs'][key]}"
            assert address in validated, f"«{key}» ({kind}) остался свободным вводом"


def test_the_limits_are_calculated_not_typed(book):
    """Лимит БРИДЖ и лимит ПФ — следствие расчёта, а не введённые числа."""
    _, evaluator, engine, meta = book
    finance = engine["finance"]

    bridge = evaluator.cell("Вводные", f"B{meta['layout']['inputs']['bridge_limit']}")
    pf = evaluator.cell("Вводные", f"B{meta['layout']['inputs']['pf_limit']}")

    assert close(bridge, finance["calculated_bridge_limit"] / 1e6)
    assert close(pf, finance["pf_limit"] / 1e6)


def test_the_book_sells_only_the_spaces_that_are_sold(book):
    """Гостевые машино-места строятся и стоят денег, но не продаются.

    Объём продаж паркинга книга брала из колонки «Единиц» — всех построенных
    мест, — и продавала гостевые вместе с остальными. На умолчаниях это
    624 млн ₽ лишней выручки: она уходила на эскроу, поднимала покрытие,
    роняла ступень ставки ПФ и меняла налог и LLCR. Ни одна из этих величин
    ошибкой не выглядела — расходились две достоверные на вид модели.
    """
    workbook, evaluator, engine, meta = book
    row = next(item for item in engine["tep"]["rows"]
               if item["key"] == "underground_parking")
    guest = core.underground_guest_spaces(core.TEP_DEFAULT["underground_parking"])
    assert guest > 0, "без гостевых мест проверка ничего не проверяет"

    sheet = workbook[core._M2_SHEETS["tep"]]
    line = next(cell.row for cell in sheet["A"]
                if cell.value == row["label"])
    assert sheet.cell(row=line, column=8).value == pytest.approx(guest)
    assert sheet.cell(row=line, column=9).value == f"=G{line}-H{line}"

    sold = evaluator.cell(core._M2_SHEETS["tep"], f"I{line}")
    assert sold == pytest.approx(core.underground_saleable_spaces(
        core.TEP_DEFAULT["underground_parking"]))
    assert sold < float(row["units"])


# --- вспомогательное -------------------------------------------------------

def _reopen(workbook):
    """Свежая копия: вычислитель кэширует значения, править надо чистую книгу."""
    buffer = io.BytesIO()
    workbook.save(buffer)
    return openpyxl.load_workbook(io.BytesIO(buffer.getvalue()))
