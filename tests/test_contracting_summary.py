"""Свод продаж действующего проекта считается по первичке, а не по своду.

Файл ЦФ несёт два листа, которые нельзя путать: «Контрактация» — обогащённый
реестр договоров (вариант оплаты, брокер, премии, график эскроу), «1С_Факт» —
проводки, где 76.06 это заключение ДДУ, а 008.01 — движение денег на эскроу.

Проверяется здесь то, что на живом файле уже подводило.

Реестра владельца в репозитории нет и не будет: в нём ФИО покупателей. Форма
листа воспроизведена синтетикой — правило проверяется правилом, а не чужими
персональными данными.

Запуск: python3 -m pytest tests/test_contracting_summary.py -q
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from market_search import contracting  # noqa: E402

HEAD = ["Квартал", "Год", "Месяц", "Объект недвижимости", "Корпус",
        "Тип объекта недвижимости", "Вид отделки", "Проектная S", "Договор",
        "Тип договора", "Состояние договора", "Сумма договор", "Дата договора",
        "Покупатель", "Шт", "Цена ДДУ за кв. м", "Вариант оплаты",
        "Оплачено премии ОП ", "Брокеры", "% брокера", "Наименование брокера",
        "% оплаты", "остаток к оплате", "оплачено эскроу всего"]

DEALS = [
    # квартира через брокера, рассрочка: оплачено меньше суммы договора
    ["3Q 2025", 2025, 45900, "Квартира", "корп.1 кв.1", "жилая", "без отделки", 40.0,
     "ДДУ №1-1-1/ГР от 12.08.2025", "ДДУ", "Действующий", 24_000_000, "12.08.2025",
     "Иванов Иван Иванович", 1, 600_000, "Рассрочка 20% ПВ и далее по 200 000 в месяц",
     0, 1_200_000, 0.05, "БРОКЕР ООО", 0.25, 18_000_000, 6_000_000],
    # квартира напрямую, 100% оплата
    ["3Q 2025", 2025, 45900, "Квартира", "корп.1 кв.2", "жилая", "без отделки", 60.0,
     "ДДУ №1-1-2/ГР от 20.08.2025", "ДДУ", "Действующий", 30_000_000, "20.08.2025",
     "Петров Пётр Петрович", 1, 500_000, "1.0", 90_000, 0, None, "", 1.0, 0, 30_000_000],
    # машино-место юрлицу через брокера без заполненной комиссии
    [None, 2025, 45930, "Машиноместа", None, "нежилая", " ", 14.0,
     "ДДУ №0001М/ГР от 05.09.2025", "ДДУ", "Действующий", 4_000_000, "05.09.2025",
     "Аэробилдинг ООО", 1, 285_714, "1.0", 0, 0, 0.05, "ТИХИЙ БРОКЕР ООО", 1.0, 0, 4_000_000],
]

LEDGER = [
    ["Год", "Квартал", "Месяц", "Дата", "Документ", "Организация", "Счет Дт",
     "Субконто1 Дт", "Субконто2 Дт", "Субконто3 Дт", "Счет Кт", "Субконто1 Кт",
     "Субконто2 Кт", "Субконто3 Кт", "Сумма", "Содержание"],
    [2025, "3Q 2025", 45900, "12.08.2025", "Операция", "СЗ", "76.06", "Иванов",
     "ДДУ №1-1-1/ГР от 12.08.2025", None, "86.02", "ЖК", None, None, 24_000_000, "корп.1 кв.1 S=40 м2"],
    [2025, "3Q 2025", 45900, "15.08.2025", "Операция", "СЗ", "008.01", "Иванов",
     "ДДУ №1-1-1/ГР от 12.08.2025", None, None, None, None, None, 6_000_000, None],
    # расторгнутый договор: начисление, деньги, сторно и возврат
    [2025, "3Q 2025", 45930, "09.09.2025", "Операция", "СЗ", "76.06", "Сидоров",
     "ДДУ №1-2-9/ГР от 09.09.2025", None, "86.02", "ЖК", None, None, 50_000_000, "корп.1 кв.9 S=80 м2"],
    [2025, "4Q 2025", 45991, "01.11.2025", "Операция", "СЗ", "008.01", "Сидоров",
     "ДДУ №1-2-9/ГР от 09.09.2025", None, None, None, None, None, 50_000_000, None],
    [2026, "1Q 2026", 46065, 46065, None, "СЗ", "76.06", "Сидоров",
     "Расторжение ДДУ №1-2-9/ГР от 09.09.2025", None, "86.02", "ЖК", None, None,
     -50_000_000, "корп.1 кв.9 S=80 м2"],
    [2026, "1Q 2026", 46071, 46071, None, "СЗ", "008.01", "Сидоров",
     "ДДУ №1-2-9/ГР от 09.09.2025", None, None, None, None, None, -49_800_000, None],
]


def _book() -> bytes:
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = contracting.SHEET_CONTRACTS
    sheet["A2"] = "Контрактация"
    sheet["A3"] = "Название проекта:"
    sheet["E3"] = "Тестовый ЖК"
    for column, title in enumerate(HEAD, start=1):
        sheet.cell(row=6, column=column, value=title)
    for index, deal in enumerate(DEALS):
        for column, value in enumerate(deal, start=1):
            sheet.cell(row=7 + index, column=column, value=value)
    ledger = book.create_sheet(contracting.SHEET_LEDGER)
    for index, row in enumerate(LEDGER, start=1):
        for column, value in enumerate(row, start=1):
            ledger.cell(row=index, column=column, value=value)
    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def _summary():
    data = _book()
    return contracting.summarise(contracting.read_contracts(data),
                                 contracting.read_ledger(data))


def test_the_file_names_its_own_project() -> None:
    """Привязка к проекту берётся из листа, а не из выбора в кабинете."""
    got = _summary()
    assert got["project"] == "Тестовый ЖК"
    assert got["missing"] == []


def test_the_commission_share_is_measured_on_its_own_deals() -> None:
    """«Процент от наполнения» по всему проекту делит на чужие сделки.

    На живом файле это давало 4,50% вместо 7,32%: в знаменатель попадал эскроу
    прямых продаж, где комиссии нет вовсе.
    """
    brokers = _summary()["brokers"]
    assert round(brokers["amount"]) == 28_000_000
    assert round(brokers["escrow"]) == 10_000_000
    assert round(brokers["fee_of_sales"], 4) == round(1_200_000 / 28_000_000, 4)
    assert round(brokers["fee_of_escrow"], 4) == round(1_200_000 / 10_000_000, 4)
    assert brokers["fee_of_escrow"] > brokers["fee_of_sales"]


def test_an_empty_commission_is_not_a_free_broker() -> None:
    """Ноль вознаграждения при названном брокере — «не заполнено»."""
    channels = {item["channel"]: item for item in _summary()["by_channel"]}
    assert channels["ТИХИЙ БРОКЕР ООО"]["fee_unknown"] is True
    assert channels["напрямую"]["fee_unknown"] is False


def test_a_reversal_is_a_termination_with_its_own_return() -> None:
    """Сторно — расторжение, а возврат эскроу — отдельное движение.

    Складывать приход и возврат в сальдо нельзя: 50 − 49,8 даёт остаток, а не
    ноль, и вернувшиеся 49,8 млн ₽ обязаны быть видны.
    """
    terminated = _summary()["terminated"]
    assert len(terminated) == 1
    only = terminated[0]
    assert only["contract"] == "1-2-9/ГР"
    assert only["on"] == "2026-02-12"
    assert round(only["escrow_paid"]) == 50_000_000
    assert round(only["escrow_returned"]) == 49_800_000


def test_a_deal_without_a_quarter_lands_by_its_contract_date() -> None:
    """У части строк колонка квартала пуста — терять их в «прочем» нельзя."""
    months = {item["month"] for item in _summary()["dynamics"]}
    assert months == {"2025-08", "2025-09"}


def test_the_payment_variant_keeps_what_it_did_not_recognise() -> None:
    """Свободный текст раскладывается правилом, неопознанное остаётся собой."""
    assert contracting.payment_variant("Рассрочка 20% ПВ и далее по 200 000 в месяц") == "рассрочка"
    assert contracting.payment_variant("Ипотека СПБ Банка") == "ипотека"
    assert contracting.payment_variant("1.0") == "100% оплата"
    assert contracting.payment_variant("5 млн на эскроу") == "5 млн на эскроу"
    assert contracting.payment_variant("") == "не указан"


def test_the_buyer_name_never_leaves_the_reader() -> None:
    """Свод агрегатный, а вопрос Платону уходит внешнему поставщику модели.

    Остаётся признак «юрлицо»: оптовая сделка тянет среднюю цену и должна быть
    видна, а фамилия для этого не нужна.
    """
    data = _book()
    rows = contracting.read_contracts(data)["rows"]
    assert all("buyer" not in row for row in rows)
    assert [row["company_buyer"] for row in rows] == [False, False, True]
    assert _summary()["company_buyers"] == 1


# --- итог по каналам и структура «свой / чужие» ------------------------------
# Таблица каналов перечисляла брокеров по одному, а сложить их было негде: сумма
# комиссий и доля своего канала считались глазами (владелец, 26.08.2026).
# Считает их тот же `_totals`, что и остальное: сложенная на экране колонка —
# это второй счёт той же величины, и однажды две суммы разойдутся, обе выглядя
# верными.


def test_the_channel_cost_is_summed_by_the_server() -> None:
    got = _summary()
    for key in ("brokers", "own_sales", "total"):
        block = got[key]
        assert "cost" in block, f"{key}: полная стоимость канала считается сервером"
        assert block["cost"] == pytest.approx(block["broker_fee"] + block["sales_bonus"])


def test_brokers_and_own_desk_add_up_to_the_project() -> None:
    got = _summary()
    total, brokers, own = got["total"], got["brokers"], got["own_sales"]
    assert brokers["contracts"] + own["contracts"] == total["contracts"]
    assert brokers["amount"] + own["amount"] == pytest.approx(total["amount"])
    assert brokers["cost"] + own["cost"] == pytest.approx(total["cost"])


def test_the_screen_shows_the_totals_it_did_not_compute() -> None:
    from market_search.cabinet import cabinet_page
    page = cabinet_page()
    assert "Итого брокеры" in page and "Итого свой отдел" in page and "Всего по проекту" in page
    assert "d.brokers.cost_of_sales" in page, "доля берётся у сервера, а не считается тут"
    assert "Свой канал против чужих" in page
    assert "function salesOwnVsBrokers(" in page


# --- планы: наша финмодель и модель банка ------------------------------------
# Свод отвечал на «что продали». Без второй половины — «сколько собирались» —
# он не говорит, идём мы по плану или отстаём (владелец, 26.08.2026). Оба плана
# лежат в той же выгрузке ЦФ, поэтому читаются тем же вызовом: просить загрузить
# один файл дважды значит однажды получить два разных файла.


def test_the_fm_plan_reads_plan_and_fact_by_month() -> None:
    data = _book()
    try:
        got = contracting.read_fm_plan(data)
    except KeyError:
        pytest.skip("в тестовой книге нет листа финмодели")
    assert got["months"], "месяцы берутся из строки дат, а не из счёта колонок"
    assert set(got) >= {"plan", "fact", "sheet"}


def test_the_bank_plan_stays_quarterly() -> None:
    """Раскладывать квартал по месяцам можно тремя способами — все наши."""
    source = (Path(__file__).resolve().parent.parent / "market_search" / "contracting.py").read_text()
    body = source[source.index("def read_bank_plan("):]
    body = body[:body.index("\n\ndef ")]
    assert "quarters" in body
    assert "month" not in body.split('"""')[2], "квартал не приводится к месяцам"


def test_a_missing_plan_is_a_reason_not_a_crash() -> None:
    """У выгрузки без листа планов есть контрактация, и это законный свод."""
    api = (Path(__file__).resolve().parent.parent / "market_search" / "api.py").read_text()
    body = api[api.index("async def cabinet_contracting("):]
    body = body[:body.index("\n    @app.post")]
    assert "read_fm_plan" in body and "read_bank_plan" in body
    assert "missing" in body, "не прочиталось — причина рядом, а не пятисотка"


def test_the_thousands_are_converted_once() -> None:
    """Лист считает в тысячах, свод — в рублях: две единицы под одним именем."""
    source = (Path(__file__).resolve().parent.parent / "market_search" / "contracting.py").read_text()
    assert "_FM_THOUSANDS" in source
    assert "value *= 1000.0" in source


def test_the_screen_says_the_plan_column_carries_fact() -> None:
    """Совпадение план-факт на прошедших месяцах — перенос, а не попадание."""
    from market_search.cabinet import cabinet_page
    page = cabinet_page()
    assert "Факт против нашей финмодели" in page
    assert "заполнена фактом" in page
    assert "Факт против плана банка" in page


def test_a_wide_table_scrolls_inside_its_own_frame() -> None:
    """Семь колонок каналов и девять кварталов банка не должны рвать карточку.

    У таблиц кабинета не было своей прокрутки: широкая таблица растягивала
    страницу за край экрана (владелец, 26.08.2026).
    """
    from market_search.cabinet import cabinet_page
    page = cabinet_page()
    assert ".tablescroll{overflow-x:auto" in page
    assert "'<div class=\"tablescroll\">'+html" in page, "таблица заворачивается в рамку"
    # Первая колонка переносится: имена брокеров длинные и в одну строку
    # выдавливают все числа за край.
    assert ".tablescroll td:first-child" in page and "white-space:normal" in page


# --- план показывается графиком, а не таблицей --------------------------------
# Таблицы с планом у владельца и так есть в книге; на экране нужно одно —
# видно ли расхождение и куда оно растёт (владелец, 26.08.2026).


def test_the_fact_is_summed_to_quarters_by_the_server() -> None:
    """План банка квартальный: сравнивать его с месяцами значит сравнивать разное."""
    got = _summary()
    assert got["by_quarter"], "факт по кварталам считает сервер, а не экран"
    total = sum(q["amount"] for q in got["by_quarter"])
    assert total == pytest.approx(got["total"]["amount"])


def test_a_quarter_is_named_as_in_the_bank_book() -> None:
    assert contracting.quarter_of("2026-07") == "2026 Q3"
    assert contracting.quarter_of("2026-01") == "2026 Q1"
    assert contracting.quarter_of("2026-12") == "2026 Q4"
    # Неразобранный месяц не превращается в выдуманный квартал.
    assert contracting.quarter_of("мусор") == "мусор"


def test_the_bank_revenue_is_summed_on_the_server_and_names_its_rows() -> None:
    data = _book()
    try:
        bank = contracting.read_bank_plan(data)
    except KeyError:
        pytest.skip("в тестовой книге нет листа банка")
    assert "revenue_by_quarter" in bank
    assert "revenue_rows" in bank, "какие строки сложены — часть ответа"


def test_the_plan_is_drawn_not_tabulated() -> None:
    from market_search.cabinet import cabinet_page
    page = cabinet_page()
    assert "function factVsPlanChart(" in page
    assert "столбики — факт, линия —" in page
    # Таблиц плана на экране больше нет: они есть в книге у владельца.
    assert "['Месяц','Факт, млн ₽','План ФМ, млн ₽','Отклонение']" not in page
    assert "Факт против плана банка" in page


def test_the_chart_does_no_economics() -> None:
    """На экране только геометрия: высота столбика и координата точки."""
    from market_search.cabinet import cabinet_page
    page = cabinet_page()
    start = page.index("function factVsPlanChart(")
    depth = 0
    for position in range(page.index("{", start), len(page)):
        if page[position] == "{":
            depth += 1
        elif page[position] == "}":
            depth -= 1
            if depth == 0:
                break
    body = page[start:position + 1]
    assert "/1e6" in body, "перевод в миллионы — оформление"
    for forbidden in ("*100", "/r.plan", "fact/plan"):
        assert forbidden not in body, f"экран считает экономику: {forbidden}"


def test_the_chart_does_not_shadow_the_market_one() -> None:
    """В кабинете уже есть planChart для плана продаж рынка.

    Вторая функция с тем же именем молча затирает первую: обе живут в одном
    скрипте страницы, и последняя выигрывает. Свой график называется своим
    именем.
    """
    from market_search.cabinet import cabinet_page
    page = cabinet_page()
    assert page.count("function planChart(") == 1
    assert page.count("function factVsPlanChart(") == 1
