"""График производства работ сшивается с деньгами и помнит базовый план.

Деньги говорят «отстаём на 275,4 млн ₽», и на этом замолкают: по ним не видно,
фасады это или инженерка. ГПР говорит «где»: у каждой работы код РСС — тот же
ключ, которым разнесены платежи и акты.

Файлов при этом два, и ни один сам по себе плана-факта не даёт. Очищенный ГПР
несёт код РСС, но не знает ни фактических дат, ни базового плана; выгрузка
планировщика знает и то и другое, но кода РСС в ней нет вовсе. Сшиваются они
по `Ид` — на Кутузове сходится 632 из 632.

Отставание считается от базового плана, а не от текущего: текущие сроки двигают
так же, как переписывают акты. На Кутузове верхний срок держится (+4 дня), а
под ним 387 работ из 454 уехали вправо с медианой 66 дней — по текущему графику
проект «в срок», по базовому нет, и видно это только сравнением двух колонок.
"""

from __future__ import annotations

import datetime
import io

import pytest

from openpyxl import Workbook

import developaid_actuals as actuals


def _cleaned_gpr(rows):
    """Очищенный ГПР в миниатюре: заголовок в 4-й строке, работы — с 5-й."""
    book = Workbook()
    sheet = book.active
    sheet.title = "ГПР"
    sheet.cell(row=1, column=1, value="Кутузов Сити — очищенный график")
    sheet.cell(row=2, column=1, value="Дата среза")
    header = ["ID", "WBS", "Раздел", "Объект", "Наименование работ",
              "Тип строки", "% выполнения", "Начало", "Окончание", "Статус",
              "Длительность, р.д.", "Предшественники", "Связанный тендер",
              "Окончание тендера", "Резерв", "Увязка", "Код РСС",
              "Статья РСС", "Основание привязки"]
    for column, title in enumerate(header, 1):
        sheet.cell(row=4, column=column, value=title)
    for offset, row in enumerate(rows):
        for column, value in enumerate(row, 1):
            sheet.cell(row=5 + offset, column=column, value=value)
    blob = io.BytesIO()
    book.save(blob)
    blob.seek(0)
    return blob


def _pm_export(rows):
    """Выгрузка планировщика: 84 колонки, из которых важны немногие."""
    book = Workbook()
    sheet = book.active
    sheet.title = "Таблица_задач1"
    titles = {1: "Ид", 3: "Название_задачи", 6: "Уровень_структуры",
              9: "Начало", 10: "Окончание", 19: "Фактическое_начало",
              20: "Фактическое_окончание", 21: "Базовое_начало",
              22: "Базовое_окончание", 36: "СДР", 39: "Суммарная_задача",
              49: "Статус"}
    for column, title in titles.items():
        sheet.cell(row=1, column=column, value=title)
    for offset, row in enumerate(rows):
        for column, value in row.items():
            sheet.cell(row=2 + offset, column=column, value=value)
    blob = io.BytesIO()
    book.save(blob)
    blob.seek(0)
    return blob


def _work(task_id, code, *, kind="Работа", status="Будущая задача",
          start="2026-07-01", finish="2026-09-30", name="Работа"):
    return [task_id, f"1.16.{task_id}", "СМР", "Корпус 1", name, kind, 0.0,
            start, finish, status, 60, "", "", "", "", "", code, "Статья",
            "Наследовано"]


def _pm_task(task_id, *, wbs="", start="01 Июль 2026 9:00",
             finish="30 Сентябрь 2026 18:00", base_start=None, base_finish=None,
             act_start="НД", act_finish="НД", level=3, summary="Нет",
             name="Работа"):
    return {1: task_id, 3: name, 6: level, 9: start, 10: finish,
            19: act_start, 20: act_finish,
            21: base_start or "НД", 22: base_finish or "НД",
            36: wbs or f"1.16.{task_id}", 39: summary, 49: ""}


def test_summary_rows_are_not_works():
    """«Сводная» — свёртка ветки: сложить её с подчинёнными — посчитать дважды."""
    schedule = actuals.read_schedule(_cleaned_gpr([
        _work("1", "2.2.1.4.", kind="Сводная"),
        _work("2", "2.2.1.4."),
    ]))

    assert len(schedule["rows"]) == 2
    assert len(schedule["works"]) == 1


def test_overdue_is_read_from_the_status():
    schedule = actuals.read_schedule(_cleaned_gpr([
        _work("1", "2.2.1.4.", status="Просрочено"),
        _work("2", "2.2.1.4.", status="Просрочено, в работе"),
        _work("3", "2.2.1.4.", status="В работе"),
    ]))

    assert len(schedule["overdue"]) == 2


def test_a_work_without_a_code_is_counted_not_lost():
    """Безкодовая работа не пришивается к деньгам — но и не исчезает молча."""
    schedule = actuals.read_schedule(_cleaned_gpr([
        _work("1", ""), _work("2", "2.2.1.4."),
    ]))

    assert len(schedule["without_code"]) == 1


def test_the_trailing_dot_of_the_code_is_stripped():
    """ГПР пишет «2.2.1.4.», реестры — «2.2.1.4»: ключ должен быть один."""
    schedule = actuals.read_schedule(_cleaned_gpr([_work("1", "2.2.1.4.")]))

    assert schedule["works"][0]["estimate_code"] == "2.2.1.4"


def test_money_gap_lands_on_the_code():
    """План из производственной программы, факт из актов — по коду РСС."""
    schedule = actuals.read_schedule(_cleaned_gpr([
        _work("1", "2.2.2.6.", status="Просрочено", name="Фасады юг"),
        _work("2", "2.2.1.4."),
    ]))
    june, july = datetime.date(2026, 6, 1), datetime.date(2026, 7, 1)
    programme = {
        "by_code": {"2.2.2.6": {july: 200e6}, "2.2.1.4": {july: 50e6}},
        "leaves": {"2.2.2.6", "2.2.1.4"},
        "months": [june, july],
    }
    works = {"rows": [
        {"code": "2.2.1.4", "amount": 45e6, "construction": True,
         "date": datetime.date(2026, 7, 10)},
        {"code": "2.2.2.6", "amount": 10e6, "construction": True,
         "date": datetime.date(2026, 5, 10)},  # до программы — не в счёт
    ]}

    against = actuals.schedule_against_money(
        schedule, programme, works, "2026-08-01")
    by_code = {row["code"]: row for row in against["by_code"]}

    assert by_code["2.2.2.6"]["gap"] == pytest.approx(-200e6)
    assert by_code["2.2.2.6"]["overdue"] == 1
    assert by_code["2.2.2.6"]["overdue_names"] == ["Фасады юг"]
    assert by_code["2.2.1.4"]["gap"] == pytest.approx(-5e6)


def test_pm_dates_come_in_two_shapes():
    """Плановые даты — словами, фактические — с днём недели, «НД» — нет даты."""
    pm = actuals.read_pm_schedule(_pm_export([_pm_task(
        "1", start="07 Июль 2020 9:00", finish="30 Сентябрь 2027 18:00",
        act_start="Вт 07.07.20", act_finish="НД")]))
    task = pm["rows"][0]

    assert task["start"] == datetime.date(2020, 7, 7)
    assert task["finish"] == datetime.date(2027, 9, 30)
    assert task["actual_start"] == datetime.date(2020, 7, 7)
    assert task["actual_finish"] is None


def test_merge_brings_baseline_and_fact_to_the_gpr():
    """У очищенного ГПР нет ни факта, ни базового плана — их приносит сшивка."""
    schedule = actuals.read_schedule(_cleaned_gpr([_work("1", "2.2.1.4.")]))
    pm = actuals.read_pm_schedule(_pm_export([_pm_task(
        "1", base_start="01 Июнь 2026 9:00", base_finish="31 Июль 2026 18:00",
        act_start="Ср 01.07.26")]))

    merged = actuals.merge_schedule(schedule, pm)
    work = merged["works"][0]

    assert merged["baseline"]["matched"] == 1
    assert work["baseline_finish"] == datetime.date(2026, 7, 31)
    assert work["actual_start"] == datetime.date(2026, 7, 1)


def test_a_wbs_conflict_is_reported_not_merged():
    """Разные WBS у одного Ид — файлы с разных срезов, чужие даты не берутся."""
    schedule = actuals.read_schedule(_cleaned_gpr([_work("1", "2.2.1.4.")]))
    pm = actuals.read_pm_schedule(_pm_export([_pm_task(
        "1", wbs="1.15.9", base_finish="31 Июль 2026 18:00")]))

    merged = actuals.merge_schedule(schedule, pm)

    assert merged["baseline"]["matched"] == 0
    assert merged["baseline"]["conflicts"][0]["id"] == "1"
    assert "baseline_finish" not in merged["works"][0]


def test_slip_is_measured_from_the_baseline_not_the_current_plan():
    """Текущий план двигают вместе с работой — от него отставания не увидеть."""
    schedule = actuals.read_schedule(_cleaned_gpr([
        _work("1", "2.2.1.4.", finish="2026-09-30")]))
    pm = actuals.read_pm_schedule(_pm_export([_pm_task(
        "1", base_finish="31 Июль 2026 18:00")]))
    actuals.merge_schedule(schedule, pm)

    gantt = actuals.gantt(schedule, "2026-08-01")

    assert gantt["bars"][0]["slip_days"] == 61
    assert gantt["slipped"] == 1


def test_a_running_work_fact_bar_ends_at_the_cut():
    """У идущей работы факт кончается срезом, а не плановым окончанием."""
    schedule = actuals.read_schedule(_cleaned_gpr([_work("1", "2.2.1.4.")]))
    pm = actuals.read_pm_schedule(_pm_export([_pm_task(
        "1", act_start="Ср 01.07.26")]))
    actuals.merge_schedule(schedule, pm)

    bar = actuals.gantt(schedule, "2026-08-15")["bars"][0]

    assert bar["running"]
    assert bar["fact"] == [datetime.date(2026, 7, 1), datetime.date(2026, 8, 15)]


def test_a_finished_work_fact_bar_ends_at_its_actual_finish():
    schedule = actuals.read_schedule(_cleaned_gpr([_work("1", "2.2.1.4.")]))
    pm = actuals.read_pm_schedule(_pm_export([_pm_task(
        "1", act_start="Ср 01.07.26", act_finish="Пт 24.07.26")]))
    actuals.merge_schedule(schedule, pm)

    bar = actuals.gantt(schedule, "2026-08-15")["bars"][0]

    assert bar["done"]
    assert bar["fact"][1] == datetime.date(2026, 7, 24)


def test_the_deadline_is_the_latest_leaf_not_the_root():
    """Базовый план корня — свёртка чужой истории: по нему выходит «опережение».

    На Кутузове корень унаследован от другого проекта с базовым окончанием в
    ноябре 2028-го — по нему проект «идёт с опережением на 401 день». Срок
    проекта — самый поздний лист со своим базовым планом.
    """
    schedule = actuals.read_schedule(_cleaned_gpr([_work("1", "2.2.1.4.")]))
    pm = actuals.read_pm_schedule(_pm_export([
        _pm_task("0", name="График", level=0, summary="Да",
                 finish="30 Сентябрь 2027 18:00",
                 base_finish="04 Ноябрь 2028 18:00"),
        _pm_task("9", name="РнВ получен", level=2,
                 finish="30 Сентябрь 2027 18:00",
                 base_finish="26 Сентябрь 2027 18:00"),
        _pm_task("1", base_finish="31 Июль 2026 18:00"),
    ]))
    actuals.merge_schedule(schedule, pm)

    deadline = actuals.schedule_deadline(schedule, pm)

    assert deadline["name"] == "РнВ получен"
    assert deadline["slip_days"] == 4
