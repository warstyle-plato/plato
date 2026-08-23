"""«Управление проектом» идёт за одними и теми же статьями во всех трёх местах.

Своего окна у статьи нет: штаб загружен тем, чем вызван, и его доля месяца
равна доле этих расходов. Список жил в трёх экземплярах — в движке, в книге v2
и сплошным диапазоном строк в шаблоне v4, — и третий включал «Сдачу и ввод»,
статью последних трёх месяцев. Книга смещала управление к сдаче: помесячный
CAPEX расходился с движком на сотую долю процента при совпадающем итоге.

Сотая доля процента оказалась дорогой. Покрытие эскроу легло ровно на ступень
лестницы ставки ПФ: 1,10002× в книге против 1,09993× у движка. Один месяц пошёл
по 3,47% вместо 1,75%, и стоимость финансирования разошлась на 28,9 млн ₽ —
при совпадающих до копейки выручке, CAPEX, EBITDA, пике БРИДЖа и пике ПФ. Ни
одно из этих чисел ошибкой не выглядело.

Запуск: python3 -m pytest tests -q
"""

from __future__ import annotations

import io
import re
import sys
import zipfile
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "tests"))

import main_legacy as core  # noqa: E402

openpyxl = pytest.importorskip("openpyxl")

BASE = {**core.DEFAULT_INPUTS, "purchase_price_mln": 4300, "land_rights_cost_mln": 0,
        "social_compensation_mln": 0, "ird_months": 1, "apartment_price_th": 650,
        "commercial_price_th": 650, "parking_price_th": 5000,
        "main_above_th_per_sqm": 190, "main_under_th_per_sqm": 120}


def _tep() -> dict[str, dict[str, float]]:
    tep = {key: dict(value) for key, value in core.TEP_DEFAULT.items()}
    tep["apartments"].update({"gns": 43201.96, "saleable": 29308.89, "units": 355,
                              "total_area": 43201.96, "useful": 29308.89})
    tep["ground_commercial"].update({"gns": 2634.85, "saleable": 2451.09,
                                     "total_area": 2634.85, "useful": 2451.09})
    tep["underground_parking"].update({"gns": 6475.0, "units": 185, "saleable": 0})
    for key in ("standalone_retail", "offices", "above_parking", "storage",
                "kindergarten", "school", "clinic"):
        tep[key].update({"gns": 0, "saleable": 0, "units": 0, "total_area": 0,
                         "useful": 0, "transfer": 0})
    return tep


def test_the_engine_has_exactly_one_declaration():
    """Копия списка в коде движка — это второе мнение об одном профиле."""
    source = Path(core.__file__).read_text(encoding="utf-8")
    assert source.count("MANAGEMENT_PROFILE_ARTICLES") >= 3, "список не используется"
    # Ни один из трёх потребителей не перечисляет статьи сам. Отличительный
    # признак профиля — «Благоустройство» и «Содержание стройплощадки» подряд:
    # сводки строительного CAPEX между ними держат «Сдачу и ввод», и их этот
    # шаблон не задевает.
    literal = re.compile(r'"landscaping",\s*"site_maintenance"')
    assert len(literal.findall(source)) == 1, "список статей профиля перечислен второй раз"


def test_the_commissioning_is_not_in_the_profile():
    """«Сдача и ввод» стоит в последних трёх месяцах — и в профиле её нет.

    Проверка именно на этой статье: разошлись книга и движок ровно из-за неё.
    """
    assert "commissioning" not in core.MANAGEMENT_PROFILE_ARTICLES
    assert "commissioning" in core._V4_CAPEX_ARTICLE_ROW, "строка шаблона потеряна"


def test_the_template_range_is_brought_to_the_engine_list():
    """Строки профиля в шаблоне идут подряд, и лишняя лежит внутри диапазона —
    значит сплошной SUM его не выражает, и формулу надо переписать."""
    first, last, excluded = core._v4_management_profile_ranges()
    assert excluded == [core._V4_CAPEX_ARTICLE_ROW["commissioning"]]
    assert first == core._V4_CAPEX_ARTICLE_ROW["ird"]
    assert last == core._V4_CAPEX_ARTICLE_ROW["site_maintenance"]


def test_an_unrecognised_formula_goes_to_missing_not_silently_through():
    """Шаблон, у которого формула другая, — не повод посчитать по-своему."""
    missing: list[str] = []
    core._v4_apply_management_profile("<x:f>ничего похожего</x:f>", missing)
    assert len(missing) == core._V4_CAPEX_PHASES
    assert all("профиль управления" in item for item in missing)


def test_every_queue_block_of_the_workbook_is_rewritten():
    content, _, meta = core.build_project_workbook(
        dict(BASE), _tep(), [], {}, project_name="Профиль")
    assert not [item for item in (meta.get("missing") or [])
                if "профиль управления" in str(item)]
    book = openpyxl.load_workbook(io.BytesIO(content))
    sheet = book["CAPEX"]
    row = core._V4_CAPEX_ARTICLE_ROW["project_management"]
    skip = core._V4_CAPEX_ARTICLE_ROW["commissioning"]
    for phase in range(core._V4_CAPEX_PHASES):
        base = core._V4_CAPEX_BLOCK_STRIDE * phase
        formula = str(sheet.cell(row=row + base, column=6).value)
        assert f"-F{skip + base}" in formula, formula
        assert f"$D${skip + base}:$DS${skip + base}" in formula, formula


def test_the_monthly_capex_of_the_book_matches_the_engine():
    """Итог сходился и раньше — расходился помесячный профиль, а по нему
    считаются покрытие, ступень ставки и проценты."""
    from openpyxl.utils import get_column_letter
    from xlsx_eval import Evaluator

    sys.setrecursionlimit(400000)
    content, _, _ = core.build_project_workbook(
        dict(BASE), _tep(), [], {}, project_name="Профиль")
    evaluator = Evaluator(openpyxl.load_workbook(io.BytesIO(content), data_only=False))
    result = core.calculate(core.CalcRequest(inputs=dict(BASE), tep=_tep(), rates=[]))
    engine = result["monthly"]["capex_total"]
    for index in range(len(engine)):
        column = get_column_letter(4 + index)
        book = evaluator.cell("CF_1", f"{column}17")
        assert book == pytest.approx(float(engine[index]) / 1e6, abs=0.01), \
            f"месяц {index}: книга {book}, движок {float(engine[index]) / 1e6}"
