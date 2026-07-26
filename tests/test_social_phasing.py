"""Тесты поздней раскладки социальных объектов по очередям.

Первая очередь несёт сети и подготовительный период, поэтому социальные
объекты по умолчанию уезжают вправо: школа в последнюю очередь, ДОУ во вторую,
поликлиника ближе к концу. Запуск: python3 -m pytest tests -q
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import main as _wrapper  # noqa: E402

main = _wrapper.core

SOCIAL_INPUTS = {
    "social_mode": "Строительство",
    "kindergarten_places": 465,
    "school_places": 975,
    "clinic_capacity": 127,
}


def phasing(count: int) -> dict:
    return {
        "enabled": True, "user_enabled": True, "phase_count": count,
        "target_size_sqm": 70000, "phase_gap_months": 12,
        "cost_inflation_pct": 8, "sales_price_inflation_pct": 8,
    }


def bundle(count: int, extra: dict | None = None, registry: list | None = None) -> dict:
    config = phasing(count)
    if registry is not None:
        config["social_objects"] = registry
    return main.calculate_phased(main.PhasedCalcRequest(
        inputs={**main.DEFAULT_INPUTS, **SOCIAL_INPUTS, **(extra or {})},
        tep=main.TEP_DEFAULT, rates=[], phasing=config,
    ))


def phases_of(result: dict) -> dict[str, int]:
    return {item["type"]: item["phase"] for item in result["social_allocation"]}


# --- раскладка по умолчанию -------------------------------------------------

@pytest.mark.parametrize("count, expected", [
    (2, {"kindergarten": 2, "school": 2, "clinic": 2}),
    (3, {"kindergarten": 2, "school": 3, "clinic": 2}),
    (4, {"kindergarten": 2, "school": 4, "clinic": 3}),
    (5, {"kindergarten": 2, "school": 5, "clinic": 4}),
])
def test_default_allocation_is_late(count, expected):
    assert phases_of(bundle(count)) == expected


@pytest.mark.parametrize("count", [2, 3, 4, 5])
def test_first_phase_carries_no_social_objects(count):
    assert all(item["phase"] > 1 for item in bundle(count)["social_allocation"])


@pytest.mark.parametrize("count", [3, 4, 5])
def test_school_goes_last(count):
    assert phases_of(bundle(count))["school"] == count


def test_reason_is_reported():
    for item in bundle(3)["social_allocation"]:
        assert "поздняя раскладка" in item["reason"]
        assert item["moved_earlier"] is False


# --- обязательства ----------------------------------------------------------

def test_obligation_pulls_the_object_earlier():
    result = bundle(3, {"school_not_later_than": 1})
    school = next(item for item in result["social_allocation"] if item["type"] == "school")
    assert school["phase"] == 1
    assert school["auto_phase"] == 3
    assert school["moved_earlier"] is True
    assert "не позже очереди 1" in school["reason"]


def test_obligation_later_than_default_does_not_move_the_object():
    result = bundle(4, {"kindergarten_not_later_than": 4})
    assert phases_of(result)["kindergarten"] == 2


def test_clinic_may_land_in_the_first_phase_by_obligation():
    result = bundle(4, {"clinic_not_later_than": 1})
    clinic = next(item for item in result["social_allocation"] if item["type"] == "clinic")
    assert clinic["phase"] == 1
    assert clinic["moved_earlier"] is True


# --- ручная раскладка -------------------------------------------------------

def test_manual_registry_is_respected():
    registry = [
        {"name": "ДОУ", "type": "kindergarten", "capacity": 465, "phase": 3},
        {"name": "СОШ", "type": "school", "capacity": 975, "phase": 1},
    ]
    result = bundle(3, registry=registry)
    allocation = {item["type"]: item for item in result["social_allocation"]}
    assert allocation["kindergarten"]["phase"] == 3
    assert allocation["school"]["phase"] == 1
    assert allocation["school"]["reason"] == "вручную"


def test_manual_phase_beyond_range_is_clamped():
    registry = [{"name": "СОШ", "type": "school", "capacity": 975, "phase": 9}]
    assert phases_of(bundle(3, registry=registry))["school"] == 3


# --- влияние на экономику ---------------------------------------------------

def test_social_cost_is_reported_per_phase():
    result = bundle(3)
    costs = {item["name"]: item["social_cost"] for item in result["comparison"]}
    assert costs["О1"] == 0
    assert costs["О2"] > 0 and costs["О3"] > 0
    total = sum(costs.values())
    expected = (
        465 * main.DEFAULT_INPUTS["kindergarten_cost_mln_per_place"]
        + 975 * main.DEFAULT_INPUTS["school_cost_mln_per_place"]
        + 127 * main.DEFAULT_INPUTS["clinic_cost_mln_per_unit"]
    ) * 1_000_000
    assert total > expected  # с учётом индексации себестоимости по очередям


def test_objects_are_listed_in_comparison():
    result = bundle(3)
    rows = {item["name"]: item["social_objects"] for item in result["comparison"]}
    assert rows["О1"] == []
    assert "СОШ" in rows["О3"]


def test_late_allocation_relieves_the_first_phase():
    late = bundle(3)
    early = bundle(3, {"kindergarten_not_later_than": 1, "school_not_later_than": 1,
                       "clinic_not_later_than": 1})
    first_late = next(item for item in late["comparison"] if item["name"] == "О1")
    first_early = next(item for item in early["comparison"] if item["name"] == "О1")
    assert first_late["social_cost"] == 0
    assert first_early["social_cost"] > 0
    # разгруженная первая очередь требует меньше пикового финансирования
    assert first_late["peak_pf"] < first_early["peak_pf"]


def test_money_compensation_mode_has_no_allocation():
    result = bundle(3, {"social_mode": "Денежная компенсация"})
    assert result["social_allocation"] == []


def test_single_phase_project_is_untouched():
    result = main.calculate_phased(main.PhasedCalcRequest(
        inputs={**main.DEFAULT_INPUTS, **SOCIAL_INPUTS}, tep=main.TEP_DEFAULT,
        rates=[], phasing={"enabled": False, "phase_count": 1},
    ))
    assert result["mode"] == "single"


# --- выгрузка ---------------------------------------------------------------

def test_comparison_sheet_shows_social_load():
    pytest.importorskip("openpyxl")
    import io
    import zipfile

    from openpyxl import load_workbook

    content, _ = main.build_model_archive(
        {**main.DEFAULT_INPUTS, **SOCIAL_INPUTS}, main.TEP_DEFAULT, [], phasing(3),
        project_name="Соц",
    )
    archive = zipfile.ZipFile(io.BytesIO(content))
    sheet = load_workbook(io.BytesIO(archive.read("00_Консолидация.xlsx")), data_only=True)["Сравнение очередей"]
    header = [cell.value for cell in sheet[4]]
    column = header.index("Социальная нагрузка, млн ₽") + 1
    objects_column = header.index("Социальные объекты") + 1
    assert sheet.cell(row=5, column=column).value == 0
    assert sheet.cell(row=5, column=objects_column).value == "—"
    assert sheet.cell(row=7, column=column).value > 0


if __name__ == "__main__":
    raise SystemExit(pytest.main([__file__, "-q"]))
