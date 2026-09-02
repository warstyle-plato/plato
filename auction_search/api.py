from __future__ import annotations

import logging
import os
import sys
import threading
import time
import io
from typing import Any
from urllib.parse import urlparse

from fastapi import FastAPI, HTTPException, Query, Request
from fastapi.concurrency import run_in_threadpool
from fastapi.responses import HTMLResponse, Response
from pydantic import BaseModel, Field
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from auction_search.adapters import (
    NistpAdapter,
    ETPGPBAdapter,
    ETPRFAdapter,
    InvestMoscowDiscoveryAdapter,
    LotOnlineAdapter,
    RoseltorgAdapter,
    SberbankASTAdapter,
)
from auction_search.adapters.torgi_gov import TorgiGovAdapter, trust_report as torgi_trust_report
from auction_search.adapters.etp_probe import (
    SLUGS as etp_slugs,
    probe as etp_probe,
    probe_browser_platform as etp_probe_browser,
)
from auction_search.adapters.fedresurs import (
    SEARCH_PAGE as FEDRESURS_SEARCH_PAGE,
    probe as fedresurs_probe, probe_browser as fedresurs_browser)
from auction_search.adapters.roseltorg_probe import (
    CATALOGUE_URL as ROSELTORG_SECTION_URL,
    probe as roseltorg_probe,
    probe_section_browser as roseltorg_section_browser,
)
from auction_search.bridge import auction_page_with_handoff, install_page_bridge
from auction_search.catalogue_quality import catalogue_quality
from auction_search import equity_stake
from auction_search.developaid_mapper import build_developaid_seed
from auction_search.documents import DocumentExtractionError
from auction_search.export_areas import export_areas
from auction_search.krt_pipeline import enrich_krt_from_official_documents
from auction_search.krt_ranking import (
    HEARTBEAT_SECONDS, NEW_FOR_SECONDS, KrtRanking, score_row)
from auction_search.krt_screening import build_krt_model_screening
from auction_search.models import LotKind
from auction_search.preset_mapper import build_project_preset
from auction_search.profile_fit import profile_fit
from auction_search.service import AuctionSearchService
from auction_search.ui import auctions_page
from market_search.krt_registry import CATALOGUE_URL, KrtRegistry
from market_search import cabinet as market_cabinet
from market_search.geocoder import GeocodingError
from market_search.http import RemoteServiceError
from market_search.subject import SubjectNotFound


logger = logging.getLogger(__name__)


class KrtRankingRequest(BaseModel):
    """Что считать: слаги отобранных площадок. Пусто — весь каталог."""

    slugs: list[str] = Field(default_factory=list, max_length=400)


class AuctionIngestRequest(BaseModel):
    url: str = Field(min_length=12, max_length=2000)
    enrich_krt_documents: bool = True
    include_raw: bool = False

class AuctionExportRequest(BaseModel):
    rows: list[dict[str, Any]] = Field(default_factory=list, max_length=2000)
    kind: str = Field(default="auctions", max_length=40)

class AuctionLotPointRequest(BaseModel):
    query: str = Field(min_length=3, max_length=500)

def _xlsx(rows: list[dict[str, Any]], kind: str = "auctions") -> bytes:
    if kind == "krt":
        columns = [
            ("name", "Проект КРТ", 48),
            ("okrug", "Округ", 10),
            ("district", "Район", 20),
            ("status", "Статус", 18),
            ("krt_area_ha", "Площадь территории, га", 20),
            ("total_gfa_sqm", "Общий объём строительства, м²", 24),
            ("housing_gfa_sqm", "Жильё, м²", 18),
            ("nonresidential_gfa_sqm", "Нежилое, м²", 18),
            ("business_gfa_sqm", "Общественно-деловое, м²", 22),
            ("jobs", "Рабочие места", 17),
            ("score", "Оценка Платона, балл", 19),
            ("traffic_light", "Светофор модели", 25),
            ("saleable_sqm", "Продаваемая площадь, м²", 22),
            ("entry_capacity_rub_per_sqm", "Потолок цены входа, ₽/м² продаваемой", 27),
            ("entry_capacity_mln", "Потолок цены входа всего, млн ₽", 25),
            ("project_llcr_x", "LLCR проекта, x", 17),
            ("weakest_phase_llcr_x", "LLCR слабейшей очереди, x", 22),
            ("margin_pct", "Маржа до неизвестных обязательств, %", 25),
            # Чьё это КРТ и не занято ли оно. Ячейка несёт цитату источника или
            # словами говорит, чего не хватает: пустая клетка читалась бы как
            # «нет», а это «не нашли» или «не читали».
            ("krt_kind", "Вид КРТ", 26),
            ("krt_city_needs", "Городские нужды — цитата", 44),
            ("krt_operator", "Оператор / застройщик", 34),
            ("url", "Источник", 42),
        ]
        obligation_columns = [
            ("name", "Проект КРТ", 48),
            ("okrug", "Округ", 10),
            ("district", "Район", 20),
            ("status", "Статус", 18),
            ("demolition_objects", "Снос, объектов", 16),
            ("demolition_area_sqm", "Снос, известная площадь, м²", 23),
            ("conditional_objects", "Снос/реконструкция, объектов", 23),
            ("conditional_area_sqm", "Снос/реконструкция, известная площадь, м²", 30),
            ("reconstruction_objects", "Реконструкция, объектов", 21),
            ("reconstruction_area_sqm", "Реконструкция, известная площадь, м²", 28),
            ("preservation_objects", "Сохранение, объектов", 20),
            ("preservation_area_sqm", "Сохранение, известная площадь, м²", 27),
            ("resettlement_mentions", "Расселение/изъятие, упоминаний", 25),
            ("url", "Источник", 42),
        ]
        numeric_keys = {
            "krt_area_ha", "total_gfa_sqm", "housing_gfa_sqm",
            "nonresidential_gfa_sqm", "business_gfa_sqm", "jobs", "score",
            "saleable_sqm", "entry_capacity_rub_per_sqm", "entry_capacity_mln",
            "project_llcr_x", "weakest_phase_llcr_x", "margin_pct",
            "demolition_objects", "demolition_area_sqm", "conditional_objects",
            "conditional_area_sqm", "reconstruction_objects", "reconstruction_area_sqm",
            "preservation_objects", "preservation_area_sqm", "resettlement_mentions",
        }
    else:
        columns = [
            ("section", "Раздел", 11),
            ("name", "Название", 48),
            ("okrug", "Округ", 10),
            ("district", "Район", 18),
            ("address", "Адрес", 34),
            ("cadastre", "Кадастровые номера", 28),
            ("type", "Тип", 18),
            ("land_area_sqm", "Площадь участка, м²", 20),
            ("building_area_sqm", "Площадь здания/ОКС, м²", 23),
            ("krt_area_ha", "Площадь КРТ, га", 18),
            ("total_gfa_sqm", "Общий объём, м²", 18),
            ("housing_gfa_sqm", "Жильё, м²", 16),
            ("price", "Цена, ₽", 18),
            ("score", "Оценка Платона", 17),
            ("status", "Статус", 22),
            ("url", "Источник", 42),
        ]
        obligation_columns = []
        numeric_keys = {
            "land_area_sqm", "building_area_sqm", "krt_area_ha", "total_gfa_sqm",
            "housing_gfa_sqm", "price", "score",
        }
    def number(value: Any) -> float | int | None:
        if value in (None, ""):
            return None
        try:
            parsed = float(value)
            return int(parsed) if parsed.is_integer() else parsed
        except (TypeError, ValueError):
            return None

    formats = {
        "land_area_sqm": '#,##0.00', "building_area_sqm": '#,##0.00',
        "krt_area_ha": '0.00', "total_gfa_sqm": '#,##0', "housing_gfa_sqm": '#,##0',
        "nonresidential_gfa_sqm": '#,##0', "business_gfa_sqm": '#,##0',
        "jobs": '#,##0', "price": '#,##0" ₽"', "score": '0',
        "saleable_sqm": '#,##0', "entry_capacity_rub_per_sqm": '#,##0" ₽/м²"',
        "entry_capacity_mln": '#,##0.0" млн ₽"', "project_llcr_x": '0.00"x"',
        "weakest_phase_llcr_x": '0.00"x"', "margin_pct": '0.0"%"',
        "demolition_objects": '0', "demolition_area_sqm": '#,##0.0',
        "conditional_objects": '0', "conditional_area_sqm": '#,##0.0',
        "reconstruction_objects": '0', "reconstruction_area_sqm": '#,##0.0',
        "preservation_objects": '0', "preservation_area_sqm": '#,##0.0',
        "resettlement_mentions": '0',
    }
    wb = Workbook()

    def fill_sheet(ws, sheet_columns, title: str, table_name: str) -> None:
        ws.title = title
        keys = tuple(column[0] for column in sheet_columns)
        ws.append([column[1] for column in sheet_columns])
        for source_row in rows[:2000]:
            row = dict(source_row)
            if row.get("section") == "Торги":
                areas = export_areas(row)
                row["land_area_sqm"] = areas.land_area_sqm
                row["building_area_sqm"] = areas.building_area_sqm
            values = [
                number(row.get(key)) if key in numeric_keys else (row.get(key) or "")
                for key in keys
            ]
            ws.append(values)

        ws.freeze_panes = "A2"
        last_column = get_column_letter(len(sheet_columns))
        ws.auto_filter.ref = f"A1:{last_column}{max(1, ws.max_row)}"
        ws.sheet_view.showGridLines = False
        header_fill = PatternFill("solid", fgColor="171717")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 38
        for index, (_, _, width) in enumerate(sheet_columns, start=1):
            ws.column_dimensions[get_column_letter(index)].width = width
        wrap_columns = {
            index for index, (key, _, _) in enumerate(sheet_columns, start=1)
            if key in {"name", "address", "cadastre", "status", "traffic_light", "url"}
        }
        numeric_columns = {
            index for index, (key, _, _) in enumerate(sheet_columns, start=1)
            if key in numeric_keys
        }
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(
                    horizontal="right" if cell.column in numeric_columns else "left",
                    vertical="top", wrap_text=cell.column in wrap_columns)
        for index, (key, _, _) in enumerate(sheet_columns, start=1):
            if key in formats:
                for cells in ws.iter_cols(min_col=index, max_col=index, min_row=2):
                    for item in cells:
                        item.number_format = formats[key]
        url_column = keys.index("url") + 1
        for row_number in range(2, ws.max_row + 1):
            cell = ws.cell(row_number, url_column)
            if cell.value:
                cell.hyperlink = str(cell.value)
                cell.style = "Hyperlink"
        if ws.max_row >= 2:
            table = Table(displayName=table_name, ref=f"A1:{last_column}{ws.max_row}")
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
                showRowStripes=True, showColumnStripes=False,
            )
            ws.add_table(table)
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True

    fill_sheet(
        wb.active, columns, "КРТ" if kind == "krt" else "Выборка",
        "DevelopAidKrtSelection" if kind == "krt" else "DevelopAidSelection",
    )
    if obligation_columns:
        fill_sheet(
            wb.create_sheet("Обязательства"), obligation_columns,
            "Обязательства", "DevelopAidKrtDuties",
        )
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _handoff_land_cadastres(preset: dict[str, Any], context: dict[str, Any]) -> list[str]:
    """Keep source lot data intact; narrow only the DevelopAid handoff."""
    land_numbers = [
        str(item.get("cadastral_number"))
        for item in (context.get("land_parcels") or [])
        if item.get("cadastral_number")
    ]
    if not land_numbers:
        return []
    preset["project"]["cadastral_numbers"] = land_numbers
    preset["project"]["cadastral_numbers_input"] = ", ".join(land_numbers)
    preset["land"]["cadastral_numbers"] = land_numbers
    preset["land"]["cadastral_numbers_csv"] = ", ".join(land_numbers)
    cadastral_import = preset["project"]["cadastral_import"]
    cadastral_import["mode"] = "bulk" if len(land_numbers) > 1 else "single"
    cadastral_import["note"] += (
        " КН зданий/ОКС исключены из площади территории; "
        "переданы только земельные участки НСПД."
    )
    return land_numbers


_LOTONLINE_PROJECT_SHARES_FLAG = "AUCTION_LOTONLINE_PROJECT_SHARES_DISCOVERY"


def _feature_enabled(name: str, default: bool = False) -> bool:
    raw = os.getenv(name, "").strip().lower()
    if not raw:
        return default
    return raw in {"1", "true", "yes", "on"}


def _lot_online_discovery_adapter() -> LotOnlineAdapter:
    # Доли в юрлицах спрашиваются по умолчанию (владелец, 01.09.2026: «надо
    # посмотреть ещё лоты по продаже долей в юр лицах»). Категория 85 РАД —
    # «Акции и доли предприятий»; выключатель остаётся на случай, когда цена
    # обхода каталога важнее полноты.
    return LotOnlineAdapter(
        include_project_shares=_feature_enabled(_LOTONLINE_PROJECT_SHARES_FLAG, True),
    )


def _adapter_for(url: str):
    host = (urlparse(url).hostname or "").lower()
    if host == "roseltorg.ru" or host.endswith(".roseltorg.ru"):
        return RoseltorgAdapter()
    if host == "lot-online.ru" or host.endswith(".lot-online.ru"):
        return LotOnlineAdapter()
    if host == ETPGPBAdapter.HOST or host.endswith("." + ETPGPBAdapter.HOST) or host.endswith(".gpb.ru"):
        return ETPGPBAdapter()
    if host == ETPRFAdapter.HOST or host.endswith("." + ETPRFAdapter.HOST):
        return ETPRFAdapter()
    if host == SberbankASTAdapter.HOST or host.endswith("." + SberbankASTAdapter.HOST):
        return SberbankASTAdapter()
    if host == NistpAdapter.HOST or host.endswith("." + NistpAdapter.HOST):
        return NistpAdapter()
    if host == TorgiGovAdapter.HOST or host.endswith("." + TorgiGovAdapter.HOST):
        return TorgiGovAdapter()
    raise ValueError(
        "Разбирать умеем только официальные адреса Росэлторга, РАД/Lot-online, "
        f"ЭТП ГПБ, ЭТП РФ, Сбербанк-АСТ и ГИС Торгов; «{host or url}» ни на один из них не похож")


def _analysis_support(url: str) -> dict[str, Any]:
    """Можно ли разобрать этот лот целиком — и если нет, то почему.

    Отвечает тот же `_adapter_for`, который потом и разбирает. Второе правило о
    том же самом однажды разошлось бы с первым, и кнопка обещала бы разбор там,
    где его нет.

    Спрашивается это ДО клика намеренно. Список показывал все лоты одинаково, а
    половина из них на «Разобрать лот» отвечала «поддерживаются только Росэлторг
    и РАД»: узнать, что карточку не открыть, можно было только нажав
    (владелец, 25.08.2026 — «зачем они тогда в списке»). Отказ, о котором
    сказано заранее, — это свойство лота; отказ после клика — поломка.
    """
    try:
        adapter = _adapter_for(url)
    except ValueError as exc:
        return {"available": False, "reason": str(exc)}
    note = getattr(adapter, "deep_parse_unavailable", "")
    if note:
        return {"available": False, "reason": str(note)}
    return {"available": True, "reason": ""}


def _discovery_adapters(source: str):
    value = (source or "all").strip().lower()
    if value in {"lot_online", "rad", "lot-online"}:
        return [_lot_online_discovery_adapter()]
    if value in {"roseltorg", "ros"}:
        return [RoseltorgAdapter()]
    if value in {"investmoscow", "moscow", "city"}:
        return [InvestMoscowDiscoveryAdapter()]
    if value in {"torgi", "torgi_gov"}:
        return [TorgiGovAdapter()]
    if value in {"etp_gpb", "etpgpb", "gpb"}:
        return [ETPGPBAdapter()]
    if value in {"etp_rf", "etprf"}:
        return [ETPRFAdapter()]
    if value in {"sberbank_ast", "sberbank", "sber", "sberbank-ast"}:
        return [SberbankASTAdapter()]
    if value in {"nistp", "nis"}:
        return [NistpAdapter()]
    if value == "all":
        # ГИС Торги — официальный источник имущественных торгов, а не
        # банкротный рынок. Это честно подписано в его отчёте, но исключать его
        # из кнопки «все источники» нельзя: именно там сейчас есть актуальные
        # московские лоты, когда две узкие ЭТП возвращают пустую выдачу.
        return [
            TorgiGovAdapter(), _lot_online_discovery_adapter(), RoseltorgAdapter(),
            InvestMoscowDiscoveryAdapter(), ETPGPBAdapter(), ETPRFAdapter(),
            SberbankASTAdapter(), NistpAdapter(),
        ]
    raise ValueError(
        "source: all, lot_online, roseltorg, investmoscow, torgi, etp_gpb, etp_rf, sberbank_ast или nistp")


def _coverage_row(adapter: Any) -> dict[str, Any]:
    """Строка охвата одного источника — с его именем.

    Отчёты у читателей разной формы: у ИнвестМосквы свои ключи про карточки
    города, у остальных «страницы / карточки / оставлено». Раньше экран читал
    ТОЛЬКО первый отчёт списка и печатал его ключами — то есть про четыре
    источника из пяти на экране не было ничего, включая их отказы. Имя
    источника проставляется здесь, а не в каждом читателе: иначе следующий
    читатель забудет его так же, как забыли эти.
    """
    name = getattr(adapter, "platform_name", adapter.__class__.__name__)
    report = getattr(adapter, "last_report", None)
    if not isinstance(report, dict):
        # Источник, который отвечал, но отчёта не прислал, из охвата исчезал —
        # и на экране становился неотличим от неопрошенного. Строка есть
        # всегда; то, чего мы про него не знаем, названо, а не показано нулём.
        return {"source": name, "reason": "отчёта об охвате не прислал"}
    row = dict(report)
    row.setdefault("source", name)
    return row


def _public_lot_dict(lot) -> dict[str, Any]:
    data = lot.to_dict()
    data.pop("raw", None)
    return data


def _plato_number(value: Any, digits: int = 0) -> str:
    try:
        return f"{float(value):,.{digits}f}".replace(",", " ").replace(".", ",")
    except (TypeError, ValueError):
        return "—"


_MARKET_DIGEST_KEYS = ("peers", "price_hint", "comparison", "analysis", "subject", "segment")


def _market_digest(report: dict[str, Any]) -> dict[str, Any]:
    """Что из отчёта рынка нужно карточке. Остальное на диске не хранится.

    Полный отчёт несёт и разбор источников, и сырые выдачи поиска. Карточке они
    не нужны, а диск у нас уже кончался — молча и до отказа выкатки.
    """
    digest = {key: report.get(key) for key in _MARKET_DIGEST_KEYS if report.get(key) is not None}
    peers = digest.get("peers")
    if isinstance(peers, list):
        digest["peers"] = peers[:12]
    analysis = digest.get("analysis")
    if isinstance(analysis, dict):
        digest["analysis"] = {key: value for key, value in analysis.items()
                              if key in {"site", "overall"}}
    verdict = report.get("verdict")
    if isinstance(verdict, dict):
        digest["verdict"] = {key: verdict.get(key) for key in
                             ("units_per_month", "sold_lot_avg", "price_per_sqm", "headline")
                             if verdict.get(key) is not None}
    return digest


def _plato_krt_prompt(stored: dict[str, Any]) -> str:
    """Что спросить у Платона про площадку.

    Числа собираются здесь и подаются готовыми: модель их не пересчитывает и
    пересчитывать не должна — экономику считает движок, а Платон читает
    посчитанное. Спрашиваем и по маркетингу, и по модели одним вопросом:
    рекомендация, разведённая по двум ответам, не сходится сама с собой.
    """
    project = stored.get("project") or {}
    screening = stored.get("screening") or {}
    report = stored.get("market") or {}
    metrics = screening.get("metrics") or {}
    market_block = screening.get("market") or {}
    phasing = screening.get("phasing") or {}
    capacity = screening.get("entry_capacity") or {}
    duties = screening.get("requirements") or {}
    verdict = report.get("verdict") or {}
    peers = report.get("peers") or []

    lines = [
        "Ты оцениваешь площадку КРТ в Москве для решения об участии в торгах.",
        "Все числа ниже посчитаны движком DevelopAid и маркетинговым модулем; "
        "пересчитывать их не надо, надо прочитать.",
        "",
        f"ПЛОЩАДКА: {project.get('name') or stored.get('slug')}",
        f"Округ и район: {' · '.join(str(v) for v in (project.get('okrug'), project.get('district')) if v) or '—'}",
        f"Статус: {project.get('status') or '—'}; площадь {project.get('area_ha') or '—'} га; "
        f"жилой объём {_plato_number(project.get('housing_gfa_sqm'))} м².",
        "",
        "МАРКЕТИНГ",
        f"Рекомендованный класс: {market_block.get('recommended_segment') or '—'}; "
        f"стартовая цена в модели {_plato_number(market_block.get('start_price_rub_sqm'))} ₽/м² "
        f"({market_block.get('price_basis') or 'основание не указано'}).",
        f"Ориентир рынка: {_plato_number(market_block.get('market_price_rub_sqm'))} ₽/м²; "
        f"сопоставимых проектов рядом: {len(peers)}.",
    ]
    if verdict.get("units_per_month"):
        lines.append(
            f"Темп рынка: {verdict.get('units_per_month')} ДДУ/мес.; "
            f"средний проданный лот {verdict.get('sold_lot_avg') or '—'} м².")
    lines += [
        "",
        "МОДЕЛЬ DEVELOPAID (предварительная, цена входа принята нулём)",
        f"Выручка {_plato_number(metrics.get('revenue_mln'))} млн ₽, "
        f"CAPEX {_plato_number(metrics.get('capex_mln'))} млн ₽, "
        f"чистая прибыль {_plato_number(metrics.get('net_profit_mln'))} млн ₽, "
        f"маржа {_plato_number(metrics.get('margin_pct'), 1)}%.",
        f"LLCR проекта {_plato_number(metrics.get('project_llcr_x'), 2)}x, "
        f"слабейшая очередь {_plato_number(metrics.get('weakest_phase_llcr_x'), 2)}x "
        f"при целевом ориентире банка 1,20x.",
        f"Очередей: {phasing.get('count') or '—'}; продаваемая площадь "
        f"{_plato_number(phasing.get('saleable_sqm'))} м².",
        f"Пик БРИДЖа {_plato_number(metrics.get('peak_bridge_mln'))} млн ₽, "
        f"пик ПФ {_plato_number(metrics.get('peak_pf_mln'))} млн ₽.",
    ]
    if capacity.get("available"):
        lines.append(
            f"Потолок цены входа при LLCR 1,20x: {_plato_number(capacity.get('amount_mln'), 1)} млн ₽.")
    else:
        lines.append(f"Потолок цены входа не подобран: {capacity.get('reason') or 'причина не названа'}.")
    if duties:
        lines += [
            "",
            "ОПУБЛИКОВАННЫЕ ОБЯЗАТЕЛЬСТВА",
            f"Безусловный снос: {_plato_number(duties.get('demolition_area_sqm'))} м² "
            f"({duties.get('demolition_objects') or 0} объектов); "
            f"снос/реконструкция: {duties.get('conditional_objects') or 0} объектов.",
            f"Расселение/изъятие: {'найдено' if duties.get('resettlement') else 'не найдено в опубликованном проекте решения'}; "
            f"дополнительные объекты: {len(duties.get('unmodelled_construction') or [])}.",
        ]
    for item in (screening.get("exclusions") or [])[:6]:
        lines.append(f"НЕ УЧТЕНО: {item}")
    lines += [
        "",
        "Ответь по-русски, коротко и по делу, без списка формул:",
        "1. Идти ли смотреть эту площадку дальше — да, нет или «при условии», и почему.",
        "2. Что в маркетинге и в модели противоречит друг другу, если противоречит.",
        "3. Три вопроса, которые надо снять до подачи заявки, по убыванию цены вопроса.",
        "Не выдумывай чисел, которых нет выше. Про неучтённое скажи прямо, что оно неучтено.",
    ]
    return "\n".join(lines)


def install(app: FastAPI) -> None:
    if getattr(app.state, "auction_search_installed", False):
        return
    app.state.auction_search_installed = True
    market = getattr(app.state, "market_discovery_service", None)
    krt_registry = getattr(market, "krt", None) or KrtRegistry(
        os.path.join(os.getenv("DATA_DIR", "data"), "market")
    )
    krt_ranking = KrtRanking(os.path.join(os.getenv("DATA_DIR", "data"), "market"))
    # Очередь новинок забирает движок (`/internal/krt/announcements`) — у него
    # общая с ботом подпись, а до api.telegram.org с ядра не дойти. Отдаём ему
    # ТОТ ЖЕ экземпляр каталога, который помечает площадки «новыми» на экране:
    # второй ответ на вопрос «это новое?» однажды разошёлся бы с первым, и в
    # чат приехало бы не то, что видно в списке.
    def _take_krt_announcements() -> list[dict[str, Any]]:
        """Новинки с именами площадок.

        В очередь ложится слаг: имя знает каталог, а `first_seen` — состав.
        Имя подставляется здесь, при выдаче, и если площадки в каталоге уже
        нет — остаётся слаг. Выдумывать имя нельзя, а молча выбросить запись
        значит потерять новость.
        """
        records = krt_ranking.take_announcements()
        if not records:
            return []
        try:
            names = {str(row.get("slug") or ""): str(row.get("name") or "")
                     for row in krt_registry.catalogue()}
        except Exception:  # noqa: BLE001
            names = {}
        return [{**record, "name": names.get(str(record.get("slug") or ""), "")}
                for record in records]

    app.state.krt_announcements_take = _take_krt_announcements

    def _weekly_ranking() -> None:
        """Раз в неделю каталог обновляется и считается сам.

        Ждать прогон каждый раз, когда открываешь торги, — это минуты на пустом
        месте (владелец, 23.08.2026). Здесь считается ВЕСЬ каталог, а не срез
        фильтра: никто не ждёт, а к утру должно быть посчитано всё.

        Срок — ночь с субботы на воскресенье, 3 часа по Москве: к утру
        воскресенья каталог свежий, а рабочая неделя начинается с посчитанного.
        Считается календарная точка, а не «неделя от прошлого раза», иначе
        расписание уползает на часы с каждой выкаткой.

        Воркеров два, память у них раздельная, поэтому работу берёт один — по
        файловому замку. Проигравший просто спит дальше. Отключается
        переменной `AUCTION_KRT_WEEKLY=0`.
        """
        while True:
            try:
                if market is not None and core is not None and krt_ranking.due():
                    if krt_ranking.claim():
                        try:
                            projects = krt_registry.projects(refresh=True)
                            rows = [row.to_dict() if hasattr(row, "to_dict") else row
                                    for row in projects]
                            if rows and not krt_ranking.start(
                                    rows, _screen_one, scheduled=True, claimed=True):
                                krt_ranking.release()
                        except Exception:
                            logger.exception("weekly KRT ranking failed")
                            krt_ranking.release()
            except Exception:
                logger.exception("weekly KRT ranking loop")
            time.sleep(HEARTBEAT_SECONDS)


    # main.py loads the canonical legacy core as `developaid_core`. Inject only the
    # same-origin handoff bootstrap; no calculation or project UI is duplicated.
    core = sys.modules.get("developaid_core")
    if core is not None:
        install_page_bridge(core)

    @app.get("/auctions", response_class=HTMLResponse)
    async def auctions_home() -> HTMLResponse:
        return HTMLResponse(
            auction_page_with_handoff(auctions_page(core)),
            headers={"Cache-Control": "no-store, must-revalidate"},
        )

    @app.get("/auctions/sources")
    async def auction_sources() -> dict[str, Any]:
        return {
            "source_policy": "official_etp_only",
            "sources": [
                {
                    "id": "lot_online",
                    "name": "Российский аукционный дом / Lot-online",
                    "direct_lot_ingest": True,
                    "moscow_discovery": True,
                    "discovery_access": "public_catalogue",
                    "project_company_shares_discovery": {
                        "enabled": _feature_enabled(_LOTONLINE_PROJECT_SHARES_FLAG, True),
                        "rollout": "on_by_default_switchable",
                    },
                },
                {
                    "id": "torgi_gov",
                    "name": "ГИС Торги (torgi.gov.ru) — приватизация и прочие торги",
                    "direct_lot_ingest": False,
                    "moscow_discovery": TorgiGovAdapter.enabled(),
                    "discovery_access": "public_api",
                    "note": ("Не входит в основную подборку: живой ответ подтвердил "
                             "178-ФЗ, а не банкротство. Доступен отдельным источником; "
                             "Москву отбираем по subjectRFCode=77."),
                    # Сертификат torgi.gov.ru выпущен Минцифры, и обычным
                    # хранилищем корней он не проверяется. Корни лежат в
                    # `certs` на машине, а не в репозитории. Пустой список
                    # здесь — самая частая причина «источник включён, а лотов
                    # нет»: соединение обрывается на проверке цепочки, и без
                    # этой строки причину пришлось бы искать в логе.
                    "trusted_roots": torgi_trust_report(),
                },
                {
                    "id": "roseltorg",
                    "name": "Росэлторг",
                    "direct_lot_ingest": True,
                    "moscow_discovery": True,
                    "discovery_access": "public_tags_search_and_property_section",
                    # Что источник СОДЕРЖИТ — часть ответа. Раздел имущества
                    # «Развитие территории» — адрес, с которого город публикует
                    # аукционы КРТ; со страницы берутся только ссылки на
                    # карточки процедур, а читает их тот же разбор, что и
                    # раньше. Чем отвечает сам раздел, показывает
                    # `/auctions/roseltorg/probe` — с ядра: из песочницы
                    # roseltorg.ru закрыт.
                    "note": ("Поиск процедур по тегам плюс раздел имущества "
                             "«Развитие территории» по Москве (ОКАТО 45000000000). "
                             "Проба раздела — /auctions/roseltorg/probe."),
                },
                {
                    "id": "etp_gpb",
                    "name": "ЭТП ГПБ",
                    "direct_lot_ingest": True,
                    "moscow_discovery": True,
                    "discovery_access": "official_public_json_api",
                },
                {
                    "id": "etp_rf",
                    "name": "ЭТП РФ",
                    "direct_lot_ingest": False,
                    "moscow_discovery": True,
                    "discovery_access": "official_public_registry",
                    "note": "Поиск подключён; детальный разбор документов площадки пока недоступен.",
                },
                {
                    "id": "sberbank_ast",
                    "name": "Сбербанк-АСТ",
                    "direct_lot_ingest": True,
                    "moscow_discovery": True,
                    "discovery_access": "official_public_html_post",
                    "note": ("Официальный реестр банкротных торгов: общий список, "
                             "имущество и раздел АП. Если площадка вернула антибот-страницу, "
                             "это отражается в охвате источника; обход защиты не используется."),
                },
                {
                    "id": "nistp",
                    "name": "НИС (nistp.ru)",
                    "direct_lot_ingest": False,
                    "moscow_discovery": True,
                    "discovery_access": "official_public_registry",
                    "note": "Публичный поиск банкротных лотов по Москве; подключён без обхода защиты.",
                },
                {
                    "id": "investmoscow",
                    "name": "Официальный портал «Торги Москвы» → фактическая ЭТП",
                    "direct_lot_ingest": False,
                    "moscow_discovery": True,
                    "discovery_access": "public_city_catalogue_then_official_etp",
                    "facts_source": "conducting_etp_only",
                },
            ],
            "document_access": "public_first_optional_service_account_session",
        }

    @app.get("/auctions/krt")
    async def auction_krt_catalogue() -> dict[str, Any]:
        projects = await run_in_threadpool(krt_registry.catalogue)
        # Каталог отмечается на каждом чтении: «новое» — это разница с прошлым
        # составом, и считать её должен тот, кто состав видит, а не человек
        # глазами по списку из ста двадцати строк.
        seen = await run_in_threadpool(
            krt_ranking.mark_seen, [str(row.get("slug") or "") for row in projects])
        projects = [
            {**row,
             "first_seen_at": seen.get(str(row.get("slug") or "")) or 0,
             "is_new": krt_ranking.is_new(seen.get(str(row.get("slug") or "")))}
            for row in projects
        ]
        # Что карточка города говорит о застройщике и реновации — бесплатный
        # официальный источник, и он больше не ждёт платного прогона: строка
        # каталога несёт то, что уже прочитано, а недостающее дочитывается
        # фоном порциями. Пока это лежало только в прогоне, у площадки с явным
        # оператором в колонке не стояло ничего, и «не спрашивали» читалось как
        # «оператора нет» (владелец, 01.09.2026).
        known = getattr(krt_registry, "card_facts_known", None)
        if callable(known):
            slugs = [str(row.get("slug") or "") for row in projects]
            try:
                facts = await run_in_threadpool(known, slugs)
            except Exception:  # noqa: BLE001
                logger.exception("KRT card facts cache failed")
                facts = {}
            if facts:
                projects = [
                    {**row, "card_facts": facts.get(str(row.get("slug") or ""))}
                    if facts.get(str(row.get("slug") or "")) else row
                    for row in projects
                ]
            filler = getattr(krt_registry, "fill_card_facts_in_background", None)
            if callable(filler):
                try:
                    filler(slugs)
                except Exception:  # noqa: BLE001
                    logger.exception("KRT card facts fill failed")

        status = krt_registry.status()
        # Неразобранная карточка называется вслух. Её значения съезжают на
        # поле — округом становится статус, статусом хвост адреса, — и такая
        # строка не проходит ни один флажок округа и пропадает с экрана без
        # единого слова. Молча выброшенная площадка читается как её
        # отсутствие, а это самая крупная площадка каталога и есть.
        unparsed = [row for row in projects if row.get("parse_problem")]
        # Площадка, у которой опубликован ПРОЕКТ решения, а карточки в
        # каталоге нет, стояла отдельным списком под таблицей — то есть в общем
        # ряду её не было («КРТ без карточки надо включать в общий список, но
        # делать пометку», владелец, 31.08.2026). Она едет строкой сюда же, с
        # меткой и датой. ТЭП у неё нет вовсе, и придумывать их нельзя: в
        # колонках стоит прочерк, а не ноль.
        #
        # Метка называет ДОКУМЕНТ, а не его отсутствие: «без карточки» — это
        # про наш каталог, а на mos.ru лежит проект решения о КРТ, и это не
        # решение (владелец, 31.08.2026: «то, что ты назвал КРТ без карточки,
        # вообще-то ПРОЕКТ решения»). Разница не в слове: проект решения город
        # публикует для сбора мнений правообладателей — решения ещё нет, и
        # назвать его принятым значит показать стадию, которой не наступало.
        # Зато это самый ранний сигнал из всех: до торгов остаётся время.
        decision_rows: list[dict[str, Any]] = []
        reader = getattr(krt_registry, "decisions", None)
        if callable(reader):
            try:
                found = await run_in_threadpool(reader)
            except Exception as exc:  # noqa: BLE001
                logger.exception("KRT decisions failed")
                found = {"decisions": [], "error": f"{type(exc).__name__}: {exc}"}
            # Дата проекта решения у площадки, которая в каталоге ЕСТЬ: колонка
            # стояла пустой, будто документа не существует, а он найден и
            # сопоставлен. «С карточкой и без — это по сути решение принято или
            # это ещё проект» (владелец, 01.09.2026): ответить на это можно
            # только датой документа, а не наличием карточки.
            by_slug = {str(one.get("slug") or ""): one
                       for one in (found.get("matched_rows") or [])}
            if by_slug:
                projects = [
                    {**row,
                     "draft_decision_at": (by_slug.get(str(row.get("slug") or "")) or {}).get("published_at") or 0,
                     "draft_decision_url": (by_slug.get(str(row.get("slug") or "")) or {}).get("url") or ""}
                    for row in projects
                ]
            for one in (found.get("decisions") or []):
                decision_rows.append({
                    "slug": "decision:" + str(one.get("id") or ""),
                    "name": str(one.get("address") or one.get("title") or ""),
                    "okrug": one.get("okrug") or "",
                    "district": "",
                    "status": "Проект решения",
                    "url": one.get("url"),
                    "no_card": True,
                    "krt_kind": one.get("kind") or "",
                    "draft_decision_at": one.get("published_at") or 0,
                    "department": one.get("department") or "",
                })
            projects = projects + decision_rows
        return {
            "source": CATALOGUE_URL,
            "geometry_status": "not_published_in_catalogue",
            **status,
            "count": len(projects),
            "unparsed_count": len(unparsed),
            "unparsed": [
                {"slug": row.get("slug"), "name": row.get("name"),
                 "problem": row.get("parse_problem")}
                for row in unparsed[:20]
            ],
            "no_card_count": len(decision_rows),
            "new_count": sum(1 for row in projects if row.get("is_new")),
            "new_for_days": NEW_FOR_SECONDS // 86400,
            "projects": projects,
        }

    def _read_open_sources(project: dict[str, Any]) -> dict[str, Any]:
        """Что об этой площадке сказано в публикациях и каналах. Один разбор.

        Зовётся и кнопкой на карточке, и еженедельным прогоном: второй такой
        разбор однажды ответил бы про одну площадку иначе, чем первый, и оба
        ответа выглядели бы верными.
        """
        from market_search import krt_open_sources

        name = str((project or {}).get("name") or "")
        if not name:
            return {"available": False, "reason": "У площадки нет имени — искать нечего"}
        # Поиск берётся у маркетингового движка — того самого, которым считаются
        # соседи по рынку. Своего второго не заводим: модуль, идущий наружу мимо
        # общего пути, однажды ответит иначе, чем весь остальной сервис.
        client = getattr(market, "search", None)
        finder = getattr(client, "search", None)
        if not callable(finder) or not getattr(client, "configured", False):
            return {"available": False,
                    "reason": "Веб-поиск не настроен — публикации спросить нечем"}
        asked = krt_open_sources.queries(
            name, str((project or {}).get("okrug") or ""),
            str((project or {}).get("district") or ""))
        docs: list[Any] = []
        errors: list[str] = []
        for query in asked:
            try:
                docs.extend(finder(query))
            except Exception as exc:  # noqa: BLE001
                errors.append(f"{type(exc).__name__}: {exc}")
        if not docs and errors:
            # Поиск не ответил — это ответ, а не «в источниках ничего нет».
            return {"available": False, "reason": "; ".join(errors[:2]), "queries": asked}
        found = krt_open_sources.read_findings(docs, name)
        # Рынок знает площадку по имени проекта, а не по адресу: «Строгино 360»
        # знают все, «Маршала Прошлякова ул., вл. 9» — никто. Имя берётся из
        # публикации, где оно стоит рядом с адресом, и вторым кругом ищется уже
        # оно. Круг ровно один: поиск платный, а бренд у площадки один.
        for brand in (found.get("brands") or [])[:1]:
            query = f'"{brand}" застройщик КРТ Москва'
            asked = asked + [query]
            try:
                docs.extend(finder(query))
            except Exception as exc:  # noqa: BLE001
                errors.append(f"{type(exc).__name__}: {exc}")
            else:
                found = krt_open_sources.read_findings(docs, name)
        # Телеграм-каналы говорят о площадке раньше прессы (владелец,
        # 01.09.2026: «а поиск информации можно через телеграм-каналы ещё
        # вести?»). Спрашиваются тем же платным индексом с ограничением по
        # домену — своего пути наружу не заводим. Круг один: у поиска цена.
        tg_asked = krt_open_sources.telegram_queries(name, found.get("brands") or [])
        tg_ok = False
        for query in tg_asked:
            asked = asked + [query]
            try:
                docs.extend(finder(query))
            except Exception as exc:  # noqa: BLE001
                errors.append(f"{type(exc).__name__}: {exc}")
            else:
                tg_ok = True
                found = krt_open_sources.read_findings(docs, name)
        found.update({"available": True, "queries": asked, "errors": errors[:2],
                      "checked_at": int(time.time()),
                      # Спросили каналы и там пусто — не то же, что не спросили.
                      "telegram_asked": tg_ok})
        return found

    # Занятость — ответ ОДНОСТОРОННИЙ: отданная площадка свободной не
    # становится (владелец, 01.09.2026: «может, один раз провести прогон как с
    # моделью, а потом только по требованию?»). Значит спрашивать её повторно
    # незачем: прогон платит за площадку один раз, дальше платит только тот,
    # кто нажал «Пересчитать сейчас». Неотвеченная и свободная спрашиваются
    # снова — там ответ ещё может измениться.
    def _open_sources_for_run(project: dict[str, Any]) -> dict[str, Any]:
        slug = str((project or {}).get("slug") or "")
        stored = {}
        try:
            stored = (krt_ranking.stored_row(slug) or {}).get("press_facts") or {}
        except Exception:  # noqa: BLE001
            logger.exception("stored press facts failed slug=%s", slug)
        if stored.get("available") and stored.get("taken"):
            return stored
        try:
            return _read_open_sources(project)
        except Exception as exc:  # noqa: BLE001
            logger.exception("KRT open sources failed slug=%s", slug)
            fresh = {"available": False, "reason": f"{type(exc).__name__}: {exc}"}
            # Неудача не затирает прежний ответ — правило «посчитанное не
            # выбрасывают» здесь то же самое.
            return stored or fresh

    def _screen_one(project: dict[str, Any]) -> dict[str, Any]:
        """Один прогон для рейтинга — тем же путём, что и открытая карточка.

        Второго скрининга не заводим: разойдись они, список и карточка
        показали бы про одну площадку разное, и оба достоверно.
        """
        if core is None or market is None:
            return {"available": False, "reason": "Финансовый движок DevelopAid не подключён"}
        report = market.build_report(
            f"krt:{project.get('slug')}", radius_km=3.0, peers_limit=12,
            city_reference=False, include_project_totals=True,
        )
        try:
            requirements = krt_registry.requirements(str(project.get("slug") or ""))
        except Exception as exc:  # noqa: BLE001
            logger.exception("KRT requirements failed slug=%s", project.get("slug"))
            requirements = {
                "available": False,
                "warning": f"Документы обязательств временно не прочитаны: {type(exc).__name__}",
            }
        # Карточка каталога — официальный и бесплатный источник застройщика и
        # реновации. Читается в прогоне, а не по нажатию: иначе фильтр по
        # оператору и городским нуждам работает только по открытым руками.
        try:
            card = krt_registry.card_facts(str(project.get("slug") or ""))
        except Exception as exc:  # noqa: BLE001
            logger.exception("KRT card facts failed slug=%s", project.get("slug"))
            card = {"available": False, "reason": f"{type(exc).__name__}: {exc}"}
        screening = build_krt_model_screening(
            project, report, core, requirements=requirements)
        screening["card_facts"] = card
        # Занятость площадки — в прогон, а не по нажатию: пока она приходила
        # только кнопкой, каталог показывал «Планируемая» там, где договор
        # давно подписан. Платится один раз на площадку: занятая больше не
        # спрашивается никогда.
        screening["press_facts"] = _open_sources_for_run(project)
        # Маркетинг едет вместе со скринингом и оседает в отчёте площадки.
        # Считать его второй раз при открытии карточки незачем: прогон уже
        # сходил к рынку, к соседям и к движку — минуты чужого ожидания на
        # готовом ответе. Кладётся выжимка, а не отчёт целиком: сто двадцать
        # площадок × полный отчёт — это десятки мегабайт на диске, который у
        # нас уже кончался, а карточка рисует ровно эти блоки.
        screening["market_report"] = _market_digest(report)
        return screening

    @app.get("/auctions/krt/decisions")
    async def auction_krt_decisions(refresh: bool = False) -> dict[str, Any]:
        """Решения о КРТ, у которых нет карточки в каталоге.

        Каталог krt.mos.ru отвечает на «какие площадки город показывает».
        Решение публикуется отдельно, и площадка может иметь опубликованное
        решение, не имея карточки вовсе — до сих пор такие были не видны ни при
        каком фильтре. Сопоставление строгое: ложная привязка прячет настоящий
        пробел, поэтому «не сопоставилось» здесь значит «карточки не нашли», а
        не «новая площадка».
        """
        reader = getattr(krt_registry, "decisions", None)
        if not callable(reader):
            raise HTTPException(status_code=503, detail="Чтение решений недоступно")
        try:
            payload = await run_in_threadpool(lambda: reader(refresh=bool(refresh)))
        except Exception as exc:  # источник закрыт — это ответ, а не пустой список
            raise HTTPException(status_code=502,
                                detail=f"Поиск mos.ru не ответил: {exc}") from exc
        return payload

    @app.get("/auctions/investmoscow/probe")
    async def auction_investmoscow_probe() -> dict[str, Any]:
        """Почему у нас нет ИнвестМосквы. Проба показывает ответ, а не догадку.

        Адаптер есть и стоит в наборе «все источники», но лот у него рождается
        длинной цепочкой: каталог → карточки города → ссылка на официальную ЭТП
        → лот у адаптера ЭТП. Оборвись цепочка на первом шаге — карточек ноль,
        лотов ноль, и в отчёте источника это видно только числами. Проба
        называет, где именно рвётся: сколько байт вернул каталог, похоже ли это
        на оболочку SPA и сколько ссылок на карточки в нём нашлось.
        """
        from .adapters.investmoscow import InvestMoscowDiscoveryAdapter as _IM

        def collect() -> list[dict[str, Any]]:
            adapter = _IM()
            out: list[dict[str, Any]] = []
            for url in _IM._search_urls()[:3]:
                row: dict[str, Any] = {"url": url}
                try:
                    html = adapter._read_html(url)
                except Exception as exc:  # noqa: BLE001
                    row["error"] = f"{type(exc).__name__}: {exc}"[:300]
                    out.append(row)
                    continue
                row["bytes"] = len(html)
                low = html[:600].lower()
                row["looks_html"] = "<html" in low or "<!doctype" in low
                try:
                    cards = _IM._city_card_urls(url, html)
                except Exception as exc:  # noqa: BLE001
                    row["parse_error"] = f"{type(exc).__name__}: {exc}"[:200]
                    cards = []
                row["city_cards"] = len(cards)
                row["first_cards"] = list(cards)[:3]
                out.append(row)
            return out

        answers = await run_in_threadpool(collect)
        return {
            "source": "investmoscow",
            "asked": len(answers),
            "answers": answers,
            "note": ("Ноль карточек при живом ответе означает, что каталог отдаёт "
                     "оболочку, а список приезжает отдельным вызовом — тогда нужен "
                     "браузер и адрес, по которому страница сама ходит за данными: "
                     "/auctions/etp/probe/browser?url=…"),
        }

    @app.post("/auctions/krt/{slug}/tender-order")
    async def auction_krt_mark_tender(slug: str, request: Request) -> dict[str, Any]:
        """Отметить: по этой площадке объявлены торги — по такому-то распоряжению.

        Машине привязать нечем. Распоряжение ДГП не называет адреса ни в
        заголовке, ни в карточке документа, а его PDF — скан: семь страниц, 199
        картинок на первой, текста только регистрационный штамп. Придумать
        привязку по номеру или по дате значило бы объявить площадку выставленной
        на торги без единого основания.

        Поэтому отметку ставит человек, открывший распоряжение, а мы храним её
        как его утверждение — с номером, ссылкой и датой отметки.
        """
        setter = getattr(krt_registry, "mark_tender", None)
        if not callable(setter):
            raise HTTPException(status_code=503, detail="Отметки недоступны")
        payload = await request.json()
        order = (payload or {}).get("order") or {}
        if order and not str(order.get("url") or "").startswith("https://www.mos.ru/"):
            raise HTTPException(status_code=422,
                                detail="Ссылка должна вести на распоряжение mos.ru")
        try:
            entry = await run_in_threadpool(
                lambda: setter(slug, order, str((payload or {}).get("who") or "")))
        except ValueError as exc:
            raise HTTPException(status_code=422, detail=str(exc)) from exc
        return {"slug": slug, "link": entry}

    @app.get("/auctions/krt/tender-links")
    async def auction_krt_tender_links() -> dict[str, Any]:
        """Все проставленные вручную привязки распоряжений к площадкам."""
        reader = getattr(krt_registry, "tender_link", None)
        if not callable(reader):
            return {"links": {}}
        return {"links": await run_in_threadpool(lambda: reader(""))}

    @app.get("/auctions/krt/map")
    async def auction_krt_map(refresh: bool = False, step_m: float = Query(default=40.0, ge=1.0, le=200.0)) -> dict[str, Any]:
        """Площадки КРТ с официальными границами — для карты Москвы.

        Контуры отдаются в метрах веб-меркатора, в тех же, в которых считает
        подложка `/land/basemap`: переводить их в браузере значило бы завести
        вторую проекцию, а перепутанный порядок пары «широта, долгота» молча
        зеркалит полигон — участок остаётся правдоподобным и встаёт не туда.
        """
        reader = getattr(krt_registry, "map_dataset", None)
        if not callable(reader):
            raise HTTPException(status_code=503, detail="Реестр карты недоступен")
        try:
            return await run_in_threadpool(
                lambda: reader(refresh=bool(refresh), step_m=float(step_m)))
        except Exception as exc:  # noqa: BLE001
            raise HTTPException(status_code=502,
                                detail=f"Карта КРТ не прочитана: {exc}") from exc

    @app.get("/auctions/krt/api-probe")
    async def auction_krt_api_probe(url: str = Query(default="")) -> dict[str, Any]:
        """Что на самом деле отдаёт api.krt.mos.ru. Разбора здесь нет.

        Каталог мы читаем с этого домена — но КАК HTML, парсером разметки и с
        текстовым фолбэком. Есть ли под ним JSON со статусами, включая
        «Проекты на торгах», мы не знаем: адреса не спрашивали ни разу.
        Гадать имена полей мы уже пробовали на ГИС Торгах — приехали гаражи,
        поэтому здесь только проба, которая ПОКАЗЫВАЕТ ответ.
        """
        from . import krt_api_probe

        return await run_in_threadpool(krt_api_probe.probe, url.strip())

    @app.get("/auctions/krt/api-probe/browser")
    async def auction_krt_api_probe_browser(
        seconds: float = Query(default=45.0, ge=5.0, le=90.0),
        url: str = Query(default=""),
    ) -> dict[str, Any]:
        """Та же проба настоящим браузером: адреса, по которым ходит сама страница.

        У SPA данные приезжают отдельными вызовами бэкенда, и списка статусов в
        разметке может не быть вовсе. Секреты проба не печатает — ни заголовков,
        ни cookies, ни хранилища.
        """
        from .adapters.browser_probe import probe_browser

        from . import krt_api_probe

        return await run_in_threadpool(
            lambda: probe_browser(url.strip() or krt_api_probe.PORTAL_URL, seconds=seconds))

    @app.post("/auctions/krt/tenders")
    async def auction_krt_tenders(request: Request) -> dict[str, Any]:
        """Торги по КРТ: распоряжения города и лоты, привязанные к площадкам.

        Лоты приходят те, что уже собраны на экране: второй прогон каталога
        стоил бы минуту чужого ожидания на готовом ответе. Правило совпадения
        одно — то же, что у решений: улица держится за своим владением, иначе
        ложная привязка объявит площадку проданной.
        """
        from . import krt_tenders

        payload = await request.json()
        lots = (payload or {}).get("lots") or []
        if not isinstance(lots, list) or len(lots) > 5000:
            raise HTTPException(status_code=422, detail="Список лотов не разобран")
        sites = await run_in_threadpool(krt_registry.catalogue)
        matched = await run_in_threadpool(krt_tenders.match, lots, sites)
        orders: dict[str, Any] = {}
        reader = getattr(krt_registry, "tender_orders", None)
        if callable(reader):
            try:
                orders = await run_in_threadpool(reader)
            except Exception as exc:  # noqa: BLE001
                logger.exception("KRT tender orders failed")
                orders = {"orders": [], "error": f"{type(exc).__name__}: {exc}"}
        # Распоряжение привязывается к площадке САМО: адрес и короткое имя КРТ
        # достаются распознаванием скана, а совпадение считает то же правило,
        # что у решений и лотов. Ручная отметка остаётся поправкой на те
        # распоряжения, где распознать не удалось.
        from market_search import krt_tender_orders as order_reader

        rows = orders.get("orders") or []
        order_by_site: dict[str, dict[str, Any]] = {}
        unbound: list[dict[str, Any]] = []
        for one in rows:
            site = await run_in_threadpool(order_reader.match_site, one, sites)
            if site:
                order_by_site.setdefault(str(site.get("slug") or ""), one)
            elif one.get("address") or one.get("krt_name"):
                unbound.append(one)
        return {**matched, "orders": rows,
                "orders_by_site": order_by_site,
                "orders_unbound": unbound[:40],
                "orders_note": orders.get("note") or "",
                "orders_complete": orders.get("complete"),
                "orders_error": orders.get("error") or ""}

    @app.get("/auctions/krt/{slug}/open-sources")
    async def auction_krt_open_sources(slug: str) -> dict[str, Any]:
        """Что об этой площадке сказано в публикациях.

        В самом проекте решения об операторе и городских нуждах не сказано
        почти ничего — проверено на восьми живых документах (31.08.2026).
        Живут эти факты в публикациях mos.ru и деловой прессы, откуда их берёт
        и ручная таблица владельца. Ни один признак не ставится без цитаты и
        ссылки, а поиск — общий крючок сервиса, своего здесь не заводим.
        """
        project = next((row for row in krt_registry.catalogue()
                        if row.get("slug") == slug), None)
        if not str((project or {}).get("name") or ""):
            raise HTTPException(status_code=404, detail="Площадка не найдена в каталоге")
        # Разбор один на кнопку и на прогон (`_read_open_sources`). Здесь
        # стояло имя `service`, а оно живёт локально внутри маршрута выдачи
        # лотов: снаружи его не существует, и кнопка «Что пишут об этой
        # площадке» отвечала пятисоткой ВСЕГДА (экран владельца, 01.09.2026).
        found = await run_in_threadpool(_read_open_sources, project or {})
        # Прочитанное остаётся на сервере, а не в памяти вкладки: поиск
        # платный, и потерянная находка — это второй счёт за тот же ответ.
        # Кладётся туда же, куда кладёт прогон, и общее для всех.
        if found.get("available") is not False:
            await run_in_threadpool(krt_ranking.remember, slug, {"press_facts": found})
        return found

    @app.get("/auctions/krt/{slug}/card-facts")
    async def auction_krt_card_facts(slug: str) -> dict[str, Any]:
        """Застройщик и реновация — со страницы самой площадки.

        Один запрос к krt.mos.ru, без поиска и без его квоты. Прогон читает то
        же самое и кладёт в строку рейтинга; здесь — для площадки, которую ещё
        не считали: открыл карточку и видишь, кто застройщик.
        """
        reader = getattr(krt_registry, "card_facts", None)
        if not callable(reader):
            raise HTTPException(status_code=503, detail="Чтение карточки не подключено")
        return await run_in_threadpool(reader, slug)

    @app.get("/auctions/krt/{slug}/requirements")
    async def auction_krt_requirements(slug: str) -> dict[str, Any]:
        """Requirements from the official decision for one planned KRT.

        This endpoint deliberately distinguishes an unpublished fact from a
        negative fact.  If the project decision says nothing about resettlement,
        the UI shows "not published", never "no resettlement".  The signed KRT
        agreement remains the final source of the developer's duties.
        """
        reader = getattr(krt_registry, "requirements", None)
        if not callable(reader):
            raise HTTPException(
                status_code=503, detail="Чтение требований КРТ не подключено")
        result = await run_in_threadpool(reader, slug)
        if result is None:
            raise HTTPException(status_code=404, detail="Территория КРТ не найдена")
        return result

    @app.get("/auctions/krt/{slug}/point")
    async def auction_krt_point(slug: str) -> dict[str, Any]:
        """Геокодированная точка территории — чтобы показать её на карте.

        Полный отчёт рынка ради одной картинки гонять незачем: он считает
        соседей, цены и модель. Здесь только адрес → координата, через тот же
        геокодер движка и тот же кэш; своего второго geocode не заводим.
        Точность возвращается вместе с точкой: центр района выглядит на карте
        так же уверенно, как настоящий адрес.
        """
        if market is None:
            raise HTTPException(status_code=503, detail="Маркетинговый движок не подключён")
        finder = getattr(krt_registry, "find", None)
        project = finder(f"krt:{slug}") if callable(finder) else None
        if project is None:
            project = next(
                (item for item in krt_registry.catalogue() if item.get("slug") == slug), None)
        if not project:
            raise HTTPException(status_code=404, detail="Территория КРТ не найдена")
        # Точка берётся тем же путём, что и точка отчёта, — `resolve_subject`.
        # Свой вызов геокодера здесь был четвёртым случаем одной ошибки: модуль
        # завёл свой путь туда, где у сервиса уже есть общий. Цена была не в
        # красоте. Во-первых, он звал `market.geocoder` напрямую, мимо крючка
        # `geocode_address`, — то есть на ядре один Nominatim вместо движковой
        # цепочки (Яндекс, DaData, Nominatim), и карта отвечала «место не
        # найдено» на территорию, которую отчёт находил. Во-вторых, он слал
        # один запрос вместо лестницы `_krt_geocode_candidates` (отдельный
        # адрес → запрос каталога → район). И в-третьих — главное: точка карты
        # и точка, на которой посчитан отчёт, могли оказаться разными. Карта —
        # то, на что смотрят; расходиться ей с расчётом нельзя.
        try:
            subject = await run_in_threadpool(market.resolve_subject, f"krt:{slug}")
        except (GeocodingError, RemoteServiceError, RuntimeError) as exc:
            raise HTTPException(status_code=422, detail=str(exc)) from exc
        data = subject.to_dict() if hasattr(subject, "to_dict") else {}
        # Ссылка на публичную карту НСПД строится движком (`_nspd_map_url`) —
        # координаты там в веб-меркаторе, и второй сборщик этого адреса в
        # браузере разошёлся бы с первым молча. Ссылка нужна не как украшение:
        # подложка — то, на чём рисуют, и когда она не пришла, первоисточник
        # остаётся единственным способом посмотреть на участок.
        nspd_url = ""
        latitude, longitude = data.get("latitude"), data.get("longitude")
        if core is not None and latitude is not None and longitude is not None:
            try:
                merc_x, merc_y = core._wgs84_to_mercator(
                    float(latitude), float(longitude))
                nspd_url = core._nspd_map_url(
                    {"merc_x": round(merc_x, 2), "merc_y": round(merc_y, 2)}, "")
            except Exception:
                logger.exception("НСПД: адрес карты не собрался slug=%s", slug)
                nspd_url = ""
        return {
            "slug": slug,
            "query": data.get("query") or "",
            "latitude": latitude,
            "longitude": longitude,
            "precision": data.get("precision"),
            "address": data.get("address"),
            # Чем опознана точка — часть ответа, а не подробность: центр района
            # выглядит на карте так же уверенно, как настоящий адрес.
            "notes": data.get("notes") or [],
            "area_ha": project.get("area_ha"),
            "nspd_url": nspd_url,
        }

    @app.get("/auctions/krt/ranking")
    async def auction_krt_ranking() -> dict[str, Any]:
        """Балл по всем КРТ: потолок цены входа на метр продаваемой.

        Отдаёт посчитанное сразу — даже на ходу прогона: половина рейтинга с
        честным ходом полезнее пустого экрана, который ничего не объясняет.
        """
        return {
            "measure": "entry_capacity_rub_per_sqm",
            "measure_label": "Потолок цены входа, ₽/м² продаваемой",
            "target_llcr": 1.20,
            "price_note": (
                "Цена аукциона в балл не входит: у проекта каталога krt.mos.ru "
                "ценового поля нет. Балл отвечает «сколько площадка выдерживает "
                "за вход», а не «проходит ли она по объявленной цене»."
            ),
            "progress": krt_ranking.progress(),
            "rows": krt_ranking.rows(),
        }

    @app.post("/auctions/krt/ranking/refresh")
    async def auction_krt_ranking_refresh(
        request: Request, body: KrtRankingRequest | None = None
    ) -> dict[str, Any]:
        """Запустить фоновый прогон. На ходу — не запускает второй.

        Считаются те площадки, которые остались после фильтра на экране, а не
        весь каталог: смотрят перспективные округа и нужный статус, а прогон по
        всем ста двадцати четырём — это минуты чужой работы и чужой нагрузки на
        рынок (владелец, 23.08.2026). Список слагов приходит со страницы; пустой
        список значит «весь каталог» — так ведёт себя вызов без тела.
        """
        market_cabinet.require_cabinet(request)
        if market is None or core is None:
            raise HTTPException(
                status_code=503,
                detail="Прогон требует и маркетингового движка, и движка DevelopAid",
            )
        projects = await run_in_threadpool(krt_registry.catalogue)
        if not projects:
            raise HTTPException(
                status_code=503,
                detail="Каталог КРТ ещё не получен — обновите каталог и повторите",
            )
        wanted = [str(slug) for slug in ((body.slugs if body else None) or []) if str(slug).strip()]
        if wanted:
            order = {slug: index for index, slug in enumerate(wanted)}
            projects = [row for row in projects if str(row.get("slug")) in order]
            projects.sort(key=lambda row: order.get(str(row.get("slug")), 0))
            if not projects:
                raise HTTPException(
                    status_code=422,
                    detail="Ни одна из выбранных площадок не найдена в каталоге",
                )
        started = krt_ranking.start(projects, _screen_one)
        progress = krt_ranking.progress()
        if started:
            reason = ""
        elif progress.get("running"):
            reason = "Прогон уже идёт"
        else:
            # Воркеров два, и ход прогона виден только тому, кто его ведёт.
            # «Прогон уже идёт» без этой оговорки читалось бы как поломка:
            # на экране ничего не движется, а кнопка отказывает.
            reason = ("Прогон уже идёт в соседнем процессе — строки появляются "
                      "в таблице по мере счёта")
        return {"started": started, "reason": reason, "progress": progress}

    def _press_only(project: dict[str, Any]) -> dict[str, Any]:
        """Прогон одних публикаций: без рынка и без модели.

        Ответ на «кто здесь собрался строить» у планируемой площадки бывает
        только один — публикации и каналы: карточка города до торгов
        застройщика не называет (0 из 30 измеренных). А поиск был заперт внутри
        полного прогона, который на каждой площадке ещё строит рыночный отчёт и
        гоняет движок — минуты на площадку, поэтому он и ходит раз в неделю.
        Дешёвый проход отвечает на один вопрос и стоит своих пяти запросов.

        Модель здесь не считается, и это сказано вслух: «не посчитали» — тоже
        ответ, а балл в строке остаётся прежним, потому что неудавшийся
        пересчёт не затирает удавшийся.
        """
        return {
            "available": False,
            "reason": "Прогон публикаций: модель и рынок в нём не считаются",
            "press_facts": _open_sources_for_run(project),
        }

    @app.post("/auctions/krt/press/run")
    async def auction_krt_press_run(request: Request) -> dict[str, Any]:
        """Прочитать публикации по всем планируемым площадкам разом.

        По планируемым — потому что вопрос «кто собрался строить» есть только у
        них: у площадки в реализации застройщика называет сама карточка города,
        бесплатно и без поиска.

        Уже спрошенные пропускаются: занятая площадка свободной не станет, и
        платить за неё второй раз незачем (владелец, 01.09.2026). Замок общий с
        прогоном модели — двум проходам по одному каталогу расходиться негде.
        """
        market_cabinet.require_cabinet(request)
        if market is None:
            raise HTTPException(
                status_code=503,
                detail="Поиск по публикациям требует маркетингового движка",
            )
        projects = await run_in_threadpool(krt_registry.catalogue)
        if not projects:
            raise HTTPException(
                status_code=503,
                detail="Каталог КРТ ещё не получен — обновите каталог и повторите",
            )
        planned = []
        for row in projects:
            if "реализац" in str(row.get("status") or "").lower():
                continue
            stored = {}
            try:
                stored = (krt_ranking.stored_row(str(row.get("slug") or "")) or {}).get(
                    "press_facts") or {}
            except Exception:  # noqa: BLE001
                logger.exception("stored press facts failed slug=%s", row.get("slug"))
            if stored.get("available") and stored.get("taken"):
                continue
            planned.append(row)
        if not planned:
            return {"started": False, "reason": "Спрашивать нечего: планируемых площадок нет",
                    "planned": 0, "progress": krt_ranking.progress()}
        started = krt_ranking.start(planned, _press_only)
        progress = krt_ranking.progress()
        if started:
            reason = ""
        elif progress.get("running"):
            reason = "Прогон уже идёт"
        else:
            reason = ("Прогон уже идёт в соседнем процессе — находки появляются "
                      "в таблице по мере чтения")
        return {"started": started, "reason": reason,
                "planned": len(planned), "progress": progress}

    def _stored_report(slug: str) -> dict[str, Any]:
        stored = krt_ranking.report(slug)
        if not stored:
            raise HTTPException(
                status_code=404,
                detail="Эта площадка ещё не считалась. Запустите прогон — или дождитесь "
                       "еженедельного: он идёт в ночь с субботы на воскресенье.",
            )
        return stored

    @app.get("/auctions/krt/{slug}/report")
    async def auction_krt_report(slug: str, request: Request) -> dict[str, Any]:
        """Готовый отчёт площадки: маркетинг, модель и рекомендация Платона.

        Ничего не считает. Прогон уже сходил и к рынку, и к движку — карточка
        показывает посчитанное, а не запускает то же самое заново.
        """
        market_cabinet.require_cabinet(request)
        return await run_in_threadpool(_stored_report, slug)

    @app.get("/auctions/krt/{slug}/handoff")
    async def auction_krt_handoff(slug: str, request: Request) -> dict[str, Any]:
        """Вводные площадки для калькулятора DevelopAid.

        Это те же самые вводные, которыми посчитан отчёт, — не собранные
        заново. Второй сборщик модели однажды разошёлся бы с первым, и карточка
        с калькулятором показывали бы про одну площадку разное.
        """
        market_cabinet.require_cabinet(request)
        stored = await run_in_threadpool(_stored_report, slug)
        screening = stored.get("screening") or {}
        model = screening.get("model_inputs")
        if not model:
            raise HTTPException(
                status_code=409,
                detail=str(screening.get("reason")
                          or "Модель по этой площадке не собрана — передавать нечего"),
            )
        project = stored.get("project") or {}
        return {
            "slug": slug,
            "name": project.get("name") or slug,
            "computed_at": stored.get("computed_at"),
            "inputs": model.get("inputs") or {},
            "tep": model.get("tep") or {},
            "phasing": model.get("phasing") or {},
            "requirements": screening.get("requirements") or {},
            "assumptions": screening.get("assumptions") or [],
            "exclusions": screening.get("exclusions") or [],
        }

    @app.post("/auctions/krt/{slug}/plato")
    async def auction_krt_plato(slug: str, request: Request) -> dict[str, Any]:
        """Рекомендация Платона по этой площадке — по маркетингу и по модели.

        Спрашивается по требованию и запоминается в том же отчёте: гонять
        модель по всему каталогу в еженедельном прогоне значит платить за сто
        двадцать ответов, из которых прочитают три. Готовый ответ возвращается
        сразу; `refresh=1` спрашивает заново.
        """
        market_cabinet.require_cabinet(request)
        stored = await run_in_threadpool(_stored_report, slug)
        refresh = str(request.query_params.get("refresh") or "").strip() in {"1", "true", "yes"}
        cached = stored.get("plato")
        if cached and not refresh:
            return {"slug": slug, "cached": True, **cached}
        ask = getattr(market, "plato_ask", None) if market is not None else None
        if ask is None:
            raise HTTPException(
                status_code=503,
                detail="Платон недоступен: модуль рынка запущен без движка DevelopAid",
            )
        prompt = _plato_krt_prompt(stored)
        try:
            answer = await run_in_threadpool(ask, prompt, request)
        except HTTPException:
            raise
        except Exception as exc:  # noqa: BLE001
            raise HTTPException(
                status_code=502, detail=f"Платон не ответил: {type(exc).__name__}: {exc}"
            ) from exc
        text = str((answer or {}).get("reply") or (answer or {}).get("text") or "").strip()
        if not text:
            raise HTTPException(status_code=502, detail="Платон вернул пустой ответ")
        payload = {"text": text, "asked_at": int(time.time())}
        stored["plato"] = payload
        rest = {key: value for key, value in stored.items()
                if key not in {"schema_version", "slug", "computed_at"}}
        await run_in_threadpool(
            lambda: krt_ranking.save_report(
                slug, rest, computed_at=stored.get("computed_at")),
        )
        return {"slug": slug, "cached": False, **payload}

    @app.get("/auctions/krt/{slug}/market")
    async def auction_krt_market(
        slug: str,
        request: Request,
        radius_km: float = Query(default=3.0, ge=0.25, le=10.0),
        peers_limit: int = Query(default=12, ge=1, le=20),
        tep_ratios: str = Query(default="", max_length=400),
    ) -> dict[str, Any]:
        """Existing market engine, embedded in the KRT opportunity card.

        `tep_ratios` — доли ТЭП в том же виде, что во вводных страницы
        («apartments: 90/65»). Пусто — наши умолчания.
        """
        market_cabinet.require_cabinet(request)
        if market is None:
            raise HTTPException(status_code=503, detail="Маркетинговый движок не подключён")
        try:
            def build_report_with_model() -> dict[str, Any]:
                report = market.build_report(
                    f"krt:{slug}", radius_km=radius_km, peers_limit=peers_limit,
                    city_reference=False, include_project_totals=True,
                )
                finder = getattr(krt_registry, "find", None)
                project = finder(f"krt:{slug}") if callable(finder) else None
                if project is None:
                    project = next(
                        (item for item in krt_registry.catalogue() if item.get("slug") == slug),
                        None,
                    )
                if core is None:
                    screening = {
                        "available": False,
                        "reason": "Финансовый движок DevelopAid не подключён",
                    }
                else:
                    try:
                        requirements = krt_registry.requirements(slug)
                        screening = build_krt_model_screening(
                            project, report, core, tep_ratios, requirements)
                    except Exception:
                        # Marketing remains useful if a preliminary model cannot
                        # be assembled.  Do not turn an optional screen into a
                        # failure of the existing market report.
                        logger.exception("KRT model screening failed slug=%s", slug)
                        screening = {
                            "available": False,
                            "reason": "Предварительный прогон модели временно недоступен",
                        }
                # Пересчёт из карточки — тот же прогон, что и еженедельный,
                # поэтому и оседает он там же. Иначе «пересчитать сейчас»
                # обновляло бы только экран: на диске остался бы прошлый отчёт,
                # в таблице — прошлый балл, и оба выглядели бы верными.
                # Расчёт на чужих долях в общий рейтинг не идёт. Балл в списке
                # и числа в карточке обязаны быть про одно: сохрани мы сюда
                # разовое допущение аналитика, таблица показывала бы его всем и
                # выглядела бы при этом посчитанной по методике.
                custom_ratios = bool((screening.get("tep_ratios") or {}).get("custom"))
                if project is not None and screening.get("available") is not None \
                        and not custom_ratios:
                    stored = dict(screening)
                    krt_ranking.save_failure_or_report(slug, stored, {
                        "project": project,
                        "market": _market_digest(report),
                        "screening": {key: value for key, value in stored.items()
                                      if key != "market_report"},
                    })
                    krt_ranking.upsert_row(score_row(project, dict(stored)))
                return {**report, "model_screening": screening}

            return await run_in_threadpool(build_report_with_model)
        except (SubjectNotFound, GeocodingError, ValueError) as exc:
            raise HTTPException(status_code=422, detail=str(exc)) from exc
        except RemoteServiceError as exc:
            raise HTTPException(status_code=502, detail=str(exc)) from exc

    @app.get("/auctions/torgi/probe")
    async def auction_torgi_probe(page: int = Query(default=0)) -> dict[str, Any]:
        """Сырой ответ ГИС Торгов и разобранный лот рядом.

        Из песочницы torgi.gov.ru закрыт сетевой политикой, как НСПД: сверить
        поля можно только с ядра. Показываем и сырое, и разобранное вместе —
        расхождение должно быть видно глазами, а не вылезти числом в отчёте.
        """
        return await run_in_threadpool(lambda: TorgiGovAdapter().probe(page))

    @app.get("/auctions/torgi/regions")
    async def auction_torgi_regions(page: int = Query(default=0)) -> dict[str, Any]:
        """Какое имя параметра действительно фильтрует регион.

        Запрос с `dynSubjRF=77,50` приносит чужие области: сервис молча
        игнорирует неизвестный параметр. Перебирать имена вслепую можно
        вечно — здесь они измеряются: «сколько из присланного наше».
        """
        return await run_in_threadpool(lambda: TorgiGovAdapter().probe_regions(page))

    @app.get("/auctions/fedresurs/probe")
    async def auction_fedresurs_probe() -> dict[str, Any]:
        """Что отвечает ЕФРСБ. Разбора нет — сначала ответ, потом код.

        ГИС Торги оказались не тем рынком: живой ответ подтвердил один код вида
        торгов — приватизацию госимущества, — а 44% выборки владельца это
        банкротство, и лежит оно на площадках, которых мы не читаем.

        Из песочницы bankrot.fedresurs.ru закрыт, как НСПД и torgi.gov.ru,
        поэтому проба ходит только с ядра. Разбор по догадке здесь запрещён
        намеренно: ровно так ГИС Торги приехали на прод с тридцатью гаражами.
        """
        return await run_in_threadpool(fedresurs_probe)

    @app.get("/auctions/fedresurs/browser")
    async def auction_fedresurs_browser(
        seconds: float = Query(default=45.0),
        url: str = Query(default=""),
    ) -> dict[str, Any]:
        """Та же проба, но настоящим браузером — и список запросов страницы.

        Простой запрос упёрся в капчу Qrator: корень отдаёт 401 со скриптом
        защиты. Обходить её мы не будем — от этого она и поставлена. Браузер
        проходит вызов штатно, как браузер человека; покажет капчу и ему —
        проба скажет это вслух, а не выдаст пустой источник за отсутствие лотов.

        Нужен не текст страницы, а адреса, по которым она сама ходит за
        данными: у SPA числа приезжают отдельными вызовами бэкенда. Гадать эти
        адреса мы уже пробовали — вышли гаражи.
        """
        # Адрес спрашивается, а не назначается: `/TradeList.aspx` — старый путь,
        # а человек открывает корень, и отказ у них может быть разный. Пустой
        # параметр оставляет страницу поиска торгов, как было.
        return await run_in_threadpool(
            lambda: fedresurs_browser(url=url.strip() or FEDRESURS_SEARCH_PAGE, seconds=seconds))

    @app.get("/auctions/roseltorg/probe")
    async def auction_roseltorg_probe(
        seconds: float = Query(default=45.0, ge=5.0, le=90.0),
    ) -> dict[str, Any]:
        """Чем отвечает раздел «Развитие территории» Росэлторга. Разбора нет.

        Владелец прислал адрес каталога и живой лот на экране: аукцион на
        право заключения договора о КРТ нежилой застройки Москвы, 14,62 га,
        2,4 млрд ₽. Наша разведка Росэлторга этот раздел не спрашивает вовсе —
        она ходит только в поиск процедур по тегам. Это утверждение о нашем
        коде, и его видно в нём; чем отвечает сам раздел, скажет проба.

        Рядом спрашивается НЫНЕШНИЙ путь разведки: «раздел отвечает» и «раздел
        отвечает лучше нынешнего» — разные утверждения, и второе без
        контрольного запроса не проверяется.

        Ответ из двух половин и обе нужны: простой запрос показывает, что
        приходит в HTML, живой браузер — за чем страница ходит сама и что в
        итоге оказалось в DOM. Пустой список запросов при видимых карточках
        значит «всё пришло первым ответом», а не «карточек нет».

        Произвольного адреса здесь нет намеренно: спрашивается ровно тот
        раздел, который открывает владелец.

        Из песочницы roseltorg.ru закрыт, поэтому проба ходит только с ядра.
        """
        return await run_in_threadpool(lambda: roseltorg_probe(seconds=float(seconds)))

    @app.get("/auctions/roseltorg/browser")
    async def auction_roseltorg_browser(
        seconds: float = Query(default=45.0, ge=5.0, le=90.0),
        url: str = Query(default=""),
        save: bool = Query(default=False),
    ) -> dict[str, Any]:
        """Тот же раздел живым браузером — и адреса, за которыми он ходит.

        Если каталог рисует приложение, в простом ответе будет оболочка, а
        числа приедут отдельными вызовами. Их адреса и нужны, чтобы написать
        читателя — не угаданные, а увиденные. `save=1` кладёт страницу файлом:
        читатель пишется по настоящей странице и ею же проверяется.
        """
        target = url.strip() or ROSELTORG_SECTION_URL
        save_to = "/tmp/roseltorg-section.html" if save else ""
        return await run_in_threadpool(
            lambda: roseltorg_section_browser(target, seconds=seconds, save_to=save_to))

    @app.get("/auctions/etp/probe")
    async def auction_etp_probe() -> dict[str, Any]:
        """Что отвечают площадки банкротства. Разбора нет — сначала ответ.

        ЕФРСБ как агрегатор закрыт: капча простому запросу, 403 живому
        браузеру. Читаем сами площадки — пять адресов из таблицы владельца
        дают больше половины банкротных лотов. Но код пишется по ответу, а не
        по догадке: ровно так ГИС Торги приехали на прод с гаражами.
        """
        return await run_in_threadpool(etp_probe)

    @app.get("/auctions/etp/probe/browser")
    async def auction_etp_probe_browser(
        platform: str = Query(default=""),
        seconds: float = Query(default=40.0, ge=5.0, le=90.0),
        save: bool = Query(default=False),
        url: str = Query(default=""),
    ) -> dict[str, Any]:
        """Каталог одной площадки, открытый настоящим браузером.

        Простой запрос показывает оболочку: каталог рисует приложение, а числа
        приезжают отдельными вызовами бэкенда. Их адреса и нужны, чтобы
        написать читателя — не угаданные, а увиденные.

        По одной площадке за вызов намеренно: пять по сорок секунд не уложатся
        ни в один шлюз, а ответ, не дошедший до человека, — это ответ, которого
        нет. Без параметра маршрут перечисляет, что можно спросить.

        `url=` спрашивает заданный адрес: наш сохранённый ведёт куда придётся,
        и это видно только по ответу — у Сбербанк-АСТ каталог банкротства увёл
        редиректом на главную. `save=1` кладёт страницу файлом: читатель
        пишется по настоящей странице и ею же проверяется.
        """
        slug = str(platform or "").strip().lower()
        if not slug:
            return {
                "ok": False,
                "reason": "укажите площадку: ?platform=<имя>",
                "known": {key: name for key, name in sorted(etp_slugs.items())},
                "parsing": "разбора нет: ни одно имя поля не сверено ответом площадки",
            }
        return await run_in_threadpool(
            lambda: etp_probe_browser(slug, seconds=float(seconds),
                                      save=bool(save), url=str(url or "")))

    # Срок на сбор каталога. Шлюз рвёт соединение на шестидесяти секундах и
    # отдаёт свою HTML-страницу; браузер разбирает её как JSON и показывает
    # поломку разбора вместо причины. Сбор ограничивался только числом страниц,
    # то есть объёмом: у ГИС Торгов это сорок страниц по восемь секунд, у
    # Росэлторга и РАД — по запросу на каждый лот по двадцать пять. Сорок
    # секунд оставляют шлюзу запас на ответ, а недособранное называется вслух
    # в строке охвата: источник, до которого не дошли, — это «не знаем», а не
    # «лотов там нет».
    DISCOVERY_BUDGET_SECONDS = 40.0

    @app.get("/auctions/discover")
    async def auction_discover(
        source: str = Query(default="all"),
        include_noise: bool = Query(default=False),
    ) -> dict[str, Any]:
        """Discover current Moscow opportunities from official public ETP catalogues.

        This endpoint does not download/parse every attachment. Full KRT document
        extraction happens only through `/auctions/ingest` for a selected lot.
        """
        try:
            adapters = _discovery_adapters(source)
            service = AuctionSearchService(adapters)
            lots = await run_in_threadpool(
                lambda: service.discover_moscow(
                    include_noise=include_noise,
                    budget_seconds=DISCOVERY_BUDGET_SECONDS))
        except ValueError as exc:
            raise HTTPException(status_code=422, detail=str(exc)) from exc
        except Exception as exc:
            raise HTTPException(status_code=502, detail=f"Не удалось получить публичный каталог ЭТП: {exc}") from exc

        return {
            "source_policy": "official_etp_only",
            "source": source,
            "count": len(lots),
            "quality": service.last_quality_report,
            "coverage": [_coverage_row(adapter) for adapter in adapters],
            "lots": [
                {
                    **_public_lot_dict(lot),
                    "analysis": _analysis_support(lot.source.lot_url),
                    # Соответствие профилю считает сервер по измеренному
                    # эталону сделок: балл, собранный в браузере, был бы
                    # вторым счётом той же величины.
                    "fit": profile_fit(_public_lot_dict(lot)),
                    "quality": catalogue_quality(lot),
                    # Доля в юрлице разбирается сервером: пороги владельца и
                    # разбор активов, объявленные один раз. Собранные на
                    # странице, они были бы вторым ответом на тот же вопрос.
                    "equity": equity_stake.screen(lot),
                    "screening": {
                        **AuctionSearchService.screen_lot(lot),
                        "documents_count": len(lot.documents),
                        "full_document_parse_deferred": True,
                    },
                }
                for lot in lots
            ],
        }

    @app.post("/auctions/export.xlsx")
    async def auction_export(req: AuctionExportRequest) -> Response:
        data = _xlsx(req.rows, req.kind)
        filename = "developaid-krt.xlsx" if req.kind == "krt" else "developaid-auctions.xlsx"
        return Response(content=data, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": f'attachment; filename="{filename}"'})

    @app.post("/auctions/lot-point")
    async def auction_lot_point(req: AuctionLotPointRequest) -> dict[str, Any]:
        if market is None:
            raise HTTPException(status_code=503, detail="Маркетинговый движок не подключён")
        try:
            subject = await run_in_threadpool(market.resolve_subject, req.query)
        except (GeocodingError, RemoteServiceError, RuntimeError) as exc:
            raise HTTPException(status_code=422, detail=str(exc)) from exc
        data = subject.to_dict() if hasattr(subject, "to_dict") else {}
        return {"latitude": data.get("latitude"), "longitude": data.get("longitude"),
                "address": data.get("address") or req.query, "precision": data.get("precision"),
                "nspd_url": data.get("nspd_url") or ""}

    @app.post("/auctions/ingest")
    async def auction_ingest(req: AuctionIngestRequest) -> dict[str, Any]:
        try:
            adapter = _adapter_for(req.url)
            lot = await run_in_threadpool(adapter.fetch_lot, req.url)
            if lot.lot_kind == LotKind.KRT and req.enrich_krt_documents:
                lot = await run_in_threadpool(enrich_krt_from_official_documents, lot)
        except ValueError as exc:
            raise HTTPException(status_code=422, detail=str(exc)) from exc
        except DocumentExtractionError as exc:
            raise HTTPException(status_code=502, detail=str(exc)) from exc
        except Exception as exc:
            # Platform outages/layout changes are upstream failures, not model errors.
            raise HTTPException(status_code=502, detail=f"Не удалось прочитать официальный лот: {exc}") from exc

        screening = AuctionSearchService.screen_lot(lot)
        normalized = lot.to_dict()
        if not req.include_raw:
            normalized.pop("raw", None)
        project_preset = build_project_preset(lot)
        # ОКС и участок — разные объекты ЕГРН. В карточках торгов часто
        # публикуют оба КН; передача обоих в ГлавАПУ заставляет калькулятор
        # принять площадь здания за площадь территории и сложить их. Для
        # имущественного комплекса/незавершёнки оставляем в handoff только
        # земельные КН, подтверждённые НСПД.
        if lot.lot_kind != LotKind.KRT and lot.cadastral_numbers and core is not None:
            try:
                context = await run_in_threadpool(core._land_lot_context, lot.cadastral_numbers)
                _handoff_land_cadastres(project_preset, context)
            except Exception:
                logger.warning("Не удалось отделить КН здания от участка для handoff", exc_info=True)
        return {
            "lot": normalized,
            "developaid_seed": build_developaid_seed(lot),
            "project_preset": project_preset,
            "project_preset_import": {
                "endpoint": "/api/project-presets/import",
                "mode": "preview_then_apply",
                "filled_inputs": project_preset.get("auction_import", {}).get("filled_inputs", {}),
            },
            "screening": {
                **screening,
                "legal_structure": lot.lot_kind.value,
                "requires_krt_terms": lot.lot_kind == LotKind.KRT,
                "krt_documents_complete": (
                    bool(lot.raw.get("krt_extraction_complete")) if lot.lot_kind == LotKind.KRT else None
                ),
                "krt_auth_required": (
                    bool(lot.raw.get("krt_auth_required")) if lot.lot_kind == LotKind.KRT else None
                ),
                "ready_for_financial_model": (
                    bool(lot.krt_program or lot.obligations) and not lot.raw.get("krt_document_warnings")
                    if lot.lot_kind == LotKind.KRT
                    else bool(lot.cadastral_numbers and (lot.current_price_rub is not None or lot.start_price_rub is not None))
                ),
            },
        }

    # Поток поднимается последним: он зовёт `_screen_one`, а замыкание
    # разрешается в момент вызова — стартуй он выше, первая же итерация
    # получила бы NameError.
    if os.getenv("AUCTION_KRT_WEEKLY", "1").strip() not in {"0", "false", "no"}:
        threading.Thread(target=_weekly_ranking, name="krt-weekly", daemon=True).start()
