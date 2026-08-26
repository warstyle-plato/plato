"""Спрос из CRM: что просят покупатели, и чем это отличается от витрины.

«Отчёт нужен для понимания, почему люди не покупают и что не устраивает»
(владелец, 26.08.2026). Прямого ответа в выгрузке нет: поля стадии и причины
отказа в ней не существует, «История касаний» и «Комментарий по касанию» пусты
во всех 573 сделках, а слово «отказ» в комментарии почти всегда означает отказ
дать контакты, отказ от контрольного звонка или отказ спамщику — не отказ от
покупки. Посчитать по нему причины значило бы выдать уверенное число, которого
никто не проверит; тридцать «отказов» на деле оказались двумя.

Поэтому берётся то, что стоит рядом с однозначной единицей и разбирается
числом: запрошенная площадь, бюджет, названная форма оплаты, упоминание
скидки. Ответ на «почему не покупают» собирается не из заявленной причины, а
из РАЗРЫВА: какую площадь и какой бюджет просят — против того, что осталось в
витрине и по какой цене.

Имена и телефоны из комментариев не выходят за пределы этого модуля: наружу
идут только числа сделки. Комментарии — пересказ BitrixGPT, а не слова
клиента, и это сказано в самом ответе: машинный пересказ, выданный за прямую
речь, — то же самое, что чужое число под нашей подписью.
"""
from __future__ import annotations

import re
from typing import Any

from .contracting import _excel_date, _number, _plural, _rows_xlsx, _text

# Лист называется по выгрузке («DEAL_20260826_…»), поэтому опознаётся по шапке,
# а не по имени: имя меняется от выгрузки к выгрузке.
HEADER_MUST_HAVE = ("Название сделки", "Комментарий")
# Где лежала бы причина, если бы её заполняли.
_WHERE_THE_REASON_WOULD_BE = ("касани", "стади", "причин", "статус", "этап")

_COLUMNS = {
    "deal": "ID",
    "kind": "Тип",
    "funnel": "Воронка",
    "created": "Дата создания",
    "comment": "Комментарий",
    "source": "Источник",
    "source_detail": "Дополнительно об источнике",
    "amount": "Сумма",
    "project": "ЖК",
    "payment": "Форма оплаты",
    "booked": "Дата брони",
}

# Число берётся ЦЕЛИКОМ, а не с середины: «26,856 млн руб.» читалось как
# 856 млн — шаблон начинал сопоставление с третьей цифры, и результат оставался
# внутри правдоподобного диапазона, то есть ошибкой не выглядел. Ровно та же
# ошибка уже была в ценах рынка: «3 306 021 ₽/м²» читался как 306 021.
_NUM = r"(?<![\d.,])\d{1,4}(?:[.,]\d{1,3})?"
_DASH = r"[-–—]"
_SQM = r"(?:кв\.?\s*м|м2|м²|метр)"
_AREA_RANGE = re.compile(rf"({_NUM})\s*{_DASH}\s*({_NUM})\s*{_SQM}", re.I)
_AREA_ONE = re.compile(rf"({_NUM})\s*{_SQM}", re.I)
# Площадь комнаты — не площадь квартиры. «Кухня-гостиная около 25 кв. м» и
# «терраса 12 м²» описывают часть лота, и в спрос они не идут.
_PART = re.compile(
    r"(кухн|гостин|лоджи|балкон|террас|санузел|санузл|спальн|прихож|кладов|"
    r"гардероб|кабинет|потолк|коридор|холл|веранд|подсобн)", re.I)
_PART_WINDOW = 40
_MLN = re.compile(rf"({_NUM})\s*{_DASH}\s*({_NUM})\s*(?:млн|миллион)", re.I)
_MLN_ONE = re.compile(rf"({_NUM})\s*(?:млн|миллион)", re.I)
_RUB = re.compile(r"(\d[\d\s  ]{6,})\s*(?:руб|₽)", re.I)

# Границы правдоподобия. Вне их число — это не площадь квартиры и не бюджет,
# а что-то другое: диапазон не заменяет разбор, но отсекает явную чушь.
AREA_RANGE = (18.0, 400.0)
BUDGET_RANGE = (5e6, 900e6)


def _dec(value: float, digits: int = 1) -> str:
    """Дробное число по-русски: запятая, а не точка."""
    return f"{value:.{digits}f}".replace(".", ",")


def _value(text: str) -> float:
    return float(text.replace(",", ".").replace(" ", "").replace(" ", ""))


def _is_part(text: str, at: int) -> bool:
    """Слово-часть считается только в СВОЁМ обороте, а не в окне назад.

    «Терраса 12 м², квартира 90 кв.м» — здесь 90 относится к квартире, и окно
    в сорок знаков назад видело бы «террасу» и выбрасывало верное число.
    Поэтому взгляд назад упирается в ближайшую запятую или точку.
    """
    window = text[max(0, at - _PART_WINDOW):at]
    cut = max(window.rfind(sign) for sign in ",.;:!?()\n")
    return bool(_PART.search(window[cut + 1:] if cut >= 0 else window))


def areas_asked(comment: str) -> list[float]:
    """Площади квартиры, названные в комментарии. Комнаты сюда не идут."""
    text = _text(comment)
    out: list[float] = []
    taken: list[tuple[int, int]] = []
    for hit in _AREA_RANGE.finditer(text):
        if _is_part(text, hit.start()):
            continue
        low, high = _value(hit.group(1)), _value(hit.group(2))
        if AREA_RANGE[0] <= low <= high <= AREA_RANGE[1]:
            out += [low, high]
            taken.append(hit.span())
    for hit in _AREA_ONE.finditer(text):
        if any(start <= hit.start() < stop for start, stop in taken):
            continue
        if _is_part(text, hit.start()):
            continue
        value = _value(hit.group(1))
        if AREA_RANGE[0] <= value <= AREA_RANGE[1]:
            out.append(value)
    return sorted(set(out))


def budgets_asked(comment: str) -> list[float]:
    """Бюджеты в рублях. «45 млн» и «45 000 000 руб.» — одно и то же число."""
    text = _text(comment)
    out: list[float] = []
    taken: list[tuple[int, int]] = []
    for hit in _MLN.finditer(text):
        low, high = _value(hit.group(1)) * 1e6, _value(hit.group(2)) * 1e6
        if BUDGET_RANGE[0] <= low <= high <= BUDGET_RANGE[1]:
            out += [low, high]
            taken.append(hit.span())
    for hit in _MLN_ONE.finditer(text):
        if any(start <= hit.start() < stop for start, stop in taken):
            continue
        value = _value(hit.group(1)) * 1e6
        if BUDGET_RANGE[0] <= value <= BUDGET_RANGE[1]:
            out.append(value)
    for hit in _RUB.finditer(text):
        value = _value(hit.group(1))
        if BUDGET_RANGE[0] <= value <= BUDGET_RANGE[1]:
            out.append(value)
    return sorted(set(out))


_WANTS = (
    ("рассрочка", re.compile(r"рассрочк", re.I)),
    ("ипотека", re.compile(r"ипотек", re.I)),
    ("100% оплата", re.compile(r"100\s*%|полн(?:ая|ой)\s+оплат", re.I)),
    ("скидка", re.compile(r"скидк", re.I)),
    ("отделка", re.compile(r"отделк", re.I)),
    ("паркинг", re.compile(r"паркинг|машиномест|машино-мест", re.I)),
)


def read_demand(data: bytes) -> dict[str, Any]:
    """Сделки CRM: только числа. Имена и телефоны отсюда не выходят."""
    import openpyxl  # noqa: PLC0415 — тяжёлый импорт нужен только здесь

    book = openpyxl.load_workbook(__import__("io").BytesIO(data), read_only=True, data_only=True)
    sheet = None
    for name in book.sheetnames:
        rows = _rows_xlsx(data, name)
        if not rows:
            continue
        header = [_text(x) for x in rows[0]]
        if all(title in header for title in HEADER_MUST_HAVE):
            sheet = name
            break
    if sheet is None:
        raise KeyError(
            "в книге нет листа сделок CRM: шапки с "
            + " и ".join(f"«{x}»" for x in HEADER_MUST_HAVE)
            + f" не нашлось ни на одном из листов {book.sheetnames}")

    rows = _rows_xlsx(data, sheet)
    header = [_text(x) for x in rows[0]]
    missing: list[str] = []
    at: dict[str, int] = {}
    for key, title in _COLUMNS.items():
        found = next((i for i, name in enumerate(header) if name.startswith(title)), None)
        if found is None:
            missing.append(title)
        else:
            at[key] = found

    deals = []
    for row in rows[1:]:
        def cell(key: str) -> Any:
            place = at.get(key)
            return row[place] if place is not None and place < len(row) else None

        if not _text(cell("deal")) and not _text(cell("comment")):
            continue
        comment = _text(cell("comment"))
        created = _excel_date(cell("created"))
        areas = areas_asked(comment)
        budgets = budgets_asked(comment)
        deals.append({
            # Номер сделки — не человек: он нужен, чтобы не считать одну сделку
            # дважды, и больше ни для чего.
            "deal": _text(cell("deal")),
            "month": created.strftime("%Y-%m") if created else "",
            "funnel": _text(cell("funnel")),
            "source": _text(cell("source")) or _text(cell("source_detail")),
            "project": _text(cell("project")),
            "amount": _number(cell("amount")),
            "booked": bool(_excel_date(cell("booked"))),
            "area_min": areas[0] if areas else None,
            "area_max": areas[-1] if areas else None,
            "budget_min": budgets[0] if budgets else None,
            "budget_max": budgets[-1] if budgets else None,
            "wants": [name for name, pattern in _WANTS if pattern.search(comment)],
        })
    # Пустая колонка — находка, а не пустяк: «История касаний» и «Комментарий
    # по касанию» есть в шапке и не заполнены ни разу, и именно в них лежал бы
    # прямой ответ на «что не устраивает». Молча пропустив их, мы бы искали
    # причину отказа там, где её нет, и нашли бы.
    # Пустыми называются не все колонки подряд, а те, в которых лежал бы ответ
    # на «что не устраивает»: «Компания» пуста и никому не мешает, а пустая
    # «История касаний» — находка. Список пустых колонок вообще был бы шумом,
    # в котором находка потерялась бы.
    empty = []
    for place, title in enumerate(header):
        if not any(word in title.lower() for word in _WHERE_THE_REASON_WOULD_BE):
            continue
        if not any(_text(row[place]) for row in rows[1:] if place < len(row)):
            empty.append(title)
    return {"sheet": sheet, "deals": deals, "missing": missing,
            "empty_columns": empty, "rows": len(rows) - 1}


# ---------------------------------------------------------------------------
# Спрос против витрины
# ---------------------------------------------------------------------------


def _overlaps(low: float | None, high: float | None, band: dict[str, Any]) -> bool:
    if low is None:
        return False
    return (high if high is not None else low) >= band["low"] and low < band["high"]


def demand_summary(deals: list[dict[str, Any]], bands: list[dict[str, Any]] | None = None,
                   read: dict[str, Any] | None = None) -> dict[str, Any]:
    """Что просят — против того, что осталось в витрине.

    Сделка попадает в КАЖДУЮ полосу, которую задевает её запрос: «55–60 м²» —
    это запрос сразу к двум полосам, и приписать его одной значит выдумать,
    к какой именно. Сколько сделок попало больше чем в одну, стоит в ответе:
    сумма по полосам больше числа сделок, и без этой строки она читается как
    завышенный спрос.
    """
    bands = list(bands or [])
    asked = [d for d in deals if d.get("area_min")]
    budgets = sorted(d["budget_max"] for d in deals if d.get("budget_max"))
    areas = sorted(d["area_min"] for d in asked)

    def middle(values: list[float]) -> float | None:
        if not values:
            return None
        half = len(values) // 2
        return values[half] if len(values) % 2 else (values[half - 1] + values[half]) / 2

    rows = []
    multi = 0
    over = 0
    top = max((b["high"] for b in bands), default=0.0)
    for deal in asked:
        hits = sum(1 for band in bands if _overlaps(deal["area_min"], deal["area_max"], band))
        if hits > 1:
            multi += 1
        if top and deal["area_min"] > top:
            over += 1
    for band in bands:
        want = [d for d in asked if _overlaps(d["area_min"], d["area_max"], band)]
        rows.append({
            "band": band["band"],
            "low": band["low"], "high": band["high"],
            "asked": float(len(want)),
            "asked_share": len(want) / len(asked) if asked else None,
            "left_units": band.get("left_units"),
            "left_share": band.get("left_share"),
            "sold_share": band.get("sold_share"),
            "price_per_sqm": band.get("book_price_per_sqm") or None,
        })
    wants: dict[str, int] = {}
    for deal in deals:
        for name in deal.get("wants") or []:
            wants[name] = wants.get(name, 0) + 1

    notes = [
        "Комментарии в выгрузке — пересказ BitrixGPT, а не слова клиента: "
        "разобранные из них числа стоит читать как порядок величины.",
    ]
    for title in (read or {}).get("empty_columns") or []:
        notes.append(f"Колонка «{title}» есть в шапке и не заполнена ни разу.")
    notes.append(
        "Причину отказа мы не считаем: поля стадии и причины в выгрузке нет, "
        "а слово «отказ» в комментарии почти всегда означает отказ дать контакты "
        "или отказ от контрольного звонка. Счёт по нему дал бы уверенное число, "
        "которого никто не проверит.")
    if multi:
        notes.append(
            f"{multi} {_plural(multi, 'запрос', 'запроса', 'запросов')} из {len(asked)} "
            "задевают больше одной полосы — "
            "сумма по полосам поэтому больше числа сделок.")
    if over:
        notes.append(
            f"{over} {_plural(over, 'запрос', 'запроса', 'запросов')} крупнее самого "
            f"большого лота проекта ({_dec(top)} м²) — в полосы они не попадают вовсе.")
    return {
        "deals": float(len(deals)),
        "with_area": float(len(asked)),
        "with_budget": float(len(budgets)),
        "area_median": middle(areas),
        "budget_median": middle(budgets),
        "budget_low": budgets[0] if budgets else None,
        "budget_high": budgets[-1] if budgets else None,
        "bands": rows,
        "wants": [{"want": name, "deals": float(count)}
                  for name, count in sorted(wants.items(), key=lambda x: -x[1])],
        "notes": notes,
    }
