"""Свод продаж действующего проекта: контрактация, эскроу, каналы, рассрочка.

Рыночный отчёт кабинета смотрит наружу — соседи, цены, темпы. Здесь смотрят
внутрь: сколько продано своим проектом, чем, кем и на каких условиях. Источник
один — файл ЦФ проекта, и в нём два листа, которые нельзя путать.

**«Контрактация»** — обогащённый реестр договоров: вариант оплаты, брокер и его
ставка, премии отдела продаж, помесячный график поступлений на эскроу. Это то,
чего нет больше нигде.

**«1С_Факт»** — проводки. Дт 76.06 / Кт 86.02 — заключение ДДУ, Дт 008.01 —
поступление на эскроу. Два разных факта на одном листе, и фильтр по счёту
обязателен: сложить их значит посчитать продажи дважды.

Правила, выведенные из живого файла (Кутузов Сити, 25.08.2026).

- **Сторно — расторжение.** Договор, у которого начисления по 76.06 сходятся в
  ноль, расторгнут: у одного это подписано словом «Расторжение», у другого нет,
  но восемь месяцев спустя это не исправление ввода. «Контрактация» таких
  договоров не показывает вовсе, и «нет строки» неотличимо от «не было».

- **Расторжение живёт в своём месяце.** Договор заключён в сентябре, расторгнут
  в феврале: сентябрь продал — это правда тогда и остаётся правдой в отчёте.
  Вычистить его из истории задним числом значит менять уже отчитанные месяцы.

- **Возврат эскроу — отдельное движение.** Между расторжением и возвратом
  прошло шесть дней, и вернулось 99,0 из 99,2 млн ₽. Сложить их в одну строку
  значит спрятать этот остаток.

- **Доля вознаграждения считается на своей выборке.** «Процент от фактического
  наполнения» по всему проекту даёт 4,50%, а по сделкам, за которые платим, —
  7,32%: в первый знаменатель попал эскроу прямых продаж, где комиссии нет.
  Показатель, посчитанный по одним сделкам и делённый на другие, не значит
  ничего.

- **Средняя цена по всем строкам — смесь.** Квартиры 660,6 тыс ₽/м², коммерция
  769,3, машино-места 376,7. Общая 652,6 не описывает ни один продукт.

- **Ноль вознаграждения бывает «не знаем».** У одного брокера 95,3 млн продаж и
  пустая комиссия. Настоящий ноль и незаполненное поле обязаны выглядеть
  по-разному.

- **ФИО покупателей наружу не выходят.** Свод агрегатный, имена в нём не нужны,
  а вопрос Платону уходит внешнему поставщику модели. Остаётся признак
  «юрлицо» — оптовая сделка тянет среднюю цену и должна быть видна.
"""

from __future__ import annotations

import datetime
import io
import re
from typing import Any, Iterable

SHEET_CONTRACTS = "Контрактация"
SHEET_LEDGER = "1С_Факт"

# Счета 1С: заключение договора и поступление на эскроу — разные факты.
ACCOUNT_CONTRACT = "76.06"
ACCOUNT_ESCROW = "008"

_HEADER_ROW = 6

# Колонки «Контрактации» по шапке. Номера не зашиты: шапка читается, и
# ненайденное имя уходит в предупреждение — лист чужой и может поехать.
_COLUMNS = {
    "quarter": "Квартал",
    "year": "Год",
    "month": "Месяц",
    "product": "Объект недвижимости",
    "unit": "Корпус",
    "kind": "Тип объекта недвижимости",
    "finish": "Вид отделки",
    "area": "Проектная S",
    "contract": "Договор",
    "contract_type": "Тип договора",
    "state": "Состояние договора",
    "amount": "Сумма договор",
    "date": "Дата договора",
    "buyer": "Покупатель",
    "units": "Шт",
    "price_per_sqm": "Цена ДДУ за кв. м",
    "payment_variant": "Вариант оплаты",
    "sales_bonus_paid": "Оплачено премии ОП ",
    "broker_fee": "Брокеры",
    "broker_rate": "% брокера",
    "broker_name": "Наименование брокера",
    "paid_share": "% оплаты",
    "left_to_pay": "остаток к оплате",
    "escrow_total": "оплачено эскроу всего",
}

# Признак юридического лица в имени покупателя. Само имя наружу не идёт.
_COMPANY = re.compile(r"\b(ООО|АО|ПАО|ЗАО|ИП|НАО|ГК)\b|\bООО\b", re.I)


def _text(value: Any) -> str:
    """Значение ячейки строкой, без висячих пробелов и переносов.

    У листов ЦФ подписи приезжают то с хвостовым пробелом («Эскроу, тыс. руб. »),
    то с переносом внутри: сверять их целиком значит терять строку на чужой
    описке.
    """
    return " ".join(str(value if value is not None else "").split())


def _excel_date(value: Any) -> datetime.date | None:
    """Дата бывает и числом, и строкой — в одной колонке.

    В «1С_Факт» обычные строки несут «09.09.2025», а сторнирующие — 46065.
    Наивный разбор даёт либо ошибку, либо 1970 год.
    """
    if value is None or value == "":
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    if isinstance(value, (int, float)):
        return datetime.date(1899, 12, 30) + datetime.timedelta(days=int(value))
    text = str(value).strip()
    match = re.search(r"(\d{2})\.(\d{2})\.(\d{4})", text)
    if match:
        day, month, year = (int(x) for x in match.groups())
        try:
            return datetime.date(year, month, day)
        except ValueError:
            return None
    if re.fullmatch(r"\d+(?:\.\d+)?", text):
        return datetime.date(1899, 12, 30) + datetime.timedelta(days=int(float(text)))
    return None


def _number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(" ", " ").replace(" ", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def _rows(data: bytes, sheet: str) -> list[list[Any]]:
    """Строки листа. Формат решает содержимое, а не расширение.

    ЦФ проекта приходит в .xlsb — openpyxl его не открывает вовсе, и все наши
    читатели факта ходят через него. Поэтому здесь два пути, и выбирается он по
    сигнатуре файла, а не по имени.
    """
    return _rows_xlsb(data, sheet) if _looks_xlsb(data) else _rows_xlsx(data, sheet)


def _looks_xlsb(data: bytes) -> bool:
    import zipfile
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            return "xl/workbook.bin" in archive.namelist()
    except Exception:  # noqa: BLE001
        return False


def _rows_xlsb(data: bytes, sheet: str) -> list[list[Any]]:
    from pyxlsb import open_workbook
    out: list[list[Any]] = []
    with open_workbook(io.BytesIO(data)) as book:
        if sheet not in book.sheets:
            raise KeyError(f"в книге нет листа «{sheet}»: {book.sheets}")
        with book.get_sheet(sheet) as page:
            for row in page.rows():
                out.append([cell.v for cell in row])
    return out


def _rows_xlsx(data: bytes, sheet: str) -> list[list[Any]]:
    from openpyxl import load_workbook
    book = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        if sheet not in book.sheetnames:
            raise KeyError(f"в книге нет листа «{sheet}»: {book.sheetnames}")
        return [list(row) for row in book[sheet].iter_rows(values_only=True)]
    finally:
        book.close()


def read_contracts(data: bytes) -> dict[str, Any]:
    """Реестр договоров листа «Контрактация».

    Шапка читается, а не зашивается номерами: лист чужой, и колонки в нём уже
    добавлялись. Не найденное имя уходит в `missing` — свод, посчитанный без
    колонки, о которой никто не сказал, выглядит исправным.
    """
    rows = _rows(data, SHEET_CONTRACTS)
    if len(rows) < _HEADER_ROW:
        return {"rows": [], "project": "", "missing": ["лист короче шапки"]}
    header = [str(x).strip() if x is not None else "" for x in rows[_HEADER_ROW - 1]]
    index = {}
    missing = []
    for key, title in _COLUMNS.items():
        wanted = title.strip()
        found = next((i for i, name in enumerate(header) if name.strip() == wanted), None)
        if found is None:
            missing.append(title)
        else:
            index[key] = found

    project = ""
    for row in rows[:_HEADER_ROW]:
        for i, cell in enumerate(row):
            if isinstance(cell, str) and "Название проекта" in cell:
                rest = [x for x in row[i + 1:] if isinstance(x, str) and x.strip()]
                if rest:
                    project = rest[0].strip()
                break
        if project:
            break

    # Помесячный график поступлений на эскроу лежит колонками с датами в
    # шапке. Их состав меняется от файла к файлу — растёт вправо по мере
    # продаж, — поэтому берутся все, чья шапка разбирается как дата.
    escrow_columns = [(i, _excel_date(name)) for i, name in enumerate(header)
                      if _excel_date(name) is not None and i > (index.get("escrow_total") or 0)]

    out: list[dict[str, Any]] = []
    for row in rows[_HEADER_ROW:]:
        def cell(key: str) -> Any:
            position = index.get(key)
            return row[position] if position is not None and position < len(row) else None

        if not cell("product"):
            continue
        buyer = str(cell("buyer") or "")
        signed = _excel_date(cell("date"))
        out.append({
            # Квартал у части строк пуст — берём из даты договора, а не теряем
            # тринадцать сделок в «прочем».
            "month": (signed.strftime("%Y-%m") if signed
                      else (_excel_date(cell("month")) or datetime.date(1900, 1, 1)).strftime("%Y-%m")),
            "signed": signed.isoformat() if signed else None,
            "product": str(cell("product") or "").strip(),
            "unit": str(cell("unit") or "").strip(),
            "kind": str(cell("kind") or "").strip(),
            "area": _number(cell("area")),
            "contract": str(cell("contract") or "").strip(),
            "contract_type": str(cell("contract_type") or "").strip(),
            "state": str(cell("state") or "").strip(),
            "amount": _number(cell("amount")),
            "units": _number(cell("units")) or 1.0,
            "payment_variant": str(cell("payment_variant") or "").strip(),
            "broker": str(cell("broker_name") or "").strip(),
            "broker_fee": _number(cell("broker_fee")),
            # Пустая ставка и ставка ноль — разные ответы.
            "broker_rate": None if cell("broker_rate") in (None, "") else _number(cell("broker_rate")),
            "sales_bonus_paid": _number(cell("sales_bonus_paid")),
            "escrow_paid": _number(cell("escrow_total")),
            # Имя покупателя наружу не идёт: остаётся только признак.
            "company_buyer": bool(_COMPANY.search(buyer)),
            "escrow_schedule": [
                {"month": when.strftime("%Y-%m"), "amount": _number(row[i])}
                for i, when in escrow_columns
                if i < len(row) and _number(row[i])
            ],
        })
    return {"rows": out, "project": project, "missing": missing}


def read_ledger(data: bytes) -> dict[str, Any]:
    """Проводки листа «1С_Факт»: контрактация и эскроу порознь.

    Договор, у которого начисления по 76.06 сходятся в ноль, расторгнут.
    «Контрактация» такие строки не показывает вовсе, и без этого листа
    расторжение неотличимо от «договора не было».
    """
    rows = _rows(data, SHEET_LEDGER)
    if not rows:
        return {"contracts": {}, "escrow": {}, "terminated": []}
    header = [str(x).strip() if x is not None else "" for x in rows[0]]

    def column(title: str) -> int | None:
        return next((i for i, name in enumerate(header) if name == title), None)

    at = {key: column(title) for key, title in (
        ("date", "Дата"), ("debit", "Счет Дт"), ("contract", "Субконто2 Дт"),
        ("amount", "Сумма"), ("content", "Содержание"))}
    contracts: dict[str, dict[str, Any]] = {}
    escrow: dict[str, dict[str, float]] = {}
    for row in rows[1:]:
        def cell(key: str) -> Any:
            position = at.get(key)
            return row[position] if position is not None and position < len(row) else None

        account = str(cell("debit") or "")
        name = str(cell("contract") or "").strip()
        if not name:
            continue
        number = _contract_number(name)
        amount = _number(cell("amount"))
        if account.startswith(ACCOUNT_CONTRACT):
            item = contracts.setdefault(number, {"amount": 0.0, "content": "", "reversed_on": None})
            item["amount"] += amount
            item["content"] = item["content"] or str(cell("content") or "").strip()
            if amount < 0:
                signed = _excel_date(cell("date"))
                item["reversed_on"] = signed.isoformat() if signed else None
        elif account.startswith(ACCOUNT_ESCROW):
            # Приход и возврат — разные факты, и складывать их в сальдо нельзя:
            # у расторгнутого договора 10 + 89 − 99 даёт ноль, и возврат в
            # 99 млн ₽ исчезает вместе с самим событием.
            side = "paid" if amount >= 0 else "returned"
            item = escrow.setdefault(number, {"paid": 0.0, "returned": 0.0})
            item[side] += abs(amount)
    terminated = [
        {"contract": number, "object": item["content"], "on": item["reversed_on"],
         "escrow_returned": (escrow.get(number) or {}).get("returned", 0.0),
         "escrow_paid": (escrow.get(number) or {}).get("paid", 0.0)}
        for number, item in contracts.items()
        if item["reversed_on"] and abs(item["amount"]) < 1.0
    ]
    return {"contracts": contracts, "escrow": escrow, "terminated": terminated}


def _contract_number(text: str) -> str:
    match = re.search(r"№\s*(\S+?)\s*от", str(text or ""))
    return match.group(1) if match else str(text or "").strip()


def payment_variant(text: str) -> str:
    """Вариант оплаты — свободный текст, и правило обязано это признавать.

    Рядом лежат «Рассрочка 20% ПВ и далее по 200 000 в месяц», «1.0», «0.1»,
    «5 млн на эскроу», «50% и по 100 000». Неопознанное остаётся собой, а не
    сваливается в «прочее»: иначе треть портфеля уедет в безымянную корзину.
    """
    value = " ".join(str(text or "").split())
    low = value.lower().replace("ё", "е")
    if not value:
        return "не указан"
    if "ипотек" in low:
        return "ипотека"
    if "рассроч" in low or "пв" in low.split() or re.search(r"по\s*\d[\d\s ]*\s*(в месяц|000)", low):
        return "рассрочка"
    if value in ("1", "1.0", "100%", "100"):
        return "100% оплата"
    share = re.fullmatch(r"(\d{1,3})\s*%\s*оплат\w*", low)
    if share:
        return f"{share.group(1)}% оплата"
    return value


_SIZE_BANDS = (("студия", 0.0, 28.0), ("1к", 28.0, 45.0), ("2к", 45.0, 70.0),
               ("3к", 70.0, 100.0), ("4к+", 100.0, float("inf")))


def size_band(area: float) -> str:
    for name, low, high in _SIZE_BANDS:
        if low <= area < high:
            return name
    return "—"


def _totals(rows: Iterable[dict[str, Any]]) -> dict[str, float]:
    rows = list(rows)
    amount = sum(r["amount"] for r in rows)
    area = sum(r["area"] for r in rows)
    escrow = sum(r["escrow_paid"] for r in rows)
    fee = sum(r["broker_fee"] for r in rows)
    # Премия своего отдела продаж лежит отдельным полем от брокерской комиссии,
    # и без неё канал «напрямую» показывал ровно ноль — то есть «бесплатно».
    # Это не бесплатно, это другая строка расходов, и считать её должен тот же
    # `_totals`, что и остальное: посчитанная на экране, она была бы вторым
    # счётом той же величины.
    bonus = sum(r["sales_bonus_paid"] for r in rows)
    return {
        "contracts": float(len(rows)),
        "units": sum(r["units"] for r in rows),
        "area": area,
        "amount": amount,
        "escrow": escrow,
        "broker_fee": fee,
        "price_per_sqm": amount / area if area else 0.0,
        "escrow_share": escrow / amount if amount else 0.0,
        # Обе базы — на СВОЕЙ выборке. Доля «от наполнения», посчитанная от
        # эскроу всего проекта, включает прямые продажи, где комиссии нет:
        # выходит 4,50% вместо 7,32%, и показатель делится не на те сделки.
        "fee_of_sales": fee / amount if amount else 0.0,
        "fee_of_escrow": fee / escrow if escrow else 0.0,
        "sales_bonus": bonus,
        "bonus_of_sales": bonus / amount if amount else 0.0,
        # Полная стоимость канала суммой, а не только долей: сложить комиссию с
        # премией на экране значит посчитать ту же величину второй раз.
        "cost": fee + bonus,
        # Полная стоимость канала — комиссия и премия вместе: у брокера это
        # почти всегда комиссия, у своего отдела — почти всегда премия, и
        # сравнивать их по одному из двух полей значит сравнивать разное.
        "cost_of_sales": (fee + bonus) / amount if amount else 0.0,
    }


# Варианты оплаты, которые мы понимаем. Всё остальное — не «прочие условия
# сделки», а дефект заполнения CRM (решение владельца, 25.08.2026): «10ПВ +10%
# через три месяца+1», «5 млн на эскроу», «0.1» и пустая ячейка описывают не
# рынок, а то, как заполнили карточку. Восемь строк по одной сделке читаются
# как разнообразие условий; одна строка с числом сделок и суммой читается как
# то, чем является, — и примеры показываются подсказкой при наведении.
KNOWN_VARIANTS = ("рассрочка", "ипотека")
CRM_DEFECT = "дефект заполнения CRM"
_CRM_EXAMPLES = 5


def _payment_structure(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    buckets: dict[str, dict[str, Any]] = {}
    for row in rows:
        name = payment_variant(row["payment_variant"])
        recognised = name in KNOWN_VARIANTS or bool(re.fullmatch(r"\d{1,3}% оплата", name))
        key = name if recognised else CRM_DEFECT
        item = buckets.setdefault(key, {
            "variant": key, "name": key, "count": 0, "area": 0.0, "amount": 0.0,
            "escrow": 0.0, "recognised": recognised, "examples": []})
        item["count"] += 1
        item["area"] += row["area"]
        item["amount"] += row["amount"]
        item["escrow"] += row["escrow_paid"]
        if not recognised:
            # Пример — то, что вписали в карточку; сколько их всего, видно по
            # count, поэтому список ограничен и это сказано числом.
            shown = item["examples"]
            text = row["payment_variant"] or "пусто"
            if text not in [x["text"] for x in shown] and len(shown) < _CRM_EXAMPLES:
                shown.append({"text": text, "contract": row["contract"],
                              "amount": row["amount"]})
    for item in buckets.values():
        item["filled"] = item["escrow"] / item["amount"] if item["amount"] else None
        item["examples_shown"] = len(item["examples"])
    return sorted(buckets.values(), key=lambda x: (x["name"] == CRM_DEFECT, -x["amount"]))


# ---------------------------------------------------------------------------
# План продаж: наша финмодель и модель банка.
#
# Свод продаж отвечает на «что продали». Без второй половины — «сколько
# собирались» — он не говорит ничего о том, идём мы по плану или отстаём
# (владелец, 26.08.2026). Оба плана лежат в той же выгрузке ЦФ, что и
# контрактация, поэтому и читаются тем же вызовом: просить загрузить один файл
# дважды значит однажды получить два разных файла и показать их как один
# проект.
# ---------------------------------------------------------------------------

FM_SHEET = "Продажи ФМ_new Банников"
BANK_SHEET = "Модель банка_new"

# Лист ФМ идёт парами колонок: строка дат общая, а строка ниже говорит, план
# это или факт. Разбирать по чётности колонок нельзя — в листе есть пустые
# столбцы-разделители, и счёт сбивается на первом же.
_FM_DATE_ROW = 3
_FM_KIND_ROW = 4
_FM_PRODUCTS = {
    "Квартира": "Квартира",
    "Кладовые": "Кладовая",
    "Машиноместа": "Машиноместо",
    "Коммерческие площади": "Коммерческие площади",
    "Итого": "Итого",
}
# Названия строк внутри продукта. Ключ — начало подписи: у листа встречаются
# висячие пробелы («Эскроу, тыс. руб. »), и сверять целиком значит терять строку
# на чужой описке.
_FM_METRICS = (
    ("Эскроу", "escrow"),
    ("Заключенные договоры", "amount"),
    ("Цена продажи", "price"),
    ("м2", "area"),
    ("шт", "units"),
)
# Лист считает в тысячах рублей, свод — в рублях. Приводим здесь, а не на
# экране: две единицы под одним именем никто не заметит.
_FM_THOUSANDS = ("escrow", "amount")


def quarter_of(month: str) -> str:
    """«2026-07» → «2026 Q3». Вид тот же, что в книге банка."""
    try:
        year, number = str(month).split("-")[:2]
        return f"{year} Q{(int(number) - 1) // 3 + 1}"
    except (ValueError, IndexError):
        return str(month)


def _fm_metric(label: str) -> str | None:
    text = _text(label)
    for prefix, name in _FM_METRICS:
        if text.startswith(prefix):
            return name
    return None


def read_fm_plan(data: bytes) -> dict[str, Any]:
    """Помесячный план и факт нашей финмодели по продуктам.

    Возвращает `{"months": [...], "plan": {...}, "fact": {...}}`, где внутри —
    продукт → месяц → показатели. Чего в листе нет, того нет и здесь: пустая
    ячейка не становится нулём, иначе «не заполнено» читалось бы как «ноль
    продаж».
    """
    rows = _rows(data, FM_SHEET)
    if len(rows) < _FM_KIND_ROW:
        raise KeyError(f"лист «{FM_SHEET}» пуст")
    dates = rows[_FM_DATE_ROW - 1]
    kinds = rows[_FM_KIND_ROW - 1]
    columns: list[tuple[int, str, str]] = []
    for index, kind in enumerate(kinds):
        name = _text(kind).lower()
        if name not in ("план", "факт"):
            continue
        moment = _excel_date(dates[index] if index < len(dates) else None)
        if moment is None:
            continue
        columns.append((index, name, moment.strftime("%Y-%m")))
    if not columns:
        raise KeyError(f"в листе «{FM_SHEET}» не нашлось колонок «план» и «факт»")

    out: dict[str, dict[str, dict[str, dict[str, float]]]] = {"план": {}, "факт": {}}
    product = ""
    for row in rows[_FM_KIND_ROW:]:
        label = _text(row[2] if len(row) > 2 else "")
        if label in _FM_PRODUCTS:
            product = _FM_PRODUCTS[label]
            continue
        if label.startswith("Продажи по годам"):
            # Годовой блок ниже месячного: те же подписи, другой горизонт.
            # Смешать их значит удвоить каждый показатель.
            break
        metric = _fm_metric(label) if product else None
        if not metric:
            continue
        for index, kind, month in columns:
            raw = row[index] if index < len(row) else None
            if raw is None or _text(raw) == "":
                continue
            value = _number(raw)
            if metric in _FM_THOUSANDS:
                value *= 1000.0
            out[kind].setdefault(product, {}).setdefault(month, {})[metric] = value
    months = sorted({month for _, _, month in columns})
    return {"sheet": FM_SHEET, "months": months,
            "plan": out["план"], "fact": out["факт"]}


# Модель банка — квартальная, и приводить её к месяцам мы не станем: разложить
# квартал по трём месяцам можно тремя способами, и любой будет нашей выдумкой,
# а не планом банка. Сравнение идёт по кварталам, и это сказано вслух.
_BANK_PERIOD_ROW = 1
# Строки выручки банка. Начало подписи, а не точное совпадение: у продуктов
# хвосты разные («… квартиры», «… Машиноместа (м.м шт)»).
BANK_REVENUE_PREFIX = "Продажи с учётом рассрочки"
_BANK_QUARTER = re.compile(r"^(20\d{2})\s*Q([1-4])$", re.I)


def read_bank_plan(data: bytes) -> dict[str, Any]:
    """Квартальный план банка: подписи строк как есть, значения по кварталам."""
    rows = _rows(data, BANK_SHEET)
    if not rows:
        raise KeyError(f"лист «{BANK_SHEET}» пуст")
    header = rows[_BANK_PERIOD_ROW - 1]
    quarters: list[tuple[int, str]] = []
    for index, cell in enumerate(header):
        match = _BANK_QUARTER.match(_text(cell))
        if match:
            quarters.append((index, f"{match.group(1)} Q{match.group(2)}"))
    if not quarters:
        raise KeyError(f"в листе «{BANK_SHEET}» не нашлось кварталов вида «2026 Q1»")
    lines: list[dict[str, Any]] = []
    for row in rows[_BANK_PERIOD_ROW:]:
        label = _text(row[1] if len(row) > 1 else "")
        if not label:
            continue
        values = {}
        for index, quarter in quarters:
            raw = row[index] if index < len(row) else None
            if raw is None or _text(raw) == "":
                continue
            values[quarter] = _number(raw)
        if values:
            lines.append({"label": label, "values": values})
    # Выручка плана банка — сумма строк «Продажи с учётом рассрочки» по всем
    # продуктам. Складываем ЗДЕСЬ, а не на экране: сложенная в браузере
    # колонка была бы вторым счётом той же величины. Какие строки сложены,
    # едет рядом — иначе число не проверить.
    revenue: dict[str, float] = {}
    summed: list[str] = []
    for line in lines:
        if not line["label"].startswith(BANK_REVENUE_PREFIX):
            continue
        summed.append(line["label"])
        for quarter, value in line["values"].items():
            # Лист банка в тысячах рублей, свод — в рублях.
            revenue[quarter] = revenue.get(quarter, 0.0) + value * 1000.0
    return {"sheet": BANK_SHEET, "quarters": [q for _, q in quarters], "lines": lines,
            "revenue_by_quarter": revenue, "revenue_rows": summed}


def summarise(contracts: dict[str, Any], ledger: dict[str, Any] | None = None) -> dict[str, Any]:
    """Свод: динамика, структура оплаты, каналы, вознаграждение, расторжения."""
    rows = contracts.get("rows") or []
    ledger = ledger or {}

    months: dict[str, dict[str, Any]] = {}
    for row in rows:
        month = months.setdefault(row["month"], {"month": row["month"], "by_product": {}})
        product = month["by_product"].setdefault(row["product"], {"units": 0.0, "area": 0.0, "amount": 0.0})
        product["units"] += row["units"]
        product["area"] += row["area"]
        product["amount"] += row["amount"]
    dynamics = []
    for month in sorted(months):
        item = months[month]
        area = sum(p["area"] for p in item["by_product"].values())
        amount = sum(p["amount"] for p in item["by_product"].values())
        item.update({
            "units": sum(p["units"] for p in item["by_product"].values()),
            "area": area,
            "amount": amount,
            # Удельное считает тот, кто считает выручку. Посчитанное на экране
            # было бы вторым счётом той же величины: разойдись они, обе строки
            # выглядели бы верными.
            "price_per_sqm": amount / area if area else 0.0,
        })
        dynamics.append(item)

    products = {}
    for row in rows:
        products.setdefault(row["product"], []).append(row)
    by_product = [{"product": key, **_totals(items)} for key, items in products.items()]
    by_product.sort(key=lambda item: -item["amount"])

    # Факт по кварталам: план банка квартальный, и сравнивать его с месяцами
    # значит сравнивать разные величины. Складываем на сервере — на экране это
    # был бы второй счёт той же выручки.
    quarters: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        quarters.setdefault(quarter_of(row["month"]), []).append(row)
    by_quarter = [{"quarter": key, **_totals(items)} for key, items in sorted(quarters.items())]

    by_payment = _payment_structure(rows)

    channels: dict[str, list] = {}
    for row in rows:
        channels.setdefault(row["broker"] or "", []).append(row)
    by_channel = []
    for name, items in channels.items():
        block = {"channel": name or "напрямую", "own": not name, **_totals(items)}
        # Ноль вознаграждения при непустой ставке — «не заполнено», а не «даром».
        block["fee_unknown"] = bool(name) and block["broker_fee"] == 0
        by_channel.append(block)
    by_channel.sort(key=lambda item: -item["amount"])

    flats = [r for r in rows if r["product"] == "Квартира"]
    bands: dict[str, list] = {}
    for row in flats:
        bands.setdefault(size_band(row["area"]), []).append(row)
    by_size = [{"band": key, **_totals(items)} for key, items in bands.items()]
    by_size.sort(key=lambda item: [b[0] for b in _SIZE_BANDS].index(item["band"])
                 if item["band"] in [b[0] for b in _SIZE_BANDS] else 99)

    broker_rows = [r for r in rows if r["broker"]]
    own_rows = [r for r in rows if not r["broker"]]
    return {
        "project": contracts.get("project") or "",
        "missing": contracts.get("missing") or [],
        "total": _totals(rows),
        "dynamics": dynamics,
        "by_quarter": by_quarter,
        "by_product": by_product,
        "by_payment": by_payment,
        "by_channel": by_channel,
        "by_size": by_size,
        "brokers": _totals(broker_rows),
        "own_sales": _totals(own_rows),
        "sales_bonus_paid": sum(r["sales_bonus_paid"] for r in rows),
        "company_buyers": sum(1 for r in rows if r["company_buyer"]),
        "terminated": ledger.get("terminated") or [],
    }


# ---------------------------------------------------------------------------
# Пул проекта: сколько всего продаётся и сколько из этого продано
# ---------------------------------------------------------------------------
#
# «Продано 3 594 м²» без второй половины — не показатель, а число: пятая часть
# проекта и половина проекта выглядят одинаково (владелец, 26.08.2026). База
# берётся из уже прочитанных источников, а не заводится третьим:
#
#   — метры, лоты и ожидаемая выручка по продуктам — из плана нашей финмодели:
#     его горизонт покрывает весь проект, и сумма плана и есть пул;
#   — квартирография в лотах — из книги финмодели, лист «график продажи_1»:
#     у квартир в плане ФМ строки «шт» нет вовсе, они планируются метрами.
#
# Оба считаются здесь, а не на экране: доля, посчитанная в браузере, была бы
# вторым счётом той же величины, и разойдись она с первым — обе выглядели бы
# верными.

POOL_SHEET = "график продажи_1"
# Подписи блока физических объёмов книги. Слева направо: продано, оплачено,
# всего. Имена продуктов — те, что стоят в книге.
_POOL_VOLUME_TITLE = "Все объекты"
_POOL_VOLUME_PRODUCTS = ("КВ", "ПСН", "М/М", "КЛД")
# Один продукт зовётся в трёх источниках по-разному: CRM пишет «Машиноместа»,
# план финмодели — «Машиноместо», книга — «М/М». Пока имена не сведены к одному,
# пул машино-мест не находится вовсе, и доля показывается пустой при полном
# наборе данных — то есть «не знаем» вместо посчитанного.
_PRODUCT_ALIASES = {
    "машиноместа": "Машиноместо",
    "машиноместо": "Машиноместо",
    "м/м": "Машиноместо",
    "кладовые": "Кладовая",
    "кладовая": "Кладовая",
    "клд": "Кладовая",
    "квартира": "Квартира",
    "квартиры": "Квартира",
    "кв": "Квартира",
    "коммерческие площади": "Коммерческие площади",
    "псн": "Коммерческие площади",
}


def product_name(label: str) -> str:
    """Имя продукта, одинаковое во всех источниках."""
    text = _text(label)
    return _PRODUCT_ALIASES.get(text.lower(), text)


_POOL_SOLD = "Продано, шт"
_POOL_PAID = "Оплачено, шт"
_POOL_PCT = "Продано, %"
_POOL_AREA = "Продано, кв.м"
_POOL_AMOUNT = "Продано, т.руб"
_POOL_PRICE = "Цена"
# Полоса вида «28,3 - 40» или «85 - 168,6»: границы с запятой в дробной части.
_BAND_RANGE = re.compile(r"^\s*(\d+(?:[.,]\d+)?)\s*[-–—]\s*(\d+(?:[.,]\d+)?)\s*$")


def _band_bounds(label: str) -> tuple[float, float] | None:
    hit = _BAND_RANGE.match(_text(label))
    if not hit:
        return None
    low = float(hit.group(1).replace(",", "."))
    high = float(hit.group(2).replace(",", "."))
    return (low, high) if high > low else None


def read_pool(data: bytes) -> dict[str, Any]:
    """Квартирография и физические объёмы проекта из книги финмодели.

    Колонка «всего» в книге без заголовка: она стоит между «Оплачено, шт» и
    «Продано, %». Угадывать её нельзя — она доказывается самой книгой: доля
    «Продано, %» обязана сойтись с «Продано, шт» делённым на неё, и сходится
    на каждой полосе. Не сошлась — это уходит в `missing`, а не считается по
    похожей колонке.
    """
    rows = _rows(data, POOL_SHEET)
    missing: list[str] = []

    header_at = None
    for number, row in enumerate(rows):
        titles = [_text(x) for x in row]
        if _POOL_SOLD in titles and _POOL_PCT in titles:
            header_at = number
            break
    bands: list[dict[str, Any]] = []
    book_sold = book_pool = 0.0
    if header_at is None:
        missing.append(f"в листе «{POOL_SHEET}» не нашлось шапки «{_POOL_SOLD}»")
    else:
        titles = [_text(x) for x in rows[header_at]]
        at = {name: titles.index(name) for name in
              (_POOL_SOLD, _POOL_PAID, _POOL_PCT, _POOL_AREA, _POOL_AMOUNT, _POOL_PRICE)
              if name in titles}
        pool_at = at[_POOL_PAID] + 1 if _POOL_PAID in at and _POOL_PCT in at else None
        if pool_at is None or pool_at >= at[_POOL_PCT]:
            missing.append(
                f"в листе «{POOL_SHEET}» между «{_POOL_PAID}» и «{_POOL_PCT}» нет колонки пула")
            pool_at = None
        # Полосы стоят под шапкой: подпись-диапазон в любой из первых колонок.
        for row in rows[header_at + 1:]:
            label = next((_text(x) for x in row[:6] if _band_bounds(_text(x))), "")
            bounds = _band_bounds(label)
            if not bounds:
                if bands:
                    # Ряд полос кончился — дальше итог и служебные строки.
                    break
                continue

            def value(name: str) -> float:
                place = at.get(name)
                return _number(row[place]) if place is not None and place < len(row) else 0.0

            in_pool = (_number(row[pool_at]) if pool_at is not None and pool_at < len(row) else 0.0)
            sold = value(_POOL_SOLD)
            share = value(_POOL_PCT)
            # Доказательство колонки: книга сама печатает долю проданного.
            if in_pool and share and abs(sold / in_pool - share) > 0.01:
                missing.append(
                    f"полоса «{label}»: «{_POOL_PCT}» {share:.3f} не сходится с "
                    f"{sold:.0f}/{in_pool:.0f} — колонка пула опознана неверно")
                bands = []
                break
            bands.append({
                "band": label,
                "low": bounds[0],
                "high": bounds[1],
                "pool_units": in_pool,
                "book_sold_units": sold,
                "book_sold_area": value(_POOL_AREA),
                # Лист считает в тысячах рублей — приводим здесь, а не на
                # экране: две единицы под одним именем никто не заметит.
                "book_sold_amount": value(_POOL_AMOUNT) * 1000.0,
                "book_price_per_sqm": value(_POOL_PRICE) * 1000.0,
            })
        book_sold = sum(b["book_sold_units"] for b in bands)
        book_pool = sum(b["pool_units"] for b in bands)

    # Физические объёмы читаются КОЛОНКАМИ своего блока, а не поиском подписи
    # по всему листу: подписи КВ/ПСН/М/М/КЛД встречаются в листе трижды — в
    # этом блоке, во вспомогательной табличке слева и ниже в «Мониторинге
    # денежных средств», где те же имена стоят над рублями. Рубли, принятые за
    # метры, выглядят как метры.
    volumes: list[dict[str, Any]] = []
    head_at = None
    for number, row in enumerate(rows):
        if not any(_text(x).startswith(_POOL_VOLUME_TITLE) for x in row):
            continue
        for candidate in rows[number:number + 3]:
            titles = [_text(x) for x in candidate]
            if "Продано" in titles and "ВСЕГО" in titles:
                head_at = rows.index(candidate, number)
                columns = (titles.index("Продано"), titles.index("Оплачено")
                           if "Оплачено" in titles else titles.index("Продано") + 1,
                           titles.index("ВСЕГО"))
                break
        if head_at is not None:
            break
    if head_at is None:
        missing.append(f"в листе «{POOL_SHEET}» не нашлось блока «{_POOL_VOLUME_TITLE}»")
    else:
        label_at = min(columns) - 1
        for row in rows[head_at + 1:]:
            label = _text(row[label_at]) if label_at < len(row) else ""
            if label not in _POOL_VOLUME_PRODUCTS:
                break
            volumes.append({
                "product": product_name(label),
                "sold": _number(row[columns[0]]) if columns[0] < len(row) else 0.0,
                "paid": _number(row[columns[1]]) if columns[1] < len(row) else 0.0,
                "pool": _number(row[columns[2]]) if columns[2] < len(row) else 0.0,
            })
    return {"sheet": POOL_SHEET, "bands": bands, "volumes": volumes,
            "book_sold_units": book_sold, "book_pool_units": book_pool,
            "missing": missing}


def plan_pool(fm: dict[str, Any]) -> dict[str, dict[str, float]]:
    """Пул по продуктам из плана финмодели: весь горизонт — это весь проект."""
    out: dict[str, dict[str, float]] = {}
    for product, by_month in (fm.get("plan") or {}).items():
        if product == "Итого":
            continue
        block = out.setdefault(product_name(product), {"amount": 0.0, "area": 0.0, "units": 0.0})
        for values in by_month.values():
            for key in ("amount", "area", "units"):
                block[key] += float(values.get(key) or 0.0)
    return out


def pool_progress(summary: dict[str, Any], rows: list[dict[str, Any]],
                  fm: dict[str, Any] | None, pool: dict[str, Any] | None) -> dict[str, Any]:
    """Продано из скольких — по продуктам, по проекту и по квартирографии.

    Считается один раз и здесь. Доля, посчитанная на экране, — второй счёт той
    же величины; в этом модуле такое уже стоило нам показанного нуля вместо
    премии своего отдела.
    """
    missing: list[str] = []
    base = plan_pool(fm) if fm else {}
    if not base:
        missing.append("план финмодели не прочитан — доли от объёма проекта показать не из чего")

    # Квартиры в плане ФМ живут метрами: строки «шт» у них в листе нет вовсе.
    # Число лотов приносит квартирография книги, и другого источника у него нет.
    flat_pool_units = float((pool or {}).get("book_pool_units") or 0.0)

    products = []
    for row in summary.get("by_product") or []:
        name = product_name(row["product"])
        have = base.get(name) or {}
        units = float(have.get("units") or 0.0)
        if not units and name == "Квартира":
            units = flat_pool_units
        products.append({
            "product": name,
            "sold_amount": row["amount"], "pool_amount": float(have.get("amount") or 0.0),
            "sold_area": row["area"], "pool_area": float(have.get("area") or 0.0),
            "sold_units": row["contracts"], "pool_units": units,
            "amount_share": row["amount"] / have["amount"] if have.get("amount") else None,
            "area_share": row["area"] / have["area"] if have.get("area") else None,
            "units_share": row["contracts"] / units if units else None,
        })
    total = summary.get("total") or {}
    pool_amount = sum(float((v or {}).get("amount") or 0.0) for v in base.values())
    pool_area = sum(float((v or {}).get("area") or 0.0) for v in base.values())
    whole = {
        "sold_amount": total.get("amount") or 0.0, "pool_amount": pool_amount,
        "sold_area": total.get("area") or 0.0, "pool_area": pool_area,
        "sold_units": total.get("contracts") or 0.0,
        "amount_share": (total.get("amount") or 0.0) / pool_amount if pool_amount else None,
        "area_share": (total.get("area") or 0.0) / pool_area if pool_area else None,
    }
    # Два источника на один пул: план финмодели и книга. Совпадать они не
    # обязаны — книга снята на свою дату, — но выбрать один и промолчать
    # нельзя: доля, посчитанная от 75 мест, и доля от 73 выглядят одинаково.
    # Считаем по плану, расхождение называем.
    for volume in (pool or {}).get("volumes") or []:
        name = product_name(volume["product"])
        planned = base.get(name) or {}
        for key, what in (("units", "лотов"), ("area", "м²")):
            mine = float(planned.get(key) or 0.0)
            theirs = float(volume.get("pool") or 0.0)
            if not mine or not theirs or key == "area":
                continue
            if abs(mine - theirs) > 0.5:
                missing.append(
                    f"пул «{name}»: план финмодели {mine:.0f} {what}, книга {theirs:.0f} — "
                    f"доли посчитаны по плану")
    bands = _absorption(rows, pool, missing)
    return {"products": products, "total": whole, "bands": bands,
            "plan_sheet": (fm or {}).get("sheet") or "", "pool_sheet": (pool or {}).get("sheet") or "",
            "missing": missing + list((pool or {}).get("missing") or [])}


def _absorption(rows: list[dict[str, Any]], pool: dict[str, Any] | None,
                missing: list[str]) -> list[dict[str, Any]]:
    """Вымывание: доля полосы в пуле против её доли в продажах.

    Полосы берутся из книги — они разбиты так, как разбит сам проект, и наши
    «студия/1к/2к» рядом с ними были бы вторым делением одной величины.
    Проданное считается по НАШИМ договорам: у них есть площадь каждой сделки,
    а колонка книги — это её собственный срез на свою дату, и она стоит рядом
    отдельной проверкой, а не подменяет наш счёт.
    """
    bands = list((pool or {}).get("bands") or [])
    if not bands:
        return []
    flats = [r for r in (rows or []) if r.get("product") == "Квартира"]
    if not flats:
        missing.append("квартирография показана без наших договоров: строк по квартирам нет")
    pool_units = sum(b["pool_units"] for b in bands)
    out = []
    unplaced = []
    counted = 0
    # Верхняя граница последней полосы — граница проекта, а не начало
    # следующей: самая большая квартира в книге ровно 168,6 м², и полуоткрытая
    # полоса теряла её вместе с её договором.
    top = max(b["high"] for b in bands)
    for band in bands:
        mine = [r for r in flats
                if band["low"] <= r["area"] < band["high"]
                or (band["high"] >= top and r["area"] == top)]
        counted += len(mine)
        out.append({
            **band,
            "sold_units": float(len(mine)),
            "sold_area": sum(r["area"] for r in mine),
            "sold_amount": sum(r["amount"] for r in mine),
            "left_units": band["pool_units"] - len(mine),
            "pool_share": band["pool_units"] / pool_units if pool_units else None,
        })
    sold_units = sum(b["sold_units"] for b in out)
    left = sum(b["left_units"] for b in out)
    for band in out:
        band["sold_share"] = band["sold_units"] / sold_units if sold_units else None
        band["skew"] = (band["sold_share"] - band["pool_share"]
                        if band["sold_share"] is not None and band["pool_share"] is not None
                        else None)
        # Остаток витрины: чем полоса представлена в том, что ещё не продано.
        band["left_share"] = band["left_units"] / left if left else None
    if flats and counted != len(flats):
        unplaced = [r["area"] for r in flats
                    if not any(b["low"] <= r["area"] < b["high"] or
                               (b["high"] >= top and r["area"] == top) for b in bands)]
        missing.append(
            f"вне полос книги осталось {len(unplaced)} договор(ов) по квартирам "
            f"(площади {', '.join(f'{a:.1f}' for a in sorted(unplaced)[:5])}"
            f"{' и др.' if len(unplaced) > 5 else ''}) — в вымывании их нет")
    return out


# ---------------------------------------------------------------------------
# Выводы под блоками
# ---------------------------------------------------------------------------
#
# «Нет выводов под блоками» (владелец, 26.08.2026). Управленцу нужна не ещё
# одна таблица — у него они есть, — а сказанное словами: что показывает
# картинка. Считаются выводы ЗДЕСЬ, рядом с числами, а не в браузере: фраза,
# собранная на экране из своей арифметики, — это второй счёт той же величины.


def _pct(value: float | None, digits: int = 1) -> str:
    return "—" if value is None else f"{value * 100:.{digits}f}%".replace(".", ",")


def _mln(value: float | None, digits: int = 1) -> str:
    return "—" if value is None else f"{value / 1e6:,.{digits}f}".replace(",", " ").replace(".", ",")


def conclusions(summary: dict[str, Any]) -> dict[str, str]:
    """По фразе на блок. Блока без числа не бывает: нечего сказать — молчим."""
    out: dict[str, str] = {}
    total = summary.get("total") or {}
    pool = summary.get("pool") or {}
    whole = pool.get("total") or {}

    if whole.get("amount_share") is not None:
        left = float(whole.get("pool_amount") or 0.0) - float(whole.get("sold_amount") or 0.0)
        out["pool"] = (
            f"Продано {_pct(whole['amount_share'])} ожидаемой выручки проекта — "
            f"{_mln(whole.get('sold_amount'))} из {_mln(whole.get('pool_amount'))} млн ₽; "
            f"впереди ещё {_mln(left)} млн ₽.")

    dynamics = [m for m in (summary.get("dynamics") or []) if m.get("amount")]
    if len(dynamics) >= 4:
        last3 = dynamics[-3:]
        before = dynamics[:-3]
        recent = sum(m["amount"] for m in last3) / len(last3)
        earlier = sum(m["amount"] for m in before) / len(before) if before else 0.0
        move = "выше" if recent > earlier else "ниже"
        share = abs(recent - earlier) / earlier if earlier else None
        out["dynamics"] = (
            f"Последние три месяца — {_mln(recent)} млн ₽ в месяц, это {move} "
            f"среднего по предыдущим {len(before)} ({_mln(earlier)} млн ₽)"
            + (f", на {_pct(share, 0)}" if share is not None else "") + ".")

    bands = pool.get("bands") or []
    if bands:
        hottest = max(bands, key=lambda b: b.get("skew") if b.get("skew") is not None else -9)
        coldest = min(bands, key=lambda b: b.get("skew") if b.get("skew") is not None else 9)
        if hottest.get("skew") is not None and coldest.get("skew") is not None:
            out["bands"] = (
                f"Вымывается полоса {hottest['band']} м²: {_pct(hottest['pool_share'])} пула и "
                f"{_pct(hottest['sold_share'])} продаж. Медленнее всего уходит {coldest['band']} м² — "
                f"{_pct(coldest['pool_share'])} пула против {_pct(coldest['sold_share'])} продаж. "
                f"В остатке витрины она уже {_pct(coldest.get('left_share'))}: чем дальше, тем "
                f"крупнее то, что остаётся показывать.")

    products = [p for p in (summary.get("by_product") or []) if p.get("amount")]
    if products:
        first = products[0]
        out["products"] = (
            f"{first['product']} — {_pct(first['amount'] / total['amount'] if total.get('amount') else None)} "
            f"выручки, всего продуктов в продажах {len(products)}.")

    payment = summary.get("by_payment") or []
    if payment:
        defect = next((x for x in payment if x.get("variant") == CRM_DEFECT), None)
        known = [x for x in payment if x.get("variant") != CRM_DEFECT]
        parts = ", ".join(f"{x['variant']} — {_pct(x['amount'] / total['amount'] if total.get('amount') else None)}"
                          for x in known)
        line = f"Условия оплаты: {parts}." if parts else ""
        if defect:
            line += (f" У {int(defect.get('count') or 0)} договоров условие в CRM не разобрать "
                     f"({_pct(defect['amount'] / total['amount'] if total.get('amount') else None)} выручки) — "
                     f"это дефект заполнения, а не рыночное условие.")
        out["payment"] = line.strip()

    brokers = summary.get("brokers") or {}
    own = summary.get("own_sales") or {}
    if brokers.get("amount") or own.get("amount"):
        out["channels"] = (
            f"Чужие каналы принесли {_pct(brokers.get('amount', 0) / total['amount'] if total.get('amount') else None)} "
            f"выручки и стоили {_pct(brokers.get('cost_of_sales'), 2)} от своих продаж; свой отдел — "
            f"{_pct(own.get('amount', 0) / total['amount'] if total.get('amount') else None)} выручки при "
            f"{_pct(own.get('cost_of_sales'), 2)}. Вознаграждения всего — "
            f"{_mln((brokers.get('cost') or 0) + (own.get('cost') or 0))} млн ₽.")

    fm = summary.get("fm_plan") or {}
    plan = (fm.get("plan") or {}).get("Итого") or (fm.get("plan") or {}).get("Квартира") or {}
    pairs = [(m["month"], float(m["amount"]), float((plan.get(m["month"]) or {}).get("amount") or 0.0))
             for m in (summary.get("dynamics") or []) if (plan.get(m["month"]) or {}).get("amount")]
    if pairs:
        behind = [p for p in pairs if p[1] < p[2]]
        gap = sum(p[1] - p[2] for p in pairs)
        out["fm"] = (
            f"Ниже плана финмодели {len(behind)} месяцев из {len(pairs)}; накопленное "
            f"{'опережение' if gap >= 0 else 'отставание'} {_mln(abs(gap))} млн ₽.")

    want = summary.get("demand") or {}
    rows = [b for b in (want.get("bands") or [])
            if b.get("asked_share") is not None and b.get("left_share") is not None]
    if rows:
        # Разрыв, а не заявленная причина: где спроса больше, чем витрины, там
        # людям нечего показать, и это ответ на «почему не покупают».
        short = max(rows, key=lambda b: b["asked_share"] - b["left_share"])
        spare = min(rows, key=lambda b: b["asked_share"] - b["left_share"])
        out["demand"] = (
            f"Разобрано {int(want.get('with_area') or 0)} запросов по площади и "
            f"{int(want.get('with_budget') or 0)} по бюджету из "
            f"{int(want.get('deals') or 0)} сделок CRM; медиана запроса — "
            f"{(want.get('area_median') or 0):.0f} м² и "
            f"{_mln(want.get('budget_median'))} млн ₽. "
            f"Больше всего не хватает полосы {short['band']} м²: "
            f"{_pct(short['asked_share'])} спроса против {_pct(short['left_share'])} витрины. "
            f"Наоборот — {spare['band']} м²: {_pct(spare['asked_share'])} спроса при "
            f"{_pct(spare['left_share'])} витрины.")

    bank = summary.get("bank_plan") or {}
    quarters = bank.get("revenue_by_quarter") or {}
    fact = {q["quarter"]: float(q["amount"]) for q in (summary.get("by_quarter") or [])}
    common = [q for q in quarters if q in fact]
    if common:
        gap = sum(fact[q] - float(quarters[q]) for q in common)
        out["bank"] = (
            f"По {len(common)} общим кварталам факт {'выше' if gap >= 0 else 'ниже'} плана банка на "
            f"{_mln(abs(gap))} млн ₽.")
    return out


# ---------------------------------------------------------------------------
# Факт против двух планов — на одном графике
# ---------------------------------------------------------------------------
#
# «Факт против ФМ и плана банка — на одном наглядном графике с ценами и
# метрами» (владелец, 26.08.2026). Общая шкала у трёх рядов ровно одна —
# квартал: план банка квартальный, и раскладывать его по месяцам мы не станем,
# сделать это можно тремя способами, и любой будет нашей выдумкой.
#
# Складывается здесь: сумма месяцев плана до квартала, посчитанная в браузере,
# была бы вторым счётом того же плана.


def plan_comparison(summary: dict[str, Any]) -> dict[str, Any]:
    """Ряд кварталов: факт, план финмодели, план банка — в ₽, м² и ₽/м²."""
    fm = summary.get("fm_plan") or {}
    want = summary.get("demand") or {}
    rows = [b for b in (want.get("bands") or [])
            if b.get("asked_share") is not None and b.get("left_share") is not None]
    if rows:
        # Разрыв, а не заявленная причина: где спроса больше, чем витрины, там
        # людям нечего показать, и это ответ на «почему не покупают».
        short = max(rows, key=lambda b: b["asked_share"] - b["left_share"])
        spare = min(rows, key=lambda b: b["asked_share"] - b["left_share"])
        out["demand"] = (
            f"Разобрано {int(want.get('with_area') or 0)} запросов по площади и "
            f"{int(want.get('with_budget') or 0)} по бюджету из "
            f"{int(want.get('deals') or 0)} сделок CRM; медиана запроса — "
            f"{(want.get('area_median') or 0):.0f} м² и "
            f"{_mln(want.get('budget_median'))} млн ₽. "
            f"Больше всего не хватает полосы {short['band']} м²: "
            f"{_pct(short['asked_share'])} спроса против {_pct(short['left_share'])} витрины. "
            f"Наоборот — {spare['band']} м²: {_pct(spare['asked_share'])} спроса при "
            f"{_pct(spare['left_share'])} витрины.")

    bank = summary.get("bank_plan") or {}
    plan = (fm.get("plan") or {}).get("Итого") or {}
    # У «Итого» финмодели нет метров: строка «м2» есть у продуктов. Метры плана
    # складываются по продуктам, деньги берутся из «Итого», если оно есть.
    by_month: dict[str, dict[str, float]] = {}
    for product, months in (fm.get("plan") or {}).items():
        for month, values in months.items():
            block = by_month.setdefault(month, {"amount": 0.0, "area": 0.0})
            if product != "Итого":
                block["area"] += float(values.get("area") or 0.0)
                if not plan:
                    block["amount"] += float(values.get("amount") or 0.0)
    for month, values in plan.items():
        by_month.setdefault(month, {"amount": 0.0, "area": 0.0})["amount"] = \
            float(values.get("amount") or 0.0)

    fact_quarter = {q["quarter"]: q for q in (summary.get("by_quarter") or [])}
    plan_quarter: dict[str, dict[str, float]] = {}
    for month, values in by_month.items():
        block = plan_quarter.setdefault(quarter_of(month), {"amount": 0.0, "area": 0.0})
        block["amount"] += values["amount"]
        block["area"] += values["area"]
    bank_quarter = {key: float(value) for key, value in
                    (bank.get("revenue_by_quarter") or {}).items()}

    names = sorted(set(fact_quarter) | set(plan_quarter) | set(bank_quarter))
    # Горизонт плана — весь проект, факт — только прошедшее. Показывать пустой
    # хвост из будущих кварталов незачем: он читается как провал продаж.
    last = max([q for q in fact_quarter], default="")
    # Незакрытый квартал против полного планового — это не провал продаж, а
    # разные отрезки времени. Сколько месяцев факта в квартале, считается
    # здесь и говорится на экране.
    months_in: dict[str, int] = {}
    for row in summary.get("dynamics") or []:
        months_in[quarter_of(row["month"])] = months_in.get(quarter_of(row["month"]), 0) + 1
    rows = []
    for name in names:
        if last and name > last:
            break
        fact = fact_quarter.get(name) or {}
        if not any((fact.get("amount"), plan_quarter.get(name, {}).get("amount"),
                    bank_quarter.get(name))):
            continue
        planned = plan_quarter.get(name) or {}
        rows.append({
            "label": name,
            "fact_amount": fact.get("amount"),
            "fact_area": fact.get("area"),
            "fact_price": fact.get("price_per_sqm") or None,
            "fm_amount": planned.get("amount") or None,
            "fm_area": planned.get("area") or None,
            "fm_price": (planned["amount"] / planned["area"]
                         if planned.get("amount") and planned.get("area") else None),
            "bank_amount": bank_quarter.get(name),
            "months": months_in.get(name, 0),
            "partial": bool(fact.get("amount")) and months_in.get(name, 0) < 3,
        })
    return {"quarters": rows, "fm_sheet": fm.get("sheet") or "",
            "bank_sheet": bank.get("sheet") or "",
            # У плана банка есть только деньги: метров и цены в его строках нет.
            # Сказать это надо вслух — пропавшая линия читается как ноль.
            "bank_metrics": ["amount"],
            "bank_rows": bank.get("revenue_rows") or []}
