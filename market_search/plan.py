"""План продаж из книги ПЛАТО: помесячно, отдельно факт и отдельно план.

Отчёт о рынке отвечает на вопрос «как у соседей». Второй вопрос, ради которого
его читают, — «а как у нас против того, что обещали». Ответ на него лежит не в
источнике, а в собственной финансовой модели проекта, на листе «План продаж_утв»:
там помесячно и штуки, и метры, и цена, а строки помечены «факт», «оперфакт»
или ничем — последнее и означает план.

Разбор нарочно терпимый к форме. Книга живая, её правят руками: лист могут
переименовать, столбцы сдвинуть, шапку поставить на другую строку. Поэтому
столбцы ищутся по подписям, а не по номерам, и любая непонятая строка
пропускается молча — но если не нашлось ни одной, разбор говорит об этом вслух,
а не возвращает пустой график.

Факт из книги и факт из «Пульса» — разные числа, и это нормально: банк считает
ДДУ по дате регистрации, книга — по дате сделки, между ними недели. Показывать
их надо рядом, а не выбирать одно.
"""

from __future__ import annotations

import datetime
import io
import re
from typing import Any


SHEET_HINTS = ("план продаж_утв", "план продаж утв", "план продаж")
FACT_MARKS = ("факт", "оперфакт")


class PlanNotFound(ValueError):
    """Книга разобрана, но плана продаж в ней не нашлось."""


def _norm(value: Any) -> str:
    return " ".join(str(value or "").split()).casefold()


# Книга в формате .xlsb хранит дату числом дней от 30.12.1899, и вернуть его
# как есть значит потерять весь помесячный ряд молча: строки просто не станут
# месяцами, а разбор скажет «плана не нашлось» на книге, где план есть.
_EXCEL_EPOCH = datetime.date(1899, 12, 30)
# Ниже этого числа это уже не дата, а количество: 20 000 дней — это 1954 год,
# раньше начала любого проекта, а «шт» в первой колонке встречается.
_EXCEL_SERIAL_MIN = 20000
_EXCEL_SERIAL_MAX = 80000


def _month(value: Any) -> str | None:
    if hasattr(value, "year") and hasattr(value, "month"):
        return f"{value.year:04d}-{value.month:02d}"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        serial = float(value)
        if _EXCEL_SERIAL_MIN <= serial <= _EXCEL_SERIAL_MAX:
            moment = _EXCEL_EPOCH + datetime.timedelta(days=int(serial))
            return f"{moment.year:04d}-{moment.month:02d}"
        return None
    text = str(value or "").strip()
    match = re.match(r"^(\d{4})-(\d{2})", text)
    if match:
        return f"{match.group(1)}-{match.group(2)}"
    match = re.match(r"^(\d{2})\.(\d{2})\.(\d{4})", text)
    if match:
        return f"{match.group(3)}-{match.group(2)}"
    return None


def _number(value: Any) -> float | None:
    if value in (None, "", "-"):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _pick_sheet_name(names: list[str], *, hints: tuple[str, ...] = ()) -> str:
    """Имя листа плана. Одинаково для обоих форматов — выбор один на разбор."""
    wanted = hints or SHEET_HINTS
    for hint in wanted:
        for name in names:
            if _norm(name) == hint:
                return name
    for hint in wanted:
        for name in names:
            if _norm(name).startswith(hint):
                return name
    if not hints:
        for name in names:
            if "план" in _norm(name) and "прод" in _norm(name):
                return name
    raise PlanNotFound("В книге нет нужного листа продаж")


# Шапка бывает не на третьей строке: во втором шаблоне отчёта над таблицей
# стоит блок с площадью квартир и составом очередей, и подписи начинаются
# только на девятнадцатой.
_HEADER_SEARCH_DEPTH = 40

# Соседние колонки называются похоже, и перепутать их значит получить число,
# которое выглядит правдоподобно. «В РЕАЛИЗАЦИИ, кв.м» — это ОСТАТОК, а не
# продажи: взятый вместо продаж, он даёт 13 428 м² за месяц и не бросается в
# глаза. «Нарастающим итогом» — сумма с начала, а не месяц.
_NOT_A_MONTHLY_SALE = ("в реализации", "нарастающ", "остат", "итого")


def _column_label(marks: dict[str, int], *, exact: tuple[str, ...] = (),
                  contains: tuple[tuple[str, ...], ...] = ()) -> int | None:
    """Колонка по подписи: сначала точное совпадение, потом по словам.

    Точное вперёд намеренно: в первом шаблоне подписи короткие («шт»), во
    втором длинные («ОБЪЕМ ПРОДАЖ,шт»). Общий поиск по вхождению нашёл бы в
    первом шаблоне не ту колонку.
    """
    for wanted in exact:
        if wanted in marks:
            return marks[wanted]
    for words in contains:
        for label, position in marks.items():
            if any(bad in label for bad in _NOT_A_MONTHLY_SALE):
                continue
            if all(word in label for word in words):
                return position
    return None


def _columns(rows: list[tuple]) -> tuple[int, dict[str, int]]:
    """Найти строку шапки и колонки по подписям, а не по номерам.

    Книга живая, шаблонов у отчёта уже два, и разные компании ведут его
    по-своему. Поэтому ищутся подписи, а не позиции: сдвинутый столбец иначе
    молча приносит соседнее число.
    """
    for index, row in enumerate(rows[:_HEADER_SEARCH_DEPTH]):
        # Подписи повторяются: сначала блок «ВСЕГО», за ним такие же колонки по
        # каждому корпусу. Нужен первый набор, поэтому запоминается первое
        # вхождение, а не последнее — иначе отчёт показывал бы один корпус
        # вместо всего проекта, и разница не бросалась бы в глаза.
        marks: dict[str, int] = {}
        for position, cell in enumerate(row):
            label = _norm(cell)
            if label and label not in marks:
                marks[label] = position
        units = _column_label(
            marks, exact=("шт",), contains=(("объем продаж", "шт"), ("продаж", "шт")))
        area = _column_label(
            marks, exact=("кв.м.", "кв.м", "м2", "м²"),
            contains=(("объем продаж", "кв.м"), ("продаж", "кв.м"), ("продаж", "м2")))
        price = _column_label(
            marks, exact=(), contains=(("руб/кв",), ("средняя стоимость",), ("средняя цена",)))
        if price is None:
            price = next((position for label, position in marks.items()
                          if label.startswith("руб/кв")), None)
        # Колонка даты тоже ищется подписью: в первом шаблоне она первая, во
        # втором — вторая, и «дата всегда слева» держалось на одной книге.
        month = _column_label(marks, exact=("дата", "период"), contains=(("период",), ("дата",)))
        if units is not None and area is not None:
            return index, {"units": units, "area": area,
                           "price": price if price is not None else -1,
                           "month": month if month is not None else 0}
    raise PlanNotFound(
        "В листе плана не найдены колонки с количеством и площадью продаж")


def _is_xlsb(data: bytes) -> bool:
    """Формат определяется содержимым, а не расширением.

    Имя файла человек меняет как угодно, а внутри .xlsb лежит `workbook.bin`
    вместо `workbook.xml`. Разбирать по расширению значит однажды сказать
    «файл не читается» на исправной книге.
    """
    try:
        import zipfile
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            return any(name.endswith("workbook.bin") for name in archive.namelist())
    except Exception:  # noqa: BLE001
        return False


def _sheet_rows(data: bytes, *, hints: tuple[str, ...] = ()) -> tuple[list[tuple], str]:
    """Строки листа плана — из .xlsx, .xlsm или .xlsb одинаковыми кортежами.

    Двоичный .xlsb — не экзотика: рабочая финмодель на двадцать семь листов в
    нём весит вдвое меньше и открывается быстрее, поэтому книги живут именно
    так. Читать его openpyxl не умеет вовсе, и раньше такая книга получала
    ответ «файл не читается как книга Excel» — верный по букве и бесполезный.
    """
    if _is_xlsb(data):
        try:
            from pyxlsb import open_workbook
        except ImportError as exc:  # pragma: no cover
            raise PlanNotFound(
                "Книга в формате .xlsb, а библиотека pyxlsb не установлена") from exc
        try:
            with open_workbook(io.BytesIO(data)) as book:
                name = _pick_sheet_name(list(book.sheets), hints=hints)
                with book.get_sheet(name) as sheet:
                    return [tuple(cell.v for cell in row) for row in sheet.rows()], name
        except PlanNotFound:
            raise
        except Exception as exc:  # noqa: BLE001
            raise PlanNotFound(f"Книга .xlsb не читается: {exc}") from exc

    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise PlanNotFound("Не установлен openpyxl — книга не читается") from exc
    try:
        book = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise PlanNotFound(f"Файл не читается как книга Excel: {exc}") from exc
    name = _pick_sheet_name(list(book.sheetnames), hints=hints)
    return [row for row in book[name].iter_rows(values_only=True)], name


def parse_plan(data: bytes) -> dict[str, Any]:
    """Помесячный план и факт из книги ПЛАТО."""
    rows, sheet_name = _sheet_rows(data)
    header, columns = _columns(rows)

    months: list[dict[str, Any]] = []
    month_at = int(columns.get("month") or 0)
    for row in rows[header + 1:]:
        if not row:
            continue
        month = _month(row[month_at]) if month_at < len(row) else None
        if not month:
            # Строка «ИТОГО» и прочие сводки датой не помечены — они не месяц.
            continue
        # Пометка «факт» стоит в соседней справа колонке в обоих шаблонах, но
        # искать её лучше во всей строке: пустая ячейка рядом ещё не значит
        # план, а ошибка здесь переносит факт в будущее.
        kind = " ".join(_norm(cell) for cell in row[:month_at + 3])
        units = _number(row[columns["units"]]) if columns["units"] < len(row) else None
        area = _number(row[columns["area"]]) if columns["area"] < len(row) else None
        price = None
        if 0 <= columns["price"] < len(row):
            price = _number(row[columns["price"]])
        if units is None and area is None:
            continue
        months.append({
            "month": month,
            "kind": "fact" if any(mark in kind for mark in FACT_MARKS) else "plan",
            "units": None if units is None else round(units, 1),
            "area": None if area is None else round(area, 1),
            "price": None if not price else int(round(price)),
        })

    if not months:
        raise PlanNotFound("В листе плана нет ни одной строки с датой и количеством")

    facts = [m for m in months if m["kind"] == "fact"]
    plans = [m for m in months if m["kind"] == "plan"]
    return {
        "months": months,
        "fact_until": facts[-1]["month"] if facts else None,
        "plan_from": plans[0]["month"] if plans else None,
        "sheet": sheet_name,
    }


def compare(plan: dict[str, Any], market: list[dict[str, Any]]) -> dict[str, Any]:
    """Сопоставить план, свой факт и факт источника по общим месяцам.

    Факт книги и факт «Пульса» расходятся на срок регистрации ДДУ, поэтому
    сравниваются не они между собой, а каждый — с планом.
    """
    by_month = {row["month"]: row for row in plan.get("months") or []}
    source = {row["month"]: row for row in market or []}
    months = sorted(set(by_month) | set(source))
    rows = []
    for month in months:
        own = by_month.get(month) or {}
        rows.append({
            "month": month,
            "kind": own.get("kind"),
            "plan_units": own.get("units") if own.get("kind") == "plan" else None,
            "fact_units": own.get("units") if own.get("kind") == "fact" else None,
            "source_units": (source.get(month) or {}).get("sold"),
            "plan_price": own.get("price"),
        })
    done = [r for r in rows if r["fact_units"] is not None]
    ahead = [r for r in rows if r["plan_units"] is not None]
    return {
        "rows": rows,
        "fact_months": len(done),
        "plan_months": len(ahead),
        "fact_total": round(sum(r["fact_units"] or 0 for r in done), 1),
        "plan_total": round(sum(r["plan_units"] or 0 for r in ahead), 1),
    }


# --- продажи по продуктам ---------------------------------------------------
# Лист «Продажи ФМ_new» устроен поперёк: месяцы по колонкам парами «план/факт»,
# продукты по строкам. Он богаче листа плана и, главное, СВЕЖЕЕ: на книге
# владельца факт по квартирам там доходит до июля 2026, тогда как на «План
# продаж_утв» пометка «факт» обрывается в мае. Читая только лист плана, мы
# показывали бы факт короче на два месяца — и это выглядело бы как «продажи
# встали», а не как «мы читаем не тот лист».
PRODUCT_SHEET_HINTS = ("продажи фм_new", "продажи фм new", "продажи фм", "продажи п-ф_new")

# Ключ продукта — наш, подпись — из книги: подписи в книгах правят руками.
PRODUCT_ROWS = (
    ("apartments", ("квартира", "квартиры")),
    ("storage", ("кладовые", "кладовки")),
    ("parking", ("машиноместа", "машино-места", "м/м")),
    ("commercial", ("коммерческие площади", "псн", "коммерция")),
)
# Меры внутри блока продукта. «Цена продажи» намеренно не берётся: она в книге
# считается делением, и второй счёт той же цены однажды разошёлся бы с первым.
PRODUCT_MEASURES = (
    ("escrow_th", ("эскроу",)),
    ("area", ("м2", "кв.м", "м²")),
    ("units", ("шт",)),
    ("contracts_th", ("заключенные договоры", "заключённые договоры")),
)
# Ниже помесячного блока в том же листе лежит такой же по подписям блок «Продажи
# по годам». Прочитать его следом значит сложить год с месяцем и не заметить.
_BLOCK_BREAK = ("продажи по годам", "продажи по кварталам")
# «Итого» стоит последней строкой блока и несёт те же подписи мер, что продукт.
# Без этого его эскроу и договоры дописывались в ПОСЛЕДНИЙ продукт — в книге
# владельца коммерция получала 114 месяцев вместо 74, и сумма выглядела просто
# крупной коммерцией.
_BLOCK_END = ("итого", "всего")


def _product_header(rows: list[tuple]) -> tuple[int, list[tuple[int, str, str]]]:
    """Строка с «план/факт» и разбор колонок: индекс, месяц, вид."""
    for index, row in enumerate(rows[:_HEADER_SEARCH_DEPTH]):
        marks = [position for position, cell in enumerate(row)
                 if _norm(cell) in ("план", "факт")]
        if len(marks) < 4:
            continue
        dates = rows[index - 1] if index else ()
        columns: list[tuple[int, str, str]] = []
        for position in marks:
            month = _month(dates[position]) if position < len(dates) else None
            if not month:
                continue
            columns.append((position, month, "fact" if _norm(row[position]) == "факт" else "plan"))
        if columns:
            return index, columns
    raise PlanNotFound("В листе продаж не найдена строка «план / факт»")


def parse_product_sales(data: bytes) -> dict[str, Any]:
    """План и факт по каждому продукту помесячно.

    Лист плана знает только квартиры; кладовые, машино-места и коммерция не
    видны в отчёте вовсе, хотя в книге они есть и по ним тоже есть факт.
    """
    rows, sheet_name = _sheet_rows(data, hints=PRODUCT_SHEET_HINTS)
    header, columns = _product_header(rows)
    products: dict[str, dict[str, Any]] = {}
    current: str | None = None
    for row in rows[header + 1:]:
        label = _norm(row[2] if len(row) > 2 else "")
        if not label:
            continue
        if any(mark in label for mark in _BLOCK_BREAK):
            break
        if label in _BLOCK_END:
            current = None
            continue
        matched = next((key for key, names in PRODUCT_ROWS if label in names), None)
        if matched:
            current = matched
            products.setdefault(current, {"label": str(row[2]).strip(), "months": {}})
            continue
        if current is None:
            continue
        measure = next((key for key, names in PRODUCT_MEASURES
                        if any(label.startswith(name) for name in names)), None)
        if measure is None:
            continue
        for position, month, kind in columns:
            value = _number(row[position]) if position < len(row) else None
            if value is None:
                continue
            slot = products[current]["months"].setdefault(
                (month, kind), {"month": month, "kind": kind})
            slot[measure] = value

    if not products:
        raise PlanNotFound("В листе продаж не найдено ни одного продукта")
    fact_until = ""
    for product in products.values():
        product["months"] = sorted(product["months"].values(),
                                   key=lambda item: (item["month"], item["kind"]))
        for item in product["months"]:
            # Ноль — это «в этом месяце не продавали», а не «факта нет»: месяц
            # с нулём и месяц без строки читаются одинаково, а значат разное.
            if item["kind"] == "fact" and any(
                    _number(item.get(key)) for key in ("area", "units", "contracts_th")):
                fact_until = max(fact_until, item["month"])
    return {"sheet": sheet_name, "products": products, "fact_until": fact_until}
