"""DevelopAid Project Monitor: fixed baseline + weekly RSS 6.1.2.

The operating workflow is intentionally small:

1. A project baseline is loaded once. It is the prepared PM/GPR workbook with
   RSS codes (or the Project Control workbook containing ``BASELINE ГПР``).
   If the same workbook contains ``CF ПЛАН-ФАКТ``, its Plan series is the
   payment baseline.
2. Every reporting week only a fresh RSS 6.1.2 workbook is uploaded.
3. Accepted construction works from RSS are mapped to baseline GPR rows by RSS
   code and produce plan/fact progress, recent pace and forecast finish.
4. ``Реестр платежей`` from the same RSS workbook is payment fact. It is
   compared with the fixed payment baseline when one is available.

RSS production programme, sales files and ad-hoc recovery proposals are not
baseline sources. Legacy storage functions remain for API compatibility, but
``build`` and ``gantt`` deliberately ignore them.
"""

from __future__ import annotations

import datetime
import io
import json
import os
import re
from pathlib import Path
from typing import Any

import developaid_actuals as actuals

_ROOT = Path(__file__).resolve().parent
_SNAPSHOT_DIR = Path(
    os.getenv("DEVELOPAID_MONITOR_DIR", "").strip()
    or (_ROOT / "data" / "monitor")
)


def _slug(name: str) -> str:
    cleaned = re.sub(r"[^0-9a-zA-Zа-яА-ЯёЁ _-]+", "", str(name or "")).strip()
    cleaned = re.sub(r"\s+", "-", cleaned).strip("-.")
    return cleaned[:64] or "project"


def _project_dir(project: str) -> Path:
    path = _SNAPSHOT_DIR / _slug(project)
    path.mkdir(parents=True, exist_ok=True)
    return path


def _iso(value: Any) -> str:
    if isinstance(value, (datetime.date, datetime.datetime)):
        return value.strftime("%Y-%m-%d")
    return str(value or "")


def _plain(value: Any) -> Any:
    if isinstance(value, dict):
        return {(_iso(k) if isinstance(k, (datetime.date, datetime.datetime)) else str(k)):
                _plain(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        return [_plain(item) for item in value]
    if isinstance(value, set):
        return sorted(str(item) for item in value)
    if isinstance(value, (datetime.date, datetime.datetime)):
        return _iso(value)
    return value


def _day(value: Any) -> datetime.date | None:
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    text = str(value or "").strip()
    try:
        return datetime.date.fromisoformat(text[:10])
    except (ValueError, TypeError):
        return None


def store_estimate(project: str, data: bytes, taken_at: Any,
                   filename: str = "") -> dict[str, Any]:
    """Store one immutable weekly RSS 6.1.2 snapshot."""
    day = _iso(taken_at)
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", day):
        raise ValueError("дата среза РСС нужна в виде ГГГГ-ММ-ДД")
    # Validate the one file that drives the whole weekly update before storing.
    try:
        actuals.read_estimate(io.BytesIO(data))
        actuals.read_payments(io.BytesIO(data))
        actuals.read_completed_works(io.BytesIO(data))
    except (KeyError, ValueError) as exc:
        raise ValueError(f"это не полная выгрузка РСС 6.1.2: {exc}") from exc
    folder = _project_dir(project) / "estimate"
    folder.mkdir(parents=True, exist_ok=True)
    path = folder / f"{day}.xlsx"
    if path.exists():
        raise FileExistsError(f"снимок РСС на {day} уже загружен")
    path.write_bytes(data)
    (folder / f"{day}.json").write_text(json.dumps({
        "taken_at": day,
        "filename": filename,
        "bytes": len(data),
        "loaded_at": datetime.datetime.now(datetime.timezone.utc).isoformat(),
    }, ensure_ascii=False), encoding="utf-8")
    return {"taken_at": day, "path": str(path), "bytes": len(data)}


def _has_sheet(data: bytes | Path, name: str) -> bool:
    from openpyxl import load_workbook
    source = io.BytesIO(data) if isinstance(data, bytes) else data
    wb = load_workbook(source, read_only=True, data_only=True)
    try:
        return name in wb.sheetnames
    finally:
        wb.close()


def store_schedule(project: str, gpr: bytes, pm: bytes | None,
                   taken_at: Any) -> dict[str, Any]:
    """Create the fixed project baseline once.

    ``gpr`` may be either the prepared cleaned GPR (sheet ``ГПР``) or our
    Project Control workbook (sheet ``BASELINE ГПР``). ``pm`` is retained as a
    compatibility argument and can carry a separate workbook with the payment
    baseline. It is never a weekly source of actual dates.
    """
    day = _iso(taken_at)
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", day):
        raise ValueError("дата baseline нужна в виде ГГГГ-ММ-ДД")
    # Validate schedule before committing it as immutable baseline.
    _read_baseline_gpr_bytes(gpr)
    folder = _project_dir(project) / "baseline"
    folder.mkdir(parents=True, exist_ok=True)
    path = folder / "gpr.xlsx"
    if path.exists():
        raise FileExistsError("baseline проекта уже загружен; он не обновляется еженедельно")
    path.write_bytes(gpr)
    if pm:
        (folder / "finance.xlsx").write_bytes(pm)
    (folder / "meta.json").write_text(json.dumps({
        "taken_at": day,
        "created_at": datetime.datetime.now(datetime.timezone.utc).isoformat(),
    }, ensure_ascii=False), encoding="utf-8")
    plan = _payment_baseline(project)
    return {
        "taken_at": day,
        "baseline": True,
        "works": len(_read_baseline_gpr(project)["works"]),
        "payment_plan": bool(plan["known"]),
    }


def _read_baseline_gpr_bytes(data: bytes) -> dict[str, Any]:
    if _has_sheet(data, "ГПР"):
        return actuals.read_schedule(io.BytesIO(data))
    if _has_sheet(data, "BASELINE ГПР"):
        return _read_control_baseline(io.BytesIO(data))
    raise ValueError("в baseline-файле нет листа «ГПР» или «BASELINE ГПР»")


def _read_control_baseline(source: Any) -> dict[str, Any]:
    """Read the compact GPR baseline from DevelopAid Project Control."""
    from openpyxl import load_workbook
    wb = load_workbook(source, read_only=True, data_only=True)
    try:
        ws = wb["BASELINE ГПР"]
        rows: list[dict[str, Any]] = []
        header: dict[str, int] = {}
        aliases = {
            "wbs": "wbs", "раздел": "section", "объект": "object",
            "работа": "name", "код рсс": "estimate_code",
            "план начало": "start", "план конец": "finish",
            "% в гпр на срез": "progress", "комментарий мэппинга": "basis",
        }
        for values in ws.iter_rows(values_only=True):
            norm = [str(v or "").strip().lower().replace("ё", "е") for v in values]
            if not header and "wbs" in norm and "работа" in norm:
                for idx, value in enumerate(norm):
                    if value in aliases:
                        header[aliases[value]] = idx
                continue
            if not header:
                continue
            def cell(key: str) -> Any:
                idx = header.get(key, -1)
                return values[idx] if 0 <= idx < len(values) else None
            wbs = str(cell("wbs") or "").strip()
            name = str(cell("name") or "").strip()
            if not wbs or not name:
                continue
            start = actuals._date(cell("start"))
            finish = actuals._date(cell("finish"))
            if not start or not finish:
                continue
            rows.append({
                "id": wbs,
                "wbs": wbs,
                "section": str(cell("section") or "").strip(),
                "object": str(cell("object") or "").strip(),
                "name": name,
                "kind": "Работа",
                "is_work": True,
                "progress": actuals._money(cell("progress")),
                "start": start,
                "finish": finish,
                "status": "",
                "overdue": False,
                "duration": (finish - start).days + 1,
                "predecessors": "",
                "estimate_code": str(cell("estimate_code") or "").strip().rstrip("."),
                "estimate_name": "",
                "basis": str(cell("basis") or "").strip(),
            })
        works = [row for row in rows if row["is_work"]]
        if not works:
            raise ValueError("на листе «BASELINE ГПР» не найдено работ")
        return {"rows": rows, "works": works, "overdue": [],
                "without_code": [r for r in works if not r["estimate_code"]]}
    finally:
        wb.close()


def _baseline_file(project: str) -> Path | None:
    modern = _project_dir(project) / "baseline" / "gpr.xlsx"
    if modern.exists():
        return modern
    # Backward-compatible migration from Claude's dated schedule snapshots:
    # freeze the earliest stored schedule rather than following later uploads.
    old = _project_dir(project) / "schedule"
    items = sorted(p for p in old.glob("*.xlsx") if not p.name.endswith(".pm.xlsx")) if old.exists() else []
    return items[0] if items else None


def store_schedule_fact(project: str, data: bytes, taken_at: Any) -> dict[str, Any]:
    """Положить еженедельный ГПР-факт снимком: проценты и статусы поверх baseline.

    План проекта зафиксирован baseline и не меняется; меняется выполнение.
    Прораб проставляет проценты в тот же файл ГПР и присылает его раз в
    неделю — вместе с РСС. Каждый снимок хранится отдельно: переписанное
    прошлое видно только парой снимков.
    """
    day = _iso(taken_at)
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", day):
        raise ValueError("дата ГПР-факта нужна в виде ГГГГ-ММ-ДД")
    parsed = _read_baseline_gpr_bytes(data)
    folder = _project_dir(project) / "schedule_fact"
    folder.mkdir(parents=True, exist_ok=True)
    path = folder / f"{day}.xlsx"
    if path.exists():
        raise FileExistsError(f"ГПР-факт на {day} уже загружен")
    path.write_bytes(data)
    works = parsed["works"]
    return {
        "taken_at": day,
        "works": len(works),
        "completed": sum(1 for row in works
                         if "заверш" in str(row.get("status") or "").lower()
                         or (row.get("progress") or 0) >= 0.999),
    }


def latest_schedule_fact(project: str, upto: str = "") -> dict[str, Any] | None:
    """Свежайший снимок ГПР-факта, не позднее `upto`. Нет снимка — None."""
    path = _latest(project, "schedule_fact", ".xlsx", upto)
    if path is None:
        return None
    parsed = _read_baseline_gpr_bytes(path.read_bytes())
    parsed["taken_at"] = path.stem
    return parsed


def _read_baseline_gpr(project: str) -> dict[str, Any]:
    path = _baseline_file(project)
    if path is None:
        raise FileNotFoundError("не загружен baseline ГПР")
    data = path.read_bytes()
    return _read_baseline_gpr_bytes(data)


def _latest(project: str, kind: str, suffix: str, upto: str = "") -> Path | None:
    folder = _project_dir(project) / kind
    if not folder.exists():
        return None
    items = sorted(folder.glob(f"*{suffix}"))
    if upto:
        items = [item for item in items if item.stem <= upto]
    return items[-1] if items else None


def _codes(expr: Any) -> list[str]:
    text = str(expr or "")
    out: list[str] = []
    for token in re.findall(r"\d+(?:\.\d+)+", text):
        token = token.rstrip(".")
        if token not in out:
            out.append(token)
    return out


def _descendants(estimate: dict[str, Any], roots: list[str]) -> set[str]:
    children: dict[str, set[str]] = {}
    for row in estimate["rows"]:
        parent = str(row.get("parent") or "")
        if parent:
            children.setdefault(parent, set()).add(row["code"])
    selected: set[str] = set()
    stack = list(roots)
    while stack:
        code = stack.pop()
        if code in selected:
            continue
        selected.add(code)
        stack.extend(children.get(code, ()))
    return selected


def _rss_metrics(estimate: dict[str, Any], works: dict[str, Any], expr: Any,
                 cut: datetime.date) -> dict[str, Any]:
    roots = _codes(expr)
    if not roots:
        return {"known": False, "codes": [], "eac": 0.0, "accepted": 0.0,
                "progress": None, "rate_3m": None, "last_act": None}
    matched = _descendants(estimate, roots)
    # EAC is read from the declared root rows. Summing descendants would double
    # count RSS hierarchy rows.
    eac = sum(float((estimate["by_code"].get(code) or {}).get("estimate") or 0.0)
              for code in roots)
    selected = [row for row in works["rows"]
                if row.get("construction") and row.get("code") in matched
                and row.get("date") and row["date"] <= cut]
    accepted = sum(float(row.get("amount") or 0.0) for row in selected)
    window = cut - datetime.timedelta(days=92)
    recent = sum(float(row.get("amount") or 0.0) for row in selected
                 if row["date"] > window)
    progress = accepted / eac if eac > 0 else None
    rate = recent / eac / 3.0 if eac > 0 else None
    last_act = max((row["date"] for row in selected), default=None)
    return {"known": eac > 0, "codes": roots, "eac": eac,
            "accepted": accepted, "progress": progress,
            "rate_3m": rate, "last_act": last_act}


def _plan_progress(start: datetime.date, finish: datetime.date,
                   cut: datetime.date) -> float:
    if cut < start:
        return 0.0
    if cut >= finish:
        return 1.0
    duration = max(1, (finish - start).days)
    return max(0.0, min(1.0, (cut - start).days / duration))


def _forecast_finish(cut: datetime.date, progress: float | None,
                     rate_3m: float | None, baseline_finish: datetime.date,
                     last_act: datetime.date | None) -> datetime.date | None:
    if progress is None:
        return None
    if progress >= 1.0:
        return last_act or cut
    if rate_3m and rate_3m > 1e-9:
        months = max(0.0, (1.0 - max(0.0, progress)) / rate_3m)
        return cut + datetime.timedelta(days=round(months * 30.4375))
    return None


def _schedule_plan_fact(project: str, rss_path: Path, cut: Any) -> dict[str, Any]:
    cut_date = _day(cut) or _day(rss_path.stem)
    if cut_date is None:
        raise ValueError("не задана дата среза")
    baseline = _read_baseline_gpr(project)
    estimate = actuals.read_estimate(rss_path)
    works = actuals.read_completed_works(rss_path)
    rows = []
    for item in baseline["works"]:
        start, finish = item.get("start"), item.get("finish")
        if not start or not finish:
            continue
        plan = _plan_progress(start, finish, cut_date)
        metrics = _rss_metrics(estimate, works, item.get("estimate_code"), cut_date)
        fact = metrics["progress"]
        rate = metrics["rate_3m"]
        forecast = _forecast_finish(cut_date, fact, rate, finish, metrics["last_act"])
        remaining_months = max((finish - cut_date).days / 30.4375, 0.0)
        required_rate = ((1.0 - max(0.0, fact or 0.0)) / remaining_months
                         if fact is not None and remaining_months > 0 else None)
        if cut_date < start:
            status = "ПО ПЛАНУ: НЕ НАЧАТО"
        elif fact is None:
            status = "НЕТ ДАННЫХ РСС"
        elif fact >= 1.0:
            status = "ЗАВЕРШЕНО"
        elif (not rate or rate <= 1e-9) and fact + 1e-9 < plan:
            status = "НЕТ ТЕМПА / РИСК"
        elif forecast and forecast > finish:
            status = "ОТСТАВАНИЕ"
        elif fact + 1e-9 >= plan:
            status = "В СРОК"
        else:
            status = "ОТСТАВАНИЕ"
        duration = max(1, (finish - start).days)
        fact_date = (start + datetime.timedelta(days=round(duration * min(max(fact or 0.0, 0.0), 1.0)))
                     if fact is not None else None)
        rows.append({
            "id": item.get("id", ""), "wbs": item.get("wbs", ""),
            "section": item.get("section", ""), "object": item.get("object", ""),
            "name": item.get("name", ""), "code": item.get("estimate_code", ""),
            "basis": item.get("basis", ""),
            "plan_start": start, "plan_finish": finish,
            "plan_progress": plan, "actual_progress": fact,
            "actual_equivalent_date": fact_date,
            "accepted": metrics["accepted"], "eac": metrics["eac"],
            "rate_3m": rate, "required_rate": required_rate,
            "forecast_finish": forecast,
            "delta_days": ((forecast - finish).days if forecast else None),
            "status": status,
        })
    risks = [r for r in rows if r["status"] in {"ОТСТАВАНИЕ", "НЕТ ТЕМПА / РИСК"}]
    forecast_end = max((r["forecast_finish"] for r in rows if r["forecast_finish"]), default=None)
    baseline_end = max((r["plan_finish"] for r in rows), default=None)
    return {"cut": cut_date, "rows": rows, "risks": len(risks),
            "baseline_end": baseline_end, "forecast_end": forecast_end}


def _payment_baseline_from_file(path: Path) -> dict[str, Any]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if "CF ПЛАН-ФАКТ" not in wb.sheetnames:
            return {"known": False, "series": {}, "source": ""}
        ws = wb["CF ПЛАН-ФАКТ"]
        rows = list(ws.iter_rows(values_only=True))
        plan_row = None
        date_row = None
        total_row = None
        for i, row in enumerate(rows):
            normalized = [str(v or "").strip().lower().replace("ё", "е") for v in row]
            if normalized.count("план") >= 2:
                plan_row = i
                # date row is usually two rows above, but find nearest preceding
                # row carrying at least two Excel dates.
                for j in range(i - 1, max(-1, i - 5), -1):
                    if sum(isinstance(v, (datetime.date, datetime.datetime)) for v in rows[j]) >= 2:
                        date_row = j
                        break
            if normalized and normalized[0] in {"итого проект", "итого"}:
                # In the four-column layout this row itself carries Plan values;
                # prefer the last matching total row before alternate series.
                total_row = i
        if plan_row is None or date_row is None:
            return {"known": False, "series": {}, "source": "CF ПЛАН-ФАКТ"}
        # Pick the explicit ИТОГО ПРОЕКТ row if present.
        for i, row in enumerate(rows):
            if str(row[0] or "").strip().lower() == "итого проект":
                total_row = i
                break
        if total_row is None:
            return {"known": False, "series": {}, "source": "CF ПЛАН-ФАКТ"}
        dates_by_col: dict[int, datetime.date] = {}
        last_date = None
        for col, value in enumerate(rows[date_row]):
            if isinstance(value, datetime.datetime):
                last_date = value.date().replace(day=1)
            elif isinstance(value, datetime.date):
                last_date = value.replace(day=1)
            if last_date is not None:
                dates_by_col[col] = last_date
        series: dict[str, float] = {}
        for col, label in enumerate(rows[plan_row]):
            if str(label or "").strip().lower() != "план":
                continue
            month = dates_by_col.get(col)
            if month is None:
                continue
            value = actuals._money(rows[total_row][col] if col < len(rows[total_row]) else None)
            series[month.isoformat()] = value * (1e6 if abs(value) < 100_000 else 1.0)
        return {"known": bool(series), "series": series, "source": "CF ПЛАН-ФАКТ"}
    finally:
        wb.close()


def _payment_baseline(project: str) -> dict[str, Any]:
    folder = _project_dir(project) / "baseline"
    for name in ("finance.xlsx", "gpr.xlsx"):
        path = folder / name
        if path.exists():
            parsed = _payment_baseline_from_file(path)
            if parsed["known"]:
                return parsed
    return {"known": False, "series": {}, "source": ""}


def _payment_plan_fact(project: str, rss_path: Path, cut: Any) -> dict[str, Any]:
    payments = actuals.read_payments(rss_path)
    fact: dict[str, float] = {}
    for row in payments["rows"]:
        date = row.get("date")
        if not date:
            continue
        month = date.replace(day=1).isoformat()
        fact[month] = fact.get(month, 0.0) + float(row.get("amount") or 0.0)
    baseline = _payment_baseline(project)
    months = sorted(set(fact) | set(baseline["series"]))
    rows = []
    for month in months:
        plan = float(baseline["series"].get(month, 0.0)) if baseline["known"] else None
        actual = float(fact.get(month, 0.0))
        rows.append({"month": month, "plan": plan, "fact": actual,
                     "delta": (actual - plan if plan is not None else None)})
    return {"known": baseline["known"], "source": baseline["source"],
            "rows": rows, "plan_total": (sum(baseline["series"].values()) if baseline["known"] else None),
            "fact_total": sum(fact.values()), "last_fact": _iso(payments.get("last"))}


def build(project: str, cut: Any, programme: dict[str, Any] | None = None,
          upto: str = "") -> dict[str, Any]:
    """Build current control view from fixed baseline + one latest RSS snapshot.

    ``programme`` is accepted only for backward call compatibility and ignored.
    """
    rss_path = _latest(project, "estimate", ".xlsx", upto)
    if rss_path is None:
        raise FileNotFoundError("нет ни одного снимка РСС")
    effective_cut = _day(cut) or _day(rss_path.stem)
    if effective_cut is None:
        raise ValueError("не задана дата среза")
    estimate = actuals.read_estimate(rss_path)
    works = actuals.read_completed_works(rss_path)
    payments = actuals.read_payments(rss_path)
    contracts = actuals.read_contracts(rss_path)
    schedule = _schedule_plan_fact(project, rss_path, effective_cut)
    cash = _payment_plan_fact(project, rss_path, effective_cut)
    total = estimate.get("total") or {}
    return _plain({
        "cut": effective_cut,
        "source": {"estimate": rss_path.stem, "baseline": _iso((_project_dir(project) / "baseline" / "meta.json").exists())},
        "money": {
            "estimate": float(total.get("estimate") or 0.0),
            "contracted": float(total.get("contracted") or 0.0),
            "paid": float(total.get("paid") or payments.get("total") or 0.0),
            "accepted": float(total.get("completed") or works.get("total") or 0.0),
            "payment_fact": float(payments.get("total") or 0.0),
            "contracts": len(contracts.get("rows") or []),
        },
        "schedule": schedule,
        "payments": cash,
        "by_code": schedule["rows"],
    })


def gantt(project: str, cut: Any, upto: str = "") -> dict[str, Any]:
    """Gantt bars: fixed plan + RSS-derived actual progress + forecast."""
    rss_path = _latest(project, "estimate", ".xlsx", upto)
    if rss_path is None:
        raise FileNotFoundError("нет ни одного снимка РСС")
    report = _schedule_plan_fact(project, rss_path, cut)
    bars = [{
        "id": r["id"], "wbs": r["wbs"], "name": r["name"],
        "section": r["section"], "object": r["object"], "code": r["code"],
        "baseline": [r["plan_start"], r["plan_finish"]],
        "plan": [r["plan_start"], r["plan_finish"]],
        # This is a cost-weighted position inferred from RSS acts, not a PM
        # actual-start/actual-finish claim.
        "fact": [r["plan_start"], r["actual_equivalent_date"]] if r["actual_equivalent_date"] else [None, None],
        "forecast_finish": r["forecast_finish"], "slip_days": r["delta_days"],
        "progress": r["actual_progress"], "plan_progress": r["plan_progress"],
        "rate_3m": r["rate_3m"], "status": r["status"],
        "overdue": r["status"] in {"ОТСТАВАНИЕ", "НЕТ ТЕМПА / РИСК"},
        "done": r["status"] == "ЗАВЕРШЕНО",
        "running": bool((r["actual_progress"] or 0) > 0 and r["status"] != "ЗАВЕРШЕНО"),
    } for r in report["rows"]]
    return _plain({
        "cut": report["cut"], "bars": bars, "by_code": report["rows"],
        "works": len(bars), "overdue": report["risks"],
        "baseline_end": report["baseline_end"], "forecast_end": report["forecast_end"],
        "source": {"schedule": "fixed-baseline", "estimate": rss_path.stem,
                   "with_baseline": True},
        "baseline": {"matched": len(bars), "total": len(bars)},
    })


def snapshots(project: str) -> dict[str, list[str]]:
    folder = _project_dir(project)
    def dates(kind: str, suffix: str) -> list[str]:
        path = folder / kind
        if not path.exists():
            return []
        return sorted(item.stem for item in path.glob(f"*{suffix}"))
    baseline = _baseline_file(project)
    return {
        "estimate": dates("estimate", ".xlsx"),
        "baseline": ["fixed"] if baseline else [],
        # Legacy keys kept so old clients do not break.
        "sales": dates("sales", ".json"),
        "schedule": [baseline.stem] if baseline else [],
        "programme": dates("programme", ".xlsx"),
        "proposal": sorted({item.split(".", 1)[0] for item in dates("proposal", ".json")}),
    }


def trend(project: str, cut: Any, programme: dict[str, Any] | None = None) -> list[dict[str, Any]]:
    points = []
    for day in snapshots(project)["estimate"]:
        try:
            point = build(project, cut=day, upto=day)
        except Exception:
            continue
        points.append({
            "snapshot": day,
            "paid": point["money"]["paid"],
            "accepted": point["money"]["accepted"],
            "payment_fact": point["money"]["payment_fact"],
            "risks": point["schedule"]["risks"],
            "forecast_end": point["schedule"].get("forecast_end"),
        })
    return points


def moved_between_snapshots(project: str, first: str, second: str) -> dict[str, Any]:
    """Audit whether prior-month accepted works were rewritten between RSS snapshots."""
    def monthly(day: str) -> dict[str, float]:
        path = _project_dir(project) / "estimate" / f"{day}.xlsx"
        if not path.exists():
            raise FileNotFoundError(f"нет снимка РСС на {day}")
        out: dict[str, float] = {}
        for row in actuals.read_completed_works(path)["rows"]:
            if row["construction"] and row["date"]:
                key = row["date"].strftime("%Y-%m")
                out[key] = out.get(key, 0.0) + row["amount"]
        return out
    before, after = monthly(first), monthly(second)
    months = sorted(set(before) | set(after))
    moved = [{"month": month, "before": before.get(month, 0.0),
              "after": after.get(month, 0.0),
              "delta": after.get(month, 0.0) - before.get(month, 0.0)}
             for month in months]
    return {"first": first, "second": second, "rows": moved,
            "rewritten": [r for r in moved if r["before"] > 0 and abs(r["delta"]) > 1e6]}


# ---------------------------------------------------------------------------
# Legacy write endpoints. They remain callable so a deployed older page does
# not fail during a rolling release, but their data no longer affects Monitor.
# ---------------------------------------------------------------------------

def store_sales(project: str, rows: list[dict[str, Any]], taken_at: Any) -> dict[str, Any]:
    """Продажи строками — то, чего книга ещё не знает.

    Книга обновляется раз в месяц и отстаёт; «в августе продано 4 лота»
    приходит словами задолго до выгрузки. Месяц из строк перекрывает тот же
    месяц книги в срезе.
    """
    day = _iso(taken_at)
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", day):
        raise ValueError("дата продаж нужна в виде ГГГГ-ММ-ДД")
    cleaned = []
    for row in rows or []:
        month = actuals._as_month(row.get("month"))
        if month is None:
            continue
        cleaned.append({
            "month": _iso(month)[:7],
            "units": float(row.get("units") or 0),
            "area": float(row.get("area") or 0),
            "revenue": float(row.get("revenue") or 0),
        })
    if not cleaned:
        raise ValueError("ни в одной строке нет месяца")
    folder = _project_dir(project) / "sales"
    folder.mkdir(parents=True, exist_ok=True)
    (folder / f"{day}.json").write_text(json.dumps(
        {"taken_at": day, "rows": cleaned}, ensure_ascii=False), encoding="utf-8")
    return {"taken_at": day, "months": len(cleaned)}


def store_sales_file(project: str, data: bytes, taken_at: Any) -> dict[str, Any]:
    return store_sales(project, [], taken_at)


def store_programme(project: str, data: bytes, start: Any, taken_at: Any) -> dict[str, Any]:
    day = _iso(taken_at)
    folder = _project_dir(project) / "programme"
    folder.mkdir(parents=True, exist_ok=True)
    (folder / f"{day}.xlsx").write_bytes(data)
    return {"taken_at": day, "ignored_by_monitor": True}


def store_proposal(project: str, data: bytes, sheet: str, start: Any,
                   code: str, taken_at: Any) -> dict[str, Any]:
    day = _iso(taken_at)
    folder = _project_dir(project) / "proposal"
    folder.mkdir(parents=True, exist_ok=True)
    slug = re.sub(r"[^0-9-]", "-", str(code))
    (folder / f"{day}.{slug}.json").write_text(json.dumps({
        "taken_at": day, "code": code, "ignored_by_monitor": True,
    }, ensure_ascii=False), encoding="utf-8")
    return {"taken_at": day, "code": code, "ignored_by_monitor": True}
