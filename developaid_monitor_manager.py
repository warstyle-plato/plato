"""Managerial hierarchy for DevelopAid Project Monitor.

Important semantics:

* RSS accepted-work acts are a cost/evidence source. ``accepted / EAC`` is
  shown as an *act-cost ratio* and is never treated as a physical WBS percent.
* Calendar forecast never extrapolates dates from the monetary pace of acts.
  Dates come from the approved GPR/PM network and approved rebaselines.
* RSS codes remain the financial crosswalk. They do not define the temporal
  phase hierarchy: one financial code may span several WBS phases.
"""
from __future__ import annotations

import datetime
import re
from pathlib import Path
from typing import Any

import developaid_actuals as actuals
import developaid_monitor as monitor
import developaid_monitor_schedule_graph as schedule_graph

_INSTALLED = False
_ORIGINAL_BUILD = None
_ORIGINAL_GANTT = None


def _natural(value: Any) -> tuple:
    return tuple(
        (0, int(part)) if part.isdigit() else (1, part.lower())
        for part in re.findall(r"\d+|[^\d]+", str(value or ""))
    )


def _norm(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower().replace("ё", "е"))


def _as_date(value: Any) -> datetime.date | None:
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    return monitor._day(value)


def _codes(value: Any) -> list[str]:
    return list(dict.fromkeys(
        item.rstrip(".")
        for item in re.findall(r"\d+(?:\.\d+)+", str(value or ""))
    ))


def _financial_control(code: str) -> str:
    if code.startswith("2.1"):
        return "Подготовка / ПОС"
    if code.startswith(("2.2", "2.3")):
        return "Основные объекты"
    if code.startswith("2.4"):
        return "Наружные сети"
    if code.startswith("2.5"):
        return "Благоустройство"
    if code.startswith("2.6"):
        return "Служба заказчика"
    if code.startswith("2.7"):
        return "Проектирование"
    if code.startswith(("2.8", "2.9")):
        return "Резерв"
    return "Прочие СМР"


def _financial_detail(code: str) -> str:
    if code.startswith("2.1"):
        return "ПОС и общеплощадочные работы"
    if code.startswith("2.2.1"):
        return "Основное строительство — подземная часть"
    if code.startswith(("2.2.2", "2.2.3", "2.3")):
        return "Основное строительство — надземная часть + ВИС"
    if code.startswith("2.4"):
        return "Наружные инженерные сети"
    if code.startswith("2.5"):
        return "Благоустройство"
    return _financial_control(code)


def _schedule_bucket(task: dict[str, Any], code: str) -> tuple[str, str]:
    """Temporal hierarchy is WBS-led; RSS code is only the finance crosswalk.

    The Grodnenskaya GPR is a good example: RSS 2.1 contains both the already
    completed territory-preparation scope and late POS/site-operation/
    demobilisation tasks. Treating all of 2.1 as one time phase makes
    "Подготовка" look open for years.
    """
    wbs = str(task.get("wbs") or "")
    name = _norm(task.get("name"))
    section = _norm(task.get("section"))
    obj = _norm(task.get("object"))

    if code.startswith("2.1"):
        if wbs.startswith("1.16.1.2"):
            return "Подготовка", "Подготовка территории"
        if wbs.startswith("1.16.1.3"):
            return "Организация стройплощадки / ПОС", "ПОС и временная инфраструктура"
        late_markers = (
            "демонтаж башенн", "демобилиз", "перебазиров", "подкран",
            "временн", "механизац", "штаб", "пос",
        )
        if any(marker in name for marker in late_markers):
            return "Организация стройплощадки / ПОС", "ПОС и временная инфраструктура"
        if "подготов" in section or "подготов" in obj or "подготов" in name:
            return "Подготовка", "Подготовка территории"

    return _financial_control(code), _financial_detail(code)


def _baseline_mapping(project: str) -> dict[str, dict[str, str]]:
    """RSS naming/crosswalk only; never a source of schedule fact."""
    path = monitor._baseline_file(project)
    if path is None:
        return {}

    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if "РСС ФАКТ" not in workbook.sheetnames:
            return {}
        result: dict[str, dict[str, str]] = {}
        header: dict[str, int] = {}
        for values in workbook["РСС ФАКТ"].iter_rows(values_only=True):
            normalized = [
                re.sub(r"\s+", " ", str(value or "").strip().lower().replace("ё", "е"))
                for value in values
            ]
            if not header and "код рсс" in normalized and "статья developaid" in normalized:
                header = {
                    "code": normalized.index("код рсс"),
                    "article": normalized.index("статья developaid"),
                }
                if "статья рсс" in normalized:
                    header["rss_name"] = normalized.index("статья рсс")
                continue
            if not header:
                continue
            code_index = header["code"]
            code = actuals._code(values[code_index] if code_index < len(values) else None)
            if not code:
                continue
            rss_index = header.get("rss_name", -1)
            result[code] = {
                "rss_name": (
                    str(values[rss_index] or "").strip()
                    if 0 <= rss_index < len(values) else ""
                ),
            }
        return result
    finally:
        workbook.close()


def _descendants(estimate: dict[str, Any], root: str) -> set[str]:
    children: dict[str, set[str]] = {}
    for row in estimate["rows"]:
        parent = str(row.get("parent") or "")
        if parent:
            children.setdefault(parent, set()).add(row["code"])
    selected: set[str] = set()
    stack = [root]
    while stack:
        code = stack.pop()
        if code in selected:
            continue
        selected.add(code)
        stack.extend(children.get(code, ()))
    return selected


def _metrics(
    estimate: dict[str, Any],
    works: dict[str, Any],
    code: str,
    cut: datetime.date,
) -> dict[str, Any]:
    """Cost/evidence metrics for one RSS article.

    ``accepted_ratio`` is deliberately named so nobody can mistake it for
    physical WBS progress.
    """
    matched = _descendants(estimate, code)
    eac = float((estimate["by_code"].get(code) or {}).get("estimate") or 0.0)
    rows = [
        row for row in works["rows"]
        if row.get("construction")
        and row.get("code") in matched
        and row.get("date")
        and row["date"] <= cut
    ]
    accepted = sum(float(row.get("amount") or 0.0) for row in rows)
    recent = sum(
        float(row.get("amount") or 0.0)
        for row in rows
        if row["date"] > cut - datetime.timedelta(days=92)
    )
    return {
        "eac": eac,
        "accepted": accepted,
        "accepted_ratio": accepted / eac if eac > 0 else None,
        "act_cost_rate_3m": recent / eac / 3 if eac > 0 else None,
        "last_act": max((row["date"] for row in rows), default=None),
    }


def _baseline_status(project: str) -> dict[str, dict[str, Any]]:
    """Approved GPR status is schedule metadata, not physical RSS fact."""
    try:
        baseline = monitor._read_baseline_gpr(project)
    except Exception:
        return {}
    out: dict[str, dict[str, Any]] = {}
    for item in baseline.get("works") or []:
        tid = str(item.get("id") or item.get("wbs") or "").strip()
        if not tid:
            continue
        status = _norm(item.get("status"))
        try:
            progress = float(item.get("progress"))
        except (TypeError, ValueError):
            progress = None
        if progress is not None and progress > 1.0:
            progress = progress / 100.0
        out[tid] = {
            "closed": "заверш" in status or (progress or 0.0) >= 0.999,
            "status": str(item.get("status") or ""),
            "progress": progress,
        }
    return out


def _sanitize_base_schedule(project: str, schedule: dict[str, Any]) -> None:
    """Remove the old money-pace calendar forecast from base Monitor rows."""
    statuses = _baseline_status(project)
    for row in schedule.get("rows") or []:
        tid = str(row.get("id") or row.get("wbs") or "").strip()
        meta = statuses.get(tid, {})
        # КС статьи — готовность всей статьи по всем захваткам и корпусам, и
        # приписывать её каждой задаче нельзя: «Разработка котлована», физически
        # пройденная, наследовала статейные 67,8% и давала +472 дня, из которых
        # +83 доезжали до РНВ по FS-цепочке. Собственный процент задачи из
        # утверждённого ГПР — нижняя граница её готовности: задача не может
        # быть менее готова, чем принято в baseline.
        article_ratio = row.get("actual_progress")
        own = meta.get("progress")
        if own is not None:
            row["rss_accepted_ratio"] = max(float(article_ratio or 0.0), own)
            row["progress_kind"] = "accepted_cost_ratio_floor_gpr"
            row["progress_label"] = "КС / EAC, не ниже % ГПР"
        else:
            row["rss_accepted_ratio"] = article_ratio
            row["progress_kind"] = "accepted_cost_ratio"
            row["progress_label"] = "КС / EAC"
        row["rss_act_cost_rate_3m"] = row.get("rate_3m")
        row["baseline_closed"] = bool(meta.get("closed"))
        row["baseline_status"] = meta.get("status", "")
        row["actual_progress"] = None
        row["actual_equivalent_date"] = None
        row["forecast_finish"] = row.get("plan_finish")
        row["delta_days"] = 0
        if row["baseline_closed"]:
            row["status"] = "ЗАВЕРШЕНО ПО УТВЕРЖДЕННОМУ ГПР"
        else:
            row["status"] = "АКТУАЛЬНЫЙ СРОК = УТВЕРЖДЕННЫЙ ГПР"


def _seed_rebaselines(project: str, schedule: dict[str, Any]) -> None:
    """Use approved article rebaseline as a calendar seed for the PM graph."""
    try:
        rebaselines = schedule_graph._rebaselines(project)
    except Exception:
        return
    if not rebaselines:
        return

    rows = schedule.get("rows") or []
    for code, rb in rebaselines.items():
        finish = rb.get("finish")
        if not finish:
            continue
        candidates = [
            row for row in rows
            if code in _codes(row.get("code"))
            and not row.get("baseline_closed")
            and _as_date(row.get("plan_finish")) is not None
        ]
        if not candidates:
            continue
        terminal = max(candidates, key=lambda row: _as_date(row.get("plan_finish")))
        terminal["forecast_finish"] = monitor._iso(finish)
        terminal["forecast_source"] = "approved_rebaseline"
        terminal["rebaseline_seed"] = {
            "code": code,
            "finish": monitor._iso(finish),
            "source": rb.get("source", ""),
        }
        plan_finish = _as_date(terminal.get("plan_finish"))
        terminal["delta_days"] = (
            (finish - plan_finish).days if plan_finish else None
        )


def _rss_row(
    code: str,
    tasks: list[dict[str, Any]],
    estimate: dict[str, Any],
    works: dict[str, Any],
    mapping: dict[str, dict[str, str]],
    cut: datetime.date,
    control: str,
    detail: str,
) -> dict[str, Any]:
    start = min(_as_date(task["plan_start"]) for task in tasks)
    finish = max(_as_date(task["plan_finish"]) for task in tasks)
    metrics = _metrics(estimate, works, code, cut)
    closed = bool(tasks) and all(bool(task.get("baseline_closed")) for task in tasks)

    forecasts = [
        _as_date(task.get("forecast_finish"))
        for task in tasks
        if _as_date(task.get("forecast_finish")) is not None
    ]
    forecast = finish if closed else (max(forecasts) if forecasts else finish)

    duration = max(1, (finish - start).days)
    ratio = metrics["accepted_ratio"]
    cost_equivalent_date = (
        start + datetime.timedelta(
            days=round(duration * min(max(ratio or 0.0, 0.0), 1.0))
        )
        if ratio is not None else None
    )
    meta = mapping.get(code, {})
    return {
        "key": f"rss:{control}:{detail}:{code}",
        "level": "rss",
        "code": code,
        "name": (
            meta.get("rss_name")
            or (estimate["by_code"].get(code) or {}).get("article")
            or code
        ),
        "control": control,
        "detail": detail,
        "plan_start": start,
        "plan_finish": finish,
        "plan_progress": None,
        "actual_progress": ratio,
        "actual_equivalent_date": cost_equivalent_date,
        "accepted": metrics["accepted"],
        "eac": metrics["eac"],
        "rate_3m": metrics["act_cost_rate_3m"],
        "progress_kind": "accepted_cost_ratio",
        "progress_label": "КС / EAC",
        "last_act": metrics["last_act"],
        "forecast_finish": forecast,
        "delta_days": (forecast - finish).days if forecast else None,
        "status": "ЗАВЕРШЕНО ПО УТВЕРЖДЕННОМУ ГПР" if closed else "В РАБОТЕ",
        "schedule_closed": closed,
        "children": tasks,
    }


def _unique_financial(children: list[dict[str, Any]]) -> tuple[float, float, float | None]:
    by_code: dict[str, dict[str, Any]] = {}
    for item in children:
        code = str(item.get("code") or "")
        if code and code not in by_code:
            by_code[code] = item
    if not by_code:
        return 0.0, 0.0, None
    eac = sum(float(item.get("eac") or 0.0) for item in by_code.values())
    accepted = sum(float(item.get("accepted") or 0.0) for item in by_code.values())
    return eac, accepted, accepted / eac if eac > 0 else None


def _summary(
    children: list[dict[str, Any]],
    name: str,
    key: str,
    level: str,
) -> dict[str, Any]:
    start = min(_as_date(item["plan_start"]) for item in children)
    finish = max(_as_date(item["plan_finish"]) for item in children)
    if start is None or finish is None:
        raise ValueError("в baseline отсутствуют даты управленческого блока")

    closed = bool(children) and all(bool(item.get("schedule_closed")) for item in children)
    forecasts = [
        _as_date(item.get("forecast_finish"))
        for item in children
        if _as_date(item.get("forecast_finish")) is not None
    ]
    forecast = finish if closed else (max(forecasts) if forecasts else finish)
    eac, accepted, accepted_ratio = _unique_financial(children)

    duration = max(1, (finish - start).days)
    cost_equivalent_date = (
        start + datetime.timedelta(
            days=round(duration * min(max(accepted_ratio or 0.0, 0.0), 1.0))
        )
        if accepted_ratio is not None else None
    )
    return {
        "key": key,
        "level": level,
        "name": name,
        "plan_start": start,
        "plan_finish": finish,
        "plan_progress": None,
        "actual_progress": accepted_ratio,
        "actual_equivalent_date": cost_equivalent_date,
        "accepted": accepted,
        "eac": eac,
        "progress_kind": "accepted_cost_ratio",
        "progress_label": "КС / EAC",
        "forecast_finish": forecast,
        "delta_days": (forecast - finish).days if forecast else None,
        "status": "ЗАВЕРШЕНО" if closed else "В РАБОТЕ",
        "schedule_closed": closed,
        "children": children,
    }


def _management(
    project: str,
    rss: Path,
    schedule: dict[str, Any],
) -> list[dict[str, Any]]:
    estimate = actuals.read_estimate(rss)
    works = actuals.read_completed_works(rss)
    mapping = _baseline_mapping(project)
    cut = _as_date(schedule.get("cut"))
    if cut is None:
        raise ValueError("не задана дата среза")

    grouped: dict[tuple[str, str, str], list[dict[str, Any]]] = {}
    for raw_task in schedule.get("rows") or []:
        task = dict(raw_task)
        start = _as_date(task.get("plan_start"))
        finish = _as_date(task.get("plan_finish"))
        if start is None or finish is None:
            continue
        task["plan_start"] = start
        task["plan_finish"] = finish
        task["forecast_finish"] = _as_date(task.get("forecast_finish")) or finish
        task["level"] = "task"
        task["schedule_closed"] = bool(task.get("baseline_closed"))
        task["actual_progress"] = None
        task["actual_equivalent_date"] = None
        task["accepted"] = None
        task["eac"] = None
        task["progress_kind"] = "wbs_schedule"

        codes = _codes(task.get("code"))
        for code in codes:
            control, detail = _schedule_bucket(task, code)
            grouped.setdefault((control, detail, code), []).append(task)

    rss_rows: list[dict[str, Any]] = []
    for (control, detail, code), tasks in grouped.items():
        rss_rows.append(_rss_row(
            code, tasks, estimate, works, mapping, cut, control, detail
        ))
    rss_rows.sort(key=lambda item: (
        _natural(item["control"]),
        _natural(item["detail"]),
        _natural(item["code"]),
    ))

    detail_buckets: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for row in rss_rows:
        detail_buckets.setdefault((row["control"], row["detail"]), []).append(row)

    details: list[dict[str, Any]] = []
    for (control, detail), children in detail_buckets.items():
        item = _summary(children, detail, f"detail:{control}:{detail}", "detail")
        item["control"] = control
        item["sort_code"] = min((child["code"] for child in children), key=_natural)
        details.append(item)

    control_buckets: dict[str, list[dict[str, Any]]] = {}
    for row in details:
        control_buckets.setdefault(row["control"], []).append(row)

    controls: list[dict[str, Any]] = []
    for control, children in control_buckets.items():
        item = _summary(children, control, f"control:{control}", "control")
        item["sort_code"] = min(
            (child["sort_code"] for child in children), key=_natural
        )
        controls.append(item)

    order = {
        "Подготовка": 0,
        "Организация стройплощадки / ПОС": 1,
        "Основные объекты": 2,
        "Наружные сети": 3,
        "Благоустройство": 4,
    }
    controls.sort(key=lambda item: (order.get(item["name"], 99), _natural(item["sort_code"])))
    return controls


def _payment_baseline(project: str) -> dict[str, Any]:
    """Read Plan from fixed ``CF ПЛАН-ФАКТ`` baseline."""
    from openpyxl import load_workbook

    folder = monitor._project_dir(project) / "baseline"
    for path in (folder / "finance.xlsx", folder / "gpr.xlsx"):
        if not path.exists():
            continue
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            if "CF ПЛАН-ФАКТ" not in workbook.sheetnames:
                continue
            rows = list(workbook["CF ПЛАН-ФАКТ"].iter_rows(values_only=True))
            plan_row = None
            date_row = None
            total_row = None
            for index, row in enumerate(rows):
                normalized = [
                    str(value or "").strip().lower().replace("ё", "е")
                    for value in row
                ]
                if plan_row is None and normalized.count("план") >= 2:
                    plan_row = index
                    for prior in range(index - 1, max(-1, index - 6), -1):
                        if sum(
                            isinstance(value, (datetime.date, datetime.datetime))
                            for value in rows[prior]
                        ) >= 2:
                            date_row = prior
                            break
                if row and str(row[0] or "").strip().lower() == "итого проект":
                    total_row = index
            if plan_row is None or date_row is None or total_row is None:
                continue

            dates: dict[int, datetime.date] = {}
            last = None
            for column, value in enumerate(rows[date_row]):
                if isinstance(value, datetime.datetime):
                    last = value.date().replace(day=1)
                elif isinstance(value, datetime.date):
                    last = value.replace(day=1)
                if last:
                    dates[column] = last
            plan_columns = [
                column for column, value in enumerate(rows[plan_row])
                if str(value or "").strip().lower() == "план"
            ]

            def series(row: tuple[Any, ...]) -> dict[str, float]:
                result: dict[str, float] = {}
                for column in plan_columns:
                    month = dates.get(column)
                    if month is None:
                        continue
                    value = actuals._money(row[column] if column < len(row) else None)
                    result[month.isoformat()] = value * (
                        1e6 if abs(value) < 100_000 else 1.0
                    )
                return result

            total = series(rows[total_row])
            by_article: dict[str, dict[str, float]] = {}
            for row in rows[plan_row + 1:total_row]:
                label = str(row[0] or "").strip() if row else ""
                if label:
                    by_article[label] = series(row)
            if total:
                return {
                    "known": True,
                    "series": total,
                    "by_article": by_article,
                }
        finally:
            workbook.close()
    return {"known": False, "series": {}, "by_article": {}}


def _payments(project: str, rss: Path) -> dict[str, Any]:
    """Payment fact strictly from current RSS payment register."""
    baseline = _payment_baseline(project)
    payments = actuals.read_payments(rss)
    total_fact: dict[str, float] = {}
    article_fact: dict[str, dict[str, float]] = {}
    code_fact: dict[str, dict[str, float]] = {}

    for row in payments["rows"]:
        date = row.get("date")
        if not date:
            continue
        month = date.replace(day=1).isoformat()
        amount = float(row.get("amount") or 0.0)
        code = str(row.get("estimate_code") or "").rstrip(".")
        article = _financial_control(code) if code else "Не сопоставлено"
        total_fact[month] = total_fact.get(month, 0.0) + amount
        article_fact.setdefault(article, {})
        article_fact[article][month] = article_fact[article].get(month, 0.0) + amount
        if code:
            code_fact.setdefault(code, {})
            code_fact[code][month] = code_fact[code].get(month, 0.0) + amount

    months = sorted(set(total_fact) | set(baseline["series"]))
    rows: list[dict[str, Any]] = []
    for month in months:
        plan = (
            float(baseline["series"].get(month, 0.0))
            if baseline["known"] else None
        )
        fact = float(total_fact.get(month, 0.0))
        rows.append({
            "month": month,
            "plan": plan,
            "fact": fact,
            "delta": fact - plan if plan is not None else None,
        })

    articles: list[dict[str, Any]] = []
    for name in sorted(set(baseline["by_article"]) | set(article_fact)):
        plan_series = baseline["by_article"].get(name, {})
        fact_series = article_fact.get(name, {})
        articles.append({
            "article": name,
            "plan": plan_series,
            "fact": fact_series,
            "plan_total": sum(plan_series.values()),
            "fact_total": sum(fact_series.values()),
        })

    return {
        "known": baseline["known"],
        "source": "CF ПЛАН-ФАКТ" if baseline["known"] else "",
        "rows": rows,
        "plan_total": (
            sum(baseline["series"].values()) if baseline["known"] else None
        ),
        "fact_total": sum(total_fact.values()),
        "last_fact": monitor._iso(payments.get("last")),
        "articles": articles,
        "by_code_fact": code_fact,
        "fact_source": "Реестр платежей",
    }


def _attach_payments(nodes: list[dict[str, Any]], cash: dict[str, Any]) -> None:
    by_article = {item["article"]: item for item in cash.get("articles", [])}
    by_code = cash.get("by_code_fact", {})

    def visit(node: dict[str, Any]) -> None:
        level = node.get("level")
        if level == "control":
            financial_name = node["name"]
            item = by_article.get(financial_name, {})
            node["payments"] = {
                "plan": item.get("plan", {}),
                "fact": item.get("fact", {}),
                "plan_total": item.get("plan_total"),
                "fact_total": item.get("fact_total", 0.0),
            }
        elif level == "rss":
            fact = by_code.get(node.get("code", ""), {})
            node["payments"] = {
                "plan": {},
                "fact": fact,
                "plan_total": None,
                "fact_total": sum(fact.values()),
            }
        for child in node.get("children", []):
            if isinstance(child, dict) and child.get("level"):
                visit(child)

    for node in nodes:
        visit(node)


def _build(
    project: str,
    cut: Any,
    programme: dict[str, Any] | None = None,
    upto: str = "",
) -> dict[str, Any]:
    view = _ORIGINAL_BUILD(project, cut, programme=programme, upto=upto)
    rss = monitor._latest(project, "estimate", ".xlsx", upto)
    if rss is None:
        return view

    _sanitize_base_schedule(project, view["schedule"])
    _seed_rebaselines(project, view["schedule"])

    management = _management(project, rss, view["schedule"])
    cash = _payments(project, rss)
    _attach_payments(management, cash)
    view["schedule"]["management"] = monitor._plain(management)
    view["schedule"]["fact_source"] = "Реестр выполненных работ (КС, стоимостное свидетельство)"
    view["schedule"]["forecast_method"] = (
        "утвержденный ГПР/PM + утвержденные rebaseline + зависимости; "
        "денежный темп актирования календарные даты не двигает"
    )
    view["payments"] = monitor._plain(cash)

    works = actuals.read_completed_works(rss)
    view["money"]["accepted"] = float(works.get("construction_dated") or 0.0)
    view["money"]["payment_fact"] = float(cash.get("fact_total") or 0.0)
    return view


def _gantt(project: str, cut: Any, upto: str = "") -> dict[str, Any]:
    view = _build(project, cut, upto=upto)
    schedule = view["schedule"]
    return {
        "cut": view["cut"],
        "management": schedule.get("management", []),
        "rows": schedule.get("rows", []),
        "works": len(schedule.get("rows", [])),
        "overdue": schedule.get("risks", 0),
        "baseline_end": schedule.get("baseline_end"),
        "forecast_end": schedule.get("forecast_end"),
        "source": {
            "schedule": "fixed-baseline",
            "estimate": view["source"]["estimate"],
            "with_baseline": True,
            "fact": "Реестр выполненных работ (КС / EAC)",
            "payments": "Реестр платежей",
        },
    }


def install() -> None:
    global _INSTALLED, _ORIGINAL_BUILD, _ORIGINAL_GANTT
    if _INSTALLED:
        return
    _ORIGINAL_BUILD = monitor.build
    _ORIGINAL_GANTT = monitor.gantt
    monitor.build = _build
    monitor.gantt = _gantt
    _INSTALLED = True
