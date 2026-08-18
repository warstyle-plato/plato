"""Факт действующего проекта: РСС, реестр договоров и сводка одного с другим.

Инвестиционный анализ считает проект с нуля. Действующий проект такого не
позволяет: полтора года уже случились, и модель обязана начинать не с нулевых
остатков, а с того, что произошло. Здесь живёт чтение этого «произошло».

Источников два, и они разные по природе.

**РСС** — банковская форма по 214-ФЗ (п. 3.3.8 НКЛ). Даёт четыре колонки на
каждую статью: смета, заключено договоров, оплачено, выполнено по принятым
КС. Это единственное место, где видно физическое выполнение.

**Реестр договоров** финансовой модели — построчный факт: договор, платёж,
акт, и при каждой строке **обе** кодировки сразу, «Код банк» (код РСС) и код
БДДС. Перекодировка между РСС и ДДС уже существует внутри этого листа, и
изобретать её не надо: 3 760 строк дают 97 связок.

Из чего собраны правила ниже.

- **Код РСС — ключ сшивки, и он пишется с точкой на конце.** «1.1.» и «1.1» —
  один и тот же код; сравнивать надо нормализованные.

- **Иерархия из кода не выводится.** В РСС Гродненской подстроки внутренних
  инженерных систем пронумерованы `2.2.3.1`…`2.2.3.5` под заголовком `2.3.`
  (суммы при этом сходятся — описка в кодах, не в деньгах). Обход «по префиксу
  кода» отнесёт их к главе 2.2 и получит двойной счёт, а `2.3` объявит листом.
  Дерево поэтому строится по порядку строк и глубине кода: родитель — ближайшая
  строка выше с меньшим числом сегментов. Итог берётся из объявленной строки
  «Всего инвестиционные расходы», а не суммированием листа.

- **Сравнивать уровень с уровнем.** Строка РСС уже агрегат, а реестр разносит
  платёж по листьям — и прямое сличение кода с кодом показывает главу 2 как
  расхождение на 1,9 млрд ₽ там, где расхождения нет вовсе. Платежи реестра
  поднимаются по дереву, и на каждом уровне сравниваются сопоставимые суммы.

- **Число в выгрузке бывает текстом.** «43 212,18» с неразрывным пробелом и
  запятой — обычное дело для 1С. `float()` на таком падает, и падает он не при
  разборе, а посреди свода.

- **Дата акта живёт не в колонке даты.** «Отчетный период» реестра выполненных
  работ заполнен у 47 строк из 1451; настоящая дата — внутри «№ и дата»,
  строкой вида «516 от 19.10.2024». Разбирать её надо от слова «от», а не
  первым похожим на дату куском: номер акта сам полон цифр с точками, и
  наивный поиск даёт «от 45485» и год 204.

- **Не всё, что названо выполнением, является актом КС.** Из 3 432,5 млн ₽
  «выполнено» 881,2 млн — платежи в ДГИ за ВРИ, комиссии банка и ФОТ, у них
  дат актов нет и быть не может. Физическое выполнение считается по датируемой
  части, а недатируемая уходит в деньги; смешивать их — значит показать
  строительную готовность там, где её нет.

- **Расхождение источников — результат, а не помеха.** Сводка ничего не
  подгоняет и не выбирает «более правильный» файл. Она печатает обе цифры и
  разницу: на Гродненской это 2,27 млрд ₽ по смете и 454,9 млн ₽ по оплате, и
  ровно это надо увидеть, а не спрятать.
"""

from __future__ import annotations

import datetime
import re
from pathlib import Path
from typing import Any

# Лист РСС и его колонки. Заголовок стоит на девятой строке, данные ниже;
# первый столбец несёт код, четвёртый — статью затрат.
_ESTIMATE_SHEET = "Расчет стоимости строительства"
_ESTIMATE_TOTAL_LABEL = "всего инвестиционные расходы"
_ESTIMATE_COLUMNS = {
    "code": 0,
    "article": 3,
    "estimate": 4,
    "contracted": 5,
    "paid": 6,
    "completed": 10,  # блок «по объекту»: выполнено по принятым КС
}

# Реестр договоров финансовой модели. Заголовок на двенадцатой строке.
_REGISTER_SHEET = "факт"
_REGISTER_HEADER_ROW = 12
_REGISTER_COLUMNS = {
    "project": 1,
    "company": 2,
    "kind": 3,
    "counterparty": 4,
    "contract": 5,
    "estimate_code": 10,   # «Код банк» — код РСС
    "bdds_code": 12,
    "article": 13,
    "contract_amount": 15,
    "plan_or_fact": 17,
    "paid_date": 18,
    "paid_amount": 19,
    "act_date": 20,
    "act_amount": 21,
    "object": 22,
    "limit": 23,
}

# Реестр актов РСС. Заголовок на шестой строке, данные с девятой.
_WORKS_SHEET = "Реестр выполненных работ"
_WORKS_FIRST_ROW = 9
_WORKS_COLUMNS = {
    "document": 1,
    "number_and_date": 2,
    "period": 3,
    "object": 4,
    "estimate_code": 6,
    "amount": 8,
    "contractor": 10,
    "contract": 11,
}

# Строка «выполнено», у которой нет и не может быть акта КС: плата городу,
# комиссии банка, фонд оплаты труда. В физическое выполнение они не идут.
_NON_CONSTRUCTION_MARKERS = ("уфк", "дги", "сбербанк", "фот")

_EXCEL_EPOCH = datetime.datetime(1899, 12, 30)


def _money(value: Any) -> float:
    """Число из выгрузки. 1С отдаёт «43 212,18» строкой — это тоже число."""
    if value is None:
        return 0.0
    if isinstance(value, bool):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def _code(value: Any) -> str:
    """Код РСС без хвостовой точки: «1.1.» и «1.1» — один код."""
    if value is None:
        return ""
    text = str(value).strip()
    if not text or text.startswith("#"):
        return ""
    text = re.sub(r"\.+$", "", text)
    return text if re.fullmatch(r"\d+(\.\d+)*", text) else ""


def _date(value: Any) -> datetime.date | None:
    """Дата из ячейки, серийного номера или строки «№ … от 19.10.2024».

    Разбор идёт от слова «от»: номер акта сам полон цифр с точками, и поиск
    первого похожего на дату куска находит номер, а не дату.
    """
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    if isinstance(value, (int, float)) and 10_000 < float(value) < 80_000:
        return (_EXCEL_EPOCH + datetime.timedelta(days=int(value))).date()
    if not isinstance(value, str):
        return None
    tail = re.split(r"\bот\b", value)[-1].strip()
    match = re.match(r"(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{2,4})", tail)
    if match:
        day, month, year = (int(part) for part in match.groups())
        if year < 100:
            year += 2000
        # Год «204» — описка источника, а не 204-й год: такую дату честнее
        # не признать, чем поставить месяц в 204-12 и увести его из свода.
        if not 2000 <= year <= 2100:
            return None
        try:
            return datetime.date(year, month, day)
        except ValueError:
            return None
    if re.fullmatch(r"\d{5}", tail):
        return (_EXCEL_EPOCH + datetime.timedelta(days=int(tail))).date()
    return None


def _month(value: datetime.date | None) -> str:
    return value.strftime("%Y-%m") if value else ""


def _normalized(text: Any) -> str:
    return re.sub(r"\s+", " ", str(text or "").replace("ё", "е")).strip().lower()


def _sheet(path: str | Path, name: str) -> Any:
    from openpyxl import load_workbook

    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        if name not in workbook.sheetnames:
            raise KeyError(f"в книге нет листа «{name}»: {workbook.sheetnames}")
        for row in workbook[name].iter_rows(values_only=True):
            yield row
    finally:
        workbook.close()


def read_estimate(path: str | Path) -> dict[str, Any]:
    """РСС: статьи с кодами и четыре колонки факта по каждой.

    Итог не суммируется — он объявлен строкой «Всего инвестиционные расходы».
    Суммировать лист нельзя: в нём и главы, и подстроки глав.
    """
    rows: list[dict[str, Any]] = []
    total: dict[str, float] = {}
    for row in _sheet(path, _ESTIMATE_SHEET):
        def cell(field: str) -> Any:
            index = _ESTIMATE_COLUMNS[field]
            return row[index] if index < len(row) else None

        article = cell("article")
        if total == {} and _normalized(article).startswith(_ESTIMATE_TOTAL_LABEL):
            total = {
                field: _money(cell(field))
                for field in ("estimate", "contracted", "paid", "completed")
            }
            continue
        code = _code(cell("code"))
        if not code:
            continue
        rows.append({
            "code": code,
            "article": str(article or "").strip(),
            "depth": code.count(".") + 1,
            **{field: _money(cell(field))
               for field in ("estimate", "contracted", "paid", "completed")},
        })
    _link_parents(rows)
    return {"rows": rows, "total": total,
            "by_code": {row["code"]: row for row in rows}}


def _link_parents(rows: list[dict[str, Any]]) -> None:
    """Родитель — ближайшая строка выше с меньшим числом сегментов кода.

    Порядок строк здесь надёжнее самого кода: в РСС Гродненской подстроки
    внутренних инженерных систем пронумерованы `2.2.3.x` под заголовком `2.3.`,
    и по префиксу они уехали бы в другую главу, а `2.3` осталось бы листом.
    """
    stack: list[dict[str, Any]] = []
    for row in rows:
        while stack and stack[-1]["depth"] >= row["depth"]:
            stack.pop()
        row["parent"] = stack[-1]["code"] if stack else ""
        stack.append(row)
    parents = {row["parent"] for row in rows if row["parent"]}
    for row in rows:
        row["is_leaf"] = row["code"] not in parents


def _rolled_up(rows: list[dict[str, Any]], amounts: dict[str, float]) -> dict[str, float]:
    """Поднять суммы реестра по дереву РСС: лист даёт вклад всем родителям."""
    parent_of = {row["code"]: row["parent"] for row in rows}
    rolled = {row["code"]: 0.0 for row in rows}
    for code, amount in amounts.items():
        cursor = code
        seen: set[str] = set()
        while cursor and cursor not in seen:
            seen.add(cursor)
            if cursor in rolled:
                rolled[cursor] += amount
            cursor = parent_of.get(cursor, "")
    return rolled


def read_register(path: str | Path) -> dict[str, Any]:
    """Реестр договоров финансовой модели: платежи, акты и обе кодировки."""
    rows: list[dict[str, Any]] = []
    unmapped_paid = 0.0
    for index, row in enumerate(_sheet(path, _REGISTER_SHEET), 1):
        if index <= _REGISTER_HEADER_ROW:
            continue

        def cell(field: str) -> Any:
            position = _REGISTER_COLUMNS[field]
            return row[position] if position < len(row) else None

        raw_code = cell("estimate_code")
        if raw_code is None and cell("bdds_code") is None:
            continue
        code = _code(raw_code)
        paid = _money(cell("paid_amount"))
        if not code:
            unmapped_paid += paid
        rows.append({
            "estimate_code": code,
            "raw_estimate_code": str(raw_code or "").strip(),
            "bdds_code": str(cell("bdds_code") or "").strip(),
            "article": str(cell("article") or "").strip(),
            "counterparty": str(cell("counterparty") or "").strip(),
            # Номер договора — вторая половина ключа сшивки с реестром
            # платежей. Без него ключ вырождается в контрагента, а у крупного
            # подрядчика договоров несколько и коды у них разные.
            "contract": str(cell("contract") or "").strip(),
            "kind": str(cell("kind") or "").strip(),
            "object": str(cell("object") or "").strip(),
            "plan_or_fact": _normalized(cell("plan_or_fact")),
            "contract_amount": _money(cell("contract_amount")),
            "limit": _money(cell("limit")),
            "paid_amount": paid,
            "paid_date": _date(cell("paid_date")),
            "act_amount": _money(cell("act_amount")),
            "act_date": _date(cell("act_date")),
        })
    crosswalk: dict[str, set[str]] = {}
    for item in rows:
        if item["estimate_code"] and item["bdds_code"]:
            crosswalk.setdefault(item["estimate_code"], set()).add(item["bdds_code"])
    return {
        "rows": rows,
        "crosswalk": {code: sorted(codes) for code, codes in crosswalk.items()},
        "unmapped_paid": unmapped_paid,
        "paid": sum(item["paid_amount"] for item in rows),
        "accepted": sum(item["act_amount"] for item in rows),
    }


def read_completed_works(path: str | Path) -> dict[str, Any]:
    """Реестр актов РСС: суммы, коды и дата, вынутая из «№ и дата».

    Строки без даты не выбрасываются: их сумма объявляется отдельно, иначе
    физическое выполнение молча недосчитается. Строки, которые актом КС не
    являются (плата городу, комиссии банка, ФОТ), помечаются здесь же.
    """
    rows: list[dict[str, Any]] = []
    for index, row in enumerate(_sheet(path, _WORKS_SHEET), 1):
        if index < _WORKS_FIRST_ROW:
            continue

        def cell(field: str) -> Any:
            position = _WORKS_COLUMNS[field]
            return row[position] if position < len(row) else None

        amount = _money(cell("amount"))
        if not amount and cell("document") is None:
            continue
        contractor = _normalized(cell("contractor"))
        rows.append({
            "code": _code(cell("estimate_code")),
            "document": str(cell("document") or "").strip(),
            "contractor": str(cell("contractor") or "").strip(),
            "contract": str(cell("contract") or "").strip(),
            "object": str(cell("object") or "").strip(),
            "amount": amount,
            "date": _date(cell("number_and_date")) or _date(cell("period")),
            "construction": not any(mark in contractor
                                    for mark in _NON_CONSTRUCTION_MARKERS),
        })
    dated = [item for item in rows if item["date"]]
    return {
        "rows": rows,
        "total": sum(item["amount"] for item in rows),
        "dated": sum(item["amount"] for item in dated),
        "undated": sum(item["amount"] for item in rows if not item["date"]),
        "construction_dated": sum(item["amount"] for item in dated
                                  if item["construction"]),
    }


def monthly(register: dict[str, Any], works: dict[str, Any]) -> dict[str, Any]:
    """Помесячный факт: деньги из реестра договоров, объёмы из актов РСС.

    Две шкалы намеренно раздельны. Деньги и выполнение расходятся во времени —
    аванс уходит раньше акта, а удержание позже, — и слитая строка не покажет
    ни того ни другого.
    """
    paid: dict[str, float] = {}
    for item in register["rows"]:
        if item["paid_amount"] and item["paid_date"]:
            key = _month(item["paid_date"])
            paid[key] = paid.get(key, 0.0) + item["paid_amount"]
    accepted: dict[str, float] = {}
    for item in works["rows"]:
        if item["amount"] and item["date"] and item["construction"]:
            key = _month(item["date"])
            accepted[key] = accepted.get(key, 0.0) + item["amount"]
    months = sorted(set(paid) | set(accepted))
    running_paid = running_accepted = 0.0
    series = []
    for month in months:
        running_paid += paid.get(month, 0.0)
        running_accepted += accepted.get(month, 0.0)
        series.append({
            "month": month,
            "paid": paid.get(month, 0.0),
            "accepted": accepted.get(month, 0.0),
            "paid_cumulative": running_paid,
            "accepted_cumulative": running_accepted,
        })
    return {"series": series,
            "first": months[0] if months else "",
            "last": months[-1] if months else ""}


def reconcile(
    estimate: dict[str, Any],
    register: dict[str, Any],
    works: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Свести РСС с реестром договоров по коду. Ничего не подгонять.

    Расхождение — это результат: обе цифры и разница печатаются рядом, а
    выбирать «более правильный» источник сводка не вправе.
    """
    register_paid: dict[str, float] = {}
    register_acts: dict[str, float] = {}
    for item in register["rows"]:
        code = item["estimate_code"]
        if not code:
            continue
        register_paid[code] = register_paid.get(code, 0.0) + item["paid_amount"]
        register_acts[code] = register_acts.get(code, 0.0) + item["act_amount"]

    # Реестр разносит платёж по листьям, строка РСС — уже агрегат. Поднимаем
    # реестр по дереву, иначе глава сравнивается с нулём и «расходится» на всю
    # свою сумму.
    rolled_paid = _rolled_up(estimate["rows"], register_paid)
    rolled_acts = _rolled_up(estimate["rows"], register_acts)
    outside = sorted(set(register_paid) - set(estimate["by_code"]),
                     key=lambda value: [int(part) for part in value.split(".")])

    by_code = []
    for row in estimate["rows"]:
        code = row["code"]
        paid_estimate = float(row.get("paid", 0.0))
        paid_register = rolled_paid.get(code, 0.0)
        by_code.append({
            "code": code,
            "article": row.get("article", ""),
            "depth": row.get("depth", 1),
            "is_leaf": bool(row.get("is_leaf")),
            "estimate": float(row.get("estimate", 0.0)),
            "contracted": float(row.get("contracted", 0.0)),
            "completed": float(row.get("completed", 0.0)),
            "paid_estimate": paid_estimate,
            "paid_register": paid_register,
            "paid_delta": paid_register - paid_estimate,
            "acts_register": rolled_acts.get(code, 0.0),
        })

    warnings: list[str] = []
    orphans = outside
    if orphans:
        warnings.append(
            "кодов реестра нет в РСС: " + ", ".join(orphans[:8]))
    if register["unmapped_paid"]:
        warnings.append(
            f"оплата без кода РСС: {register['unmapped_paid'] / 1e6:,.1f} млн ₽")
    total = estimate.get("total") or {}
    paid_estimate_total = float(total.get("paid", 0.0))
    paid_gap = register["paid"] - paid_estimate_total
    # Порог относительный. Зашитый «миллион рублей» молчал бы на любой книге,
    # где суммы ведутся не в рублях, а расхождение там такое же настоящее.
    if abs(paid_gap) > max(1.0, 0.0001 * abs(paid_estimate_total)):
        warnings.append(
            f"оплачено: РСС {paid_estimate_total / 1e6:,.1f} против реестра "
            f"{register['paid'] / 1e6:,.1f} млн ₽, разрыв {paid_gap / 1e6:,.1f}")
    if works and works["undated"]:
        warnings.append(
            f"выполнение без даты: {works['undated'] / 1e6:,.1f} млн ₽ "
            "(плата городу, комиссии банка, ФОТ — не акты КС)")

    return {
        "by_code": by_code,
        "total": {
            "estimate": float(total.get("estimate", 0.0)),
            "contracted": float(total.get("contracted", 0.0)),
            "paid_estimate": float(total.get("paid", 0.0)),
            "paid_register": register["paid"],
            "completed": float(total.get("completed", 0.0)),
            "acts_register": register["accepted"],
        },
        "crosswalk": register["crosswalk"],
        "warnings": warnings,
    }


def _report(estimate_path: str, register_path: str) -> str:
    """Свод в текст: чтобы прогнать очередную выгрузку, не написав ни строки."""
    estimate = read_estimate(estimate_path)
    register = read_register(register_path)
    works = read_completed_works(estimate_path)
    report = reconcile(estimate, register, works)
    lines = [
        f'{"":<34}{"смета":>10}{"договоры":>10}{"опл.РСС":>10}'
        f'{"опл.реестр":>11}{"Δ":>9}{"КС":>10}{"%вып":>7}',
    ]

    def line(label: str, row: dict[str, Any], delta: float) -> str:
        done = row["completed"] / row["estimate"] * 100 if row["estimate"] else 0.0
        return (f'{label[:32]:<34}{row["estimate"] / 1e6:>10,.1f}'
                f'{row["contracted"] / 1e6:>10,.1f}{row["paid_estimate"] / 1e6:>10,.1f}'
                f'{row["paid_register"] / 1e6:>11,.1f}{delta / 1e6:>9,.1f}'
                f'{row["completed"] / 1e6:>10,.1f}{done:>6,.1f}%')

    for row in report["by_code"]:
        if row["depth"] == 1:
            lines.append(line(row["article"], row, row["paid_delta"]))
    total = report["total"]
    lines.append(line("ИТОГО", {**total, "paid_estimate": total["paid_estimate"],
                                "paid_register": total["paid_register"]},
                      total["paid_register"] - total["paid_estimate"]))
    series = monthly(register, works)["series"]
    lines.append("")
    lines.append(f"помесячно: {len(series)} мес., "
                 f"{series[0]['month'] if series else '—'} … "
                 f"{series[-1]['month'] if series else '—'}")
    lines.append(f"связок код РСС ↔ код БДДС: {len(report['crosswalk'])}")
    for warning in report["warnings"]:
        lines.append(f"  • {warning}")
    return "\n".join(lines)


if __name__ == "__main__":  # pragma: no cover
    import sys

    if len(sys.argv) != 3:
        print("использование: python3 developaid_actuals.py РСС.xlsx финмодель.xlsx")
        raise SystemExit(2)
    print(_report(sys.argv[1], sys.argv[2]))


# ---------------------------------------------------------------------------
# Наложение факта на план.
#
# Движок разворачивает весь проект от даты старта: график продаж, освоение
# CAPEX, выборку долга. Для действующего проекта прошлое выдумывать не надо —
# оно случилось. Наложение подменяет ряды до даты среза фактическими, а
# плановый хвост перенормирует на остаток.
#
# Два правила, из которых всё следует.
#
# **Итог не меняется от того, что часть его уже случилась.** Если подменить
# прошлое фактом и оставить будущее как было, проект посчитается дважды: разом
# и по факту, и по плану. Поэтому хвост масштабируется так, чтобы факт плюс
# хвост давали прежний итог. Форма хвоста при этом сохраняется — движок про
# сроки знает больше, чем мы про их пересмотр.
#
# **Остаток, которому некуда лечь, не исчезает.** Если планового хвоста нет
# вовсе, а остаток есть, деньги молча растворились бы в перенормировке. Такой
# остаток кладётся в месяц среза и объявляется в отчёте: лучше спорная дата,
# чем пропавшая сумма.
# ---------------------------------------------------------------------------


def _as_month(value: Any) -> datetime.date | None:
    """Ключ месяца к виду движка: первое число, тип `date`."""
    if isinstance(value, datetime.datetime):
        return value.date().replace(day=1)
    if isinstance(value, datetime.date):
        return value.replace(day=1)
    if isinstance(value, str):
        match = re.match(r"(\d{4})-(\d{2})", value.strip())
        if match:
            return datetime.date(int(match.group(1)), int(match.group(2)), 1)
    return None


def _months(series: dict[Any, float]) -> dict[datetime.date, float]:
    out: dict[datetime.date, float] = {}
    for key, value in (series or {}).items():
        month = _as_month(key)
        if month is None:
            continue
        out[month] = out.get(month, 0.0) + float(value or 0.0)
    return out


def _blend(
    plan: dict[datetime.date, float],
    fact: dict[datetime.date, float],
    cut: datetime.date,
    label: str,
    notes: list[str],
) -> dict[datetime.date, float]:
    """Факт до среза, перенормированный на остаток план после."""
    planned_total = sum(plan.values())
    blended = {month: value for month, value in fact.items() if month < cut}

    # Ряд фактом покрыт не до самого среза: выгрузки обрываются в разное время.
    # На Гродненской оплаты идут до 04.2026, акты КС — до 06.2026, продажи — до
    # 03.2026, а срез по РСС стоит на 01.07.2026. Месяц, до которого факт не
    # дотянулся, — это «не знаем», а не «было ноль»: прочитать его нулём значит
    # объявить, что три месяца ничего не платили и ничего не продали, и дальше
    # уедут эскроу, покрытие, ставка ПФ и LLCR — все правдоподобно.
    observed = sorted(blended)
    if observed:
        first, last = observed[0], observed[-1]
        carried = {month: value for month, value in plan.items()
                   if month < cut and not (first <= month <= last)}
        if carried:
            blended.update(carried)
            notes.append(
                f"{label}: факт покрывает {first:%Y-%m}…{last:%Y-%m}, а срез стоит "
                f"на {cut:%Y-%m} — за {len(carried)} мес. вне покрытия оставлен план "
                f"({sum(carried.values()) / 1e6:,.1f} млн ₽)")

    fact_total = sum(blended.values())
    tail = {month: value for month, value in plan.items() if month >= cut}
    tail_total = sum(tail.values())
    remainder = planned_total - fact_total
    if remainder <= 0:
        if remainder < -0.005 * max(abs(planned_total), 1.0):
            notes.append(
                f"{label}: факт до среза {fact_total / 1e6:,.1f} млн ₽ превысил план "
                f"{planned_total / 1e6:,.1f} — остаток обнулён, перерасход "
                f"{-remainder / 1e6:,.1f} млн ₽")
        return blended
    if tail_total > 0:
        factor = remainder / tail_total
        for month, value in tail.items():
            blended[month] = blended.get(month, 0.0) + value * factor
    else:
        blended[cut] = blended.get(cut, 0.0) + remainder
        notes.append(
            f"{label}: планового хвоста после среза нет, остаток "
            f"{remainder / 1e6:,.1f} млн ₽ отнесён на месяц среза")
    return blended


def overlay(op: dict[str, Any], actuals: dict[str, Any]) -> dict[str, Any]:
    """Подменить ряды операционной модели фактом до даты среза.

    `actuals` несёт `cut` — первый месяц прогноза — и ряды факта: `capex`,
    `operating`, `revenue` и `quantity` по продуктам. Чего нет, то остаётся
    плановым: отсутствующий ряд — это «не знаем», а не «было ноль».

    Возвращает новую модель и отчёт о наложении; сам `op` не меняется.
    """
    cut = _as_month(actuals.get("cut"))
    if cut is None:
        raise ValueError("не задана дата среза (`cut`)")

    notes: list[str] = []
    report: dict[str, Any] = {"cut": cut.isoformat(), "series": {}, "notes": notes}
    updated = dict(op)

    def record(label: str, plan: dict, blended: dict) -> None:
        report["series"][label] = {
            "plan": sum(plan.values()),
            "fact": sum(v for m, v in blended.items() if m < cut),
            "forecast": sum(v for m, v in blended.items() if m >= cut),
        }

    # --- CAPEX -------------------------------------------------------------
    if "capex" in actuals:
        plan_capex = dict(op.get("capex") or {})
        fact_capex = _months(actuals["capex"])
        blended = _blend(plan_capex, fact_capex, cut, "CAPEX", notes)
        updated["capex"] = blended
        record("capex", plan_capex, blended)
        updated["capex_by_article"] = _split_by_article(
            op.get("capex_by_article") or {}, plan_capex, blended, cut, notes)

    # Бюджет вместо норматива. На действующем проекте расходную часть считать
    # незачем — она посчитана и утверждена, и удельные ставки движка ей заведомо
    # проигрывают: на Гродненской норматив даёт 176,1 тыс ₽/м² ГНС против 302,7
    # фактических, а ПИР и технадзор, наоборот, вдвое-втрое дороже норматива.
    #
    # Бюджет задаёт сумму, движок — календарь: форма кривой освоения остаётся
    # его, потому что про сроки он знает, а смета их не несёт. Статья, которой в
    # бюджете нет, остаётся нормативной: отсутствие строки — это «не знаем», а
    # не «ноль».
    if "capex_budget" in actuals:
        planned = {a: dict(s) for a, s in (op.get("capex_by_article") or {}).items()}
        overall = dict(op.get("capex") or {})
        overall_total = sum(overall.values())
        rescaled: dict[str, dict[datetime.date, float]] = {}
        for article, series in planned.items():
            rescaled[article] = dict(series)
        for article, amount in (actuals["capex_budget"] or {}).items():
            shape = planned.get(article) or {}
            shape_total = sum(shape.values())
            if shape_total > 0:
                factor = amount / shape_total
                rescaled[article] = {m: v * factor for m, v in shape.items()}
            elif overall_total > 0:
                rescaled[article] = {m: v * amount / overall_total
                                     for m, v in overall.items()}
                notes.append(
                    f"бюджет {article}: {amount / 1e6:,.1f} млн ₽ разложен по общей "
                    "кривой освоения — своей у статьи в плане нет")
            else:
                notes.append(
                    f"бюджет {article}: {amount / 1e6:,.1f} млн ₽ разложить не по "
                    "чему — плановой кривой нет вовсе")
        updated["capex_by_article"] = rescaled
        combined_budget: dict[datetime.date, float] = {}
        for series in rescaled.values():
            for month, value in series.items():
                combined_budget[month] = combined_budget.get(month, 0.0) + value
        updated["capex"] = combined_budget
        op = {**op, "capex_by_article": rescaled, "capex": combined_budget}

    # Факт, уже разнесённый по статьям (карта «код БДДС → статья движка»), —
    # это лучше, чем доли плана: каждая статья складывается из своих платежей.
    # Итог тогда выводится из статей, а не считается вторым способом: два
    # способа на одно число рано или поздно разойдутся.
    if "capex_by_article" in actuals:
        planned_articles = {a: dict(s) for a, s in (op.get("capex_by_article") or {}).items()}
        facts = {a: _months(s) for a, s in (actuals["capex_by_article"] or {}).items()}
        merged: dict[str, dict[datetime.date, float]] = {}
        for article in sorted(set(planned_articles) | set(facts)):
            merged[article] = _blend(
                planned_articles.get(article, {}), facts.get(article, {}),
                cut, f"CAPEX:{article}", notes)
        updated["capex_by_article"] = merged
        combined: dict[datetime.date, float] = {}
        for series in merged.values():
            for month, value in series.items():
                combined[month] = combined.get(month, 0.0) + value
        plan_capex = dict(op.get("capex") or {})
        updated["capex"] = combined
        record("capex", plan_capex, combined)
        blended = combined

    if "capex" in actuals or "capex_by_article" in actuals:
        # Долговой CAPEX — тот же расход за вычетом доли ВРИ, которую закрывает
        # капитал. Пересобираем из нового CAPEX, а не масштабируем отдельно:
        # иначе связь между ними разъедется на второй же правке.
        equity = _months(op.get("vri_equity") or {})
        updated["debt_capex"] = {
            month: max(0.0, value - equity.get(month, 0.0))
            for month, value in updated["capex"].items()
        }

    # --- коммерческие расходы ---------------------------------------------
    if "operating" in actuals:
        plan_operating = dict(op.get("operating") or {})
        blended = _blend(plan_operating, _months(actuals["operating"]), cut,
                         "Коммерческие расходы", notes)
        updated["operating"] = blended
        record("operating", plan_operating, blended)

    # --- продажи -----------------------------------------------------------
    for field, source in (("revenue_product_schedules", "revenue"),
                          ("quantity_product_schedules", "quantity")):
        if source not in actuals:
            continue
        planned = {key: dict(value) for key, value in (op.get(field) or {}).items()}
        facts = actuals[source] or {}
        for product, fact_series in facts.items():
            plan_series = planned.get(product, {})
            planned[product] = _blend(plan_series, _months(fact_series), cut,
                                      f"{source}:{product}", notes)
            record(f"{source}:{product}", plan_series, planned[product])
        updated[field] = planned

    # Выручка не сохраняет свой итог, в отличие от CAPEX. У расходов инвариант —
    # бюджет: сколько бы ни было потрачено до среза, проект стоит столько, во
    # сколько его оценили. У продаж инвариант — метры: проданное дёшево нельзя
    # продать ещё раз. Поэтому объём перенормируется на остаток (метров всего
    # столько, сколько в ТЭП), а выручка хвоста считается заново — плановой
    # ценой на новый объём. Иначе факт продаж по 657 тыс ₽/м² не двигает итог
    # вовсе: он просто отбирает выручку у будущих месяцев, и модель делает вид,
    # что недополученное вернётся.
    if "quantity" in actuals and "revenue" in actuals:
        plan_quantity = {a: dict(s) for a, s in (op.get("quantity_product_schedules") or {}).items()}
        plan_revenue = {a: dict(s) for a, s in (op.get("revenue_product_schedules") or {}).items()}
        rebuilt = {product: dict(series)
                   for product, series in updated["revenue_product_schedules"].items()}
        for product, quantity in updated["quantity_product_schedules"].items():
            if product not in (actuals["quantity"] or {}):
                continue
            fact_revenue = _months((actuals["revenue"] or {}).get(product) or {})
            series = {month: value for month, value in fact_revenue.items() if month < cut}
            planned_q = plan_quantity.get(product, {})
            planned_r = plan_revenue.get(product, {})
            for month, volume in quantity.items():
                if month < cut:
                    continue
                planned_volume = planned_q.get(month, 0.0)
                price = (planned_r.get(month, 0.0) / planned_volume
                         if planned_volume else 0.0)
                if price:
                    series[month] = series.get(month, 0.0) + volume * price
            rebuilt[product] = series
            before = sum(plan_revenue.get(product, {}).values())
            after = sum(series.values())
            if abs(after - before) > 0.005 * max(abs(before), 1.0):
                notes.append(
                    f"выручка {product}: план {before / 1e6:,.1f} → "
                    f"{after / 1e6:,.1f} млн ₽ — факт продаж по своей цене, "
                    "хвост по плановой")
        updated["revenue_product_schedules"] = rebuilt

    if "revenue" in actuals:
        # Свод выручки должен идти из тех же рядов, иначе итог и детализация
        # разойдутся — обе достоверные на вид.
        combined: dict[datetime.date, float] = {}
        for series in updated["revenue_product_schedules"].values():
            for month, value in series.items():
                combined[month] = combined.get(month, 0.0) + value
        updated["revenue"] = combined
        updated["revenue_by_product"] = {
            product: sum(series.values())
            for product, series in updated["revenue_product_schedules"].items()
        }

    # Финансирование до среза — факт. Ряды кладутся в модель как есть: движок
    # берёт с них остатки и не начисляет ничего до среза, а с месяца среза
    # считает от них вперёд.
    if "financing" in actuals:
        finance = actuals["financing"] or {}
        updated["financing_fact"] = {
            "cut": cut,
            "escrow": _months(finance.get("escrow") or {}),
            "pf_balance": _months(finance.get("pf_balance") or {}),
            "bridge_balance": _months(finance.get("bridge_balance") or {}),
            "interest": _months(finance.get("interest") or {}),
        }
        carried = updated["financing_fact"]
        last = max((max(series) for series in
                    (carried["pf_balance"], carried["escrow"]) if series),
                   default=None)
        if last is not None and last < _add_month(cut, -1):
            notes.append(
                f"финансирование: факт остатков идёт по {last:%Y-%m}, а срез "
                f"стоит на {cut:%Y-%m} — за {(cut.year - last.year) * 12 + cut.month - last.month - 1}"
                " мес. остаток держится последним известным")
        report["financing"] = {
            "pf_balance": max((v for m, v in carried["pf_balance"].items()
                               if m < cut), default=0.0),
            "escrow": max((v for m, v in carried["escrow"].items() if m < cut),
                          default=0.0),
            "interest_paid": sum(v for m, v in carried["interest"].items()
                                 if m < cut),
        }

    return {"op": updated, "report": report}


def _split_by_article(
    planned_articles: dict[str, dict[datetime.date, float]],
    plan_capex: dict[datetime.date, float],
    blended: dict[datetime.date, float],
    cut: datetime.date,
    notes: list[str],
) -> dict[str, dict[datetime.date, float]]:
    """Разнести новый CAPEX по статьям, сохранив согласие с итогом.

    Факт приходит по кодам РСС, а не по статьям движка, поэтому до среза
    разнос делается долями плана — того же месяца, а при его отсутствии
    общими долями проекта. Это приближение, и оно объявляется: детализация по
    статьям до среза показывает структуру плана, а не структуру факта. Молча
    оставить старые ряды нельзя — тогда сумма по статьям разойдётся с итогом,
    и обе цифры будут выглядеть достоверно.
    """
    if not planned_articles:
        return {}
    overall_total = sum(sum(series.values()) for series in planned_articles.values())
    if overall_total <= 0:
        return {article: dict(series) for article, series in planned_articles.items()}
    overall_share = {
        article: sum(series.values()) / overall_total
        for article, series in planned_articles.items()
    }

    result: dict[str, dict[datetime.date, float]] = {a: {} for a in planned_articles}
    approximated = False
    for month, amount in blended.items():
        if month >= cut:
            planned_month = plan_capex.get(month, 0.0)
            shares = (
                {a: s.get(month, 0.0) / planned_month for a, s in planned_articles.items()}
                if planned_month > 0 else overall_share
            )
        else:
            approximated = True
            planned_month = plan_capex.get(month, 0.0)
            shares = (
                {a: s.get(month, 0.0) / planned_month for a, s in planned_articles.items()}
                if planned_month > 0 else overall_share
            )
        for article, share in shares.items():
            if share:
                result[article][month] = result[article].get(month, 0.0) + amount * share
    if approximated:
        notes.append(
            "структура CAPEX по статьям до среза разнесена долями плана: факт "
            "приходит по кодам РСС, соответствия статьям движка пока нет")
    return result


# ---------------------------------------------------------------------------
# Карта «код БДДС → статья движка».
#
# Ключ — код БДДС, а не код РСС, и это не вкусовщина. РСС сваливает в один код
# разное: в 2.6 сидят технадзор, ФОТ, временные сооружения и получение РВЭ; в
# 2.2.1.10 — вознаграждение генподрядчика вместе с содержанием площадки; в 2.7
# — ИРД, стадия П, стадия РД и авторский надзор разом. Разложить это по нашим
# статьям, зная только код РСС, нельзя. Кодификатор БДДС иерархичен и разделяет
# ровно то, что нам нужно, а реестр договоров несёт оба кода в одной строке —
# значит выбор ключа ничего не стоит.
#
# Правило разбора — самый длинный совпавший префикс. Так исключение живёт рядом
# с правилом: `2.2.2.3` целиком подземная часть, но `2.2.2.3.8` — содержание
# стройплощадки, и оно уходит в свою статью.
#
# Соответствие снято с листа «статьи_БДДС» финансовой модели Гродненской
# (369 строк, 259 из них несут код РСС). Всё, чему статьи не нашлось,
# возвращается как `None` и попадает в отчёт: статья, разнесённая наугад,
# выглядит на экране так же, как разнесённая верно.
# ---------------------------------------------------------------------------

_BDDS_TO_ARTICLE: tuple[tuple[str, str], ...] = (
    # --- земля и права -----------------------------------------------------
    ("2.1.1.1", "purchase"),            # приобретение прав застройщика
    ("2.1.1.2", "purchase"),            # приобретение ЗУ под сети и прочих
    ("2.1.1.3", "land_rights"),         # межевание, смена ВРИ, нотариат
    ("2.1.1.4", "land_rights"),         # аренда ЗУ, земельный налог
    ("2.1.2", "social"),                # обременения и стройка соцобъектов

    # --- строительство -----------------------------------------------------
    ("2.2.1", "preparation"),
    ("2.2.2.2", "gc_fee"),
    ("2.2.2.3", "main_under"),
    ("2.2.2.3.8", "site_maintenance"),  # содержание площадки внутри подземной
    ("2.2.2.4", "main_above"),
    ("2.2.2.4.5", "site_maintenance"),  # то же в надземной
    ("2.2.2.5", "main_above"),          # фасады и внешнее остекление
    # Отделка МОП разнесена по частям здания: 2.2.2.6.1 — МОП, технических
    # помещений и автопарковок (РСС 2.2.1.7, подземная), 2.2.2.6.2 — надземная.
    ("2.2.2.6", "main_above"),
    ("2.2.2.6.1", "main_under"),
    ("2.2.2.7", "main_above"),          # внутренние инженерные системы
    ("2.2.2.8", "main_above"),          # отделка продаваемых помещений
    ("2.2.2.8.1", "design_p"),          # дизайн-проект отделки — это ПИР
    ("2.2.3", "main_above"),            # подсобные и обслуживающие объекты
    ("2.2.4", "utilities"),             # транспортное хозяйство и связь
    ("2.2.5", "utilities"),             # наружные сети, СМР и техприсоединение
    ("2.2.5.2", "ird"),                 # получение ТУ
    ("2.2.6", "landscaping"),
    ("2.2.7", "site_maintenance"),      # временные здания и сооружения
    ("2.2.8", "site_maintenance"),      # охрана, коммуналка, инвентарь стройки
    ("2.2.8.2", "commissioning"),       # ЗОС, РВЭ, сдача объекта
    ("2.2.9", "technical_supervision"),
    ("2.2.10.1", "ird"),                # АГК, ДПТ, изыскания, массинг
    ("2.2.10.2", "design_p"),
    ("2.2.10.2.2", "design_rd"),
    ("2.2.10.3", "design_p"),           # ПиР сетей
    ("2.2.10.4", "author_supervision"),
    ("2.2.11", "reserve"),

    # --- коммерческие расходы: у движка это не CAPEX, а operating ----------
    ("2.3.1", "selling"),               # персонал отдела продаж
    ("2.3.2", "selling"),               # содержание офиса продаж
    ("2.3.3", "marketing"),             # реклама и PR
    ("2.3.4", "selling"),               # расходы на продажу

    # --- административно-хозяйственные ------------------------------------
    ("2.4", "project_management"),
    ("2.4.5", "tax"),                   # налоги — ни CAPEX, ни коммерческие

    # --- финансовая деятельность ------------------------------------------
    ("3", "financing"),
)

# Коды, которые приходят одной суммой на две статьи. «Строительство ЖК
# (внешний генподряд)» не разделено на наземную и подземную части, а модель
# считает их по разным удельным ставкам. Делится по ГНС — решение владельца,
# 18.08.2026.
#
# Пропорция приходит снаружи, из ТЭП проекта, и не выводится здесь: доля,
# посчитанная модулем разбора выгрузок, была бы второй реализацией ТЭП. Не дали
# пропорцию — код остаётся неразнесённым и называет причину, а не делится
# пополам «чтобы сошлось».
_BDDS_GNS_SPLIT: dict[str, tuple[str, str]] = {
    "2.2.2.1": ("main_above", "main_under"),
}

# Коды, для которых статьи движка нет и придумывать её нельзя.
_BDDS_UNRESOLVED: dict[str, str] = {}

# Статьи, которых в модели нет как CAPEX: они уходят в свои разделы расчёта.
_NON_CAPEX_ARTICLES = frozenset({"selling", "marketing", "tax", "financing"})


def article_for(bdds_code: Any) -> tuple[str | None, str]:
    """Статья движка по коду БДДС. Возвращает статью и причину, если её нет.

    Самый длинный совпавший префикс: исключение из правила живёт рядом с
    правилом и побеждает его.
    """
    code = str(bdds_code or "").strip().rstrip(".")
    if not code:
        return None, "код БДДС не указан"
    for prefix, reason in sorted(_BDDS_UNRESOLVED.items(), key=lambda i: -len(i[0])):
        if code == prefix or code.startswith(prefix + "."):
            return None, reason
    if _gns_split_for(code):
        return None, (f"код {code} делится по ГНС между наземной и подземной "
                      "частями — нужна пропорция из ТЭП")
    best: tuple[int, str | None] = (-1, None)
    for prefix, article in _BDDS_TO_ARTICLE:
        if code == prefix or code.startswith(prefix + "."):
            if len(prefix) > best[0]:
                best = (len(prefix), article)
    if best[1] is None:
        return None, f"коду {code} не сопоставлена статья"
    return best[1], ""


def _gns_split_for(code: str) -> tuple[str, str] | None:
    """Пара статей, между которыми код делится по ГНС, если он такой."""
    best: tuple[int, tuple[str, str] | None] = (-1, None)
    for prefix, pair in _BDDS_GNS_SPLIT.items():
        if code == prefix or code.startswith(prefix + "."):
            if len(prefix) > best[0]:
                best = (len(prefix), pair)
    return best[1]


def _gns_shares(gns: dict[str, Any] | None) -> tuple[float, float] | None:
    """Доли наземной и подземной частей по ГНС. Нулевой итог — не пропорция."""
    if not gns:
        return None
    above = max(0.0, float(gns.get("above") or 0.0))
    under = max(0.0, float(gns.get("under") or 0.0))
    total = above + under
    if total <= 0:
        return None
    return above / total, under / total


def articles_from_register(
    register: dict[str, Any], gns: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Разнести фактические платежи реестра по статьям движка.

    Возвращает помесячные ряды по статьям — те же, что строит движок, — и
    отчёт: сколько денег статьи не нашло и почему. Нерасписанное не
    растворяется в «прочем», а называется суммой и причиной.
    """
    capex: dict[str, dict[datetime.date, float]] = {}
    other: dict[str, dict[datetime.date, float]] = {}
    unresolved: dict[str, float] = {}
    shares = _gns_shares(gns)

    def put(article: str, month: datetime.date, amount: float) -> None:
        target = other if article in _NON_CAPEX_ARTICLES else capex
        target.setdefault(article, {})
        target[article][month] = target[article].get(month, 0.0) + amount

    for item in register["rows"]:
        amount = item["paid_amount"]
        if not amount:
            continue
        code = str(item["bdds_code"] or "").strip().rstrip(".")
        month = item["paid_date"]
        if month is None:
            unresolved["платёж без даты"] = unresolved.get("платёж без даты", 0.0) + amount
            continue
        month = month.replace(day=1)

        split = _gns_split_for(code)
        if split and shares:
            above_article, under_article = split
            put(above_article, month, amount * shares[0])
            put(under_article, month, amount * shares[1])
            continue

        article, reason = article_for(code)
        if article is None:
            unresolved[reason] = unresolved.get(reason, 0.0) + amount
            continue
        put(article, month, amount)

    mapped = sum(sum(s.values()) for s in capex.values()) + \
        sum(sum(s.values()) for s in other.values())
    return {
        "capex_by_article": capex,
        "other_by_article": other,
        "unresolved": unresolved,
        "mapped": mapped,
        "total": register["paid"],
    }


# ---------------------------------------------------------------------------
# Реестр платежей РСС и сшивка его с реестром договоров.
#
# Оплаты есть в двух местах, и ни одно не годится в одиночку.
#
# Реестр платежей РСС полон: 2 251 платёж, 4 077,2 млн ₽, 32 месяца по июнь
# 2026, ни одной строки без даты, плюс источник платежа — свои или заёмные. Но
# он несёт только код ССР, а тот сваливает разное в один код: 2.6, 2.7, 1.8 и
# 2.2.1.10 держат 941,1 млн ₽, которые по нашим статьям однозначно не ложатся.
#
# Реестр договоров финансовой модели несёт код БДДС и раскладывается без
# остатка — но отстаёт на два месяца: 3 622,3 млн ₽ против 4 077,2. Ровно эта
# разница и была «расхождением источников», которое мы искали в методике.
#
# Поэтому: суммы и даты — из реестра платежей, код БДДС — подтягивается по
# договору из реестра договоров. По паре «контрагент + номер договора» сходится
# 97,2% денег. Остальное добирается по коду ССР там, где он однозначен, а что
# не добралось — уходит в отчёт с причиной.
#
# Соответствие «код ССР → статья» здесь не пишется руками: оно выводится из
# того же реестра договоров, где обе кодировки стоят в одной строке. Вторая
# рукописная карта разошлась бы с первой на первой же правке.
# ---------------------------------------------------------------------------

_CROSSWALK_SHEET = "статьи_БДДС"
_CROSSWALK_FIRST_ROW = 3
_CROSSWALK_COLUMNS = {"bdds": 6, "name": 7, "estimate_code": 8, "estimate_name": 9}


def read_article_crosswalk(path: str | Path) -> dict[str, Any]:
    """Лист «статьи_БДДС»: официальное соответствие кода БДДС коду РСС.

    Выводить это соответствие из реестра договоров нельзя: там «Код банк»
    местами разошёлся с листом. Фундаментная плита (БДДС 2.2.2.3.4) помечена в
    реестре кодом РСС 2.2.2.1, хотя по листу это 2.2.1.4 — и код 2.2.2.1
    начинает вести к двум разным статьям сразу, унося 284,6 млн ₽ в
    «неоднозначное». Лист — источник, реестр — данные.
    """
    rows: list[dict[str, Any]] = []
    for index, row in enumerate(_sheet(path, _CROSSWALK_SHEET), 1):
        if index < _CROSSWALK_FIRST_ROW:
            continue

        def cell(field: str) -> Any:
            position = _CROSSWALK_COLUMNS[field]
            return row[position] if position < len(row) else None

        bdds = str(cell("bdds") or "").strip()
        if not bdds:
            continue
        rows.append({
            "bdds_code": bdds,
            "name": str(cell("name") or "").strip(),
            "estimate_code": _code(cell("estimate_code")),
            "estimate_name": str(cell("estimate_name") or "").strip(),
        })
    articles: dict[str, set[str]] = {}
    for item in rows:
        if not item["estimate_code"]:
            continue
        article, _ = article_for(item["bdds_code"])
        if article:
            articles.setdefault(item["estimate_code"], set()).add(article)
    return {"rows": rows, "articles_by_estimate_code": articles}


_PAYMENTS_SHEET = "Реестр платежей"
_PAYMENTS_FIRST_ROW = 10
_PAYMENTS_COLUMNS = {
    "contractor": 1,
    "contract": 2,
    "contract_date": 3,
    "object": 4,
    "purpose": 5,
    "estimate_code": 6,
    "article_name": 7,
    "date": 8,
    "amount": 9,
    "source": 10,
    "counterparty": 11,
}

# Источник платежа: собственные средства процентов не несут (решение владельца
# 06.08.2026), заёмные — несут. Различие важно для финансовой части, поэтому
# оно читается, а не отбрасывается.
_OWN_FUNDS_MARKER = "собствен"


def _party_key(contractor: Any, contract: Any = None) -> str:
    """Ключ сшивки: контрагент и номер договора без пунктуации и регистра."""
    def clean(value: Any) -> str:
        return re.sub(r"[^0-9a-zа-яё]", "", str(value or "").lower().replace("ё", "е"))

    return clean(contractor) + ("|" + clean(contract) if contract is not None else "")


def read_payments(path: str | Path) -> dict[str, Any]:
    """Реестр платежей РСС: дата, сумма, код ССР, договор и источник средств."""
    rows: list[dict[str, Any]] = []
    for index, row in enumerate(_sheet(path, _PAYMENTS_SHEET), 1):
        if index < _PAYMENTS_FIRST_ROW:
            continue

        def cell(field: str) -> Any:
            position = _PAYMENTS_COLUMNS[field]
            return row[position] if position < len(row) else None

        amount = _money(cell("amount"))
        if not amount:
            continue
        source = str(cell("source") or "").strip()
        rows.append({
            "contractor": str(cell("contractor") or "").strip(),
            "contract": str(cell("contract") or "").strip(),
            "object": str(cell("object") or "").strip(),
            "purpose": str(cell("purpose") or "").strip(),
            "estimate_code": _code(cell("estimate_code")),
            "article_name": str(cell("article_name") or "").strip(),
            "date": _date(cell("date")),
            "amount": amount,
            "source": source,
            "own_funds": _OWN_FUNDS_MARKER in _normalized(source),
        })
    dated = [item for item in rows if item["date"]]
    return {
        "rows": rows,
        "total": sum(item["amount"] for item in rows),
        "undated": sum(item["amount"] for item in rows if not item["date"]),
        "own_funds": sum(item["amount"] for item in rows if item["own_funds"]),
        "first": min((item["date"] for item in dated), default=None),
        "last": max((item["date"] for item in dated), default=None),
    }


def payments_by_article(
    payments: dict[str, Any],
    register: dict[str, Any],
    gns: dict[str, Any] | None = None,
    crosswalk: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Разнести платежи по статьям движка, подтянув код БДДС по договору.

    Порядок попыток — от точного к приблизительному, и каждый шаг называется в
    отчёте: пара «контрагент + договор», затем контрагент с единственным кодом,
    затем код ССР там, где он даёт одну статью. Что не добралось — с причиной.
    """
    # Ключ сшивки — пара «контрагент + договор», но один договор нередко
    # покрывает несколько статей, и тогда пары мало. Сужаем кодом ССР самого
    # платежа: он в реестре платежей есть у каждой строки.
    by_pair: dict[tuple[str, str], set[str]] = {}
    by_pair_any: dict[str, set[str]] = {}
    by_party: dict[str, set[str]] = {}
    ssr_articles: dict[str, set[str]] = {}
    for item in register["rows"]:
        bdds = str(item["bdds_code"] or "").strip()
        if not bdds:
            continue
        pair = _party_key(item["counterparty"], item.get("contract"))
        by_pair.setdefault((pair, item["estimate_code"]), set()).add(bdds)
        by_pair_any.setdefault(pair, set()).add(bdds)
        by_party.setdefault(_party_key(item["counterparty"]), set()).add(bdds)
    # Соответствие «код РСС → статья» берётся с листа «статьи_БДДС», а из
    # реестра выводится только если листа не дали: в реестре «Код банк» местами
    # расходится с листом, и сборный по ошибке код уносит деньги в
    # «неоднозначное».
    if crosswalk:
        ssr_articles = {code: set(articles) for code, articles
                        in crosswalk["articles_by_estimate_code"].items()}
    else:
        for item in register["rows"]:
            bdds = str(item["bdds_code"] or "").strip()
            article, _ = article_for(bdds) if bdds else (None, "")
            if article and item["estimate_code"]:
                ssr_articles.setdefault(item["estimate_code"], set()).add(article)

    shares = _gns_shares(gns)
    capex: dict[str, dict[datetime.date, float]] = {}
    other: dict[str, dict[datetime.date, float]] = {}
    unresolved: dict[str, float] = {}
    matched: dict[str, float] = {}

    observed: dict[str, dict[str, float]] = {}
    current_code = {"value": ""}

    def put(article: str, month: datetime.date, amount: float) -> None:
        target = other if article in _NON_CAPEX_ARTICLES else capex
        target.setdefault(article, {})
        target[article][month] = target[article].get(month, 0.0) + amount
        code = current_code["value"]
        if code:
            observed.setdefault(code, {})
            observed[code][article] = observed[code].get(article, 0.0) + amount

    def place(code: str, month: datetime.date, amount: float, how: str) -> bool:
        split = _gns_split_for(code)
        if split:
            if not shares:
                return False
            put(split[0], month, amount * shares[0])
            put(split[1], month, amount * shares[1])
            matched[how] = matched.get(how, 0.0) + amount
            return True
        article, _ = article_for(code)
        if article is None:
            return False
        put(article, month, amount)
        matched[how] = matched.get(how, 0.0) + amount
        return True

    for item in payments["rows"]:
        amount, month = item["amount"], item["date"]
        if month is None:
            unresolved["платёж без даты"] = unresolved.get("платёж без даты", 0.0) + amount
            continue
        month = month.replace(day=1)
        current_code["value"] = item["estimate_code"]

        pair = _party_key(item["contractor"], item["contract"])
        attempts = (
            ("по договору и коду РСС", by_pair.get((pair, item["estimate_code"]), set())),
            ("по договору", by_pair_any.get(pair, set())),
            ("по контрагенту", by_party.get(_party_key(item["contractor"]), set())),
        )
        placed = False
        for how, codes in attempts:
            if len(codes) == 1 and place(next(iter(codes)), month, amount, how):
                placed = True
                break
        if placed:
            continue

        # Договор не опознан — остаётся код ССР, и он годится только там, где
        # ведёт к одной статье. Сборный код (2.6, 2.7, 1.8, 2.2.1.10) разносить
        # долями нельзя: это уже не факт.
        articles = ssr_articles.get(item["estimate_code"], set())
        if len(articles) == 1:
            put(next(iter(articles)), month, amount)
            matched["по коду РСС"] = matched.get("по коду РСС", 0.0) + amount
            continue
        reason = (f"код РСС {item['estimate_code'] or '—'} ведёт к нескольким статьям"
                  if articles else
                  f"договор не опознан, код РСС {item['estimate_code'] or '—'} неизвестен")
        unresolved[reason] = unresolved.get(reason, 0.0) + amount

    return {
        "capex_by_article": capex,
        "other_by_article": other,
        "unresolved": unresolved,
        "matched": matched,
        "mapped": sum(matched.values()),
        "total": payments["total"],
        # Как деньги одного кода РСС легли по нашим статьям на самом деле.
        # Этим потом делится смета того же кода: пропорция взята из факта, а не
        # выдумана.
        "observed_by_estimate_code": {
            code: dict(shares) for code, shares in observed.items()},
    }


# Реестр договоров РСС: по каждому договору законтрактовано, оплачено, авансы,
# остаток к оплате и выполнено. Это контрактный план-факт — уровень между
# сметой и платежом, которого не было ни в одном другом источнике.
_CONTRACTS_SHEET = "Реестр договоров"
_CONTRACTS_FIRST_ROW = 9
_CONTRACTS_COLUMNS = {
    "contractor": 1,
    "contract": 2,
    "subject": 4,
    "estimate_code": 5,
    "article_name": 6,
    "amount": 7,
    "paid": 8,
    "advances": 9,
    "outstanding": 10,
    "completed": 11,
}


def read_contracts(path: str | Path) -> dict[str, Any]:
    """Реестр договоров РСС: сколько заключено, оплачено и выполнено."""
    rows: list[dict[str, Any]] = []
    for index, row in enumerate(_sheet(path, _CONTRACTS_SHEET), 1):
        if index < _CONTRACTS_FIRST_ROW:
            continue

        def cell(field: str) -> Any:
            position = _CONTRACTS_COLUMNS[field]
            return row[position] if position < len(row) else None

        amount = _money(cell("amount"))
        contractor = str(cell("contractor") or "").strip()
        if not contractor and not amount:
            continue
        rows.append({
            "contractor": contractor,
            "contract": str(cell("contract") or "").strip(),
            "subject": str(cell("subject") or "").strip(),
            "estimate_code": _code(cell("estimate_code")),
            "article_name": str(cell("article_name") or "").strip(),
            "amount": amount,
            "paid": _money(cell("paid")),
            "advances": _money(cell("advances")),
            "outstanding": _money(cell("outstanding")),
            "completed": _money(cell("completed")),
        })
    return {
        "rows": rows,
        "amount": sum(item["amount"] for item in rows),
        "paid": sum(item["paid"] for item in rows),
        "advances": sum(item["advances"] for item in rows),
        "outstanding": sum(item["outstanding"] for item in rows),
        "completed": sum(item["completed"] for item in rows),
    }


def works_by_article(
    works: dict[str, Any],
    register: dict[str, Any],
    gns: dict[str, Any] | None = None,
    crosswalk: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Разнести принятые КС по статьям движка — тем же ключом, что и платежи.

    Деньги и объёмы идут по раздельным шкалам: аванс уходит раньше акта, а
    удержание позже. Но статья у них одна, и определяется она одинаково —
    договором, суженным кодом РСС. Строки, которые актом КС не являются (плата
    городу, комиссии банка, ФОТ), сюда не идут: они уже посчитаны деньгами.
    """
    rows = [{"contractor": item["contractor"], "contract": item.get("contract", ""),
             "estimate_code": item["code"], "amount": item["amount"],
             "date": item["date"]}
            for item in works["rows"] if item["construction"]]
    return payments_by_article(
        {"rows": rows, "total": sum(item["amount"] for item in rows)},
        register, gns=gns, crosswalk=crosswalk)


# Лист «План продаж» финансовой модели: помесячный объём и средняя цена, с
# меткой ФАКТ до среза и планом после. Цену модель выводит формулой из
# стартовой и темпа роста, а на действующем проекте она известна — договоры
# подписаны. Брать её расчётом, имея факт, значит спорить с реальностью.
_SALES_SHEET = "План продаж"
_SALES_FIRST_ROW = 22
_SALES_COLUMNS = {
    "period": 1,
    "mark": 2,
    "units": 5,
    "area": 6,
    "cumulative": 7,
    "price": 8,
}
_SALES_FACT_MARK = "факт"


def read_sales(path: str | Path) -> dict[str, Any]:
    """Помесячные продажи: объём, средняя цена и метка «факт»/«план»."""
    rows: list[dict[str, Any]] = []
    for index, row in enumerate(_sheet(path, _SALES_SHEET), 1):
        if index < _SALES_FIRST_ROW:
            continue

        def cell(field: str) -> Any:
            position = _SALES_COLUMNS[field]
            return row[position] if position < len(row) else None

        month = _as_month(cell("period"))
        if month is None:
            continue
        area = _money(cell("area"))
        price = _money(cell("price"))
        rows.append({
            "month": month,
            "fact": _SALES_FACT_MARK in _normalized(cell("mark")),
            "units": _money(cell("units")),
            "area": area,
            "price": price,
            "revenue": area * price,
        })
    fact = [item for item in rows if item["fact"]]
    sold = sum(item["area"] for item in fact)
    revenue = sum(item["revenue"] for item in fact)
    return {
        "rows": rows,
        "fact_area": sold,
        "fact_units": sum(item["units"] for item in fact),
        "fact_revenue": revenue,
        "average_price": revenue / sold if sold else 0.0,
        "last_fact": max((item["month"] for item in fact), default=None),
    }


# Коды РСС, которых нет в перекодировке БДДС, потому что по ним не платят
# подрядчикам: резервы существуют в смете и исчезают при заключении договоров.
# В карту БДДС им попасть неоткуда, а в смете они есть и в бюджет обязаны войти.
_ESTIMATE_ONLY_CODES: dict[str, str] = {
    "2.8": "reserve",   # резерв на непредвиденные расходы
    "2.9": "reserve",   # резерв на инфляционное удорожание
}


def budget_by_article(
    estimate: dict[str, Any],
    observed: dict[str, dict[str, float]] | None = None,
    crosswalk: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Смета РСС, разложенная по статьям движка.

    На действующем проекте расходную часть считать незачем — она посчитана и
    утверждена. Смета приходит по кодам РСС, а код сваливает разное в одну
    строку, поэтому доли берутся из уже сшитого факта по тому же коду: как
    деньги этого кода легли на самом деле, так делится и его смета.

    Где факта ещё нет, годится перекодировка — но только если код ведёт к одной
    статье. Что не разложилось, называется суммой и причиной: смета, размазанная
    наугад, на экране неотличима от разложенной верно.

    Суммируются только листья дерева: строки-главы РСС уже агрегаты, и сложить
    их с подстроками значит посчитать смету дважды.
    """
    observed = observed or {}
    by_code_articles = (crosswalk or {}).get("articles_by_estimate_code") or {}
    budget: dict[str, float] = {}
    unresolved: dict[str, float] = {}
    for row in estimate["rows"]:
        if not row.get("is_leaf"):
            continue
        amount = float(row.get("estimate") or 0.0)
        if amount <= 0:
            continue
        code = row["code"]
        shares = observed.get(code) or {}
        total = sum(shares.values())
        if total > 0:
            for article, value in shares.items():
                budget[article] = budget.get(article, 0.0) + amount * value / total
            continue
        if code in _ESTIMATE_ONLY_CODES:
            article = _ESTIMATE_ONLY_CODES[code]
            budget[article] = budget.get(article, 0.0) + amount
            continue
        articles = by_code_articles.get(code) or set()
        if len(articles) == 1:
            article = next(iter(articles))
            budget[article] = budget.get(article, 0.0) + amount
            continue
        reason = (f"код РСС {code} ведёт к нескольким статьям, а факта по нему нет"
                  if articles else f"коду РСС {code} не сопоставлена статья")
        unresolved[reason] = unresolved.get(reason, 0.0) + amount
    return {
        "budget": budget,
        "unresolved": unresolved,
        "mapped": sum(budget.values()),
        "total": sum(float(r.get("estimate") or 0.0)
                     for r in estimate["rows"] if r.get("is_leaf")),
    }


# ---------------------------------------------------------------------------
# Производственная программа РСС и ответ «отстаём или нет».
#
# В РСС от 08.07.2026 справа от сметы стоит шахматка: по каждому коду сколько
# работ планируется принять в июле, августе, сентябре и так далее. Это и есть
# график, с которым сравнивается выполнение.
#
# Год в шапке не написан — только «июль», «август». Поэтому первый месяц
# программы задаётся вызовом: вывести его из имени файла значило бы гадать, а
# ошибка на год не заметна ни в одной сумме.
# ---------------------------------------------------------------------------

_PROGRAMME_MONTHS = (
    "январь", "февраль", "март", "апрель", "май", "июнь",
    "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь",
)
_PROGRAMME_HEADER_ROW = 9
_PROGRAMME_CODE_COLUMN = 0


def _add_month(month: datetime.date, shift: int) -> datetime.date:
    total = month.year * 12 + (month.month - 1) + shift
    return datetime.date(total // 12, total % 12 + 1, 1)


def read_programme(path: str | Path, start: Any) -> dict[str, Any]:
    """Производственная программа РСС: план приёмки работ по кодам и месяцам.

    `start` — первый месяц программы. Он приходит снаружи, потому что в шапке
    стоит «июль» без года, и додумывать год нельзя: ошибка на двенадцать
    месяцев не видна ни в одной сумме.
    """
    first = _as_month(start)
    if first is None:
        raise ValueError("не задан первый месяц программы (`start`)")
    columns: list[tuple[int, datetime.date]] = []
    by_code: dict[str, dict[datetime.date, float]] = {}
    order: list[dict[str, Any]] = []
    for index, row in enumerate(_sheet(path, _ESTIMATE_SHEET), 1):
        if index == _PROGRAMME_HEADER_ROW:
            names = [(position, _normalized(value))
                     for position, value in enumerate(row)]
            found = [position for position, name in names
                     if name in _PROGRAMME_MONTHS]
            columns = [(position, _add_month(first, offset))
                       for offset, position in enumerate(sorted(found))]
            continue
        if not columns:
            continue
        code = _code(row[_PROGRAMME_CODE_COLUMN]
                     if _PROGRAMME_CODE_COLUMN < len(row) else None)
        if not code:
            continue
        series = {}
        for position, month in columns:
            amount = _money(row[position] if position < len(row) else None)
            if amount:
                series[month] = amount
        order.append({"code": code, "depth": code.count(".") + 1})
        if series:
            by_code[code] = series
    # Лист несёт и главы, и их подстроки. Сложить всё подряд — посчитать
    # программу дважды: код «2» уже содержит 2.2 и 2.3. Дерево строится тем же
    # правилом, что в смете, — по порядку строк и глубине кода.
    _link_parents(order)
    leaves = {row["code"] for row in order if row["is_leaf"]}
    months = [month for _, month in columns]
    return {
        "by_code": by_code,
        "leaves": leaves,
        "months": months,
        "first": months[0] if months else None,
        "last": months[-1] if months else None,
        "total": sum(sum(series.values()) for code, series in by_code.items()
                     if code in leaves),
    }


def monitor(
    estimate: dict[str, Any],
    payments: dict[str, Any],
    works: dict[str, Any],
    contracts: dict[str, Any],
    cut: Any,
    programme: dict[str, Any] | None = None,
    sales: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Недельный срез: где проект по деньгам, по объёмам и по графику.

    Считается из одного набора выгрузок, поэтому обновляется их заменой — в
    этом весь смысл: раз в неделю кладутся свежие РСС и книга, и картина
    пересобирается сама, без ручной сверки.

    Отставание меряется по производственной программе: сколько работ должно
    было быть принято к срезу и сколько принято на самом деле. Программа
    начинается там, где кончается факт, поэтому в первую неделю сравнивать
    нечего — и это говорится прямо, а не показывается нулём.
    """
    cut_month = _as_month(cut)
    if cut_month is None:
        raise ValueError("не задана дата среза (`cut`)")

    paid = sum(item["amount"] for item in payments["rows"]
               if item["date"] and item["date"] < cut_month)
    accepted = sum(item["amount"] for item in works["rows"]
                   if item["construction"] and item["date"]
                   and item["date"] < cut_month)
    budget = float((estimate.get("total") or {}).get("estimate") or 0.0)

    money = {
        "budget": budget,
        "contracted": contracts["amount"],
        "paid": paid,
        "advances": contracts["advances"],
        "outstanding": contracts["outstanding"],
        "accepted": accepted,
        # Оплачено больше принятого — это авансы, выданные вперёд. Деньги из
        # кассы вышли и проценты по ним идут, а работы ещё не приняты.
        "paid_ahead": max(0.0, paid - accepted),
        "left_to_budget": max(0.0, budget - paid),
    }

    schedule: dict[str, Any] = {"comparable": False}
    if programme and programme.get("months"):
        due_months = [m for m in programme["months"] if m < cut_month]
        leaves = programme.get("leaves") or set(programme["by_code"])
        due = sum(series.get(month, 0.0)
                  for code, series in programme["by_code"].items()
                  if code in leaves for month in due_months)
        done = sum(item["amount"] for item in works["rows"]
                   if item["construction"] and item["date"]
                   and programme["months"][0] <= item["date"] < cut_month)
        if due_months:
            schedule = {
                "comparable": True,
                "from": programme["months"][0],
                "months_due": len(due_months),
                "due": due,
                "done": done,
                "gap": done - due,
                "ratio": done / due if due else 0.0,
            }
        else:
            schedule["reason"] = (
                f"программа начинается {programme['months'][0]:%Y-%m}, "
                f"а срез стоит на {cut_month:%Y-%m} — сравнивать ещё нечего")

    sold: dict[str, Any] = {}
    if sales:
        fact = [row for row in sales["rows"]
                if row["fact"] and row["month"] < cut_month]
        plan = [row for row in sales["rows"]
                if not row["fact"] and row["month"] < cut_month]
        sold = {
            "area": sum(row["area"] for row in fact),
            "units": sum(row["units"] for row in fact),
            "revenue": sum(row["revenue"] for row in fact),
            "average_price": (sum(row["revenue"] for row in fact)
                              / sum(row["area"] for row in fact)
                              if sum(row["area"] for row in fact) else 0.0),
            "last_fact": max((row["month"] for row in fact), default=None),
            # Месяцы до среза, где факта нет, а план есть: выгрузка продаж
            # отстаёт от среза, и это не «не продавали».
            "months_without_fact": len(plan),
        }

    return {"cut": cut_month, "money": money, "schedule": schedule, "sales": sold}


# ---------------------------------------------------------------------------
# Фактическое финансирование из листа «КРЕДИТЫ» книги.
#
# До среза всё берётся фактом — это правило, а финансирование было последним,
# что ему не подчинялось: проценты и остатки движок выводил из долга, ставки и
# покрытия, хотя они уже случились и заплачены. Считать их заново значит
# спорить с банковской выпиской.
#
# Остатки идут помесячно из книги, уплаченные проценты — из реестра платежей по
# кодам БДДС 3.1.4.x. Два источника здесь не спорят: книга даёт остаток, реестр
# — движение денег, и это разные величины.
# ---------------------------------------------------------------------------

_CREDITS_SHEET = "КРЕДИТЫ"
_CREDITS_HEADER_ROW = 3
_CREDITS_ROWS = {
    "escrow": 54,          # Счета эскроу (нараст)
    "pf_drawn": 55,        # Проектное финансирование (нараст)
    "pf_draw": 61,         # Получение кредита ПФ (ОД)
    "pf_repayment": 62,    # Погашение кредита ПФ (ОД)
    "bridge_draw": 33,     # Получение кредита БРИДЖ (ОД)
    "bridge_repayment": 36,
    # Начисленные проценты, а не уплаченные. До раскрытия эскроу проценты по ПФ
    # капитализируются: кассой уходит малая часть, и подставить «выплату
    # процентов» вместо начисления значит занизить стоимость денег в разы. На
    # Гродненской это 174,8 млн уплаченных против начисленных за тот же срок.
    "pf_interest_accrued": 67,
    "pf_interest_capitalized": 68,
    "pf_limit_fee": 69,
    "bridge_interest_accrued": 42,
}


def read_credits(path: str | Path) -> dict[str, Any]:
    """Помесячные ряды листа «КРЕДИТЫ»: эскроу, выборка и погашение долга."""
    months: list[tuple[int, datetime.date]] = []
    series: dict[str, dict[datetime.date, float]] = {name: {} for name in _CREDITS_ROWS}
    wanted = {row: name for name, row in _CREDITS_ROWS.items()}
    for index, row in enumerate(_sheet(path, _CREDITS_SHEET), 1):
        if index == _CREDITS_HEADER_ROW:
            months = [(position, value.date().replace(day=1))
                      for position, value in enumerate(row)
                      if isinstance(value, datetime.datetime)]
            continue
        name = wanted.get(index)
        if not name or not months:
            continue
        for position, month in months:
            amount = _money(row[position] if position < len(row) else None)
            if amount:
                series[name][month] = amount
    # Остаток долга — выборка минус погашение нарастающим итогом. В книге
    # строка 55 накопительная, но погашение в неё не заходит, и брать её за
    # остаток нельзя: после раскрытия эскроу долг падает, а строка нет.
    balance: dict[datetime.date, float] = {}
    running = 0.0
    for month in sorted({*series["pf_draw"], *series["pf_repayment"]}):
        running += series["pf_draw"].get(month, 0.0)
        running += series["pf_repayment"].get(month, 0.0)   # погашение отрицательное
        balance[month] = max(0.0, running)
    bridge: dict[datetime.date, float] = {}
    running = 0.0
    for month in sorted({*series["bridge_draw"], *series["bridge_repayment"]}):
        running += series["bridge_draw"].get(month, 0.0)
        running += series["bridge_repayment"].get(month, 0.0)
        bridge[month] = max(0.0, running)
    return {
        **series,
        "pf_balance": balance,
        "bridge_balance": bridge,
        "months": [month for _, month in months],
    }
