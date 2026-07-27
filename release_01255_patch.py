from __future__ import annotations

import copy
import io
from datetime import datetime
from typing import Any

VERSION = "0.12.55"

_EXPERT_NOTE = (
    "Оценочная дата — за месяц до РнС; на этапе инвестиционного анализа "
    "точная дата соглашения обычно неизвестна, после появления утверждённых "
    "документов и графика её необходимо заменить на фактическую"
)


def _patch_vri_date(core: Any) -> None:
    core.DEFAULT_INPUTS["vri_obligation_date_mode"] = "before_rns_1m"
    if hasattr(core, "VRI_DEFAULTS"):
        core.VRI_DEFAULTS["vri_obligation_date_mode"] = "before_rns_1m"

    def vri_obligation_date(x: dict[str, Any], permit: Any) -> tuple[Any, str, bool]:
        mode = str(x.get("vri_obligation_date_mode") or "before_rns_1m").strip().lower()
        raw = str(x.get("vri_obligation_date") or "").strip()
        if raw:
            try:
                return core.d(raw), "Дата известна — введена вручную", False
            except Exception:
                pass
        if mode == "at_rns":
            return permit, "Оценочная дата — в дату РнС", True
        if mode == "before_rns_3m":
            return core.add_months(permit, -3), "Оценочная дата — за три месяца до РнС", True
        if mode == "after_purchase":
            months = max(0, int(core.n(x, "vri_months_after_purchase", 12)))
            start = core.d(x.get("project_start", "2027-01-01"))
            return core.add_months(start, months), f"Оценочная дата — через {months} мес. после покупки", True
        if mode == "manual":
            return core.add_months(permit, -1), (
                "Точная дата не задана — на этапе инвестиционного анализа принята "
                "экспертная оценка за месяц до РнС; после появления документов дату нужно заменить"
            ), True
        return core.add_months(permit, -1), _EXPERT_NOTE, True

    core.vri_obligation_date = vri_obligation_date

    def split_payment(amount: float, before_pf: bool, settings: dict[str, Any]) -> tuple[float, float, float]:
        def fallback() -> tuple[float, float, float]:
            if not settings["in_bank_budget"]:
                return 0.0, 0.0, amount
            if before_pf:
                return amount, 0.0, 0.0
            return 0.0, amount, 0.0

        if settings["financing_mode"] != "shares":
            return fallback()
        total = settings["share_bridge"] + settings["share_pf"] + settings["share_equity"]
        if total <= 0:
            return fallback()
        bridge = amount * settings["share_bridge"] / total
        pf = amount * settings["share_pf"] / total
        equity = amount * settings["share_equity"] / total
        if pf and (before_pf or not settings["in_bank_budget"]):
            if before_pf and settings["in_bank_budget"]:
                bridge += pf
            else:
                equity += pf
            pf = 0.0
        if not settings["in_bank_budget"] and bridge:
            equity += bridge
            bridge = 0.0
        return bridge, pf, equity

    core._vri_split_payment = split_payment

    for group_name, fields in core.FIELD_GROUPS:
        if group_name != "Смена ВРИ и земельные права":
            continue
        for field in fields:
            if field[0] == "vri_obligation_date_mode":
                field[4] = [
                    ["before_rns_1m", "За месяц до РнС — экспертная оценка"],
                    ["at_rns", "В дату РнС"],
                    ["before_rns_3m", "За три месяца до РнС"],
                    ["after_purchase", "Через N мес. после покупки"],
                    ["manual", "Задана вручную"],
                ]
            elif field[0] == "vri_obligation_date":
                field[2] = "точная дата по документу; пусто — экспертная оценка"


def _land_text(value: Any) -> str:
    return str(value or "").strip()


def _plato_land_parcel(inputs: dict[str, Any]) -> dict[str, Any]:
    lookup = (inputs.get("_land_lookup") or {}).get("results") or []
    parcel = next((item for item in lookup if item.get("kind") == "land"), None)
    if parcel:
        return parcel
    parcels = ((inputs.get("_mo_calc") or {}).get("vri") or {}).get("parcels") or []
    return parcels[0] if parcels else {}


def _fill_land_sheet(core: Any, workbook: Any, inputs: dict[str, Any], filled: list[dict[str, Any]]) -> None:
    if "ЗУ" not in workbook.sheetnames:
        return
    sheet = workbook["ЗУ"]

    def put(row: int, value: Any, label: str) -> None:
        sheet.cell(row=row, column=3).value = value
        filled.append({"sheet": "ЗУ", "row": row, "label": label, "value": value})

    blocks = ((22, 31), (40, 49))
    land_rows = (
        "Кад.№", "Адрес", "Площадь, кв.м", "ВРИ", "УПКС жилье",
        "Кадастровая стоимость", "План. кад.стоимость, руб.",
        "Ставка арендной платы, %", "Смена ВРИ, руб.", "Дата первого платежа:",
    )
    parcel = _plato_land_parcel(inputs)
    first_block = True
    for start, end in blocks:
        labels = {
            core._plato_normalize(sheet.cell(row=row, column=2).value): row
            for row in range(start, end + 1)
        }
        for name in land_rows:
            row = labels.get(core._plato_normalize(name))
            if row:
                sheet.cell(row=row, column=3).value = None
        if not first_block:
            continue
        first_block = False
        for name, value in (
            ("Кад.№", _land_text(parcel.get("cadastral_number"))),
            ("Адрес", _land_text(parcel.get("address"))),
            ("Площадь, кв.м", core._land_float(parcel.get("area_sqm"))),
            ("ВРИ", _land_text(parcel.get("permitted_use"))),
            ("Кадастровая стоимость", core._land_float(parcel.get("cadastral_value_rub"))),
        ):
            row = labels.get(core._plato_normalize(name))
            if row and value not in (None, ""):
                put(row, value, name)

    permit = core.add_months(core.d(inputs.get("project_start", "2027-01-01")), int(core.n(inputs, "ird_months", 18)))
    gross = core.n(inputs, "land_rights_cost_mln") * 1_000_000
    _relief, net = core.vri_relief(inputs, gross)
    schedule = core.build_vri_schedule(inputs, net, permit)
    labels = {
        core._plato_normalize(sheet.cell(row=row, column=2).value): row
        for row in range(56, 70)
    }
    share = (net / gross) if gross else 1.0
    row = labels.get(core._plato_normalize("Доля оплаты"))
    if row:
        put(row, round(share, 6), "Доля оплаты по ВРИ")
    row = labels.get(core._plato_normalize("%% за рассрочку"))
    if row:
        put(row, round(core.n(inputs, "vri_interest_spread_pp", 3.0) / 100.0, 6), "Спред по рассрочке ВРИ")

    rows = schedule.get("rows") or []
    if not rows:
        return
    first = core.d(rows[0]["date"])
    last = core.d(rows[-1]["date"])
    row = labels.get(core._plato_normalize("Первый"))
    if row:
        put(row, datetime(first.year, first.month, first.day), "Первый платёж ВРИ")
    row = labels.get(core._plato_normalize("Последний"))
    if row:
        put(row, datetime(last.year, last.month, last.day), "Последний платёж ВРИ")
    row = labels.get(core._plato_normalize("В месяц"))
    if row:
        months = max(1, core.months_between(first, last))
        sheet.cell(row=row, column=3).value = f"=C{row - 4}/{months}"
        filled.append({"sheet": "ЗУ", "row": row, "label": "Ежемесячный платёж ВРИ", "value": f"1/{months}"})


def _patch_plato_export(core: Any) -> None:
    if hasattr(core, "_plato_fill_land_sheet"):
        return
    original = core.fill_plato_template

    def fill_plato_template(inputs: dict[str, Any], tep: dict[str, dict[str, Any]], **kwargs: Any):
        content, report = original(inputs, tep, **kwargs)
        from openpyxl import load_workbook

        merged = {**copy.deepcopy(core.DEFAULT_INPUTS), **(inputs or {})}
        workbook = load_workbook(io.BytesIO(content), data_only=False, keep_vba=False)
        filled = report.setdefault("filled", [])
        _fill_land_sheet(core, workbook, merged, filled)
        workbook.calculation.fullCalcOnLoad = True
        buffer = io.BytesIO()
        workbook.save(buffer)
        report["filled_count"] = len(filled)
        return buffer.getvalue(), report

    fill_plato_template._developaid_01255 = True
    core.fill_plato_template = fill_plato_template


def _patch_page(core: Any) -> None:
    page = str(getattr(core, "PAGE", ""))
    if not page:
        return
    page = page.replace("0.12.52", VERSION).replace("0.12.53", VERSION)
    page = page.replace('"vri_obligation_date_mode": "at_rns"', '"vri_obligation_date_mode": "before_rns_1m"')
    page = page.replace(
        '[["at_rns", "В дату РнС"], ["before_rns_1m", "За месяц до РнС"], ["before_rns_3m", "За три месяца до РнС"], ["after_purchase", "Через N мес. после покупки"], ["manual", "Задана вручную"]]',
        '[["before_rns_1m", "За месяц до РнС — экспертная оценка"], ["at_rns", "В дату РнС"], ["before_rns_3m", "За три месяца до РнС"], ["after_purchase", "Через N мес. после покупки"], ["manual", "Задана вручную"]]',
    )
    page = page.replace(
        '"Дата возникновения обязательства", "дата (пусто — РнС)"',
        '"Дата возникновения обязательства", "точная дата по документу; пусто — экспертная оценка"',
    )
    old = (
        '<div class="note">Платежи до открытия ПФ несёт БРИДЖ или собственный капитал, '
        'после — ПФ, и только если ВРИ включена в банковский бюджет. Проценты по '
        'рассрочке считаются отдельно от процентов по кредитам.</div>'
    )
    new = (
        '<div class="note"><b>Дата обязательства по умолчанию — экспертно за 1 месяц до РнС.</b> '
        'На этапе инвестиционного анализа точная дата соглашения обычно неизвестна; после появления '
        'утверждённых документов и графика её необходимо заменить на фактическую. Платежи до открытия '
        'ПФ несёт БРИДЖ или собственный капитал, после — ПФ, и только если ВРИ включена в банковский '
        'бюджет. Проценты по рассрочке считаются отдельно от процентов по кредитам.</div>'
    )
    page = page.replace(old, new)
    core.PAGE = page


def apply(runtime: Any) -> None:
    core = runtime.core
    _patch_vri_date(core)
    _patch_plato_export(core)
    _patch_page(core)
    runtime._RUNTIME_VERSION = VERSION
    runtime.app.version = VERSION
