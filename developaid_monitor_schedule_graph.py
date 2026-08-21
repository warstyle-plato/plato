"""Dependency/float forecast layer for DevelopAid Project Monitor.

The PM workbook is a fixed source of relationships only. Physical progress is
never read from PM: it remains RSS accepted acts. This module propagates delays
through the actual PM dependency graph and preserves the original schedule while
allowing approved article rebaselines to be stored separately.
"""
from __future__ import annotations

import datetime
import json
import re
from pathlib import Path
from typing import Any

import developaid_monitor as monitor

_RU_MONTHS = {
    "январь": 1, "февраль": 2, "март": 3, "апрель": 4,
    "май": 5, "июнь": 6, "июль": 7, "август": 8,
    "сентябрь": 9, "октябрь": 10, "ноябрь": 11, "декабрь": 12,
}


def _norm(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower().replace("ё", "е"))


def _date(value: Any) -> datetime.date | None:
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    text = str(value or "").strip().lower().replace("ё", "е")
    if not text or text in {"нд", "na", "n/a"}:
        return None
    m = re.search(r"(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{2,4})", text)
    if m:
        d, mo, y = map(int, m.groups())
        if y < 100:
            y += 2000
        try:
            return datetime.date(y, mo, d)
        except ValueError:
            return None
    m = re.search(r"(\d{1,2})\s+([а-я]+)\s+(\d{4})", text)
    if m:
        d, month, y = m.groups()
        mo = _RU_MONTHS.get(month)
        if mo:
            try:
                return datetime.date(int(y), mo, int(d))
            except ValueError:
                return None
    m = re.search(r"(\d{1,2})\s+([а-я]+)\s+(\d{2})\b", text)
    if m:
        d, month, y = m.groups()
        mo = _RU_MONTHS.get(month)
        if mo:
            try:
                return datetime.date(2000 + int(y), mo, int(d))
            except ValueError:
                return None
    return None


def _days(value: Any) -> int:
    m = re.search(r"-?\d+(?:[.,]\d+)?", str(value or ""))
    return int(round(float(m.group(0).replace(",", ".")))) if m else 0


def _parse_predecessors(value: Any) -> list[dict[str, Any]]:
    text = str(value or "").strip()
    if not text:
        return []
    out: list[dict[str, Any]] = []
    for token in re.split(r"[;,]", text):
        token = token.strip().replace(" ", "")
        if not token:
            continue
        m = re.match(r"^(\d+)(ОН|НН|ОО|НО)?([+-]\d+(?:[.,]\d+)?)?[а-яА-Я]*$", token)
        if not m:
            continue
        pid, ru_type, lag = m.groups()
        kind = {None: "FS", "ОН": "FS", "НН": "SS", "ОО": "FF", "НО": "SF"}[ru_type]
        out.append({"id": pid, "type": kind, "lag_days": int(round(float((lag or "0").replace(",", "."))))})
    return out


def store_reference(project: str, data: bytes, sheet: str, start: Any,
                    code: str, taken_at: Any) -> dict[str, Any]:
    day = monitor._iso(taken_at)
    base = monitor._project_dir(project) / "baseline"
    base.mkdir(parents=True, exist_ok=True)
    if code == "__PM__":
        path = base / "pm.xlsx"
        path.write_bytes(data)
        return {"taken_at": day, "code": code, "stored": "pm", "bytes": len(data)}
    folder = base / "rebaseline"
    folder.mkdir(parents=True, exist_ok=True)
    slug = re.sub(r"[^0-9A-Za-zА-Яа-я_-]+", "-", str(code)).strip("-") or "article"
    path = folder / f"{day}.{slug}.xlsx"
    path.write_bytes(data)
    (folder / f"{day}.{slug}.json").write_text(json.dumps({
        "taken_at": day, "code": code, "sheet": sheet, "start": monitor._iso(start),
    }, ensure_ascii=False), encoding="utf-8")
    return {"taken_at": day, "code": code, "stored": "rebaseline", "bytes": len(data)}


def _load_pm(project: str) -> dict[str, Any]:
    path = monitor._project_dir(project) / "baseline" / "pm.xlsx"
    if not path.exists():
        return {"known": False, "tasks": {}, "rnv_id": "", "source": ""}
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["Таблица_задач1"] if "Таблица_задач1" in wb.sheetnames else wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header_values = next(rows, ())
        header = {_norm(v).replace(" ", "_"): i for i, v in enumerate(header_values)}
        aliases = {
            "id": ("ид", "id"), "name": ("название_задачи", "наименование_работ"),
            "pred": ("предшественники",), "start": ("начало",), "finish": ("окончание",),
            "free": ("свободный_временной_резерв",), "total": ("общий_временной_резерв",),
        }
        idx: dict[str, int] = {}
        for key, names in aliases.items():
            for name in names:
                if name in header:
                    idx[key] = header[name]
                    break
        tasks: dict[str, dict[str, Any]] = {}
        rnv_id = ""
        for values in rows:
            def cell(key: str) -> Any:
                i = idx.get(key, -1)
                return values[i] if 0 <= i < len(values) else None
            tid = str(cell("id") or "").strip()
            name = str(cell("name") or "").strip()
            start = _date(cell("start")); finish = _date(cell("finish"))
            if not tid or not name or not start or not finish:
                continue
            task = {
                "id": tid, "name": name, "start": start, "finish": finish,
                "duration_days": max(0, (finish - start).days),
                "predecessors": _parse_predecessors(cell("pred")),
                "free_float_days": _days(cell("free")),
                "total_float_days": _days(cell("total")),
            }
            tasks[tid] = task
            low = _norm(name)
            if (("рнв" in low and "получ" in low) or
                ("разрешен" in low and "ввод" in low and ("получ" in low or "выдан" in low))):
                rnv_id = tid
        return {"known": bool(tasks), "tasks": tasks, "rnv_id": rnv_id, "source": path.name}
    finally:
        wb.close()


def _rebaselines(project: str) -> dict[str, dict[str, Any]]:
    folder = monitor._project_dir(project) / "baseline" / "rebaseline"
    if not folder.exists():
        return {}
    from openpyxl import load_workbook
    result: dict[str, dict[str, Any]] = {}
    for meta_path in sorted(folder.glob("*.json")):
        try:
            meta = json.loads(meta_path.read_text(encoding="utf-8"))
        except Exception:
            continue
        code = str(meta.get("code") or "").rstrip(".")
        xlsx = meta_path.with_suffix(".xlsx")
        if not code or not xlsx.exists():
            continue
        finish: datetime.date | None = None
        wb = load_workbook(xlsx, read_only=True, data_only=True)
        try:
            sheet = str(meta.get("sheet") or "")
            ws = wb[sheet] if sheet in wb.sheetnames else wb[wb.sheetnames[0]]
            for row in ws.iter_rows(values_only=True):
                for value in row:
                    if isinstance(value, str) and "заверш" in _norm(value):
                        d = _date(value)
                        if d and (finish is None or d > finish):
                            finish = d
        finally:
            wb.close()
        result[code] = {**meta, "finish": finish, "source": xlsx.name}
    return result


def _constraint_start(edge: dict[str, Any], pred: dict[str, Any], succ: dict[str, Any]) -> datetime.date:
    lag = datetime.timedelta(days=edge["lag_days"])
    duration = datetime.timedelta(days=succ["duration_days"])
    kind = edge["type"]
    if kind == "SS":
        return pred["forecast_start"] + lag
    if kind == "FF":
        return pred["forecast_finish"] + lag - duration
    if kind == "SF":
        return pred["forecast_start"] + lag - duration
    return pred["forecast_finish"] + lag


def _propagate(pm: dict[str, Any], seeds: dict[str, datetime.date]) -> dict[str, dict[str, Any]]:
    tasks = {tid: {**item, "forecast_start": item["start"], "forecast_finish": item["finish"],
                   "own_delay_days": 0, "inherited_delay_days": 0}
             for tid, item in pm["tasks"].items()}
    successors: dict[str, list[dict[str, Any]]] = {}
    indegree = {tid: 0 for tid in tasks}
    for sid, succ in tasks.items():
        valid = []
        for edge in succ["predecessors"]:
            if edge["id"] in tasks:
                valid.append(edge)
                successors.setdefault(edge["id"], []).append({**edge, "successor_id": sid})
                indegree[sid] += 1
        succ["predecessors"] = valid
    for tid, finish in seeds.items():
        task = tasks.get(tid)
        if not task or not finish or finish <= task["finish"]:
            continue
        shift = (finish - task["finish"]).days
        task["forecast_start"] = task["start"] + datetime.timedelta(days=shift)
        task["forecast_finish"] = finish
        task["own_delay_days"] = shift
    queue = [tid for tid, deg in indegree.items() if deg == 0]
    order: list[str] = []
    while queue:
        tid = queue.pop()
        order.append(tid)
        for edge in successors.get(tid, []):
            sid = edge["successor_id"]
            indegree[sid] -= 1
            if indegree[sid] == 0:
                queue.append(sid)
    if len(order) < len(tasks):
        ordered = set(order)
        order.extend(tid for tid in tasks if tid not in ordered)
    for _ in range(2):
        for sid in order:
            succ = tasks[sid]
            required = succ["forecast_start"]
            for edge in succ["predecessors"]:
                pred = tasks[edge["id"]]
                required = max(required, _constraint_start(edge, pred, succ))
            if required > succ["forecast_start"]:
                delta = (required - succ["forecast_start"]).days
                succ["forecast_start"] += datetime.timedelta(days=delta)
                succ["forecast_finish"] += datetime.timedelta(days=delta)
                succ["inherited_delay_days"] += delta
    for tid, task in tasks.items():
        task["successors"] = successors.get(tid, [])
        shift = max(0, (task["forecast_finish"] - task["finish"]).days)
        task["current_float_days"] = task["total_float_days"] - shift
    return tasks


def _walk_nodes(nodes: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for node in nodes:
        out.append(node)
        children = [c for c in node.get("children", []) if isinstance(c, dict)]
        out.extend(_walk_nodes(children))
    return out


def apply(project: str, view: dict[str, Any]) -> dict[str, Any]:
    schedule = view.get("schedule") or {}
    pm = _load_pm(project)
    if not pm["known"]:
        schedule["dependency_graph"] = {"known": False, "reason": "не загружен сырой PM/ГРП"}
        return view
    seeds: dict[str, datetime.date] = {}
    for row in schedule.get("rows") or []:
        tid = str(row.get("id") or "").strip()
        forecast = monitor._day(row.get("forecast_finish"))
        if tid in pm["tasks"] and forecast:
            seeds[tid] = forecast
    tasks = _propagate(pm, seeds)
    rnv = tasks.get(pm["rnv_id"]) if pm["rnv_id"] else None
    baseline_rnv = rnv["finish"] if rnv else max((t["finish"] for t in tasks.values()), default=None)
    forecast_rnv = rnv["forecast_finish"] if rnv else max((t["forecast_finish"] for t in tasks.values()), default=None)
    project_delay = max(0, (forecast_rnv - baseline_rnv).days) if baseline_rnv and forecast_rnv else 0
    ancestors: set[str] = set()
    if pm["rnv_id"] in tasks:
        stack = [pm["rnv_id"]]
        while stack:
            tid = stack.pop()
            if tid in ancestors:
                continue
            ancestors.add(tid)
            stack.extend(edge["id"] for edge in tasks[tid]["predecessors"])
    flat = {str(row.get("id") or "").strip(): row for row in schedule.get("rows") or []}
    for tid, row in flat.items():
        task = tasks.get(tid)
        if not task:
            continue
        propagated = task["forecast_finish"]
        own = monitor._day(row.get("forecast_finish"))
        if propagated and (own is None or propagated > own):
            row["forecast_finish"] = monitor._iso(propagated)
            finish = monitor._day(row.get("plan_finish"))
            row["delta_days"] = (propagated - finish).days if finish else None
        row["dependencies"] = {
            "free_float_days": task["free_float_days"],
            "total_float_days": task["total_float_days"],
            "current_float_days": task["current_float_days"],
            "own_delay_days": task["own_delay_days"],
            "inherited_delay_days": task["inherited_delay_days"],
            "impact_rnv_days": project_delay if tid in ancestors and task["current_float_days"] <= 0 else 0,
            "predecessors": [
                {"id": e["id"], "name": tasks[e["id"]]["name"], "type": e["type"], "lag_days": e["lag_days"]}
                for e in task["predecessors"]
            ],
            "successors": [
                {"id": e["successor_id"], "name": tasks[e["successor_id"]]["name"], "type": e["type"], "lag_days": e["lag_days"]}
                for e in task["successors"]
            ],
        }
    management = schedule.get("management") or []
    for node in _walk_nodes(management):
        if node.get("id"):
            src = flat.get(str(node.get("id") or "").strip())
            if src:
                node.update({k: src.get(k) for k in ("forecast_finish", "delta_days", "dependencies")})
                node.setdefault("level", "task")
    rebaselines = _rebaselines(project)
    def aggregate(node: dict[str, Any]) -> None:
        children = [c for c in node.get("children", []) if isinstance(c, dict)]
        for child in children:
            aggregate(child)
        if node.get("level") == "rss":
            code = str(node.get("code") or "").rstrip(".")
            rb = rebaselines.get(code)
            if rb and rb.get("finish"):
                node["original_plan_finish"] = node.get("plan_finish")
                node["plan_finish"] = monitor._iso(rb["finish"])
                node["rebaseline"] = {**rb, "finish": monitor._iso(rb["finish"])}
        if children:
            forecasts = [monitor._day(c.get("forecast_finish")) for c in children]
            forecasts = [d for d in forecasts if d]
            if forecasts:
                node["forecast_finish"] = monitor._iso(max(forecasts))
            finish = monitor._day(node.get("plan_finish"))
            forecast = monitor._day(node.get("forecast_finish"))
            if finish and forecast:
                node["delta_days"] = (forecast - finish).days
            deps = [c.get("dependencies") for c in children if c.get("dependencies")]
            if deps:
                node["dependencies"] = {
                    "current_float_days": min((d.get("current_float_days", 10**9) for d in deps), default=None),
                    "impact_rnv_days": max((d.get("impact_rnv_days", 0) for d in deps), default=0),
                    "predecessors": [], "successors": [],
                }
    for root in management:
        aggregate(root)
    schedule["dependency_graph"] = {
        "known": True, "source": pm["source"], "tasks": len(tasks),
        "rnv_baseline": monitor._iso(baseline_rnv), "rnv_forecast": monitor._iso(forecast_rnv),
        "rnv_delay_days": project_delay,
    }
    if forecast_rnv:
        schedule["forecast_end"] = monitor._iso(forecast_rnv)
    if baseline_rnv:
        schedule["approved_end"] = monitor._iso(baseline_rnv)
    return view
