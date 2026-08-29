"""Гарантийные удержания: скрытый резерв, которого нет в лимите РСС.

«В РСС банка сумма по договору берётся общая, но ГУ до момента погашения ПФ не
заплатятся — по сути это скрытый резерв: общий лимит РСС не учитывает, что
потрачен будет не весь» (владелец, 29.08.2026, с реестром ГУ на руках).

На присланном реестре это 70,4 млн ₽ к выплате при 2,79 млрд ₽ договоров
(2,65%), и 97% этой суммы имеет срок выплаты 2030–2032 — то есть за горизонтом
стройки и после погашения ПФ. В потребности до конца стройки эти деньги стоят,
а потрачены в ней не будут.

Что здесь считается: суммы реестра складываются и делятся на «до» и «после»
названной даты. Ничего сверх этого — потребность, лимиты и дефицит считает
дашборд, и второго счёта тех же величин здесь нет.

Читается по ЗАГОЛОВКАМ, а не по номерам колонок: реестр ведут руками, колонку
вставят — и разбор по позициям молча возьмёт соседнюю. Не нашлась колонка —
это отказ с именем колонки, а не ноль.
"""

from __future__ import annotations

import datetime
import io
import re
from typing import Any

# Заголовки реестра. Слева — наше имя поля, справа — по каким словам искать
# колонку. Сравнение по «начинается с»: в живом файле у заголовков хвосты
# («Выплачено ГУ по состточнию на …») и переносы строк.
COLUMNS: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("counterparty", ("контрагент",)),
    ("contract", ("реквизиты договора", "договор")),
    ("contract_amount", ("стоимость договора",)),
    ("share_pct", ("размер гу",)),
    ("amount", ("сумма гу",)),
    ("paid", ("выплачено",)),
    ("left", ("остаток к выплате", "остаток")),
    ("due", ("дата выплаты",)),
    ("note", ("примечание",)),
    ("works_end", ("дата окончания работ",)),
)
REQUIRED = ("contract_amount", "amount", "left", "due")

_DATE = re.compile(r"(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{4})")
_ISO = re.compile(r"(\d{4})-(\d{2})-(\d{2})")


class RetentionUnreadable(ValueError):
    """Реестр не разобрался. Это отказ с причиной, а не пустой список."""


def _clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip().lower()


def _money(value: Any) -> float | None:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = re.sub(r"[^\d,.\-]", "", str(value or "")).replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _date(value: Any) -> datetime.date | None:
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    text = str(value or "")
    iso = _ISO.search(text)
    if iso:
        year, month, day = (int(part) for part in iso.groups())
        try:
            return datetime.date(year, month, day)
        except ValueError:
            return None
    match = _DATE.search(text)
    if not match:
        return None
    day, month, year = (int(part) for part in match.groups())
    try:
        return datetime.date(year, month, day)
    except ValueError:
        return None


def read_retention(data: bytes) -> dict[str, Any]:
    """Реестр ГУ: строки договоров и итоги.

    Итоговая строка файла в строки не берётся: у неё нет контрагента, а сумма
    её равна сумме остальных — сложенная дважды, она удвоила бы резерв.
    """
    from openpyxl import load_workbook

    try:
        book = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        raise RetentionUnreadable(f"файл не открылся как книга Excel: {exc}") from exc
    try:
        sheet = book[book.sheetnames[0]]
        grid = [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        book.close()

    header_at, columns = None, {}
    for index, row in enumerate(grid[:20]):
        found: dict[str, int] = {}
        for position, cell in enumerate(row):
            text = _clean(cell)
            if not text:
                continue
            for name, marks in COLUMNS:
                if name in found:
                    continue
                if any(text.startswith(mark) for mark in marks):
                    found[name] = position
                    break
        if all(name in found for name in REQUIRED):
            header_at, columns = index, found
            break
    if header_at is None:
        raise RetentionUnreadable(
            "в файле не нашлось шапки реестра ГУ: нужны колонки "
            "«Стоимость Договора», «Сумма ГУ», «Остаток к выплате», «Дата выплаты ГУ»")

    rows: list[dict[str, Any]] = []
    total_row: dict[str, float] | None = None
    for row in grid[header_at + 1:]:
        def cell(name: str) -> Any:
            position = columns.get(name)
            return row[position] if position is not None and position < len(row) else None

        counterparty = str(cell("counterparty") or "").strip()
        contract_amount = _money(cell("contract_amount"))
        amount = _money(cell("amount"))
        left = _money(cell("left"))
        due = _date(cell("due"))
        share = _money(cell("share_pct"))
        if not counterparty and not amount and not left:
            continue
        # Строка без контрагента бывает двух видов, и путать их дорого:
        # продолжение договора выше (второй этап выплаты — у него есть свой
        # срок или процент) и ИТОГОВАЯ строка реестра, у которой нет ни того
        # ни другого. Итог, посчитанный вместе со строками, удваивает резерв —
        # на присланном реестре это 140 млн ₽ вместо 70.
        continued = bool(not counterparty and (due or share))
        if not counterparty and not continued:
            # Итог файла не выбрасывается, а запоминается: свой счёт, не
            # сошедшийся с итогом источника, — это находка, а не мелочь.
            total_row = {"contract_amount": contract_amount or 0.0,
                         "amount": amount or 0.0, "paid": _money(cell("paid")) or 0.0,
                         "left": left or 0.0}
            continue
        if continued:
            # Стоимость договора у продолжения та же, что у строки выше:
            # сложенная второй раз, она удваивает объём подряда.
            contract_amount = 0.0
        rows.append({
            "counterparty": counterparty,
            "contract": str(cell("contract") or "").strip(),
            "contract_amount": contract_amount or 0.0,
            "share_pct": share,
            "continued": continued,
            "amount": amount or 0.0,
            "paid": _money(cell("paid")) or 0.0,
            "left": left if left is not None else (amount or 0.0),
            "due": due.isoformat() if due else "",
            "due_note": str(cell("note") or "").strip(),
            "works_end": (_date(cell("works_end")).isoformat()
                          if _date(cell("works_end")) else ""),
        })
    if not rows:
        raise RetentionUnreadable("шапка реестра ГУ нашлась, а строк под ней нет")
    return {"rows": rows, "columns": sorted(columns), "file_total": total_row}


def summary(register: dict[str, Any], *, horizon: Any = None) -> dict[str, Any]:
    """Итоги реестра и та его часть, что лежит за горизонтом стройки.

    `horizon` — дата, после которой выплата ГУ стройку уже не касается: ввод
    объекта или погашение ПФ. Не задана — «отложенное» не считается вовсе, а
    не берётся нулём: без горизонта вопрос «что за ним» не имеет ответа.
    """
    rows = register.get("rows") or []
    edge = _date(horizon)
    total = sum(float(row.get("amount") or 0.0) for row in rows)
    paid = sum(float(row.get("paid") or 0.0) for row in rows)
    left = sum(float(row.get("left") or 0.0) for row in rows)
    contracts = sum(float(row.get("contract_amount") or 0.0) for row in rows)
    deferred = None
    undated = sum(float(row.get("left") or 0.0) for row in rows if not row.get("due"))
    if edge is not None:
        deferred = sum(float(row.get("left") or 0.0) for row in rows
                       if (_date(row.get("due")) or edge) > edge)
    # Свой счёт против итога файла. Совпало — молчим; разошлось — говорим, на
    # сколько и по чему: реестр ведут руками, и строка, выпавшая из его
    # собственной формулы, иначе не видна никому.
    stated = register.get("file_total") or {}
    mismatch = {}
    for name, ours in (("contract_amount", contracts), ("amount", total), ("left", left)):
        said = float(stated.get(name) or 0.0)
        if said and abs(said - ours) > 1.0:
            mismatch[name] = {"file": said, "rows": ours, "delta": ours - said}
    return {
        "known": True,
        "contracts": len(rows),
        "file_total_mismatch": mismatch,
        "contract_amount": contracts,
        "amount": total,
        "paid": paid,
        "left": left,
        "share_of_contracts": (total / contracts) if contracts else None,
        "horizon": edge.isoformat() if edge else "",
        # Отложенное за горизонт — тот самый скрытый резерв: в потребности эти
        # деньги стоят, а в стройке потрачены не будут.
        "deferred_after_horizon": deferred,
        "undated_left": undated,
    }
