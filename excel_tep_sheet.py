"""Читаемый и редактируемый ТЭП в полной Excel-модели DevelopAid.

Старый шаблон содержит рабочую таблицу ТЭП внутри технической раскладки.
Переставлять её строки нельзя: фиксированные ссылки книги и openpyxl при
insert_rows — плохое сочетание. Поэтому здесь используется безопасная схема:

1. создаётся отдельный лист ``ТЭП проекта`` с нормальной продуктовой таблицей;
2. существующая техническая таблица остаётся ровно на своих строках;
3. её ячейки B:G становятся простыми ссылками на новый лист.

Все старые формулы продолжают смотреть туда же, куда смотрели, но человек
правит ТЭП в одном понятном месте. Это также делает кладовки видимой вводной:
их количество находится в том же ряду, что остальные продукты, и уже
существующая модель продаж берёт его через техническую колонку ``Количество``.
"""

from __future__ import annotations

import io
from functools import wraps
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter

SHEET_NAME = "ТЭП проекта"

PRODUCT_ORDER = (
    "apartments",
    "ground_commercial",
    "underground_parking",
    "storage",
    "standalone_retail",
    "offices",
    "above_parking",
    "kindergarten",
    "school",
    "clinic",
)

FIELDS = (
    ("gns", "ГНС, м²"),
    ("total_area", "Общая площадь, м²"),
    ("useful", "Полезная площадь, м²"),
    ("saleable", "Продаваемая площадь, м²"),
    ("transfer", "Передаваемая площадь, м²"),
    ("units", "Количество, шт."),
)

# Техническая таблица в действующей книге: A = продукт, B:G = те же шесть
# показателей. Именно поэтому модель продаж читает продаваемую из колонки 5,
# а паркинг/кладовки — количество из колонки 7.
TECH_FIELD_COLUMNS = {
    "gns": 2,
    "total_area": 3,
    "useful": 4,
    "saleable": 5,
    "transfer": 6,
    "units": 7,
}


def _text(value: Any) -> str:
    return " ".join(str(value or "").replace("\n", " ").split()).strip().casefold()


def _ordered_keys(core: Any, tep: dict[str, Any]) -> list[str]:
    known = set(getattr(core, "TEP_DEFAULT", {}) or {}) | set(tep)
    ordered = [key for key in PRODUCT_ORDER if key in known]
    ordered.extend(key for key in known if key not in ordered)
    return ordered


def _labels(core: Any, tep: dict[str, Any]) -> dict[str, str]:
    defaults = getattr(core, "TEP_DEFAULT", {}) or {}
    result: dict[str, str] = {}
    for key in _ordered_keys(core, tep):
        current = tep.get(key) or {}
        base = defaults.get(key) or {}
        result[key] = str(current.get("label") or base.get("label") or key)
    return result


def find_technical_tep_table(book: Any, labels: dict[str, str]) -> tuple[Any | None, dict[str, int]]:
    """Находит существующую таблицу ТЭП без знания номера строки.

    Требуем одновременно заголовок ``Продукт / ГНС`` и несколько известных
    продуктовых строк. Так мы не перепутаем её с отчётной таблицей, где те же
    названия могут встречаться повторно.
    """
    expected = {_text(label): key for key, label in labels.items() if label}
    best: tuple[int, Any, dict[str, int]] | None = None

    for ws in book.worksheets:
        if ws.title == SHEET_NAME:
            continue
        header_rows: list[int] = []
        for row in range(1, ws.max_row + 1):
            left = _text(ws.cell(row=row, column=1).value)
            second = _text(ws.cell(row=row, column=2).value)
            if "продукт" in left and "гнс" in second:
                header_rows.append(row)
        if not header_rows:
            continue

        rows: dict[str, int] = {}
        for row in range(1, ws.max_row + 1):
            key = expected.get(_text(ws.cell(row=row, column=1).value))
            if key:
                rows[key] = row
        if len(rows) < 3:
            continue

        # Продуктовые строки должны находиться ниже хотя бы одного заголовка.
        below = [header for header in header_rows if min(rows.values()) > header]
        if not below:
            continue
        score = len(rows) * 10 - (min(rows.values()) - max(below))
        if best is None or score > best[0]:
            best = (score, ws, rows)

    return (best[1], best[2]) if best else (None, {})


def _style_sheet(ws: Any) -> None:
    navy = "162235"
    navy2 = "24364F"
    blue = "2F8CFF"
    pale = "EAF2FB"
    pale_blue = "EAF4FF"
    grid = "D8E1EB"
    dark = "1E293B"
    muted = "64748B"
    white = "FFFFFF"

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B5"
    ws.auto_filter.ref = f"A4:J{ws.max_row}"

    ws.merge_cells("A1:J1")
    ws["A1"].fill = PatternFill("solid", fgColor=navy)
    ws["A1"].font = Font(name="Aptos Display", size=18, bold=True, color=white)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"].fill = PatternFill("solid", fgColor=navy)
    ws["A2"].font = Font(name="Aptos", size=10, color="C9D7E8")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 34

    for cell in ws[4]:
        cell.fill = PatternFill("solid", fgColor=navy2)
        cell.font = Font(name="Aptos", size=9, bold=True, color=white)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 42

    thin = Side(style="thin", color=grid)
    for row in range(5, ws.max_row + 1):
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.border = Border(bottom=thin)
            cell.font = Font(name="Aptos", size=10, color=dark)
            cell.alignment = Alignment(vertical="center", wrap_text=(col in (1, 10)))
        if row % 2 == 0:
            for col in range(1, 11):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="F8FAFC")
        ws.row_dimensions[row].height = 26

    # Вводимые фактические ТЭП визуально отделены от расчётных долей.
    for row in range(5, ws.max_row + 1):
        for col in range(2, 8):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=pale_blue)
            ws.cell(row=row, column=col).protection = Protection(locked=False)
        for col in (8, 9):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=pale)
            ws.cell(row=row, column=col).font = Font(name="Aptos", size=10, color=muted)

    widths = {1: 25, 2: 15, 3: 18, 4: 18, 5: 20, 6: 20, 7: 16, 8: 15, 9: 20, 10: 34}
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # Небольшой акцент на кладовках: раньше цена была, а сама вводная терялась.
    for row in range(5, ws.max_row + 1):
        if _text(ws.cell(row=row, column=1).value) in {"кладовки", "кладовые"}:
            ws.cell(row=row, column=1).font = Font(name="Aptos", size=10, bold=True, color=blue)
            break


def add_tep_project_sheet(data: bytes, tep: dict[str, dict[str, Any]], core: Any) -> bytes:
    """Добавляет лист ТЭП и связывает с ним старую техническую таблицу."""
    if not data or not isinstance(tep, dict):
        return data

    source = io.BytesIO(data)
    book = load_workbook(source, data_only=False, keep_vba=False)
    if SHEET_NAME in book.sheetnames:
        del book[SHEET_NAME]

    labels = _labels(core, tep)
    technical_ws, technical_rows = find_technical_tep_table(book, labels)

    # Ставим лист сразу после «Вводные», если она есть: он относится к исходным
    # данным, а не к отчётной части книги.
    index = book.sheetnames.index("Вводные") + 1 if "Вводные" in book.sheetnames else 0
    ws = book.create_sheet(SHEET_NAME, index=index)
    ws["A1"] = "ТЭП ПРОЕКТА"
    ws["A2"] = (
        "Редактируйте фактические площади и количества здесь. Существующая техническая "
        "таблица книги остаётся на прежних строках и связана с этим листом формулами — "
        "поэтому фиксированные ссылки модели не сдвигаются."
    )

    headers = ["Продукт", *(label for _, label in FIELDS), "Общая / ГНС", "Продаваемая / общей", "Примечание"]
    for col, value in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=value)

    row_by_key: dict[str, int] = {}
    defaults = getattr(core, "TEP_DEFAULT", {}) or {}
    ratios = getattr(core, "TEP_RATIOS", {}) or {}
    for offset, key in enumerate(_ordered_keys(core, tep), start=5):
        row_by_key[key] = offset
        current = {**(defaults.get(key) or {}), **(tep.get(key) or {})}
        ws.cell(row=offset, column=1, value=labels[key])
        for col, (field, _label) in enumerate(FIELDS, start=2):
            value = current.get(field, 0)
            ws.cell(row=offset, column=col, value=float(value or 0))
            ws.cell(row=offset, column=col).number_format = '#,##0.0'

        ws.cell(row=offset, column=8, value=f'=IFERROR(B{offset}/B{offset},0)')
        # Исправляем формулу ниже сразу: H — общая / ГНС, I — продаваемая / общей.
        ws.cell(row=offset, column=8, value=f'=IFERROR(C{offset}/B{offset},0)')
        ws.cell(row=offset, column=9, value=f'=IFERROR(E{offset}/C{offset},0)')
        ws.cell(row=offset, column=8).number_format = '0.0%'
        ws.cell(row=offset, column=9).number_format = '0.0%'

        note = ""
        if key == "storage":
            note = "Выручка = количество × цена кладовой во «Вводных»"
        elif key in ratios:
            source_note = str((ratios.get(key) or {}).get("source") or "")
            if source_note:
                note = f"Базовая пропорция: {source_note}"
        ws.cell(row=offset, column=10, value=note)

    _style_sheet(ws)

    # Старую таблицу НЕ двигаем. Только заменяем её входные значения ссылками.
    # Все формулы книги, уже ссылающиеся на B:G этой таблицы, остаются прежними.
    if technical_ws is not None:
        for key, old_row in technical_rows.items():
            new_row = row_by_key.get(key)
            if not new_row:
                continue
            for field, old_col in TECH_FIELD_COLUMNS.items():
                new_col = 2 + [name for name, _ in FIELDS].index(field)
                technical_ws.cell(row=old_row, column=old_col, value=(
                    f"='{SHEET_NAME}'!${get_column_letter(new_col)}${new_row}"
                ))

    output = io.BytesIO()
    book.save(output)
    return output.getvalue()


def install(core: Any) -> None:
    """Подключает лист к действующему генератору полной PLATO-книги."""
    original = getattr(core, "build_plato_model_v2", None)
    if not callable(original) or getattr(original, "_developaid_tep_sheet", False):
        return

    @wraps(original)
    def wrapped(*args: Any, **kwargs: Any):
        result = original(*args, **kwargs)
        if not (isinstance(result, tuple) and result and isinstance(result[0], (bytes, bytearray))):
            return result
        tep = kwargs.get("tep")
        if tep is None and len(args) >= 2:
            tep = args[1]
        if not isinstance(tep, dict):
            return result
        data = add_tep_project_sheet(bytes(result[0]), tep, core)
        return (data, *result[1:])

    wrapped._developaid_tep_sheet = True  # type: ignore[attr-defined]
    core.build_plato_model_v2 = wrapped
