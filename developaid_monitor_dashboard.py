"""Management KPI, funding-risk and recurring-input layer for Project Monitor.

Funding is calculated article-by-article. A free balance in one RSS article is
not silently allowed to finance another article. Once an article's own balance
is exhausted, that month's need consumes the explicit 2.8/2.9 reserve; only
after the reserve is exhausted does the model report uncovered financing.
"""
from __future__ import annotations

import datetime
import io
import re
from pathlib import Path
from typing import Any

import developaid_actuals as actuals
import developaid_monitor as monitor
import developaid_monitor_schedule_graph as schedule_graph

_INSTALLED = False
_ORIGINAL_BUILD = None
_ORIGINAL_STORE_SALES_FILE = None
_ORIGINAL_STORE_PROPOSAL = None


def _norm(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower().replace("ё", "е"))


def _finance_file(project: str) -> Path:
    return monitor._project_dir(project) / "baseline" / "finance.xlsx"


def _sales_dir(project: str) -> Path:
    path = monitor._project_dir(project) / "sales"
    path.mkdir(parents=True, exist_ok=True)
    return path


def _latest_sales(project: str, upto: datetime.date | None = None) -> Path | None:
    folder = monitor._project_dir(project) / "sales"
    if not folder.exists():
        return None
    rows = sorted(folder.glob("*.xlsx"))
    if upto:
        rows = [row for row in rows if row.stem[:10] <= upto.isoformat()]
    return rows[-1] if rows else None


def _find_header(ws: Any, needle: str, max_rows: int = 20) -> tuple[int, int] | None:
    wanted = _norm(needle)
    for r, row in enumerate(
        ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_rows), values_only=True), 1
    ):
        for c, value in enumerate(row, 1):
            if wanted in _norm(value):
                return r, c
    return None


def _month_value(value: Any, year: int, previous_month: int) -> tuple[datetime.date | None, int, int]:
    if isinstance(value, datetime.datetime):
        day = value.date().replace(day=1)
        return day, day.year, day.month
    if isinstance(value, datetime.date):
        day = value.replace(day=1)
        return day, day.year, day.month
    ru_months = {
        "январь": 1, "февраль": 2, "март": 3, "апрель": 4,
        "май": 5, "июнь": 6, "июль": 7, "август": 8,
        "сентябрь": 9, "октябрь": 10, "ноябрь": 11, "декабрь": 12,
    }
    month = ru_months.get(_norm(value))
    if not month:
        return None, year, previous_month
    if previous_month and month < previous_month:
        year += 1
    return datetime.date(year, month, 1), year, month


_FINANCE_CACHE: dict[tuple, dict[str, Any]] = {}


def _finance_baseline(project: str) -> dict[str, Any]:
    """Финансовый baseline. Результат помнится по (путь, mtime, размер).

    Читается один раз одним проходом: на read-only книге каждый `ws.cell()`
    перечитывает XML листа целиком, и 904 обращения стоили 23 секунды на
    срез — при том что baseline по определению неизменен.
    """
    import copy as _copy
    import os as _os

    path = _finance_file(project)
    if not path.exists():
        return {"known": False, "reason": "не загружен финансовый baseline"}
    stat = _os.stat(path)
    key = (str(path), stat.st_mtime_ns, stat.st_size)
    if key in _FINANCE_CACHE:
        return _copy.deepcopy(_FINANCE_CACHE[key])
    result = _read_finance_baseline(path)
    if len(_FINANCE_CACHE) > 8:
        _FINANCE_CACHE.clear()
    _FINANCE_CACHE[key] = result
    return _copy.deepcopy(result)


class _Grid:
    """Лист, прочитанный одним проходом, с интерфейсом cell(row, col)."""

    def __init__(self, ws: Any) -> None:
        self.rows = list(ws.iter_rows(values_only=True))
        self.max_row = len(self.rows)
        self.max_column = max((len(row) for row in self.rows), default=0)

    class _Cell:
        __slots__ = ("value",)

        def __init__(self, value: Any) -> None:
            self.value = value

    def cell(self, row: int, column: int) -> "_Grid._Cell":
        values = self.rows[row - 1] if 0 < row <= len(self.rows) else ()
        return self._Cell(values[column - 1] if 0 < column <= len(values) else None)

    def iter_rows(self, min_row: int = 1, max_row: int | None = None,
                  values_only: bool = True) -> Any:
        stop = max_row if max_row is not None else len(self.rows)
        return iter(self.rows[min_row - 1:stop])


def _read_finance_baseline(path: Path) -> dict[str, Any]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if "Расчет стоимости строительства" not in wb.sheetnames:
            return {"known": False, "reason": "нет листа «Расчет стоимости строительства»"}
        ws = _Grid(wb["Расчет стоимости строительства"])
        approved_hdr = _find_header(ws, "Утвержденная фин.модель проекта")
        need_hdr = _find_header(ws, "Средства на завершение строительства") or _find_header(
            ws, "Средства на завершение согласно бюджету"
        )
        paid_hdr = _find_header(ws, "Оплачено по состояни")
        programme_hdr = _find_header(ws, "производств")
        tail_hdr = _find_header(ws, "Остаток к выполнению на 01.04.27")
        rss_limit_hdr = _find_header(ws, "Общая сметная стоимость")
        if not approved_hdr or not need_hdr:
            return {"known": False, "reason": "не найдены колонки утвержденной модели/остатка"}

        approved_col = approved_hdr[1]
        need_col = need_hdr[1]
        paid_col = paid_hdr[1] if paid_hdr else 8
        rss_limit_col = rss_limit_hdr[1] if rss_limit_hdr else 4
        programme_col = programme_hdr[1] if programme_hdr else 10
        tail_col = tail_hdr[1] if tail_hdr else 20

        total_row = None
        reserve_rows: list[int] = []
        source_rows: list[dict[str, Any]] = []
        stack: list[dict[str, Any]] = []
        for r, values in enumerate(ws.iter_rows(values_only=True), 1):
            code = actuals._code(values[0] if values else None)
            article = _norm(values[1] if len(values) > 1 else None)
            if total_row is None and "всего инвестиционные расходы глава 2, 3" in article:
                total_row = r
            if code in {"2.8", "2.9"}:
                reserve_rows.append(r)
            if not code:
                continue
            depth = code.count(".") + 1
            while stack and stack[-1]["depth"] >= depth:
                stack.pop()
            row = {
                "row": r,
                "code": code,
                "depth": depth,
                "parent": stack[-1] if stack else None,
                "has_children": False,
            }
            if stack:
                stack[-1]["has_children"] = True
            source_rows.append(row)
            stack.append(row)

        if total_row is None:
            return {"known": False, "reason": "не найдена итоговая строка глав 2+3"}

        approved = actuals._money(ws.cell(total_row, approved_col).value)
        completion_need = actuals._money(ws.cell(total_row, need_col).value)
        paid_at_baseline = actuals._money(ws.cell(total_row, paid_col).value)
        tail_after_apr = actuals._money(ws.cell(total_row, tail_col).value)

        reserve = 0.0
        reserve_parts: dict[str, float] = {}
        # Резерв ищется по трём колонкам по убыванию доверия: «на завершение
        # согласно бюджету», «согласно лимитам», «Общая сметная стоимость
        # Увеличенная». Первоначальная смета — последней: на Кутузове бюджетная
        # колонка резервов пуста, и запасной ход на «первоначальную» терял
        # 2.8 целиком и показывал 209,7 млн вместо 306,1.
        limits_hdr = _find_header(ws, "Средства на завершение согласно лимитам")
        increased_hdr = _find_header(ws, "Увеличенная")
        for r in reserve_rows:
            code = actuals._code(ws.cell(r, 1).value)
            amount = 0.0
            for col in (need_col,
                        limits_hdr[1] if limits_hdr else None,
                        increased_hdr[1] if increased_hdr else None,
                        rss_limit_col):
                if col is None:
                    continue
                amount = actuals._money(ws.cell(r, col).value)
                if amount > 0:
                    break
            reserve += max(0.0, amount)
            reserve_parts[code] = max(0.0, amount)

        month_row = (programme_hdr[0] + 1) if programme_hdr else 9
        month_columns: list[tuple[int, datetime.date]] = []
        year = 2026
        previous_month = 0
        for c in range(programme_col, min(ws.max_column, programme_col + 18) + 1):
            month, year, previous_month = _month_value(
                ws.cell(month_row, c).value, year, previous_month
            )
            if month:
                month_columns.append((c, month))

        monthly_total: dict[str, float] = {}
        for c, month in month_columns:
            monthly_total[month.isoformat()] = actuals._money(ws.cell(total_row, c).value)

        # Only terminal rows are financing buckets. Parent rows (2.2, 2.2.2,
        # 2.3...) repeat the same programme and would double-count need.
        articles: dict[str, dict[str, Any]] = {}
        for meta in source_rows:
            code = meta["code"]
            if meta["has_children"] or not code.startswith("2") or code in {"2.8", "2.9"}:
                continue
            r = meta["row"]
            month_need = {
                month.isoformat(): max(0.0, actuals._money(ws.cell(r, c).value))
                for c, month in month_columns
            }
            if not any(month_need.values()):
                continue
            item = articles.setdefault(code, {
                "code": code,
                "name": str(ws.cell(r, 2).value or "").strip(),
                "rss_limit": 0.0,
                "paid_at_baseline": 0.0,
                "monthly_need": {month.isoformat(): 0.0 for _, month in month_columns},
            })
            item["rss_limit"] += actuals._money(ws.cell(r, rss_limit_col).value)
            item["paid_at_baseline"] += actuals._money(ws.cell(r, paid_col).value)
            for month, amount in month_need.items():
                item["monthly_need"][month] = item["monthly_need"].get(month, 0.0) + amount

        return {
            "known": True,
            "source": path.name,
            "approved": approved,
            "completion_need_at_baseline": completion_need,
            "paid_at_baseline": paid_at_baseline,
            "reserve": reserve,
            "reserve_parts": reserve_parts,
            "monthly_need": monthly_total,
            "tail_after_apr": tail_after_apr,
            "articles": articles,
        }
    finally:
        wb.close()


def _rss_ch23(estimate: dict[str, Any]) -> dict[str, float]:
    rows = estimate.get("by_code") or {}
    return {
        "limit": sum(float((rows.get(code) or {}).get("estimate") or 0.0) for code in ("2", "3")),
        "paid_bank_sheet": sum(float((rows.get(code) or {}).get("paid") or 0.0) for code in ("2", "3")),
        "contracted": sum(float((rows.get(code) or {}).get("contracted") or 0.0) for code in ("2", "3")),
    }


def _payment_total_ch23(rss: Path, estimate: dict[str, Any]) -> float:
    payments = actuals.read_payments(rss)
    parents = {row["code"]: row.get("parent") for row in estimate.get("rows") or []}

    def root(code: str) -> str:
        seen = set()
        while code and code not in seen:
            seen.add(code)
            parent = str(parents.get(code) or "")
            if not parent:
                return code.split(".")[0]
            code = parent
        return code.split(".")[0] if code else ""

    return sum(
        float(row.get("amount") or 0.0)
        for row in payments.get("rows") or []
        if root(str(row.get("estimate_code") or "").rstrip(".")) in {"2", "3"}
    )


def _payment_by_code(rss: Path) -> dict[str, float]:
    result: dict[str, float] = {}
    for row in actuals.read_payments(rss).get("rows") or []:
        code = actuals._code(row.get("estimate_code"))
        if not code:
            continue
        result[code] = result.get(code, 0.0) + float(row.get("amount") or 0.0)
    return result


def _interpolated_crossing(
    month: datetime.date, month_amount: float, before: float, threshold: float
) -> datetime.date:
    if month_amount <= 0:
        return month
    ratio = max(0.0, min(1.0, (threshold - before) / month_amount))
    nxt = (
        datetime.date(month.year + 1, 1, 1)
        if month.month == 12
        else datetime.date(month.year, month.month + 1, 1)
    )
    days = max(1, (nxt - month).days)
    return month + datetime.timedelta(days=max(0, min(days - 1, round(ratio * days))))


def _article_waterfall(
    articles: dict[str, dict[str, Any]],
    reserve: float,
    cut: datetime.date,
    current_paid: dict[str, float] | None = None,
) -> dict[str, Any]:
    """Monthly article-limit -> reserve -> uncovered waterfall."""
    current_paid = current_paid or {}
    cut_month = cut.replace(day=1)
    months = sorted({
        monitor._day(month)
        for item in articles.values()
        for month in (item.get("monthly_need") or {})
        if monitor._day(month) is not None and monitor._day(month) >= cut_month
    })
    months = [month for month in months if month is not None]

    state: dict[str, float] = {}
    opening_raw: dict[str, float] = {}
    for code, item in articles.items():
        paid = current_paid.get(code)
        if paid is None:
            paid = float(item.get("paid_at_baseline") or 0.0)
        raw = float(item.get("rss_limit") or 0.0) - float(paid or 0.0)
        opening_raw[code] = raw
        state[code] = max(0.0, raw)

    reserve_balance = max(0.0, float(reserve or 0.0))
    reserve_start = None
    reserve_exhaustion = None
    cumulative_reserve_need = 0.0
    cumulative_unfunded = 0.0
    monthly_need: dict[str, float] = {}
    monthly_reserve_draw: dict[str, float] = {}
    monthly_unfunded: dict[str, float] = {}
    monthly_reserve_balance: dict[str, float] = {}
    article_rows: list[dict[str, Any]] = []

    first_shortfall: dict[str, datetime.date] = {}
    article_reserve_take: dict[str, float] = {}
    article_unfunded: dict[str, float] = {}
    for month in months:
        month_key = month.isoformat()
        month_need_total = 0.0
        month_shortfall = 0.0
        month_shortages: dict[str, float] = {}
        for code, item in articles.items():
            need = max(0.0, float((item.get("monthly_need") or {}).get(month_key, 0.0) or 0.0))
            month_need_total += need
            available = state.get(code, 0.0)
            own_funding = min(available, need)
            shortage = max(0.0, need - own_funding)
            state[code] = max(0.0, available - need)
            month_shortfall += shortage
            if shortage > 0:
                month_shortages[code] = shortage
                if code not in first_shortfall:
                    first_shortfall[code] = month

        monthly_need[month_key] = month_need_total
        before_need = cumulative_reserve_need
        cumulative_reserve_need += month_shortfall
        if month_shortfall > 0 and reserve_start is None:
            reserve_start = month

        draw = min(reserve_balance, month_shortfall)
        uncovered = max(0.0, month_shortfall - draw)
        reserve_balance -= draw
        cumulative_unfunded += uncovered
        monthly_reserve_draw[month_key] = draw
        monthly_unfunded[month_key] = uncovered
        monthly_reserve_balance[month_key] = reserve_balance

        # В месяц, когда резерва на всех не хватает, покрытая доля у всех
        # статей одна: резерв — общий, отдельной очереди статей внутри
        # месяца в ДДС нет.
        covered_share = draw / month_shortfall if month_shortfall > 0 else 0.0
        for code, shortage in month_shortages.items():
            covered = shortage * covered_share
            article_reserve_take[code] = article_reserve_take.get(code, 0.0) + covered
            article_unfunded[code] = article_unfunded.get(code, 0.0) + (shortage - covered)

        if reserve_exhaustion is None and month_shortfall > 0 and cumulative_reserve_need > reserve:
            reserve_exhaustion = _interpolated_crossing(
                month, month_shortfall, before_need, reserve
            )

    for code, item in sorted(articles.items()):
        limit = float(item.get("rss_limit") or 0.0)
        need_total = sum(max(0.0, float(v or 0.0))
                         for v in (item.get("monthly_need") or {}).values())
        first_month = first_shortfall.get(code)
        reserve_left_after_first = None
        if first_month is not None:
            reserve_left_after_first = monthly_reserve_balance.get(first_month.isoformat())
        article_rows.append({
            "code": code,
            "name": item.get("name", ""),
            "limit": limit,
            "need_total": need_total,
            "opening_limit_raw": opening_raw.get(code, 0.0),
            "opening_limit": max(0.0, opening_raw.get(code, 0.0)),
            "remaining_limit": state.get(code, 0.0),
            "first_reserve_month": monitor._iso(first_shortfall.get(code)),
            "reserve_take": article_reserve_take.get(code, 0.0),
            "unfunded_take": article_unfunded.get(code, 0.0),
            "reserve_left_after_first": reserve_left_after_first,
        })

    # Когда дофинансирование понадобится. Потребность без срока — половина
    # ответа: «нужно 3,6 млрд» и «нужно 3,6 млрд с марта» — разные новости
    # (владелец, 29.08.2026: «потребность в доп финансировании нужно выделить
    # по сумме и сроку»). Месяц выбирается, а не считается: это первый месяц с
    # непокрытой потребностью.
    unfunded_from = next((month for month in sorted(monthly_unfunded)
                          if monthly_unfunded[month] > 0), None)
    return {
        "opening_bank_remaining": sum(max(0.0, value) for value in opening_raw.values()),
        "opening_article_deficit": sum(max(0.0, -value) for value in opening_raw.values()),
        "additional_financing_from": unfunded_from,
        "remaining_article_limits": sum(state.values()),
        "reserve_start": reserve_start,
        "reserve_exhaustion": reserve_exhaustion,
        "reserve_balance": reserve_balance,
        "reserve_need": cumulative_reserve_need,
        "additional_financing": cumulative_unfunded,
        "monthly_need": monthly_need,
        "monthly_reserve_draw": monthly_reserve_draw,
        "monthly_unfunded": monthly_unfunded,
        "monthly_reserve_balance": monthly_reserve_balance,
        "articles": article_rows,
    }


def _retention(project: str, horizon: Any) -> dict[str, Any] | None:
    """Гарантийные удержания против горизонта стройки.

    «В РСС банка сумма по договору берётся общая, но ГУ до момента погашения ПФ
    не заплатятся — по сути это скрытый резерв» (владелец, 29.08.2026). Горизонт
    — прогнозный ввод: за ним выплата ГУ стройку уже не касается.

    Реестра нет — это `None`, а не нулевой резерв: «не загружали» и «удержаний
    нет» на экране значат разное.
    """
    import developaid_monitor_retention as retention

    register = monitor.latest_retention(project)
    if not register:
        return None
    if register.get("known") is False:
        return register
    got = retention.summary(register, horizon=horizon)
    got["taken_at"] = register.get("taken_at", "")
    return got


def _unspent(project: str, rss: Path, estimate: dict[str, Any],
             horizon: Any, waterfall: dict[str, Any]) -> dict[str, Any] | None:
    """Постатейно: что до ввода израсходовано не будет.

    Реестра ГУ может не быть, реестра договоров в РСС может не быть — свободное
    от договоров считается и без них, а причина, по которой ГУ не разложены,
    называется вслух. Пусто здесь бывает только когда нечего показать вовсе.
    """
    import developaid_monitor_unspent as unspent_mod

    register = monitor.latest_retention(project)
    if register and register.get("known") is False:
        register = None
    try:
        contracts = actuals.read_contracts(rss)
    except Exception:  # noqa: BLE001 — листа договоров в РСС может не быть
        contracts = None
    needy = {str(row.get("code") or ""): float(row.get("unfunded_take") or 0.0)
             for row in (waterfall.get("articles") or [])}
    got = unspent_mod.unspent(estimate, contracts=contracts, retention=register,
                              horizon=horizon, needy=needy)
    return got if (got["articles"] or got["retention"]["reason"]) else None


def _funding_risk(project: str, rss: Path, cut: datetime.date, view: dict[str, Any]) -> dict[str, Any]:
    baseline = _finance_baseline(project)
    if not baseline.get("known"):
        return {"known": False, "reason": baseline.get("reason", "нет финансового baseline")}

    estimate = actuals.read_estimate(rss)
    current = _rss_ch23(estimate)
    paid_actual = _payment_total_ch23(rss, estimate)
    paid_delta = max(0.0, paid_actual - float(baseline["paid_at_baseline"] or 0.0))
    remaining_need = max(0.0, float(baseline["completion_need_at_baseline"] or 0.0) - paid_delta)

    articles = baseline.get("articles") or {}
    if not articles:
        return {
            "known": False,
            "reason": "в финансовом baseline не найдена постатейная программа потребности",
            "source": baseline.get("source"),
        }

    waterfall = _article_waterfall(
        articles,
        float(baseline.get("reserve") or 0.0),
        cut,
        _payment_by_code(rss),
    )
    rnv = monitor._day(_current_forecast_end(view))
    if rnv is None:
        rnv = monitor._day((view.get("schedule") or {}).get("approved_end"))

    reserve_start = waterfall["reserve_start"]
    reserve_exhaustion = waterfall["reserve_exhaustion"]
    return {
        "known": True,
        "source": baseline["source"],
        "bank_limit": current["limit"],
        "paid_actual": paid_actual,
        # Второй остаток потребности по тем же главам 2–3: утверждённый бюджет
        # минус оплаченное. С «средствами на завершение» они расходятся, и обе
        # величины считает сервер — экран их только показывает.
        "approved": float(baseline.get("approved") or 0.0),
        "approved_remaining": max(0.0, float(baseline.get("approved") or 0.0) - paid_actual),
        "bank_remaining": waterfall["opening_bank_remaining"],
        "opening_article_deficit": waterfall["opening_article_deficit"],
        "remaining_need": remaining_need,
        "reserve": float(baseline.get("reserve") or 0.0),
        "reserve_parts": baseline.get("reserve_parts") or {},
        "reserve_start": monitor._iso(reserve_start),
        "reserve_exhaustion": monitor._iso(reserve_exhaustion),
        # Backward-compatible field used by the red marker/card. It now means
        # the first uncovered need after article limits + explicit reserve.
        "bank_exhaustion": monitor._iso(reserve_exhaustion),
        "reserve_balance": waterfall["reserve_balance"],
        "reserve_need": waterfall["reserve_need"],
        "additional_financing": waterfall["additional_financing"],
        "additional_financing_from": waterfall["additional_financing_from"],
        "retention": _retention(project, rnv),
        # Что не будет выбрано до ввода: свободное от договоров считается точно,
        # ГУ раскладываются по статьям оценкой. Считает это отдельный модуль,
        # здесь только вход — лимиты и потребность живут выше и второй раз не
        # считаются.
        "unspent": _unspent(project, rss, estimate, rnv, waterfall),
        "forecast_to": monitor._iso(rnv),
        "monthly_need": waterfall["monthly_need"],
        "monthly_reserve_draw": waterfall["monthly_reserve_draw"],
        "monthly_unfunded": waterfall["monthly_unfunded"],
        "monthly_reserve_balance": waterfall["monthly_reserve_balance"],
        "articles": waterfall["articles"],
        "tail_after_apr": max(0.0, float(baseline.get("tail_after_apr") or 0.0)),
        "method": (
            "постатейный waterfall: потребность месяца → остаток лимита статьи RSS → "
            "резервы 2.8/2.9 → непокрытая потребность; свободные лимиты других статей "
            "автоматически не кросс-финансируют дефицит"
        ),
    }


def _physical_smr(rss: Path, estimate: dict[str, Any], cut: datetime.date) -> float:
    works = actuals.read_completed_works(rss)
    return sum(
        float(row.get("amount") or 0.0)
        for row in works.get("rows") or []
        if row.get("construction")
        and row.get("date")
        and row["date"] <= cut
        and str(row.get("code") or "").startswith("2")
    )


def _latest_sales_rows(project: str, upto: datetime.date | None = None) -> Path | None:
    folder = monitor._project_dir(project) / "sales"
    if not folder.exists():
        return None
    rows = sorted(folder.glob("*.json"))
    if upto:
        rows = [row for row in rows if row.stem[:10] <= upto.isoformat()]
    return rows[-1] if rows else None


def _sales_snapshot(project: str, cut: datetime.date) -> dict[str, Any]:
    """Продажи: числа, а не факт наличия файла.

    Источника два, и они дополняют друг друга. Книга обновляется раз в месяц
    и несёт историю (лист «План продаж», строки ФАКТ); строки руками несут
    то, чего книга ещё не знает — «в августе продано 4 лота» приходит словами
    за месяц до выгрузки. Месяц из строк перекрывает тот же месяц книги.
    """
    import json as _json

    monthly: dict[str, dict[str, float]] = {}
    sources: list[str] = []

    book = _latest_sales(project, cut)
    if book is not None:
        try:
            parsed = actuals.read_sales(book)
        except Exception:
            parsed = None
        if parsed:
            for row in parsed["rows"]:
                if not row.get("fact") or not row.get("month"):
                    continue
                monthly[monitor._iso(row["month"])[:7]] = {
                    "units": float(row.get("units") or 0.0),
                    "area": float(row.get("area") or 0.0),
                    "revenue": float(row.get("revenue") or 0.0),
                }
            sources.append(book.name)
        else:
            sources.append(f"{book.name} (лист продаж не разобран)")

    manual = _latest_sales_rows(project, cut)
    if manual is not None:
        stored = _json.loads(manual.read_text(encoding="utf-8"))
        for row in stored.get("rows") or []:
            month = monitor._iso(actuals._as_month(row.get("month")))[:7]
            if not month:
                continue
            monthly[month] = {
                "units": float(row.get("units") or 0.0),
                "area": float(row.get("area") or 0.0),
                "revenue": float(row.get("revenue") or 0.0),
            }
        sources.append(f"строки на {stored.get('taken_at', manual.stem)}")

    if not monthly:
        return {"known": False, "reason": "не загружены ни книга, ни строки продаж"}
    months = sorted(monthly)
    recent = [{"month": month, **monthly[month]} for month in months[-3:]]
    return {
        "known": True,
        "source": " + ".join(sources),
        "last_fact": months[-1],
        "total_units": sum(row["units"] for row in monthly.values()),
        "total_area": sum(row["area"] for row in monthly.values()),
        "total_revenue": sum(row["revenue"] for row in monthly.values()),
        "recent": recent,
    }


def _fmt_money(value: Any) -> str:
    number = float(value or 0.0)
    if abs(number) >= 1e9:
        text = f"{number/1e9:,.2f}".replace(",", " ").replace(".", ",")
        return f"{text} млрд ₽"
    text = f"{number/1e6:,.1f}".replace(",", " ").replace(".", ",")
    return f"{text} млн ₽"


def _month_name(value: Any) -> str:
    day = monitor._day(value)
    if day is None:
        return str(value or "")
    names = ("январь", "февраль", "март", "апрель", "май", "июнь", "июль",
             "август", "сентябрь", "октябрь", "ноябрь", "декабрь")
    return f"{names[day.month - 1]} {day.year}"


def _summary(view: dict[str, Any], funding: dict[str, Any],
             sales: dict[str, Any]) -> list[str]:
    """Управленческое резюме словами — из тех же чисел, что и графики.

    Лимит, утверждённая модель, дефицит и его дата не должны собираться в
    голове из столбиков: сервер, который их посчитал, обязан их и сказать.
    """
    out: list[str] = []
    schedule = (view.get("dashboard") or {}).get("schedule") or view.get("schedule") or {}
    approved = monitor._day(schedule.get("approved_finish") or schedule.get("approved_end"))
    forecast = monitor._day(schedule.get("forecast_finish"))
    delay = schedule.get("rnv_delay_days")
    if approved and forecast:
        drift = (f"на {delay} дней позже плана" if isinstance(delay, (int, float))
                 and delay and delay > 0 else "в срок")
        out.append(
            f"Срок. Утверждённый ввод — {approved.strftime('%d.%m.%Y')}; при "
            f"нынешнем темпе актов КС модель ждёт "
            f"{forecast.strftime('%d.%m.%Y')} — {drift}.")

    if not funding.get("known"):
        out.append("Деньги. Финансовая книга не загружена — потребность и "
                   "дефицит посчитать не из чего.")
        return out

    need = float(funding.get("remaining_need") or 0.0)
    bank = float(funding.get("bank_remaining") or 0.0)
    reserve = float(funding.get("reserve") or 0.0)
    fuel = bank + reserve
    # РСС — это то, что даёт банк; утверждённая модель — сколько надо реально,
    # чтобы построить (владелец, 30.08.2026). Значит потребность берётся из
    # модели, а лимиты банка с резервом — это источник, и главный дефицит есть
    # разница между ними. Прежде фраза брала потребностью банковскую колонку
    # «Средства на завершение» — то есть взгляд банка на остаток, — и дефицит
    # выходил вчетверо меньше настоящего.
    model_need = float(funding.get("approved_remaining") or 0.0)
    if model_need > 0:
        gap = max(0.0, model_need - fuel)
        money = (
            f"Деньги. По утверждённой модели достроить стоит "
            f"{_fmt_money(model_need)} (модель минус оплаченное). По РСС "
            f"осталось {_fmt_money(fuel)}: {_fmt_money(bank)} лимитов статей и "
            f"{_fmt_money(reserve)} резерва 2.8/2.9 — оставшийся лимит и есть "
            f"то, что банк готов дать; сам РСС при этом бюджет всей стройки, а "
            f"не банковская доля. Дефицит {_fmt_money(gap)}.")
        if need > 0:
            money += (f" Справочно: сам банк считает остаток к завершению в "
                      f"{_fmt_money(need)} — это его взгляд по РСС, а не "
                      f"потребность стройки.")
    else:
        money = (
            f"Деньги. Утверждённая модель не прочитана — сколько реально надо "
            f"достроить, сказать нечем. Банк даёт {_fmt_money(fuel)}; его "
            f"остаток к завершению по РСС — {_fmt_money(need)}.")
    out.append(money)

    start = funding.get("reserve_start")
    exhaustion = funding.get("reserve_exhaustion")
    unfunded = funding.get("monthly_unfunded") or {}
    first_gap = next((month for month in sorted(unfunded)
                      if float(unfunded[month] or 0.0) > 1e6), None)
    gap_total = float(funding.get("additional_financing") or 0.0)
    if gap_total > 1e6:
        pieces = []
        if start:
            pieces.append(f"с {_month_name(start)} статьи начинают тратить резерв")
        if exhaustion:
            pieces.append(f"{monitor._day(exhaustion).strftime('%d.%m.%Y')} резерв "
                          f"кончается")
        if first_gap:
            pieces.append(f"с {_month_name(first_gap)} платить нечем — "
                          f"в этот месяц не хватает "
                          f"{_fmt_money(unfunded[first_gap])}")
        out.append("Когда возникает дыра. " + "; ".join(pieces) + ". "
                   f"Всего до конца не хватает {_fmt_money(gap_total)} — это и "
                   "есть потребность в дофинансировании.")
        deadline = exhaustion or first_gap
        if deadline:
            out.append(
                f"Решение. Дофинансирование или увеличение лимитов на "
                f"{_fmt_money(gap_total)} нужно согласовать до "
                f"{monitor._day(deadline).strftime('%d.%m.%Y')} — дальше стройка "
                "платит только тем, что осталось, и темп станет падать.")
    else:
        out.append("Дыры по РСС нет: лимиты статей и резерв "
                   "закрывают потребность до конца.")

    rows = funding.get("articles") or []
    short = sorted((row for row in rows
                    if float(row.get("need_total") or 0.0)
                    > float(row.get("opening_limit") or 0.0) + 1e6),
                   key=lambda row: float(row.get("need_total") or 0.0)
                   - float(row.get("opening_limit") or 0.0), reverse=True)[:3]
    if short:
        names = "; ".join(
            f"{row['code']} {str(row.get('name') or '').strip()} — не хватает "
            f"{_fmt_money(float(row.get('need_total') or 0.0) - float(row.get('opening_limit') or 0.0))}"
            for row in short)
        out.append(f"Где не хватает больше всего: {names}. Резерв общий и "
                   "закрывает их по очереди месяцев, пока не кончится.")

    if sales.get("known"):
        out.append(
            f"Продажи. Продано {sales.get('total_units', 0):.0f} лотов на "
            f"{_fmt_money(sales.get('total_revenue'))}; последний факт — "
            f"{sales.get('last_fact', '')}.")
    return out


def _dashboard(project: str, rss: Path, cut: datetime.date, view: dict[str, Any]) -> dict[str, Any]:
    estimate = actuals.read_estimate(rss)
    finance = _finance_baseline(project)
    current = _rss_ch23(estimate)
    physical = _physical_smr(rss, estimate, cut)
    approved = float(finance.get("approved") or 0.0)
    funding = _funding_risk(project, rss, cut, view)
    graph = (view.get("schedule") or {}).get("dependency_graph") or {}
    return {
        "physical": {
            "accepted": physical,
            "completion": physical / approved if approved > 0 else None,
        },
        "construction": {
            "approved": approved or None,
            "limit": current["limit"],
            "contracted": current["contracted"],
            "remaining_need": funding.get("remaining_need") if funding.get("known") else None,
        },
        "schedule": {
            "approved_finish": (view.get("schedule") or {}).get("approved_end"),
            "forecast_finish": _current_forecast_end(view),
            "forecast_known": graph.get("forecast_known", True),
            "forecast_source": graph.get("forecast_source", ""),
            "rnv_delay_days": graph.get("rnv_delay_days"),
        },
        "sales": _sales_snapshot(project, cut),
        "funding": funding,
        "summary": _summary(view, funding, _sales_snapshot(project, cut)),
        "sources": {
            "rss": rss.name,
            "physical_fact": "Реестр выполненных работ",
            "payment_fact": "Реестр платежей",
            "financial_baseline": finance.get("source"),
        },
    }


def _store_sales_file(project: str, data: bytes, taken_at: Any) -> dict[str, Any]:
    day = monitor._iso(taken_at)
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        known = ("Продажи П-Ф", "Продажи", "Дашборд", "План продаж")
        if not any(name in wb.sheetnames for name in known):
            raise ValueError(
                "в книге не найден лист продаж — жду «Продажи П-Ф», «Продажи», "
                "«Дашборд» или «План продаж»")
    finally:
        wb.close()
    path = _sales_dir(project) / f"{day}.xlsx"
    if path.exists():
        raise FileExistsError(f"отчет продаж на {day} уже загружен")
    path.write_bytes(data)
    return {"taken_at": day, "stored": True, "path": str(path), "bytes": len(data)}


def _store_proposal(
    project: str, data: bytes, sheet: str, start: Any, code: str, taken_at: Any
) -> dict[str, Any]:
    return schedule_graph.store_reference(project, data, sheet, start, code, taken_at)


def _current_forecast_end(view: dict[str, Any]) -> Any:
    """Текущий прогноз РНВ — сеть по темпу, а не плоский максимум по строкам.

    `schedule.forecast_end` — это `max(forecast_finish)` по WBS: он не знает
    PM-зависимостей и потому отвечает на другой вопрос. Pace считает сеть и
    кладёт ответ в `pace_forecast_end`, но дашборд собирался заново и брал
    плоский максимум — на Кутузов Сити шапка показывала 04.01.2029, а
    «Платон · управленческий прогноз» 11.02.2028 на том же срезе и том же
    снимке РСС. Подпись шапки при этом обещала «КС / EAC proxy + PM
    dependencies» — обещание, которого она не исполняла.

    Плоский максимум остаётся запасным ответом: без PM-графа сети нет, и
    честнее показать позднейшую строку, чем ничего.
    """
    schedule = view.get("schedule") or {}
    return schedule.get("pace_forecast_end") or schedule.get("forecast_end")


def _build(
    project: str, cut: Any, programme: dict[str, Any] | None = None, upto: str = ""
) -> dict[str, Any]:
    if _ORIGINAL_BUILD is None:
        raise RuntimeError("dashboard layer is not installed")
    view = _ORIGINAL_BUILD(project, cut, programme=programme, upto=upto)
    view = schedule_graph.apply(project, view)
    rss = monitor._latest(project, "estimate", ".xlsx", upto or monitor._iso(cut))
    cut_date = monitor._day(cut)
    if rss is not None and cut_date is not None:
        dashboard = _dashboard(project, rss, cut_date, view)
        view["dashboard"] = dashboard
        view["financing"] = dashboard["funding"]
    return view


def install() -> None:
    global _INSTALLED, _ORIGINAL_BUILD, _ORIGINAL_STORE_SALES_FILE, _ORIGINAL_STORE_PROPOSAL
    if _INSTALLED:
        return
    _ORIGINAL_BUILD = monitor.build
    _ORIGINAL_STORE_SALES_FILE = getattr(monitor, "store_sales_file", None)
    _ORIGINAL_STORE_PROPOSAL = getattr(monitor, "store_proposal", None)
    monitor.build = _build
    monitor.store_sales_file = _store_sales_file
    monitor.store_proposal = _store_proposal
    _INSTALLED = True
